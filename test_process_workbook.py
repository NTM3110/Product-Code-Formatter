import json
import re
import os
import unittest
from collections import Counter
from io import BytesIO
from pathlib import Path
from tempfile import TemporaryDirectory
from uuid import uuid4

from openpyxl import Workbook, load_workbook
import xlrd

import app
from product_code.workflow_runtime import process_is_running
from workflows.estimate_extractor.logic import analyze_estimate_workbook, list_estimate_workbook_sheets
from app import build_vietmax_ban_ra_purchase_matches, create_up_ban_ra_workbook, create_up_mua_vao_workbook, default_config, excel_col_to_index, license_allows_company, license_allows_profile, license_has_local_activation, make_product_part, normalize_config, process_workbook, profile_key, resolve_output_path, suggest_prefix, up_ban_ra_output_path, up_mua_vao_output_path, vietmax_ban_ra_sales_products_from_workbook, vietmax_product_review_rows, vietmax_purchase_match_export_rows, vietmax_purchase_products_from_workbook


class XlsCellAdapter:
    def __init__(self, value):
        self.value = value


class XlsSheetAdapter:
    def __init__(self, sheet):
        self._sheet = sheet
        self.max_column = sheet.ncols
        self.max_row = sheet.nrows

    def cell(self, row, col):
        return XlsCellAdapter(self._sheet.cell_value(row - 1, col - 1))


class XlsWorkbookAdapter:
    def __init__(self, data):
        self._workbook = xlrd.open_workbook(file_contents=data)
        self.sheetnames = self._workbook.sheet_names()

    def __getitem__(self, name):
        return XlsSheetAdapter(self._workbook.sheet_by_name(name))

    def close(self):
        self._workbook.release_resources()


class ProcessWorkbookTests(unittest.TestCase):
    def test_ho_guom_defaults_define_three_purchase_forms_and_groups(self):
        profile = app.normalize_profile_config("ho_guom", {})
        groups = {group["id"]: group for group in profile["processing_groups"]}
        self.assertEqual(
            set(groups),
            {"materials", "services", "payment_voucher", "ignored"},
        )
        self.assertTrue(groups["materials"]["uses_product_code"])
        self.assertTrue(groups["services"]["uses_product_code"])
        self.assertTrue(groups["payment_voucher"]["uses_product_code"])
        self.assertFalse(groups["ignored"]["uses_product_code"])

        forms = app.load_system_default_form_mappings(profile="ho_guom")
        self.assertEqual(
            {(form["id"], form["group_id"], form["sheet"]) for form in forms},
            {
                ("ho_guom_imexpng", "materials", "IMEXPNG"),
                ("ho_guom_imexpn1", "services", "IMEXPN1"),
                ("ho_guom_imexpc1", "payment_voucher", "IMEXPC1"),
            },
        )

        profile["disabled_default_form_ids"] = ["ho_guom_imexpn1"]
        restored = app.apply_system_default_forms_to_config({"profiles": {"ho_guom": profile}})
        self.assertNotIn(
            "ho_guom_imexpn1",
            {form["id"] for form in restored["profiles"]["ho_guom"]["form_mapping_presets"]},
        )

    def test_ho_guom_purchase_forms_use_seller_identity_columns(self):
        forms = {
            form["id"]: form
            for form in app.load_system_default_form_mappings(profile="ho_guom")
        }

        def source_columns(form_id):
            return {
                mapping["target_col"]: mapping.get("source_col")
                for mapping in forms[form_id]["mappings"]
                if mapping.get("source_type") == "source_column"
            }

        self.assertEqual(
            {key: source_columns("ho_guom_imexpng")[key] for key in ("C", "D", "X", "Y", "Z", "AA")},
            {"C": "G", "D": "F", "X": "G", "Y": "F", "Z": "H", "AA": "G"},
        )
        self.assertEqual(
            {key: source_columns("ho_guom_imexpng")[key] for key in ("J", "P")},
            {"J": "MA_KHO", "P": "TK_VAT_TU"},
        )
        self.assertEqual(
            {column["letter"] for column in app.ho_guom_default_source_columns()} & {"MA_KHO", "TK_VAT_TU"},
            {"MA_KHO", "TK_VAT_TU"},
        )
        self.assertEqual(
            {key: source_columns("ho_guom_imexpn1")[key] for key in ("A", "B")},
            {"A": "G", "B": "F"},
        )
        self.assertEqual(
            {key: source_columns("ho_guom_imexpc1")[key] for key in ("A", "J", "T", "U", "V", "W")},
            {"A": "F", "J": "F", "T": "G", "U": "F", "V": "H", "W": "G"},
        )

    def test_ho_guom_old_default_inventory_mapping_is_migrated_without_touching_custom_rules(self):
        old_form = next(
            form for form in app.ho_guom_default_form_mapping_presets()
            if form["id"] == "ho_guom_imexpng"
        )
        old_form["mappings"] = [
            rule for rule in old_form["mappings"]
            if rule.get("target_col") != "P"
        ]
        for rule in old_form["mappings"]:
            if rule.get("target_col") == "J":
                rule.clear()
                rule.update({
                    "target_col": "J",
                    "source_type": "constant",
                    "source_phase": "purchase",
                    "value": "KHO1",
                })
            if rule.get("target_col") == "S":
                rule["value"] = "CUSTOM"

        merged = app.merge_system_default_form_mappings([old_form], profile="ho_guom")
        migrated = next(form for form in merged if form["id"] == "ho_guom_imexpng")
        mappings = {rule["target_col"]: rule for rule in migrated["mappings"]}
        self.assertEqual(mappings["J"].get("source_col"), "MA_KHO")
        self.assertEqual(mappings["P"].get("source_col"), "TK_VAT_TU")
        self.assertEqual(mappings["S"].get("value"), "CUSTOM")

    def test_ho_guom_formatter_excludes_zero_material_quantity_but_keeps_other_groups(self):
        with TemporaryDirectory() as tmpdir:
            source = Path(tmpdir) / "ho_guom.xlsx"
            output = Path(tmpdir) / "ho_guom_fdi.xls"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Invoices"
            for letter, header in app.HO_GUOM_FDI_SOURCE_HEADERS.items():
                sheet.cell(2, app.excel_col_to_index(letter) + 1, header)
            sheet.cell(2, app.excel_col_to_index("AZ") + 1, "TK vật tư")
            sheet.cell(2, app.excel_col_to_index("BA") + 1, "Mã kho")

            def add_row(row_number, *, invoice, company, mst, product, quantity, amount=100, tax=8):
                values = {
                    "B": "1C24TAA",
                    "C": invoice,
                    "D": "01/06/2026",
                    "F": company,
                    "G": mst,
                    "L": 0,
                    "M": product,
                    "N": "kg",
                    "O": quantity,
                    "P": 10,
                    "V": "08",
                    "W": 0,
                    "X": 0,
                    "Y": 0,
                    "Z": amount,
                    "AA": tax,
                    "AU": amount,
                    "AV": tax,
                    "AZ": "156",
                    "BA": "KHHVT",
                }
                for letter, value in values.items():
                    sheet.cell(row_number, app.excel_col_to_index(letter) + 1, value)

            add_row(3, invoice="237", company="Công ty Dương Châu", mst="010A", product="Vật tư bỏ", quantity=0)
            add_row(4, invoice="237", company="Công ty Dương Châu", mst="010A", product="Vật tư giữ", quantity=5)
            add_row(5, invoice="237", company="Công ty Dương Châu", mst="010A", product="Vật tư giữ 2", quantity=7)
            add_row(6, invoice="237", company="Công ty Khác", mst="010D", product="Vật tư công ty khác", quantity=9)
            add_row(7, invoice="501", company="Công ty Dịch vụ", mst="010B", product="Dịch vụ kiểm định", quantity=0)
            add_row(8, invoice="601", company="Công ty Phiếu chi", mst="010C", product="Xăng E10", quantity=0)
            workbook.save(source)
            workbook.close()

            payload = {
                "profile": "ho_guom",
                "include_company_prefix": True,
                "all_mst": ["010A", "010D", "010B", "010C"],
                "process_mst": ["010A", "010D", "010B", "010C"],
                "mst_safe_id": ["010A|||0", "010D|||1", "010B|||2", "010C|||3"],
                "prefix_0": "DC",
                "prefix_1": "KH",
                "prefix_2": "DV",
                "prefix_3": "PC",
                "selected_products_0": ["Vật tư bỏ", "Vật tư giữ", "Vật tư giữ 2"],
                "selected_products_1": ["Vật tư công ty khác"],
                "selected_products_2": ["Dịch vụ kiểm định"],
                "selected_products_3": ["Xăng E10"],
                "company_group_assignments": {
                    "010A": "materials",
                    "010D": "materials",
                    "010B": "services",
                    "010C": "payment_voucher",
                },
            }
            processed = app.process_workbook(source, output, payload)
            code_index = app.excel_col_to_index("L")
            product_index = app.excel_col_to_index("M")
            group_index = processed.shape[1] - 1
            data = {
                app.raw_text(row.iloc[product_index]): (
                    app.raw_text(row.iloc[code_index]),
                    app.raw_text(row.iloc[group_index]),
                )
                for _, row in processed.iloc[2:].iterrows()
            }

            self.assertNotIn("Vật tư bỏ", data)
            self.assertEqual(data["Vật tư giữ"], ("VT.DC.237.01", "materials"))
            self.assertEqual(data["Vật tư giữ 2"], ("VT.DC.237.02", "materials"))
            self.assertEqual(data["Vật tư công ty khác"], ("VT.KH.237.01", "materials"))
            self.assertEqual(data["Dịch vụ kiểm định"], ("VT.DV.501.01", "services"))
            self.assertEqual(data["Xăng E10"], ("VT.PC.601.01", "payment_voucher"))

            sheets = app.fast_import_sheet_rows(
                processed,
                None,
                form_mapping_presets=app.load_system_default_form_mappings(profile="ho_guom"),
                require_inventory_columns=False,
                include_duplicate_report=False,
                profile="ho_guom",
            )
            material_sheets = {name: value for name, value in sheets.items() if name.startswith("IMEXPNG")}
            self.assertEqual(set(sheets) - set(material_sheets), {"IMEXPN1", "IMEXPC1"})
            self.assertEqual(len(material_sheets), 2)
            self.assertEqual(sum(len(rows) for _, rows in material_sheets.values()), 3)
            for _, rows in material_sheets.values():
                self.assertTrue(all(row[app.excel_col_to_index("J")] == "KHHVT" for row in rows))
                self.assertTrue(all(row[app.excel_col_to_index("P")] == "156" for row in rows))
            self.assertEqual(len(sheets["IMEXPN1"][1]), 1)
            self.assertEqual(len(sheets["IMEXPC1"][1]), 1)

    def test_ho_guom_workflow_preflight_does_not_require_inventory_pairs(self):
        import web_api
        from product_code.workflow_runtime import WorkflowFailure

        with TemporaryDirectory() as tmpdir:
            source = Path(tmpdir) / "ho_guom.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            for column in range(1, 16):
                sheet.cell(1, column, f"Cột {column}")
            workbook.save(source)
            workbook.close()

            payload = {
                "profile": "ho_guom",
                "process_mst": ["010A"],
                "company_col": "F",
                "mst_col": "G",
                "product_col": "M",
                "qty_col": "O",
                "inventory_pairs": [],
            }
            web_api.validate_workflow_process_payload(source, payload, app.VIETMAX_PHASE_PURCHASE)

            payload["profile"] = "quang_thinh"
            with self.assertRaises(WorkflowFailure) as raised:
                web_api.validate_workflow_process_payload(source, payload, app.VIETMAX_PHASE_PURCHASE)
            self.assertEqual(raised.exception.code, "INVENTORY_PAIR_REQUIRED")

    def test_son_phuong_inventory_defaults_are_scoped_and_preserve_custom_values(self):
        config = normalize_config({
            "profiles": {
                "son_phuong": {
                    "inventory_allocation_config": {
                        "scenario_count": 250,
                        "sales_inventory_pairs": [
                            {"id": "custom-materials", "role": "materials", "ma_kho": "KHH-CUSTOM", "tk_vat_tu": "1561"},
                        ],
                    },
                    "scopes": {
                        "purchase": {
                            "inventory_pairs": [
                                {"id": "purchase-custom", "role": "materials", "ma_kho": "KHH-MUA", "tk_vat_tu": "1562"},
                            ],
                        },
                    },
                },
            },
        })

        profile = config["profiles"]["son_phuong"]
        purchase_pairs = profile["scopes"]["purchase"]["inventory_pairs"]
        sales_pairs = profile["inventory_allocation_config"]["sales_inventory_pairs"]
        self.assertEqual(len(purchase_pairs), 1)
        self.assertEqual(profile["inventory_allocation_config"]["policy"]["generic_min_type_count"], 2)

        self.assertEqual(purchase_pairs[0]["ma_kho"], "KHH-MUA")
        self.assertEqual(purchase_pairs[0]["tk_vat_tu"], "1562")
        self.assertEqual(profile["inventory_allocation_config"]["scenario_count"], 250)
        self.assertEqual(next(item for item in sales_pairs if item["role"] == "materials")["ma_kho"], "KHH-CUSTOM")
        self.assertEqual(next(item for item in sales_pairs if item["role"] == "finished_goods")["ma_kho"], "KTP")
        self.assertEqual(next(item for item in sales_pairs if item["role"] == "fallback")["ma_kho"], "KHOCK")

    def test_estimate_detector_supports_raw_layout_and_excel_control_markers(self):
        with TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "raw_estimate.xlsx"
            workbook = Workbook()
            bid = workbook.active
            bid.title = "D\u1ef1 th\u1ea7u"
            detail = workbook.create_sheet("\u0110\u01a1n gi\u00e1 chi ti\u1ebft")
            bid.append([])
            bid.append([])
            bid.append([])
            bid.append([])
            bid.append(["STT", "M\u00e3 hi\u1ec7u", "C\u1ed9t \u1ea9n", "T\u00ean c\u00f4ng t\u00e1c", "\u0110\u01a1n v\u1ecb", "Kh\u1ed1i l\u01b0\u1ee3ng", "\u0110\u01a1n gi\u00e1_x000D_\n(\u0111)", "Th\u00e0nh ti\u1ec1n_x000D_\n(\u0111)"])
            bid.append([1, "AA.100", None, "C\u00f4ng t\u00e1c A", "m", 2, 10, 20])
            detail.append([])
            detail.append([])
            detail.append([])
            detail.append([])
            detail.append(["STT", "M\u00e3 hi\u1ec7u \u0111\u01a1n gi\u00e1", "T\u00ean c\u00f4ng t\u00e1c", "\u0110\u01a1n v\u1ecb", "\u0110\u1ecbnh m\u1ee9c", "\u0110\u01a1n gi\u00e1_x000D_\n(\u0111)", "H\u1ec7 s\u1ed1", "Th\u00e0nh ti\u1ec1n_x000D_\n(\u0111)"])
            detail.append([1, "AA.100", "C\u00f4ng t\u00e1c A", "m", 1, None, None, None])
            detail.append([None, None, "V\u1eadt li\u1ec7u", None, None, None, None, 10])
            detail.append([None, "VL.1", "V\u1eadt li\u1ec7u A", "kg", 1, 10, 1, 10])
            workbook.save(path)

            info = list_estimate_workbook_sheets(str(path))
            suggested = info["suggested_sheets"]
            self.assertEqual(suggested["bid_header_row"], 5)
            self.assertEqual(suggested["detail_header_row"], 5)
            self.assertEqual(suggested["bid_columns"]["name"], "D")
            self.assertEqual(suggested["bid_columns"]["unit_total"], "G")
            self.assertEqual(suggested["detail_columns"]["name"], "C")
            result = analyze_estimate_workbook(str(path))
            self.assertTrue(result.ok)
            self.assertEqual((result.bid_rows, result.detail_blocks), (1, 1))

    def test_workflow_runtime_detects_current_process(self):
        self.assertTrue(process_is_running(os.getpid()))
        self.assertFalse(process_is_running(-1))
        self.assertFalse(process_is_running("invalid"))

    def test_product_code_replacement_ignores_company_prefix(self):
        replacements = app.normalize_product_code_replacements({"TU100": "TU100.001"})
        self.assertEqual(app.apply_product_code_replacement("TU100", replacements), "TU100.001")
        self.assertEqual(app.apply_product_code_replacement("TA.TU100", replacements), "TA.TU100.001")
        self.assertEqual(app.apply_product_code_replacement("TP.TU100", replacements), "TP.TU100.001")

    def test_invoice_status_filter_keeps_new_invoices(self):
        self.assertFalse(app.ignored_invoice_status("Hóa đơn mới"))
        self.assertFalse(app.ignored_invoice_status("Hóa đơn mới", ["Hóa đơn đã bị hủy"]))
        self.assertTrue(app.ignored_invoice_status("Hóa đơn đã bị hủy"))
        self.assertTrue(app.ignored_invoice_status("Hóa đơn đã bị hủy", ["Hóa đơn đã bị hủy"]))
    def test_analyze_includes_missing_mst_new_invoices_as_companies(self):
        with TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "missing_mst.xlsx"
            wb = Workbook()
            ws = wb.active

            def write_row(row, invoice_no, company, mst, product, qty, status):
                ws.cell(row=row, column=3, value=invoice_no)
                ws.cell(row=row, column=6, value=company)
                ws.cell(row=row, column=7, value=mst)
                ws.cell(row=row, column=13, value=product)
                ws.cell(row=row, column=15, value=qty)
                ws.cell(row=row, column=36, value=status)

            write_row(1, "283", "Công ty Không MST", "", "Hàng hóa mới", 1, "Hóa đơn mới")
            write_row(2, "302", "Công ty Không MST", None, "Hàng hóa mới 2", 2, "Hóa đơn mới")
            write_row(3, "999", "Công ty Không MST", "", "Hàng bị hủy", 1, "Hóa đơn đã bị hủy")
            write_row(4, "100", "Công ty Có MST", "0100100100", "Hàng xử lý", 1, "Hóa đơn mới")
            wb.save(path)

            result = app.analyze(str(path), "F", "G", "H", "M", "O", "", {}, invoice_status_col="AJ", invoice_status_skip_values=["Hóa đơn đã bị hủy"])

        self.assertEqual(result["rows_to_process"], 3)
        self.assertEqual(result["company_count"], 2)
        self.assertEqual(result["missing_mst_row_count"], 0)
        self.assertEqual(result["missing_mst_companies"], [])
        missing_company = next(company for company in result["companies"] if not company["mst"])
        self.assertEqual(missing_company["company"], "Công ty Không MST")
        self.assertTrue(missing_company["missing_mst"])
        self.assertTrue(missing_company["company_id"].startswith("__NO_MST__:"))
        self.assertEqual(missing_company["count"], 2)
        self.assertEqual({product["name"] for product in missing_company["all_products"]}, {"Hàng hóa mới", "Hàng hóa mới 2"})

    def test_process_workbook_can_select_or_skip_missing_mst_company(self):
        outputs = Path(__file__).parent / "outputs"
        outputs.mkdir(exist_ok=True)
        run_id = uuid4().hex
        source = outputs / f"_test_missing_mst_source_{run_id}.xlsx"
        selected_result = outputs / f"_test_missing_mst_selected_{run_id}.xlsx"
        skipped_result = outputs / f"_test_missing_mst_skipped_{run_id}.xlsx"
        try:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Invoices"
            sheet.cell(2, 12, "Ma VT")
            sheet.cell(2, 13, "Ten hang")
            sheet.cell(2, 15, "So luong")
            sheet.cell(3, 3, "HD-NO-MST")
            sheet.cell(3, 6, "Cong ty Khong MST")
            sheet.cell(3, 7, "")
            sheet.cell(3, 12, 0)
            sheet.cell(3, 13, "Hang Khong MST")
            sheet.cell(3, 15, 1)
            workbook.save(source)
            workbook.close()

            company_id = app.company_selection_key("Cong ty Khong MST", "")
            process_workbook(source, selected_result, {
                "company_col": "F",
                "mst_col": "G",
                "product_col": "M",
                "qty_col": "O",
                "output_col": "L",
                "profile": "vietmax",
                "vietmax_phase": "sales",
                "include_company_prefix": True,
                "all_mst": [company_id],
                "process_mst": [company_id],
                "mst_safe_id": [f"{company_id}|||0"],
                "prefix_0": "CTKM",
                "selected_products_0": ["Hang Khong MST"],
            })
            selected_output = load_workbook(selected_result, data_only=True)
            try:
                selected_sheet = selected_output["Invoices"]
                self.assertEqual(selected_sheet.max_row, 3)
                self.assertTrue(str(selected_sheet.cell(3, 12).value).startswith("CTKM."))
                self.assertIsNone(selected_sheet.cell(3, 7).value)
                customer_code_col = next(
                    col
                    for col in range(1, selected_sheet.max_column + 1)
                    if app.normalized_header_label(selected_sheet.cell(2, col).value) == "ma khach hang"
                )
                self.assertEqual(selected_sheet.cell(3, customer_code_col).value, "CTKM")
            finally:
                selected_output.close()

            process_workbook(source, skipped_result, {
                "company_col": "F",
                "mst_col": "G",
                "product_col": "M",
                "qty_col": "O",
                "output_col": "L",
                "profile": "vietmax",
                "vietmax_phase": "sales",
                "include_company_prefix": True,
                "all_mst": [company_id],
                "process_mst": [],
                "mst_safe_id": [f"{company_id}|||0"],
                "prefix_0": "CTKM",
                "selected_products_0": ["Hang Khong MST"],
            })
            skipped_output = load_workbook(skipped_result, data_only=True)
            try:
                skipped_sheet = skipped_output["Invoices"]
                self.assertEqual(skipped_sheet.max_row, 2)
            finally:
                skipped_output.close()
        finally:
            source.unlink(missing_ok=True)
            selected_result.unlink(missing_ok=True)
            skipped_result.unlink(missing_ok=True)
    def test_vietmax_profiles_are_available_by_default(self):
        config = default_config()
        self.assertEqual(profile_key("vietmax"), "vietmax")
        self.assertIn("vietmax", config["profiles"])
        self.assertIn("vietmax_mua_vao", config["profiles"])
        self.assertIn("vietmax_ban_ra", config["profiles"])

    def test_unified_vietmax_profile_config_is_canonical(self):
        config = normalize_config({
            "selected_profile": "vietmax",
            "profiles": {
                "vietmax": {
                    "manual_code_overrides": {"MST|||Hang A": "VM.A"},
                    "inventory_pairs": [{"id": "one", "ma_kho": "KHO", "tk_vat_tu": "152"}],
                }
            },
        })

        self.assertEqual(config["selected_profile"], "vietmax")
        self.assertEqual(config["profiles"]["vietmax"]["manual_code_overrides"], {"MST|||Hang A": "VM.A"})
        self.assertEqual(config["profiles"]["vietmax"]["inventory_pairs"][0]["id"], "one")

    def test_license_config_round_trips_allowed_profiles(self):
        config = normalize_config({
            "license": {
                "server_url": "http://license-server.local:3000/",
                "account_id": "acct",
                "license_key": "KEY",
                "activated": True,
                "allowed_profiles": "son_phuong; Cao Thành\nvietmax_ban_ra",
            }
        })

        self.assertEqual(config["license"]["server_url"], "http://license-server.local:3000/")
        self.assertTrue(config["license"]["activated"])
        self.assertEqual(config["license"]["allowed_profiles"], ["son_phuong", "Cao Thành", "vietmax_ban_ra"])

    def test_license_allowed_profiles_match_key_or_label(self):
        self.assertTrue(license_allows_profile("cao_thanh", ["Cao Thành"]))
        self.assertTrue(license_allows_profile("vietmax_ban_ra", ["vietmax_ban_ra"]))
        self.assertFalse(license_allows_profile("quang_thinh", ["son_phuong", "cao_thanh"]))

    def test_license_local_activation_trusts_same_machine_without_server(self):
        original_fingerprint = app.local_machine_fingerprint
        try:
            app.local_machine_fingerprint = lambda: "TEST-FINGERPRINT"
            self.assertTrue(license_has_local_activation({"activated": True, "machine_fingerprint": "TEST-FINGERPRINT"}))
            self.assertFalse(license_has_local_activation({"activated": True, "machine_fingerprint": "OTHER-FINGERPRINT"}))
            self.assertFalse(license_has_local_activation({"activated": False, "machine_fingerprint": "TEST-FINGERPRINT"}))
        finally:
            app.local_machine_fingerprint = original_fingerprint

    def test_license_local_activation_accepts_legacy_same_computer_fingerprint(self):
        original_fingerprint = app.local_machine_fingerprint
        original_legacy_fingerprint = app.legacy_machine_fingerprint
        original_machine_name = app.current_machine_name
        try:
            app.local_machine_fingerprint = lambda: "DESKTOP-SGFC2SA-WIN-STABLE-GUID"
            app.legacy_machine_fingerprint = lambda: "DESKTOP-SGFC2SA-28A06B93D8B0"
            app.current_machine_name = lambda: "DESKTOP-SGFC2SA"
            self.assertTrue(license_has_local_activation({"activated": True, "machine_fingerprint": "DESKTOP-SGFC2SA-28A06B93D8AD"}))
            self.assertFalse(license_has_local_activation({"activated": True, "machine_fingerprint": "OTHERPC-28A06B93D8AD"}))
        finally:
            app.local_machine_fingerprint = original_fingerprint
            app.legacy_machine_fingerprint = original_legacy_fingerprint
            app.current_machine_name = original_machine_name

    def test_load_config_recovers_local_license_from_backup(self):
        original_config_path = app.CONFIG_PATH
        original_license_path = app.LICENSE_PATH
        original_default_mappings_path = app.DEFAULT_FORM_MAPPINGS_PATH
        original_fingerprint = app.local_machine_fingerprint
        original_legacy_fingerprint = app.legacy_machine_fingerprint
        try:
            with TemporaryDirectory() as tmp:
                config_path = Path(tmp) / "product_code_config.json"
                backup_path = Path(tmp) / "product_code_config.manual_code_overrides_backup.json"
                current = default_config()
                backup = default_config()
                backup["license"].update({
                    "server_url": "http://license-server.local:3000",
                    "account_id": "acct",
                    "license_key": "KEY",
                    "activated": True,
                    "machine_fingerprint": "TEST-FINGERPRINT",
                    "status": "ACTIVE",
                    "allowed_profiles": ["vietmax"],
                })
                config_path.write_text(json.dumps(current), encoding="utf-8")
                backup_path.write_text(json.dumps(backup), encoding="utf-8")
                app.CONFIG_PATH = config_path
                app.LICENSE_PATH = Path(tmp) / "license.json"
                app.DEFAULT_FORM_MAPPINGS_PATH = Path(tmp) / "default_form_mappings.json"
                app.local_machine_fingerprint = lambda: "TEST-FINGERPRINT"
                app.legacy_machine_fingerprint = lambda: "LEGACY-FINGERPRINT"

                loaded = app.load_config()
                persisted = json.loads(config_path.read_text(encoding="utf-8"))

                self.assertTrue(loaded["license"]["activated"])
                self.assertEqual(loaded["license"]["license_key"], "KEY")
                self.assertEqual(loaded["license"]["allowed_profiles"], ["vietmax"])
                self.assertNotIn("license", persisted)
                persisted_license = json.loads(app.LICENSE_PATH.read_text(encoding="utf-8"))
                self.assertTrue(persisted_license["activated"])
                self.assertEqual(len(loaded["profiles"]["son_phuong"]["form_mapping_presets"]), 5)
                self.assertTrue(app.DEFAULT_FORM_MAPPINGS_PATH.exists())

                loaded["profiles"]["son_phuong"]["form_mapping_presets"][0]["enabled"] = False
                loaded["profiles"]["son_phuong"]["form_mapping_initialized"] = True
                app.save_config(loaded)
                reloaded = app.load_config()["profiles"]["son_phuong"]["form_mapping_presets"]
                self.assertFalse(reloaded[0]["enabled"])
                self.assertEqual(len(reloaded), 5)

                restored_forms = app.load_system_default_form_mappings()
                restored_forms[0].update({
                    "type": "template_mapping",
                    "template_saved_name": "missing-cache.xls",
                    "output_columns": [],
                })
                loaded = app.load_config()
                loaded["profiles"]["son_phuong"]["form_mapping_presets"] = restored_forms
                app.save_config(loaded)
                hydrated = app.load_config()["profiles"]["son_phuong"]["form_mapping_presets"]
                self.assertEqual(len(hydrated), 5)
                self.assertEqual(hydrated[0]["type"], "builtin")
                self.assertNotIn("template_saved_name", hydrated[0])
                self.assertTrue(hydrated[0]["output_columns"])
        finally:
            app.CONFIG_PATH = original_config_path
            app.LICENSE_PATH = original_license_path
            app.DEFAULT_FORM_MAPPINGS_PATH = original_default_mappings_path
            app.local_machine_fingerprint = original_fingerprint
            app.legacy_machine_fingerprint = original_legacy_fingerprint

    def test_keygen_metadata_accepts_camel_case_profile_keys(self):
        response = {
            "data": {
                "attributes": {
                    "metadata": {
                        "allowedProfiles": ["cao_thanh", "vietmax_ban_ra"]
                    }
                }
            }
        }
        metadata = app.keygen_license_metadata(response)
        self.assertEqual(app.extract_allowed_profiles(metadata), ["cao_thanh", "vietmax_ban_ra"])

    def test_keygen_company_metadata_does_not_limit_profiles(self):
        metadata = {"allowed_companies": ["BAC MAI"]}

        self.assertEqual(app.extract_allowed_companies(metadata), ["BAC MAI"])
        self.assertEqual(app.extract_allowed_profiles(metadata), [])

    def test_license_allowed_companies_match_mst_or_name(self):
        company = {"mst": "0303062904", "company": "Công ty A", "all_names": ["Tên khác"]}

        self.assertTrue(license_allows_company(company, ["0303062904"]))
        self.assertTrue(license_allows_company({"mst": "0303062904.0", "company": "X"}, ["03030629040"]))
        self.assertTrue(license_allows_company(company, ["cong ty a"]))
        self.assertFalse(license_allows_company(company, ["030857608"]))

    def test_suggest_prefix_skips_viet_nam_and_provinces(self):
        self.assertEqual(suggest_prefix("Công ty TNHH In Bao Bì Việt Nam"), "BB")
        self.assertEqual(suggest_prefix("Công ty TNHH In Bao Bì Hà Nội"), "BB")

    def test_son_phuong_decimal_comma_dimensions_keep_original_normalization(self):
        self.assertEqual(
            app.make_code("", "Dầu Điêzen 0,001S Mức 5", 1, {}, "son_phuong", {}, require_qty=True, include_company_prefix=False),
            "DD0.001SM5",
        )
        self.assertEqual(
            app.make_code("", "ống thép mạ tròn 21,2 x 2,1 x 6.0", 1, {}, "son_phuong", {}, require_qty=True, include_company_prefix=False),
            "OTMT21.2X2.1X6.0",
        )
        self.assertEqual(
            app.make_code("", "Φ88.3x1.8", 1, {}, "son_phuong", {}, require_qty=True, include_company_prefix=False),
            "F88.3X1.8",
        )

    def test_son_phuong_sample_product_codes_match_legacy_output(self):
        base = Path("E:/Excel Mom") / "D\u00f3c" / "Son Phuong" / "input_ouput"
        source = base / "Chi_tiet_HD_mua_vao_theo_tung_san_pham_01-01-2026_30-05-2026.xlsx"
        expected = base / "Chi_tiet_HD_mua_vao_theo_tung_san_pham_01-01-2026_30-05-2026_ket_qua_xu_ly" / "Chi_tiet_HD_mua_vao_theo_tung_san_pham_01-01-2026_30-05-2026_formatted.xlsx"
        if not source.exists() or not expected.exists():
            self.skipTest("Sơn Phương sample input/output files are not available on this machine.")

        _, source_df = app.read_workbook(source)
        _, expected_df = app.read_workbook(expected)
        word_rules = {"t\u00f4n": "TON", "\u0111en": "DEN"}

        def row_key(df, row_index):
            return (
                app.raw_text(app.cell(df, row_index, 2)),
                app.raw_text(app.cell(df, row_index, 5)),
                app.raw_text(app.cell(df, row_index, 6)),
                app.raw_text(app.cell(df, row_index, 12)),
                app.raw_text(app.cell(df, row_index, 14)),
            )

        source_keys = Counter(
            row_key(source_df, row_index)
            for row_index in range(2, len(source_df))
            if app.raw_text(app.cell(source_df, row_index, 2)) and app.raw_text(app.cell(source_df, row_index, 12))
        )
        prefix_map = {}
        expected_rows = []
        for row_index in range(2, len(expected_df)):
            key = row_key(expected_df, row_index)
            invoice_no, seller, mst, product, qty = key
            expected_code = app.raw_text(app.cell(expected_df, row_index, 11))
            if not invoice_no or not product or not qty or not expected_code:
                continue
            company_id = app.company_selection_key(seller, mst)
            prefix = expected_code.split(".", 1)[0] if "." in expected_code else ""
            if prefix:
                prefix_map.setdefault(company_id, prefix)
            expected_rows.append((row_index + 1, key, company_id, product, app.cell(expected_df, row_index, 14), expected_code))

        missing = []
        mismatches = []
        compared_rows = 0
        skipped_decimal_comma_rows = 0
        for excel_row, key, company_id, product, qty, expected_code in expected_rows:
            if source_keys[key] <= 0:
                missing.append((excel_row, key))
                if len(missing) >= 5:
                    break
            else:
                source_keys[key] -= 1
            if re.search(r"(?<=\d),(?=\d)", app.raw_text(product)):
                skipped_decimal_comma_rows += 1
                continue
            compared_rows += 1
            actual_code = app.make_code(company_id, product, qty, prefix_map, "son_phuong", word_rules, {}, require_qty=True, include_company_prefix=True)
            if actual_code != expected_code:
                mismatches.append((excel_row, expected_code, actual_code, product))
                if len(mismatches) >= 5:
                    break

        self.assertEqual(len(expected_rows), 6753)
        self.assertEqual(skipped_decimal_comma_rows, 60)
        self.assertEqual(compared_rows, 6693)
        self.assertFalse(missing, [(row, tuple(str(value).encode("unicode_escape").decode() for value in key)) for row, key in missing])
        self.assertFalse(mismatches, [(row, expected_code, actual_code, str(product).encode("unicode_escape").decode()) for row, expected_code, actual_code, product in mismatches])

    def test_keygen_url_allows_lan_http(self):
        self.assertTrue(app.keygen_url("http://license-server.local:3000", "acct", "licenses/actions/validate-key").startswith("http://license-server.local:3000/"))
        self.assertTrue(app.keygen_url("http://192.168.1.228:3000", "acct", "machines").startswith("http://192.168.1.228:3000/"))
        self.assertTrue(app.keygen_url("http://localhost:3000", "acct", "machines").startswith("http://localhost:3000/"))
        with self.assertRaises(ValueError):
            app.keygen_url("http://example.com", "acct", "machines")

    def test_keygen_request_marks_lan_http_as_forwarded_https(self):
        captured = []

        class FakeResponse:
            def __enter__(self):
                return self

            def __exit__(self, exc_type, exc, tb):
                return False

            def read(self):
                return b"{}"

        def fake_urlopen(request_obj, timeout=10):
            captured.append(request_obj)
            return FakeResponse()

        original_urlopen = app.urlopen
        try:
            app.urlopen = fake_urlopen
            app.keygen_request("POST", "http://license-server.local:3000/v1/accounts/acct/licenses/actions/validate-key", {})
        finally:
            app.urlopen = original_urlopen

        self.assertEqual(captured[0].get_header("X-forwarded-proto"), "https")

    def test_keygen_activation_creates_machine_then_revalidates(self):
        calls = []

        def fake_request(method, url, payload=None, license_key=None, timeout=10):
            calls.append((method, url, payload, license_key))
            if url.endswith("licenses/actions/validate-key") and len(calls) == 1:
                return {
                    "meta": {"valid": False, "code": "NO_MACHINE", "detail": "machine required"},
                    "data": {"id": "lic-1", "attributes": {"metadata": {"allowed_profiles": ["cao_thanh"]}}},
                }
            if url.endswith("machines"):
                self.assertEqual(payload["data"]["relationships"]["license"]["data"], {"type": "licenses", "id": "lic-1"})
                return {"data": {"id": "machine-1"}}
            return {
                "meta": {"valid": True, "code": "VALID"},
                "data": {"id": "lic-1", "attributes": {"status": "ACTIVE", "metadata": {"allowed_profiles": ["cao_thanh"]}}},
            }

        original_request = app.keygen_request
        original_fingerprint = app.local_machine_fingerprint
        try:
            app.keygen_request = fake_request
            app.local_machine_fingerprint = lambda: "TEST-FINGERPRINT"
            license_cfg = app.activate_keygen_license("https://license-server.local", "acct", "KEY")
        finally:
            app.keygen_request = original_request
            app.local_machine_fingerprint = original_fingerprint

        self.assertEqual(license_cfg["machine_id"], "machine-1")
        self.assertEqual(license_cfg["allowed_profiles"], ["cao_thanh"])
        self.assertEqual([call[0] for call in calls], ["POST", "POST", "POST"])

    def test_vietmax_product_code_matches_workbook_examples(self):
        examples = {
            "Bản nhôm CTP (D-L) 400x530x0.3 (50 tấm) 2 lớp": "BNCTP400x530x0.3",
            "Giấy Cacbon CF yellow 56/650*860_TL (R500)": "GCCFYELLOW56650*860_TL(R500)",
            "Màng tự dính SYNWK-F1840N": "MANGTUDINH",
            "Mực in - TK Mark V T Cyan (VN) - 2Kg (xanh)": "MUCINXANH",
            "Véc ni phủ bóng bề mặt OP Varnish (new)": "VECNI",
        }
        for product_name, expected in examples.items():
            with self.subTest(product_name=product_name):
                self.assertEqual(make_product_part("vietmax_mua_vao", product_name, {}), expected)

    def test_vietmax_product_code_falls_back_for_unknown_products(self):
        self.assertEqual(
            make_product_part("vietmax_mua_vao", "Pin SYNWK-F1840N dung lượng", {}),
            "PINSYNWK-F1840NDULU",
        )

    def test_vietmax_ban_ra_fallback_uses_first_word_then_initials(self):
        self.assertEqual(
            make_product_part("vietmax_ban_ra", "Pin SYNWK-F1840N 500ML dung lượng", {}),
            "PINSYNWK-F1840N500MLDL",
        )
        self.assertEqual(
            make_product_part("vietmax_ban_ra", "Giấy An Hòa PP 92/70", {}),
            "GIAYAHPP92/70",
        )
        self.assertEqual(
            make_product_part("vietmax_ban_ra", "Giấy in offset 100gsm", {}),
            "GIAYIO100GSM",
        )
        self.assertEqual(
            make_product_part("vietmax_ban_ra", "Màng tự dính SYNWK-F1840N", {}),
            "MANGTDSYNWK-F1840N",
        )
        self.assertEqual(
            make_product_part("vietmax_ban_ra", "Supercalifragilisticexpialidocious alpha beta gamma delta epsilon", {}),
            "SUPERCALIFRAGILISTICEXPIALIDOCIOUSABGDE"[:50],
        )

    def test_vietmax_product_code_uses_all_words_for_note_patterns(self):
        examples = {
            "Giấy Couche": "GIAYCOUCHE",
            "Giấy offset": "GIAYOFFSET",
            "Giấy Cacbon": "GIAYCACBON",
            "Giấy in BB58/92gsm": "GIAYINBB58/92GSM",
            "Giấy ivory ningbo 230gsm khổ 62cm": "GIVORYNINGBO230GSM62CM",
        }
        for product_name, expected in examples.items():
            with self.subTest(product_name=product_name):
                self.assertEqual(make_product_part("vietmax_mua_vao", product_name, {}), expected)

    def test_vietmax_product_code_uses_workbook_prefix_rules_for_variants(self):
        examples = {
            "Giấy An Hoà 95/60gsm/620x860": "GAH9560GSM620x860",
            "Giấy An Hoà 95/70gsm/620x860": "GAH9570GSM620x860",
            "Giấy In An Hòa 92/70gsm/790 x 1090": "GIAYINANHO9270GSM790X1090",
            "Giấy Couche 250gsm (79x47) cm": "GCOUCHE250GSM79X47CM",
            "Giấy Cacbon CB white 56/610*860_TL (R500)": "GCCBWHITE56610*860_TL(R500)",
            "Giấy offset 100gsm khổ 62x86cm": "GOFFSET100GSM62X86CM",
            "Giấy In Offset 120gsm khổ 79x109cm": "GIAYINOFFSET120GSMKHO79X109CM",
            "Giấy Couche định lượng 80gsm, khổ 620x860mm": "GCOUCHE80GSM620x860MM",
            "Giấy Ivory định lượng 250gsm, khổ 790 mm": "GIVORY250GSM790MM",
            "Màng tự dính BLWK-Z0585MW": "MANGTUDINH",
            "Tấm bản in bằng nhôm CTP SR 530x400": "TAMBANNHOMCTPSR530x400",
        }
        for product_name, expected in examples.items():
            with self.subTest(product_name=product_name):
                self.assertEqual(make_product_part("vietmax_mua_vao", product_name, {}), expected)

    def test_vietmax_product_code_keeps_hl_prefix_marker(self):
        self.assertEqual(
            make_product_part("vietmax_mua_vao", "Tấm bản in bằng nhôm CTP HL 1030x800", {}),
            "TAMBANNHOMCTPHL1030x800",
        )

    def test_vietmax_tam_ban_nhom_hyphen_variants_use_base_prefix(self):
        examples = {
            "T\u1ea5m b\u1ea3n in b\u1eb1ng nh\u00f4m CTP-HL I 1030x800": (
                "TAMBANNHOMCTP-HLI1030x800",
                "TAMBIBNCTP-HLI1030x800",
            ),
            "T\u1ea5m b\u1ea3n in b\u1eb1ng nh\u00f4m CTCP-HL I530x400x0.3mm": (
                "TAMBANNHOMCTCP-HLI530x400x0.3MM",
                "TAMBIBNCTCP-HLI530x400x0.3MM",
            ),
        }
        for product_name, (purchase_expected, sales_expected) in examples.items():
            with self.subTest(product_name=product_name):
                self.assertEqual(make_product_part("vietmax_mua_vao", product_name, {}), purchase_expected)
                self.assertEqual(make_product_part("vietmax_ban_ra", product_name, {}), sales_expected)

    def test_vietmax_mang_tu_dinh_keeps_customer_hardcode(self):
        self.assertEqual(make_product_part("vietmax_mua_vao", "Màng tự dính BLWK-Z0585MW", {}), "MANGTUDINH")
        self.assertEqual(make_product_part("vietmax_mua_vao", "Màng tự dính SYNWK-F1840N", {}), "MANGTUDINH")
        self.assertEqual(make_product_part("vietmax_ban_ra", "Màng tự dính SYNWK-F1840N", {}), "MANGTDSYNWK-F1840N")

    def test_vietmax_duplex_g_m2_uses_single_uppercase_gm2(self):
        self.assertEqual(make_product_part("vietmax_mua_vao", "Giấy Duplex ĐL 300 g/m2", {}), "GDL300GM2")
        self.assertEqual(make_product_part("vietmax_mua_vao", "Giấy Duplex ĐL 400 g/m2", {}), "GDL400GM2")

    def test_vietmax_product_code_honors_custom_word_rules(self):
        self.assertEqual(
            make_product_part("vietmax_mua_vao", "Mực pantone 2009 C", {"pantone": "PANTONE"}),
            "MUCPANTONE2009C",
        )

    def test_vietmax_product_code_normalizes_formatting_only_tokens(self):
        self.assertEqual(
            make_product_part("vietmax_mua_vao", "Giấy couche hikote 120gsm khổ62x86cm", {}),
            make_product_part("vietmax_mua_vao", "Giấy Couche Hikote 120gsm khổ 62x86cm", {}),
        )
        self.assertEqual(
            make_product_part("vietmax_mua_vao", "Mực pantone2009C", {}),
            make_product_part("vietmax_mua_vao", "Mực pantone 2009 C", {}),
        )

    def test_vietmax_ban_ra_purchase_match_accepts_near_one_character_difference(self):
        matches = build_vietmax_ban_ra_purchase_matches(
            ["Giấy Couche 300 gsm"],
            [{"purchase_product": "Giấy Couche 300 gms", "purchase_code": "GC300", "purchase_row": 5}],
        )

        self.assertEqual(len(matches), 1)
        self.assertEqual(matches[0]["purchase_code"], "GC300")
        self.assertGreaterEqual(matches[0]["score"], 0.92)

    def test_vietmax_sales_products_include_invoice_metadata(self):
        outputs = Path(__file__).parent / "outputs"
        outputs.mkdir(exist_ok=True)
        source = outputs / f"_test_vietmax_sales_{uuid4().hex}.xlsx"
        try:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Invoices"
            sheet.cell(2, 3, "Số HD")
            sheet.cell(2, 4, "Ngày có hàng bán ra")
            sheet.cell(2, 13, "Tên hàng")
            sheet.cell(2, 14, "ĐVT")
            sheet.cell(2, 15, "Số lượng")
            sheet.cell(3, 3, "HD-001")
            sheet.cell(3, 4, "06/06/2026")
            sheet.cell(3, 13, "Giấy Couche 300 gsm")
            sheet.cell(3, 14, "Tờ")
            sheet.cell(3, 15, 2)
            sheet.cell(3, 16, 12345)
            workbook.save(source)
            workbook.close()

            products = vietmax_ban_ra_sales_products_from_workbook(source, price_col="P")
            matches = build_vietmax_ban_ra_purchase_matches(
                products,
                [{"purchase_product": "Giấy Couche 300 gsm", "purchase_code": "GC300", "purchase_row": 5, "purchase_unit": "Tờ"}],
            )

            self.assertEqual(products, [{"sales_product": "Giấy Couche 300 gsm", "sales_unit": "Tờ", "sales_price": "12345", "invoice_no": "HD-001", "invoice_date": "06/06/2026"}])
            self.assertEqual(matches[0]["invoice_no"], "HD-001")
            self.assertEqual(matches[0]["invoice_date"], "06/06/2026")
            self.assertEqual(matches[0]["sales_unit"], "Tờ")
            self.assertEqual(matches[0]["sales_price"], "12345")
            self.assertFalse(matches[0]["unit_mismatch"])
        finally:
            source.unlink(missing_ok=True)

    def test_vietmax_purchase_match_prefers_duplicate_rows_with_prices(self):
        sales_products = [
            {"sales_product": "Giấy Couche", "sales_unit": "Tờ", "sales_price": "", "invoice_no": "HD-001", "invoice_date": "06/06/2026"},
            {"sales_product": "Giấy Couche", "sales_unit": "Tờ", "sales_price": "10500", "invoice_no": "HD-002", "invoice_date": "07/06/2026"},
        ]
        purchase_products = [
            {"purchase_product": "Giấy Couche", "purchase_code": "GIAYCOUCHE", "purchase_row": 5, "purchase_unit": "Kg", "purchase_price": ""},
            {"purchase_product": "Giấy Couche", "purchase_code": "GIAYCOUCHE", "purchase_row": 6, "purchase_unit": "Kg", "purchase_price": "21000"},
        ]

        matches = build_vietmax_ban_ra_purchase_matches(sales_products, purchase_products)

        self.assertEqual(len(matches), 1)
        self.assertEqual(matches[0]["sales_price"], "10500")
        self.assertEqual(matches[0]["purchase_price"], "21000")
        self.assertEqual(matches[0]["invoice_no"], "HD-002")

    def test_vietmax_exact_khh_match_keeps_sales_and_purchase_prices(self):
        matches = app.build_vietmax_khh_exact_purchase_matches(
            [{"sales_product": "Giấy Couche", "sales_unit": "Kg", "sales_price": "21,000", "invoice_no": "HD-001"}],
            [{"purchase_product": "Giấy Couche", "purchase_code": "GC", "purchase_unit": "kg", "purchase_price": "19,500"}],
        )

        self.assertEqual(len(matches), 1)
        self.assertEqual(matches[0]["sales_price"], "21,000")
        self.assertEqual(matches[0]["purchase_price"], "19,500")

    def test_vietmax_purchase_match_export_rows_match_display_columns(self):
        rows = vietmax_purchase_match_export_rows([
            {
                "confirmed": False,
                "sales_product": "Giấy Couche 300 gsm",
                "sales_unit": "Tờ",
                "invoice_no": "HD-001",
                "invoice_date": "06/06/2026",
                "purchase_code": "GC300",
                "purchase_product": "Giấy Couche 300 gms",
                "purchase_unit": "Ram",
                "unit_mismatch": True,
                "unit_warning": "Khác đơn vị tính",
                "conversion_mode": app.VIETMAX_CONVERSION_MODE_QTY_AND_UNIT,
                "conversion_formula": "1 ram = 500 tờ",
                "purchase_row": 5,
                "score": 0.925,
            }
        ])

        self.assertEqual(rows[0], ["Dùng", "Hàng bán ra", "ĐVT bán ra", "Số HD", "Ngày có hàng bán ra", "Mã VT mua vào", "Hàng mua vào", "ĐVT mua vào", "Cảnh báo", "Quy đổi", "Khác biệt", "Độ giống"])
        self.assertEqual(rows[1], ["Không", "Giấy Couche 300 gsm", "Tờ", "HD-001", "06/06/2026", "GC300", "Giấy Couche 300 gms", "Ram", "Khác đơn vị tính", "1 ram = 500 tờ", "Giấy Couche 300 gsm -> Giấy Couche 300 gms", "92.5%"])

    def test_vietmax_purchase_match_flags_unit_mismatch(self):
        matches = build_vietmax_ban_ra_purchase_matches(
            [{"sales_product": "Giấy Couche 300 gsm", "sales_unit": "Tờ"}],
            [{"purchase_product": "Giấy Couche 300 gsm", "purchase_code": "GC300", "purchase_row": 5, "purchase_unit": "Ram"}],
        )

        self.assertEqual(len(matches), 1)
        self.assertEqual(matches[0]["sales_unit"], "Tờ")
        self.assertEqual(matches[0]["purchase_unit"], "Ram")
        self.assertTrue(matches[0]["unit_mismatch"])
        self.assertEqual(matches[0]["unit_warning"], "Khác đơn vị tính")
        self.assertEqual(matches[0]["conversion_formula"], "")

    def test_vietmax_product_review_rows_show_near_matches_not_exact_duplicates(self):
        rows = vietmax_product_review_rows(
            [
                {"purchase_product": "Giấy Couche 300 gsm", "purchase_unit": "Tờ", "invoice_no": "MV-001", "invoice_date": "01/06/2026"},
                {"purchase_product": "Giấy Couche 300 gms", "purchase_unit": "Ram", "invoice_no": "MV-002", "invoice_date": "02/06/2026"},
                {"purchase_product": "Giấy Couche 300 gsm", "purchase_unit": "Tờ", "invoice_no": "MV-003", "invoice_date": "03/06/2026"},
                {"purchase_product": "Mực in xanh", "purchase_unit": "Hộp"},
            ],
            "purchase_product",
            "purchase_unit",
        )

        self.assertEqual([row["product"] for row in rows], ["Giấy Couche 300 gms"])
        self.assertEqual(rows[0]["invoice_no"], "MV-002")
        self.assertEqual(rows[0]["invoice_date"], "02/06/2026")
        self.assertEqual(rows[0]["similar_product"], "Giấy Couche 300 gsm")
        self.assertEqual(rows[0]["similar_invoice_no"], "MV-001")
        self.assertEqual(rows[0]["similar_invoice_date"], "01/06/2026")
        self.assertNotIn("Mực in xanh", [row["product"] for row in rows])

    def test_vietmax_product_review_rows_ignores_formatting_only_differences(self):
        rows = vietmax_product_review_rows(
            [
                {"purchase_product": "Giấy couche hikote 120gsm khổ62x86cm", "purchase_unit": "Kg"},
                {"purchase_product": "Giấy Couche Hikote 120gsm khổ 62 x 86cm", "purchase_unit": "Kg"},
                {"purchase_product": "Mực pantone2009C", "purchase_unit": "kg"},
                {"purchase_product": "Mực pantone 2009 C", "purchase_unit": "kg"},
            ],
            "purchase_product",
            "purchase_unit",
        )

        self.assertEqual(rows, [])

    def test_vietmax_product_review_rows_ignores_case_accent_space_only_differences(self):
        products = [
            "Gi\u1ea5y in",
            "Gi\u1ea5y In",
            "GI\u1ea4Y   IN",
            "Giay in",
        ]
        rows = vietmax_product_review_rows(
            [{"purchase_product": value, "purchase_unit": "Kg"} for value in products],
            "purchase_product",
            "purchase_unit",
        )

        self.assertEqual(rows, [])
        codes = {
            app.make_code("", value, 1, {}, "vietmax_mua_vao", {}, require_qty=False, include_company_prefix=False)
            for value in products
        }
        self.assertEqual(len(codes), 1)

    def test_vietmax_product_review_rows_groups_volume_unit_spelling_differences(self):
        left = "In Decal N\u01b0\u1edbc gi\u1eb7t Hi5 Ng\u00e0n hoa 10L"
        right = "In decal n\u01b0\u1edbc gi\u1eb7t hi5 Ng\u00e0n hoa 10 lit"

        purchase_rows = vietmax_product_review_rows(
            [
                {"purchase_product": left, "purchase_unit": "B\u1ed9"},
                {"purchase_product": right, "purchase_unit": "B\u1ed9"},
            ],
            "purchase_product",
            "purchase_unit",
        )
        sales_rows = vietmax_product_review_rows(
            [
                {"sales_product": left, "sales_unit": "B\u1ed9", "code": "INDNGHI5NH10L"},
                {"sales_product": right, "sales_unit": "B\u1ed9", "code": "INDNGHI5NH10L"},
            ],
            "sales_product",
            "sales_unit",
            allow_same_code_split=True,
        )

        self.assertEqual(len(purchase_rows), 1)
        self.assertEqual(purchase_rows[0]["review_group"], "unit_spelling_diff")
        self.assertEqual(len(sales_rows), 1)
        self.assertEqual(sales_rows[0]["review_group"], "unit_spelling_diff")

    def test_vietmax_code_generation_normalizes_decimal_comma_and_dot(self):
        codes = {
            app.make_code("", value, 1, {}, "vietmax_mua_vao", {}, require_qty=False, include_company_prefix=False)
            for value in ["In tem chai 3.5L", "In tem chai 3,5L"]
        }

        self.assertEqual(len(codes), 1)

    def test_vietmax_product_review_rows_groups_optional_in_form_variants(self):
        rows = vietmax_product_review_rows(
            [
                {"purchase_product": "Gi\u1ea5y in Couche", "purchase_unit": "Kg", "code": "GIAYINCOUCHE"},
                {"purchase_product": "Gi\u1ea5y Couche", "purchase_unit": "Kg", "code": "GIAYCOUCHE"},
                {"purchase_product": "Gi\u1ea5y In Duplex", "purchase_unit": "Kg", "code": "GIAYINDUPLEX"},
                {"purchase_product": "Gi\u1ea5y Duplex", "purchase_unit": "Kg", "code": "GIAYDUPLEX"},
            ],
            "purchase_product",
            "purchase_unit",
        )

        similar_rows = [row for row in rows if row["review_group"] == "similar_form"]
        pairs = {tuple(sorted([row["product"], row["similar_product"]])) for row in similar_rows}
        self.assertIn(tuple(sorted(["Gi\u1ea5y in Couche", "Gi\u1ea5y Couche"])), pairs)
        self.assertIn(tuple(sorted(["Gi\u1ea5y In Duplex", "Gi\u1ea5y Duplex"])), pairs)
        self.assertTrue(all(row["confirmed"] is False for row in similar_rows))

    def test_vietmax_product_review_rows_marks_same_code_split_only_when_enabled(self):
        products = [
            {"sales_product": "Giay in Offset", "sales_unit": "Kg", "code": "GIAYOFFSET"},
            {"sales_product": "Giay Offset", "sales_unit": "Kg", "code": "GIAYOFFSET"},
        ]

        default_rows = vietmax_product_review_rows(products, "sales_product", "sales_unit")
        split_rows = vietmax_product_review_rows(products, "sales_product", "sales_unit", allow_same_code_split=True)

        self.assertFalse(any(row.get("review_type") == "same_code_split" for row in default_rows))
        self.assertTrue(any(row.get("review_group") == "same_code_split" and row.get("review_type") == "same_code_split" for row in split_rows))

    def test_vietmax_product_review_rows_skips_same_code_from_user_config(self):
        rows = vietmax_product_review_rows(
            [
                {"purchase_product": "Giay in Couche", "purchase_unit": "Kg", "code": "GIAYINCOUCHE"},
                {"purchase_product": "Giay Couche", "purchase_unit": "Kg", "code": "GIAYINCOUCHE", "code_from_user_config": True},
                {"purchase_product": "Giay in Duplex", "purchase_unit": "Kg", "code": "GIAYINDUPLEX"},
                {"purchase_product": "Giay Duplex", "purchase_unit": "Kg", "code": "GIAYDUPLEX"},
            ],
            "purchase_product",
            "purchase_unit",
        )

        pairs = {tuple(sorted([row["product"], row["similar_product"]])) for row in rows}
        self.assertNotIn(tuple(sorted(["Giay in Couche", "Giay Couche"])), pairs)
        self.assertIn(tuple(sorted(["Giay in Duplex", "Giay Duplex"])), pairs)

    def test_review_merge_config_strips_normal_codes_but_keeps_split_codes(self):
        rules = app.sanitize_review_merge_rules_for_config([
            {
                "product": "Giay in Offset",
                "similar_product": "Giay Offset",
                "confirmed": True,
                "code_choice": "similar",
                "code": "STALE1",
                "similar_code": "STALE2",
                "review_group": "similar_form",
            },
            {
                "product": "Giay in Couche",
                "similar_product": "Giay Couche",
                "confirmed": True,
                "code_choice": "split",
                "split_code": "GIAYINCOUCHE(1)",
                "similar_split_code": "GIAYCOUCHE/2",
                "review_group": "same_code_split",
                "review_type": "same_code_split",
            },
        ])

        self.assertNotIn("code", rules[0])
        self.assertNotIn("similar_code", rules[0])
        self.assertEqual(rules[1]["split_code"], "GIAYINCOUCHE1")
        self.assertEqual(rules[1]["similar_split_code"], "GIAYCOUCHE2")

    def test_sanitize_product_code_removes_fast_illegal_characters(self):
        self.assertEqual(app.sanitize_product_code('TT173.INTG(ABC)/A+B? "X"'), "TT173.INTGABCABX")
        self.assertNotIn("(", app.sanitize_product_code("BNCTP(ABC)"))
        self.assertNotIn("/", app.sanitize_product_code("BNCTP/ABC"))

    def test_vietmax_split_review_rule_applies_without_manual_code_overrides(self):
        outputs = Path(__file__).parent / "outputs"
        outputs.mkdir(exist_ok=True)
        run_id = uuid4().hex
        source = outputs / f"_test_vietmax_split_source_{run_id}.xlsx"
        result = outputs / f"_test_vietmax_split_result_{run_id}.xlsx"
        try:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Purchase"
            sheet.cell(2, 6, "Ten nguoi ban")
            sheet.cell(2, 7, "MST")
            sheet.cell(2, 12, "Ma VT")
            sheet.cell(2, 13, "Ten hang")
            sheet.cell(2, 15, "So luong")
            sheet.cell(3, 6, "Cong ty A")
            sheet.cell(3, 7, "111")
            sheet.cell(3, 12, 0)
            sheet.cell(3, 13, "In decal Tumax")
            sheet.cell(3, 15, 1)
            sheet.cell(4, 6, "Cong ty A")
            sheet.cell(4, 7, "111")
            sheet.cell(4, 12, 0)
            sheet.cell(4, 13, "In decal Fly-pro")
            sheet.cell(4, 15, 1)
            workbook.save(source)
            workbook.close()

            process_workbook(source, result, {
                "company_col": "F",
                "mst_col": "G",
                "product_col": "M",
                "qty_col": "O",
                "output_col": "L",
                "profile": "vietmax",
                "vietmax_phase": app.VIETMAX_PHASE_PURCHASE,
                "include_company_prefix": False,
                "all_mst": ["111"],
                "process_mst": ["111"],
                "mst_safe_id": ["111|||0"],
                "selected_products_0": ["In decal Tumax", "In decal Fly-pro"],
                "vietmax_mua_vao_internal_merges": [
                    {
                        "product": "In decal Tumax",
                        "similar_product": "In decal Fly-pro",
                        "confirmed": True,
                        "code_choice": "split",
                        "review_group": "same_code_split",
                        "review_type": "same_code_split",
                        "split_code": "INDT(1)",
                        "similar_split_code": "INDF/2",
                    },
                ],
            })

            output = load_workbook(result, data_only=True)
            try:
                result_sheet = output["Purchase"]
                self.assertEqual(result_sheet.cell(3, 12).value, "INDT1")
                self.assertEqual(result_sheet.cell(4, 12).value, "INDF2")
            finally:
                output.close()
        finally:
            source.unlink(missing_ok=True)
            result.unlink(missing_ok=True)

    def test_saved_review_choices_restore_by_scope_and_section(self):
        from web_api import apply_saved_review_choices

        rows = [{
            "product": "Giay Couche",
            "similar_product": "Giay in Couche",
            "review_group": "similar_form",
            "confirmed": False,
        }]
        wrong_section = [{
            "product": "Giay Couche",
            "similar_product": "Giay in Couche",
            "review_group": "other",
            "comparison_scope": app.VIETMAX_COMPARISON_SCOPE_ALL_COMPANIES,
            "confirmed": True,
            "code_choice": "similar",
        }]
        wrong_scope = [{
            "product": "Giay Couche",
            "similar_product": "Giay in Couche",
            "review_group": "similar_form",
            "comparison_scope": app.VIETMAX_COMPARISON_SCOPE_SAME_COMPANY,
            "confirmed": True,
            "code_choice": "similar",
        }]
        right_rule = [{
            "product": "Giay Couche",
            "similar_product": "Giay in Couche",
            "review_group": "similar_form",
            "comparison_scope": app.VIETMAX_COMPARISON_SCOPE_ALL_COMPANIES,
            "confirmed": True,
            "code_choice": "similar",
        }]

        self.assertFalse(apply_saved_review_choices(rows, wrong_section, app.VIETMAX_COMPARISON_SCOPE_ALL_COMPANIES)[0]["confirmed"])
        self.assertFalse(apply_saved_review_choices(rows, wrong_scope, app.VIETMAX_COMPARISON_SCOPE_ALL_COMPANIES)[0]["confirmed"])
        self.assertTrue(apply_saved_review_choices(rows, right_rule, app.VIETMAX_COMPARISON_SCOPE_ALL_COMPANIES)[0]["confirmed"])

    def test_vietmax_product_review_rows_groups_obvious_size_or_volume_differences(self):
        rows = vietmax_product_review_rows(
            [
                {"purchase_product": "Bản nhôm CTP BOCICA 560x670x0.3 (50 tấm)", "purchase_unit": "Hộp"},
                {"purchase_product": "Bản nhôm CTP BOCICA 600x730x0.3 (50 tấm)", "purchase_unit": "Hộp"},
                {"purchase_product": "In decal Lau sàn Fineday huong ly 3.6 lít", "purchase_unit": "Chai"},
                {"purchase_product": "In decal Lau sàn Fineday huong ly 1.2 lít", "purchase_unit": "Chai"},
                {"purchase_product": "Tem nhãn A5", "purchase_unit": "Tờ"},
                {"purchase_product": "Tem nhãn A4", "purchase_unit": "Tờ"},
            ],
            "purchase_product",
            "purchase_unit",
        )

        self.assertGreaterEqual(len(rows), 2)
        self.assertTrue(all(row["review_group"] == "dimension_diff" for row in rows))
        self.assertTrue(all(row["dimension_only"] is True for row in rows))

    def test_vietmax_product_review_rows_keeps_hl_merge_candidate_separate(self):
        rows = vietmax_product_review_rows(
            [
                {"purchase_product": "Tấm bản in bằng nhôm CTP 1030 x 800", "purchase_unit": "Tấm", "code": "TAMBANNHOMCTP1030x800"},
                {"purchase_product": "Tấm bản in bằng nhôm CTP HL 1030 x 800", "purchase_unit": "Tấm", "code": "TAMBANNHOMCTPHL1030x800"},
            ],
            "purchase_product",
            "purchase_unit",
        )

        self.assertEqual(len(rows), 1)
        self.assertIn("HL", rows[0]["similar_product"] + rows[0]["product"])
        self.assertNotEqual(rows[0].get("code"), rows[0].get("similar_code"))

    def test_vietmax_product_review_rows_forces_special_purchase_pairs_into_other(self):
        rows = vietmax_product_review_rows(
            [
                {"purchase_product": "Giấy An Hòa PP 92/70", "purchase_unit": "Tờ", "code": "GAHPP9270"},
                {"purchase_product": "Giấy An Hòa 92/70gms/790x1090", "purchase_unit": "Tờ", "code": "GAH9270GMS790x1090"},
                {"purchase_product": "Giấy in offset 100gsm", "purchase_unit": "Tờ", "code": "GIAYINOFFSET100GSM"},
                {"purchase_product": "Giấy In Offset", "purchase_unit": "Tờ", "code": "GIAYINOFFSET"},
            ],
            "purchase_product",
            "purchase_unit",
        )

        pairs = {(row["product"], row["similar_product"]): row for row in rows}
        self.assertIn(("Giấy An Hòa PP 92/70", "Giấy An Hòa 92/70gms/790 x 1090"), pairs)
        self.assertIn(("Giấy in offset 100gsm", "Giấy In Offset"), pairs)
        self.assertTrue(all(row["review_group"] == "other" for row in pairs.values()))
        self.assertTrue(all(row["dimension_only"] is False for row in pairs.values()))
        self.assertTrue(all(row.get("forced_review") is True for row in pairs.values()))

    def test_vietmax_product_review_rows_same_company_scope_filters_near_matches(self):
        rows = vietmax_product_review_rows(
            [
                {"sales_product": "Giấy Couche 300 gsm", "sales_company": "Cong ty A", "sales_mst": "111"},
                {"sales_product": "Giấy Couche 300 gms", "sales_company": "Cong ty B", "sales_mst": "222"},
                {"sales_product": "Giấy Couche 300 gms", "sales_company": "Cong ty A", "sales_mst": "111"},
            ],
            "sales_product",
            "sales_unit",
            comparison_scope=app.VIETMAX_COMPARISON_SCOPE_SAME_COMPANY,
        )

        self.assertEqual(len(rows), 1)
        self.assertTrue(all(row["company_key"] == "mst:111" for row in rows))

    def test_vietmax_purchase_internal_merge_updates_processed_product_name(self):
        outputs = Path(__file__).parent / "outputs"
        outputs.mkdir(exist_ok=True)
        run_id = uuid4().hex
        source = outputs / f"_test_vietmax_purchase_merge_source_{run_id}.xlsx"
        result = outputs / f"_test_vietmax_purchase_merge_result_{run_id}.xlsx"
        try:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Purchase"
            sheet.cell(2, 6, "Ten nguoi ban")
            sheet.cell(2, 7, "MST")
            sheet.cell(2, 12, "Ma VT")
            sheet.cell(2, 13, "Ten hang")
            sheet.cell(2, 14, "DVT")
            sheet.cell(2, 15, "So luong")
            sheet.cell(3, 6, "Cong ty A")
            sheet.cell(3, 7, "111")
            sheet.cell(3, 12, 0)
            sheet.cell(3, 13, "Giấy An Hòa PP 92/70")
            sheet.cell(3, 14, "Tờ")
            sheet.cell(3, 15, 1)
            workbook.save(source)
            workbook.close()

            process_workbook(source, result, {
                "company_col": "F",
                "mst_col": "G",
                "product_col": "M",
                "qty_col": "O",
                "output_col": "L",
                "profile": "vietmax",
                "vietmax_phase": app.VIETMAX_PHASE_PURCHASE,
                "include_company_prefix": False,
                "all_mst": ["111"],
                "process_mst": ["111"],
                "mst_safe_id": ["111|||0"],
                "selected_products_0": ["Giấy An Hòa PP 92/70"],
                "vietmax_mua_vao_internal_merges": [
                    {
                        "product": "Giấy An Hòa PP 92/70",
                        "similar_product": "Giấy An Hòa 92/70gms/790x1090",
                        "similar_code": "GAH9270GMS790x1090",
                        "confirmed": True,
                    },
                ],
            })

            output = load_workbook(result, data_only=True)
            try:
                result_sheet = output["Purchase"]
                self.assertEqual(result_sheet.cell(3, 12).value, "GAH9270GMS790x1090")
                self.assertEqual(result_sheet.cell(3, 13).value, "Giấy An Hòa 92/70gms/790x1090")
            finally:
                output.close()
        finally:
            source.unlink(missing_ok=True)
            result.unlink(missing_ok=True)

    def test_vietmax_internal_purchase_merge_affects_cross_file_matching(self):
        purchase_products = [{"purchase_product": "Giấy Couche 300 gms", "purchase_code": "GC300", "purchase_unit": "Ram"}]
        merges = app.normalize_vietmax_internal_merges([
            {"product": "Giấy Couche 300 gms", "similar_product": "Giấy Couche 300 gsm", "confirmed": True},
        ])
        effective_purchase_products = app.apply_vietmax_internal_merges_to_products(
            purchase_products,
            "purchase_product",
            "purchase_unit",
            merges,
        )
        matches = build_vietmax_ban_ra_purchase_matches(
            [{"sales_product": "Giấy Couche 300 gsm", "sales_unit": "Ram"}],
            effective_purchase_products,
        )

        self.assertEqual(len(matches), 1)
        self.assertEqual(matches[0]["purchase_code"], "GCOUCHE300GSM")
        self.assertEqual(matches[0]["purchase_product"], "Giấy Couche 300 gsm")

    def test_vietmax_internal_purchase_merge_uses_representative_code(self):
        purchase_products = [{"purchase_product": "Giấy Couche 300 gms", "purchase_code": "OLD300", "purchase_unit": "Ram"}]
        purchase_products.append({
            "purchase_product": purchase_products[0]["purchase_product"].replace("gms", "gsm"),
            "purchase_code": "COMMON300",
            "purchase_unit": "Ram",
        })
        merges = app.normalize_vietmax_internal_merges([
            {
                "product": "Giấy Couche 300 gms",
                "similar_product": "Giấy Couche 300 gsm",
                "similar_code": "COMMON300",
                "confirmed": True,
            },
        ])
        effective_purchase_products = app.apply_vietmax_internal_merges_to_products(
            purchase_products,
            "purchase_product",
            "purchase_unit",
            merges,
        )
        matches = build_vietmax_ban_ra_purchase_matches(
            [{"sales_product": "Giấy Couche 300 gsm", "sales_unit": "Ram"}],
            effective_purchase_products,
        )

        self.assertEqual(len(matches), 1)
        self.assertEqual(matches[0]["purchase_code"], "COMMON300")
        self.assertEqual(matches[0]["purchase_product"], "Giấy Couche 300 gsm")

    def test_vietmax_similar_form_merge_uses_selected_ma_vt_2(self):
        purchase_products = [
            {"purchase_product": "Giấy In An Hòa 92/70gsm/790 x 1090", "purchase_code": "GIAYINANHO9270GSM790X1090", "purchase_unit": "Ram"},
            {"purchase_product": "Giấy An Hòa 92/70gsm/790x1090", "purchase_code": "GAH9270GSM790x1090", "purchase_unit": "Ram"},
        ]
        merges = app.normalize_vietmax_internal_merges([
            {
                "product": "Giấy In An Hòa 92/70gsm/790 x 1090",
                "similar_product": "Giấy An Hòa 92/70gsm/790x1090",
                "code": "GIAYINANHO9270GSM790X1090",
                "similar_code": "GAH9270GSM790x1090",
                "code_choice": "similar",
                "review_group": "similar_form",
                "confirmed": True,
            },
        ])
        effective_purchase_products = app.apply_vietmax_internal_merges_to_products(
            purchase_products,
            "purchase_product",
            "purchase_unit",
            merges,
        )

        self.assertEqual(len(effective_purchase_products), 1)
        self.assertEqual(effective_purchase_products[0]["purchase_product"], "Giấy An Hòa 92/70gsm/790x1090")
        self.assertEqual(effective_purchase_products[0]["purchase_code"], "GAH9270GSM790x1090")

    def test_vietmax_similar_form_merge_uses_selected_ma_vt_1(self):
        purchase_products = [
            {"purchase_product": "Giấy In An Hòa 92/70gsm/790 x 1090", "purchase_code": "GIAYINANHO9270GSM790X1090", "purchase_unit": "Ram"},
            {"purchase_product": "Giấy An Hòa 92/70gsm/790x1090", "purchase_code": "GAH9270GSM790x1090", "purchase_unit": "Ram"},
        ]
        merges = app.normalize_vietmax_internal_merges([
            {
                # This is the shape saved by the React review table when the user chooses "Mã VT 1".
                "product": "Giấy An Hòa 92/70gsm/790x1090",
                "similar_product": "Giấy In An Hòa 92/70gsm/790 x 1090",
                "code": "GAH9270GSM790x1090",
                "similar_code": "GIAYINANHO9270GSM790X1090",
                "code_choice": "current",
                "review_group": "similar_form",
                "confirmed": True,
            },
        ])
        effective_purchase_products = app.apply_vietmax_internal_merges_to_products(
            purchase_products,
            "purchase_product",
            "purchase_unit",
            merges,
        )

        self.assertEqual(len(effective_purchase_products), 1)
        self.assertEqual(effective_purchase_products[0]["purchase_product"], "Giấy In An Hòa 92/70gsm/790 x 1090")
        self.assertEqual(effective_purchase_products[0]["purchase_code"], "GIAYINANHO9270GSM790X1090")

    def test_unchecked_vietmax_internal_purchase_merge_is_ignored(self):
        purchase_products = [{"purchase_product": "Tên khác hoàn toàn", "purchase_code": "WRONG", "purchase_unit": "Ram"}]
        merges = app.normalize_vietmax_internal_merges([
            {"product": "Tên khác hoàn toàn", "similar_product": "Sản phẩm mục tiêu 123", "confirmed": False},
        ])
        effective_purchase_products = app.apply_vietmax_internal_merges_to_products(
            purchase_products,
            "purchase_product",
            "purchase_unit",
            merges,
        )
        matches = build_vietmax_ban_ra_purchase_matches(
            [{"sales_product": "Sản phẩm mục tiêu 123", "sales_unit": "Ram"}],
            effective_purchase_products,
        )

        self.assertEqual(matches, [])

    def test_vietmax_ban_ra_purchase_match_focus_list_bypasses_threshold(self):
        matches = build_vietmax_ban_ra_purchase_matches(
            ["Cabon 1 liên kích thước A4"],
            [{"purchase_product": "Tên mua vào khác hoàn toàn", "purchase_code": "FOCUS-CODE", "purchase_row": 8}],
        )

        self.assertEqual(len(matches), 1)
        self.assertEqual(matches[0]["purchase_code"], "FOCUS-CODE")
        self.assertTrue(matches[0]["focus"])
        self.assertLess(matches[0]["score"], 0.92)

    def test_vietmax_purchase_match_default_scope_matches_across_companies(self):
        matches = build_vietmax_ban_ra_purchase_matches(
            [{"sales_product": "Giấy Couche 300 gsm", "sales_company": "Cong ty A", "sales_mst": "111"}],
            [{"purchase_product": "Giấy Couche 300 gsm", "purchase_code": "GC300", "purchase_company": "Cong ty B", "purchase_mst": "222"}],
        )

        self.assertEqual(len(matches), 1)
        self.assertEqual(matches[0]["purchase_code"], "GC300")

    def test_vietmax_purchase_match_same_company_scope_filters_candidates(self):
        different_company_matches = build_vietmax_ban_ra_purchase_matches(
            [{"sales_product": "Giấy Couche 300 gsm", "sales_company": "Cong ty A", "sales_mst": "111"}],
            [{"purchase_product": "Giấy Couche 300 gsm", "purchase_code": "WRONG", "purchase_company": "Cong ty B", "purchase_mst": "222"}],
            comparison_scope=app.VIETMAX_COMPARISON_SCOPE_SAME_COMPANY,
        )
        same_mst_matches = build_vietmax_ban_ra_purchase_matches(
            [{"sales_product": "Giấy Couche 300 gsm", "sales_company": "Ten ban ra", "sales_mst": "111"}],
            [{"purchase_product": "Giấy Couche 300 gsm", "purchase_code": "SAME-MST", "purchase_company": "Ten mua vao", "purchase_mst": "111"}],
            comparison_scope=app.VIETMAX_COMPARISON_SCOPE_SAME_COMPANY,
        )
        same_name_matches = build_vietmax_ban_ra_purchase_matches(
            [{"sales_product": "Màng tự dính SYNWK-F1840N", "sales_company": "Công ty In Bao Bì A"}],
            [{"purchase_product": "Màng tự dính SYNWK-F1840N", "purchase_code": "SAME-NAME", "purchase_company": "Cong ty In Bao Bi A"}],
            comparison_scope=app.VIETMAX_COMPARISON_SCOPE_SAME_COMPANY,
        )
        focus_wrong_company_matches = build_vietmax_ban_ra_purchase_matches(
            [{"sales_product": "Cabon 1 liên kích thước A4", "sales_company": "Cong ty A", "sales_mst": "111"}],
            [{"purchase_product": "Tên mua vào khác hoàn toàn", "purchase_code": "FOCUS-WRONG", "purchase_company": "Cong ty B", "purchase_mst": "222"}],
            comparison_scope=app.VIETMAX_COMPARISON_SCOPE_SAME_COMPANY,
        )

        self.assertEqual(different_company_matches, [])
        self.assertEqual(same_mst_matches[0]["purchase_code"], "SAME-MST")
        self.assertEqual(same_name_matches[0]["purchase_code"], "SAME-NAME")
        self.assertEqual(focus_wrong_company_matches, [])

    def test_vietmax_ban_ra_purchase_match_skips_print_service_rows(self):
        matches = build_vietmax_ban_ra_purchase_matches(
            ["In tờ rơi A4 giấy coucher", "Công in giấy couche", "Gia công in giấy Couche", "Giấy in couche"],
            [{"purchase_product": "Giấy in Couche", "purchase_code": "GIAYINCOUCHE", "purchase_row": 5}],
        )

        self.assertEqual([item["sales_product"] for item in matches], ["Giấy in couche"])
        self.assertEqual(matches[0]["purchase_code"], "GIAYINCOUCHE")

    def test_vietmax_purchase_products_generate_code_when_purchase_code_is_zero(self):
        outputs = Path(__file__).parent / "outputs"
        outputs.mkdir(exist_ok=True)
        source = outputs / f"_test_vietmax_purchase_{uuid4().hex}.xlsx"
        try:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Invoices"
            sheet.cell(2, 3, "Số HD")
            sheet.cell(2, 4, "Ngày HD")
            sheet.cell(2, 12, "Mã VT")
            sheet.cell(2, 13, "Tên hàng hóa, dịch vụ")
            sheet.cell(2, 14, "ĐVT")
            sheet.cell(3, 3, "MV-001")
            sheet.cell(3, 4, "06/06/2026")
            sheet.cell(3, 12, 0)
            sheet.cell(3, 13, "Giấy Couche")
            sheet.cell(3, 14, "Ram")
            sheet.cell(3, 16, 67890)
            workbook.save(source)
            workbook.close()

            products = vietmax_purchase_products_from_workbook(source)

            self.assertEqual(products, [{"purchase_product": "Giấy Couche", "purchase_code": "GIAYCOUCHE", "purchase_row": 3, "purchase_unit": "Ram", "purchase_price": "67890", "invoice_no": "MV-001", "invoice_date": "06/06/2026", "purchase_invoice_no": "MV-001", "purchase_invoice_date": "06/06/2026"}])
        finally:
            source.unlink(missing_ok=True)

    def test_vietmax_purchase_products_read_price_by_header(self):
        outputs = Path(__file__).parent / "outputs"
        outputs.mkdir(exist_ok=True)
        source = outputs / f"_test_vietmax_purchase_price_{uuid4().hex}.xlsx"
        try:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Invoices"
            sheet.cell(2, 3, "Số HD")
            sheet.cell(2, 4, "Ngày HD")
            sheet.cell(2, 12, "Mã VT")
            sheet.cell(2, 13, "Tên hàng hóa, dịch vụ")
            sheet.cell(2, 14, "ĐVT")
            sheet.cell(2, 18, "Ðon giá")
            sheet.cell(3, 3, "MV-002")
            sheet.cell(3, 4, "07/06/2026")
            sheet.cell(3, 12, "GC300")
            sheet.cell(3, 13, "Giấy Couche 300 gsm")
            sheet.cell(3, 14, "Ram")
            sheet.cell(3, 16, 11111)
            sheet.cell(3, 18, 67890)
            workbook.save(source)
            workbook.close()

            products = vietmax_purchase_products_from_workbook(source)

            self.assertEqual(products[0]["purchase_price"], "67890")
        finally:
            source.unlink(missing_ok=True)

    def test_vietmax_purchase_products_read_selected_price_column(self):
        outputs = Path(__file__).parent / "outputs"
        outputs.mkdir(exist_ok=True)
        source = outputs / f"_test_vietmax_purchase_selected_price_{uuid4().hex}.xlsx"
        try:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Invoices"
            sheet.cell(2, 3, "Số HD")
            sheet.cell(2, 4, "Ngày HD")
            sheet.cell(2, 12, "Mã VT")
            sheet.cell(2, 13, "Tên hàng hóa, dịch vụ")
            sheet.cell(2, 14, "ĐVT")
            sheet.cell(2, 16, "Ðon giá")
            sheet.cell(3, 3, "MV-003")
            sheet.cell(3, 4, "08/06/2026")
            sheet.cell(3, 12, "GC300")
            sheet.cell(3, 13, "Giấy Couche 300 gsm")
            sheet.cell(3, 14, "Ram")
            sheet.cell(3, 16, 11111)
            sheet.cell(3, 18, 67890)
            workbook.save(source)
            workbook.close()

            products = vietmax_purchase_products_from_workbook(source, price_col="R")

            self.assertEqual(products[0]["purchase_price"], "67890")
        finally:
            source.unlink(missing_ok=True)

    def test_removes_invoice_rows_that_are_not_assigned_a_processed_code(self):
        outputs = Path(__file__).parent / "outputs"
        outputs.mkdir(exist_ok=True)
        run_id = uuid4().hex
        source = outputs / f"_test_process_source_{run_id}.xlsx"
        result = outputs / f"_test_process_result_{run_id}.xlsx"
        try:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Invoices"
            sheet.cell(2, 13, "Ma VT")
            sheet.cell(2, 14, "Ten hang")
            sheet.cell(2, 16, "So luong")
            sheet.cell(2, 18, "Don gia")
            sheet.cell(3, 3, "HD-BO-QUA")
            sheet.cell(3, 13, 0)
            sheet.cell(4, 3, "HD-XU-LY")
            sheet.cell(4, 6, "Cong ty A")
            sheet.cell(4, 7, "MST")
            sheet.cell(4, 13, 0)
            sheet.cell(4, 14, "Hang A")
            sheet.cell(4, 16, 2)
            sheet.cell(4, 18, 100)
            workbook.save(source)
            workbook.close()

            process_workbook(source, result, {
                "company_col": "F",
                "mst_col": "G",
                "product_col": "N",
                "qty_col": "P",
                "price_col": "R",
                "output_col": "M",
                "profile": "cao_thanh",
                "include_company_prefix": False,
                "all_mst": ["MST"],
                "process_mst": ["MST"],
                "mst_safe_id": ["MST|||0"],
                "selected_products_0": ["Hang A"],
            })

            output = load_workbook(result, data_only=True)
            try:
                result_sheet = output["Invoices"]
                self.assertEqual(result_sheet.max_row, 3)
                self.assertEqual(result_sheet.cell(3, 3).value, "HD-XU-LY")
                self.assertNotIn(result_sheet.cell(3, 13).value, (None, "", 0, "0"))
            finally:
                output.close()
        finally:
            source.unlink(missing_ok=True)
            result.unlink(missing_ok=True)


    def inventory_base_payload(self, selected_products):
        return {
            "company_col": "F",
            "mst_col": "G",
            "product_col": "N",
            "qty_col": "P",
            "output_col": "M",
            "profile": "son_phuong",
            "include_company_prefix": False,
            "all_mst": ["MST"],
            "process_mst": ["MST"],
            "mst_safe_id": ["MST|||0"],
            "selected_products_0": selected_products,
        }

    def run_inventory_case(self, rows, payload_updates=None, existing_inventory_headers=False):
        outputs = Path(__file__).parent / "outputs"
        outputs.mkdir(exist_ok=True)
        run_id = uuid4().hex
        source = outputs / f"_test_inventory_source_{run_id}.xlsx"
        result = outputs / f"_test_inventory_result_{run_id}.xlsx"
        try:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Invoices"
            sheet.cell(2, 13, "Ma VT")
            sheet.cell(2, 14, "Ten hang")
            sheet.cell(2, 15, "Don vi")
            sheet.cell(2, 16, "So luong")
            if existing_inventory_headers:
                sheet.cell(2, 17, "TK vat tu")
                sheet.cell(2, 18, "Ma kho")
                sheet.cell(3, 17, "OLD-TK")
                sheet.cell(3, 18, "OLD-KHO")
            for offset, row in enumerate(rows, start=3):
                sheet.cell(offset, 3, f"HD-{offset}")
                sheet.cell(offset, 6, "Cong ty A")
                sheet.cell(offset, 7, "MST")
                sheet.cell(offset, 13, 0)
                sheet.cell(offset, 14, row["product"])
                sheet.cell(offset, 15, row.get("unit", "Cai"))
                sheet.cell(offset, 16, row.get("qty", 1))
            workbook.save(source)
            workbook.close()

            payload = self.inventory_base_payload([row["product"] for row in rows])
            payload.update(payload_updates or {})
            process_workbook(source, result, payload)

            output = load_workbook(result, data_only=True)
            try:
                result_sheet = output["Invoices"]
                values = []
                for row_index in range(3, 3 + len(rows)):
                    values.append((result_sheet.cell(row_index, 17).value, result_sheet.cell(row_index, 18).value))
                headers = (result_sheet.cell(2, 17).value, result_sheet.cell(2, 18).value)
                return headers, values, result_sheet.max_column
            finally:
                output.close()
        finally:
            source.unlink(missing_ok=True)
            result.unlink(missing_ok=True)

    def test_inventory_default_pair_updates_existing_columns(self):
        headers, values, max_column = self.run_inventory_case(
            [{"product": "Hang A"}],
            {
                "inventory_pairs": [{"id": "default", "ma_kho": "KHO-1", "tk_vat_tu": "1561"}],
                "use_default_inventory_pair": True,
                "default_inventory_pair_id": "default",
            },
            existing_inventory_headers=True,
        )

        self.assertEqual(headers, ("TK vật tư", "Mã kho"))
        self.assertEqual(values, [("1561", "KHO-1")])
        self.assertEqual(max_column, 18)

    def test_inventory_rule_priority_wins_over_broad_rule(self):
        headers, values, _ = self.run_inventory_case(
            [{"product": "Gia cong in va hoan thien"}, {"product": "In tui giay"}],
            {
                "inventory_pairs": [
                    {"id": "ktp", "ma_kho": "KTP", "tk_vat_tu": "1551"},
                    {"id": "kgci", "ma_kho": "KGCI", "tk_vat_tu": "1552"},
                ],
                "inventory_pair_rules": [
                    {"source_col": "N", "operator": "contains", "value": "In", "pair_id": "ktp", "enabled": True, "priority": 1},
                    {"source_col": "N", "operator": "contains", "value": "Gia cong in", "pair_id": "kgci", "enabled": True, "priority": 10},
                ],
            },
        )

        self.assertEqual(headers, ("TK vật tư", "Mã kho"))
        self.assertEqual(values, [("1552", "KGCI"), ("1551", "KTP")])

    def test_inventory_unit_column_equals_rule(self):
        _, values, _ = self.run_inventory_case(
            [{"product": "Hang Kg", "unit": " kg "}, {"product": "Hang Cai", "unit": "Cai"}],
            {
                "inventory_pairs": [
                    {"id": "kg", "ma_kho": "KHO-KG", "tk_vat_tu": "152-KG"},
                    {"id": "other", "ma_kho": "KHO-OTHER", "tk_vat_tu": "152-OTHER"},
                ],
                "inventory_pair_rules": [
                    {"source_col": "O", "operator": "equals", "value": "KG", "pair_id": "kg", "enabled": True},
                ],
            },
        )

        self.assertEqual(values, [("152-KG", "KHO-KG"), (None, None)])

    def test_inventory_appends_missing_headers(self):
        headers, values, max_column = self.run_inventory_case(
            [{"product": "Hang A"}],
            {
                "inventory_pairs": [{"id": "default", "ma_kho": "KHO-APPEND", "tk_vat_tu": "1562"}],
                "use_default_inventory_pair": True,
                "default_inventory_pair_id": "default",
            },
        )

        self.assertEqual(headers, ("TK vật tư", "Mã kho"))
        self.assertEqual(values, [("1562", "KHO-APPEND")])
        self.assertEqual(max_column, 18)

    def test_inventory_single_pair_auto_defaults_all_processed_rows(self):
        headers, values, _ = self.run_inventory_case(
            [{"product": "Hang A"}, {"product": "Hang B"}],
            {
                "inventory_pairs": [{"id": "only", "ma_kho": "KHO-ONLY", "tk_vat_tu": "156-ONLY"}],
                "use_default_inventory_pair": False,
                "default_inventory_pair_id": "",
            },
        )

        self.assertEqual(headers, ("TK vật tư", "Mã kho"))
        self.assertEqual(values, [("156-ONLY", "KHO-ONLY"), ("156-ONLY", "KHO-ONLY")])

    def test_inventory_normalizes_reversed_saved_pair_values(self):
        headers, values, _ = self.run_inventory_case(
            [{"product": "Hang A"}],
            {
                "inventory_pairs": [{"id": "reversed", "ma_kho": "152", "tk_vat_tu": "KVT"}],
                "use_default_inventory_pair": False,
                "default_inventory_pair_id": "",
            },
        )

        self.assertEqual(headers, ("TK vật tư", "Mã kho"))
        self.assertEqual(values, [("152", "KVT")])


    def test_vietmax_ban_ra_purchase_match_converts_mismatched_unit_rows(self):
        outputs = Path(__file__).parent / "outputs"
        outputs.mkdir(exist_ok=True)
        run_id = uuid4().hex
        source = outputs / f"_test_vietmax_convert_source_{run_id}.xlsx"
        result = outputs / f"_test_vietmax_convert_result_{run_id}.xlsx"
        try:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Invoices"
            sheet.cell(2, 9, "Ten nguoi mua")
            sheet.cell(2, 10, "MST nguoi mua")
            sheet.cell(2, 12, "Ma VT")
            sheet.cell(2, 13, "Ten hang")
            sheet.cell(2, 14, "DVT")
            sheet.cell(2, 15, "So luong")
            sheet.cell(2, 16, "Don gia")
            sheet.cell(2, 17, "TK vat tu")
            sheet.cell(2, 18, "Ma kho")
            for row, quantity in [(3, 500), (4, 250)]:
                sheet.cell(row, 9, "Cong ty A")
                sheet.cell(row, 10, "MST")
                sheet.cell(row, 12, 0)
                sheet.cell(row, 13, "Giay in A4")
                sheet.cell(row, 14, "To")
                sheet.cell(row, 15, quantity)
                sheet.cell(row, 16, 100)
            workbook.save(source)
            workbook.close()

            process_workbook(source, result, {
                "company_col": "I",
                "mst_col": "J",
                "product_col": "M",
                "qty_col": "O",
                "output_col": "L",
                "profile": "vietmax_ban_ra",
                "include_company_prefix": False,
                "all_mst": ["MST"],
                "process_mst": ["MST"],
                "mst_safe_id": ["MST|||0"],
                "selected_products_0": ["Giay in A4", "Giay my thuat"],
                "inventory_pairs": [{"id": "default", "ma_kho": "KVT", "tk_vat_tu": "152"}],
                "use_default_inventory_pair": True,
                "default_inventory_pair_id": "default",
                "vietmax_ban_ra_purchase_matches": [
                    {
                        "sales_product": "Giay in A4",
                        "purchase_code": "GIAY-RAM",
                        "sales_unit": "To",
                        "purchase_unit": "Ram",
                        "unit_mismatch": True,
                        "conversion_mode": app.VIETMAX_CONVERSION_MODE_QTY_AND_UNIT,
                        "conversion_formula": "1 ram = 500 to",
                        "confirmed": True,
                    },
                ],
            })

            output = load_workbook(result, data_only=True)
            try:
                result_sheet = output["Invoices"]
                self.assertEqual(result_sheet.cell(3, 12).value, make_product_part("vietmax_ban_ra", "Giay in A4", {}))
                self.assertEqual(result_sheet.cell(4, 12).value, make_product_part("vietmax_ban_ra", "Giay in A4", {}))
                self.assertEqual(result_sheet.cell(3, 14).value, "Ram")
                self.assertEqual(result_sheet.cell(4, 14).value, "Ram")
                self.assertEqual(result_sheet.cell(3, 15).value, 1)
                self.assertEqual(result_sheet.cell(4, 15).value, 0.5)
                self.assertEqual(result_sheet.cell(3, 16).value, 50000)
                self.assertEqual(result_sheet.cell(4, 16).value, 50000)
                self.assertEqual((result_sheet.cell(3, 17).value, result_sheet.cell(3, 18).value), ("152", "KVT"))
                self.assertEqual((result_sheet.cell(4, 17).value, result_sheet.cell(4, 18).value), ("152", "KVT"))
            finally:
                output.close()
        finally:
            source.unlink(missing_ok=True)
            result.unlink(missing_ok=True)

    def test_vietmax_ban_ra_saved_purchase_match_rule_converts_with_exact_khh_match(self):
        outputs = Path(__file__).parent / "outputs"
        outputs.mkdir(exist_ok=True)
        run_id = uuid4().hex
        source = outputs / f"_test_vietmax_saved_rule_source_{run_id}.xlsx"
        purchase = outputs / f"_test_vietmax_saved_rule_purchase_{run_id}.xlsx"
        result = outputs / f"_test_vietmax_saved_rule_result_{run_id}.xlsx"
        try:
            sales_workbook = Workbook()
            sales_sheet = sales_workbook.active
            sales_sheet.title = "Invoices"
            sales_sheet.cell(2, 9, "Ten nguoi mua")
            sales_sheet.cell(2, 10, "MST nguoi mua")
            sales_sheet.cell(2, 12, "Ma VT")
            sales_sheet.cell(2, 13, "Ten hang")
            sales_sheet.cell(2, 14, "DVT")
            sales_sheet.cell(2, 15, "So luong")
            sales_sheet.cell(2, 16, "Don gia")
            sales_sheet.cell(2, 18, "Thue suat")
            sales_sheet.cell(2, 19, "Tien hang")
            sales_sheet.cell(2, 20, "Tien thue")
            sales_sheet.cell(2, 21, "Tong TT")
            sales_sheet.cell(3, 9, "Cong ty A")
            sales_sheet.cell(3, 10, "MST")
            sales_sheet.cell(3, 12, 0)
            sales_sheet.cell(3, 13, "Giay in A4")
            sales_sheet.cell(3, 14, "To")
            sales_sheet.cell(3, 15, 35)
            sales_sheet.cell(3, 16, 600)
            sales_sheet.cell(3, 18, 8)
            sales_sheet.cell(3, 19, 21000)
            sales_sheet.cell(3, 20, 1680)
            sales_sheet.cell(3, 21, 22680)
            sales_sheet.cell(4, 9, "Cong ty A")
            sales_sheet.cell(4, 10, "MST")
            sales_sheet.cell(4, 12, 0)
            sales_sheet.cell(4, 13, "Giay my thuat")
            sales_sheet.cell(4, 14, "Kg")
            sales_sheet.cell(4, 15, 0.5)
            sales_sheet.cell(4, 16, 35000)
            sales_sheet.cell(4, 18, 8)
            sales_sheet.cell(4, 19, 17500)
            sales_sheet.cell(4, 20, 1400)
            sales_sheet.cell(4, 21, 18900)
            sales_sheet.cell(5, 9, "Cong ty A")
            sales_sheet.cell(5, 10, "MST")
            sales_sheet.cell(5, 12, 0)
            sales_sheet.cell(5, 13, "Giay in A4")
            sales_sheet.cell(5, 14, "Kg")
            sales_sheet.cell(5, 15, 850)
            sales_sheet.cell(5, 16, 25000)
            sales_sheet.cell(5, 18, 8)
            sales_sheet.cell(5, 19, 21250000)
            sales_sheet.cell(5, 20, 1700000)
            sales_sheet.cell(5, 21, 22950000)
            sales_workbook.save(source)
            sales_workbook.close()

            purchase_workbook = Workbook()
            purchase_sheet = purchase_workbook.active
            purchase_sheet.title = "ProcessedPurchase"
            purchase_sheet.cell(2, 12, "Ma VT")
            purchase_sheet.cell(2, 13, "Ten hang")
            purchase_sheet.cell(2, 14, "DVT")
            purchase_sheet.cell(2, 16, "Don gia")
            purchase_sheet.cell(3, 12, "GIAY-KG")
            purchase_sheet.cell(3, 13, "Giay in A4")
            purchase_sheet.cell(3, 14, "Kg")
            purchase_sheet.cell(3, 16, 21000)
            purchase_sheet.cell(4, 12, "GIAYMYTHUAT")
            purchase_sheet.cell(4, 13, "Giay my thuat")
            purchase_sheet.cell(4, 14, "Xap")
            purchase_sheet.cell(4, 16, 1925000)
            purchase_workbook.save(purchase)
            purchase_workbook.close()

            process_workbook(source, result, {
                "company_col": "I",
                "mst_col": "J",
                "product_col": "M",
                "qty_col": "O",
                "output_col": "L",
                "profile": "vietmax_ban_ra",
                "include_company_prefix": False,
                "all_mst": ["MST"],
                "process_mst": ["MST"],
                "mst_safe_id": ["MST|||0"],
                "selected_products_0": ["Giay in A4", "Giay my thuat"],
                "inventory_pairs": [{"id": "default", "ma_kho": "KVT", "tk_vat_tu": "152"}],
                "use_default_inventory_pair": True,
                "default_inventory_pair_id": "default",
                "vietmax_processed_purchase_path": str(purchase),
                "vietmax_ban_ra_purchase_match_rules": [
                    {
                        "sales_product": "Giay in A4",
                        "purchase_code": "GIAY-KG",
                        "purchase_product": "Giay in A4",
                        "sales_unit": "To",
                        "purchase_unit": "Kg",
                        "conversion_mode": app.VIETMAX_CONVERSION_MODE_QTY_AND_UNIT,
                        "conversion_formula": "35 To = 1 Kg",
                        "confirmed": True,
                    },
                    {
                        "sales_product": "Giay my thuat",
                        "purchase_code": "GIAYMYTHUAT",
                        "purchase_product": "Giay my thuat",
                        "sales_unit": "Kg",
                        "purchase_unit": "Xap",
                        "conversion_mode": app.VIETMAX_CONVERSION_MODE_QTY_AND_UNIT,
                        "conversion_formula": "55 Kg = 1 Xap",
                        "confirmed": True,
                    },
                ],
            })

            output = load_workbook(result, data_only=True)
            try:
                result_sheet = output["Invoices"]
                self.assertEqual(result_sheet.cell(3, 14).value, "Kg")
                self.assertEqual(result_sheet.cell(3, 15).value, 1)
                self.assertEqual(result_sheet.cell(3, 16).value, 21000)
                self.assertEqual(result_sheet.cell(3, 19).value, 21000)
                self.assertEqual(result_sheet.cell(3, 20).value, 1680)
                self.assertEqual(result_sheet.cell(3, 21).value, 22680)
                self.assertEqual((result_sheet.cell(3, 22).value, result_sheet.cell(3, 23).value), ("152", "KVT"))
                self.assertEqual(result_sheet.cell(4, 14).value, "Xap")
                self.assertAlmostEqual(result_sheet.cell(4, 15).value, 0.009091, places=6)
                self.assertAlmostEqual(result_sheet.cell(4, 16).value * result_sheet.cell(4, 15).value, 17500, places=2)
                self.assertEqual(result_sheet.cell(4, 19).value, 17500)
                self.assertEqual(result_sheet.cell(4, 20).value, 1400)
                self.assertEqual(result_sheet.cell(4, 21).value, 18900)
                self.assertEqual((result_sheet.cell(4, 22).value, result_sheet.cell(4, 23).value), ("152", "KVT"))
                self.assertEqual(result_sheet.cell(5, 14).value, "Kg")
                self.assertEqual(result_sheet.cell(5, 15).value, 850)
                self.assertEqual(result_sheet.cell(5, 16).value, 25000)
                self.assertEqual(result_sheet.cell(5, 19).value, 21250000)
                self.assertEqual(result_sheet.cell(5, 20).value, 1700000)
                self.assertEqual(result_sheet.cell(5, 21).value, 22950000)
                self.assertEqual((result_sheet.cell(5, 22).value, result_sheet.cell(5, 23).value), ("152", "KVT"))
            finally:
                output.close()
        finally:
            source.unlink(missing_ok=True)
            purchase.unlink(missing_ok=True)
            result.unlink(missing_ok=True)

    def test_vietmax_ban_ra_unchecked_purchase_match_does_not_override_convert_or_force_inventory(self):
        outputs = Path(__file__).parent / "outputs"
        outputs.mkdir(exist_ok=True)
        run_id = uuid4().hex
        source = outputs / f"_test_vietmax_unchecked_source_{run_id}.xlsx"
        result = outputs / f"_test_vietmax_unchecked_result_{run_id}.xlsx"
        try:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Invoices"
            sheet.cell(2, 9, "Ten nguoi mua")
            sheet.cell(2, 10, "MST nguoi mua")
            sheet.cell(2, 12, "Ma VT")
            sheet.cell(2, 13, "Ten hang")
            sheet.cell(2, 14, "DVT")
            sheet.cell(2, 15, "So luong")
            sheet.cell(2, 17, "TK vat tu")
            sheet.cell(2, 18, "Ma kho")
            sheet.cell(3, 9, "Cong ty A")
            sheet.cell(3, 10, "MST")
            sheet.cell(3, 12, 0)
            sheet.cell(3, 13, "Unchecked Product")
            sheet.cell(3, 14, "To")
            sheet.cell(3, 15, 500)
            workbook.save(source)
            workbook.close()

            process_workbook(source, result, {
                "company_col": "I",
                "mst_col": "J",
                "product_col": "M",
                "qty_col": "O",
                "output_col": "L",
                "profile": "vietmax_ban_ra",
                "include_company_prefix": False,
                "all_mst": ["MST"],
                "process_mst": ["MST"],
                "mst_safe_id": ["MST|||0"],
                "selected_products_0": ["Unchecked Product"],
                "inventory_pairs": [{"id": "default", "ma_kho": "KVT", "tk_vat_tu": "152"}],
                "use_default_inventory_pair": True,
                "default_inventory_pair_id": "default",
                "vietmax_ban_ra_purchase_matches": [
                    {
                        "sales_product": "Unchecked Product",
                        "purchase_code": "PURCHASE-CODE",
                        "sales_unit": "To",
                        "purchase_unit": "Ram",
                        "unit_mismatch": True,
                        "conversion_mode": app.VIETMAX_CONVERSION_MODE_QTY_AND_UNIT,
                        "conversion_formula": "1 ram = 500 to",
                        "confirmed": False,
                    },
                ],
            })

            output = load_workbook(result, data_only=True)
            try:
                result_sheet = output["Invoices"]
                self.assertNotEqual(result_sheet.cell(3, 12).value, "PURCHASE-CODE")
                self.assertEqual(result_sheet.cell(3, 14).value, "To")
                self.assertEqual(result_sheet.cell(3, 15).value, 500)
                self.assertEqual((result_sheet.cell(3, 17).value, result_sheet.cell(3, 18).value), ("152", "KVT"))
            finally:
                output.close()
        finally:
            source.unlink(missing_ok=True)
            result.unlink(missing_ok=True)

    def test_vietmax_ban_ra_purchase_match_overrides_code_and_inventory(self):
        outputs = Path(__file__).parent / "outputs"
        outputs.mkdir(exist_ok=True)
        run_id = uuid4().hex
        source = outputs / f"_test_vietmax_match_source_{run_id}.xlsx"
        result = outputs / f"_test_vietmax_match_result_{run_id}.xlsx"
        try:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Invoices"
            sheet.cell(2, 9, "Ten nguoi mua")
            sheet.cell(2, 10, "MST nguoi mua")
            sheet.cell(2, 12, "Ma VT")
            sheet.cell(2, 13, "Ten hang")
            sheet.cell(2, 15, "So luong")
            sheet.cell(2, 17, "TK vat tu")
            sheet.cell(2, 18, "Ma kho")
            sheet.cell(3, 9, "Cong ty A")
            sheet.cell(3, 10, "MST")
            sheet.cell(3, 12, 0)
            sheet.cell(3, 13, "Matched Product")
            sheet.cell(3, 15, 1)
            sheet.cell(4, 9, "Cong ty A")
            sheet.cell(4, 10, "MST")
            sheet.cell(4, 12, 0)
            sheet.cell(4, 13, "Other Product")
            sheet.cell(4, 15, 1)
            workbook.save(source)
            workbook.close()

            process_workbook(source, result, {
                "company_col": "I",
                "mst_col": "J",
                "product_col": "M",
                "qty_col": "O",
                "output_col": "L",
                "profile": "vietmax_ban_ra",
                "include_company_prefix": False,
                "all_mst": ["MST"],
                "process_mst": ["MST"],
                "mst_safe_id": ["MST|||0"],
                "selected_products_0": ["Matched Product", "Other Product"],
                "inventory_pairs": [{"id": "default", "ma_kho": "KVT", "tk_vat_tu": "152"}],
                "use_default_inventory_pair": True,
                "default_inventory_pair_id": "default",
                "vietmax_ban_ra_purchase_matches": [
                    {"sales_product": "Matched Product", "purchase_code": "PURCHASE-CODE", "confirmed": True},
                ],
            })

            output = load_workbook(result, data_only=True)
            try:
                result_sheet = output["Invoices"]
                self.assertEqual(result_sheet.cell(3, 12).value, make_product_part("vietmax_ban_ra", "Matched Product", {}))
                self.assertNotEqual(result_sheet.cell(4, 12).value, "PURCHASE-CODE")
                self.assertEqual((result_sheet.cell(3, 17).value, result_sheet.cell(3, 18).value), ("152", "KVT"))
                self.assertEqual((result_sheet.cell(4, 17).value, result_sheet.cell(4, 18).value), ("152", "KVT"))
            finally:
                output.close()
        finally:
            source.unlink(missing_ok=True)
            result.unlink(missing_ok=True)

    def test_unified_vietmax_sales_uses_processed_purchase_for_khh_and_excludes_from_default(self):
        outputs = Path(__file__).parent / "outputs"
        outputs.mkdir(exist_ok=True)
        run_id = uuid4().hex
        purchase = outputs / f"_test_vietmax_processed_purchase_{run_id}.xlsx"
        sales = outputs / f"_test_vietmax_sales_{run_id}.xlsx"
        result = outputs / f"_test_vietmax_sales_result_{run_id}.xlsx"
        try:
            purchase_wb = Workbook()
            purchase_ws = purchase_wb.active
            purchase_ws.title = "Purchase"
            purchase_ws.cell(2, 12, "Ma VT")
            purchase_ws.cell(2, 13, "Ten hang")
            purchase_ws.cell(2, 14, "DVT")
            purchase_ws.cell(2, 17, "TK vat tu")
            purchase_ws.cell(2, 18, "Ma kho")
            purchase_ws.cell(3, 12, "KVT-PROC-001")
            purchase_ws.cell(3, 13, "Giay in A4")
            purchase_ws.cell(3, 14, "Ram")
            purchase_ws.cell(3, 17, "152")
            purchase_ws.cell(3, 18, "KVT")
            purchase_wb.save(purchase)
            purchase_wb.close()

            info = app.validate_vietmax_processed_purchase_workbook(purchase)
            self.assertEqual(info["code_col"], "L")
            self.assertEqual(info["valid_rows"], 1)

            sales_wb = Workbook()
            sales_ws = sales_wb.active
            sales_ws.title = "Sales"
            sales_ws.cell(2, 9, "Ten nguoi mua")
            sales_ws.cell(2, 10, "MST nguoi mua")
            sales_ws.cell(2, 12, "Ma VT")
            sales_ws.cell(2, 13, "Ten hang")
            sales_ws.cell(2, 14, "DVT")
            sales_ws.cell(2, 15, "So luong")
            sales_ws.cell(2, 17, "TK vat tu")
            sales_ws.cell(2, 18, "Ma kho")
            for row, product in [(3, "Giay in A4"), (4, "In decal tem")]:
                sales_ws.cell(row, 9, "Cong ty A")
                sales_ws.cell(row, 10, "MST")
                sales_ws.cell(row, 12, 0)
                sales_ws.cell(row, 13, product)
                sales_ws.cell(row, 14, "Ram")
                sales_ws.cell(row, 15, 1)
            sales_wb.save(sales)
            sales_wb.close()

            process_workbook(sales, result, {
                "company_col": "I",
                "mst_col": "J",
                "product_col": "M",
                "qty_col": "O",
                "output_col": "L",
                "profile": "vietmax",
                "vietmax_phase": app.VIETMAX_PHASE_SALES,
                "vietmax_processed_purchase_path": str(purchase),
                "include_company_prefix": False,
                "all_mst": ["MST"],
                "process_mst": ["MST"],
                "mst_safe_id": ["MST|||0"],
                "selected_products_0": ["Giay in A4", "In decal tem"],
                "inventory_pairs": [
                    {"id": "ktp", "ma_kho": "KTP", "tk_vat_tu": "154"},
                    {"id": "default", "ma_kho": "KTP", "tk_vat_tu": "154"},
                ],
                "use_default_inventory_pair": True,
                "default_inventory_pair_id": "default",
                "inventory_pair_rules": [
                    {"enabled": True, "source_col": "M", "operator": "contains", "value": "In", "pair_id": "ktp"},
                ],
            })

            output = load_workbook(result, data_only=True)
            try:
                result_sheet = output["Sales"]
                self.assertEqual(result_sheet.cell(3, 12).value, make_product_part("vietmax_ban_ra", "Giay in A4", {}))
                self.assertEqual((result_sheet.cell(3, 17).value, result_sheet.cell(3, 18).value), ("152", "KVT"))
                self.assertNotEqual(result_sheet.cell(4, 12).value, "KVT-PROC-001")
                self.assertEqual((result_sheet.cell(4, 17).value, result_sheet.cell(4, 18).value), ("154", "KTP"))
            finally:
                output.close()
        finally:
            purchase.unlink(missing_ok=True)
            sales.unlink(missing_ok=True)
            result.unlink(missing_ok=True)

    def test_unified_vietmax_sales_exact_khh_overrides_stale_manual_match(self):
        outputs = Path(__file__).parent / "outputs"
        outputs.mkdir(exist_ok=True)
        run_id = uuid4().hex
        purchase = outputs / f"_test_vietmax_exact_priority_purchase_{run_id}.xlsx"
        sales = outputs / f"_test_vietmax_exact_priority_sales_{run_id}.xlsx"
        result = outputs / f"_test_vietmax_exact_priority_result_{run_id}.xlsx"
        try:
            purchase_wb = Workbook()
            purchase_ws = purchase_wb.active
            purchase_ws.title = "Purchase"
            purchase_ws.cell(2, 12, "Ma VT")
            purchase_ws.cell(2, 13, "Ten hang")
            purchase_ws.cell(2, 14, "DVT")
            purchase_ws.cell(2, 17, "TK vat tu")
            purchase_ws.cell(2, 18, "Ma kho")
            purchase_ws.cell(3, 12, "KVT-EXACT")
            purchase_ws.cell(3, 13, "Giay in A4")
            purchase_ws.cell(3, 14, "Ram")
            purchase_ws.cell(3, 17, "152")
            purchase_ws.cell(3, 18, "KVT")
            purchase_wb.save(purchase)
            purchase_wb.close()

            sales_wb = Workbook()
            sales_ws = sales_wb.active
            sales_ws.title = "Sales"
            sales_ws.cell(2, 9, "Ten nguoi mua")
            sales_ws.cell(2, 10, "MST nguoi mua")
            sales_ws.cell(2, 12, "Ma VT")
            sales_ws.cell(2, 13, "Ten hang")
            sales_ws.cell(2, 14, "DVT")
            sales_ws.cell(2, 15, "So luong")
            sales_ws.cell(2, 17, "TK vat tu")
            sales_ws.cell(2, 18, "Ma kho")
            sales_ws.cell(3, 9, "Cong ty A")
            sales_ws.cell(3, 10, "MST")
            sales_ws.cell(3, 12, 0)
            sales_ws.cell(3, 13, "Giay in A4")
            sales_ws.cell(3, 14, "Ram")
            sales_ws.cell(3, 15, 1)
            sales_wb.save(sales)
            sales_wb.close()

            process_workbook(sales, result, {
                "company_col": "I",
                "mst_col": "J",
                "product_col": "M",
                "qty_col": "O",
                "output_col": "L",
                "profile": "vietmax",
                "vietmax_phase": app.VIETMAX_PHASE_SALES,
                "vietmax_processed_purchase_path": str(purchase),
                "include_company_prefix": False,
                "all_mst": ["MST"],
                "process_mst": ["MST"],
                "mst_safe_id": ["MST|||0"],
                "selected_products_0": ["Giay in A4", "Giay my thuat"],
                "vietmax_ban_ra_purchase_matches": [
                    {"sales_product": "Giay in A4", "purchase_code": "STALE-MATCH", "confirmed": True},
                ],
                "inventory_pairs": [{"id": "default", "ma_kho": "KTP", "tk_vat_tu": "154"}],
                "use_default_inventory_pair": True,
                "default_inventory_pair_id": "default",
            })

            output = load_workbook(result, data_only=True)
            try:
                result_sheet = output["Sales"]
                self.assertEqual(result_sheet.cell(3, 12).value, make_product_part("vietmax_ban_ra", "Giay in A4", {}))
                self.assertEqual((result_sheet.cell(3, 17).value, result_sheet.cell(3, 18).value), ("152", "KVT"))
            finally:
                output.close()
        finally:
            purchase.unlink(missing_ok=True)
            sales.unlink(missing_ok=True)
            result.unlink(missing_ok=True)

    def test_vietmax_sales_internal_merge_affects_final_purchase_match_lookup(self):
        outputs = Path(__file__).parent / "outputs"
        outputs.mkdir(exist_ok=True)
        run_id = uuid4().hex
        source = outputs / f"_test_vietmax_sales_internal_merge_source_{run_id}.xlsx"
        result = outputs / f"_test_vietmax_sales_internal_merge_result_{run_id}.xlsx"
        try:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Invoices"
            sheet.cell(2, 9, "Ten nguoi mua")
            sheet.cell(2, 10, "MST nguoi mua")
            sheet.cell(2, 12, "Ma VT")
            sheet.cell(2, 13, "Ten hang")
            sheet.cell(2, 15, "So luong")
            sheet.cell(3, 9, "Cong ty A")
            sheet.cell(3, 10, "MST")
            sheet.cell(3, 12, 0)
            sheet.cell(3, 13, "Giấy Couche 300 gms")
            sheet.cell(3, 15, 1)
            workbook.save(source)
            workbook.close()

            process_workbook(source, result, {
                "company_col": "I",
                "mst_col": "J",
                "product_col": "M",
                "qty_col": "O",
                "output_col": "L",
                "profile": "vietmax_ban_ra",
                "include_company_prefix": False,
                "all_mst": ["MST"],
                "process_mst": ["MST"],
                "mst_safe_id": ["MST|||0"],
                "selected_products_0": ["Giấy Couche 300 gms"],
                "vietmax_ban_ra_sales_internal_merges": [
                    {"product": "Giấy Couche 300 gms", "similar_product": "Giấy Couche 300 gsm", "confirmed": True},
                ],
                "vietmax_ban_ra_purchase_matches": [
                    {"sales_product": "Giấy Couche 300 gsm", "purchase_code": "PURCHASE-CODE", "confirmed": True},
                ],
            })

            output = load_workbook(result, data_only=True)
            try:
                result_sheet = output["Invoices"]
                self.assertEqual(result_sheet.cell(3, 12).value, make_product_part("vietmax_ban_ra", "Giấy Couche 300 gsm", {}))
            finally:
                output.close()
        finally:
            source.unlink(missing_ok=True)
            result.unlink(missing_ok=True)

    def test_vietmax_ban_ra_repeated_phrase_rule_affects_processed_excel(self):
        outputs = Path(__file__).parent / "outputs"
        outputs.mkdir(exist_ok=True)
        run_id = uuid4().hex
        source = outputs / f"_test_vietmax_repeat_source_{run_id}.xlsx"
        result = outputs / f"_test_vietmax_repeat_result_{run_id}.xlsx"
        try:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Invoices"
            sheet.cell(2, 9, "Ten nguoi mua")
            sheet.cell(2, 10, "MST nguoi mua")
            sheet.cell(2, 12, "Ma VT")
            sheet.cell(2, 13, "Ten hang")
            sheet.cell(2, 15, "So luong")
            sheet.cell(3, 9, "Cong ty A")
            sheet.cell(3, 10, "111")
            sheet.cell(3, 12, 0)
            sheet.cell(3, 13, "Danh sách số 01 số 02 số 03")
            sheet.cell(3, 15, 1)
            workbook.save(source)
            workbook.close()

            process_workbook(source, result, {
                "company_col": "I",
                "mst_col": "J",
                "product_col": "M",
                "qty_col": "O",
                "output_col": "L",
                "profile": "vietmax_ban_ra",
                "include_company_prefix": False,
                "all_mst": ["111"],
                "process_mst": ["111"],
                "mst_safe_id": ["111|||0"],
                "selected_products_0": ["Danh sách số 01 số 02 số 03"],
                "repeated_phrase_removals": ["số"],
            })

            output = load_workbook(result, data_only=True)
            try:
                result_sheet = output["Invoices"]
                self.assertEqual(result_sheet.cell(3, 12).value, "DANHSS010203")
                self.assertLessEqual(len(result_sheet.cell(3, 12).value), app.MAX_CODE_LENGTH)
            finally:
                output.close()
        finally:
            source.unlink(missing_ok=True)
            result.unlink(missing_ok=True)

    def test_vietmax_sales_internal_merge_reuses_similar_product_code(self):
        outputs = Path(__file__).parent / "outputs"
        outputs.mkdir(exist_ok=True)
        run_id = uuid4().hex
        source = outputs / f"_test_vietmax_sales_merge_source_{run_id}.xlsx"
        result = outputs / f"_test_vietmax_sales_merge_result_{run_id}.xlsx"
        try:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Invoices"
            sheet.cell(2, 9, "Ten nguoi mua")
            sheet.cell(2, 10, "MST nguoi mua")
            sheet.cell(2, 12, "Ma VT")
            sheet.cell(2, 13, "Ten hang")
            sheet.cell(2, 15, "So luong")
            sheet.cell(3, 9, "Cong ty A")
            sheet.cell(3, 10, "111")
            sheet.cell(3, 12, 0)
            sheet.cell(3, 13, "Giấy Couche 300 gms")
            sheet.cell(3, 15, 1)
            sheet.cell(4, 9, "Cong ty A")
            sheet.cell(4, 10, "111")
            sheet.cell(4, 12, 0)
            sheet.cell(4, 13, "Giấy Couche 300 gsm")
            sheet.cell(4, 15, 1)
            workbook.save(source)
            workbook.close()

            process_workbook(source, result, {
                "company_col": "I",
                "mst_col": "J",
                "product_col": "M",
                "qty_col": "O",
                "output_col": "L",
                "profile": "vietmax_ban_ra",
                "include_company_prefix": False,
                "all_mst": ["111"],
                "process_mst": ["111"],
                "mst_safe_id": ["111|||0"],
                "selected_products_0": ["Giấy Couche 300 gms", "Giấy Couche 300 gsm"],
                "comparison_scope": app.VIETMAX_COMPARISON_SCOPE_SAME_COMPANY,
                "vietmax_ban_ra_sales_internal_merges": [
                    {
                        "product": "Giấy Couche 300 gms",
                        "similar_product": "Giấy Couche 300 gsm",
                        "company": "Cong ty A",
                        "mst": "111",
                        "company_key": "111",
                        "comparison_scope": app.VIETMAX_COMPARISON_SCOPE_SAME_COMPANY,
                        "confirmed": True,
                    },
                ],
            })

            output = load_workbook(result, data_only=True)
            try:
                result_sheet = output["Invoices"]
                self.assertEqual(result_sheet.cell(3, 12).value, result_sheet.cell(4, 12).value)
                self.assertEqual(result_sheet.cell(3, 12).value, make_product_part("vietmax_ban_ra", "Giấy Couche 300 gsm", {}))
            finally:
                output.close()
        finally:
            source.unlink(missing_ok=True)
            result.unlink(missing_ok=True)

    def test_vietmax_same_company_confirmed_match_affects_only_matching_company_rows(self):
        outputs = Path(__file__).parent / "outputs"
        outputs.mkdir(exist_ok=True)
        run_id = uuid4().hex
        source = outputs / f"_test_vietmax_same_company_source_{run_id}.xlsx"
        result = outputs / f"_test_vietmax_same_company_result_{run_id}.xlsx"
        try:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Invoices"
            sheet.cell(2, 9, "Ten nguoi mua")
            sheet.cell(2, 10, "MST nguoi mua")
            sheet.cell(2, 12, "Ma VT")
            sheet.cell(2, 13, "Ten hang")
            sheet.cell(2, 14, "DVT")
            sheet.cell(2, 15, "So luong")
            sheet.cell(2, 17, "TK vat tu")
            sheet.cell(2, 18, "Ma kho")
            for row, company, mst in [(3, "Cong ty A", "111"), (4, "Cong ty B", "222")]:
                sheet.cell(row, 9, company)
                sheet.cell(row, 10, mst)
                sheet.cell(row, 12, 0)
                sheet.cell(row, 13, "Matched Product")
                sheet.cell(row, 14, "To")
                sheet.cell(row, 15, 500)
            workbook.save(source)
            workbook.close()

            process_workbook(source, result, {
                "company_col": "I",
                "mst_col": "J",
                "product_col": "M",
                "qty_col": "O",
                "output_col": "L",
                "profile": "vietmax_ban_ra",
                "include_company_prefix": False,
                "all_mst": ["111", "222"],
                "process_mst": ["111", "222"],
                "mst_safe_id": ["111|||0", "222|||1"],
                "selected_products_0": ["Matched Product"],
                "selected_products_1": ["Matched Product"],
                "inventory_pairs": [{"id": "default", "ma_kho": "KVT", "tk_vat_tu": "152"}],
                "use_default_inventory_pair": True,
                "default_inventory_pair_id": "default",
                "comparison_scope": app.VIETMAX_COMPARISON_SCOPE_SAME_COMPANY,
                "vietmax_ban_ra_purchase_matches": [
                    {
                        "sales_product": "Matched Product",
                        "sales_company": "Cong ty A",
                        "sales_mst": "111",
                        "purchase_code": "PURCHASE-CODE",
                        "sales_unit": "To",
                        "purchase_unit": "Ram",
                        "unit_mismatch": True,
                        "conversion_mode": app.VIETMAX_CONVERSION_MODE_QTY_AND_UNIT,
                        "conversion_formula": "1 ram = 500 to",
                        "comparison_scope": app.VIETMAX_COMPARISON_SCOPE_SAME_COMPANY,
                        "confirmed": True,
                    },
                ],
            })

            output = load_workbook(result, data_only=True)
            try:
                result_sheet = output["Invoices"]
                self.assertEqual(result_sheet.cell(3, 12).value, make_product_part("vietmax_ban_ra", "Matched Product", {}))
                self.assertNotEqual(result_sheet.cell(4, 12).value, "PURCHASE-CODE")
                self.assertEqual(result_sheet.cell(3, 14).value, "Ram")
                self.assertEqual(result_sheet.cell(4, 14).value, "To")
                self.assertEqual(result_sheet.cell(3, 15).value, 1)
                self.assertEqual(result_sheet.cell(4, 15).value, 500)
                self.assertEqual((result_sheet.cell(3, 17).value, result_sheet.cell(3, 18).value), ("152", "KVT"))
                self.assertEqual((result_sheet.cell(4, 17).value, result_sheet.cell(4, 18).value), ("152", "KVT"))
            finally:
                output.close()
        finally:
            source.unlink(missing_ok=True)
            result.unlink(missing_ok=True)


    def test_output_suffixes_are_fdi_and_up_files(self):
        output = resolve_output_path("hoa_don.xlsx", "")
        self.assertEqual(output.name, "hoa_don_fdi.xls")
        self.assertEqual(up_mua_vao_output_path(output).name, "hoa_don_fdi_UP_mua_vao.xls")
        self.assertEqual(up_ban_ra_output_path(output).name, "hoa_don_fdi_UP_ban_ra.xls")

    def test_selected_output_filename_is_normalized_to_fdi_suffix(self):
        output = resolve_output_path("hoa_don.xlsx", "outputs/custom_name.xlsx")
        self.assertEqual(output.name, "custom_name_fdi.xls")
        self.assertEqual(resolve_output_path("hoa_don.xlsx", "outputs/custom_name_fdi.xlsx").name, "custom_name_fdi.xls")

    def test_up_workbooks_use_inventory_columns(self):
        outputs = Path(__file__).parent / "outputs"
        outputs.mkdir(exist_ok=True)
        run_id = uuid4().hex
        source = outputs / f"_test_nhap_kho_source_{run_id}.xlsx"
        result = outputs / f"_test_nhap_kho_result_{run_id}.xlsx"
        result_no_inventory = outputs / f"_test_nhap_kho_result_no_inventory_{run_id}.xlsx"
        try:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Invoices"
            sheet.cell(2, 3, "So hoa don")
            sheet.cell(2, 4, "Ngay hoa don")
            sheet.cell(2, 6, "Cong ty")
            sheet.cell(2, 7, "MST")
            sheet.cell(2, 10, "Nguoi mua")
            sheet.cell(2, 12, "Ma VT")
            sheet.cell(2, 14, "Ten hang")
            sheet.cell(2, 15, "So luong")
            sheet.cell(2, 16, "So luong xu ly")
            sheet.cell(2, 17, "Don gia")
            sheet.cell(2, 18, "Thue")
            sheet.cell(2, 23, "Ghi chu")
            sheet.cell(3, 3, "HD-1")
            sheet.cell(3, 4, "01/01/2026")
            sheet.cell(3, 6, "Cong ty A")
            sheet.cell(3, 7, "MST")
            sheet.cell(3, 10, "Nguoi mua A")
            sheet.cell(3, 12, 0)
            sheet.cell(3, 14, "Hang A")
            sheet.cell(3, 15, 2)
            sheet.cell(3, 16, 2)
            sheet.cell(3, 17, 100)
            sheet.cell(3, 18, 0.1)
            sheet.cell(3, 23, "Ghi chu")
            workbook.save(source)
            workbook.close()

            processed = process_workbook(source, result, {
                "company_col": "F",
                "mst_col": "G",
                "product_col": "N",
                "qty_col": "P",
                "output_col": "L",
                "profile": "son_phuong",
                "include_company_prefix": False,
                "all_mst": ["MST"],
                "process_mst": ["MST"],
                "mst_safe_id": ["MST|||0"],
                "selected_products_0": ["Hang A"],
                "inventory_pairs": [{"id": "only", "ma_kho": "KVT", "tk_vat_tu": "152"}],
            })
            purchase_stream = create_up_mua_vao_workbook(processed)
            purchase_output = xlrd.open_workbook(file_contents=purchase_stream.getvalue())
            try:
                purchase_sheet = purchase_output.sheet_by_index(0)
                self.assertEqual(purchase_sheet.cell_value(1, excel_col_to_index("J")), "KVT")
                self.assertEqual(purchase_sheet.cell_value(1, excel_col_to_index("O")), "152")
            finally:
                purchase_output.release_resources()

            sales_stream = create_up_ban_ra_workbook(processed)
            sales_output = xlrd.open_workbook(file_contents=sales_stream.getvalue())
            try:
                sales_sheet = sales_output.sheet_by_index(0)
                self.assertEqual(sales_sheet.cell_value(1, excel_col_to_index("AB")), "152")
                self.assertEqual(sales_sheet.cell_value(1, excel_col_to_index("AF")), "KVT")
                self.assertNotEqual(sales_sheet.cell_value(1, excel_col_to_index("AB")), "KVT")
                self.assertNotEqual(sales_sheet.cell_value(1, excel_col_to_index("AF")), "152")
            finally:
                sales_output.release_resources()

            processed_without_inventory = process_workbook(source, result_no_inventory, {
                "company_col": "F",
                "mst_col": "G",
                "product_col": "N",
                "qty_col": "P",
                "output_col": "L",
                "profile": "son_phuong",
                "include_company_prefix": False,
                "all_mst": ["MST"],
                "process_mst": ["MST"],
                "mst_safe_id": ["MST|||0"],
                "selected_products_0": ["Hang A"],
            })
            with self.assertRaises(ValueError):
                create_up_mua_vao_workbook(processed_without_inventory)
            optional_purchase_stream = create_up_mua_vao_workbook(processed_without_inventory, require_inventory_columns=False)
            optional_purchase_output = xlrd.open_workbook(file_contents=optional_purchase_stream.getvalue())
            try:
                optional_sheet = optional_purchase_output.sheet_by_index(0)
                self.assertEqual(optional_sheet.cell_value(1, excel_col_to_index("K")), "HA")
                self.assertEqual(optional_sheet.cell_value(1, excel_col_to_index("J")), "")
                self.assertEqual(optional_sheet.cell_value(1, excel_col_to_index("O")), "")
            finally:
                optional_purchase_output.release_resources()
        finally:
            source.unlink(missing_ok=True)
            result.unlink(missing_ok=True)
            result_no_inventory.unlink(missing_ok=True)

    def test_fast_import_package_uses_processed_fdi_sources(self):
        outputs = Path(__file__).parent / "outputs"
        outputs.mkdir(exist_ok=True)
        run_id = uuid4().hex
        purchase_source = outputs / f"_test_fast_purchase_{run_id}.xlsx"
        sales_source = outputs / f"_test_fast_sales_{run_id}.xlsx"

        def write_processed_fdi(path, row_values):
            if isinstance(row_values, dict):
                row_values = [row_values]
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Invoices"
            for col, header in {
                "A": "Mẫu HD",
                "B": "Ký hiệu",
                "C": "Số HĐ",
                "D": "Ngày HD",
                "F": "Tên người bán",
                "G": "MST người bán",
                "H": "Địa chỉ người bán",
                "I": "Tên người mua",
                "J": "MST người mua",
                "K": "Địa chỉ người mua",
                "L": "Mã VT",
                "M": "Tên hàng hóa",
                "N": "ĐVT",
                "O": "Số lượng",
                "P": "Đơn giá",
                "R": "Thuế suất",
                "V": "Tiền chưa thuế",
                "W": "Tiền thuế",
                "AQ": "Mã ngoại tệ",
                "AR": "Tỷ giá",
                "AS": "TK vật tư",
                "AT": "Mã kho",
                "AU": "Mã khách hàng",
            }.items():
                sheet.cell(2, app.excel_col_to_index(col) + 1, header)
            for row_index, values in enumerate(row_values, start=3):
                for col, value in values.items():
                    sheet.cell(row_index, app.excel_col_to_index(col) + 1, value)
            workbook.save(path)
            workbook.close()

        try:
            write_processed_fdi(purchase_source, {
                "A": "01GTKT0",
                "B": "AA/26E",
                "C": "312",
                "D": "16/01/2026",
                "F": "CÔNG TY BÁN",
                "G": "010SELL",
                "H": "Địa chỉ bán",
                "I": "VIETMAX",
                "J": "010BUY",
                "K": "Địa chỉ mua",
                "L": "GIAYIN",
                "M": "Giấy in",
                "N": "Kg",
                "O": 12.5,
                "P": 21000,
                "R": 0.08,
                "V": 262500,
                "W": 21000,
                "AQ": "VND",
                "AR": 1,
                "AS": "152",
                "AT": "KVT",
            })
            write_processed_fdi(sales_source, [{
                "A": "01GTKT0",
                "B": "BB/26E",
                "C": "140",
                "D": "13/01/2026",
                "F": "VIETMAX",
                "G": "010SELL",
                "H": "Địa chỉ bán",
                "I": "CÔNG TY MUA",
                "J": "010CUSTOMER",
                "K": "Địa chỉ khách",
                "L": "INTG",
                "M": "In túi giấy",
                "N": "Chiếc",
                "O": 10,
                "P": 10000,
                "R": 0.08,
                "V": 100000,
                "W": 8000,
                "AQ": "VND",
                "AR": 1,
                "AS": "155",
                "AT": "KTP",
            }, {
                "A": "01GTKT0",
                "B": "BB/26E",
                "C": "141",
                "D": "14/01/2026",
                "F": "VIETMAX",
                "G": "010SELL",
                "H": "Địa chỉ bán",
                "I": "CÔNG TY GIA CÔNG",
                "J": "010CUSTOMER2",
                "K": "Địa chỉ khách 2",
                "L": "CONGIN",
                "M": "Công in",
                "N": "Lần",
                "O": 5,
                "P": 100,
                "R": 0.08,
                "V": 500,
                "W": 40,
                "AQ": "VND",
                "AR": 1,
                "AS": "1552",
                "AT": "KGCI",
            }, {
                "A": "01GTKT0",
                "B": "BB/26E",
                "C": "142",
                "D": "15/01/2026",
                "F": "VIETMAX",
                "G": "010SELL",
                "H": "Äá»‹a chá»‰ bÃ¡n",
                "I": "CÃ”NG TY KHONG MST",
                "J": "",
                "K": "Äá»‹a chá»‰ khÃ¡ch 3",
                "L": "KHM.HANG",
                "M": "HÃ ng khÃ´ng MST",
                "N": "CÃ¡i",
                "O": 2,
                "P": 5000,
                "R": 0.08,
                "V": 10000,
                "W": 800,
                "AQ": "VND",
                "AR": 1,
                "AS": "155",
                "AT": "KTP",
                "AU": "KHM",
            }])
            _, purchase_df = app.read_workbook(purchase_source)
            _, sales_df = app.read_workbook(sales_source)

            stream = app.fast_import_multi_sheet_workbook(purchase_df, sales_df)
            workbook = XlsWorkbookAdapter(stream.getvalue())
            try:
                self.assertEqual(
                    workbook.sheetnames,
                    ["Hoadonmuahang", "Bao_cao_trung_so_ct", "Hoadonbanhang", "DMvat_tu", "DMkhachhang"],
                )

                sheet = workbook["Hoadonmuahang"]
                self.assertEqual(sheet.max_column, len(app.FAST_HOA_DON_MUA_HANG_HEADERS))
                self.assertEqual(sheet.cell(2, 1).value, "010SELL")
                self.assertEqual(sheet.cell(2, 6).value, "312")
                self.assertEqual(sheet.cell(2, 10).value, "KVT")
                self.assertEqual(sheet.cell(2, 11).value, "GIAYIN")
                self.assertEqual(sheet.cell(2, 15).value, "152")
                for col in [2, 4, 16, 20, 21]:
                    self.assertIn(sheet.cell(2, col).value, (None, ""))

                sheet = workbook["Hoadonbanhang"]
                self.assertEqual(sheet.max_column, len(app.FAST_HOA_DON_BAN_HANG_HEADERS))
                self.assertEqual(sheet.cell(2, 1).value, "010CUSTOMER")
                self.assertEqual(sheet.cell(2, 5).value, "140")
                self.assertEqual(sheet.cell(2, 22).value, "08")
                self.assertEqual(sheet.cell(2, 28).value, "155")
                self.assertEqual(sheet.cell(2, 32).value, "KTP")
                self.assertEqual(sheet.cell(2, 33).value, "INTG")
                self.assertEqual(sheet.cell(2, 34).value, "Xuất bán hàng")
                self.assertEqual(sheet.cell(3, 1).value, "010CUSTOMER2")
                self.assertEqual(sheet.cell(3, 28).value, "1552")
                self.assertEqual(sheet.cell(3, 32).value, "KGCI")
                self.assertEqual(sheet.cell(3, 33).value, "CONGIN")
                self.assertEqual(sheet.cell(4, 1).value, "KHM")
                self.assertEqual(sheet.cell(4, 5).value, "142")
                for col in [2, 3, 4, 7, 9, 10, 11, 12, 13, 14, 17, 18, 19, 20, 21, 23, 24, 30, 35, 36, 38, 39, 40, 41, 42, 44, 46, 47, 48]:
                    self.assertIn(sheet.cell(2, col).value, (None, ""))

                sheet = workbook["DMvat_tu"]
                codes = {sheet.cell(row, 1).value for row in range(2, sheet.max_row + 1)}
                self.assertEqual(codes, {"GIAYIN", "INTG", "CONGIN", "KHM.HANG"})
                for row in range(2, sheet.max_row + 1):
                    for col in [8, 9, 12, 13, 14, 15, 16, 18, 19, 20, 21, 22, 24, 25, 26, 27, 28, 29, 30]:
                        self.assertIn(sheet.cell(row, col).value, (None, ""))

                sheet = workbook["DMkhachhang"]
                self.assertEqual(sheet.max_column, len(app.FAST_DM_KHACH_HANG_HEADERS))
                customers = {sheet.cell(row, 1).value for row in range(2, sheet.max_row + 1)}
                self.assertEqual(customers, {"010SELL", "010CUSTOMER", "010CUSTOMER2", "KHM"})
                for row in range(2, sheet.max_row + 1):
                    mst = sheet.cell(row, 1).value
                    if mst == "KHM":
                        self.assertIn(sheet.cell(row, 4).value, (None, ""))
                        continue
                    self.assertEqual(sheet.cell(row, 4).value, mst)
                    self.assertIn(sheet.cell(row, 2).value, {"CÔNG TY BÁN", "CÔNG TY MUA", "CÔNG TY GIA CÔNG"})
                    self.assertIn(sheet.cell(row, 5).value, {"Địa chỉ bán", "Địa chỉ khách", "Địa chỉ khách 2"})
                    for col in [3, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19]:
                        self.assertIn(sheet.cell(row, col).value, (None, ""))
            finally:
                workbook.close()
        finally:
            purchase_source.unlink(missing_ok=True)
            sales_source.unlink(missing_ok=True)

    def test_fast_import_package_adds_custom_form_mapping_sheet(self):
        outputs = Path(__file__).parent / "outputs"
        outputs.mkdir(exist_ok=True)
        run_id = uuid4().hex
        source = outputs / f"_test_fast_custom_mapping_{run_id}.xlsx"
        try:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Invoices"
            for column, header in {
                "C": "So HD",
                "L": "Ma VT",
                "O": "So luong",
                "AS": "TK vat tu",
                "AT": "Ma kho",
            }.items():
                sheet.cell(2, app.excel_col_to_index(column) + 1, header)
            for column, value in {
                "C": "777",
                "D": "20/01/2026",
                "F": "CONG TY DICH VU",
                "G": "010SERVICE",
                "L": "DV.TEST",
                "M": "Dich vu test",
                "N": "Lan",
                "O": 1,
                "P": 1000,
                "V": 1000,
                "AS": "642",
                "AT": "KDV",
            }.items():
                sheet.cell(3, app.excel_col_to_index(column) + 1, value)
            workbook.save(source)
            workbook.close()
            _, df = app.read_workbook(source)
            stream = app.fast_import_multi_sheet_workbook(
                df,
                None,
                form_mapping_presets=[{
                    "id": "service_form",
                    "label": "PhieuDichVu",
                    "scope": "purchase",
                    "group_id": "services",
                    "output_columns": [
                        {"letter": "A", "header": "Ma khach hang"},
                        {"letter": "B", "header": "Loai"},
                        {"letter": "C", "header": "Dien giai"},
                    ],
                    "mappings": [
                        {"target_col": "A", "source_col": "G", "source_phase": "purchase", "source_type": "source_column"},
                        {"target_col": "B", "source_phase": "purchase", "source_type": "constant", "value": "DV"},
                        {"target_col": "C", "source_phase": "purchase", "source_type": "text_template", "value": "HD{C}"},
                    ],
                }],
                purchase_company_group_assignments={"010SERVICE": "services"},
            )
            result = XlsWorkbookAdapter(stream.getvalue())
            try:
                self.assertIn("PhieuDichVu", result.sheetnames)
                mapped_sheet = result["PhieuDichVu"]
                self.assertEqual(mapped_sheet.cell(2, 1).value, "010SERVICE")
                self.assertEqual(mapped_sheet.cell(2, 2).value, "DV")
                self.assertEqual(mapped_sheet.cell(2, 3).value, "HD777")
            finally:
                result.close()
        finally:
            source.unlink(missing_ok=True)

    def test_fast_import_custom_group_matches_missing_mst_company_key(self):
        with TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "no_mst_service.xlsx"
            wb = Workbook()
            sheet = wb.active
            sheet.title = "Invoices"
            for column, value in {
                "C": "778",
                "D": "21/01/2026",
                "F": "CONG TY KHONG MST",
                "G": "",
                "L": "DV.NOMST",
                "M": "Dich vu khong mst",
                "N": "Lan",
                "O": 1,
                "P": 2000,
                "V": 2000,
            }.items():
                sheet.cell(3, app.excel_col_to_index(column) + 1, value)
            wb.save(path)
            wb.close()
            _, df = app.read_workbook(path)

        service_key = app.company_selection_key("CONG TY KHONG MST", "")
        stream = app.fast_import_multi_sheet_workbook(
            df,
            None,
            form_mapping_presets=[{
                "id": "service_no_mst",
                "label": "DichVuKhongMst",
                "scope": "purchase",
                "group_id": "services",
                "output_columns": [
                    {"letter": "A", "header": "Cong ty"},
                    {"letter": "B", "header": "So HD"},
                ],
                "mappings": [
                    {"target_col": "A", "source_col": "F", "source_phase": "purchase", "source_type": "source_column"},
                    {"target_col": "B", "source_col": "C", "source_phase": "purchase", "source_type": "source_column"},
                ],
            }],
            purchase_company_group_assignments={service_key: "services"},
            require_inventory_columns=False,
        )
        result = XlsWorkbookAdapter(stream.getvalue())
        try:
            self.assertIn("DichVuKhongMst", result.sheetnames)
            mapped_sheet = result["DichVuKhongMst"]
            self.assertEqual(mapped_sheet.cell(2, 1).value, "CONG TY KHONG MST")
            self.assertEqual(mapped_sheet.cell(2, 2).value, "778")
        finally:
            result.close()

    def test_fast_mapping_transform_rules_are_configurable(self):
        revenue_rules = [
            {"match_type": "regex", "value": "^33", "result": "515"},
            {"match_type": "contains", "value": "711", "result": "7111"},
            {"match_type": "gte", "value": "900", "result": "9999"},
            {"match_type": "default", "result": "5119"},
        ]
        self.assertEqual(app.fast_apply_mapping_transform("revenue_account_from_inventory_account", "331", {}, revenue_rules), "515")
        self.assertEqual(app.fast_apply_mapping_transform("revenue_account_from_inventory_account", "711", {}, revenue_rules), "7111")
        self.assertEqual(app.fast_apply_mapping_transform("revenue_account_from_inventory_account", "950", {}, revenue_rules), "9999")
        self.assertEqual(app.fast_apply_mapping_transform("revenue_account_from_inventory_account", "642", {}, revenue_rules), "5119")
        self.assertEqual(
            app.fast_apply_mapping_transform(
                "revenue_account_from_inventory_account",
                "642",
                {},
                [{"match_type": "starts_with", "value": "155", "result": "5112"}],
            ),
            "642",
        )
        self.assertEqual(
            app.fast_apply_mapping_transform(
                "material_type_from_inventory_account",
                "331",
                {},
                [{"match_type": "default", "result": "99"}],
            ),
            "99",
        )
        self.assertEqual(
            app.fast_mapping_rule_value(
                {
                    "source_type": "conditional_rules",
                    "source_col": "B",
                    "transform_rules": [
                        {"match_type": "contains", "value": "dịch vụ", "result": "DV"},
                        {"match_type": "default", "result": "VT"},
                    ],
                },
                {"source_values": {"B": "Phí dịch vụ tháng 1"}},
            ),
            "DV",
        )
        self.assertEqual(
            app.fast_mapping_rule_value(
                {
                    "source_type": "conditional_rules",
                    "source_col": "C",
                    "transform_rules": [{"match_type": "starts_with", "value": "HD", "result": "invoice"}],
                },
                {"source_values": {"C": "PX001"}},
            ),
            "PX001",
        )
        self.assertEqual(
            app.fast_mapping_rule_value(
                {
                    "source_type": "if_rules",
                    "transform_rules": [
                        {"source_col": "B", "match_type": "starts_with", "value": "152", "result": "21"},
                        {"match_type": "default", "result": "51"},
                    ],
                },
                {"source_values": {"A": "Giữ giá trị gốc", "B": "1521"}},
            ),
            "21",
        )
        self.assertEqual(
            app.fast_mapping_rule_value(
                {
                    "source_type": "if_rules",
                    "source_col": "A",
                    "transform_rules": [{"source_col": "B", "match_type": "starts_with", "value": "155", "result": "51"}],
                },
                {"source_values": {"A": "Giữ giá trị gốc", "B": "1521"}},
            ),
            "Giữ giá trị gốc",
        )

        self.assertEqual(
            app.fast_mapping_rule_value(
                {
                    "source_type": "if_rules",
                    "transform_rules": [{"source_col": "B", "match_type": "equals", "value": "OK", "result": "HD{A}"}],
                },
                {"source_values": {"A": "123", "B": "OK"}},
            ),
            "HD123",
        )

    def test_fast_import_splits_sales_duplicate_invoice_numbers_by_customer(self):
        max_col = app.excel_col_to_index("AU") + 1
        rows = [[""] * max_col for _ in range(4)]
        for col, header in {
            "C": "So HD",
            "D": "Ngay HD",
            "I": "Ten nguoi mua",
            "J": "MST nguoi mua",
            "L": "Ma VT",
            "M": "Ten hang hoa",
            "N": "DVT",
            "O": "So luong",
            "P": "Don gia",
            "R": "Thue",
            "V": "Tien hang",
            "W": "Tien thue",
            "AQ": "Ma ngoai te",
            "AR": "Ty gia",
            "AS": "TK vat tu",
            "AT": "Ma kho",
            "AU": "Ma khach hang",
        }.items():
            rows[1][app.excel_col_to_index(col)] = header

        for row_index, customer_code, customer_mst, customer_name, ma_vt in [
            (2, "KH-A", "010A", "Customer A", "A.HANG"),
            (3, "KH-B", "010B", "Customer B", "B.HANG"),
        ]:
            for col, value in {
                "C": "777",
                "D": "20/01/2026",
                "I": customer_name,
                "J": customer_mst,
                "L": ma_vt,
                "M": f"Hang {customer_code}",
                "N": "Cai",
                "O": 1,
                "P": 1000,
                "R": 0.08,
                "V": 1000,
                "W": 80,
                "AQ": "VND",
                "AR": 1,
                "AS": "155",
                "AT": "KTP",
                "AU": customer_code,
            }.items():
                rows[row_index][app.excel_col_to_index(col)] = value

        sheets = app.fast_import_sheet_rows(None, app.pd.DataFrame(rows))
        self.assertIn("Hoadonbanhang-1", sheets)
        self.assertIn("Hoadonbanhang-2", sheets)
        self.assertNotIn("Hoadonbanhang", sheets)
        report_rows = sheets["Bao_cao_trung_so_ct"][1]
        self.assertEqual(len(report_rows), 2)
        self.assertEqual({row[0] for row in report_rows}, {"777"})
        self.assertEqual({row[1] for row in report_rows}, {"KH-A", "KH-B"})

    def test_son_phuong_allocation_builds_split_processed_sales_fdi(self):
        from web_api import son_phuong_processed_sales_dataframe

        max_col = app.excel_col_to_index("AU") + 1
        rows = [[""] * max_col for _ in range(3)]
        for col, header in {
            "C": "Số HĐ",
            "J": "MST người mua",
            "L": "Mã VT",
            "O": "Số lượng",
            "S": "Tiền hàng quy đổi",
            "V": "Tiền hàng",
            "W": "Tiền thuế",
            "AS": "TK vật tư",
            "AT": "Mã kho",
            "AU": "Mã khách hàng",
        }.items():
            rows[1][app.excel_col_to_index(col)] = header

        for col, value in {
            "C": "SP-001",
            "J": "010TEST",
            "L": "ONGHOP.CACLOAI",
            "O": 10,
            "S": 1000,
            "V": 1000,
            "W": 80,
        }.items():
            rows[2][app.excel_col_to_index(col)] = value

        result = {
            "allocations": [{
                "row_number": 3,
                "warehouse_code": "KHHVT",
                "warehouse_account": "156",
                "used": [
                    {"ledger_variant_code": "ONG.20X20", "quantity": 4},
                    {"ledger_variant_code": "ONG.30X30", "quantity": 6},
                ],
            }],
        }
        processed = son_phuong_processed_sales_dataframe(
            app.pd.DataFrame(rows),
            {"sales": {"header_row": 2, "data_start_row": 3, "code_col": "L", "qty_col": "O"}},
            result,
        )

        self.assertEqual(len(processed), 4)
        data_rows = processed.iloc[2:]
        self.assertEqual(set(data_rows.iloc[:, app.excel_col_to_index("L")]), {"ONG.20X20", "ONG.30X30"})
        self.assertAlmostEqual(float(data_rows.iloc[:, app.excel_col_to_index("O")].sum()), 10)
        self.assertAlmostEqual(float(data_rows.iloc[:, app.excel_col_to_index("S")].sum()), 1000)
        self.assertAlmostEqual(float(data_rows.iloc[:, app.excel_col_to_index("V")].sum()), 1000)
        self.assertAlmostEqual(float(data_rows.iloc[:, app.excel_col_to_index("W")].sum()), 80)
        self.assertEqual(set(data_rows.iloc[:, app.excel_col_to_index("AS")]), {"156"})
        self.assertEqual(set(data_rows.iloc[:, app.excel_col_to_index("AT")]), {"KHHVT"})
        self.assertEqual(set(data_rows.iloc[:, app.excel_col_to_index("AU")]), {"010TEST"})
        self.assertEqual(set(data_rows.iloc[:, -1]), {"materials"})

    def test_son_phuong_processed_sales_fdi_rejects_unresolved_material_shortage(self):
        from web_api import son_phuong_processed_sales_dataframe

        rows = [[""] * (app.excel_col_to_index("O") + 1) for _ in range(3)]
        rows[1][app.excel_col_to_index("L")] = "Mã VT"
        rows[1][app.excel_col_to_index("O")] = "Số lượng"
        rows[2][app.excel_col_to_index("L")] = "ONGHOP.CACLOAI"
        rows[2][app.excel_col_to_index("O")] = 10

        with self.assertRaisesRegex(ValueError, "còn thiếu KHHVT thực"):
            son_phuong_processed_sales_dataframe(
                app.pd.DataFrame(rows),
                {"sales": {"header_row": 2, "data_start_row": 3, "code_col": "L", "qty_col": "O"}},
                {"allocations": [{
                    "row_number": 3,
                    "allocation_role": "materials",
                    "unresolved_material_quantity": 2,
                    "used": [{"ledger_variant_code": "ONG.20X20", "quantity": 8}],
                }]},
            )

    def test_fast_import_validation_requires_inventory_columns(self):
        outputs = Path(__file__).parent / "outputs"
        outputs.mkdir(exist_ok=True)
        source = outputs / f"_test_fast_missing_inventory_{uuid4().hex}.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        for col, header in {
            "C": "Số HĐ",
            "L": "Mã VT",
            "O": "Số lượng",
        }.items():
            sheet.cell(2, app.excel_col_to_index(col) + 1, header)
        sheet.cell(3, app.excel_col_to_index("C") + 1, "1")
        sheet.cell(3, app.excel_col_to_index("L") + 1, "GIAYIN")
        sheet.cell(3, app.excel_col_to_index("O") + 1, 10)
        try:
            workbook.save(source)
            workbook.close()
            _, df = app.read_workbook(source)
            with self.assertRaises(ValueError) as context:
                app.validate_fast_import_processed_dataframe(df, "FDI test")
            message = str(context.exception)
            self.assertIn("TK vật tư", message)
            self.assertIn("Mã kho", message)
        finally:
            workbook.close()
            source.unlink(missing_ok=True)

    def test_fast_import_generic_mapping_can_skip_inventory_requirement(self):
        rows = [[""] * (app.excel_col_to_index("O") + 1) for _ in range(4)]
        for col, header in {
            "C": "Số HĐ",
            "F": "Tên người bán",
            "G": "MST người bán",
            "L": "Mã VT",
            "M": "Tên hàng",
            "O": "Số lượng",
        }.items():
            rows[1][app.excel_col_to_index(col)] = header
        for col, value in {
            "C": "SP-001",
            "F": "Công ty Sơn Phương Test",
            "G": "010TEST",
            "L": "SP.TEST",
            "M": "Hàng test",
            "O": 1,
        }.items():
            rows[2][app.excel_col_to_index(col)] = value

        stream = app.fast_import_multi_sheet_workbook(
            app.pd.DataFrame(rows),
            None,
            form_mapping_presets=[{
                "id": "generic_form",
                "label": "GenericForm",
                "scope": "purchase",
                "group_id": "materials",
                "output_columns": [
                    {"letter": "A", "header": "Mã khách hàng"},
                    {"letter": "B", "header": "Mã VT"},
                ],
                "mappings": [
                    {"target_col": "A", "source_col": "G", "source_phase": "purchase", "source_type": "source_column"},
                    {"target_col": "B", "source_col": "L", "source_phase": "purchase", "source_type": "source_column"},
                ],
            }],
            require_inventory_columns=False,
        )
        workbook = XlsWorkbookAdapter(stream.getvalue())
        try:
            self.assertIn("GenericForm", workbook.sheetnames)
            sheet = workbook["GenericForm"]
            self.assertEqual(sheet.cell(2, 1).value, "010TEST")
            self.assertEqual(sheet.cell(2, 2).value, "SP.TEST")
        finally:
            workbook.close()



if __name__ == "__main__":
    unittest.main()
