import json
import base64
import ipaddress
import os
import re
import shutil
import subprocess
import sys
import tempfile
import threading
import time
import unicodedata
import uuid
import webbrowser
import zipfile
from collections import Counter
from copy import copy
from difflib import SequenceMatcher
from functools import lru_cache
from io import BytesIO
from pathlib import Path
from urllib.error import HTTPError, URLError
from urllib.parse import urlparse
from urllib.request import Request, urlopen

import numpy as np
import pandas as pd
from flask import Flask, jsonify, request, send_file, send_from_directory
from flask.json.provider import DefaultJSONProvider

from product_code import code_normalization, config_store
from product_code import license_client as product_license_client
from product_code.release_version import NOTES as RELEASE_NOTES, VERSION as RELEASE_VERSION

APP_VERSION = RELEASE_VERSION
APP_RELEASE_NOTES = RELEASE_NOTES
CONFIG_SCHEMA_VERSION = 2
LEGACY_VIETMAX_STORAGE_PROFILES = {"vietmax_mua_vao", "vietmax_ban_ra"}
LOCAL_KEYGEN_HOSTS = {"localhost", "127.0.0.1", "::1"}
XLS_MAX_ROWS = 65536
XLS_MAX_COLUMNS = 256
XLS_MEDIA_TYPE = "application/vnd.ms-excel"
XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

class CustomJSONProvider(DefaultJSONProvider):
    def default(self, o):
        if isinstance(o, (np.integer,)):
            return int(o)
        if isinstance(o, (np.floating,)):
            return float(o)
        if isinstance(o, np.ndarray):
            return o.tolist()
        if isinstance(o, pd.Timestamp):
            return str(o)
        return super().default(o)


APP_DATA_DIR = Path(os.environ.get("LOCALAPPDATA") or Path.home() / "AppData" / "Local") / "ProductCodeFormatter"
LEGACY_REPO_CONFIG_PATH = Path(__file__).resolve().parent / "product_code_config.json"

if getattr(sys, "frozen", False):
    BASE_DIR = APP_DATA_DIR
    STATIC_DIR = Path(getattr(sys, "_MEIPASS", str(BASE_DIR))) / "static"
    RESOURCE_DIR = Path(getattr(sys, "_MEIPASS", str(BASE_DIR)))
else:
    BASE_DIR = Path(__file__).resolve().parent
    STATIC_DIR = BASE_DIR / "static"
    RESOURCE_DIR = BASE_DIR

UPLOAD_DIR = BASE_DIR / "uploads"
OUTPUT_DIR = BASE_DIR / "outputs"
CONFIG_PATH = APP_DATA_DIR / "product_code_config.json"
LICENSE_PATH = APP_DATA_DIR / "license.json"
DEFAULT_FORM_MAPPINGS_PATH = APP_DATA_DIR / "default_form_mappings.json"
CONFIG_BACKUP_DIR = APP_DATA_DIR / "config_backups"
TEMPLATE_DIR = APP_DATA_DIR / "templates"
ICON_PATH = RESOURCE_DIR / "app_icon.ico"
BASE_DIR.mkdir(parents=True, exist_ok=True)
APP_DATA_DIR.mkdir(parents=True, exist_ok=True)
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
TEMPLATE_DIR.mkdir(parents=True, exist_ok=True)

app = Flask(__name__, static_folder=None)
app.json_provider_class = CustomJSONProvider
app.json = CustomJSONProvider(app)

if not getattr(sys, "frozen", False):
    try:
        from flask_cors import CORS

        CORS(app)
    except ImportError:
        pass


PROFILE_LABELS = {
    "son_phuong": "Sơn Phương",
    "cao_thanh": "Cao Thành",
    "quang_thinh": "Quang Thịnh",
    "vietmax": "Vietmax",
    "ho_guom": "Hồ Gươm",
    "viet_hung": "Việt Hưng",
    "vietmax_mua_vao": "Vietmax mua vào",
    "vietmax_ban_ra": "Vietmax bán ra",
}

PROFILE_ALIASES = {
    "quang_thinh_1": "quang_thinh",
    "quang_thinh_2": "quang_thinh",
    "boc_tach_du_toan": "ho_guom",
}

VIETMAX_PROFILE = "vietmax"
VIETMAX_PHASE_PURCHASE = "purchase"
VIETMAX_PHASE_SALES = "sales"
VIETMAX_PROFILES = {"vietmax", "vietmax_mua_vao", "vietmax_ban_ra"}
PREFIX_STRATEGIES = ("last_2_words", "last_3_mst", "2_words_mst", "all_name_words")
VIETMAX_BAN_RA_MATCH_MA_KHO = "KVT"
VIETMAX_BAN_RA_MATCH_TK_VAT_TU = "152"
VIETMAX_COMPARISON_SCOPE_ALL_COMPANIES = "all_companies"
VIETMAX_COMPARISON_SCOPE_SAME_COMPANY = "same_company"
VIETMAX_COMPARISON_SCOPES = {VIETMAX_COMPARISON_SCOPE_ALL_COMPANIES, VIETMAX_COMPARISON_SCOPE_SAME_COMPANY}
VIETMAX_CONVERSION_MODE_NONE = "none"
VIETMAX_CONVERSION_MODE_QTY_AND_UNIT = "qty_and_unit"
VIETMAX_CONVERSION_MODE_QTY_ONLY = "qty_only"
VIETMAX_CONVERSION_MODE_UNIT_ONLY = "unit_only"
VIETMAX_CONVERSION_MODE_LABELS = {
    VIETMAX_CONVERSION_MODE_NONE: "Chỉ ghi chú, không đổi file",
    VIETMAX_CONVERSION_MODE_QTY_AND_UNIT: "Đổi số lượng + ĐVT",
    VIETMAX_CONVERSION_MODE_QTY_ONLY: "Chỉ đổi số lượng",
    VIETMAX_CONVERSION_MODE_UNIT_ONLY: "Chỉ đổi ĐVT",
}
VIETMAX_CONVERSION_MODES = set(VIETMAX_CONVERSION_MODE_LABELS)
VIETMAX_BAN_RA_FOCUS_PRODUCTS = [
    "Cabon 1 liên kích thước A4",
    "Cabon CB white 56/610*860_TL (R500)",
    "Cabon CF pink 56/610*860_TL (R500)",
    "Cabon CF yellow 56/650*860_TL (R500)",
    "Cabon CFB blue 50/610*860_TL (R500)",
    "Giấy An Hòa 92/70gms/790x1090",
    "Giấy An Hòa PP 92/70",
    "Cabon kích thước A4",
    "Cabon kích thước A5",
    "Giấy Couche",
    "Giấy Couche 300 gsm (62x86) cm",
    "Giấy in",
    "Giấy in couche",
    "Giấy in logo Desylia",
    "Giấy Ivory Ningbo 300gsm khổ 62 cm",
    "Giấy in offset 100gsm",
    "Giấy in offset 120gsm (790x1090mm)",
    "Giấy Ivory 300 gsm (79x109) cm",
    "Giấy không in",
    "Giấy không in KT 7,4x43,9cm",
    "Giấy mỹ thuật",
    "Giấy nháp",
    "Giấy Offset",
    "Giấy thi",
]

VIETMAX_PURCHASE_REVIEW_FORCED_PAIRS = (
    ("Giấy An Hòa PP 92/70", "Giấy An Hòa 92/70gms/790x1090"),
    ("Giấy in offset 100gsm", "Giấy In Offset"),
    ("Giấy Offset", "Giấy In Offset"),
    ("Giấy Couche", "Giấy in Couche"),
    ("Cabon CB white 56/610*860_TL (R500)", "Giấy Cacbon CB white 56/610*860_TL (R500)"),
    ("Cabon CF pink 56/610*860_TL (R500)", "Giấy Cacbon CF pink 56/610*860_TL (R500)"),
    ("Cabon CF yellow 56/650*860_TL (R500)", "Giấy Cacbon CF yellow 56/650*860_TL (R500)"),
    ("Cabon CFB blue 50/610*860_TL (R500)", "Giấy Cacbon CFB blue 50/610*860_TL (R500)"),
)
KEYGEN_ACTIVATION_REQUIRED_CODES = {
    "NO_MACHINE",
    "NO_MACHINES",
    "FINGERPRINT_SCOPE_MISMATCH",
    "MACHINE_SCOPE_MISMATCH",
}
DEFAULT_LICENSE_SERVER_IP = os.environ.get("PRODUCT_CODE_FORMATTER_LICENSE_SERVER_IP", "192.168.1.210")
DEFAULT_LICENSE_SERVER_IPS = (
    DEFAULT_LICENSE_SERVER_IP,
    "192.168.1.210",
    "192.168.101.13",
)
DEFAULT_LICENSE_SERVER_URL = f"http://{DEFAULT_LICENSE_SERVER_IP}:8080"
DEFAULT_KEYGEN_SERVER_URL = os.environ.get("PRODUCT_CODE_FORMATTER_KEYGEN_SERVER", DEFAULT_LICENSE_SERVER_URL)
DEFAULT_KEYGEN_ACCOUNT_ID = os.environ.get("PRODUCT_CODE_FORMATTER_KEYGEN_ACCOUNT", "6f1f56e8-3b6f-4a86-9a31-9e0e7f62c001")
DEFAULT_KEYGEN_PUBLIC_HOST = os.environ.get("PRODUCT_CODE_FORMATTER_KEYGEN_HOST", "license-server.local")

MAX_CODE_LENGTH = 50
DIAMETER_CHARS = "\u03a6\u03c6\u03d5\u00d8\u00f8\u2205\u2300\u0424\u0444\uff06"
DEFAULT_INVOICE_STATUS_COL = "AJ"
DEFAULT_INVOICE_STATUS_SKIP_VALUES = [
    "Hóa đơn đã bị hủy",
]
IGNORED_INVOICE_STATUSES = {
    "hoa don da bi huy",
}


def empty_profile_config(profile_key_name=None, include_scopes=True):
    config = {
        "prefixes": {},
        "selected_products": {},
        "removed_companies": {},
        "word_rules": {"\u0111en": "DEN", "t\u00f4n": "TON"},
        "first_word_rules": {},
        "repeated_phrase_removals": ["inox"] if profile_key_name == "cao_thanh" else [],
        "price_group_rules": {},
        "price_range_rules": {},
        "price_adjust_all_percent": 0,
        "manual_code_overrides": {},
        "product_code_replacements": {},
        "product_review_merges": [],
        "vietmax_mua_vao_internal_merges": [],
        "vietmax_ban_ra_sales_internal_merges": [],
        "vietmax_ban_ra_purchase_match_rules": [],
        "inventory_pairs": [],
        "use_default_inventory_pair": False,
        "default_inventory_pair_id": "",
        "inventory_pair_rules": [],
        "inventory_allocation_config": {},
        "include_company_prefix": True,
        "prefix_strategy": "last_2_words",
        "prefix_mst_digits": 3,
        "prefix_name_words": 2,
        "prefix_name_chars": 1,
        "prefix_missing_mst_strategy": "all_name_words",
        "prefix_strategy_values": {},
        "processing_groups": [],
        "company_group_assignments": {},
        "form_mapping_presets": [],
        "form_mapping_initialized": False,
        "columns": {},
    }
    if include_scopes:
        config["scopes"] = {
            VIETMAX_PHASE_PURCHASE: empty_profile_config(profile_key_name, include_scopes=False),
            VIETMAX_PHASE_SALES: empty_profile_config(profile_key_name, include_scopes=False),
        }
    return config


def default_config():
    return {
        "app_version": APP_VERSION,
        "config_schema_version": CONFIG_SCHEMA_VERSION,
        "selected_profile": "son_phuong",
        "license_server_ip": DEFAULT_LICENSE_SERVER_IP,
        "license_server_ips": list(dict.fromkeys(DEFAULT_LICENSE_SERVER_IPS)),
        "license": empty_license_config(),
        "profiles": {key: empty_profile_config(key) for key in PROFILE_LABELS},
        "columns": {
            "company_col": "F",
            "mst_col": "G",
            "address_col": "H",
            "product_col": "M",
            "qty_col": "O",
            "price_col": "",
            "purchase_price_col": "",
            "output_col": "L",
            "invoice_status_col": DEFAULT_INVOICE_STATUS_COL,
            "invoice_status_skip_values": DEFAULT_INVOICE_STATUS_SKIP_VALUES[:],
        },
    }


def empty_license_config():
    return {
        "server_url": "",
        "account_id": "",
        "license_key": "",
        "machine_fingerprint": "",
        "machine_id": "",
        "activated": False,
        "last_validated_at": "",
        "status": "",
        "product_code": "",
        "application": "",
        "allowed_companies": [],
        "allowed_profiles": [],
        "supported_profiles": [],
    }


def configured_license_server_urls(config=None):
    config = config if isinstance(config, dict) else {}
    candidates = []
    configured_ips = config.get("license_server_ips")
    if isinstance(configured_ips, list):
        candidates.extend(configured_ips)
    if not candidates:
        candidates.extend(DEFAULT_LICENSE_SERVER_IPS)
    legacy_ip = str(config.get("license_server_ip") or "").strip()
    if legacy_ip and legacy_ip not in candidates:
        candidates.append(legacy_ip)
    candidates.extend(DEFAULT_LICENSE_SERVER_IPS)
    urls = []
    seen = set()
    for candidate in candidates:
        value = str(candidate or "").strip().rstrip("/")
        if not value:
            continue
        if not value.startswith(("http://", "https://")):
            value = f"http://{value}:8080"
        if value not in seen:
            seen.add(value)
            urls.append(value)
    return urls or [DEFAULT_LICENSE_SERVER_URL]


def configured_license_server_url(config=None):
    return configured_license_server_urls(config)[0]
def profile_key(value):
    key = PROFILE_ALIASES.get(value, value)
    return key if key in PROFILE_LABELS else "son_phuong"


def normalize_vietmax_phase(value):
    value = str(value or "").strip().casefold()
    return VIETMAX_PHASE_SALES if value in {VIETMAX_PHASE_SALES, "ban_ra", "bán ra", "sales"} else VIETMAX_PHASE_PURCHASE


def effective_processing_profile(profile, vietmax_phase=None):
    profile = profile_key(profile)
    if profile != VIETMAX_PROFILE:
        return profile
    return "vietmax_ban_ra" if normalize_vietmax_phase(vietmax_phase) == VIETMAX_PHASE_SALES else "vietmax_mua_vao"


def normalize_removed_companies(value):
    if isinstance(value, dict):
        return {str(mst): bool(enabled) for mst, enabled in value.items() if enabled}
    if isinstance(value, list):
        return {str(mst): True for mst in value if str(mst).strip()}
    return {}


def normalize_products_map(value):
    result = {}
    if not isinstance(value, dict):
        return result
    for key, code in value.items():
        key = str(key)
        clean_code = sanitize_product_code(code)
        if "|||" in key:
            result[key] = clean_code
        elif "|" in key:
            mst, product = key.split("|", 1)
            result[product_key(mst, product)] = clean_code
    return result


def normalize_product_code_replacements(value):
    result = {}
    if not isinstance(value, dict):
        return result
    for old_code, new_code in value.items():
        old_clean = sanitize_product_code(old_code)
        new_clean = sanitize_product_code(new_code)
        if old_clean and new_clean:
            result[old_clean] = new_clean
    return result


def apply_product_code_replacement(code, replacements):
    clean_code = sanitize_product_code(code)
    normalized = normalize_product_code_replacements(replacements)
    if not clean_code or not normalized:
        return clean_code
    if clean_code in normalized:
        return sanitize_product_code(normalized[clean_code])
    if "." in clean_code:
        prefix, body = clean_code.rsplit(".", 1)
        replacement = normalized.get(body)
        if replacement:
            return sanitize_product_code(f"{prefix}.{replacement}")
    return clean_code


def legacy_word_rules(profile_key_name, profile):
    formula_options = profile.get("formula_options") or {}
    keep_words = formula_options.get(f"{profile_key_name}_keep_words")
    if keep_words is None and profile_key_name == "quang_thinh":
        keep_words = formula_options.get("son_phuong_keep_words")
    return keep_words


def normalize_price_group_rules(value, products=None):
    result = {}
    products = products or {}
    if not isinstance(value, dict):
        return result
    for key, rule in value.items():
        if not isinstance(rule, dict):
            continue
        normalized_key = str(key)
        if "|||" not in normalized_key and "|" in normalized_key:
            mst, product = normalized_key.split("|", 1)
            normalized_key = product_key(mst, product)
        try:
            min_price = float(rule.get("min_price"))
            max_price = float(rule.get("max_price"))
            percent = float(rule.get("percent", 8))
        except Exception:
            continue
        if min_price > max_price:
            min_price, max_price = max_price, min_price
        base_code = rule.get("base_code") or products.get(normalized_key) or ""
        groups = []
        for group in rule.get("groups") or []:
            if not isinstance(group, dict):
                continue
            try:
                group_min = float(group.get("min_price"))
                group_max = float(group.get("max_price"))
                adjust_percent = float(group.get("adjust_percent", 0))
            except Exception:
                continue
            if group_min > group_max:
                group_min, group_max = group_max, group_min
            groups.append({
                "index": int(group.get("index") or len(groups) + 1),
                "label": str(group.get("label") or f"Nhóm {len(groups) + 1}"),
                "min_price": group_min,
                "max_price": group_max,
                "average_price": float(group.get("average_price")) if group.get("average_price") is not None else None,
                "adjust_percent": adjust_percent,
            })
        result[normalized_key] = {
            "base_code": str(base_code),
            "min_price": min_price,
            "max_price": max_price,
            "percent": percent,
            "groups": groups,
        }
    return result


def normalize_phrase_list(value):
    result = []
    seen = set()
    if not isinstance(value, list):
        return result
    for item in value:
        phrase = str(item or "").strip()
        key = normalize_rule_key(phrase)
        if not phrase or not key or key in seen:
            continue
        seen.add(key)
        result.append(phrase)
    return result


def normalize_inventory_pairs(value):
    result = []
    seen = set()
    if not isinstance(value, list):
        return result
    for item in value:
        if not isinstance(item, dict):
            continue
        pair_id = str(item.get("id") or "").strip()
        if not pair_id or pair_id in seen:
            continue
        seen.add(pair_id)
        ma_kho = str(item.get("ma_kho") or "").strip()
        tk_vat_tu = str(item.get("tk_vat_tu") or "").strip()
        if re.fullmatch(r"\d+(?:\.\d+)?", ma_kho) and re.search(r"[A-Za-z]", tk_vat_tu):
            ma_kho, tk_vat_tu = tk_vat_tu, ma_kho
        result.append({
            "id": pair_id,
            "ma_kho": ma_kho,
            "tk_vat_tu": tk_vat_tu,
        })
    return result


def normalize_inventory_rule_priority(value):
    try:
        return int(float(value or 0))
    except (TypeError, ValueError):
        return 0


def normalize_inventory_pair_rules(value):
    result = []
    if not isinstance(value, list):
        return result
    for item in value:
        if not isinstance(item, dict):
            continue
        operator = str(item.get("operator") or "contains").strip().casefold()
        if operator not in {"contains", "equals"}:
            operator = "contains"
        result.append({
            "source_col": str(item.get("source_col") or "").strip().upper(),
            "operator": operator,
            "value": str(item.get("value") or "").strip(),
            "pair_id": str(item.get("pair_id") or "").strip(),
            "enabled": item.get("enabled") is not False,
            "priority": normalize_inventory_rule_priority(item.get("priority")),
        })
    return result


def normalize_string_list(value):
    if isinstance(value, str):
        items = re.split(r"[,;\n]+", value)
    elif isinstance(value, list):
        items = value
    else:
        items = []
    result = []
    seen = set()
    for item in items:
        text = str(item or "").strip()
        key = normalize_rule_key(text)
        if not text or not key or key in seen:
            continue
        seen.add(key)
        result.append(text)
    return result


def normalize_license_config(value):
    defaults = empty_license_config()
    value = value if isinstance(value, dict) else {}
    result = dict(defaults)
    for key in ["server_url", "account_id", "license_key", "machine_fingerprint", "machine_id", "last_validated_at", "status", "product_code", "application"]:
        result[key] = str(value.get(key) or "").strip()
    result["activated"] = bool(value.get("activated"))
    result["allowed_companies"] = normalize_string_list(value.get("allowed_companies") or [])
    result["allowed_profiles"] = normalize_string_list(value.get("allowed_profiles") or [])
    result["supported_profiles"] = normalize_string_list(value.get("supported_profiles") or [])
    return result


def current_machine_name():
    name = os.environ.get("COMPUTERNAME") or os.environ.get("HOSTNAME") or "windows"
    return str(name or "windows").strip() or "windows"


def legacy_machine_fingerprint():
    return f"{current_machine_name()}-{uuid.getnode():012x}".upper()


def windows_machine_guid():
    if sys.platform != "win32":
        return ""
    try:
        import winreg
        with winreg.OpenKey(winreg.HKEY_LOCAL_MACHINE, r"SOFTWARE\Microsoft\Cryptography") as key:
            value, _ = winreg.QueryValueEx(key, "MachineGuid")
    except Exception:
        return ""
    guid = re.sub(r"[^A-Za-z0-9-]+", "", raw_text(value)).strip("-")
    return guid.upper()


def local_machine_fingerprint():
    machine_guid = windows_machine_guid()
    if machine_guid:
        return f"{current_machine_name()}-WIN-{machine_guid}".upper()
    return legacy_machine_fingerprint()


def license_fingerprint_matches(saved_fingerprint):
    saved = str(saved_fingerprint or "").strip().upper()
    if not saved:
        return False
    if saved == local_machine_fingerprint():
        return True
    if saved == legacy_machine_fingerprint():
        return True
    machine_name = re.escape(current_machine_name().upper())
    return bool(re.fullmatch(rf"{machine_name}-[0-9A-F]{{12}}", saved))


def license_has_local_activation(value):
    license_cfg = normalize_license_config(value)
    if not license_cfg.get("activated"):
        return False
    return license_fingerprint_matches(license_cfg.get("machine_fingerprint"))


def license_company_match(value, allowed_companies):
    if not allowed_companies:
        return True
    key = normalize_rule_key(value)
    allowed = {normalize_rule_key(item) for item in allowed_companies if normalize_rule_key(item)}
    digits = re.sub(r"\D+", "", raw_text(value))
    allowed_digits = {re.sub(r"\D+", "", raw_text(item)) for item in allowed_companies}
    allowed_digits = {item for item in allowed_digits if item}
    return key in allowed or (bool(digits) and digits in allowed_digits)


def license_allows_company(company, allowed_companies):
    if not allowed_companies:
        return True
    names = [company.get("mst", ""), company.get("company", ""), *(company.get("all_names") or [])]
    return any(license_company_match(name, allowed_companies) for name in names)


def profile_match_value(value):
    key = profile_key(value)
    if key != "son_phuong" or normalize_rule_key(value) in {"son phuong", "son_phuong"}:
        return key
    normalized = normalize_rule_key(value).replace("_", " ")
    for profile, label in PROFILE_LABELS.items():
        if normalized in {normalize_rule_key(profile).replace("_", " "), normalize_rule_key(label)}:
            return profile
    return ""


def license_allows_profile(profile, allowed_profiles):
    if not allowed_profiles:
        return True
    current = profile_match_value(profile)
    allowed = {profile_match_value(item) for item in allowed_profiles}
    return bool(current and current in allowed)


def extract_allowed_companies(metadata):
    if not isinstance(metadata, dict):
        return []
    for key in ["allowed_companies", "allowedCompanies", "companies", "allowed_company_mst", "allowedCompanyMst", "allowed_mst", "allowedMst", "mst", "msts"]:
        if key in metadata:
            return normalize_string_list(metadata.get(key))
    return []


def extract_allowed_profiles(metadata):
    if not isinstance(metadata, dict):
        return []
    for key in ["allowed_profiles", "allowedProfiles", "profiles", "company_profiles", "companyProfiles"]:
        if key in metadata:
            return normalize_string_list(metadata.get(key))
    return []


def extract_supported_profiles(metadata):
    if not isinstance(metadata, dict):
        return []
    for key in ["supported_profiles", "supportedProfiles", "available_profiles", "availableProfiles"]:
        if key in metadata:
            return normalize_string_list(metadata.get(key))
    return []


def keygen_license_metadata(response):
    data = response.get("data") if isinstance(response, dict) else {}
    attributes = data.get("attributes") if isinstance(data, dict) else {}
    metadata = attributes.get("metadata") if isinstance(attributes, dict) else {}
    return metadata if isinstance(metadata, dict) else {}


def normalize_keygen_server_url(server_url):
    base = str(server_url or DEFAULT_KEYGEN_SERVER_URL).strip().rstrip("/")
    if base and not re.match(r"^[a-zA-Z][a-zA-Z0-9+.-]*://", base):
        base = f"http://{base}"
    return base


def normalize_keygen_account_id(account_id):
    return str(account_id or DEFAULT_KEYGEN_ACCOUNT_ID).strip().strip("/")


def keygen_url(server_url, account_id, path):
    base = normalize_keygen_server_url(server_url)
    account = normalize_keygen_account_id(account_id)
    if not base or not account:
        raise ValueError("Cần nhập địa chỉ server và account của Keygen.")
    parsed = urlparse(base)
    if parsed.scheme not in {"http", "https"}:
        raise ValueError("License server phải bắt đầu bằng http:// hoặc https://.")
    if parsed.scheme == "http" and not keygen_allows_http_host(parsed.hostname):
        raise ValueError("License server HTTP chỉ được phép cho localhost hoặc máy trong mạng LAN.")
    return f"{base}/v1/accounts/{account}/{path.lstrip('/')}"


def keygen_allows_http_host(hostname):
    host = str(hostname or "").strip().lower()
    if host in LOCAL_KEYGEN_HOSTS:
        return True
    try:
        address = ipaddress.ip_address(host)
    except ValueError:
        return host.endswith(".local") or "." not in host
    return address.is_private or address.is_loopback or address.is_link_local


def keygen_is_local_http_url(url):
    parsed = urlparse(str(url or ""))
    return parsed.scheme == "http" and keygen_allows_http_host(parsed.hostname)


def keygen_host_header(url):
    parsed = urlparse(str(url or ""))
    if parsed.scheme != "http" or not keygen_allows_http_host(parsed.hostname):
        return ""
    public_host = str(DEFAULT_KEYGEN_PUBLIC_HOST or "").strip()
    if not public_host:
        return ""
    public_parsed = urlparse(public_host if re.match(r"^[a-zA-Z][a-zA-Z0-9+.-]*://", public_host) else f"http://{public_host}")
    host = public_parsed.hostname or public_host.strip("/")
    port = public_parsed.port or parsed.port
    return f"{host}:{port}" if port else host


def keygen_forbidden_hint(url):
    parsed = urlparse(str(url or ""))
    host = str(parsed.hostname or "").strip().lower()
    if host in {"127.0.0.1", "localhost", "::1"}:
        return (
            " Keygen đang chặn host này. Hãy nhập IP thật của máy license server "
            "hoặc kiểm tra license_server/.env có KEYGEN_HOST=license-server.local."
        )
    return (
        " Kiểm tra License server có đúng KEYGEN_HOST trong license_server/.env "
        "và Account có đúng KEYGEN_ACCOUNT_ID không."
    )


def keygen_request(method, url, payload=None, license_key=None, timeout=10):
    data = None if payload is None else json.dumps(payload).encode("utf-8")
    headers = {
        "Accept": "application/vnd.api+json",
        "Content-Type": "application/vnd.api+json",
    }
    if keygen_is_local_http_url(url):
        headers["X-Forwarded-Proto"] = "https"
        host_header = keygen_host_header(url)
        if host_header:
            headers["Host"] = host_header
    if license_key:
        headers["Authorization"] = f"License {license_key}"
    request_obj = Request(url, data=data, headers=headers, method=method)
    try:
        with urlopen(request_obj, timeout=timeout) as response:
            body = response.read().decode("utf-8")
    except HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="replace")
        hint = keygen_forbidden_hint(url) if exc.code == 403 else ""
        raise ValueError(f"Keygen trả về lỗi {exc.code}: {detail or exc.reason}.{hint}") from exc
    except URLError as exc:
        raise ValueError(f"Không kết nối được license server: {exc.reason}") from exc
    return json.loads(body) if body else {}


def keygen_validate_license(server_url, account_id, license_key, fingerprint=None, timeout=10):
    key = str(license_key or "").strip()
    if not key:
        raise ValueError("Cần nhập license key.")
    fingerprint = str(fingerprint or local_machine_fingerprint()).strip()
    payload = {"meta": {"key": key, "scope": {"fingerprint": fingerprint}}}
    response = keygen_request("POST", keygen_url(server_url, account_id, "licenses/actions/validate-key"), payload, timeout=timeout)
    meta = response.get("meta") if isinstance(response, dict) else {}
    data = response.get("data") if isinstance(response, dict) else {}
    attributes = data.get("attributes") if isinstance(data, dict) else {}
    metadata = keygen_license_metadata(response)
    return {
        "valid": meta.get("valid") is not False,
        "code": str(meta.get("code") or ""),
        "detail": str(meta.get("detail") or ""),
        "license_id": str(data.get("id") or "") if isinstance(data, dict) else "",
        "status": str((attributes or {}).get("status") or meta.get("code") or "valid"),
        "product_code": raw_text(metadata.get("product_code") or metadata.get("productCode")),
        "application": raw_text(metadata.get("application")),
        "allowed_companies": extract_allowed_companies(metadata),
        "allowed_profiles": extract_allowed_profiles(metadata),
        "supported_profiles": extract_supported_profiles(metadata),
        "raw": response,
    }


def keygen_activate_machine(server_url, account_id, license_key, license_id, fingerprint=None, timeout=10):
    fingerprint = str(fingerprint or local_machine_fingerprint()).strip()
    license_id = str(license_id or "").strip()
    if not license_id:
        raise ValueError("Không tìm thấy license id để kích hoạt máy.")
    payload = {
        "data": {
            "type": "machines",
            "attributes": {
                "fingerprint": fingerprint,
                "name": os.environ.get("COMPUTERNAME") or "ProductCodeFormatter",
                "platform": sys.platform,
            },
            "relationships": {
                "license": {
                    "data": {"type": "licenses", "id": license_id}
                }
            },
        }
    }
    try:
        response = keygen_request("POST", keygen_url(server_url, account_id, "machines"), payload, license_key=license_key, timeout=timeout)
    except ValueError as exc:
        if "409" not in str(exc):
            raise
        return ""
    data = response.get("data") if isinstance(response, dict) else {}
    return str(data.get("id") or "") if isinstance(data, dict) else ""


def activate_keygen_license(server_url, account_id, license_key, timeout=10):
    fingerprint = local_machine_fingerprint()
    validation = keygen_validate_license(server_url, account_id, license_key, fingerprint, timeout)
    machine_id = ""
    if not validation["valid"]:
        if validation["code"] not in KEYGEN_ACTIVATION_REQUIRED_CODES:
            raise ValueError(validation["detail"] or validation["code"] or "License không hợp lệ.")
        machine_id = keygen_activate_machine(server_url, account_id, license_key, validation["license_id"], fingerprint, timeout)
        validation = keygen_validate_license(server_url, account_id, license_key, fingerprint, timeout)
    if not validation["valid"]:
        raise ValueError(validation["detail"] or validation["code"] or "License chưa hợp lệ sau khi kích hoạt máy.")
    return {
        "server_url": normalize_keygen_server_url(server_url),
        "account_id": normalize_keygen_account_id(account_id),
        "license_key": str(license_key or "").strip(),
        "machine_fingerprint": fingerprint,
        "machine_id": machine_id,
        "activated": True,
        "last_validated_at": time.strftime("%Y-%m-%d %H:%M:%S"),
        "status": validation["status"],
        "product_code": validation["product_code"],
        "application": validation["application"],
        "allowed_companies": validation["allowed_companies"],
        "allowed_profiles": validation["allowed_profiles"],
        "supported_profiles": validation["supported_profiles"],
    }


# Compatibility facade: runtime Keygen behavior now lives in
# product_code/license_client.py. Keep the legacy function names stable for
# existing routes/tests while future edits can target the smaller module.
current_machine_name = product_license_client.current_machine_name
legacy_machine_fingerprint = product_license_client.legacy_machine_fingerprint
windows_machine_guid = product_license_client.windows_machine_guid
local_machine_fingerprint = product_license_client.local_machine_fingerprint
keygen_license_metadata = product_license_client.keygen_license_metadata
normalize_keygen_server_url = product_license_client.normalize_keygen_server_url
normalize_keygen_account_id = product_license_client.normalize_keygen_account_id
keygen_url = product_license_client.keygen_url
keygen_allows_http_host = product_license_client.keygen_allows_http_host
keygen_is_local_http_url = product_license_client.keygen_is_local_http_url
keygen_host_header = product_license_client.keygen_host_header
keygen_forbidden_hint = product_license_client.keygen_forbidden_hint
extract_allowed_companies = product_license_client.extract_allowed_companies
extract_allowed_profiles = product_license_client.extract_allowed_profiles
extract_supported_profiles = product_license_client.extract_supported_profiles


def keygen_request(method, url, payload=None, license_key=None, timeout=10):
    return product_license_client.keygen_request(
        method,
        url,
        payload=payload,
        license_key=license_key,
        timeout=timeout,
        opener=urlopen,
    )


def keygen_validate_license(server_url, account_id, license_key, fingerprint=None, timeout=10):
    return product_license_client.keygen_validate_license(
        server_url,
        account_id,
        license_key,
        fingerprint=fingerprint,
        timeout=timeout,
        requester=keygen_request,
        fingerprint_func=local_machine_fingerprint,
    )


def keygen_activate_machine(server_url, account_id, license_key, license_id, fingerprint=None, timeout=10):
    return product_license_client.keygen_activate_machine(
        server_url,
        account_id,
        license_key,
        license_id,
        fingerprint=fingerprint,
        timeout=timeout,
        requester=keygen_request,
        fingerprint_func=local_machine_fingerprint,
    )


def activate_keygen_license(server_url, account_id, license_key, timeout=10):
    return product_license_client.activate_keygen_license(
        server_url,
        account_id,
        license_key,
        timeout=timeout,
        requester=keygen_request,
        fingerprint_func=local_machine_fingerprint,
    )


def price_ranges_from_groups(price_group_rules):
    ranges = {}
    for rule in price_group_rules.values():
        base_code = str(rule.get("base_code") or "").strip()
        if not base_code:
            continue
        ranges[base_code] = {
            "min_price": rule.get("min_price"),
            "max_price": rule.get("max_price"),
            "percent": rule.get("percent", 8),
        }
    return normalize_price_range_rules(ranges)


def legacy_price_adjust_all_percent(profile, price_groups, price_ranges):
    if "price_adjust_all_percent" in profile:
        return float(profile.get("price_adjust_all_percent") or 0)
    margins = set()
    rule_sets = [price_ranges.values(), price_groups.values()]
    for rules in rule_sets:
        for rule in rules:
            for group in rule.get("groups") or []:
                try:
                    margins.add(float(group.get("adjust_percent") or 0))
                except Exception:
                    continue
    return margins.pop() if len(margins) == 1 else 0


def normalize_prefix_strategy(value):
    value = str(value or "").strip()
    return value if value in PREFIX_STRATEGIES else "last_2_words"


def normalize_missing_mst_prefix_strategy(value):
    value = str(value or "").strip()
    return value if value in {"last_2_words", "all_name_words"} else "all_name_words"


def normalize_prefix_mst_digits(value):
    try:
        digits = int(value or 3)
    except Exception:
        digits = 3
    return max(1, min(10, digits))


def normalize_prefix_name_words(value):
    try:
        words = int(value or 2)
    except Exception:
        words = 2
    return max(1, min(20, words))


def normalize_prefix_name_chars(value):
    try:
        chars = int(value or 1)
    except Exception:
        chars = 1
    return max(1, min(10, chars))


def normalize_prefix_strategy_values(raw):
    values = {strategy: {} for strategy in PREFIX_STRATEGIES}
    if not isinstance(raw, dict):
        return values
    for strategy in PREFIX_STRATEGIES:
        strategy_values = raw.get(strategy)
        if not isinstance(strategy_values, dict):
            continue
        values[strategy] = {
            str(key): str(value).strip().upper()
            for key, value in strategy_values.items()
            if key and value is not None and str(value).strip()
        }
    return values


REVIEW_MERGE_CONFIG_TEXT_FIELDS = (
    "product",
    "similar_product",
    "unit",
    "similar_unit",
    "company",
    "mst",
    "company_key",
    "similar_company",
    "similar_mst",
    "similar_company_key",
    "product_key",
    "similar_product_key",
    "review_group",
    "review_type",
)


def _config_text(value):
    return str(value or "").strip()


def _config_review_scope(value, fallback=VIETMAX_COMPARISON_SCOPE_ALL_COMPANIES):
    value = _config_text(value)
    fallback = fallback if fallback in VIETMAX_COMPARISON_SCOPES else VIETMAX_COMPARISON_SCOPE_ALL_COMPANIES
    return value if value in VIETMAX_COMPARISON_SCOPES else fallback


def _config_review_code_choice(value):
    value = _config_text(value).lower()
    return value if value in {"current", "similar", "split"} else "current"


def sanitize_review_merge_rule_for_config(item, comparison_scope=VIETMAX_COMPARISON_SCOPE_ALL_COMPANIES):
    if not isinstance(item, dict):
        return None
    product = _config_text(item.get("product"))
    similar_product = _config_text(item.get("similar_product"))
    if not product or not similar_product:
        return None
    code_choice = _config_review_code_choice(item.get("code_choice"))
    rule = {
        "product": product,
        "similar_product": similar_product,
        "confirmed": item.get("confirmed") is True,
        "code_choice": code_choice,
        "comparison_scope": _config_review_scope(item.get("comparison_scope"), comparison_scope),
    }
    for field in REVIEW_MERGE_CONFIG_TEXT_FIELDS:
        if field in {"product", "similar_product"}:
            continue
        text = _config_text(item.get(field))
        if text:
            rule[field] = text
    if item.get("dimension_only"):
        rule["dimension_only"] = True
    # Literal manual codes are only valid for the explicit split action.
    # Normal merge rules persist the selected side via code_choice and
    # product -> similar_product, then regenerate current codes at runtime.
    if code_choice == "split":
        split_code = sanitize_product_code(item.get("split_code"))
        similar_split_code = sanitize_product_code(item.get("similar_split_code"))
        if split_code:
            rule["split_code"] = split_code
        if similar_split_code:
            rule["similar_split_code"] = similar_split_code
    return rule


def sanitize_review_merge_rules_for_config(value, comparison_scope=VIETMAX_COMPARISON_SCOPE_ALL_COMPANIES):
    if not isinstance(value, list):
        return []
    result = []
    seen = set()
    for item in value:
        rule = sanitize_review_merge_rule_for_config(item, comparison_scope)
        if not rule:
            continue
        key = (
            rule.get("comparison_scope"),
            rule.get("review_group", ""),
            rule.get("product", ""),
            rule.get("similar_product", ""),
            rule.get("company_key", ""),
            rule.get("similar_company_key", ""),
        )
        if key in seen:
            continue
        seen.add(key)
        result.append(rule)
    return result


BUILTIN_PROCESSING_GROUP_LABELS = {
    "materials": "Nhóm vật tư",
    "services": "Nhóm dịch vụ",
    "ignored": "Không xử lý",
}


def normalize_processing_groups_for_config(value):
    if not isinstance(value, list):
        return []
    result = []
    seen = set()
    for item in value:
        if not isinstance(item, dict):
            continue
        group_id = _config_text(item.get("id"))
        if not group_id or group_id in seen:
            continue
        seen.add(group_id)
        group = dict(item)
        group["id"] = group_id
        if group_id in BUILTIN_PROCESSING_GROUP_LABELS:
            group["label"] = BUILTIN_PROCESSING_GROUP_LABELS[group_id]
            group["builtin"] = True
        else:
            group["label"] = _config_text(group.get("label")) or group_id
        group["forms"] = list(group.get("forms") or []) if isinstance(group.get("forms"), list) else []
        group["uses_product_code"] = group.get("uses_product_code") is not False
        result.append(group)
    return result


def normalize_profile_config(profile_key_name, profile, include_scopes=True):
    defaults = empty_profile_config(profile_key_name, include_scopes=include_scopes)
    profile = profile if isinstance(profile, dict) else {}
    prefixes = profile.get("prefixes") or profile.get("companies") or {}
    products = normalize_products_map(profile.get("manual_code_overrides") or profile.get("products") or {})
    product_code_replacements = normalize_product_code_replacements(profile.get("product_code_replacements") or {})
    word_rules_source = profile.get("word_rules")
    first_word_rules_source = profile.get("first_word_rules") or {}
    if word_rules_source is None:
        word_rules_source = legacy_word_rules(profile_key_name, profile)
    if word_rules_source is None:
        word_rules_source = defaults["word_rules"]
    price_groups = normalize_price_group_rules(profile.get("price_group_rules") or {}, products)
    price_ranges = normalize_price_range_rules(profile.get("price_range_rules") or {})
    if not price_ranges:
        price_ranges = price_ranges_from_groups(price_groups)
    normalized = {
        "prefixes": {str(mst): str(prefix) for mst, prefix in dict(prefixes).items()} if isinstance(prefixes, dict) else {},
        "selected_products": dict(profile.get("selected_products") or {}),
        "removed_companies": normalize_removed_companies(profile.get("removed_companies")),
        "word_rules": normalize_word_rules(word_rules_source),
        "first_word_rules": normalize_word_rules(first_word_rules_source),
        "repeated_phrase_removals": normalize_phrase_list(profile.get("repeated_phrase_removals", defaults["repeated_phrase_removals"])),
        "price_group_rules": price_groups,
        "price_range_rules": price_ranges,
        "price_adjust_all_percent": legacy_price_adjust_all_percent(profile, price_groups, price_ranges),
        "manual_code_overrides": products,
        "product_code_replacements": product_code_replacements,
        "product_review_merges": sanitize_review_merge_rules_for_config(profile.get("product_review_merges") or []),
        "vietmax_mua_vao_internal_merges": sanitize_review_merge_rules_for_config(profile.get("vietmax_mua_vao_internal_merges") or []),
        "vietmax_ban_ra_sales_internal_merges": sanitize_review_merge_rules_for_config(profile.get("vietmax_ban_ra_sales_internal_merges") or []),
        "vietmax_ban_ra_purchase_match_rules": list(profile.get("vietmax_ban_ra_purchase_match_rules") or profile.get("sales_match_rules") or []),
        "inventory_pairs": normalize_inventory_pairs(profile.get("inventory_pairs") or []),
        "use_default_inventory_pair": bool(profile.get("use_default_inventory_pair")),
        "default_inventory_pair_id": str(profile.get("default_inventory_pair_id") or "").strip(),
        "inventory_pair_rules": normalize_inventory_pair_rules(profile.get("inventory_pair_rules") or []),
        "inventory_allocation_config": dict(profile.get("inventory_allocation_config") or {}) if isinstance(profile.get("inventory_allocation_config"), dict) else {},
        "include_company_prefix": profile.get("include_company_prefix") is not False,
        "prefix_strategy": normalize_prefix_strategy(profile.get("prefix_strategy") or defaults["prefix_strategy"]),
        "prefix_mst_digits": normalize_prefix_mst_digits(profile.get("prefix_mst_digits", defaults["prefix_mst_digits"])),
        "prefix_name_words": normalize_prefix_name_words(profile.get("prefix_name_words", defaults["prefix_name_words"])),
        "prefix_name_chars": normalize_prefix_name_chars(profile.get("prefix_name_chars", defaults["prefix_name_chars"])),
        "prefix_missing_mst_strategy": normalize_missing_mst_prefix_strategy(profile.get("prefix_missing_mst_strategy", defaults["prefix_missing_mst_strategy"])),
        "prefix_strategy_values": normalize_prefix_strategy_values(profile.get("prefix_strategy_values") or defaults["prefix_strategy_values"]),
        "processing_groups": normalize_processing_groups_for_config(profile.get("processing_groups") or []),
        "company_group_assignments": dict(profile.get("company_group_assignments") or {}),
        "form_mapping_presets": list(profile.get("form_mapping_presets") or []),
        "form_mapping_initialized": bool(profile.get("form_mapping_initialized")) or bool(profile.get("form_mapping_presets")),
        "columns": dict(profile.get("columns") or {}) if isinstance(profile.get("columns"), dict) else {},
    }
    if include_scopes:
        raw_scopes = profile.get("scopes") if isinstance(profile.get("scopes"), dict) else {}
        normalized["scopes"] = {
            VIETMAX_PHASE_PURCHASE: normalize_profile_config(
                profile_key_name,
                raw_scopes.get(VIETMAX_PHASE_PURCHASE) or raw_scopes.get("mua_vao") or raw_scopes.get("muavao") or {},
                include_scopes=False,
            ),
            VIETMAX_PHASE_SALES: normalize_profile_config(
                profile_key_name,
                raw_scopes.get(VIETMAX_PHASE_SALES) or raw_scopes.get("ban_ra") or raw_scopes.get("banra") or {},
                include_scopes=False,
            ),
        }
    return normalized


def normalize_config(data):
    cfg = default_config()
    cfg["app_version"] = APP_VERSION
    cfg["config_schema_version"] = CONFIG_SCHEMA_VERSION
    if isinstance(data, dict):
        selected = data.get("selected_profile") or data.get("active_profile") or data.get("format_rule")
        cfg["selected_profile"] = profile_key(selected) if selected else cfg["selected_profile"]
        configured_ips = data.get("license_server_ips")
        if isinstance(configured_ips, list):
            configured_ips = [str(item or "").strip() for item in configured_ips if str(item or "").strip()]
        else:
            configured_ips = []
        if not configured_ips:
            configured_ips.extend(DEFAULT_LICENSE_SERVER_IPS)
        legacy_ip = str(data.get("license_server_ip") or "").strip()
        if legacy_ip and legacy_ip not in configured_ips:
            configured_ips.append(legacy_ip)
        configured_ips.extend(item for item in DEFAULT_LICENSE_SERVER_IPS if item not in configured_ips)
        cfg["license_server_ips"] = list(dict.fromkeys(configured_ips))
        cfg["license_server_ip"] = cfg["license_server_ips"][0] if cfg["license_server_ips"] else DEFAULT_LICENSE_SERVER_IP
        cfg["license"] = normalize_license_config(data.get("license") or {})
        cfg["columns"].update(data.get("columns") or {})
        cfg["columns"].setdefault("invoice_status_col", DEFAULT_INVOICE_STATUS_COL)
        if not isinstance(cfg["columns"].get("invoice_status_skip_values"), list):
            cfg["columns"]["invoice_status_skip_values"] = DEFAULT_INVOICE_STATUS_SKIP_VALUES[:]
        profiles = data.get("profiles") or {}
        if not isinstance(profiles, dict):
            profiles = {}
        if not profiles and any(key in data for key in ["companies", "products", "removed_companies", "formula_options", "price_group_rules"]):
            profiles = {cfg["selected_profile"]: data}
        for alias, target in PROFILE_ALIASES.items():
            if alias in profiles and target not in profiles:
                profiles[target] = profiles[alias]
        for key in PROFILE_LABELS:
            cfg["profiles"][key].update(normalize_profile_config(key, profiles.get(key) or {}))
        vietmax_profile = cfg["profiles"].get(VIETMAX_PROFILE) or empty_profile_config(VIETMAX_PROFILE)
        raw_vietmax = profiles.get(VIETMAX_PROFILE) if isinstance(profiles.get(VIETMAX_PROFILE), dict) else {}
        raw_scopes = raw_vietmax.get("scopes") if isinstance(raw_vietmax.get("scopes"), dict) else {}
        for phase, legacy_key in (
            (VIETMAX_PHASE_PURCHASE, "vietmax_mua_vao"),
            (VIETMAX_PHASE_SALES, "vietmax_ban_ra"),
        ):
            raw_scope = raw_scopes.get(phase) if isinstance(raw_scopes.get(phase), dict) else {}
            legacy_profile = profiles.get(legacy_key) if isinstance(profiles.get(legacy_key), dict) else {}
            if not profile_config_has_user_data(raw_scope) and profile_config_has_user_data(legacy_profile):
                vietmax_profile["scopes"][phase] = normalize_profile_config(
                    VIETMAX_PROFILE,
                    legacy_profile,
                    include_scopes=False,
                )
        cfg["profiles"][VIETMAX_PROFILE] = vietmax_profile
    if cfg["selected_profile"] not in PROFILE_LABELS:
        cfg["selected_profile"] = "son_phuong"
    cfg["columns"].setdefault("invoice_status_col", DEFAULT_INVOICE_STATUS_COL)
    if not isinstance(cfg["columns"].get("invoice_status_skip_values"), list):
        cfg["columns"]["invoice_status_skip_values"] = DEFAULT_INVOICE_STATUS_SKIP_VALUES[:]
    return cfg


def profile_config_has_user_data(profile):
    if not isinstance(profile, dict):
        return False
    ignored = {
        "scopes",
        "prefix_strategy",
        "prefix_mst_digits",
        "prefix_name_words",
        "prefix_name_chars",
        "prefix_missing_mst_strategy",
        "include_company_prefix",
        "price_adjust_all_percent",
        "use_default_inventory_pair",
        "default_inventory_pair_id",
    }
    for key, value in profile.items():
        if key in ignored:
            continue
        if isinstance(value, (dict, list, tuple, set)) and value:
            return True
        if isinstance(value, str) and value.strip():
            return True
        if isinstance(value, (int, float)) and value:
            return True
    return False


def canonical_config_for_storage(cfg):
    stored = json.loads(json.dumps(cfg, ensure_ascii=False, default=raw_text))
    stored["config_schema_version"] = CONFIG_SCHEMA_VERSION
    stored.pop("license", None)
    strip_system_default_forms_from_profiles(stored)
    profiles = stored.get("profiles") if isinstance(stored.get("profiles"), dict) else {}
    for key in LEGACY_VIETMAX_STORAGE_PROFILES:
        profiles.pop(key, None)
    stored["profiles"] = profiles
    return stored


def persist_configured_form_templates(cfg):
    profiles = cfg.get("profiles") if isinstance(cfg, dict) and isinstance(cfg.get("profiles"), dict) else {}

    def persist_profile(profile):
        if not isinstance(profile, dict):
            return
        forms = profile.get("form_mapping_presets") if isinstance(profile.get("form_mapping_presets"), list) else []
        for form in forms:
            if not isinstance(form, dict):
                continue
            saved_name = Path(str(form.get("template_saved_name") or "")).name
            if not saved_name:
                continue
            persistent_path = TEMPLATE_DIR / saved_name
            legacy_path = UPLOAD_DIR / saved_name
            if not persistent_path.exists() and legacy_path.exists() and legacy_path.is_file():
                try:
                    shutil.copy2(legacy_path, persistent_path)
                except OSError:
                    pass
        scopes = profile.get("scopes") if isinstance(profile.get("scopes"), dict) else {}
        for scoped in scopes.values():
            persist_profile(scoped)

    for profile in profiles.values():
        persist_profile(profile)
    return cfg


def license_backup_config_paths():
    paths = []
    backup_names = [
        f"{CONFIG_PATH.stem}.manual_code_overrides_backup.json",
        f"{CONFIG_PATH.stem}.backup.json",
    ]
    for name in backup_names:
        paths.append(CONFIG_PATH.with_name(name))
        paths.append(LEGACY_REPO_CONFIG_PATH.with_name(name))
    paths.append(LEGACY_REPO_CONFIG_PATH)
    try:
        paths.extend(CONFIG_PATH.parent.glob(f"{CONFIG_PATH.stem}*.json"))
    except Exception:
        pass
    try:
        paths.extend(LEGACY_REPO_CONFIG_PATH.parent.glob(f"{LEGACY_REPO_CONFIG_PATH.stem}*.json"))
    except Exception:
        pass
    unique = []
    seen = {str(CONFIG_PATH.resolve()).casefold()} if CONFIG_PATH.exists() else {str(CONFIG_PATH).casefold()}
    for path in paths:
        try:
            key = str(path.resolve()).casefold()
        except Exception:
            key = str(path).casefold()
        if key in seen:
            continue
        seen.add(key)
        unique.append(path)
    return sorted(unique, key=lambda item: item.stat().st_mtime if item.exists() else 0, reverse=True)


def recover_license_from_backup_configs(cfg):
    if license_has_local_activation((cfg or {}).get("license")):
        return cfg, False
    for path in license_backup_config_paths():
        if not path.exists() or not path.is_file():
            continue
        try:
            backup = json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            continue
        if not isinstance(backup, dict):
            continue
        license_cfg = normalize_license_config(backup.get("license") or {})
        if not license_has_local_activation(license_cfg):
            continue
        recovered = dict(cfg or default_config())
        recovered["license"] = license_cfg
        return recovered, True
    return cfg, False


def migrate_config_file_if_needed():
    if not CONFIG_PATH.exists() or not CONFIG_PATH.is_file():
        return None
    try:
        raw = json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
    except Exception:
        return None
    try:
        schema_version = int(raw.get("config_schema_version") or 0)
    except (TypeError, ValueError):
        schema_version = 0
    if schema_version >= CONFIG_SCHEMA_VERSION:
        return None
    backup_path = config_store.backup_config_file(CONFIG_PATH, CONFIG_BACKUP_DIR)
    persist_configured_form_templates(raw)
    if not LICENSE_PATH.exists() and isinstance(raw.get("license"), dict):
        save_license_config(raw.get("license"))
    config_store.save_config_file(
        CONFIG_PATH,
        raw,
        normalize_config,
        serializer=canonical_config_for_storage,
    )
    return backup_path


def load_config():
    if not CONFIG_PATH.exists() and LEGACY_REPO_CONFIG_PATH.exists():
        try:
            legacy = json.loads(LEGACY_REPO_CONFIG_PATH.read_text(encoding="utf-8"))
            config_store.backup_config_file(LEGACY_REPO_CONFIG_PATH, CONFIG_BACKUP_DIR)
            persist_configured_form_templates(legacy)
            if not LICENSE_PATH.exists() and isinstance(legacy.get("license"), dict):
                save_license_config(legacy.get("license"))
            config_store.save_config_file(
                CONFIG_PATH,
                legacy,
                normalize_config,
                serializer=canonical_config_for_storage,
            )
        except Exception:
            pass
    migrate_config_file_if_needed()
    cfg = config_store.load_config_file(CONFIG_PATH, default_config, normalize_config)
    if LICENSE_PATH.exists():
        license_cfg = load_license_config()
    else:
        cfg, _ = recover_license_from_backup_configs(cfg)
        license_cfg = save_license_config(cfg.get("license") or {})

    embedded_license = False
    if CONFIG_PATH.exists():
        try:
            embedded_license = "license" in json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
        except Exception:
            embedded_license = False
    if embedded_license:
        config_store.backup_config_file(CONFIG_PATH, CONFIG_BACKUP_DIR)
        config_store.save_config_file(
            CONFIG_PATH,
            cfg,
            normalize_config,
            serializer=canonical_config_for_storage,
        )
    cfg["license"] = license_cfg
    return apply_system_default_forms_to_config(cfg)


def save_config(cfg):
    cfg = apply_system_default_forms_to_config(cfg)
    persist_configured_form_templates(cfg)
    if isinstance(cfg, dict) and "license" in cfg:
        save_license_config(cfg.get("license") or {})
    saved = config_store.save_config_file(
        CONFIG_PATH,
        cfg,
        normalize_config,
        serializer=canonical_config_for_storage,
    )
    saved["license"] = load_license_config()
    return apply_system_default_forms_to_config(saved)


def normalize_word_rules(value):
    if isinstance(value, list):
        result = {}
        for item in value:
            if isinstance(item, dict):
                word = str(item.get("word") or "").strip()
                output = str(item.get("output") or "").strip()
            else:
                word = str(item or "").strip()
                output = ""
            if word:
                result[word] = normalize_token(output or word)
        return result
    if isinstance(value, dict):
        result = {}
        for word, output in value.items():
            word = "đen" if str(word).upper() == "DEN" else "tôn" if str(word).upper() == "TON" else str(word).strip()
            if word:
                result[word] = normalize_token(str(output or word))
        return result
    return {}


def normalize_price_range_rules(value):
    result = {}
    if not isinstance(value, dict):
        return result
    for key, rule in value.items():
        if not isinstance(rule, dict):
            continue
        try:
            min_price = float(rule.get("min_price"))
            max_price = float(rule.get("max_price"))
            percent = float(rule.get("percent", 8))
        except Exception:
            continue
        if min_price > max_price:
            min_price, max_price = max_price, min_price
        groups = []
        for group in rule.get("groups") or []:
            if not isinstance(group, dict):
                continue
            try:
                group_min = float(group.get("min_price"))
                group_max = float(group.get("max_price"))
                adjust_percent = float(group.get("adjust_percent", 0))
            except Exception:
                continue
            if group_min > group_max:
                group_min, group_max = group_max, group_min
            groups.append({
                "index": int(group.get("index") or len(groups) + 1),
                "label": str(group.get("label") or f"Nhóm {len(groups) + 1}"),
                "min_price": group_min,
                "max_price": group_max,
                "average_price": float(group.get("average_price")) if group.get("average_price") is not None else None,
                "adjust_percent": adjust_percent,
            })
        result[str(key)] = {
            "min_price": min_price,
            "max_price": max_price,
            "percent": percent,
            "groups": groups,
        }
    return result


@app.errorhandler(Exception)
def handle_exception(e):
    return jsonify(error=str(e)), 500


def excel_col_to_index(col):
    col = str(col).strip().upper()
    if not re.fullmatch(r"[A-Z]+", col):
        raise ValueError(f"Invalid Excel column: {col}")
    n = 0
    for ch in col:
        n = n * 26 + ord(ch) - 64
    return n - 1


def index_to_excel_col(idx):
    idx += 1
    s = ""
    while idx:
        idx, rem = divmod(idx - 1, 26)
        s = chr(65 + rem) + s
    return s


def rm_accents(text):
    text = str(text)
    text = text.replace("Ð", "D").replace("ð", "d")
    text = text.replace("Đ", "D").replace("đ", "d").replace("\u00c4\u0090", "D").replace("\u00c4\u2018", "d")
    for ch in DIAMETER_CHARS:
        text = text.replace(ch, "F")
    text = unicodedata.normalize("NFD", text)
    return "".join(ch for ch in text if unicodedata.category(ch) != "Mn")


def normalize_code_decimal_separators(text):
    return re.sub(r"(?<=\d),(?=\d)", ".", str(text or ""))


def normalize_token(text):
    text = normalize_code_decimal_separators(rm_accents(text).upper())
    text = re.sub(r"[^A-Z0-9.]+", "", text)
    return text


def normalize_code_token(text, keep_slash=False, keep_hyphen=False):
    text = normalize_code_decimal_separators(rm_accents(text).upper())
    allowed = "A-Z0-9."
    if keep_slash:
        allowed += "/"
    if keep_hyphen:
        allowed += "-"
    return re.sub(rf"[^{allowed}]+", "", text)


def is_upper_code_token(text):
    raw = rm_accents(str(text or "")).strip()
    return bool(re.fullmatch(r"[A-Z0-9./-]+", raw) and re.search(r"[A-Z]", raw))


def normalize_rule_key(text):
    text = rm_accents(str(text or "")).casefold().strip()
    return re.sub(r"\s+", " ", text)


# Exact Vietmax examples from "mã vietmax.xlsx". Keep these before the
# generic fallback because several rows intentionally use custom abbreviations.
VIETMAX_PRODUCT_CODE_OVERRIDES = (
    ('Bản nhôm CTP (D-L) 400x530x0.3 (50 tấm) 2 lớp', 'BNCTP400x530x0.3'),
    ('Bản nhôm CTP (D-L) 470x560x0.3 (50 tấm) 2 lớp', 'BCTP470x560x0.3'),
    ('Bản nhôm CTP BOCICA 470x560x0.3 (50 tấm)', 'BNCTPBOCICA470x560x0.3'),
    ('Bản nhôm CTP BOCICA 560x670x0.3 (50 tấm)', 'BNCTPBOCICA560x670x0.3'),
    ('Bản nhôm CTP BOCICA 600x730x0.3 (50 tấm)', 'BNCTPBOCICA600x730x0.3'),
    ('Bột khô F-20', 'BKF-20'),
    ('Cao su KINYO MC0877  khổ 0.730x0.710', 'CSKINYOMC0877.0.730x0.710'),
    ('Cao su KINYO MC0877  khổ 1.030x0.920', 'CSKINYOMC0877.1.030x0.920'),
    ('Cồn Công Nghiệp', 'CCN'),
    ('Dung dịch hiện bản CTP', 'DDHBCTP'),
    ('Dung dịch làm ẩm stabilat PR2105', 'DDLAstabilatPR2105'),
    ('Giấy  Couche 250 gsm (79x47) cm', 'GCOUCHE250GSM79X47'),
    ('Giấy  Couche 300 gsm (43x62) cm', 'GCOUCHE300GSM43X62'),
    ('Giấy  Couche 300 gsm (62x88) cm', 'GCOUCHE300GSM62X88'),
    ('Giấy  Couche Pindo 300 gsm (86x70) cm', 'GCOUCHEPINDO300GSM86X70'),
    ('Giấy  Couche Pindo 350 gsm (43x65) cm', 'GCOUCHEPINDO350GSM43X65'),
    ('Giấy An Hòa 92/70gms/790x1090', 'GAH9270GMS790x1090'),
    ('Giấy An Hòa 95/60gms/620x860', 'GAH9560GMS620x860'),
    ('Giấy An Hòa 95/70gms/620x860', 'GAH9570GMS620x860'),
    ('Giấy Bãi Bằng 70/92 gsm', 'GBB7092GSM'),
    ('Giấy Cacbon CB white 56/610*860_TL (R500)', 'GCCBWHITE56610*860_TL(R500)'),
    ('Giấy Cacbon CB white 56/650*860 _TL (R500)', 'GCCBWHITE56650*860 _TL(R500)'),
    ('Giấy Cacbon CF Blue 56/610*860_TL (R500)', 'GCCFBLUE56610*860_TL(R500)'),
    ('Giấy Cacbon CF pink 56/610*860_TL (R500)', 'GCCFPINK56610*860_TL(R500)'),
    ('Giấy Cacbon CF pink 56/650*860_TL (R500)', 'GCCFPINK56650*860_TL(R500)'),
    ('Giấy Cacbon CF White 56/610*860_TL (R500)', 'GCCFWHITE56610*860_TL(R500)'),
    ('Giấy Cacbon CF yellow 56/610*860_TL (R500)', 'GCCFYELLOW56610*860_TL(R500)'),
    ('Giấy Cacbon CF yellow 56/650*860_TL (R500)', 'GCCFYELLOW56650*860_TL(R500)'),
    ('Giấy Cacbon CFB blue 50/790*1094_ TL (R500)', 'GCCFBBLUE56790*1094_ TL(R500)'),
    ('Giấy Cacbon CFB pink 50/610*860_TL (R500)', 'GCCFBPINK56610*860_TL(R500)'),
    ('Giấy Cacbon CFB pink 50/650*860_TL (R500)', 'GCCFBPINK56650*860_TL(R500)'),
    ('Giấy Cacbon CFB yellow 50/650*860_TL (R500)', 'GCCFBYELLOW56650*860_TL(R500)'),
    ('Giấy Couche 100 gsm (65x86) cm', 'GCOUCHE100GSM65X86'),
    ('Giấy Couche 150 gsm (62x86) cm', 'GCOUCHE150GSM62X86'),
    ('Giấy Couche 80 gsm (43x65) cm', 'GCOUCHE80GSM43X65'),
    ('Giấy Couche định lượng 300gsm, khổ 620mm', 'GCOUCHE300620MM'),
    ('Giấy Couche định lượng 300gsm, khổ 790mm', 'GCOUCHE300790MM'),
    ('Giấy Couche Gloss 148/32.5x63cm', 'GCOUCHEGLOSS14832.5X63CM'),
    ('Giấy Couche Gloss 148/56x86cm', 'GCOUCHEGLOSSGLOSS148/56X86CM'),
    ('Giấy couche hikote 120gsm                 khổ 62x86cm', 'GCOUCHEHIKOTE120GSMKHO62X86CM'),
    ('Giấy couche hikote 148gsm                 khổ 72cm', 'GCOUCHEHIKOTE120GSMKHO62X86CM'),
    ('Giấy couche hikote 148gsm           khổ 62x86cm', 'GCOUCHEHIKOTE148GSMKHO62X86CM'),
    ('Giấy couche hikote 200gsm           khổ 62x86cm', 'GCOUCHEHIKOTE200GSMKHO62X86CM'),
    ('Giấy couche hikote 200gsm           khổ 79x109cm', 'GCOUCHEHIKOTE200GSMKHO79X109CM'),
    ('Giấy couche hikote 250gsm           khổ 62x86cm', 'GCOUCHEHIKOTE250GSMKHO62X86CM'),
    ('Giấy couche hikote 300gsm           khổ 62x86cm', 'GCOUCHEHIKOTE300GSMKHO62X86CM'),
    ('Giấy couche hikote 300gsm           khổ 79x109cm', 'GCOUCHEHIKOTE300GSMKHO79X109CM'),
    ('Giấy Couche matt 148/52x72cm', 'GCOUCHEMATTMATT148/52X72CM'),
    ('Giấy Couche matt 80/46x79cm', 'GCOUCHEMATTMATT80/46X79CM'),
    ('Giấy decal tự dính', 'GIAYDECELTUDINH'),
    ('Giấy Duplex', 'GIAYDUPLEX'),
    ('Giấy Duplex ĐL 300 g/m2', 'GDL300GM2'),
    ('Giấy Duplex ĐL 400 g/m2', 'GDL400GM2'),
    ('GIẤY DUPLEX LION 250GSM - 650MM', 'GDUPLEXLION250GSM650MM'),
    ('Giấy in', 'GIAYIN'),
    ('GIẤY IVORY (SLIVER PARK) 250GSM - 650MM', 'GIVORYSLIVERPARK350GSM720MM'),
    ('GIẤY IVORY (SLIVER PARK) 350GSM - 720MM', 'GIVORYSLIVERPARK250GSM651MM'),
    ('Giấy Ivory (S-Pak Plus) - 300gsm Khổ 72cm', 'GIVORYSPALPLUS30GSM72CM'),
    ('Giấy Ivory 230 gsm (100x54) cm', 'GIVORY230GSM100X54'),
    ('Giấy Ivory 350 gsm (36x61) cm', 'GIVORY350GSM36X61'),
    ('Giấy Ivory SP 250/43x61cm', 'GIVORYSPSP250/43X61CM'),
    ('Giấy mỹ thuật 120gsm', 'GMYTHUAT120gsm'),
    ('Giấy mỹ thuật 200 gsm 79x109cm', 'GMYTHUAT200GSM79X109CM'),
    ('Giấy Offset 100 gsm (62x86) cm', 'GOFFSET100GSM62X86'),
    ('Giấy Offset 100 gsm (79x109) cm', 'GOFFSET100GSM79X109'),
    ('Giấy Offset 100/43.5x62cm', 'GOFFSET10010043.5X62CM'),
    ('Giấy offset 100gsm khổ 62x86cm', 'GOFFSET100GSM62X86CM'),
    ('Giấy offset 120gsm khổ 79x109cm', 'GOFFSET12079X109CM'),
    ('Giấy offset 140gsm khổ 79x109cm', 'GOFFSET140GSM79X109CM'),
    ('Giấy Offset 180 gsm (39x54) cm', 'GOFFSET180GSM39X54CM'),
    ('Giấy Offset 80/62x86cm', 'GOFFSET80OFFSET80/62X86CM'),
    ('Giấy offset 80gsm khổ 62x86cm', 'GOFFSET8062X86CM'),
    ('Keo 3035', 'KEO3035'),
    ('Keo làm từ Polyme ( Gôm bản)', 'KEOPOLYMR'),
    ('Màng bóng BOPP - 20 mic', 'MANGBONGBOPP'),
    ('Màng Bopp bóng', 'MANGBOPPBONG'),
    ('Màng Bopp mờ', 'MANGBOPPMO'),
    ('Màng tự dính BLWK-Z0585MW', 'MANGTUDINH'),
    ('Màng tự dính SYNWK-F1840N', 'MANGTUDINH'),
    ('Mực in - TK Mark V T Black (LT) (đen)', 'MUCINDEN'),
    ('Mực in - TK Mark V T Cyan (VN) - 2Kg (xanh)', 'MUCINXANH'),
    ('Mực in - TK Mark V T Magenta (VN) - 2kg (đỏ)', 'MUCINDO'),
    ('Mực in - TK Mark V T Yellow (VN) - 2kg (vàng)', 'MUCINVANG'),
    ('Mực in Peony đỏ cờ (KELE-04)', 'MUCINPEONYDO'),
    ('Tấm bản in bằng nhôm CTP 1130x930', 'TAMBANNHOMCTP1130X930'),
    ('Véc ni phủ bóng bề mặt OP Varnish (new)', 'VECNI'),
)


def normalize_vietmax_exact_product_key(product):
    return " ".join(normalize_rule_key(word) for word in code_words(product))


def vietmax_product_code_override(product):
    key = normalize_vietmax_exact_product_key(product)
    for source, code in VIETMAX_PRODUCT_CODE_OVERRIDES:
        if normalize_vietmax_exact_product_key(source) == key:
            return code
    return None


def vietmax_preserve_dimension_unit_code(product, code):
    text = normalize_sep(product)
    match = re.search(r"\d+(?:[.,]\d+)?\s*x\s*\d+(?:[.,]\d+)?(?:\s*x\s*\d+(?:[.,]\d+)?)?\s*\)?\s*(cm|mm|m)\b", text, re.IGNORECASE)
    if not match:
        return code
    dimension = extract_dimensions(text)
    if not dimension:
        return code
    compact_dimension = normalize_code_token(dimension)
    unit = normalize_code_token(match.group(1))
    value = raw_text(code)
    if value.upper().endswith(f"{compact_dimension}{unit}"):
        return value
    if value.upper().endswith(compact_dimension):
        return f"{value}{unit}"
    return value


VIETMAX_ALL_WORD_PHRASES = (
    "Giấy Couche",
    "Giấy Cacbon",
    "Giấy Duplex",
    "Giấy Ivory",
    "Giấy Offset",
    "Giấy in 70/76",
    "Giấy in BB58/92gsm",
    "Giấy in BB60/92gsm",
    "Giấy in BB80/92gsm",
    "Giấy in Cacbon",
    "Giấy in Couche",
    "Giấy in Duplex",
    "Giấy in Ivory",
    "Giấy In Offset",
)

VIETMAX_ALL_WORD_PREFIXES = (
    "Giấy in offset",
    "Giấy ivory ningbo",
)


VIETMAX_FALLBACK_PREFIX_RULES = (
    ("Bản nhôm CTP BOCICA", "BNCTPBOCICA"),
    ("Bản nhôm CTP", "BNCTP"),
    ("Giấy Cacbon CFB", "GCCFB"),
    ("Giấy Cacbon CF", "GCCF"),
    ("Giấy Cacbon CB", "GCCB"),
    ("Giấy An Hòa", "GAH"),
    ("Giấy Ivory", "GIVORY"),
    ("Giấy Couche", "GCOUCHE"),
    ("Giấy mỹ thuật", "GMYTHUAT"),
    ("Giấy Offset", "GOFFSET"),
    ("Màng tự dính", "MANGTUDINH"),
    ("Tấm bản in bằng nhôm", "TAMBANNHOM"),
    ("Giấy tự dính", "GIAYTUDINH"),
)


VIETMAX_RULE_WORKBOOK_CANDIDATES = (
    Path(r"E:\Excel Mom\Dóc\VIETMAX\mã vietmax.xlsx"),
    BASE_DIR.parent / "Dóc" / "VIETMAX" / "mã vietmax.xlsx",
    BASE_DIR / "Dóc" / "VIETMAX" / "mã vietmax.xlsx",
    RESOURCE_DIR / "mã vietmax.xlsx",
)


def vietmax_suffix_token(token, keep_slash=False):
    text = rm_accents(str(token or "")).strip()
    parenthesized_dimension = bool(re.fullmatch(r"\(\d+(?:[xX.]*\d+)+\)", text, re.IGNORECASE))
    text = re.sub(r"\s+", "", text)
    if keep_slash:
        text = re.sub(r"[^A-Za-z0-9.*()_/-]+", "", text)
    else:
        text = re.sub(r"[^A-Za-z0-9.*()_-]+", "", text.replace("/", ""))
    if re.fullmatch(r"\(?\d+(?:[xX.]*\d+)+\)?(?:MM|CM|M)?", text, re.IGNORECASE):
        text = text.strip("()")
    if parenthesized_dimension:
        return text.upper()
    dimension_token = bool(re.search(r"\d+[xX]\d+", text))
    result = []
    for char in text:
        result.append("x" if dimension_token and char in {"x", "X"} else char.upper())
    return "".join(result)


def vietmax_suffix_is_dimension(value):
    return bool(re.search(r"\d+[xX.]\d+", str(value or "")))


def vietmax_rule_marker(value):
    key = normalize_rule_key(value)
    return key if "+" in key and ("ki tu sau" in key or "kitu sau" in key or "kich thuoc" in key or "cac ki tu sau" in key) else ""


def vietmax_rule_prefix(value):
    text = raw_text(value)
    if "+" not in text:
        return ""
    return normalize_code_token(text.split("+", 1)[0])


def vietmax_target_code_prefix(value):
    match = re.match(r"[A-Za-z]+", raw_text(value))
    return match.group(0) if match else ""


def vietmax_workbook_rule_path():
    for path in VIETMAX_RULE_WORKBOOK_CANDIDATES:
        if path.exists():
            return path
    return None


def vietmax_common_prefix(values):
    if not values:
        return ""
    prefix = str(values[0] or "")
    for value in values[1:]:
        text = str(value or "")
        while prefix and not text.startswith(prefix):
            prefix = prefix[:-1]
    return prefix


def vietmax_candidate_phrase_for_rule(name, prefix):
    words = code_words(name)
    if not words:
        return ""
    best = ""
    for index in range(1, len(words) + 1):
        candidate_words = words[:index]
        acronym = "".join((normalize_code_token(word) if is_upper_code_token(word) else normalize_code_token(word)[:1]) for word in candidate_words if normalize_code_token(word))
        compact = "".join(normalize_code_token(word) for word in candidate_words)
        without_giay = "".join(normalize_code_token(word) for word in candidate_words[1:]) if normalize_rule_key(candidate_words[0]) == "giay" else compact
        if prefix in {acronym, compact, without_giay}:
            best = " ".join(candidate_words)
    if best:
        return best
    for index, word in enumerate(words):
        if has_number(word):
            return " ".join(words[:index])
    return " ".join(words[:1])


@lru_cache(maxsize=1)
def vietmax_workbook_prefix_rules():
    path = vietmax_workbook_rule_path()
    if path is None:
        return VIETMAX_FALLBACK_PREFIX_RULES
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, data_only=True, read_only=True)
        ws = wb.active
        rules = []
        for row in range(2, ws.max_row + 1):
            name = raw_text(ws.cell(row, 1).value)
            rule_value = ws.cell(row, 2).value
            target_prefix = vietmax_target_code_prefix(ws.cell(row, 3).value)
            marker = vietmax_rule_marker(rule_value)
            if marker:
                prefix = vietmax_rule_prefix(rule_value)
                if target_prefix and not target_prefix.startswith(prefix):
                    prefix = target_prefix
                phrase = vietmax_candidate_phrase_for_rule(name, prefix)
                if phrase and prefix:
                    rules.append((phrase, prefix))
        wb.close()
        seen = set()
        unique_rules = []
        for phrase, prefix in sorted(rules, key=lambda item: len(code_words(item[0])), reverse=True):
            key = (normalize_rule_key(phrase), prefix)
            if key not in seen:
                seen.add(key)
                unique_rules.append((phrase, prefix))
        seen_phrases = {normalize_rule_key(phrase) for phrase, _prefix in unique_rules}
        for phrase, prefix in VIETMAX_FALLBACK_PREFIX_RULES:
            phrase_key = normalize_rule_key(phrase)
            if phrase_key not in seen_phrases:
                unique_rules.append((phrase, prefix))
                seen_phrases.add(phrase_key)
        unique_rules.sort(key=lambda item: len(code_words(item[0])), reverse=True)
        return tuple(unique_rules) or VIETMAX_FALLBACK_PREFIX_RULES
    except Exception:
        return VIETMAX_FALLBACK_PREFIX_RULES


def vietmax_prefix_phrase_match_end(words, phrase_words):
    if not words or not phrase_words:
        return None
    phrase_keys = [normalize_rule_key(word) for word in phrase_words]
    if len(words) < len(phrase_keys):
        return None
    for index, phrase_key in enumerate(phrase_keys):
        if normalize_rule_key(words[index]) != phrase_key:
            return None
    return len(phrase_keys)


def vietmax_prefix_rule_code(product, sales_style=False):
    words = code_words(product)
    for phrase, prefix in vietmax_workbook_prefix_rules():
        phrase_words = code_words(phrase)
        match_end = vietmax_prefix_phrase_match_end(words, phrase_words)
        if match_end is None:
            continue
        suffix_parts = []
        for word in words[match_end:]:
            word_key = normalize_rule_key(word)
            if word_key in {"dinh", "luong"}:
                continue
            if word_key == "kho" and prefix in {"GOFFSET", "GCOUCHE", "GIVORY"}:
                continue
            part = vietmax_sales_suffix_token(word) if sales_style else vietmax_suffix_token(word)
            suffix_parts.append(part)
        suffix = "".join(suffix_parts)
        if suffix:
            return f"{prefix}{suffix}"
    return None


def vietmax_early_prefix_rule_code(product):
    words = code_words(product)
    for phrase, prefix in vietmax_workbook_prefix_rules():
        if normalize_rule_key(phrase) != "giay an hoa":
            continue
        phrase_words = code_words(phrase)
        match_end = vietmax_prefix_phrase_match_end(words, phrase_words)
        if match_end is not None:
            suffix = "".join(vietmax_suffix_token(word) for word in words[match_end:])
            return f"{prefix}{suffix}" if suffix else None
    return None


def vietmax_manual_prefix_rule_code(product):
    rules = (
        ("Giấy Ivory", "GIVORY"),
        ("Màng tự dính", "MANGTUDINH"),
        ("Tấm bản in bằng nhôm", "TAMBANNHOM"),
    )
    words = code_words(product)
    for phrase, prefix in rules:
        phrase_words = code_words(phrase)
        match_end = vietmax_prefix_phrase_match_end(words, phrase_words)
        if match_end is None:
            continue
        suffix_parts = []
        for word in words[match_end:]:
            word_key = normalize_rule_key(word)
            if word_key in {"dinh", "luong", "kho"}:
                continue
            suffix_parts.append(vietmax_suffix_token(word))
        suffix = "".join(suffix_parts)
        return f"{prefix}{suffix}" if suffix else prefix
    return None


def vietmax_all_words_code(product):
    key = normalize_rule_key(product)
    phrase_keys = {normalize_rule_key(phrase) for phrase in VIETMAX_ALL_WORD_PHRASES}
    prefix_keys = tuple(normalize_rule_key(prefix) for prefix in VIETMAX_ALL_WORD_PREFIXES)
    if key not in phrase_keys and not any(key.startswith(prefix_key + " ") or key == prefix_key for prefix_key in prefix_keys):
        return None
    return "".join(normalize_code_token(word, keep_hyphen=True, keep_slash=True) for word in code_words(product))


def vietmax_sales_suffix_token(token, first_word=False):
    part = vietmax_suffix_token(token, keep_slash=True)
    if not part:
        return ""
    if first_word or has_number(token) or is_upper_code_token(token):
        return part
    return part[:1]


def vietmax_ban_ra_fallback_code(words):
    parts = [vietmax_sales_suffix_token(word, index == 0) for index, word in enumerate(words)]
    return "".join(part for part in parts if part)[:MAX_CODE_LENGTH]

VIETNAM_LOCATION_PHRASES = sorted(
    [
        tuple(normalize_token(word) for word in name.split())
        for name in [
            "An Giang", "Bà Rịa Vũng Tàu", "Bắc Giang", "Bắc Kạn", "Bạc Liêu", "Bắc Ninh", "Bến Tre",
            "Bình Dương", "Bình Định", "Bình Phước", "Bình Thuận", "Cà Mau", "Cao Bằng", "Cần Thơ",
            "Đà Nẵng", "Đắk Lắk", "Đắk Nông", "Điện Biên", "Đồng Nai", "Đồng Tháp", "Gia Lai", "Hà Giang",
            "Hà Nam", "Hà Nội", "Hà Tĩnh", "Hải Dương", "Hải Phòng", "Hậu Giang", "Hòa Bình", "Hồ Chí Minh",
            "Hưng Yên", "Khánh Hòa", "Kiên Giang", "Kon Tum", "Lai Châu", "Lâm Đồng", "Lạng Sơn", "Lào Cai",
            "Long An", "Nam Định", "Nghệ An", "Ninh Bình", "Ninh Thuận", "Phú Thọ", "Phú Yên", "Quảng Bình",
            "Quảng Nam", "Quảng Ngãi", "Quảng Ninh", "Quảng Trị", "Sóc Trăng", "Sơn La", "Tây Ninh",
            "Thái Bình", "Thái Nguyên", "Thanh Hóa", "Thành phố Huế", "Thừa Thiên Huế", "Huế", "Tiền Giang",
            "Trà Vinh", "Tuyên Quang", "Vĩnh Long", "Vĩnh Phúc", "Yên Bái", "Việt Nam",
        ]
    ],
    key=len,
    reverse=True,
)


def remove_company_location_phrases(tokens):
    result = []
    index = 0
    while index < len(tokens):
        phrase = next(
            (items for items in VIETNAM_LOCATION_PHRASES if tuple(tokens[index:index + len(items)]) == items),
            None,
        )
        if phrase:
            index += len(phrase)
        else:
            result.append(tokens[index])
            index += 1
    return result


def suggest_prefix(company):
    words = re.sub(r"[^A-Za-z0-9À-ỹĐđ ]+", " ", str(company)).split()
    meaningful = remove_company_location_phrases([normalize_token(w) for w in words if normalize_token(w)])
    tail_words = meaningful[-2:]
    prefix = "".join(w[:1] for w in tail_words)
    return prefix or normalize_token(company)[:2]


def prefix_last_n_words(company, n=2):
    words = re.sub(r"[^A-Za-z0-9À-ỹĐđ ]+", " ", str(company)).split()
    meaningful = remove_company_location_phrases([normalize_token(w) for w in words if normalize_token(w)])
    if not meaningful:
        meaningful = [normalize_token(w) for w in words if normalize_token(w)]
    selected = meaningful[-max(1, int(n or 2)):]
    prefix = "".join(w[:1] for w in selected)
    return prefix or normalize_token(company)[:2]


def prefix_last_n_words_chars(company, word_count=2, chars_per_word=1):
    words = re.sub(r"[^A-Za-z0-9À-ỹĐđ ]+", " ", str(company)).split()
    meaningful = remove_company_location_phrases([normalize_token(w) for w in words if normalize_token(w)])
    if not meaningful:
        meaningful = [normalize_token(w) for w in words if normalize_token(w)]
    selected = meaningful[-max(1, int(word_count or 2)):]
    chars = max(1, int(chars_per_word or 1))
    prefix = "".join(w[:chars] for w in selected)
    return prefix or normalize_token(company)[:2]


def prefix_all_name_words(company):
    words = re.sub(r"[^A-Za-z0-9À-ỹĐđ ]+", " ", str(company)).split()
    meaningful = remove_company_location_phrases([normalize_token(w) for w in words if normalize_token(w)])
    if not meaningful:
        meaningful = [normalize_token(w) for w in words if normalize_token(w)]
    prefix = "".join(w[:1] for w in meaningful)
    return prefix or normalize_token(company)[:2]


def prefix_last_2_words(company):
    return prefix_last_n_words(company, 2)


def prefix_last_n_mst(mst, n=3):
    digits = re.sub(r"\D", "", str(mst))
    return digits[-n:] if len(digits) >= n else digits


def prefix_2_words_mst(company, mst, mst_digits=3):
    words_prefix = prefix_last_2_words(company)
    mst_suffix = prefix_last_n_mst(mst, mst_digits)
    return f"{words_prefix}{mst_suffix}" if words_prefix and mst_suffix else (words_prefix or mst_suffix or "")


def compute_prefix_strategies(company, mst):
    return {
        "last_2_words": prefix_last_2_words(company),
        "last_3_mst": prefix_last_n_mst(mst, 3),
        "2_words_mst": prefix_2_words_mst(company, mst, 3),
        "all_name_words": prefix_all_name_words(company),
    }


def normalize_sep(text):
    return re.sub(r"(?<=\d)\s*([xX*])\s*(?=\d)", "x", str(text).strip())


def norm_num(value, is_last_numeric=False):
    s = str(value).replace(",", ".")
    try:
        f = float(s)
        if f.is_integer():
            return str(int(f))
        return str(f).rstrip("0").rstrip(".")
    except Exception:
        return s


def extract_dimensions(name):
    text = normalize_sep(name)
    m = re.search(r"(\d+(?:[.,]\d+)?)x(\d+(?:[.,]\d+)?)(?:x(\d+(?:[.,]\d+)?))?", text, re.I)
    if not m:
        return ""
    parts = [p for p in m.groups() if p]
    return "x".join(norm_num(p) for p in parts)


def has_number(text):
    return bool(re.search(r"\d", str(text)))


def simple_number_text(text):
    text = normalize_sep(text)
    text = rm_accents(text).upper()
    text = re.sub(r"\b(MM|M|CM|C)\b", "", text)
    text = re.sub(r"[^A-Z0-9.x,]+", "", text)
    return text.replace(",", ".")


def code_words(name):
    text = str(name or "")
    for ch in DIAMETER_CHARS:
        text = text.replace(ch, "F")
    mojibake_kho = "\u006b\u0068\u00e1\u00bb\u2022"
    text = re.sub(rf"(?i)(khổ|{mojibake_kho}|kho)(?=\d)", lambda match: match.group(1) + " ", text)
    text = normalize_sep(text)
    return [w for w in re.split(r"\s+", text.strip()) if w]


def trim_code(value):
    value = re.sub(r"\.+", ".", str(value or "")).strip(". ")
    return value[:MAX_CODE_LENGTH].rstrip(".")


def sanitize_product_code(value):
    return code_normalization.sanitize_product_code(value, accent_normalizer=rm_accents, trim=trim_code)


def is_volume_token(token):
    token = str(token or "").strip()
    return bool(re.fullmatch(r"\d+(?:[,.]\d+)?[lL]", token))


def is_dimension_token(token):
    token = normalize_sep(token)
    return bool(re.search(r"\d", token)) or token.lower() == "x"


def son_phuong_dimension_pair(words, idx):
    cur = normalize_token(words[idx])
    nxt = normalize_token(words[idx + 1]) if idx + 1 < len(words) else ""
    if cur == "C" and re.fullmatch(r"\d+(?:MM)?", nxt):
        return "Cx" + re.sub(r"MM$", "", nxt)
    return ""


def should_process_qty(v):
    if pd.isna(v):
        return False
    s = str(v).strip()
    if not s:
        return False
    try:
        return float(s.replace(",", ".")) > 0
    except Exception:
        return False


def parse_price(value):
    if pd.isna(value):
        return None
    if isinstance(value, (int, float, np.integer, np.floating)):
        return float(value)
    s = str(value).strip()
    if not s:
        return None
    s = re.sub(r"[^\d,.\-]", "", s)
    if "," in s and "." in s:
        if s.rfind(".") > s.rfind(","):
            s = s.replace(",", "")
        else:
            s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        parts = s.split(",")
        if len(parts[-1]) == 3 and len(parts) > 1:
            s = "".join(parts)
        else:
            s = s.replace(",", ".")
    else:
        s = s
    try:
        return float(s)
    except Exception:
        return None


def raw_text(value):
    return "" if pd.isna(value) else str(value).strip()


def normalized_invoice_status(value):
    return rm_accents(raw_text(value)).casefold()


def normalized_invoice_status_set(values):
    if not isinstance(values, list):
        return IGNORED_INVOICE_STATUSES
    selected = {normalized_invoice_status(value) for value in values if raw_text(value)}
    return selected


def ignored_invoice_status(value, skip_statuses=None):
    status = rm_accents(raw_text(value)).casefold()
    return status in normalized_invoice_status_set(skip_statuses)


def parse_quantity(value):
    parsed = parse_price(value)
    return parsed if parsed is not None and parsed > 0 else 1


def fmt_price(value):
    if value is None:
        return ""
    if abs(value - int(value)) < 0.000001:
        return f"{int(value):,}"
    return f"{value:,.2f}".rstrip("0").rstrip(".")


def cell(df, row, col):
    return "" if col < 0 or col >= df.shape[1] else df.iat[row, col]


def normalized_inventory_match_value(value):
    return raw_text(value).casefold().strip()


def inventory_rule_matches(source_value, rule_value, operator):
    source_value = normalized_inventory_match_value(source_value)
    rule_value = normalized_inventory_match_value(rule_value)
    if operator == "equals":
        return source_value == rule_value
    return rule_value in source_value


def inventory_rule_priority(rule, order):
    return int(rule.get("priority") or 0), -order


def normalized_header_label(value):
    text = raw_text(value)
    try:
        text = text.encode("latin1").decode("utf-8")
    except (UnicodeEncodeError, UnicodeDecodeError):
        pass
    return rm_accents(text).casefold().strip()


def inventory_column_indexes(df, header_index):
    labels = {"tk vat tu": None, "ma kho": None}
    if 0 <= header_index < len(df):
        for col in range(df.shape[1]):
            key = normalized_header_label(cell(df, header_index, col))
            if key in labels and labels[key] is None:
                labels[key] = col
    for key in labels:
        if labels[key] is None:
            labels[key] = df.shape[1]
            df[df.shape[1]] = ""
    df.iat[header_index, labels["tk vat tu"]] = "TK vật tư"
    df.iat[header_index, labels["ma kho"]] = "Mã kho"
    return labels["tk vat tu"], labels["ma kho"]


def customer_code_column_index(df, header_index):
    label_key = "ma khach hang"
    column_index = None
    if 0 <= header_index < len(df):
        for col in range(df.shape[1]):
            if normalized_header_label(cell(df, header_index, col)) == label_key:
                column_index = col
                break
    if column_index is None:
        column_index = df.shape[1]
        df[df.shape[1]] = ""
    df.iat[header_index, column_index] = "Mã khách hàng"
    return column_index


def apply_fast_customer_codes(df, header_index, row_customer_codes):
    if not row_customer_codes:
        return df
    customer_index = customer_code_column_index(df, header_index)
    for row_index, customer_code in row_customer_codes.items():
        if 0 <= row_index < len(df):
            df.iat[row_index, customer_index] = raw_text(customer_code)
    return df


def apply_inventory_pairs(df, header_index, processed_row_indexes, data, excluded_row_indexes=None):
    pairs = {pair["id"]: pair for pair in normalize_inventory_pairs(data.get("inventory_pairs") or [])}
    rules = normalize_inventory_pair_rules(data.get("inventory_pair_rules") or [])
    default_enabled = bool(data.get("use_default_inventory_pair"))
    default_pair_id = str(data.get("default_inventory_pair_id") or "").strip()
    fallback_pair = None
    if default_enabled and default_pair_id in pairs:
        fallback_pair = pairs[default_pair_id]
    elif len(pairs) == 1:
        fallback_pair = next(iter(pairs.values()))

    prepared_rules = []
    for order, rule in enumerate(rules):
        if not rule.get("enabled"):
            continue
        pair_id = rule.get("pair_id")
        if pair_id not in pairs:
            continue
        source_col = rule.get("source_col")
        try:
            source_index = excel_col_to_index(source_col)
        except ValueError as exc:
            raise ValueError(f"Invalid inventory rule source column: {source_col}") from exc
        if source_index >= df.shape[1]:
            raise ValueError(f"Inventory rule source column exceeds the number of columns in the sheet: {source_col}")
        prepared_rules.append((source_index, rule, order))

    has_inventory_config = bool(pairs or rules or default_enabled or default_pair_id)
    if not has_inventory_config:
        return df

    tk_index, ma_kho_index = inventory_column_indexes(df, header_index)
    excluded_row_indexes = set(excluded_row_indexes or [])
    for row_index in processed_row_indexes:
        if row_index in excluded_row_indexes:
            continue
        selected_pair = None
        selected_score = None
        for source_index, rule, order in prepared_rules:
            if inventory_rule_matches(cell(df, row_index, source_index), rule.get("value"), rule.get("operator")):
                score = inventory_rule_priority(rule, order)
                if selected_score is None or score > selected_score:
                    selected_pair = pairs[rule["pair_id"]]
                    selected_score = score
        if selected_pair is None:
            selected_pair = fallback_pair
        df.iat[row_index, tk_index] = selected_pair["tk_vat_tu"] if selected_pair else ""
        df.iat[row_index, ma_kho_index] = selected_pair["ma_kho"] if selected_pair else ""
    return df


def word_piece(token, word_rules, keep_numeric=True, keep_liter=False, default_len=1, preserve_upper_code=False, keep_slash=False, keep_hyphen=False):
    key = normalize_rule_key(token)
    rule_key = next((rule for rule in word_rules if normalize_rule_key(rule) == key), None)
    if rule_key is not None:
        return normalize_token(word_rules[rule_key])
    compact = normalize_code_token(token, keep_slash=keep_slash, keep_hyphen=keep_hyphen)
    if preserve_upper_code and is_upper_code_token(token):
        return compact
    compact_descriptor = vietmax_compact_descriptor_piece(token, default_len, keep_slash=keep_slash, keep_hyphen=keep_hyphen)
    if compact_descriptor:
        return compact_descriptor
    if keep_numeric and has_number(token):
        return compact
    if keep_liter and re.search(r"\d+\s*[lL]\b|[lL]\s*$", str(token)):
        return compact
    return compact[:default_len]


def vietmax_compact_descriptor_piece(token, default_len=2, keep_slash=False, keep_hyphen=False):
    raw = str(token or "").strip()
    if is_upper_code_token(raw):
        return ""
    match = re.fullmatch(r"([A-Za-zÀ-ỹĐđ]{3,})(\d.*)", raw)
    if not match:
        return ""
    prefix, suffix = match.groups()
    normalized_prefix = normalize_code_token(prefix)[:default_len]
    normalized_suffix = normalize_code_token(suffix, keep_slash=keep_slash, keep_hyphen=keep_hyphen)
    return f"{normalized_prefix}{normalized_suffix}" if normalized_prefix and normalized_suffix else ""


def phrase_rule_piece(words, start, word_rules):
    if not word_rules:
        return None
    max_len = max((len(code_words(rule)) for rule in word_rules), default=0)
    max_len = min(max_len, len(words) - start)
    for length in range(max_len, 0, -1):
        phrase = " ".join(words[start:start + length])
        rule_key = next((rule for rule in word_rules if normalize_rule_key(rule) == normalize_rule_key(phrase)), None)
        if rule_key is not None:
            return normalize_token(word_rules[rule_key]), length
    return None


def word_pieces(words, word_rules, fallback):
    parts = []
    i = 0
    while i < len(words):
        matched = phrase_rule_piece(words, i, word_rules)
        if matched:
            part, length = matched
            parts.append(part)
            i += length
            continue
        parts.append(fallback(words[i]))
        i += 1
    return parts


def apply_word_rules_to_words(words, word_rules):
    if not word_rules:
        return list(words)
    result = []
    i = 0
    while i < len(words):
        matched = phrase_rule_piece(words, i, word_rules)
        if matched:
            part, length = matched
            result.append(part)
            i += length
            continue
        result.append(apply_inline_word_rules_to_word(words[i], word_rules))
        i += 1
    return result


def apply_inline_word_rules_to_word(word, word_rules):
    text = str(word or "")
    for rule, replacement in (word_rules or {}).items():
        rule_text = str(rule or "").strip()
        if not rule_text or " " in rule_text:
            continue
        replacement_text = normalize_token(replacement)
        if not replacement_text:
            continue
        text = re.sub(rf"(?<![A-Za-zÀ-ỹ]){re.escape(rule_text)}(?![A-Za-zÀ-ỹ])", replacement_text, text, flags=re.IGNORECASE)
    return text


def remove_repeated_phrases(words, repeated_phrases):
    phrase_words = [code_words(phrase) for phrase in normalize_phrase_list(repeated_phrases)]
    phrase_words = [items for items in phrase_words if items]
    if not phrase_words:
        return words
    phrase_words.sort(key=len, reverse=True)
    seen = set()
    result = []
    i = 0
    while i < len(words):
        matched = None
        for items in phrase_words:
            length = len(items)
            if length > len(words) - i:
                continue
            phrase = " ".join(words[i:i + length])
            if normalize_rule_key(phrase) == normalize_rule_key(" ".join(items)):
                matched = (phrase, length)
                break
        if matched:
            phrase, length = matched
            key = normalize_rule_key(phrase)
            if key not in seen:
                result.extend(words[i:i + length])
                seen.add(key)
            i += length
        else:
            result.append(words[i])
            i += 1
    return result


def normalize_cao_thanh_inox_grade_words(words):
    result = []
    for word in words:
        current = normalize_code_token(word)
        previous = normalize_code_token(result[-1]) if result else ""
        if current == "304" and previous == "INOX":
            continue
        result.append(word)
    return result


def normalize_cao_thanh_con_reducer_words(words):
    if not words:
        return words
    first = normalize_code_token(words[0])
    last = normalize_code_token(words[-1], keep_slash=True)
    if first != "CON" or not re.search(r"\d+(?:[.,]\d+)?/\d+", last):
        return words
    if any(normalize_code_token(word) == "THU" for word in words):
        return words
    return [words[0], "thu", *words[1:]]


def make_product_part(profile, product, word_rules, first_word_rules=None, repeated_phrase_removals=None):
    first_word_rules = first_word_rules or {}
    name = "" if pd.isna(product) else str(product).strip()

    if profile == "cao_thanh":
        name = re.sub(r"\([^)]*\)", " ", name)
        words = remove_repeated_phrases(code_words(name), repeated_phrase_removals)
        words = normalize_cao_thanh_inox_grade_words(words)
        words = normalize_cao_thanh_con_reducer_words(words)
        first = word_pieces(
            words[:2],
            first_word_rules,
            lambda w: word_piece(w, {}, keep_numeric=True, default_len=len(str(w)), preserve_upper_code=True, keep_slash=True),
        )
        rest = word_pieces(
            words[2:],
            word_rules,
            lambda w: word_piece(w, {}, keep_numeric=True, preserve_upper_code=True, keep_slash=True),
        )
        return "".join(first + rest)

    words = remove_repeated_phrases(code_words(name), repeated_phrase_removals)

    if profile == "vietmax_ban_ra":
        if word_rules:
            words = apply_word_rules_to_words(words, word_rules)
            name = " ".join(words)
        return vietmax_preserve_dimension_unit_code(name, vietmax_ban_ra_fallback_code(words))

    if profile == "vietmax_mua_vao":
        if word_rules:
            words = apply_word_rules_to_words(words, word_rules)
            name = " ".join(words)
        override = vietmax_product_code_override(name)
        if override is not None:
            return vietmax_preserve_dimension_unit_code(name, override)
        early_prefix_rule_code = vietmax_early_prefix_rule_code(name)
        if early_prefix_rule_code is not None:
            return vietmax_preserve_dimension_unit_code(name, early_prefix_rule_code)
        manual_prefix_rule_code = vietmax_manual_prefix_rule_code(name)
        if manual_prefix_rule_code is not None:
            return vietmax_preserve_dimension_unit_code(name, manual_prefix_rule_code)
        prefix_rule_code = vietmax_prefix_rule_code(name)
        if prefix_rule_code is not None:
            return vietmax_preserve_dimension_unit_code(name, prefix_rule_code)
        all_words_code = vietmax_all_words_code(name)
        if all_words_code is not None:
            return vietmax_preserve_dimension_unit_code(name, all_words_code)
        parts = []
        for index, word in enumerate(words):
            parts.append(word_piece(word, {}, keep_numeric=True, default_len=(len(str(word)) if index == 0 else 2), preserve_upper_code=True, keep_hyphen=True))
        return vietmax_preserve_dimension_unit_code(name, "".join(p for p in parts if p))

    if profile == "quang_thinh":
        filtered = []
        for w in words:
            if normalize_token(w) == "HANG":
                break
            if normalize_token(w) == "SON":
                continue
            filtered.append(w)
        parts = []
        i = 0
        while i < len(filtered):
            matched = phrase_rule_piece(filtered, i, word_rules)
            if matched:
                part, length = matched
                parts.append(part)
                i += length
                continue
            w = filtered[i]
            if is_volume_token(w):
                parts.append(normalize_token(w))
            else:
                parts.append(word_piece(w, {}, keep_numeric=True, default_len=2))
            i += 1
        return ".".join(p for p in parts if p)

    parts = []
    i = 0
    while i < len(words):
        paired = son_phuong_dimension_pair(words, i)
        if paired:
            parts.append(paired)
            i += 2
            continue
        w = words[i]
        if normalize_token(w) == "X":
            parts.append("x")
        elif is_dimension_token(w):
            parts.append(normalize_token(w))
        else:
            matched = phrase_rule_piece(words, i, word_rules)
            if matched:
                part, length = matched
                parts.append(part)
                i += length
                continue
            parts.append(word_piece(w, {}, keep_numeric=False))
        i += 1
    return "".join(parts)


def make_code(mst, product, qty, prefix_map, profile, word_rules, first_word_rules=None, require_qty=True, include_company_prefix=True, repeated_phrase_removals=None):
    if require_qty and not should_process_qty(qty):
        return ""
    mst = "" if pd.isna(mst) else str(mst).strip()
    if include_company_prefix and (not mst or mst not in prefix_map):
        return ""
    body = make_product_part(profile, product, word_rules, first_word_rules, repeated_phrase_removals)
    if not include_company_prefix:
        return sanitize_product_code(body)
    prefix = normalize_token(prefix_map[mst])
    return sanitize_product_code(f"{prefix}.{body}")


def raw_price_group(price, rule):
    min_price = float(rule.get("min_price") or 0)
    percent = float(rule.get("percent") or 8)
    if price is None or min_price <= 0 or percent <= 0:
        return None
    step = min_price * percent / 100
    if step <= 0:
        return None
    if price < min_price:
        return 1
    return int((price - min_price) // step) + 1


def price_group_suffix(price, rule, occupied_groups=None):
    group = raw_price_group(price, rule)
    if group is None:
        return ""
    if occupied_groups:
        ordered = sorted(g for g in occupied_groups if g is not None)
        if group in ordered:
            group = ordered.index(group) + 1
    return f".{group:03d}"


def merge_price_ranges(old_rules, new_rules):
    merged = normalize_price_range_rules(old_rules or {})
    for key, rule in normalize_price_range_rules(new_rules or {}).items():
        if key in merged:
            merged[key]["min_price"] = min(merged[key]["min_price"], rule["min_price"])
            merged[key]["max_price"] = max(merged[key]["max_price"], rule["max_price"])
            merged[key]["percent"] = rule.get("percent") or merged[key].get("percent") or 8
            merged[key]["groups"] = rule.get("groups") or merged[key].get("groups") or []
        else:
            merged[key] = rule
    return merged


def product_key(mst, product):
    return f"{mst}|||{product}"


def powershell_single_quote(value):
    return "'" + str(value).replace("'", "''") + "'"


def excel_com_convert_xls_to_xlsx_bytes(path):
    source = str(Path(path).resolve())
    with tempfile.TemporaryDirectory(prefix="pcf_xls_convert_") as temp_dir:
        target = str(Path(temp_dir) / f"{Path(path).stem}_{uuid.uuid4().hex}.xlsx")
        script = f"""
$ErrorActionPreference = 'Stop'
$src = {powershell_single_quote(source)}
$dst = {powershell_single_quote(target)}
$excel = $null
$workbook = $null
try {{
    $excel = New-Object -ComObject Excel.Application
    $excel.Visible = $false
    $excel.DisplayAlerts = $false
    $workbook = $excel.Workbooks.Open($src, 0, $true)
    $workbook.SaveAs($dst, 51)
    $workbook.Close($false)
}} finally {{
    if ($workbook -ne $null) {{ [void][System.Runtime.InteropServices.Marshal]::ReleaseComObject($workbook) }}
    if ($excel -ne $null) {{
        $excel.Quit()
        [void][System.Runtime.InteropServices.Marshal]::ReleaseComObject($excel)
    }}
}}
"""
        encoded = base64.b64encode(script.encode("utf-16le")).decode("ascii")
        result = subprocess.run(
            ["powershell.exe", "-NoProfile", "-NonInteractive", "-ExecutionPolicy", "Bypass", "-EncodedCommand", encoded],
            capture_output=True,
            text=True,
            timeout=90,
        )
        if result.returncode != 0:
            detail = (result.stderr or result.stdout or "").strip()
            raise ValueError(detail or "Excel could not convert legacy .xls file.")
        return Path(target).read_bytes()


def read_first_non_empty_excel_sheet(raw, engine):
    first_sheet = ""
    first_df = None
    with pd.ExcelFile(BytesIO(raw), engine=engine) as xl:
        for sheet in xl.sheet_names:
            df = pd.read_excel(xl, sheet_name=sheet, header=None, dtype=object)
            if first_df is None:
                first_sheet = sheet
                first_df = df
            if df is not None and not df.dropna(how="all").dropna(axis=1, how="all").empty:
                return sheet, df
    return first_sheet, first_df if first_df is not None else pd.DataFrame()


def read_workbook(path):
    with open(path, "rb") as fh:
        raw = fh.read()
    suffix = Path(path).suffix.lower()
    legacy_xls = raw.startswith(b"\xd0\xcf\x11\xe0") or (suffix == ".xls" and not raw.startswith(b"PK"))
    engine = "xlrd" if legacy_xls else "openpyxl"
    try:
        sheet, df = read_first_non_empty_excel_sheet(raw, engine)
    except Exception as exc:
        if not legacy_xls:
            raise
        try:
            converted = excel_com_convert_xls_to_xlsx_bytes(path)
            sheet, df = read_first_non_empty_excel_sheet(converted, "openpyxl")
        except Exception as fallback_exc:
            raise ValueError(f"{exc}. Excel fallback also failed: {fallback_exc}") from fallback_exc
    return sheet, df


def report_loop_progress(progress_callback, done, total, label, interval=200):
    if progress_callback and (done == 1 or done == total or done % interval == 0):
        progress_callback(done, total, label)


NO_MST_COMPANY_KEY_PREFIX = "__NO_MST__:"


def company_selection_key(company="", mst=""):
    mst_text = raw_text(mst)
    if mst_text:
        return mst_text
    company_text = raw_text(company)
    company_key = rm_accents(company_text).casefold()
    company_key = re.sub(r"[^a-z0-9]+", "_", company_key).strip("_")
    return f"{NO_MST_COMPANY_KEY_PREFIX}{company_key}" if company_key else ""


def company_rows(df, company_col, mst_col, product_col, qty_col=None, price_col=None, address_col=None, invoice_status_col=DEFAULT_INVOICE_STATUS_COL, invoice_status_skip_values=None, progress_callback=None):
    ci, mi, pi = map(excel_col_to_index, [company_col, mst_col, product_col])
    qi = excel_col_to_index(qty_col) if str(qty_col or "").strip() else None
    indexes = [ci, mi, pi]
    if qi is not None:
        indexes.append(qi)
    pri = excel_col_to_index(price_col) if price_col else None
    if pri is not None:
        indexes.append(pri)
    ai = excel_col_to_index(address_col) if address_col else None
    if ai is not None:
        indexes.append(ai)
    status_col = str(invoice_status_col or "").strip().upper()
    status_index = excel_col_to_index(status_col) if status_col else None
    if status_index is not None and df.shape[1] > status_index:
        indexes.append(status_index)
    if df.shape[1] <= max(indexes):
        raise ValueError("Selected columns exceed the number of columns in the sheet.")
    rows = []
    for i in range(len(df)):
        report_loop_progress(progress_callback, i + 1, len(df), "Đang quét dòng hóa đơn")
        if status_index is not None and df.shape[1] > status_index and ignored_invoice_status(cell(df, i, status_index), invoice_status_skip_values):
            continue
        if qi is not None and not should_process_qty(cell(df, i, qi)):
            continue
        company = "" if pd.isna(cell(df, i, ci)) else str(cell(df, i, ci)).strip()
        mst = "" if pd.isna(cell(df, i, mi)) else str(cell(df, i, mi)).strip()
        product = "" if pd.isna(cell(df, i, pi)) else str(cell(df, i, pi)).strip()
        selection_key = company_selection_key(company, mst)
        if not product or not selection_key:
            continue
        quantity = parse_quantity(cell(df, i, qi)) if qi is not None else 1
        price = parse_price(cell(df, i, pri)) if pri is not None else None
        unit_index = pi + 1
        amount_index = pri + 1 if pri is not None else -1
        amount = parse_price(cell(df, i, amount_index)) if amount_index >= 0 else None
        if (amount is None or amount <= 0) and price is not None:
            amount = price * quantity
        rows.append({
            "excel_row": i + 1,
            "stt": raw_text(cell(df, i, 0)),
            "invoice_no": raw_text(cell(df, i, 2)),
            "invoice_date": raw_text(cell(df, i, 3)),
            "mst": mst,
            "selection_key": selection_key,
            "company_key": vietmax_company_identity_key(company, mst),
            "company": company,
            "address": "" if ai is None or pd.isna(cell(df, i, ai)) else str(cell(df, i, ai)).strip(),
            "product": product,
            "unit": raw_text(cell(df, i, unit_index)),
            "quantity": quantity,
            "price": price,
            "amount": amount,
        })
    return rows


def missing_mst_company_warnings(df, company_col, mst_col, product_col, qty_col=None, invoice_status_col=DEFAULT_INVOICE_STATUS_COL, invoice_status_skip_values=None):
    ci, mi, pi = map(excel_col_to_index, [company_col, mst_col, product_col])
    qi = excel_col_to_index(qty_col) if str(qty_col or "").strip() else None
    status_col = str(invoice_status_col or "").strip().upper()
    status_index = excel_col_to_index(status_col) if status_col else None
    indexes = [ci, mi, pi]
    if qi is not None:
        indexes.append(qi)
    if status_index is not None and df.shape[1] > status_index:
        indexes.append(status_index)
    if df.shape[1] <= max(indexes):
        return []
    warnings = {}
    for i in range(len(df)):
        if status_index is not None and df.shape[1] > status_index and ignored_invoice_status(cell(df, i, status_index), invoice_status_skip_values):
            continue
        if qi is not None and not should_process_qty(cell(df, i, qi)):
            continue
        company = raw_text(cell(df, i, ci))
        mst = raw_text(cell(df, i, mi))
        product = raw_text(cell(df, i, pi))
        if mst or not company or not product:
            continue
        item = warnings.setdefault(company, {"company": company, "count": 0, "rows": [], "invoice_nos": []})
        item["count"] += 1
        if len(item["rows"]) < 20:
            item["rows"].append(i + 1)
        invoice_no = raw_text(cell(df, i, 2)) if df.shape[1] > 2 else ""
        if invoice_no and invoice_no not in item["invoice_nos"] and len(item["invoice_nos"]) < 20:
            item["invoice_nos"].append(invoice_no)
    return sorted(warnings.values(), key=lambda item: item["company"].casefold())


def vietmax_ban_ra_match_key(value):
    text = rm_accents(str(value or "")).casefold()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    text = re.sub(r"\bcabon\b", "cacbon", text)
    return re.sub(r"\s+", " ", text).strip()


def vietmax_normalize_unit_spelling_aliases(text):
    text = str(text or "")
    return re.sub(
        r"(\d+(?:[,.]\d+)?)\s*(?:l|lit|liter|litre|liters|litres)\b",
        lambda match: f"{norm_num(match.group(1))}l",
        text,
    )


def vietmax_equivalent_product_key(value):
    text = normalize_sep(str(value or ""))
    text = rm_accents(text).casefold()
    text = re.sub(r"\bcabon\b", "cacbon", text)
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"[^a-z0-9]+", "", text)


def vietmax_unit_spelling_product_key(value):
    text = normalize_sep(str(value or ""))
    text = rm_accents(text).casefold()
    text = vietmax_normalize_unit_spelling_aliases(text)
    text = re.sub(r"\bcabon\b", "cacbon", text)
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"[^a-z0-9]+", "", text)


def vietmax_is_unit_spelling_only_diff(current, target):
    current_norm = vietmax_equivalent_product_key(current)
    target_norm = vietmax_equivalent_product_key(target)
    if not current_norm or not target_norm or current_norm == target_norm:
        return False
    return vietmax_unit_spelling_product_key(current) == vietmax_unit_spelling_product_key(target)


def vietmax_similar_form_key(value):
    key = vietmax_ban_ra_match_key(value)
    if key.startswith("giay in "):
        tail = key[len("giay in "):].strip()
        return f"giay {tail}" if tail else ""
    if key.startswith("giay "):
        tail = key[len("giay "):].strip()
        if tail and not tail.startswith("in "):
            return key
    return ""


def vietmax_is_optional_in_form(value):
    return vietmax_ban_ra_match_key(value).startswith("giay in ")


def vietmax_review_product_name(value):
    text = " ".join(code_words(value))
    text = re.sub(
        r"(?i)(\d+(?:[.,]\d+)?)x(\d+(?:[.,]\d+)?)(cm|mm|m)?\b",
        lambda match: f"{match.group(1)} x {match.group(2)}{match.group(3) or ''}",
        text,
    )
    return text


def normalize_vietmax_comparison_scope(value):
    return value if value in VIETMAX_COMPARISON_SCOPES else VIETMAX_COMPARISON_SCOPE_ALL_COMPANIES


def vietmax_company_identity_key(company="", mst=""):
    mst_digits = re.sub(r"\D+", "", raw_text(mst))
    if mst_digits:
        return f"mst:{mst_digits}"
    company_key = vietmax_ban_ra_match_key(company)
    return f"company:{company_key}" if company_key else ""


def add_vietmax_company_identity(row, prefix, company="", mst=""):
    company_text = raw_text(company)
    mst_text = raw_text(mst)
    row[f"{prefix}_company"] = company_text
    row[f"{prefix}_mst"] = mst_text
    row[f"{prefix}_company_key"] = vietmax_company_identity_key(company_text, mst_text) or raw_text(row.get(f"{prefix}_company_key"))
    return row


def vietmax_ban_ra_match_score(sales_product, purchase_product):
    sales_key = vietmax_ban_ra_match_key(sales_product)
    purchase_key = vietmax_ban_ra_match_key(purchase_product)
    if not sales_key or not purchase_key:
        return 0.0
    return SequenceMatcher(None, sales_key, purchase_key).ratio()


def vietmax_ban_ra_focus_match(product):
    product_key_value = vietmax_ban_ra_match_key(product)
    if not product_key_value:
        return False
    return any(vietmax_ban_ra_match_key(item) in product_key_value for item in VIETMAX_BAN_RA_FOCUS_PRODUCTS)


def vietmax_ban_ra_skip_purchase_suggestion(product):
    product_key_value = vietmax_ban_ra_match_key(product)
    return product_key_value.startswith(("in ", "cong in ", "gia cong in "))


def vietmax_sales_product_row(value):
    if isinstance(value, dict):
        product = raw_text(value.get("sales_product") or value.get("product") or value.get("name"))
        row = dict(value)
        row["sales_product"] = product
        row["invoice_no"] = raw_text(value.get("invoice_no") or value.get("so_hd"))
        row["invoice_date"] = raw_text(value.get("invoice_date") or value.get("date") or value.get("ngay_co_hang_ban_ra"))
        if "sales_company" in row or "sales_mst" in row or "sales_company_key" in row:
            add_vietmax_company_identity(row, "sales", row.get("sales_company"), row.get("sales_mst"))
        return row
    return raw_text(value)


def unique_vietmax_sales_products(values, comparison_scope=VIETMAX_COMPARISON_SCOPE_ALL_COMPANIES):
    scope = normalize_vietmax_comparison_scope(comparison_scope)
    result = []
    seen = {}
    for value in values:
        normalized = vietmax_sales_product_row(value)
        product = raw_text(normalized.get("sales_product") if isinstance(normalized, dict) else normalized)
        key = vietmax_ban_ra_match_key(product)
        company_key = raw_text(normalized.get("sales_company_key")) if isinstance(normalized, dict) else ""
        unique_key = (company_key, key) if scope == VIETMAX_COMPARISON_SCOPE_SAME_COMPANY else key
        if not product or not key:
            continue
        row = normalized if isinstance(normalized, dict) else product
        if unique_key in seen:
            existing_index = seen[unique_key]
            if vietmax_display_metadata_score(row, "sales") > vietmax_display_metadata_score(result[existing_index], "sales"):
                result[existing_index] = row
            continue
        seen[unique_key] = len(result)
        result.append(row)
    return result


def vietmax_display_metadata_score(row, prefix):
    if not isinstance(row, dict):
        return 0
    score = 0
    if raw_text(row.get(f"{prefix}_price")):
        score += 8
    if raw_text(row.get("invoice_no") or row.get(f"{prefix}_invoice_no")):
        score += 2
    if raw_text(row.get("invoice_date") or row.get(f"{prefix}_invoice_date")):
        score += 2
    if raw_text(row.get(f"{prefix}_unit") or row.get("unit")):
        score += 1
    return score


def normalized_vietmax_unit(value):
    return normalize_rule_key(value).replace(" ", "")


def vietmax_unit_mismatch(sales_unit, purchase_unit):
    sales_key = normalized_vietmax_unit(sales_unit)
    purchase_key = normalized_vietmax_unit(purchase_unit)
    return bool(sales_key and purchase_key and sales_key != purchase_key)


def normalize_vietmax_conversion_mode(value):
    mode = raw_text(value)
    return mode if mode in VIETMAX_CONVERSION_MODES else VIETMAX_CONVERSION_MODE_NONE


def vietmax_conversion_mode_label(mode):
    return VIETMAX_CONVERSION_MODE_LABELS.get(normalize_vietmax_conversion_mode(mode), VIETMAX_CONVERSION_MODE_LABELS[VIETMAX_CONVERSION_MODE_NONE])


def vietmax_conversion_mode_requires_formula(mode):
    return normalize_vietmax_conversion_mode(mode) in {
        VIETMAX_CONVERSION_MODE_QTY_AND_UNIT,
        VIETMAX_CONVERSION_MODE_QTY_ONLY,
        VIETMAX_CONVERSION_MODE_UNIT_ONLY,
    }


def parse_vietmax_formula_number(value):
    text = raw_text(value).replace(" ", "")
    if not re.fullmatch(r"\d+(?:[,.]\d+)?", text):
        return None
    try:
        number = float(text.replace(",", "."))
    except Exception:
        return None
    return number if number > 0 else None


def parse_vietmax_conversion_formula(formula):
    text = raw_text(formula)
    match = re.fullmatch(r"\s*(\d+(?:[,.]\d+)?)\s+([^=]+?)\s*=\s*(\d+(?:[,.]\d+)?)\s+([^=]+?)\s*", text)
    if not match:
        return None
    left_qty = parse_vietmax_formula_number(match.group(1))
    right_qty = parse_vietmax_formula_number(match.group(3))
    left_unit = raw_text(match.group(2))
    right_unit = raw_text(match.group(4))
    if not left_qty or not right_qty or not left_unit or not right_unit:
        return None
    return {
        "left_qty": left_qty,
        "left_unit": left_unit,
        "right_qty": right_qty,
        "right_unit": right_unit,
    }


def vietmax_conversion_quantity_factor(sales_unit, purchase_unit, parsed_formula):
    if not parsed_formula:
        return None
    sales_key = normalized_vietmax_unit(sales_unit)
    purchase_key = normalized_vietmax_unit(purchase_unit)
    left_key = normalized_vietmax_unit(parsed_formula.get("left_unit"))
    right_key = normalized_vietmax_unit(parsed_formula.get("right_unit"))
    left_qty = parsed_formula.get("left_qty")
    right_qty = parsed_formula.get("right_qty")
    if not sales_key or not purchase_key or not left_key or not right_key or not left_qty or not right_qty:
        return None
    if sales_key == left_key and purchase_key == right_key:
        return right_qty / left_qty
    if sales_key == right_key and purchase_key == left_key:
        return left_qty / right_qty
    return None


def format_vietmax_converted_quantity(value):
    if value is None:
        return value
    if abs(value - round(value)) < 0.000001:
        return int(round(value))
    return round(value, 6)


def parse_vietmax_tax_rate_ratio(value):
    rate = parse_price(value)
    if rate is None:
        return None
    return rate if abs(rate) <= 1 else rate / 100


def apply_vietmax_match_conversion_to_row(
    df,
    row_index,
    qty_index,
    unit_index,
    match,
    amount_index=None,
    tax_rate_index=None,
    tax_amount_index=None,
    total_index=None,
):
    mode = normalize_vietmax_conversion_mode(match.get("conversion_mode"))
    if not vietmax_conversion_mode_requires_formula(mode):
        return False
    current_unit_key = normalized_vietmax_unit(cell(df, row_index, unit_index)) if unit_index < df.shape[1] else ""
    sales_unit_key = normalized_vietmax_unit(match.get("sales_unit"))
    purchase_unit_key = normalized_vietmax_unit(match.get("purchase_unit"))
    if not current_unit_key or not sales_unit_key or current_unit_key != sales_unit_key:
        return False
    if purchase_unit_key and current_unit_key == purchase_unit_key:
        return False
    parsed_formula = parse_vietmax_conversion_formula(match.get("conversion_formula"))
    factor = vietmax_conversion_quantity_factor(match.get("sales_unit"), match.get("purchase_unit"), parsed_formula)
    if factor is None:
        return False
    if unit_index >= df.shape[1]:
        df[df.shape[1]] = ""

    price_index = qty_index + 1 if qty_index is not None else None
    original_quantity = parse_price(cell(df, row_index, qty_index)) if qty_index is not None else None
    original_unit_price = parse_price(cell(df, row_index, price_index)) if price_index is not None and price_index < df.shape[1] else None
    original_line_amount = parse_price(cell(df, row_index, amount_index)) if amount_index is not None and amount_index < df.shape[1] else None
    if original_line_amount is None and original_quantity is not None and original_unit_price is not None:
        original_line_amount = original_quantity * original_unit_price

    converted_quantity = None
    if mode in {VIETMAX_CONVERSION_MODE_QTY_AND_UNIT, VIETMAX_CONVERSION_MODE_QTY_ONLY} and qty_index is not None:
        if original_quantity is None:
            return False
        converted_quantity = format_vietmax_converted_quantity(original_quantity * factor)
        converted_quantity_number = parse_price(converted_quantity)
        if converted_quantity_number is None or converted_quantity_number <= 0:
            return False
        df.iat[row_index, qty_index] = converted_quantity
        if price_index is not None and price_index < df.shape[1]:
            converted_unit_price = None
            if original_line_amount is not None:
                converted_unit_price = original_line_amount / converted_quantity_number
            elif original_unit_price is not None:
                converted_unit_price = original_unit_price / factor
            if converted_unit_price is not None:
                df.iat[row_index, price_index] = format_vietmax_converted_quantity(converted_unit_price)

    if converted_quantity is not None and original_line_amount is not None:
        line_amount = format_vietmax_converted_quantity(original_line_amount)
        if amount_index is not None and amount_index < df.shape[1]:
            df.iat[row_index, amount_index] = line_amount

        tax_amount = parse_price(cell(df, row_index, tax_amount_index)) if tax_amount_index is not None and tax_amount_index < df.shape[1] else None
        if tax_amount is None and tax_rate_index is not None and tax_rate_index < df.shape[1]:
            tax_rate = parse_vietmax_tax_rate_ratio(cell(df, row_index, tax_rate_index))
            if tax_rate is not None:
                tax_amount = original_line_amount * tax_rate
        if tax_amount is not None and tax_amount_index is not None and tax_amount_index < df.shape[1]:
            df.iat[row_index, tax_amount_index] = format_vietmax_converted_quantity(tax_amount)

        total_amount = parse_price(cell(df, row_index, total_index)) if total_index is not None and total_index < df.shape[1] else None
        if total_amount is None:
            total_amount = original_line_amount + (tax_amount or 0)
        if total_index is not None and total_index < df.shape[1]:
            df.iat[row_index, total_index] = format_vietmax_converted_quantity(total_amount)

    if mode in {VIETMAX_CONVERSION_MODE_QTY_AND_UNIT, VIETMAX_CONVERSION_MODE_UNIT_ONLY}:
        df.iat[row_index, unit_index] = raw_text(match.get("purchase_unit"))
    return True

_VIETMAX_DIMENSION_RE = re.compile(r"^(D\d+|[A-Z]\d+|\d+(?:[,.]\d+)?\s*(l|lit|lít|ml)|\d+\s*gsm|\d+x\d+(?:x\d+(?:[,.]\d+)?)?[\w]*|D\d+x\d+|\d+/\d+[*\w]*|BB\d+[\w/]*|\d{1,3}\s*(cm|mm|m|inch|kg|g)|\d+(?:[,.]\d+)?|\d{1,3}|loại\d+|type\d+|mẫu\d+|hạng\d+)$", re.IGNORECASE)


def vietmax_is_case_space_only_diff(current, target):
    current_norm = vietmax_equivalent_product_key(current)
    target_norm = vietmax_equivalent_product_key(target)
    return current_norm == target_norm and str(current or "") != str(target or "")


def vietmax_has_dimension_diff(current, target):
    if vietmax_is_case_space_only_diff(current, target):
        return False
    current_dimension = extract_dimensions(current)
    target_dimension = extract_dimensions(target)
    if current_dimension and target_dimension and current_dimension != target_dimension:
        return True
    current_parts = str(current or "").split()
    target_parts = str(target or "").split()
    matcher = SequenceMatcher(None, current_parts, target_parts)
    for tag, i1, i2, j1, j2 in matcher.get_opcodes():
        if tag == "equal":
            continue
        current_words = current_parts[i1:i2]
        target_words = target_parts[j1:j2]
        if tag == "replace":
            current_joined = "".join(current_words)
            target_joined = "".join(target_words)
            if current_joined.lower().replace(" ", "") == target_joined.lower().replace(" ", ""):
                continue
            if _VIETMAX_DIMENSION_RE.match(current_joined) and _VIETMAX_DIMENSION_RE.match(target_joined):
                return True
        for word in current_words + target_words:
            if _VIETMAX_DIMENSION_RE.match(word):
                return True
    return False


def vietmax_review_pair_row(left, right, similarity, review_group="other", dimension_only=False, extra=None):
    review_row = dict(left)
    review_row["similar_product"] = right.get("product", "")
    review_row["similar_unit"] = right.get("unit", "")
    review_row["similar_invoice_no"] = right.get("invoice_no", "")
    review_row["similar_invoice_date"] = right.get("invoice_date", "")
    review_row["similar_company"] = right.get("company", "")
    review_row["similar_mst"] = right.get("mst", "")
    review_row["similar_company_key"] = right.get("company_key", "")
    review_row["similar_code"] = right.get("code", "")
    review_row["similar_product_key"] = right.get("product_key", "")
    review_row["similar_company_index"] = right.get("company_index")
    review_row["similar_product_index"] = right.get("product_index")
    review_row["similarity"] = f"{similarity * 100:.1f}%"
    review_row["review_group"] = review_group
    review_row["dimension_only"] = dimension_only
    review_row["confirmed"] = False
    if extra:
        review_row.update(extra)
    return review_row


def vietmax_same_generated_code(left, right):
    left_code = sanitize_product_code((left or {}).get("code"))
    right_code = sanitize_product_code((right or {}).get("code"))
    return bool(left_code and right_code and left_code == right_code)


def vietmax_code_from_user_config(row):
    row = row or {}
    return bool(
        row.get("code_from_user_config")
        or row.get("purchase_code_from_user_config")
        or row.get("sales_code_from_user_config")
    )


def vietmax_same_code_from_user_config(left, right):
    return vietmax_same_generated_code(left, right) and (
        vietmax_code_from_user_config(left) or vietmax_code_from_user_config(right)
    )


def vietmax_review_group_for_pair(left, right, allow_same_code_split=False):
    if vietmax_same_code_from_user_config(left, right):
        return None, False, None
    if vietmax_is_unit_spelling_only_diff((left or {}).get("product"), (right or {}).get("product")):
        return "unit_spelling_diff", False, {"review_type": "unit_spelling_diff"}
    if allow_same_code_split and vietmax_same_generated_code(left, right):
        return "same_code_split", False, {"review_type": "same_code_split"}
    if vietmax_has_dimension_diff((left or {}).get("product"), (right or {}).get("product")):
        return "dimension_diff", True, None
    return "other", False, None


def vietmax_product_review_rows(products, product_key, unit_key, comparison_scope=VIETMAX_COMPARISON_SCOPE_ALL_COMPANIES, threshold=0.92, progress_callback=None, allow_same_code_split=False):
    scope = normalize_vietmax_comparison_scope(comparison_scope)
    normalized = []
    key_counts = {}
    token_index = {}
    token_counts = {}
    source_products = products or []
    for item_index, item in enumerate(source_products):
        report_loop_progress(progress_callback, item_index + 1, max(1, len(source_products)), "Đang chuẩn hóa hàng hóa review", interval=100)
        if isinstance(item, dict):
            raw_product = raw_text(item.get(product_key) or item.get("product") or item.get("name"))
            product = vietmax_review_product_name(raw_product)
            unit = raw_text(item.get(unit_key) or item.get("unit"))
            invoice_no = raw_text(item.get("invoice_no") or item.get("so_hd"))
            invoice_date = raw_text(item.get("invoice_date") or item.get("date") or item.get("ngay_hd") or item.get("ngay_co_hang_ban_ra"))
            company = raw_text(item.get("purchase_company") or item.get("sales_company"))
            mst = raw_text(item.get("purchase_mst") or item.get("sales_mst"))
            company_key = raw_text(item.get("purchase_company_key") or item.get("sales_company_key")) or vietmax_company_identity_key(company, mst)
        else:
            raw_product = raw_text(item)
            product = vietmax_review_product_name(raw_product)
            unit = ""
            invoice_no = ""
            invoice_date = ""
            company = ""
            mst = ""
            company_key = ""
        key = vietmax_ban_ra_match_key(product)
        equivalent_key = vietmax_equivalent_product_key(product)
        if product and key:
            row = {"product": product, "unit": unit, "invoice_no": invoice_no, "invoice_date": invoice_date, "company": company, "mst": mst, "company_key": company_key, "match_key": key, "equivalent_key": equivalent_key}
            for extra_key in ("code", "product_key", "company_index", "product_index", "code_from_user_config", "purchase_code_from_user_config", "sales_code_from_user_config"):
                if extra_key in item:
                    row[extra_key] = item.get(extra_key)
            row_index = len(normalized)
            normalized.append(row)
            key_scope = company_key if scope == VIETMAX_COMPARISON_SCOPE_SAME_COMPANY else ""
            key_counts[(key_scope, equivalent_key)] = key_counts.get((key_scope, equivalent_key), 0) + 1
            for token in set(key.split()):
                if len(token) >= 2:
                    token_index.setdefault((key_scope, token), set()).add(row_index)
                    token_counts[(key_scope, token)] = token_counts.get((key_scope, token), 0) + 1
    rows = []
    seen_products = set()
    for row_index, row in enumerate(normalized):
        report_loop_progress(progress_callback, row_index + 1, max(1, len(normalized)), "Đang so sánh hàng hóa review", interval=25)
        row_scope = row.get("company_key") if scope == VIETMAX_COMPARISON_SCOPE_SAME_COMPANY else ""
        if key_counts.get((row_scope, row.get("equivalent_key")), 0) > 1:
            continue
        candidate_indexes = set()
        row_tokens = {token for token in str(row.get("match_key") or "").split() if len(token) >= 2}
        rare_tokens = sorted(row_tokens, key=lambda token: token_counts.get((row_scope, token), len(normalized)))[:4]
        for token in rare_tokens:
            if len(token) >= 2:
                candidate_indexes.update(token_index.get((row_scope, token), set()))
        best = None
        best_score = 0.0
        for other_index in candidate_indexes:
            if row_index == other_index:
                continue
            other = normalized[other_index]
            if scope == VIETMAX_COMPARISON_SCOPE_SAME_COMPANY and row.get("company_key") != other.get("company_key"):
                continue
            if row.get("equivalent_key") == other.get("equivalent_key"):
                continue
            row_form_key = vietmax_similar_form_key(row.get("product"))
            other_form_key = vietmax_similar_form_key(other.get("product"))
            if row_form_key and row_form_key == other_form_key and vietmax_is_optional_in_form(row.get("product")) != vietmax_is_optional_in_form(other.get("product")):
                continue
            other_tokens = {token for token in str(other.get("match_key") or "").split() if len(token) >= 2}
            shared_token_count = len(row_tokens & other_tokens)
            required_shared_tokens = min(len(row_tokens), len(other_tokens))
            if required_shared_tokens >= 3 and shared_token_count < required_shared_tokens - 1:
                continue
            if required_shared_tokens == 2 and shared_token_count < 1:
                continue
            row_key_length = len(str(row.get("match_key") or ""))
            other_key_length = len(str(other.get("match_key") or ""))
            if min(row_key_length, other_key_length) / max(row_key_length, other_key_length) < threshold:
                continue
            score = vietmax_ban_ra_match_score(row.get("product"), other.get("product"))
            if score >= threshold and score > best_score:
                best = other
                best_score = score
        if not best:
            continue
        pair_key = tuple(sorted((row.get("match_key"), best.get("match_key"))))
        unique_key = (row.get("company_key") if scope == VIETMAX_COMPARISON_SCOPE_SAME_COMPANY else "", pair_key)
        if unique_key in seen_products:
            continue
        seen_products.add(unique_key)
        review_group, dimension_only, extra = vietmax_review_group_for_pair(row, best, allow_same_code_split)
        if not review_group:
            continue
        review_row = vietmax_review_pair_row(row, best, best_score, review_group, dimension_only, extra)
        rows.append(review_row)
    rows.extend(vietmax_similar_form_review_rows(normalized, seen_products, scope, allow_same_code_split=allow_same_code_split))
    rows.extend(vietmax_forced_purchase_review_rows(normalized, seen_products, scope))
    return rows


def vietmax_similar_form_review_rows(normalized, seen_products, scope, allow_same_code_split=False):
    by_form = {}
    for row in normalized:
        row_scope = row.get("company_key") if scope == VIETMAX_COMPARISON_SCOPE_SAME_COMPANY else ""
        form_key = vietmax_similar_form_key(row.get("product"))
        if form_key:
            by_form.setdefault((row_scope, form_key), []).append(row)
    rows = []
    for (row_scope, _), candidates in by_form.items():
        if len(candidates) < 2:
            continue
        with_optional = [row for row in candidates if vietmax_is_optional_in_form(row.get("product"))]
        without_optional = [row for row in candidates if not vietmax_is_optional_in_form(row.get("product"))]
        for left in with_optional:
            best = None
            best_score = 0.0
            for right in without_optional:
                if left.get("equivalent_key") == right.get("equivalent_key"):
                    continue
                score = vietmax_ban_ra_match_score(left.get("product"), right.get("product"))
                if score > best_score:
                    best = right
                    best_score = score
            if not best:
                continue
            pair_key = tuple(sorted((left.get("match_key"), best.get("match_key"))))
            unique_key = (row_scope, pair_key)
            if unique_key in seen_products:
                continue
            seen_products.add(unique_key)
            if vietmax_same_code_from_user_config(left, best):
                continue
            review_group = "similar_form"
            extra = {"confirmed": False}
            if allow_same_code_split and vietmax_same_generated_code(left, best):
                review_group = "same_code_split"
                extra["review_type"] = "same_code_split"
            rows.append(vietmax_review_pair_row(
                left,
                best,
                best_score,
                review_group,
                False,
                extra,
            ))
    return rows


def vietmax_forced_purchase_review_rows(normalized, seen_products, scope):
    by_match_key = {}
    for row in normalized:
        row_scope = row.get("company_key") if scope == VIETMAX_COMPARISON_SCOPE_SAME_COMPANY else ""
        by_match_key.setdefault((row_scope, row.get("match_key")), row)
    rows = []
    scopes = sorted({row.get("company_key") if scope == VIETMAX_COMPARISON_SCOPE_SAME_COMPANY else "" for row in normalized}) or [""]
    for left_name, right_name in VIETMAX_PURCHASE_REVIEW_FORCED_PAIRS:
        left_key = vietmax_ban_ra_match_key(vietmax_review_product_name(left_name))
        right_key = vietmax_ban_ra_match_key(vietmax_review_product_name(right_name))
        for row_scope in scopes:
            left = by_match_key.get((row_scope, left_key))
            right = by_match_key.get((row_scope, right_key))
            if not left or not right:
                continue
            pair_key = tuple(sorted((left.get("match_key"), right.get("match_key"))))
            unique_key = (row_scope, pair_key)
            if unique_key in seen_products:
                continue
            seen_products.add(unique_key)
            if vietmax_same_code_from_user_config(left, right):
                continue
            score = vietmax_ban_ra_match_score(left.get("product"), right.get("product"))
            review_row = vietmax_review_pair_row(left, right, score, "other", False, {"forced_review": True})
            rows.append(review_row)
    return rows


def vietmax_ban_ra_sales_products_from_workbook(path, product_col="M", qty_col="O", invoice_status_col=DEFAULT_INVOICE_STATUS_COL, invoice_status_skip_values=None, progress_callback=None, company_col=None, mst_col=None, comparison_scope=VIETMAX_COMPARISON_SCOPE_ALL_COMPANIES, price_col=None):
    _, df = read_workbook(path)
    scope = normalize_vietmax_comparison_scope(comparison_scope)
    product_index = excel_col_to_index(product_col or "M")
    unit_index = product_index + 1
    qty_index = excel_col_to_index(qty_col) if str(qty_col or "").strip() else None
    price_index = excel_col_to_index(price_col) if str(price_col or "").strip() else None
    company_index = excel_col_to_index(company_col) if str(company_col or "").strip() else None
    mst_index = excel_col_to_index(mst_col) if str(mst_col or "").strip() else None
    invoice_no_index = excel_col_to_index("C")
    invoice_date_index = excel_col_to_index("D")
    status_col = str(invoice_status_col or "").strip().upper()
    status_index = excel_col_to_index(status_col) if status_col else None
    indexes = [product_index, invoice_no_index, invoice_date_index]
    if qty_index is not None:
        indexes.append(qty_index)
    if price_index is not None:
        indexes.append(price_index)
    if company_index is not None:
        indexes.append(company_index)
    if mst_index is not None:
        indexes.append(mst_index)
    if status_index is not None and df.shape[1] > status_index:
        indexes.append(status_index)
    if df.shape[1] <= max(indexes):
        raise ValueError("Selected columns exceed the number of columns in the sheet.")
    products = []
    for row_index in range(len(df)):
        report_loop_progress(progress_callback, row_index + 1, len(df), "Đang đọc hàng bán ra")
        if status_index is not None and df.shape[1] > status_index and ignored_invoice_status(cell(df, row_index, status_index), invoice_status_skip_values):
            continue
        if qty_index is not None and not should_process_qty(cell(df, row_index, qty_index)):
            continue
        product = raw_text(cell(df, row_index, product_index))
        if product and "ten" not in vietmax_ban_ra_match_key(product):
            row = {
                "sales_product": product,
                "sales_unit": raw_text(cell(df, row_index, unit_index)) if df.shape[1] > unit_index else "",
                "sales_price": raw_text(cell(df, row_index, price_index)) if price_index is not None and df.shape[1] > price_index else "",
                "invoice_no": raw_text(cell(df, row_index, invoice_no_index)),
                "invoice_date": raw_text(cell(df, row_index, invoice_date_index)),
            }
            if company_index is not None or mst_index is not None:
                add_vietmax_company_identity(
                    row,
                    "sales",
                    cell(df, row_index, company_index) if company_index is not None else "",
                    cell(df, row_index, mst_index) if mst_index is not None else "",
                )
            products.append(row)
    return unique_vietmax_sales_products(products, scope)


def vietmax_purchase_column_score(df, column_index, labels):
    score = 0
    for row_index in range(min(10, len(df))):
        key = normalized_header_label(cell(df, row_index, column_index))
        if key in labels:
            score += 10
        elif any(label in key for label in labels):
            score += 5
    nonempty = sum(1 for row_index in range(min(len(df), 20)) if raw_text(cell(df, row_index, column_index)))
    return score + min(nonempty, 5)


def detect_vietmax_purchase_columns(df):
    preferred_code_columns = [excel_col_to_index(col) for col in ["L", "M"] if excel_col_to_index(col) < df.shape[1]]
    preferred_product_columns = [excel_col_to_index(col) for col in ["M", "N"] if excel_col_to_index(col) < df.shape[1]]
    code_labels = {"ma vt", "ma vat tu", "ma hang", "ma hang hoa"}
    product_labels = {"ten hang", "ten hang hoa", "hang hoa", "ten vat tu"}
    code_candidates = preferred_code_columns or list(range(df.shape[1]))
    product_candidates = preferred_product_columns or list(range(df.shape[1]))
    code_index = max(code_candidates, key=lambda index: vietmax_purchase_column_score(df, index, code_labels))
    product_index = max(product_candidates, key=lambda index: vietmax_purchase_column_score(df, index, product_labels))
    if product_index == code_index and len(product_candidates) > 1:
        alternatives = [index for index in product_candidates if index != code_index]
        product_index = max(alternatives, key=lambda index: vietmax_purchase_column_score(df, index, product_labels))
    header_index = 0
    for row_index in range(min(10, len(df))):
        code_header = normalized_header_label(cell(df, row_index, code_index))
        product_header = normalized_header_label(cell(df, row_index, product_index))
        if code_header in code_labels or any(label in product_header for label in product_labels):
            header_index = row_index
            break
    return code_index, product_index, header_index


def detect_vietmax_purchase_company_columns(df):
    company_labels = {
        "ten nguoi mua",
        "ten nguoi ban",
        "nguoi mua",
        "nguoi ban",
        "ten don vi",
        "don vi ban hang",
        "don vi mua hang",
        "cong ty",
    }
    mst_labels = {
        "mst",
        "ma so thue",
        "mst nguoi mua",
        "mst nguoi ban",
        "ma so thue nguoi mua",
        "ma so thue nguoi ban",
    }
    company_index = None
    mst_index = None
    for row_index in range(min(10, len(df))):
        for column_index in range(df.shape[1]):
            label = normalized_header_label(cell(df, row_index, column_index))
            if company_index is None and (label in company_labels or any(item in label for item in company_labels)):
                company_index = column_index
            if mst_index is None and (label in mst_labels or any(item in label for item in mst_labels)):
                mst_index = column_index
        if company_index is not None and mst_index is not None:
            return company_index, mst_index
    for company_col, mst_col in [("F", "G"), ("I", "J")]:
        company_fallback = excel_col_to_index(company_col)
        mst_fallback = excel_col_to_index(mst_col)
        if df.shape[1] > max(company_fallback, mst_fallback):
            return company_index if company_index is not None else company_fallback, mst_index if mst_index is not None else mst_fallback
    return company_index, mst_index


def detect_vietmax_price_column(df, fallback_index=None):
    price_labels = {"don gia", "don gia ban", "don gia mua", "unit price", "price"}
    for row_index in range(min(10, len(df))):
        for column_index in range(df.shape[1]):
            label = normalized_header_label(cell(df, row_index, column_index))
            if label in price_labels:
                return column_index
            if "don gia" in label and "thanh tien" not in label:
                return column_index
    return fallback_index


def vietmax_purchase_products_from_workbook(path, code_col=None, product_col=None, price_col=None, progress_callback=None, comparison_scope=VIETMAX_COMPARISON_SCOPE_ALL_COMPANIES, require_existing_code=False):
    _, df = read_workbook(path)
    if df.empty:
        return []
    scope = normalize_vietmax_comparison_scope(comparison_scope)
    code_index, product_index, header_index = detect_vietmax_purchase_columns(df)
    company_index, mst_index = detect_vietmax_purchase_company_columns(df) if scope == VIETMAX_COMPARISON_SCOPE_SAME_COMPANY else (None, None)
    if code_col:
        code_index = excel_col_to_index(code_col)
    if product_col:
        product_index = excel_col_to_index(product_col)
    unit_index = product_index + 1
    price_index = excel_col_to_index(price_col) if str(price_col or "").strip() else detect_vietmax_price_column(df, product_index + 3)
    invoice_no_index = excel_col_to_index("C")
    invoice_date_index = excel_col_to_index("D")
    indexes = [code_index, product_index]
    if price_index is not None and price_index < df.shape[1]:
        indexes.append(price_index)
    if company_index is not None:
        indexes.append(company_index)
    if mst_index is not None:
        indexes.append(mst_index)
    if df.shape[1] <= max(indexes):
        raise ValueError("Selected purchase columns exceed the number of columns in the sheet.")
    products = []
    seen = {}
    for row_index in range(header_index + 1, len(df)):
        report_loop_progress(progress_callback, row_index - header_index, max(1, len(df) - header_index - 1), "Đang đọc hàng mua vào")
        code = raw_text(cell(df, row_index, code_index))
        product = raw_text(cell(df, row_index, product_index))
        key = vietmax_ban_ra_match_key(product)
        if not product or not key:
            continue
        if normalized_header_label(code) in {"ma vt", "ma vat tu"} or "ten hang" in normalized_header_label(product):
            continue
        if require_existing_code and (not code or code in {"0", "0.0"}):
            continue
        if not code or code in {"0", "0.0"}:
            code = make_product_part("vietmax_mua_vao", product, {})
        code = sanitize_product_code(code)
        if not code:
            continue
        invoice_no = raw_text(cell(df, row_index, invoice_no_index)) if df.shape[1] > invoice_no_index else ""
        invoice_date = raw_text(cell(df, row_index, invoice_date_index)) if df.shape[1] > invoice_date_index else ""
        row = {
            "purchase_product": product,
            "purchase_code": code,
            "purchase_row": row_index + 1,
            "purchase_unit": raw_text(cell(df, row_index, unit_index)) if df.shape[1] > unit_index else "",
            "purchase_price": raw_text(cell(df, row_index, price_index)) if price_index is not None and df.shape[1] > price_index else "",
            "invoice_no": invoice_no,
            "invoice_date": invoice_date,
            "purchase_invoice_no": invoice_no,
            "purchase_invoice_date": invoice_date,
        }
        company_key = ""
        if scope == VIETMAX_COMPARISON_SCOPE_SAME_COMPANY:
            add_vietmax_company_identity(
                row,
                "purchase",
                cell(df, row_index, company_index) if company_index is not None else "",
                cell(df, row_index, mst_index) if mst_index is not None else "",
            )
            company_key = row.get("purchase_company_key", "")
        seen_key = (company_key, key) if scope == VIETMAX_COMPARISON_SCOPE_SAME_COMPANY else key
        if seen_key in seen:
            existing_index = seen[seen_key]
            if vietmax_display_metadata_score(row, "purchase") > vietmax_display_metadata_score(products[existing_index], "purchase"):
                products[existing_index] = row
            continue
        seen[seen_key] = len(products)
        products.append(row)
    return products


def validate_vietmax_processed_purchase_workbook(path):
    _, df = read_workbook(path)
    if df.empty:
        raise ValueError("File HD mua vào đã xử lý không có dữ liệu.")
    code_index, product_index, header_index = detect_vietmax_purchase_columns(df)
    unit_index = product_index + 1
    tk_index = ma_kho_index = None
    try:
        inventory_indexes = processed_inventory_column_indexes(df, code_index)
        tk_index = inventory_indexes.get("tk_vat_tu")
        ma_kho_index = inventory_indexes.get("ma_kho")
    except Exception:
        pass
    valid_rows = 0
    for row_index in range(header_index + 1, len(df)):
        code = raw_text(cell(df, row_index, code_index))
        product = raw_text(cell(df, row_index, product_index))
        if code and code not in {"0", "0.0"} and product and normalized_header_label(code) not in {"ma vt", "ma vat tu"}:
            valid_rows += 1
    if not valid_rows:
        raise ValueError("File HD mua vào đã xử lý chưa có dòng nào có Mã VT hợp lệ. Hãy xuất file mua vào đã xử lý trước khi tiếp tục bán ra.")
    return {
        "code_col": index_to_excel_col(code_index),
        "product_col": index_to_excel_col(product_index),
        "unit_col": index_to_excel_col(unit_index) if unit_index < df.shape[1] else "",
        "tk_vat_tu_col": index_to_excel_col(tk_index) if tk_index is not None else "",
        "ma_kho_col": index_to_excel_col(ma_kho_index) if ma_kho_index is not None else "",
        "header_row": header_index + 1,
        "valid_rows": valid_rows,
    }


def build_vietmax_ban_ra_purchase_matches(sales_products, purchase_products, threshold=0.92, progress_callback=None, comparison_scope=VIETMAX_COMPARISON_SCOPE_ALL_COMPANIES):
    scope = normalize_vietmax_comparison_scope(comparison_scope)
    normalized_purchase_products = []
    purchase_token_index = {}
    purchase_source = purchase_products or []
    for item_index, item in enumerate(purchase_source):
        report_loop_progress(progress_callback, item_index + 1, len(purchase_source), "Đang lập chỉ mục hàng mua", interval=100)
        if isinstance(item, dict):
            product = raw_text(item.get("purchase_product") or item.get("product") or item.get("name"))
            code = raw_text(item.get("purchase_code") or item.get("code"))
            row = item.get("purchase_row") or item.get("row") or ""
            purchase_unit = raw_text(item.get("purchase_unit") or item.get("unit"))
            purchase_price = raw_text(item.get("purchase_price"))
            purchase_company = raw_text(item.get("purchase_company"))
            purchase_mst = raw_text(item.get("purchase_mst"))
            purchase_company_key = raw_text(item.get("purchase_company_key")) or vietmax_company_identity_key(purchase_company, purchase_mst)
        else:
            product = raw_text(item)
            code = ""
            row = ""
            purchase_unit = ""
            purchase_price = ""
            purchase_company = ""
            purchase_mst = ""
            purchase_company_key = ""
        if product and code:
            purchase_key_value = vietmax_ban_ra_match_key(product)
            purchase_index = len(normalized_purchase_products)
            normalized_purchase_products.append({"purchase_product": product, "purchase_code": code, "purchase_row": row, "purchase_unit": purchase_unit, "purchase_price": purchase_price, "purchase_company": purchase_company, "purchase_mst": purchase_mst, "purchase_company_key": purchase_company_key, "match_key": purchase_key_value})
            for token in set(purchase_key_value.split()):
                if len(token) >= 2:
                    purchase_token_index.setdefault(token, set()).add(purchase_index)
    rows = []
    unique_sales = unique_vietmax_sales_products(sales_products or [], scope)
    for sales_index, sales_item in enumerate(unique_sales):
        report_loop_progress(progress_callback, sales_index + 1, len(unique_sales), "Đang so khớp hàng bán/mua", interval=25)
        if isinstance(sales_item, dict):
            sales_product = raw_text(sales_item.get("sales_product") or sales_item.get("product") or sales_item.get("name"))
            sales_unit = raw_text(sales_item.get("sales_unit") or sales_item.get("unit"))
            sales_price = raw_text(sales_item.get("sales_price"))
            invoice_no = raw_text(sales_item.get("invoice_no") or sales_item.get("so_hd"))
            invoice_date = raw_text(sales_item.get("invoice_date") or sales_item.get("date") or sales_item.get("ngay_co_hang_ban_ra"))
            sales_company = raw_text(sales_item.get("sales_company"))
            sales_mst = raw_text(sales_item.get("sales_mst"))
            sales_company_key = raw_text(sales_item.get("sales_company_key")) or vietmax_company_identity_key(sales_company, sales_mst)
        else:
            sales_product = raw_text(sales_item)
            sales_unit = ""
            sales_price = ""
            invoice_no = ""
            invoice_date = ""
            sales_company = ""
            sales_mst = ""
            sales_company_key = ""
        if vietmax_ban_ra_skip_purchase_suggestion(sales_product):
            continue
        sales_key_value = vietmax_ban_ra_match_key(sales_product)
        candidate_indexes = set()
        for token in set(sales_key_value.split()):
            if len(token) >= 2:
                candidate_indexes.update(purchase_token_index.get(token, set()))
        best = None
        best_score = 0.0
        focus = vietmax_ban_ra_focus_match(sales_product)
        candidates = [normalized_purchase_products[index] for index in candidate_indexes]
        if scope == VIETMAX_COMPARISON_SCOPE_SAME_COMPANY:
            candidates = [purchase for purchase in candidates if sales_company_key and purchase.get("purchase_company_key") == sales_company_key]
        if focus and not candidates:
            candidates = normalized_purchase_products
            if scope == VIETMAX_COMPARISON_SCOPE_SAME_COMPANY:
                candidates = [purchase for purchase in candidates if sales_company_key and purchase.get("purchase_company_key") == sales_company_key]
        for purchase in candidates:
            score = vietmax_ban_ra_match_score(sales_product, purchase["purchase_product"])
            if score > best_score or (score == best_score and vietmax_display_metadata_score(purchase, "purchase") > vietmax_display_metadata_score(best, "purchase")):
                best = purchase
                best_score = score
        if best and (best_score >= threshold or focus):
            purchase_unit = raw_text(best.get("purchase_unit"))
            unit_mismatch = vietmax_unit_mismatch(sales_unit, purchase_unit)
            rows.append({
                "sales_product": sales_product,
                "sales_unit": sales_unit,
                "sales_price": sales_price,
                "sales_company": sales_company,
                "sales_mst": sales_mst,
                "sales_company_key": sales_company_key,
                "invoice_no": invoice_no,
                "invoice_date": invoice_date,
                "purchase_product": best["purchase_product"],
                "purchase_code": best["purchase_code"],
                "purchase_row": best.get("purchase_row", ""),
                "purchase_unit": purchase_unit,
                "purchase_price": raw_text(best.get("purchase_price")),
                "purchase_company": raw_text(best.get("purchase_company")),
                "purchase_mst": raw_text(best.get("purchase_mst")),
                "purchase_company_key": raw_text(best.get("purchase_company_key")),
                "comparison_scope": scope,
                "unit_mismatch": unit_mismatch,
                "unit_warning": "Khác đơn vị tính" if unit_mismatch else "",
                "conversion_mode": VIETMAX_CONVERSION_MODE_NONE,
                "conversion_mode_label": vietmax_conversion_mode_label(VIETMAX_CONVERSION_MODE_NONE),
                "conversion_formula": "",
                "score": best_score,
                "focus": focus,
                "confirmed": True,
            })
    return rows


def vietmax_purchase_match_export_rows(matches):
    rows = [["Dùng", "Hàng bán ra", "ĐVT bán ra", "Số HD", "Ngày có hàng bán ra", "Mã VT mua vào", "Hàng mua vào", "ĐVT mua vào", "Cảnh báo", "Quy đổi", "Khác biệt", "Độ giống"]]
    for match in matches or []:
        if not isinstance(match, dict):
            continue
        try:
            score = float(match.get("score") or 0) * 100
        except Exception:
            score = 0
        rows.append([
            "Có" if match.get("confirmed") is not False else "Không",
            raw_text(match.get("sales_product")),
            raw_text(match.get("sales_unit")),
            raw_text(match.get("invoice_no")),
            raw_text(match.get("invoice_date")),
            raw_text(match.get("purchase_code")),
            raw_text(match.get("purchase_product")),
            raw_text(match.get("purchase_unit")),
            raw_text(match.get("unit_warning") or ("Khác đơn vị tính" if match.get("unit_mismatch") else "")),
            raw_text(match.get("conversion_formula")),
            f"{raw_text(match.get('sales_product'))} -> {raw_text(match.get('purchase_product'))}",
            f"{score:.1f}%",
        ])
    return rows

def vietmax_purchase_match_lookup_key(product, company="", mst="", comparison_scope=VIETMAX_COMPARISON_SCOPE_ALL_COMPANIES):
    product_match_key = vietmax_ban_ra_match_key(product)
    if not product_match_key:
        return ""
    if normalize_vietmax_comparison_scope(comparison_scope) == VIETMAX_COMPARISON_SCOPE_SAME_COMPANY:
        company_key = vietmax_company_identity_key(company, mst)
        return f"{company_key}|||{product_match_key}" if company_key else ""
    return product_match_key


def vietmax_internal_merge_lookup_key(product, company="", mst="", company_key="", comparison_scope=VIETMAX_COMPARISON_SCOPE_ALL_COMPANIES):
    product_match_key = vietmax_ban_ra_match_key(product)
    if not product_match_key:
        return ""
    if normalize_vietmax_comparison_scope(comparison_scope) == VIETMAX_COMPARISON_SCOPE_SAME_COMPANY:
        identity_key = raw_text(company_key) or vietmax_company_identity_key(company, mst)
        return f"{identity_key}|||{product_match_key}" if identity_key else ""
    return product_match_key


def normalize_vietmax_internal_merges(value, comparison_scope=VIETMAX_COMPARISON_SCOPE_ALL_COMPANIES):
    result = {}
    if not isinstance(value, list):
        return result
    scope = normalize_vietmax_comparison_scope(comparison_scope)
    for item in value:
        item = sanitize_review_merge_rule_for_config(item, scope)
        if not item or item.get("confirmed") is not True:
            continue
        if raw_text(item.get("code_choice")).lower() == "split":
            continue
        product = raw_text(item.get("product"))
        similar_product = raw_text(item.get("similar_product"))
        if not product or not similar_product:
            continue
        key = vietmax_internal_merge_lookup_key(
            product,
            item.get("company"),
            item.get("mst"),
            item.get("company_key"),
            item.get("comparison_scope") or scope,
        )
        if key:
            result[key] = dict(item, product=product, similar_product=similar_product)
    return result


def normalize_vietmax_split_code_overrides(value, comparison_scope=VIETMAX_COMPARISON_SCOPE_ALL_COMPANIES):
    result = {}
    if not isinstance(value, list):
        return result
    scope = normalize_vietmax_comparison_scope(comparison_scope)
    for item in value:
        item = sanitize_review_merge_rule_for_config(item, scope)
        if not item or item.get("confirmed") is not True:
            continue
        if raw_text(item.get("code_choice")).lower() != "split":
            continue
        sides = (
            ("product", "company", "mst", "company_key", "split_code"),
            ("similar_product", "similar_company", "similar_mst", "similar_company_key", "similar_split_code"),
        )
        for product_field, company_field, mst_field, company_key_field, code_field in sides:
            product = raw_text(item.get(product_field))
            code = sanitize_product_code(item.get(code_field))
            if not product or not code:
                continue
            key = vietmax_internal_merge_lookup_key(
                product,
                item.get(company_field),
                item.get(mst_field),
                item.get(company_key_field),
                item.get("comparison_scope") or scope,
            )
            if key:
                result[key] = code
    return result


def find_vietmax_split_code_override(overrides, product, company="", mst="", company_key="", comparison_scope=VIETMAX_COMPARISON_SCOPE_ALL_COMPANIES):
    if not overrides:
        return ""
    scope = normalize_vietmax_comparison_scope(comparison_scope)
    key = vietmax_internal_merge_lookup_key(product, company, mst, company_key, scope)
    return overrides.get(key) if key else ""


def find_vietmax_internal_merge(merges, product, company="", mst="", company_key="", comparison_scope=VIETMAX_COMPARISON_SCOPE_ALL_COMPANIES):
    if not merges:
        return None
    scope = normalize_vietmax_comparison_scope(comparison_scope)
    candidate_scopes = [scope]
    for extra_scope in (VIETMAX_COMPARISON_SCOPE_SAME_COMPANY, VIETMAX_COMPARISON_SCOPE_ALL_COMPANIES):
        if extra_scope not in candidate_scopes:
            candidate_scopes.append(extra_scope)
    for candidate_scope in candidate_scopes:
        key = vietmax_internal_merge_lookup_key(product, company, mst, company_key, candidate_scope)
        if key and key in merges:
            return merges[key]
    return None


def apply_vietmax_internal_merges_to_products(products, product_key_name, unit_key_name, merges, comparison_scope=VIETMAX_COMPARISON_SCOPE_ALL_COMPANIES):
    if not merges:
        return list(products or [])
    scope = normalize_vietmax_comparison_scope(comparison_scope)
    product_rows = []
    row_by_lookup = {}
    for item in products or []:
        row = dict(item) if isinstance(item, dict) else {product_key_name: raw_text(item)}
        product = raw_text(row.get(product_key_name) or row.get("product") or row.get("name"))
        company = raw_text(row.get("purchase_company") or row.get("sales_company"))
        mst = raw_text(row.get("purchase_mst") or row.get("sales_mst"))
        company_key = raw_text(row.get("purchase_company_key") or row.get("sales_company_key"))
        product_rows.append((item, row, product, company, mst, company_key))
        for candidate_scope in (scope, VIETMAX_COMPARISON_SCOPE_ALL_COMPANIES, VIETMAX_COMPARISON_SCOPE_SAME_COMPANY):
            lookup = vietmax_internal_merge_lookup_key(product, company, mst, company_key, candidate_scope)
            if lookup:
                row_by_lookup.setdefault((candidate_scope, lookup), row)

    def representative_row(merge):
        merge_scope = normalize_vietmax_comparison_scope(merge.get("comparison_scope") or scope)
        product = raw_text(merge.get("similar_product"))
        company = raw_text(merge.get("similar_company"))
        mst = raw_text(merge.get("similar_mst"))
        company_key = raw_text(merge.get("similar_company_key"))
        for candidate_scope in (merge_scope, scope, VIETMAX_COMPARISON_SCOPE_ALL_COMPANIES, VIETMAX_COMPARISON_SCOPE_SAME_COMPANY):
            lookup = vietmax_internal_merge_lookup_key(product, company, mst, company_key, candidate_scope)
            found = row_by_lookup.get((candidate_scope, lookup)) if lookup else None
            if found:
                return found
        return None

    def row_code(row):
        return raw_text(row.get("purchase_code") or row.get("sales_code") or row.get("code"))

    def generated_merge_code(merge):
        target_product = raw_text((merge or {}).get("similar_product"))
        if not target_product:
            return ""
        target_profile = "vietmax_ban_ra" if product_key_name == "sales_product" else "vietmax_mua_vao"
        return sanitize_product_code(make_product_part(target_profile, target_product, {}, {}, []))

    merged_products = []
    seen = set()
    for item, original_row, product, company, mst, company_key in product_rows:
        row = dict(original_row)
        merge_key = vietmax_internal_merge_lookup_key(product, company, mst, company_key, scope)
        merge = merges.get(merge_key) if merge_key else None
        if merge:
            row[f"original_{product_key_name}"] = product
            row[product_key_name] = raw_text(merge.get("similar_product"))
            target_row = representative_row(merge)
            target_code = row_code(target_row or {}) or generated_merge_code(merge)
            if target_code:
                if product_key_name == "purchase_product":
                    row["original_purchase_code"] = raw_text(row.get("purchase_code") or row.get("code"))
                    if "purchase_code" in row:
                        row["purchase_code"] = target_code
                    elif "code" in row:
                        row["code"] = target_code
                elif product_key_name == "sales_product":
                    row["original_sales_code"] = raw_text(row.get("sales_code") or row.get("code"))
                    if "sales_code" in row:
                        row["sales_code"] = target_code
                    elif "code" in row:
                        row["code"] = target_code
            target_unit = raw_text((target_row or {}).get(unit_key_name)) or raw_text(merge.get("similar_unit"))
            if target_unit:
                row[f"original_{unit_key_name}"] = raw_text(row.get(unit_key_name))
                row[unit_key_name] = target_unit
        product_after_merge = raw_text(row.get(product_key_name))
        merged_company_key = raw_text(row.get("purchase_company_key") or row.get("sales_company_key")) if scope == VIETMAX_COMPARISON_SCOPE_SAME_COMPANY else ""
        unique_key = (merged_company_key, vietmax_ban_ra_match_key(product_after_merge))
        if unique_key in seen:
            continue
        seen.add(unique_key)
        merged_products.append(row if isinstance(item, dict) else product_after_merge)
    return merged_products


def build_vietmax_khh_exact_purchase_matches(sales_products, processed_purchase_products, comparison_scope=VIETMAX_COMPARISON_SCOPE_ALL_COMPANIES):
    scope = normalize_vietmax_comparison_scope(comparison_scope)
    purchase_by_key = {}
    for purchase in processed_purchase_products or []:
        code = raw_text(purchase.get("purchase_code"))
        product = raw_text(purchase.get("purchase_product"))
        if not code or code in {"0", "0.0"} or not product:
            continue
        company_key = raw_text(purchase.get("purchase_company_key")) if scope == VIETMAX_COMPARISON_SCOPE_SAME_COMPANY else ""
        key = (company_key, vietmax_ban_ra_match_key(product)) if scope == VIETMAX_COMPARISON_SCOPE_SAME_COMPANY else vietmax_ban_ra_match_key(product)
        purchase_by_key.setdefault(key, purchase)
    matches = []
    seen = set()
    for sales in sales_products or []:
        product = raw_text(sales.get("sales_product"))
        if not product:
            continue
        company_key = raw_text(sales.get("sales_company_key")) if scope == VIETMAX_COMPARISON_SCOPE_SAME_COMPANY else ""
        key = (company_key, vietmax_ban_ra_match_key(product)) if scope == VIETMAX_COMPARISON_SCOPE_SAME_COMPANY else vietmax_ban_ra_match_key(product)
        purchase = purchase_by_key.get(key)
        if not purchase or key in seen:
            continue
        seen.add(key)
        sales_unit = raw_text(sales.get("sales_unit"))
        purchase_unit = raw_text(purchase.get("purchase_unit"))
        unit_mismatch = vietmax_unit_mismatch(sales_unit, purchase_unit)
        matches.append({
            "confirmed": True,
            "khh_exact_match": True,
            "sales_product": product,
            "sales_unit": sales_unit,
            "sales_price": raw_text(sales.get("sales_price")),
            "invoice_no": raw_text(sales.get("invoice_no")),
            "invoice_date": raw_text(sales.get("invoice_date")),
            "purchase_code": raw_text(purchase.get("purchase_code")),
            "purchase_product": raw_text(purchase.get("purchase_product")),
            "purchase_unit": purchase_unit,
            "purchase_price": raw_text(purchase.get("purchase_price")),
            "purchase_company": raw_text(purchase.get("purchase_company")),
            "purchase_mst": raw_text(purchase.get("purchase_mst")),
            "purchase_company_key": raw_text(purchase.get("purchase_company_key")),
            "unit_mismatch": unit_mismatch,
            "unit_warning": "Khác đơn vị tính" if unit_mismatch else "Từ HD mua vào đã xử lý (KVT)",
            "conversion_mode": VIETMAX_CONVERSION_MODE_NONE,
            "conversion_formula": "",
            "score": 1,
            "sales_company": raw_text(sales.get("sales_company")),
            "sales_mst": raw_text(sales.get("sales_mst")),
            "sales_company_key": raw_text(sales.get("sales_company_key")),
        })
    return matches


def normalize_vietmax_ban_ra_purchase_matches(value, comparison_scope=VIETMAX_COMPARISON_SCOPE_ALL_COMPANIES):
    result = {}
    default_scope = normalize_vietmax_comparison_scope(comparison_scope)
    if isinstance(value, dict):
        iterable = [dict({"sales_product": key, "purchase_code": code}) for key, code in value.items()]
    elif isinstance(value, list):
        iterable = value
    else:
        iterable = []
    for item in iterable:
        if not isinstance(item, dict) or item.get("confirmed") is False:
            continue
        sales_product = raw_text(item.get("sales_product"))
        purchase_code = raw_text(item.get("purchase_code"))
        scope = normalize_vietmax_comparison_scope(item.get("comparison_scope") or default_scope)
        sales_company = raw_text(item.get("sales_company"))
        sales_mst = raw_text(item.get("sales_mst"))
        sales_company_key = raw_text(item.get("sales_company_key")) or vietmax_company_identity_key(sales_company, sales_mst)
        key = vietmax_purchase_match_lookup_key(sales_product, sales_company, sales_mst, scope)
        if scope == VIETMAX_COMPARISON_SCOPE_SAME_COMPANY and sales_company_key:
            key = f"{sales_company_key}|||{vietmax_ban_ra_match_key(sales_product)}"
        if key and purchase_code:
            sales_unit = raw_text(item.get("sales_unit"))
            purchase_unit = raw_text(item.get("purchase_unit"))
            unit_mismatch = item.get("unit_mismatch")
            if unit_mismatch is None:
                unit_mismatch = vietmax_unit_mismatch(sales_unit, purchase_unit)
            result[key] = {
                "sales_product": sales_product,
                "purchase_code": purchase_code,
                "purchase_product": raw_text(item.get("purchase_product")),
                "comparison_scope": scope,
                "sales_company": sales_company,
                "sales_mst": sales_mst,
                "sales_company_key": sales_company_key,
                "purchase_company": raw_text(item.get("purchase_company")),
                "purchase_mst": raw_text(item.get("purchase_mst")),
                "purchase_company_key": raw_text(item.get("purchase_company_key")),
                "sales_unit": sales_unit,
                "purchase_unit": purchase_unit,
                "unit_mismatch": bool(unit_mismatch),
                "conversion_mode": normalize_vietmax_conversion_mode(item.get("conversion_mode")),
                "conversion_formula": raw_text(item.get("conversion_formula")),
            }
    return result




def vietmax_purchase_match_row_key(item, comparison_scope=VIETMAX_COMPARISON_SCOPE_ALL_COMPANIES):
    if not isinstance(item, dict):
        return ""
    sales_product = raw_text(item.get("sales_product"))
    sales_company = raw_text(item.get("sales_company"))
    sales_mst = raw_text(item.get("sales_mst"))
    sales_company_key = raw_text(item.get("sales_company_key")) or vietmax_company_identity_key(sales_company, sales_mst)
    scope = normalize_vietmax_comparison_scope(item.get("comparison_scope") or comparison_scope)
    key = vietmax_purchase_match_lookup_key(sales_product, sales_company, sales_mst, scope)
    if scope == VIETMAX_COMPARISON_SCOPE_SAME_COMPANY and sales_company_key:
        key = f"{sales_company_key}|||{vietmax_ban_ra_match_key(sales_product)}"
    return key


def merge_vietmax_purchase_match_rows(match_rows, rule_rows=None, comparison_scope=VIETMAX_COMPARISON_SCOPE_ALL_COMPANIES):
    merged = {}
    for item in list(rule_rows or []) + list(match_rows or []):
        key = vietmax_purchase_match_row_key(item, comparison_scope)
        if not key:
            continue
        if item.get("confirmed") is False:
            merged[key] = None
            continue
        if raw_text(item.get("purchase_code")):
            merged[key] = dict(item)
    return [item for item in merged.values() if item]


def same_vietmax_purchase_match_target(left, right):
    if not left or not right:
        return False
    left_code = raw_text(left.get("purchase_code"))
    right_code = raw_text(right.get("purchase_code"))
    if left_code and right_code and left_code == right_code:
        return True
    return (
        vietmax_ban_ra_match_key(left.get("purchase_product")) ==
        vietmax_ban_ra_match_key(right.get("purchase_product"))
    )


def preserve_vietmax_conversion_settings(target, source):
    if not target or not source or not same_vietmax_purchase_match_target(target, source):
        return target
    mode = normalize_vietmax_conversion_mode(source.get("conversion_mode"))
    if not vietmax_conversion_mode_requires_formula(mode):
        return target
    if (
        normalized_vietmax_unit(target.get("sales_unit")) != normalized_vietmax_unit(source.get("sales_unit")) or
        normalized_vietmax_unit(target.get("purchase_unit")) != normalized_vietmax_unit(source.get("purchase_unit"))
    ):
        return target
    target = dict(target)
    target["conversion_mode"] = mode
    target["conversion_formula"] = raw_text(source.get("conversion_formula"))
    return target


def apply_vietmax_ban_ra_match_inventory(df, header_index, row_indexes):
    if not row_indexes:
        return df
    tk_index, ma_kho_index = inventory_column_indexes(df, header_index)
    for row_index in row_indexes:
        df.iat[row_index, tk_index] = VIETMAX_BAN_RA_MATCH_TK_VAT_TU
        df.iat[row_index, ma_kho_index] = VIETMAX_BAN_RA_MATCH_MA_KHO
    return df


def choose_company(names):
    names = [n for n in names if str(n).strip()]
    return Counter(names).most_common(1)[0][0] if names else ""


def unique_values(values):
    seen = set()
    result = []
    for v in values:
        v = "" if pd.isna(v) else str(v).strip()
        if not v or v in seen:
            continue
        seen.add(v)
        result.append(v)
    return result


def analyze(path, company_col, mst_col, address_col, product_col, qty_col, price_col, profile_cfg, invoice_status_col=DEFAULT_INVOICE_STATUS_COL, invoice_status_skip_values=None, progress_callback=None):
    _, df = read_workbook(path)
    rows = company_rows(df, company_col, mst_col, product_col, qty_col, price_col, address_col, invoice_status_col, invoice_status_skip_values, progress_callback=progress_callback)
    by = {}
    addresses = {}
    products = {}
    msts = {}
    for row_index, r in enumerate(rows):
        report_loop_progress(progress_callback, row_index + 1, len(rows), "Đang gom công ty và hàng hóa")
        company_id = r.get("selection_key") or company_selection_key(r.get("company"), r.get("mst"))
        by.setdefault(company_id, []).append(r["company"])
        msts.setdefault(company_id, []).append(r["mst"])
        if r["address"]:
            addresses.setdefault(company_id, []).append(r["address"])
        products.setdefault(company_id, {}).setdefault(r["product"], []).append(r)

    mst_company = {company_id: choose_company(names) for company_id, names in by.items()}
    suggestions = {company_id: suggest_prefix(comp) for company_id, comp in mst_company.items()}
    cnt = Counter(s for s in suggestions.values() if s)
    companies_data = []
    saved_prefixes = profile_cfg.get("prefixes") or {}
    saved_selected = profile_cfg.get("selected_products") or {}

    for company_index, company_id in enumerate(mst_company):
        report_loop_progress(progress_callback, company_index + 1, len(mst_company), "Đang dựng danh sách công ty", interval=50)
        mst_values = unique_values(msts.get(company_id, []))
        mst = raw_text(mst_values[0]) if mst_values else ""
        suggested = saved_prefixes.get(company_id) or (saved_prefixes.get(mst) if mst else "") or suggestions[company_id]
        prefix_strategies = compute_prefix_strategies(mst_company[company_id], mst)
        product_items = []
        for name, product_rows in products.get(company_id, {}).items():
            prices = [r["price"] for r in product_rows if r["price"] is not None]
            price_rows = [{
                "price": r["price"],
                "quantity": r["quantity"],
                "amount": r["amount"],
                "excelRow": r["excel_row"],
                "stt": r["stt"],
                "invoiceNo": r["invoice_no"],
                "invoiceDate": r["invoice_date"],
                "unit": r["unit"],
                "name": r["product"],
            } for r in product_rows]
            product_items.append({
                "name": name,
                "count": len(product_rows),
                "minPrice": min(prices) if prices else None,
                "maxPrice": max(prices) if prices else None,
                "priceCount": len(set(prices)),
                "priceRows": price_rows,
            })
        skipped_list = saved_selected.get(company_id)
        if skipped_list is None and mst:
            skipped_list = saved_selected.get(mst)
        skipped_set = set(skipped_list) if isinstance(skipped_list, list) else set()
        companies_data.append({
            "mst": str(mst),
            "company_id": str(company_id),
            "missing_mst": not bool(mst),
            "company": str(mst_company[company_id]),
            "address": unique_values(addresses.get(company_id, []))[0] if addresses.get(company_id) else "",
            "addresses": unique_values(addresses.get(company_id, [])),
            "all_names": [str(n) for n in unique_values(by[company_id])],
            "all_products": product_items,
            "selected_product_names": [p["name"] for p in product_items if p["name"] not in skipped_set],
            "count": int(len(by[company_id])),
            "default_prefix": str(suggestions[company_id]),
            "suggested": str(suggested),
            "value": str(suggested),
            "needs_manual": (not suggested) or cnt[suggestions[company_id]] > 1,
            "status": "OK" if suggested and cnt[suggestions[company_id]] <= 1 else "Need manual check",
            "prefix_strategies": prefix_strategies,
        })

    companies_data.sort(key=lambda x: (x["mst"] or "~~~~", x["company"]))
    for idx, item in enumerate(companies_data):
        item["safe_id"] = str(idx)
    return {"rows_to_process": int(len(rows)), "company_count": int(len(companies_data)), "companies": companies_data, "missing_mst_companies": [], "missing_mst_row_count": 0}


def resolve_output_path(original, requested_path):
    default_name = f"{Path(original).stem}_fdi.xls"
    requested_path = str(requested_path or "").strip().strip('"')
    if not requested_path:
        return OUTPUT_DIR / default_name
    p = Path(requested_path)
    if p.suffix.lower() in {".xls", ".xlsx", ".xlsm"}:
        p.parent.mkdir(parents=True, exist_ok=True)
        stem = p.stem if p.stem.casefold().endswith("_fdi") else f"{p.stem}_fdi"
        return p.with_name(f"{stem}.xls")
    p.mkdir(parents=True, exist_ok=True)
    return p / default_name


def validate_payload(data):
    all_mst = [str(x).strip() for x in data.get("all_mst", [])]
    selected = set(str(x).strip() for x in data.get("process_mst", []))
    include_company_prefix = data.get("include_company_prefix") is not False
    safe = {}
    for item in data.get("mst_safe_id", []):
        if "|||" in item:
            mst, sid = item.split("|||", 1)
            safe[mst] = sid
    prefix_map = {}
    selected_products = {}
    used = {}
    errors = []
    for mst in all_mst:
        if not mst or mst not in selected:
            continue
        sid = safe.get(mst, "")
        if include_company_prefix:
            prefix = str(data.get(f"prefix_{sid}", "") or "").strip().upper()
            if not prefix:
                errors.append(f"{mst}: prefix is empty.")
                continue
            if not re.fullmatch(r"[A-Z0-9]{1,20}", prefix):
                errors.append(f"{mst}: prefix must use only A-Z or 0-9.")
                continue
            if prefix in used and used[prefix] != mst:
                errors.append(f"{mst}: prefix {prefix} duplicates MST {used[prefix]}.")
                continue
            used[prefix] = mst
            prefix_map[mst] = prefix
        selected_products[mst] = set(data.get(f"selected_products_{sid}", []))
    if errors:
        raise ValueError("\n".join(errors))
    return prefix_map, selected_products


def process_workbook(path, out, data, progress_callback=None):
    if progress_callback:
        progress_callback(0, 1, "Đang đọc workbook nguồn")
    sheet, df = read_workbook(path)
    if progress_callback:
        progress_callback(1, 1, f"Đã đọc workbook nguồn: {len(df)} dòng")
    company_col = data.get("company_col", "F").upper()
    mst_col = data.get("mst_col", "G").upper()
    product_col = data.get("product_col", "N").upper()
    qty_col = str(data.get("qty_col", "P") or "").upper()
    profile = profile_key(data.get("profile", "son_phuong"))
    vietmax_phase = normalize_vietmax_phase(data.get("vietmax_phase"))
    effective_profile = effective_processing_profile(profile, vietmax_phase)
    uses_price_rules = effective_profile == "cao_thanh"
    price_col = str(data.get("price_col", "R") or "").upper() if uses_price_rules else ""
    output_col = data.get("output_col", "M").upper()
    invoice_status_col = str(data.get("invoice_status_col", DEFAULT_INVOICE_STATUS_COL) or "").upper()
    invoice_status_skip_values = data.get("invoice_status_skip_values")
    ci, mi, pi, oi = map(excel_col_to_index, [company_col, mst_col, product_col, output_col])
    pri = excel_col_to_index(price_col) if price_col else None
    qi = excel_col_to_index(qty_col) if qty_col else None
    indexes = [ci, mi, pi, oi]
    if pri is not None:
        indexes.append(pri)
    if qi is not None:
        indexes.append(qi)
    if df.shape[1] <= max(indexes):
        raise ValueError("Selected columns exceed the number of columns in the sheet.")
    word_rules = data.get("word_rules") or {}
    first_word_rules = data.get("first_word_rules") or {}
    repeated_phrase_removals = normalize_phrase_list(data.get("repeated_phrase_removals") or [])
    include_company_prefix = data.get("include_company_prefix") is not False
    price_rules = (data.get("price_group_rules") or {}) if uses_price_rules else {}
    price_range_rules = (data.get("price_range_rules") or {}) if uses_price_rules else {}
    manual_code_overrides = data.get("manual_code_overrides") or {}
    product_code_replacements = normalize_product_code_replacements(data.get("product_code_replacements") or {})
    vietmax_comparison_scope = normalize_vietmax_comparison_scope(data.get("comparison_scope") or data.get("vietmax_ban_ra_comparison_scope"))
    uses_vietmax_sales = effective_profile == "vietmax_ban_ra"
    uses_vietmax_purchase = effective_profile == "vietmax_mua_vao"
    vietmax_purchase_match_rows = []
    if uses_vietmax_sales:
        vietmax_purchase_match_rows = merge_vietmax_purchase_match_rows(
            data.get("vietmax_ban_ra_purchase_matches") or [],
            data.get("vietmax_ban_ra_purchase_match_rules") or [],
            vietmax_comparison_scope,
        )
    vietmax_purchase_matches = normalize_vietmax_ban_ra_purchase_matches(vietmax_purchase_match_rows, vietmax_comparison_scope) if uses_vietmax_sales else {}
    vietmax_sales_internal_merges = normalize_vietmax_internal_merges(data.get("vietmax_ban_ra_sales_internal_merges") or [], vietmax_comparison_scope) if uses_vietmax_sales else {}
    vietmax_purchase_internal_merges = normalize_vietmax_internal_merges(data.get("vietmax_mua_vao_internal_merges") or [], vietmax_comparison_scope) if uses_vietmax_purchase else {}
    vietmax_sales_split_code_overrides = normalize_vietmax_split_code_overrides(data.get("vietmax_ban_ra_sales_internal_merges") or [], vietmax_comparison_scope) if uses_vietmax_sales else {}
    vietmax_purchase_split_code_overrides = normalize_vietmax_split_code_overrides(data.get("vietmax_mua_vao_internal_merges") or [], vietmax_comparison_scope) if uses_vietmax_purchase else {}
    prefix_map, selected_products = validate_payload(data)
    rows = company_rows(df, company_col, mst_col, product_col, qty_col, price_col, invoice_status_col=invoice_status_col, invoice_status_skip_values=invoice_status_skip_values, progress_callback=progress_callback)
    company_group_assignments = data.get("company_group_assignments") if isinstance(data.get("company_group_assignments"), dict) else {}
    selected_company_ids = set(str(x).strip() for x in data.get("process_mst", []))
    tag_processing_groups = bool(company_group_assignments)

    def row_processing_group(row):
        mst = row.get("mst")
        company = row.get("company")
        candidates = [
            row.get("selection_key"),
            row.get("company_key"),
            row.get("safe_id"),
            mst,
            company,
            company_selection_key(company, mst),
            vietmax_company_identity_key(company, mst),
        ]
        for candidate in candidates:
            key = raw_text(candidate)
            if key and key in company_group_assignments:
                return raw_text(company_group_assignments.get(key)) or FAST_MATERIALS_GROUP_ID
        company_id = row.get("selection_key") or company_selection_key(company, mst)
        if company_id in selected_company_ids:
            return FAST_MATERIALS_GROUP_ID
        return FAST_IGNORED_GROUP_ID

    processed_purchase_path = raw_text(data.get("vietmax_processed_purchase_path"))
    if uses_vietmax_sales and processed_purchase_path:
        processed_purchase_products = vietmax_purchase_products_from_workbook(
            processed_purchase_path,
            price_col=data.get("purchase_price_col") or "P",
            comparison_scope=vietmax_comparison_scope,
            require_existing_code=True,
        )
        sales_products_for_khh = []
        for row in rows:
            sales_row = {
                "sales_product": row.get("product", ""),
                "sales_unit": row.get("unit", ""),
                "invoice_no": row.get("invoice_no", ""),
                "invoice_date": row.get("invoice_date", ""),
                "sales_company": row.get("company", ""),
                "sales_mst": row.get("mst", ""),
            }
            add_vietmax_company_identity(sales_row, "sales", row.get("company", ""), row.get("mst", ""))
            sales_products_for_khh.append(sales_row)
        exact_khh_matches = build_vietmax_khh_exact_purchase_matches(sales_products_for_khh, processed_purchase_products, vietmax_comparison_scope)
        if exact_khh_matches:
            current_purchase_matches = normalize_vietmax_ban_ra_purchase_matches(vietmax_purchase_match_rows, vietmax_comparison_scope)
            enriched_exact_matches = []
            for exact_match in exact_khh_matches:
                exact_key = vietmax_purchase_match_row_key(exact_match, vietmax_comparison_scope)
                enriched_exact_matches.append(preserve_vietmax_conversion_settings(exact_match, current_purchase_matches.get(exact_key)))
            vietmax_purchase_match_rows = merge_vietmax_purchase_match_rows(
                enriched_exact_matches,
                vietmax_purchase_match_rows,
                vietmax_comparison_scope,
            )
            vietmax_purchase_matches = normalize_vietmax_ban_ra_purchase_matches(vietmax_purchase_match_rows, vietmax_comparison_scope)

    occupied = {}
    for row_index, r in enumerate(rows):
        report_loop_progress(progress_callback, row_index + 1, len(rows), "Đang kiểm tra quy tắc giá")
        mst = r["mst"]
        company_id = r.get("selection_key") or company_selection_key(r.get("company"), mst)
        prod = r["product"]
        if (include_company_prefix and company_id not in prefix_map) or prod not in selected_products.get(company_id, set()):
            continue
        key = product_key(company_id, prod)
        purchase_merge = find_vietmax_internal_merge(
            vietmax_purchase_internal_merges,
            prod,
            r.get("company"),
            mst,
            r.get("company_key"),
            vietmax_comparison_scope,
        ) if uses_vietmax_purchase else None
        effective_purchase_prod = raw_text(purchase_merge.get("similar_product")) if purchase_merge else prod
        base_code = find_vietmax_split_code_override(
            vietmax_purchase_split_code_overrides,
            prod,
            r.get("company"),
            mst,
            r.get("company_key"),
            vietmax_comparison_scope,
        ) if uses_vietmax_purchase else ""
        if not base_code:
            base_code = str(manual_code_overrides.get(key) or "").strip()
        if not base_code:
            base_code = make_code(company_id, effective_purchase_prod, 1, prefix_map, effective_profile, word_rules, first_word_rules, require_qty=False, include_company_prefix=include_company_prefix, repeated_phrase_removals=repeated_phrase_removals)
        base_code = sanitize_product_code(base_code)
        base_code = apply_product_code_replacement(base_code, product_code_replacements)
        rule = price_rules.get(key) or price_range_rules.get(base_code)
        if not rule:
            continue
        group = raw_price_group(r["price"], rule)
        if group is not None:
            occupied.setdefault(key, set()).add(group)

    processed_row_indexes = set()
    retained_group_row_indexes = set()
    processing_group_by_row_index = {}
    vietmax_purchase_match_row_indexes = set()
    sales_customer_codes = {}
    sales_amount_index = qi + 4 if uses_vietmax_sales and qi is not None else None
    sales_tax_rate_index = qi + 3 if uses_vietmax_sales and qi is not None else None
    sales_tax_amount_index = qi + 5 if uses_vietmax_sales and qi is not None else None
    sales_total_index = qi + 6 if uses_vietmax_sales and qi is not None else None
    for row_index, row in enumerate(rows):
        report_loop_progress(progress_callback, row_index + 1, len(rows), "Đang tạo mã VT cho dòng xử lý")
        i = row["excel_row"] - 1
        mst = row["mst"]
        company_id = row.get("selection_key") or company_selection_key(row.get("company"), mst)
        processing_group = row_processing_group(row) if tag_processing_groups else FAST_MATERIALS_GROUP_ID
        if tag_processing_groups:
            processing_group_by_row_index[i] = processing_group
            retained_group_row_indexes.add(i)
        prod = row["product"]
        key = product_key(company_id, prod)
        qty = cell(df, i, qi) if qi is not None else 1
        purchase_merge = find_vietmax_internal_merge(
            vietmax_purchase_internal_merges,
            prod,
            row.get("company"),
            mst,
            row.get("company_key"),
            vietmax_comparison_scope,
        ) if uses_vietmax_purchase else None
        purchase_effective_prod = raw_text(purchase_merge.get("similar_product")) if purchase_merge else prod
        sales_merge = find_vietmax_internal_merge(
            vietmax_sales_internal_merges,
            prod,
            row.get("company"),
            mst,
            row.get("company_key"),
            vietmax_comparison_scope,
        )
        effective_prod = raw_text(sales_merge.get("similar_product")) if sales_merge else purchase_effective_prod
        same_company_match_key = vietmax_purchase_match_lookup_key(effective_prod, row.get("company"), mst, VIETMAX_COMPARISON_SCOPE_SAME_COMPANY)
        matched_purchase_match = vietmax_purchase_matches.get(same_company_match_key) if same_company_match_key else None
        if not matched_purchase_match and vietmax_comparison_scope != VIETMAX_COMPARISON_SCOPE_SAME_COMPANY:
            matched_purchase_match = vietmax_purchase_matches.get(vietmax_ban_ra_match_key(effective_prod))
        matched_purchase_code = raw_text(matched_purchase_match.get("purchase_code")) if matched_purchase_match else ""
        selected_for_processing = not ((include_company_prefix and company_id not in prefix_map) or prod not in selected_products.get(company_id, set()))
        if not selected_for_processing and not matched_purchase_match:
            continue
        code = matched_purchase_code if matched_purchase_match and not selected_for_processing else ""
        if not code and uses_vietmax_purchase:
            code = find_vietmax_split_code_override(
                vietmax_purchase_split_code_overrides,
                prod,
                row.get("company"),
                mst,
                row.get("company_key"),
                vietmax_comparison_scope,
            )
        if not code and uses_vietmax_sales:
            code = find_vietmax_split_code_override(
                vietmax_sales_split_code_overrides,
                prod,
                row.get("company"),
                mst,
                row.get("company_key"),
                vietmax_comparison_scope,
            )
        if not code:
            code = str(manual_code_overrides.get(key) or "").strip()
        code = sanitize_product_code(code)
        if not code:
            code = make_code(company_id, effective_prod, qty, prefix_map, effective_profile, word_rules, first_word_rules, require_qty=(qi is not None), include_company_prefix=include_company_prefix, repeated_phrase_removals=repeated_phrase_removals)
        code = apply_product_code_replacement(code, product_code_replacements)
        rule = price_rules.get(key) or price_range_rules.get(code)
        if code and rule and pri is not None:
            code += price_group_suffix(parse_price(cell(df, i, pri)), rule, occupied.get(key))
        code = sanitize_product_code(code)
        if code:
            df.iat[i, oi] = code
            if uses_vietmax_sales:
                customer_code = raw_text(mst) or raw_text(prefix_map.get(company_id))
                if customer_code:
                    sales_customer_codes[i] = customer_code
            if purchase_merge:
                df.iat[i, pi] = purchase_effective_prod
                similar_unit = raw_text(purchase_merge.get("similar_unit"))
                unit_index = pi + 1
                if similar_unit and unit_index < df.shape[1]:
                    df.iat[i, unit_index] = similar_unit
            processed_row_indexes.add(i)
            if matched_purchase_match:
                vietmax_purchase_match_row_indexes.add(i)
                apply_vietmax_match_conversion_to_row(
                    df,
                    i,
                    qi,
                    pi + 1,
                    matched_purchase_match,
                    amount_index=sales_amount_index,
                    tax_rate_index=sales_tax_rate_index,
                    tax_amount_index=sales_tax_amount_index,
                    total_index=sales_total_index,
                )

    header_index = None
    for i in range(len(df)):
        output_header = rm_accents(raw_text(cell(df, i, oi))).upper()
        product_header = rm_accents(raw_text(cell(df, i, pi))).upper()
        qty_header = rm_accents(raw_text(cell(df, i, qi))).upper() if qi is not None else ""
        if "MA VT" in output_header or ("TEN" in product_header and "HANG" in product_header) or "SO LUONG" in qty_header:
            header_index = i
            break
    if header_index is None:
        header_index = min(processed_row_indexes, default=len(df)) - 1

    if tag_processing_groups:
        group_col_index = df.shape[1]
        df[group_col_index] = ""
        if 0 <= header_index < len(df):
            df.iat[header_index, group_col_index] = FAST_PROCESSING_GROUP_HEADER
        for row_index, group_id in processing_group_by_row_index.items():
            if 0 <= row_index < len(df):
                df.iat[row_index, group_col_index] = group_id

    df = apply_inventory_pairs(df, header_index, processed_row_indexes, data, excluded_row_indexes=vietmax_purchase_match_row_indexes)
    df = apply_vietmax_ban_ra_match_inventory(df, header_index, vietmax_purchase_match_row_indexes)
    if uses_vietmax_sales:
        df = apply_fast_customer_codes(df, header_index, sales_customer_codes)

    kept_data_row_indexes = processed_row_indexes | retained_group_row_indexes
    keep_indexes = [
        i for i in range(len(df))
        if i <= header_index or i in kept_data_row_indexes
    ]
    df = df.iloc[keep_indexes].reset_index(drop=True)

    if progress_callback:
        progress_callback(1, 2, "Đang ghi file Excel chính")
    write_dataframe_workbook(df, sheet, out)
    if progress_callback:
        progress_callback(2, 2, "Đã ghi file Excel chính")
    return df


def up_ban_ra_tax_values(value):
    rate = parse_price(value)
    if rate is None:
        return "", ""
    percent = rate * 100 if abs(rate) <= 1 else rate
    rounded = int(round(percent))
    if abs(percent - rounded) < 0.000001:
        return f"{rounded:02d}", rounded
    return f"{percent:02g}", percent


def up_template_sheet(workbook, expected_title=None):
    expected = rm_accents(expected_title).casefold() if expected_title else ""
    for ws in workbook.worksheets:
        if expected and rm_accents(ws.title).casefold() == expected:
            return ws
    return workbook.worksheets[0]


def up_ban_ra_template_sheet(workbook):
    return up_template_sheet(workbook, "UP Bán ra")


def up_mua_vao_template_sheet(workbook):
    return up_template_sheet(workbook, "UP Mua vào")


def processed_inventory_column_indexes(processed_df, output_code_index):
    labels = {"tk vat tu": None, "ma kho": None}
    header_index = None
    for row in range(len(processed_df)):
        if "ma vt" in normalized_header_label(cell(processed_df, row, output_code_index)):
            header_index = row
            break
    if header_index is None:
        return {"tk_vat_tu": None, "ma_kho": None}
    for col in range(processed_df.shape[1]):
        key = normalized_header_label(cell(processed_df, header_index, col))
        if key in labels and labels[key] is None:
            labels[key] = col
    return {"tk_vat_tu": labels["tk vat tu"], "ma_kho": labels["ma kho"]}


def processed_customer_code_column_index(processed_df, output_code_index):
    header_index = None
    for row in range(len(processed_df)):
        if "ma vt" in normalized_header_label(cell(processed_df, row, output_code_index)):
            header_index = row
            break
    if header_index is None:
        return None
    for col in range(processed_df.shape[1]):
        if normalized_header_label(cell(processed_df, header_index, col)) == "ma khach hang":
            return col
    return None


def up_workbook_template_path(*names):
    for name in names:
        path = RESOURCE_DIR / name
        if path.exists():
            return path
    return None


def prepare_up_workbook(template_path, sheet_title, expected_title=None):
    from openpyxl import load_workbook

    wb = load_workbook(template_path)
    ws = up_template_sheet(wb, expected_title or sheet_title)
    ws.title = sheet_title
    for other in list(wb.worksheets):
        if other is not ws:
            wb.remove(other)

    max_column = ws.max_column
    style_row = 2 if ws.max_row >= 2 else 1
    data_styles = [copy(ws.cell(style_row, col)._style) for col in range(1, max_column + 1)]
    data_number_formats = [ws.cell(style_row, col).number_format for col in range(1, max_column + 1)]
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    return wb, ws, max_column, data_styles, data_number_formats


def up_template_headers(template_path, sheet_title):
    from openpyxl import load_workbook

    wb = load_workbook(template_path, read_only=True, data_only=True)
    try:
        ws = up_template_sheet(wb, sheet_title)
        max_column = ws.max_column or 1
        first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        headers = []
        for index in range(max_column):
            value = first_row[index] if index < len(first_row) else ""
            headers.append(raw_text(value) or index_to_excel_col(index))
        return headers
    finally:
        wb.close()


def clean_up_cell_value(value, default=""):
    try:
        if pd.isna(value):
            return default
    except (TypeError, ValueError):
        pass
    return value


def ensure_xls_dimensions(row_count, column_count, label="Workbook"):
    if row_count > XLS_MAX_ROWS:
        raise ValueError(f"{label} có {row_count} dòng, vượt giới hạn {XLS_MAX_ROWS} dòng của định dạng .xls.")
    if column_count > XLS_MAX_COLUMNS:
        raise ValueError(f"{label} có {column_count} cột, vượt giới hạn {XLS_MAX_COLUMNS} cột của định dạng .xls.")


def coerce_xls_value(value):
    value = clean_up_cell_value(value)
    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime()
    if isinstance(value, (np.integer,)):
        return int(value)
    if isinstance(value, (np.floating,)):
        return float(value)
    if value is None:
        return ""
    return value


def write_dataframe_xls(df, sheet_name, out):
    import xlwt

    ensure_xls_dimensions(len(df), int(df.shape[1]), Path(out).name)
    workbook = xlwt.Workbook(encoding="utf-8")
    used_names = set()
    ws = workbook.add_sheet(safe_excel_sheet_name(sheet_name, used_names))
    for row_index in range(len(df)):
        for column_index in range(int(df.shape[1])):
            ws.write(row_index, column_index, coerce_xls_value(df.iat[row_index, column_index]))
    workbook.save(str(out))


def write_dataframe_workbook(df, sheet_name, out):
    if Path(out).suffix.lower() == ".xls":
        write_dataframe_xls(df, sheet_name, out)
        return
    with pd.ExcelWriter(out, engine="openpyxl") as w:
        df.to_excel(w, sheet_name=sheet_name, index=False, header=False)


def openpyxl_workbook_to_xls_stream(path_or_workbook):
    import xlwt
    from datetime import date, datetime
    from openpyxl import load_workbook

    if isinstance(path_or_workbook, (str, Path)):
        source = load_workbook(path_or_workbook, data_only=True)
        source_label = Path(path_or_workbook).name
        should_close = True
    else:
        source = path_or_workbook
        source_label = "workbook"
        should_close = False
    output = BytesIO()
    try:
        workbook = xlwt.Workbook(encoding="utf-8")
        used_names = set()
        date_style = xlwt.easyxf(num_format_str="DD/MM/YYYY")
        datetime_style = xlwt.easyxf(num_format_str="DD/MM/YYYY HH:MM:SS")
        for sheet in source.worksheets:
            ensure_xls_dimensions(sheet.max_row, sheet.max_column, f"{source_label}/{sheet.title}")
            target = workbook.add_sheet(safe_excel_sheet_name(sheet.title, used_names))
            for row in sheet.iter_rows():
                for cell_obj in row:
                    value = coerce_xls_value(cell_obj.value)
                    if isinstance(value, datetime):
                        target.write(cell_obj.row - 1, cell_obj.column - 1, value, datetime_style)
                    elif isinstance(value, date):
                        target.write(cell_obj.row - 1, cell_obj.column - 1, value, date_style)
                    else:
                        target.write(cell_obj.row - 1, cell_obj.column - 1, value)
        workbook.save(output)
    finally:
        if should_close:
            source.close()
    output.seek(0)
    return output


def write_up_output_row(ws, output_row, max_column, data_styles, data_number_formats, values):
    for column, value in values.items():
        column_index = excel_col_to_index(column) + 1
        if column_index > max_column:
            continue
        output_cell = ws.cell(output_row, column_index)
        output_cell._style = copy(data_styles[column_index - 1])
        output_cell.number_format = data_number_formats[column_index - 1]
        output_cell.value = clean_up_cell_value(value)


def up_values_to_row(headers, values):
    row = [""] * len(headers)
    for column, value in values.items():
        column_index = excel_col_to_index(column)
        if 0 <= column_index < len(row):
            row[column_index] = clean_up_cell_value(value)
    return row


def is_processed_product_data_row(processed_df, row, code_col, invoice_col, qty_col, parse_quantity=True):
    code = raw_text(cell(processed_df, row, code_col))
    invoice_no = raw_text(cell(processed_df, row, invoice_col))
    quantity_value = cell(processed_df, row, qty_col)
    if parse_quantity:
        quantity = parse_price(quantity_value)
        has_quantity = quantity is not None
    else:
        has_quantity = bool(raw_text(quantity_value))
    if not code or not invoice_no or not has_quantity:
        return False
    return rm_accents(code).casefold() != "ma vt"


def up_source_value(processed_df, row, col, default=""):
    return clean_up_cell_value(cell(processed_df, row, col), default)


def up_source_number(processed_df, row, col, default=""):
    value = parse_price(cell(processed_df, row, col))
    return default if value is None else value


def create_up_mua_vao_sheet_data(processed_df, progress_callback=None, require_inventory_columns=True):
    template_path = up_workbook_template_path("mau_mua_vao_up.xlsx")
    if not template_path:
        raise ValueError("Không tìm thấy file mẫu mau_mua_vao_up.xlsx để tạo file UP mua vào.")
    validate_fast_import_processed_dataframe(processed_df, "FDI mua vào đã xử lý", require_inventory_columns=require_inventory_columns)

    headers = up_template_headers(template_path, "UP Mua vào")
    source_indexes = {name: excel_col_to_index(name) for name in ["A", "B", "C", "D", "F", "G", "L", "O", "P", "R", "V", "AQ", "AR"]}
    inventory_indexes = processed_inventory_column_indexes(processed_df, source_indexes["L"])
    rows = []
    for row in range(len(processed_df)):
        report_loop_progress(progress_callback, row + 1, len(processed_df), "Đang dựng file UP mua vào", interval=200)
        if not is_processed_product_data_row(processed_df, row, source_indexes["L"], source_indexes["C"], source_indexes["O"]):
            continue

        seller_mst = up_source_value(processed_df, row, source_indexes["G"])
        invoice_no = up_source_value(processed_df, row, source_indexes["C"])
        invoice_date = up_source_value(processed_df, row, source_indexes["D"])
        tax_code, _ = up_ban_ra_tax_values(cell(processed_df, row, source_indexes["R"]))
        values = {
            "A": seller_mst,
            "B": "",
            "F": invoice_no,
            "G": invoice_date,
            "H": "",
            "I": "",
            "J": up_source_value(processed_df, row, inventory_indexes["ma_kho"]) if inventory_indexes["ma_kho"] is not None else "",
            "K": up_source_value(processed_df, row, source_indexes["L"]),
            "L": up_source_value(processed_df, row, source_indexes["O"]),
            "M": up_source_value(processed_df, row, source_indexes["P"]),
            "N": up_source_value(processed_df, row, source_indexes["V"]),
            "O": up_source_value(processed_df, row, inventory_indexes["tk_vat_tu"]) if inventory_indexes["tk_vat_tu"] is not None else "",
            "R": invoice_no,
            "S": invoice_date,
            "T": up_source_value(processed_df, row, source_indexes["A"]),
            "U": up_source_value(processed_df, row, source_indexes["B"]),
            "V": tax_code,
        }
        rows.append(up_values_to_row(headers, values))

    return headers, rows


def create_up_mua_vao_workbook(processed_df, progress_callback=None, require_inventory_columns=True):
    headers, rows = create_up_mua_vao_sheet_data(processed_df, progress_callback, require_inventory_columns=require_inventory_columns)
    if progress_callback:
        progress_callback(1, 2, "Đang lưu bộ nhớ file UP mua vào")
    stream = tabular_xls_workbook({"UP Mua vào": (headers, rows)})
    if progress_callback:
        progress_callback(2, 2, "Đã tạo file UP mua vào")
    return stream


def create_up_ban_ra_workbook(processed_df, progress_callback=None, require_inventory_columns=True):
    template_path = up_workbook_template_path("mau_ban_ra_up.xlsx", "mau HD ban ra.xlsx")
    if not template_path:
        raise ValueError("Không tìm thấy file mẫu UP bán ra để tạo file UP bán ra.")
    validate_fast_import_processed_dataframe(processed_df, "FDI bán ra đã xử lý", require_inventory_columns=require_inventory_columns)

    headers = up_template_headers(template_path, "UP Bán ra")
    source_indexes = {name: excel_col_to_index(name) for name in ["C", "D", "I", "J", "L", "M", "O", "P", "R", "S", "V", "W", "AQ", "AR"]}
    inventory_indexes = processed_inventory_column_indexes(processed_df, source_indexes["L"])
    rows = []
    for row in range(len(processed_df)):
        report_loop_progress(progress_callback, row + 1, len(processed_df), "Đang dựng file UP bán ra", interval=200)
        if not is_processed_product_data_row(processed_df, row, source_indexes["L"], source_indexes["C"], source_indexes["O"]):
            continue

        tax_code, tax_rate = up_ban_ra_tax_values(cell(processed_df, row, source_indexes["R"]))
        values = {
            "A": up_source_value(processed_df, row, source_indexes["J"]),
            "B": "",
            "E": up_source_value(processed_df, row, source_indexes["C"]),
            "F": up_source_value(processed_df, row, source_indexes["D"]),
            "H": up_source_value(processed_df, row, source_indexes["O"]),
            "I": up_source_value(processed_df, row, source_indexes["P"]),
            "J": up_source_value(processed_df, row, source_indexes["S"]),
            "M": "",
            "N": "",
            "O": up_source_value(processed_df, row, source_indexes["P"]),
            "P": up_source_value(processed_df, row, source_indexes["V"]),
            "V": tax_code,
            "W": tax_rate,
            "Y": up_source_value(processed_df, row, source_indexes["W"]),
            "AB": up_source_value(processed_df, row, inventory_indexes["tk_vat_tu"]) if inventory_indexes["tk_vat_tu"] is not None else "",
            "AF": up_source_value(processed_df, row, inventory_indexes["ma_kho"]) if inventory_indexes["ma_kho"] is not None else "",
            "AG": up_source_value(processed_df, row, source_indexes["L"]),
            "AH": up_source_value(processed_df, row, source_indexes["M"]),
        }
        rows.append(up_values_to_row(headers, values))

    if progress_callback:
        progress_callback(1, 2, "Đang lưu bộ nhớ file UP bán ra")
    stream = tabular_xls_workbook({"UP Bán ra": (headers, rows)})
    if progress_callback:
        progress_callback(2, 2, "Đã tạo file UP bán ra")
    return stream


FAST_HOA_DON_BAN_HANG_HEADERS = [
    "Mã khách hàng (ma_kh)",
    "Người mua hàng (ong_ba)",
    "Quyển sổ (ma_qs)",
    "Số seri (so_seri)",
    "Số chứng từ (so_ct)",
    "Ngày chứng từ (ngay_ct)",
    "NVBH (ma_bp)",
    "Số lượng:Q (so_luong)",
    "Giá bán n.tệ:P1 (gia_nt2)",
    "Tiền bán n.tệ:N1 (tien_nt2)",
    "Giá vốn n.tệ:P1 (gia_nt)",
    "Tiền vốn n.tệ:N1 (tien_nt)",
    "Mã n.tệ (ma_nt)",
    "Tỷ giá:R (ty_gia)",
    "Giá bán:P0 (gia2)",
    "Tiền bán:N0 (tien2)",
    "Giá vốn:P0 (gia)",
    "Tiền vốn:N0 (tien)",
    "Tỷ lệ chiết khấu (tl_ck)",
    "Tiền chiết khấu n.tệ:N1 (tien_ck_nt)",
    "Tiền chiết khấu:N0 (tien_ck)",
    "Mã thuế (ma_thue)",
    "Thuế suất (thue_suat)",
    "Tiền thuế n.tệ:N1 (tien_thue_nt)",
    "Tiền thuế:N0 (tien_thue)",
    "Mã nx (Tk nợ) (ma_nx)",
    "Tk doanh thu (tk_dt)",
    "Tk vật tư (tk_vt)",
    "Tk giá vốn (tk_gv)",
    "Tk chiết khấu (tk_ck)",
    "Tài khoản thuế (tk_thue_co)",
    "Mã kho  (ma_kho)",
    "Mã vật tư (ma_vt)",
    "Diễn giải (dien_giai)",
    "Hạn thanh toán:N (han_tt)",
    "Hình thức tt (ht_tt)",
    "Loại hoá đơn (ma_gd)",
    "Mã dự án (ma_vv_i)",
    "Mã phí (ma_phi_i)",
    "Mã bpht (ma_bpht_i)",
    "Khuyến mại (km_ck)",
    "Tk cp km (tk_km_i)",
    "Mã ĐVCS (ma_dvcs)",
    "Sử dụng HĐĐT (sd_hddt_yn)",
    "Tỷ giá hđ:R (ty_gia_hd)",
    "Thêm thông tin (xu_ly_hddt)",
    "Thông tin thêm (gc_dc_hddt)",
    "Mã loại (ma_loai)",
]


FAST_HOA_DON_MUA_HANG_HEADERS = [
    "Mã khách hàng (ma_kh)",
    "Người mua hàng (ong_ba)",
    "Diễn giải (dien_giai)",
    "Quyển sổ (ma_qs)",
    "Mã nx (Tk có) (ma_nx)",
    "Số chứng từ (so_ct)",
    "Ngày chứng từ (ngay_ct)",
    "Mã ngoại tệ (ma_nt)",
    "Tỷ giá (ty_gia)",
    "Mã kho  (ma_kho)",
    "Mã vật tư (ma_vt)",
    "Số lượng:Q (so_luong)",
    "Giá mua chưa thuế:P0 (gia0)",
    "Tiền mua:N0 (tien0)",
    "Tk vật tư (tk_vt)",
    "Mã dự án (ma_vv_i)",
    "Mã ĐVCS (ma_dvcs)",
    "Số HĐ thuế (so_ct0)",
    "Ngày HĐ (ngay_ct0)",
    "Mẫu HĐ (kh_mau_hd)",
    "Số seri (so_seri0)",
    "Mã thuế (ma_thue)",
]


FAST_DM_VAT_TU_HEADERS = [
    "Mã vật tư (ma_vt)",
    "Tên vật tư (ten_vt)",
    "Đơn vị tính (dvt)",
    "Theo dõi tồn kho (vt_ton_kho)",
    "Tk vật tư (tk_vt)",
    "Tk giá vốn (tk_gv)",
    "Tk doanh thu (tk_dt)",
    "Tk hàng bán bị trả lại (tk_tl)",
    "Tk sp dở dang (tk_spdd)",
    "Loại vật tư (loai_vt)",
    "Cho sửa tk kho (sua_tk_vt)",
    "Tk NVL (tk_nvl)",
    "Tk chiết khấu (tk_ck)",
    "Tk khuyến mại (tk_km)",
    "Mã phụ (part_no)",
    "Tên 2 (ten_vt2)",
    "Cách tính giá tồn kho (gia_ton)",
    "Nhóm vt 1 (nh_vt1)",
    "Nhóm vt 2 (nh_vt2)",
    "Nhóm vt 3 (nh_vt3)",
    "Số lượng tồn tối thiểu:Q (sl_min)",
    "Số lượng tồn tối đa:Q (sl_max)",
    "Tk chênh lệch vật tư (tk_cl_vt)",
    "Tk doanh thu nội bộ (tk_dtnb)",
    "Ghi chú (ghi_chu)",
    "Mã tra cứu (ma_tra_cuu)",
    "Mã kho (ma_kho_def)",
    "Mã t.suất (ma_thue)",
    "Thuế suất (thue_suat)",
    "Chỉ có ở kho này (chi_co_kho_nay)",
]


FAST_DM_KHACH_HANG_HEADERS = [
    "Mã khách hàng (ma_kh)",
    "Tên khách hàng (ten_kh)",
    "Tên 2 (ten_kh2)",
    "Mã số thuế (ma_so_thue)",
    "Địa chỉ (dia_chi)",
    "Điện thoại (dien_thoai)",
    "Fax (fax)",
    "Mail (e_mail)",
    "Đối tác (doi_tac)",
    "Tên ngân hàng (ten_nh)",
    "Ghi chú (ghi_chu)",
    "Mã điều khoản thanh toán ngầm định (ma_thck)",
    "Tk mặc định (tk)",
    "Nhóm kh 1 (nh_kh1)",
    "Nhóm kh 2 (nh_kh2)",
    "Nhóm kh 3 (nh_kh3)",
    "Tài khoản ngân hàng (tk_nh)",
    "Tỉnh thành (tinh_thanh)",
    "Mã tra cứu (ma_tra_cuu)",
]

FAST_DUPLICATE_PURCHASE_REPORT_HEADERS = [
    "Số chứng từ",
    "Mã khách hàng",
    "Tên công ty",
    "Ngày chứng từ",
    "Dòng FDI",
    "Mã VT",
    "Mã kho",
    "TK vật tư",
    "Sheet FAST",
]


FAST_MATERIALS_GROUP_ID = "materials"
FAST_IGNORED_GROUP_ID = "ignored"
FAST_PROCESSING_GROUP_HEADER = "Nhóm xử lý"
FAST_REVENUE_ACCOUNT_RULES = [
    {"match_type": "starts_with", "value": "152", "result": "5111"},
    {"match_type": "starts_with", "value": "155", "result": "5112"},
    {"match_type": "default", "value": "", "result": "5112"},
]
FAST_MATERIAL_TYPE_RULES = [
    {"match_type": "starts_with", "value": "152", "result": "21"},
    {"match_type": "default", "value": "", "result": "51"},
]


def fast_output_columns(headers):
    return [
        {
            "letter": index_to_excel_col(index),
            "header": raw_text(header),
            "label": f"{index_to_excel_col(index)}. {raw_text(header)}",
        }
        for index, header in enumerate(headers)
    ]


VIETMAX_FDI_SOURCE_HEADERS = {
    "A": "Mẫu HĐ",
    "B": "Số seri",
    "C": "Số chứng từ",
    "D": "Ngày chứng từ",
    "F": "Người bán",
    "G": "MST người bán",
    "H": "Địa chỉ người bán",
    "I": "Người mua",
    "J": "MST người mua",
    "K": "Địa chỉ người mua",
    "L": "Mã VT",
    "M": "Tên hàng",
    "N": "ĐVT",
    "O": "Số lượng",
    "P": "Đơn giá",
    "R": "Mã thuế / thuế suất",
    "S": "Thành tiền",
    "V": "Tiền hàng",
    "W": "Tiền thuế",
    "AQ": "Mã ngoại tệ",
    "AR": "Tỷ giá",
    "AS": "TK vật tư",
    "AT": "Mã kho",
    "AU": "Mã khách hàng",
}


def vietmax_default_source_columns(phase=None):
    return [
        {"letter": letter, "header": header, "label": f"{letter}. {header}"}
        for letter, header in VIETMAX_FDI_SOURCE_HEADERS.items()
    ]


def fast_mapping_rule(target_col, source_col="", source_phase="purchase", source_type="source_column", **extra):
    rule = {
        "target_col": target_col,
        "source_type": source_type,
        "source_phase": source_phase,
    }
    if source_col:
        rule["source_col"] = source_col
    rule.update({key: value for key, value in extra.items() if value is not None})
    return rule


def default_vietmax_form_mapping_presets(phase=None):
    purchase_form = {
        "id": "fast_hoadonmuahang",
        "label": "Hoadonmuahang",
        "scope": "purchase",
        "type": "builtin",
        "enabled": True,
        "builtin_exporter": "fast_hoadonmuahang",
        "group_id": "materials",
        "input_phase": "purchase",
        "sheet": "Hoadonmuahang",
        "output_columns": fast_output_columns(FAST_HOA_DON_MUA_HANG_HEADERS),
        "mappings": [
            fast_mapping_rule("A", "G", "purchase", transform="mst_or_prefix"),
            fast_mapping_rule("C", source_type="text_template", source_phase="purchase", value="Mua hàng nhập kho HD{C}"),
            fast_mapping_rule("E", source_type="constant", source_phase="purchase", value="331"),
            fast_mapping_rule("F", "C", "purchase"),
            fast_mapping_rule("G", "D", "purchase"),
            fast_mapping_rule("H", "AQ", "purchase"),
            fast_mapping_rule("I", "AR", "purchase"),
            fast_mapping_rule("J", "AT", "purchase"),
            fast_mapping_rule("K", "L", "purchase"),
            fast_mapping_rule("L", "O", "purchase"),
            fast_mapping_rule("M", "P", "purchase"),
            fast_mapping_rule("N", "V", "purchase"),
            fast_mapping_rule("O", "AS", "purchase"),
            fast_mapping_rule("Q", source_type="constant", source_phase="purchase", value="CTY"),
            fast_mapping_rule("R", "C", "purchase"),
            fast_mapping_rule("S", "D", "purchase"),
            fast_mapping_rule("V", "R", "purchase", transform="tax_code"),
        ],
    }
    sales_form = {
        "id": "fast_hoadonbanhang",
        "label": "Hoadonbanhang",
        "scope": "sales",
        "type": "builtin",
        "enabled": True,
        "builtin_exporter": "fast_hoadonbanhang",
        "group_id": "materials",
        "input_phase": "sales",
        "sheet": "Hoadonbanhang",
        "output_columns": fast_output_columns(FAST_HOA_DON_BAN_HANG_HEADERS),
        "mappings": [
            fast_mapping_rule("A", "AU", "sales", transform="mst_or_prefix"),
            fast_mapping_rule("E", "C", "sales"),
            fast_mapping_rule("F", "D", "sales"),
            fast_mapping_rule("H", "O", "sales"),
            fast_mapping_rule("O", "P", "sales"),
            fast_mapping_rule("P", "V", "sales"),
            fast_mapping_rule("V", "R", "sales", transform="tax_code"),
            fast_mapping_rule("Y", "W", "sales"),
            fast_mapping_rule("Z", source_type="constant", source_phase="sales", value="131"),
            fast_mapping_rule("AA", "", "sales", source_type="if_rules", condition_source_col="AS", transform_rules=FAST_REVENUE_ACCOUNT_RULES),
            fast_mapping_rule("AB", "AS", "sales"),
            fast_mapping_rule("AC", source_type="constant", source_phase="sales", value="632"),
            fast_mapping_rule("AE", source_type="constant", source_phase="sales", value="33311"),
            fast_mapping_rule("AF", "AT", "sales"),
            fast_mapping_rule("AG", "L", "sales"),
            fast_mapping_rule("AH", source_type="constant", source_phase="sales", value="Xuất bán hàng"),
            fast_mapping_rule("AK", source_type="constant", source_phase="sales", value="1"),
            fast_mapping_rule("AQ", source_type="constant", source_phase="sales", value="CTY"),
            fast_mapping_rule("AS", "AR", "sales"),
        ],
    }
    material_form = {
        "id": "fast_dm_vat_tu",
        "label": "Danh sách vật tư",
        "scope": "both",
        "type": "builtin",
        "enabled": True,
        "builtin_exporter": "fast_dm_vat_tu",
        "group_id": "materials",
        "input_phase": "both",
        "sheet": "DMvat_tu",
        "output_columns": fast_output_columns(FAST_DM_VAT_TU_HEADERS),
        "mappings": [
            fast_mapping_rule("A", "L", "both"),
            fast_mapping_rule("B", "M", "both"),
            fast_mapping_rule("C", "N", "both"),
            fast_mapping_rule("D", source_type="constant", source_phase="both", value="1"),
            fast_mapping_rule("E", "AS", "both"),
            fast_mapping_rule("F", source_type="constant", source_phase="both", value="632"),
            fast_mapping_rule("G", "", "both", source_type="if_rules", condition_source_col="AS", transform_rules=FAST_REVENUE_ACCOUNT_RULES),
            fast_mapping_rule("J", "", "both", source_type="if_rules", condition_source_col="AS", transform_rules=FAST_MATERIAL_TYPE_RULES),
            fast_mapping_rule("K", source_type="constant", source_phase="both", value="1"),
            fast_mapping_rule("Q", source_type="constant", source_phase="both", value="4"),
            fast_mapping_rule("W", source_type="constant", source_phase="both", value="632"),
        ],
    }
    customer_form = {
        "id": "fast_dm_khach_hang",
        "label": "Danh sách khách hàng",
        "scope": "both",
        "type": "builtin",
        "enabled": True,
        "builtin_exporter": "fast_dm_khach_hang",
        "group_id": "materials",
        "input_phase": "both",
        "sheet": "DMkhachhang",
        "output_columns": fast_output_columns(FAST_DM_KHACH_HANG_HEADERS),
        "mappings": [
            fast_mapping_rule("A", "G", "purchase", transform="mst_or_prefix"),
            fast_mapping_rule("B", "F", "purchase"),
            fast_mapping_rule("D", "G", "purchase"),
            fast_mapping_rule("E", "H", "purchase"),
            fast_mapping_rule("A", "AU", "sales", transform="mst_or_prefix"),
            fast_mapping_rule("B", "I", "sales"),
            fast_mapping_rule("D", "J", "sales"),
            fast_mapping_rule("E", "K", "sales"),
        ],
    }
    duplicate_report_form = {
        "id": "fast_duplicate_invoice_report",
        "label": "B\u00e1o c\u00e1o tr\u00f9ng s\u1ed1 ch\u1ee9ng t\u1eeb",
        "scope": "both",
        "type": "builtin",
        "enabled": True,
        "builtin_exporter": "fast_duplicate_report",
        "group_id": "materials",
        "input_phase": "both",
        "sheet": "Bao_cao_trung_so_ct",
        "system_generated": True,
        "output_columns": fast_output_columns(FAST_DUPLICATE_PURCHASE_REPORT_HEADERS),
        "mappings": [],
    }
    forms = [purchase_form, duplicate_report_form, sales_form, material_form, customer_form]
    normalized_phase = raw_text(phase)
    if normalized_phase == VIETMAX_PHASE_PURCHASE:
        return [form for form in forms if form.get("scope") in {"purchase", "both"}]
    if normalized_phase == VIETMAX_PHASE_SALES:
        return [form for form in forms if form.get("scope") in {"sales", "both"}]
    return forms


SYSTEM_DEFAULT_FORM_MAPPING_VERSION = 2


def default_form_mapping_document():
    return {
        "schema_version": SYSTEM_DEFAULT_FORM_MAPPING_VERSION,
        "product": "ProductCodeFormatter",
        "forms": default_vietmax_form_mapping_presets(),
    }


def normalize_default_form_mapping_document(value):
    value = value if isinstance(value, dict) else {}
    forms = value.get("forms")
    try:
        schema_version = int(value.get("schema_version") or 0)
    except (TypeError, ValueError):
        schema_version = 0
    return {
        "schema_version": schema_version,
        "product": "ProductCodeFormatter",
        "forms": list(forms) if isinstance(forms, list) and forms else default_vietmax_form_mapping_presets(),
    }


def load_system_default_form_mappings(phase=None):
    if not DEFAULT_FORM_MAPPINGS_PATH.exists():
        document = config_store.save_config_file(
            DEFAULT_FORM_MAPPINGS_PATH,
            default_form_mapping_document(),
            normalize_default_form_mapping_document,
        )
    else:
        document = config_store.load_config_file(
            DEFAULT_FORM_MAPPINGS_PATH,
            default_form_mapping_document,
            normalize_default_form_mapping_document,
        )
        if (
            document.get("schema_version") != SYSTEM_DEFAULT_FORM_MAPPING_VERSION
            or not isinstance(document.get("forms"), list)
            or len(document.get("forms") or []) < 5
        ):
            document = config_store.save_config_file(
                DEFAULT_FORM_MAPPINGS_PATH,
                default_form_mapping_document(),
                normalize_default_form_mapping_document,
            )
    forms = json.loads(json.dumps(document.get("forms") or [], ensure_ascii=False))
    normalized_phase = raw_text(phase)
    if normalized_phase == VIETMAX_PHASE_PURCHASE:
        return [form for form in forms if form.get("scope") in {"purchase", "both"}]
    if normalized_phase == VIETMAX_PHASE_SALES:
        return [form for form in forms if form.get("scope") in {"sales", "both"}]
    return forms


def _form_without_runtime_markers(form):
    clean = json.loads(json.dumps(form, ensure_ascii=False, default=raw_text))
    clean.pop("_system_default", None)
    return clean


def merge_system_default_form_mappings(forms, phase=None):
    defaults = load_system_default_form_mappings(phase)
    configured = [form for form in (forms or []) if isinstance(form, dict)]
    configured_by_id = {
        raw_text(form.get("id")): form
        for form in configured
        if raw_text(form.get("id"))
    }
    result = []
    default_ids = set()
    for default in defaults:
        form_id = raw_text(default.get("id"))
        default_ids.add(form_id)
        selected = configured_by_id.get(form_id)
        if selected is None:
            selected = dict(default)
            selected["_system_default"] = True
        result.append(selected)
    result.extend(
        form for form in configured
        if raw_text(form.get("id")) not in default_ids
    )
    return result


def hydrate_builtin_default_form_assets(forms, phase=None):
    defaults = {
        raw_text(form.get("id")): form
        for form in load_system_default_form_mappings(phase)
        if raw_text(form.get("id"))
    }
    hydrated = []
    for form in forms or []:
        if not isinstance(form, dict):
            continue
        default = defaults.get(raw_text(form.get("id")))
        if default is None:
            hydrated.append(form)
            continue
        current = dict(form)
        for key in ("type", "builtin_exporter", "output_columns", "sheet", "system_generated"):
            if key == "output_columns" and isinstance(current.get(key), list) and current.get(key):
                continue
            if key in default:
                current[key] = json.loads(json.dumps(default[key], ensure_ascii=False))
            else:
                current.pop(key, None)
        current.pop("template_saved_name", None)
        current.pop("template_original_name", None)
        current.pop("output_preview", None)
        hydrated.append(current)
    return hydrated


def apply_system_default_forms_to_config(cfg):
    profiles = cfg.get("profiles") if isinstance(cfg, dict) and isinstance(cfg.get("profiles"), dict) else {}

    def initialize_profile(profile, phase=None):
        if not isinstance(profile, dict):
            return
        profile["form_mapping_presets"] = merge_system_default_form_mappings(
            profile.get("form_mapping_presets"), phase
        )
        profile["form_mapping_initialized"] = True
        profile["form_mapping_presets"] = hydrate_builtin_default_form_assets(
            profile.get("form_mapping_presets"), phase
        )

    for profile in profiles.values():
        initialize_profile(profile)
        scopes = profile.get("scopes") if isinstance(profile, dict) and isinstance(profile.get("scopes"), dict) else {}
        for phase in (VIETMAX_PHASE_PURCHASE, VIETMAX_PHASE_SALES):
            initialize_profile(scopes.get(phase), phase)
    return cfg


def strip_system_default_forms_from_profiles(cfg):
    profiles = cfg.get("profiles") if isinstance(cfg, dict) and isinstance(cfg.get("profiles"), dict) else {}
    system_by_phase = {
        "all": {raw_text(form.get("id")): _form_without_runtime_markers(form) for form in load_system_default_form_mappings()},
        VIETMAX_PHASE_PURCHASE: {raw_text(form.get("id")): _form_without_runtime_markers(form) for form in load_system_default_form_mappings(VIETMAX_PHASE_PURCHASE)},
        VIETMAX_PHASE_SALES: {raw_text(form.get("id")): _form_without_runtime_markers(form) for form in load_system_default_form_mappings(VIETMAX_PHASE_SALES)},
    }

    def strip_profile(profile, phase="all"):
        if not isinstance(profile, dict):
            return
        kept = []
        for form in profile.get("form_mapping_presets") or []:
            if not isinstance(form, dict):
                continue
            clean = _form_without_runtime_markers(form)
            system = system_by_phase[phase].get(raw_text(clean.get("id")))
            if system is not None and clean == system:
                continue
            kept.append(clean)
        profile["form_mapping_presets"] = kept
        scopes = profile.get("scopes") if isinstance(profile.get("scopes"), dict) else {}
        for scope_phase in (VIETMAX_PHASE_PURCHASE, VIETMAX_PHASE_SALES):
            strip_profile(scopes.get(scope_phase), scope_phase)

    for profile in profiles.values():
        strip_profile(profile)
    return cfg


def load_license_config():
    return config_store.load_config_file(
        LICENSE_PATH,
        empty_license_config,
        normalize_license_config,
    )


def save_license_config(value):
    return config_store.save_config_file(
        LICENSE_PATH,
        value or {},
        normalize_license_config,
    )


def processed_vietmax_source_indexes():
    return {name: excel_col_to_index(name) for name in ["A", "B", "C", "D", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "R", "S", "V", "W", "AQ", "AR"]}


def processed_group_column_index(processed_df):
    if processed_df is None or processed_df.empty:
        return None
    header_scan_rows = min(len(processed_df), 10)
    for row in range(header_scan_rows):
        for col in range(processed_df.shape[1]):
            header = normalized_header_label(cell(processed_df, row, col))
            if header in {"nhom xu ly", "processing group", "processing_group"}:
                return col
    return None


def processed_row_group(processed_df, row, group_index=None):
    index = processed_group_column_index(processed_df) if group_index is None else group_index
    return raw_text(cell(processed_df, row, index)) if index is not None else ""


def fast_import_cell_value(processed_df, row, column, indexes):
    return up_source_value(processed_df, row, indexes[column])


def fast_import_cell_number(processed_df, row, column, indexes):
    return up_source_number(processed_df, row, indexes[column], 0)


def fast_import_tax_info(processed_df, row, indexes):
    tax_code, tax_rate = up_ban_ra_tax_values(cell(processed_df, row, indexes["R"]))
    return tax_code, tax_rate or 0


def validate_fast_import_processed_dataframe(processed_df, label="FDI đã xử lý", require_inventory_columns=True):
    indexes = processed_vietmax_source_indexes()
    if processed_df is None or processed_df.empty:
        raise ValueError(f"File {label} không có dữ liệu.")
    required_indexes = [indexes["C"], indexes["L"], indexes["O"]]
    if processed_df.shape[1] <= max(required_indexes):
        raise ValueError(f"File {label} không đủ cột FDI mặc định để đọc số hóa đơn, mã VT và số lượng.")
    inventory_indexes = processed_inventory_column_indexes(processed_df, indexes["L"])
    missing_columns = []
    if inventory_indexes["tk_vat_tu"] is None:
        missing_columns.append("TK vật tư")
    if inventory_indexes["ma_kho"] is None:
        missing_columns.append("Mã kho")
    if missing_columns and require_inventory_columns:
        raise ValueError(f"File {label} thiếu cột {', '.join(missing_columns)}. Hãy dùng file FDI đã xử lý từ app, có đủ TK vật tư và Mã kho.")

    total_rows = 0
    missing_rows = []
    group_index = processed_group_column_index(processed_df)
    for row in range(len(processed_df)):
        if not is_processed_product_data_row(processed_df, row, indexes["L"], indexes["C"], indexes["O"], parse_quantity=False):
            continue
        row_group = processed_row_group(processed_df, row, group_index)
        if row_group and row_group != FAST_MATERIALS_GROUP_ID:
            continue
        total_rows += 1
        tk_vat_tu = raw_text(cell(processed_df, row, inventory_indexes["tk_vat_tu"])) if inventory_indexes["tk_vat_tu"] is not None else ""
        ma_kho = raw_text(cell(processed_df, row, inventory_indexes["ma_kho"])) if inventory_indexes["ma_kho"] is not None else ""
        if not tk_vat_tu or not ma_kho:
            missing_rows.append({
                "row": row + 1,
                "invoice_no": raw_text(cell(processed_df, row, indexes["C"])),
                "ma_vt": raw_text(cell(processed_df, row, indexes["L"])),
                "tk_vat_tu": tk_vat_tu,
                "ma_kho": ma_kho,
            })
    if not total_rows and group_index is None:
        raise ValueError(f"File {label} chưa có dòng hàng hóa FDI đã xử lý hợp lệ.")
    if missing_rows and require_inventory_columns:
        examples = ", ".join(
            f"dòng {item['row']} HD {item['invoice_no']} mã {item['ma_vt']}"
            for item in missing_rows[:5]
        )
        raise ValueError(f"File {label} có {len(missing_rows)}/{total_rows} dòng hàng hóa thiếu TK vật tư hoặc Mã kho. Ví dụ: {examples}.")
    return {
        "valid_rows": total_rows,
        "tk_vat_tu_col": index_to_excel_col(inventory_indexes["tk_vat_tu"]) if inventory_indexes["tk_vat_tu"] is not None else "",
        "ma_kho_col": index_to_excel_col(inventory_indexes["ma_kho"]) if inventory_indexes["ma_kho"] is not None else "",
    }


def fast_tk_doanh_thu(tk_vat_tu):
    tk = raw_text(tk_vat_tu)
    if tk.startswith("152"):
        return "5111"
    if tk.startswith("155"):
        return "5112"
    return "5112"


def fast_loai_vat_tu(tk_vat_tu):
    tk = raw_text(tk_vat_tu)
    return "21" if tk.startswith("152") else "51"


def fast_processed_rows(processed_df, progress_callback=None, label="Đang đọc file đã xử lý", phase="", require_inventory_columns=True):
    indexes = processed_vietmax_source_indexes()
    validate_fast_import_processed_dataframe(processed_df, label, require_inventory_columns=require_inventory_columns)
    inventory_indexes = processed_inventory_column_indexes(processed_df, indexes["L"])
    customer_code_index = processed_customer_code_column_index(processed_df, indexes["L"])
    group_index = processed_group_column_index(processed_df)
    rows = []
    for row in range(len(processed_df)):
        report_loop_progress(progress_callback, row + 1, len(processed_df), label, interval=500)
        if not is_processed_product_data_row(processed_df, row, indexes["L"], indexes["C"], indexes["O"], parse_quantity=False):
            continue
        source_values = fast_source_values_for_row(processed_df, row)
        processing_group = processed_row_group(processed_df, row, group_index)
        tk_vat_tu = up_source_value(processed_df, row, inventory_indexes["tk_vat_tu"]) if inventory_indexes["tk_vat_tu"] is not None else ""
        ma_kho = up_source_value(processed_df, row, inventory_indexes["ma_kho"]) if inventory_indexes["ma_kho"] is not None else ""
        customer_code = up_source_value(processed_df, row, customer_code_index) if customer_code_index is not None else ""
        source_values["AS"] = tk_vat_tu
        source_values["AT"] = ma_kho
        source_values["AU"] = customer_code
        rows.append({
            "row_index": row,
            "source_phase": phase,
            "source_values": source_values,
            "processing_group": processing_group,
            "mau_hd": source_values.get("A", ""),
            "seri": source_values.get("B", ""),
            "invoice_no": source_values.get("C", ""),
            "invoice_date": source_values.get("D", ""),
            "seller_name": source_values.get("F", ""),
            "seller_mst": source_values.get("G", ""),
            "seller_address": source_values.get("H", ""),
            "buyer_name": source_values.get("I", ""),
            "buyer_mst": source_values.get("J", ""),
            "buyer_customer_code": customer_code,
            "seller_customer_code": customer_code,
            "customer_code": customer_code,
            "buyer_address": source_values.get("K", ""),
            "ma_vt": source_values.get("L", ""),
            "product": source_values.get("M", ""),
            "unit": source_values.get("N", ""),
            "qty": source_values.get("O", ""),
            "price": source_values.get("P", ""),
            "line_total": source_values.get("S", ""),
            "amount": source_values.get("V", ""),
            "tax_amount": source_values.get("W", ""),
            "currency": source_values.get("AQ", "") or "VND",
            "exchange_rate": source_values.get("AR", "") or 1,
            "tax_code": source_values.get("R", ""),
            "tax_rate": source_values.get("R", ""),
            "tk_vat_tu": tk_vat_tu,
            "ma_kho": ma_kho,
        })
    return rows


def fast_import_fill_sheet(workbook, headers, rows, title):
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    ws = workbook.create_sheet(title[:31])
    ws.title = title[:31]
    header_fill = PatternFill("solid", fgColor="D9EAD3")
    for col_index, header in enumerate(headers, start=1):
        header_cell = ws.cell(1, col_index, header)
        header_cell.fill = header_fill
        header_cell.font = Font(bold=True)
        header_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row_index, row_values in enumerate(rows, start=2):
        for col_index, value in enumerate(row_values, start=1):
            ws.cell(row_index, col_index, clean_up_cell_value(value))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col_index, header in enumerate(headers, start=1):
        letter = get_column_letter(col_index)
        width = min(max(len(str(header)) + 2, 12), 32)
        ws.column_dimensions[letter].width = width
    return ws


def write_tabular_xls_sheet(workbook, used_names, headers, rows, title):
    import xlwt

    rows = rows if isinstance(rows, list) else []
    headers = headers if isinstance(headers, list) else []
    ensure_xls_dimensions(len(rows) + 1, len(headers), title)
    ws = workbook.add_sheet(safe_excel_sheet_name(title, used_names))
    header_style = xlwt.easyxf("font: bold on; align: horiz center, vert center, wrap on; pattern: pattern solid, fore_colour pale_blue;")
    for col_index, header in enumerate(headers):
        ws.write(0, col_index, coerce_xls_value(header), header_style)
    width_samples = [len(str(header or "")) for header in headers]
    for row_index, row_values in enumerate(rows, start=1):
        for col_index, value in enumerate(row_values):
            ws.write(row_index, col_index, coerce_xls_value(value))
            if col_index < len(width_samples) and row_index <= 200:
                width_samples[col_index] = max(width_samples[col_index], len(str(clean_up_cell_value(value, ""))))
    for col_index, width in enumerate(width_samples):
        ws.col(col_index).width = min(max(width + 2, 10), 36) * 256
    return ws


def write_raw_xls_sheet(workbook, used_names, rows, title):
    rows = rows if isinstance(rows, list) else []
    column_count = max([len(row) for row in rows if isinstance(row, list)] or [0])
    ensure_xls_dimensions(len(rows), column_count, title)
    ws = workbook.add_sheet(safe_excel_sheet_name(title, used_names))
    width_samples = [0] * column_count
    for row_index, row_values in enumerate(rows):
        row_values = row_values if isinstance(row_values, list) else []
        for col_index, value in enumerate(row_values):
            ws.write(row_index, col_index, coerce_xls_value(value))
            if col_index < len(width_samples) and row_index <= 200:
                width_samples[col_index] = max(width_samples[col_index], len(str(clean_up_cell_value(value, ""))))
    for col_index, width in enumerate(width_samples):
        ws.col(col_index).width = min(max(width + 2, 10), 36) * 256
    return ws


def tabular_xls_workbook(sheets):
    import xlwt

    workbook = xlwt.Workbook(encoding="utf-8")
    used_names = set()
    iterable = sheets.items() if hasattr(sheets, "items") else sheets
    for title, (headers, rows) in iterable:
        write_tabular_xls_sheet(workbook, used_names, headers, rows, title)
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


def mixed_xls_workbook(raw_sheets=None, tabular_sheets=None):
    import xlwt

    workbook = xlwt.Workbook(encoding="utf-8")
    used_names = set()
    for title, rows in (raw_sheets or []):
        write_raw_xls_sheet(workbook, used_names, rows, title)
    iterable = tabular_sheets.items() if hasattr(tabular_sheets, "items") else (tabular_sheets or [])
    for title, (headers, rows) in iterable:
        write_tabular_xls_sheet(workbook, used_names, headers, rows, title)
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


def fast_import_workbook(headers, rows, title):
    return tabular_xls_workbook({title: (headers, rows)})


def fast_import_row(headers, values_by_column):
    row = [""] * len(headers)
    for column, value in values_by_column.items():
        column_index = int(column) - 1
        if 0 <= column_index < len(row):
            row[column_index] = value
    return row


FAST_MAPPING_SOURCE_FIELD_BY_COL = {
    "A": "mau_hd",
    "B": "seri",
    "C": "invoice_no",
    "D": "invoice_date",
    "F": "seller_name",
    "G": "seller_mst",
    "H": "seller_address",
    "I": "buyer_name",
    "J": "buyer_mst",
    "K": "buyer_address",
    "L": "ma_vt",
    "M": "product",
    "N": "unit",
    "O": "qty",
    "P": "price",
    "R": "tax_code",
    "S": "line_total",
    "V": "amount",
    "W": "tax_amount",
    "AQ": "currency",
    "AR": "exchange_rate",
    "AS": "tk_vat_tu",
    "AT": "ma_kho",
    "AU": "customer_code",
}


def fast_source_values_for_row(df, row):
    values = {}
    if df is None:
        return values
    if row < 0 or row >= len(df):
        return values
    row_values = df.iloc[row]
    for index, value in enumerate(row_values):
        values[index_to_excel_col(index)] = clean_up_cell_value(value, "")
    return values


def fast_source_number(source_values, letter, default=0):
    value = parse_price(source_values.get(letter))
    return default if value is None else value


def fast_row_tax_info(source_values):
    tax_code, tax_rate = up_ban_ra_tax_values(source_values.get("R"))
    return tax_code, tax_rate or 0


def fast_mapping_source_value(source_row, source_col):
    col = raw_text(source_col).upper()
    if not col:
        return ""
    field = FAST_MAPPING_SOURCE_FIELD_BY_COL.get(col)
    if field and field in source_row:
        return source_row.get(field)
    return (source_row.get("source_values") or {}).get(col, "")


def fast_mapping_template_value(template, source_row):
    text = raw_text(template)
    def replace(match):
        return raw_text(fast_mapping_source_value(source_row, match.group(1)))
    return re.sub(r"\{([A-Za-z]{1,3})\}", replace, text)


def fast_mapping_prefix_before_dot(value, delimiter="."):
    text = raw_text(value)
    delimiter = raw_text(delimiter) or "."
    return text.split(delimiter, 1)[0] if delimiter in text else text


def fast_mapping_customer_code(source_row, current_value=""):
    value = raw_text(current_value)
    if value:
        return value
    for key in ("customer_code", "buyer_customer_code", "seller_customer_code", "buyer_mst", "seller_mst"):
        value = raw_text(source_row.get(key))
        if value:
            return value
    return fast_mapping_prefix_before_dot(source_row.get("ma_vt"))


def fast_default_transform_rules(transform):
    transform = raw_text(transform)
    if transform == "revenue_account_from_inventory_account":
        return FAST_REVENUE_ACCOUNT_RULES
    if transform == "material_type_from_inventory_account":
        return FAST_MATERIAL_TYPE_RULES
    return []


def fast_rule_matches_value(value, rule):
    text = raw_text(value)
    number_value = parse_price(text)
    match_type = raw_text(rule.get("match_type") or "starts_with")
    match_value = raw_text(rule.get("value"))
    if match_type == "default":
        return None
    if match_type == "equals":
        return text == match_value
    if match_type == "not_equals":
        return text != match_value
    if match_type == "contains":
        return bool(match_value and match_value in text)
    if match_type == "not_contains":
        return bool(match_value and match_value not in text)
    if match_type == "starts_with":
        return bool(match_value and text.startswith(match_value))
    if match_type == "ends_with":
        return bool(match_value and text.endswith(match_value))
    if match_type == "blank":
        return not text
    if match_type == "not_blank":
        return bool(text)
    if match_type == "regex" and match_value:
        try:
            return bool(re.search(match_value, text))
        except re.error:
            return False
    if match_type in {"gt", "gte", "lt", "lte"}:
        compare_value = parse_price(match_value)
        if number_value is None or compare_value is None:
            return False
        if match_type == "gt":
            return number_value > compare_value
        if match_type == "gte":
            return number_value >= compare_value
        if match_type == "lt":
            return number_value < compare_value
        if match_type == "lte":
            return number_value <= compare_value
    return False


def fast_rule_output_value(result, source_row=None):
    return fast_mapping_template_value(result, source_row) if isinstance(source_row, dict) else result


def fast_transform_rule_result(value, rules, source_row=None, default_source_col=""):
    default_result = None
    for rule in rules if isinstance(rules, list) else []:
        if not isinstance(rule, dict):
            continue
        match_type = raw_text(rule.get("match_type") or "starts_with")
        result = rule.get("result", "")
        rule_value = value
        if isinstance(source_row, dict) and raw_text(rule.get("source_col") or default_source_col):
            rule_value = fast_mapping_source_value(source_row, rule.get("source_col") or default_source_col)
        if match_type == "default":
            default_result = result
        elif fast_rule_matches_value(rule_value, rule):
            return fast_rule_output_value(result, source_row)
    return fast_rule_output_value(default_result, source_row) if default_result is not None else None


def fast_apply_mapping_transform(transform, value, source_row, transform_rules=None):
    transform = raw_text(transform)
    if not transform:
        return value
    if transform == "mst_or_prefix":
        return fast_mapping_customer_code(source_row, value)
    if transform == "tax_code":
        tax_code, _ = up_ban_ra_tax_values(value)
        return tax_code or raw_text(source_row.get("tax_code")) or value
    if transform == "tax_rate":
        _, tax_rate = up_ban_ra_tax_values(value)
        return tax_rate if tax_rate is not None else source_row.get("tax_rate") or value
    if transform == "revenue_account_from_inventory_account":
        custom_result = fast_transform_rule_result(value, transform_rules if transform_rules is not None else fast_default_transform_rules(transform))
        if custom_result is not None:
            return custom_result
        if transform_rules is not None:
            return value
        return fast_tk_doanh_thu(value)
    if transform == "material_type_from_inventory_account":
        custom_result = fast_transform_rule_result(value, transform_rules if transform_rules is not None else fast_default_transform_rules(transform))
        if custom_result is not None:
            return custom_result
        if transform_rules is not None:
            return value
        return fast_loai_vat_tu(value)
    return value


def fast_mapping_rule_value(rule, source_row):
    source_type = raw_text(rule.get("source_type") or "source_column")
    if source_type == "empty":
        value = ""
    elif source_type in {"constant", "fixed", "text"}:
        value = rule.get("value", "")
    elif source_type == "text_template":
        value = fast_mapping_template_value(rule.get("value", ""), source_row)
    elif source_type == "conditional_rules":
        value = fast_mapping_source_value(source_row, rule.get("source_col"))
        mapped_value = fast_transform_rule_result(value, rule.get("transform_rules"), source_row)
        if mapped_value is not None:
            value = mapped_value
    elif source_type == "if_rules":
        value = fast_mapping_source_value(source_row, rule.get("source_col"))
        condition_value = fast_mapping_source_value(source_row, rule.get("condition_source_col") or rule.get("fallback_source_col") or rule.get("source_col"))
        mapped_value = fast_transform_rule_result(condition_value, rule.get("transform_rules"), source_row, rule.get("condition_source_col") or rule.get("fallback_source_col"))
        if mapped_value is not None:
            value = mapped_value
    else:
        value = fast_mapping_source_value(source_row, rule.get("source_col"))
        if source_type == "if_empty_prefix_before_dot" and not raw_text(value):
            fallback = fast_mapping_source_value(source_row, rule.get("fallback_source_col"))
            value = fast_mapping_prefix_before_dot(fallback, rule.get("fallback_delimiter") or ".")
    return fast_apply_mapping_transform(rule.get("transform"), value, source_row, rule.get("transform_rules"))


def fast_rule_matches_row_phase(rule, source_row):
    rule_phase = raw_text(rule.get("source_phase") or source_row.get("source_phase") or "")
    row_phase = raw_text(source_row.get("source_phase") or "")
    return not rule_phase or rule_phase == "both" or not row_phase or rule_phase == row_phase


def fast_safe_excel_index(letter):
    try:
        return excel_col_to_index(letter)
    except Exception:
        return None


def fast_form_output_headers(form):
    output_columns = form.get("output_columns") if isinstance(form, dict) else []
    mappings = form.get("mappings") if isinstance(form, dict) else []
    max_index = -1
    headers_by_index = {}
    if isinstance(output_columns, list):
        for column in output_columns:
            if not isinstance(column, dict):
                continue
            index = fast_safe_excel_index(column.get("letter"))
            if index is None:
                continue
            max_index = max(max_index, index)
            headers_by_index[index] = raw_text(column.get("header") or column.get("label") or column.get("letter"))
    if isinstance(mappings, list):
        for rule in mappings:
            if not isinstance(rule, dict):
                continue
            index = fast_safe_excel_index(rule.get("target_col"))
            if index is not None:
                max_index = max(max_index, index)
    if max_index < 0:
        max_index = 0
    return [headers_by_index.get(index) or index_to_excel_col(index) for index in range(max_index + 1)]


def fast_apply_form_mapping(form, source_row, headers=None):
    headers = headers or fast_form_output_headers(form)
    row = [""] * len(headers)
    for rule in form.get("mappings") or []:
        if not isinstance(rule, dict) or not fast_rule_matches_row_phase(rule, source_row):
            continue
        target_index = fast_safe_excel_index(rule.get("target_col"))
        if target_index is None:
            continue
        while target_index >= len(row):
            row.append("")
        row[target_index] = fast_mapping_rule_value(rule, source_row)
    return row


def fast_mapping_forms_from_payload(forms):
    if not isinstance(forms, list):
        return []
    result = []
    for form in forms:
        if not isinstance(form, dict) or form.get("enabled") is False:
            continue
        mappings = form.get("mappings")
        if not isinstance(mappings, list) or not mappings:
            continue
        result.append(form)
    return result


def fast_merge_mapping_forms(*form_lists):
    merged = {}
    for form_list in form_lists:
        for form in fast_mapping_forms_from_payload(form_list):
            form_id = raw_text(form.get("id")) or raw_text(form.get("label"))
            if not form_id:
                continue
            merged[form_id] = form
    return list(merged.values())


def fast_row_company_group(source_row, assignments, default_group=FAST_MATERIALS_GROUP_ID):
    row_group = raw_text(source_row.get("processing_group"))
    if row_group:
        return row_group
    assignments = assignments if isinstance(assignments, dict) else {}
    seller_name = source_row.get("seller_name")
    seller_mst = source_row.get("seller_mst")
    buyer_name = source_row.get("buyer_name")
    buyer_mst = source_row.get("buyer_mst")
    candidates = [
        source_row.get("customer_code"),
        source_row.get("buyer_customer_code"),
        source_row.get("seller_customer_code"),
        source_row.get("seller_company_key"),
        source_row.get("buyer_company_key"),
        source_row.get("seller_mst"),
        source_row.get("buyer_mst"),
        source_row.get("seller_name"),
        source_row.get("buyer_name"),
        company_selection_key(seller_name, seller_mst),
        company_selection_key(buyer_name, buyer_mst),
        vietmax_company_identity_key(seller_name, seller_mst),
        vietmax_company_identity_key(buyer_name, buyer_mst),
    ]
    for candidate in candidates:
        key = raw_text(candidate)
        if key and key in assignments:
            return raw_text(assignments.get(key)) or default_group
    return default_group


def fast_rows_for_form(form, purchase_rows, sales_rows, purchase_raw_rows, sales_raw_rows, purchase_assignments=None, sales_assignments=None):
    group_id = raw_text(form.get("group_id")) or FAST_MATERIALS_GROUP_ID
    if group_id == FAST_IGNORED_GROUP_ID:
        return []
    scope = raw_text(form.get("scope") or form.get("input_phase") or "both")
    use_processed = group_id == FAST_MATERIALS_GROUP_ID
    purchase_source = purchase_rows if use_processed else purchase_raw_rows
    sales_source = sales_rows if use_processed else sales_raw_rows
    rows = []
    if scope in {"purchase", "both", ""}:
        rows.extend(row for row in purchase_source if fast_row_company_group(row, purchase_assignments) == group_id)
    if scope in {"sales", "both", ""}:
        rows.extend(row for row in sales_source if fast_row_company_group(row, sales_assignments) == group_id)
    return rows


def fast_unique_mapped_rows(form, source_rows, key_func):
    headers = fast_form_output_headers(form)
    rows_by_key = {}
    for source_row in source_rows:
        mapped = fast_apply_form_mapping(form, source_row, headers)
        key = raw_text(key_func(source_row, mapped))
        if not key or key in rows_by_key:
            continue
        rows_by_key[key] = mapped
    return headers, list(rows_by_key.values())


def fast_import_invoice_sort_key(value):
    text = raw_text(value)
    numeric = parse_price(text)
    if numeric is not None:
        return (0, numeric, text)
    return (1, text.casefold())


def split_fast_invoice_rows(items, sheet_base_name, company_key_fields, company_name_field):
    invoice_companies = {}
    for item in items:
        source = item.get("source") or {}
        invoice_no = raw_text(source.get("invoice_no"))
        company_key = ""
        for field in company_key_fields:
            company_key = raw_text(source.get(field))
            if company_key:
                break
        if not invoice_no or not company_key:
            continue
        invoice_companies.setdefault(invoice_no, [])
        if company_key not in invoice_companies[invoice_no]:
            invoice_companies[invoice_no].append(company_key)

    conflict_invoices = {
        invoice_no
        for invoice_no, companies in invoice_companies.items()
        if len(companies) > 1
    }
    if not conflict_invoices:
        return {sheet_base_name: [item["row"] for item in items]}, []

    max_sheets = max(len(invoice_companies[invoice_no]) for invoice_no in conflict_invoices)
    sheet_items = [[] for _ in range(max_sheets)]
    target_sheet_by_item = {}
    for index, item in enumerate(items):
        source = item.get("source") or {}
        invoice_no = raw_text(source.get("invoice_no"))
        company_key = ""
        for field in company_key_fields:
            company_key = raw_text(source.get(field))
            if company_key:
                break
        if invoice_no in conflict_invoices:
            companies = sorted(invoice_companies[invoice_no], key=lambda value: value.casefold())
            sheet_index = companies.index(company_key) if company_key in companies else 0
        else:
            sheet_index = 0
        sheet_name = f"{sheet_base_name}-{sheet_index + 1}"
        sheet_items[sheet_index].append(item["row"])
        target_sheet_by_item[index] = sheet_name

    report_rows = []
    for invoice_no in sorted(conflict_invoices, key=fast_import_invoice_sort_key):
        conflict_items = [
            (index, item)
            for index, item in enumerate(items)
            if raw_text((item.get("source") or {}).get("invoice_no")) == invoice_no
        ]
        conflict_items.sort(key=lambda pair: (
            raw_text((pair[1].get("source") or {}).get(company_key_fields[0])).casefold(),
            int((pair[1].get("source") or {}).get("row_index") or 0),
        ))
        for index, item in conflict_items:
            source = item.get("source") or {}
            company_code = ""
            for field in company_key_fields:
                company_code = raw_text(source.get(field))
                if company_code:
                    break
            report_rows.append([
                source.get("invoice_no"),
                company_code,
                source.get(company_name_field),
                source.get("invoice_date"),
                int(source.get("row_index") or 0) + 1,
                source.get("ma_vt"),
                source.get("ma_kho"),
                source.get("tk_vat_tu"),
                target_sheet_by_item.get(index, f"{sheet_base_name}-1"),
            ])

    invoice_sheets = {
        f"{sheet_base_name}-{index + 1}": rows
        for index, rows in enumerate(sheet_items)
        if rows or index == 0
    }
    return invoice_sheets, report_rows


def split_fast_purchase_invoice_rows(items):
    return split_fast_invoice_rows(items, "Hoadonmuahang", ["seller_mst", "seller_name"], "seller_name")


def split_fast_sales_invoice_rows(items):
    return split_fast_invoice_rows(items, "Hoadonbanhang", ["buyer_customer_code", "buyer_mst", "buyer_name"], "buyer_name")


def fast_import_sheet_rows(
    purchase_df=None,
    sales_df=None,
    progress_callback=None,
    form_mapping_presets=None,
    purchase_company_group_assignments=None,
    sales_company_group_assignments=None,
    require_inventory_columns=True,
    include_duplicate_report=True,
    use_default_forms_when_empty=True,
):
    purchase_rows = fast_processed_rows(purchase_df, progress_callback, "Đang đọc FDI mua vào", VIETMAX_PHASE_PURCHASE, require_inventory_columns=require_inventory_columns) if purchase_df is not None else []
    sales_rows = fast_processed_rows(sales_df, progress_callback, "Đang đọc FDI bán ra", VIETMAX_PHASE_SALES, require_inventory_columns=require_inventory_columns) if sales_df is not None else []
    purchase_raw_rows = purchase_rows
    sales_raw_rows = sales_rows
    forms = fast_mapping_forms_from_payload(form_mapping_presets)
    if not forms and use_default_forms_when_empty:
        forms = load_system_default_form_mappings()
    total_steps = 4
    done_steps = 0

    def mark(label):
        nonlocal done_steps
        done_steps += 1
        if progress_callback:
            progress_callback(done_steps, total_steps, label)

    result = {}
    duplicate_purchase_report_rows = []
    duplicate_sales_report_rows = []

    purchase_form = next((form for form in forms if raw_text(form.get("builtin_exporter")) == "fast_hoadonmuahang"), None)
    if purchase_form:
        headers = fast_form_output_headers(purchase_form)
        source_rows = fast_rows_for_form(purchase_form, purchase_rows, sales_rows, purchase_raw_rows, sales_raw_rows, purchase_company_group_assignments, sales_company_group_assignments)
        mua_vao_items = [{"row": fast_apply_form_mapping(purchase_form, row, headers), "source": row} for row in source_rows]
        purchase_sheet_rows, duplicate_purchase_report_rows = split_fast_purchase_invoice_rows(mua_vao_items)
        for title, rows in purchase_sheet_rows.items():
            result[title] = (headers, rows)
    if include_duplicate_report:
        result["Bao_cao_trung_so_ct"] = (FAST_DUPLICATE_PURCHASE_REPORT_HEADERS, duplicate_purchase_report_rows)
    mark("Đã dựng dữ liệu Hoadonmuahang")

    sales_form = next((form for form in forms if raw_text(form.get("builtin_exporter")) == "fast_hoadonbanhang"), None)
    if sales_form:
        headers = fast_form_output_headers(sales_form)
        source_rows = fast_rows_for_form(sales_form, purchase_rows, sales_rows, purchase_raw_rows, sales_raw_rows, purchase_company_group_assignments, sales_company_group_assignments)
        ban_ra_items = [{"row": fast_apply_form_mapping(sales_form, row, headers), "source": row} for row in source_rows]
        sales_sheet_rows, duplicate_sales_report_rows = split_fast_sales_invoice_rows(ban_ra_items)
        for title, rows in sales_sheet_rows.items():
            result[title] = (headers, rows)
    mark("Đã dựng dữ liệu Hoadonbanhang")

    material_form = next((form for form in forms if raw_text(form.get("builtin_exporter")) == "fast_dm_vat_tu"), None)
    if material_form:
        source_rows = fast_rows_for_form(material_form, purchase_rows, sales_rows, purchase_raw_rows, sales_raw_rows, purchase_company_group_assignments, sales_company_group_assignments)
        headers, dm_vat_tu_rows = fast_unique_mapped_rows(material_form, source_rows, lambda source, mapped: mapped[0] if mapped else source.get("ma_vt"))
        result[raw_text(material_form.get("sheet") or material_form.get("label") or "DMvat_tu")] = (headers, dm_vat_tu_rows)
    mark("Đã dựng dữ liệu DMvat tu")

    customer_form = next((form for form in forms if raw_text(form.get("builtin_exporter")) == "fast_dm_khach_hang"), None)
    if customer_form:
        source_rows = fast_rows_for_form(customer_form, purchase_rows, sales_rows, purchase_raw_rows, sales_raw_rows, purchase_company_group_assignments, sales_company_group_assignments)
        headers, dm_khach_hang_rows = fast_unique_mapped_rows(customer_form, source_rows, lambda source, mapped: mapped[0] if mapped else fast_mapping_customer_code(source))
        result[raw_text(customer_form.get("sheet") or customer_form.get("label") or "DMkhachhang")] = (headers, dm_khach_hang_rows)
    mark("Đã dựng dữ liệu DMkhachhang")

    if include_duplicate_report:
        result["Bao_cao_trung_so_ct"] = (FAST_DUPLICATE_PURCHASE_REPORT_HEADERS, duplicate_purchase_report_rows + duplicate_sales_report_rows)

    builtin_ids = {"fast_hoadonmuahang", "fast_hoadonbanhang", "fast_dm_vat_tu", "fast_dm_khach_hang", "fast_duplicate_report"}
    for form in forms:
        if raw_text(form.get("builtin_exporter")) in builtin_ids:
            continue
        source_rows = fast_rows_for_form(form, purchase_rows, sales_rows, purchase_raw_rows, sales_raw_rows, purchase_company_group_assignments, sales_company_group_assignments)
        headers = fast_form_output_headers(form)
        rows = [fast_apply_form_mapping(form, row, headers) for row in source_rows]
        title = raw_text(form.get("sheet") or form.get("label") or form.get("id") or "CustomForm")
        result[title] = (headers, rows)
    return result


def fast_import_multi_sheet_workbook(purchase_df=None, sales_df=None, progress_callback=None, **mapping_options):
    sheets = fast_import_sheet_rows(purchase_df, sales_df, progress_callback, **mapping_options)
    return tabular_xls_workbook(sheets)


def fast_import_package_zip(purchase_df=None, sales_df=None, progress_callback=None, **mapping_options):
    sheets = fast_import_sheet_rows(purchase_df, sales_df, progress_callback, **mapping_options)

    stream = BytesIO()
    with zipfile.ZipFile(stream, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for title, (headers, rows) in sheets.items():
            archive.writestr(f"{title}.xls", fast_import_workbook(headers, rows, title).getvalue())
    stream.seek(0)
    return stream


def up_mua_vao_output_path(formatted_path):
    return formatted_path.with_name(f"{formatted_path.stem}_UP_mua_vao.xls")


def up_ban_ra_output_path(formatted_path):
    return formatted_path.with_name(f"{formatted_path.stem}_UP_ban_ra.xls")


def process_zip_stream(*paths):
    stream = BytesIO()
    with zipfile.ZipFile(stream, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for path in paths:
            archive.write(path, path.name)
    stream.seek(0)
    return stream


def preview_data(df):
    preview_indexes = list(range(df.shape[1]))
    view = df.iloc[:8, preview_indexes].copy()
    view.columns = [index_to_excel_col(i) for i in preview_indexes]
    view = view.fillna("")
    return [{col: str(val) if val != "" else "" for col, val in row.items()} for _, row in view.iterrows()]


def invoice_status_options(df, invoice_status_col=DEFAULT_INVOICE_STATUS_COL, skip_statuses=None, progress_callback=None):
    status_col = str(invoice_status_col or "").strip().upper()
    if not status_col:
        return []
    status_index = excel_col_to_index(status_col)
    if df.shape[1] <= status_index:
        return []
    counts = Counter()
    display_values = {}
    for i in range(len(df)):
        report_loop_progress(progress_callback, i + 1, len(df), "Đang đọc trạng thái hóa đơn")
        value = raw_text(cell(df, i, status_index))
        if not value:
            continue
        normalized = normalized_invoice_status(value)
        if normalized == "trang thai hoa don":
            continue
        counts[normalized] += 1
        display_values.setdefault(normalized, value)
    selected = normalized_invoice_status_set(skip_statuses)
    return [
        {
            "value": display_values[key],
            "count": count,
            "skip": key in selected,
        }
        for key, count in counts.most_common()
    ]


def safe_excel_sheet_name(value, used_names):
    name = re.sub(r"[:\\/?*\[\]]", " ", str(value or "Sheet")).strip() or "Sheet"
    name = name[:31]
    base = name
    counter = 2
    while name in used_names:
        suffix = f" {counter}"
        name = f"{base[:31 - len(suffix)]}{suffix}"
        counter += 1
    used_names.add(name)
    return name


def safe_excel_filename(value):
    stem = Path(str(value or "bao_cao_ban_hang.xls")).stem or "bao_cao_ban_hang"
    stem = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", stem).strip(" .") or "bao_cao_ban_hang"
    return f"{stem}.xls"


def coerce_excel_value(value):
    if value is None:
        return ""
    if isinstance(value, (int, float, np.integer, np.floating)):
        return float(value) if isinstance(value, (float, np.floating)) else int(value)
    return str(value)


def make_excel_workbook(sheets):
    tabular_sheets = []
    for sheet in sheets:
        rows = sheet.get("rows") if isinstance(sheet, dict) else []
        rows = rows if isinstance(rows, list) else []
        title = sheet.get("name") if isinstance(sheet, dict) else "Sheet"
        headers = sheet.get("headers") if isinstance(sheet, dict) and isinstance(sheet.get("headers"), list) else []
        headers = [str(header) for header in headers if str(header).strip()]
        if not headers:
            headers = list(rows[0].keys()) if rows and isinstance(rows[0], dict) else []
        if not headers:
            tabular_sheets.append((title, (["Không có dữ liệu"], [])))
            continue
        table_rows = []
        for row in rows:
            table_rows.append([coerce_excel_value(row.get(header)) if isinstance(row, dict) else "" for header in headers])
        tabular_sheets.append((title, (headers, table_rows)))
    if not tabular_sheets:
        tabular_sheets.append(("Bao cao", (["Không có dữ liệu"], [])))
    return tabular_xls_workbook(tabular_sheets)


@app.route("/api/config", methods=["GET"])
def get_config():
    return jsonify(load_config())


@app.route("/api/config", methods=["POST"])
def set_config():
    return jsonify(save_config(request.get_json() or {}))


@app.route("/api/config/profile/<profile>", methods=["POST"])
def import_profile_config(profile):
    target = profile_key(profile)
    incoming = normalize_config(request.get_json() or {})
    source = target if target in incoming.get("profiles", {}) else incoming.get("selected_profile")
    source = source if source in PROFILE_LABELS else target
    current = load_config()
    current["selected_profile"] = target
    current["columns"].update(incoming.get("columns") or {})
    current["profiles"][target] = incoming["profiles"].get(source, empty_profile_config(target))
    return jsonify(save_config(current))


@app.route("/api/mapping", methods=["POST"])
def mapping():
    f = request.files.get("file")
    if not f:
        return jsonify({"error": "No file uploaded."}), 400
    original = f.filename
    ext = Path(original).suffix.lower()
    if ext not in [".xls", ".xlsx", ".xlsm"]:
        return jsonify({"error": "Please upload .xls, .xlsx or .xlsm only."}), 400
    saved = f"{uuid.uuid4().hex}{ext}"
    path = UPLOAD_DIR / saved
    f.save(path)
    try:
        _, df = read_workbook(path)
    except Exception as e:
        return jsonify({"error": str(e)}), 400
    cols = []
    for idx in range(df.shape[1]):
        letter = index_to_excel_col(idx)
        samples = []
        for r in range(min(6, len(df))):
            v = df.iat[r, idx]
            if not pd.isna(v) and str(v).strip():
                samples.append(str(v).strip())
        label = letter + ((" - " + " | ".join(samples[:2])[:45]) if samples else "")
        cols.append({"letter": letter, "label": label})
    return jsonify({
        "original_name": original,
        "saved_name": saved,
        "columns": cols,
        "preview": preview_data(df),
        "invoice_statuses": invoice_status_options(df, DEFAULT_INVOICE_STATUS_COL, DEFAULT_INVOICE_STATUS_SKIP_VALUES),
    })


@app.route("/api/invoice_statuses", methods=["POST"])
def invoice_statuses():
    data = request.get_json() or {}
    saved = data.get("saved_name", "")
    path = UPLOAD_DIR / saved
    if not path.exists():
        return jsonify({"error": "Uploaded file was not found. Please upload again."}), 400
    try:
        _, df = read_workbook(path)
        statuses = invoice_status_options(
            df,
            str(data.get("invoice_status_col", DEFAULT_INVOICE_STATUS_COL) or "").upper(),
            data.get("invoice_status_skip_values"),
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 400
    return jsonify({"invoice_statuses": statuses})


@app.route("/api/check", methods=["POST"])
def check():
    data = request.get_json() or {}
    saved = data.get("saved_name", "")
    profile = data.get("profile", "son_phuong")
    cfg = load_config()["profiles"].get(profile, empty_profile_config())
    path = UPLOAD_DIR / saved
    if not path.exists():
        return jsonify({"error": "Uploaded file was not found. Please upload again."}), 400
    try:
        result = analyze(
            path,
            data.get("company_col", "F").upper(),
            data.get("mst_col", "G").upper(),
            data.get("address_col", "H").upper(),
            data.get("product_col", "N").upper(),
            data.get("qty_col", "P").upper(),
            data.get("price_col", "R").upper(),
            cfg,
            str(data.get("invoice_status_col", DEFAULT_INVOICE_STATUS_COL) or "").upper(),
            data.get("invoice_status_skip_values"),
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 400
    result.update({
        "original_name": data.get("original_name", "output.xlsx"),
        "saved_name": saved,
    })
    return jsonify(result)


@app.route("/api/export_price_report", methods=["POST"])
def export_price_report():
    data = request.get_json() or {}
    sheets = data.get("sheets") or []
    if not isinstance(sheets, list) or not sheets:
        return jsonify({"error": "Không có dữ liệu để xuất Excel."}), 400
    try:
        workbook = make_excel_workbook(sheets)
        filename = safe_excel_filename(data.get("filename") or "bao_cao_ban_hang.xls")
    except Exception as e:
        return jsonify({"error": str(e)}), 400
    return send_file(
        workbook,
        as_attachment=True,
        download_name=filename,
        mimetype=XLS_MEDIA_TYPE,
    )


@app.route("/api/process", methods=["POST"])
def process():
    data = request.get_json() or {}
    saved = data.get("saved_name", "")
    original = data.get("original_name", "output.xlsx")
    path = UPLOAD_DIR / saved
    if not path.exists():
        return jsonify({"error": "Uploaded file was not found. Please upload again."}), 400
    try:
        out = resolve_output_path(original, "")
        process_workbook(path, out, data)
        cfg = load_config()
        profile = profile_key(data.get("profile", cfg.get("selected_profile", "son_phuong")))
        uses_price_rules = profile == "cao_thanh"
        cfg["selected_profile"] = profile
        cfg["columns"].update({k: data.get(k, v) for k, v in cfg["columns"].items()})
        cfg["profiles"].setdefault(profile, empty_profile_config(profile))
        old_profile = cfg["profiles"].get(profile) or empty_profile_config(profile)
        merged_price_ranges = merge_price_ranges(old_profile.get("price_range_rules"), data.get("price_range_rules", {})) if uses_price_rules else {}
        cfg["profiles"][profile].update({
            "prefixes": data.get("prefixes", {}),
            "selected_products": data.get("skipped_products_map", {}),
            "removed_companies": data.get("removed_companies", old_profile.get("removed_companies", {})),
            "word_rules": data.get("word_rules", {}),
            "first_word_rules": data.get("first_word_rules", {}),
            "repeated_phrase_removals": normalize_phrase_list(data.get("repeated_phrase_removals", old_profile.get("repeated_phrase_removals", []))),
            "price_group_rules": data.get("price_group_rules", {}) if uses_price_rules else {},
            "price_range_rules": merged_price_ranges,
            "price_adjust_all_percent": float(data.get("price_adjust_all_percent") or 0) if uses_price_rules else 0,
            "manual_code_overrides": data.get("manual_code_overrides", {}),
            "product_code_replacements": normalize_product_code_replacements(data.get("product_code_replacements") or old_profile.get("product_code_replacements") or {}),
            "inventory_pairs": normalize_inventory_pairs(data.get("inventory_pairs") or old_profile.get("inventory_pairs") or []),
            "use_default_inventory_pair": bool(data.get("use_default_inventory_pair", old_profile.get("use_default_inventory_pair", False))),
            "default_inventory_pair_id": str(data.get("default_inventory_pair_id", old_profile.get("default_inventory_pair_id", "")) or "").strip(),
            "inventory_pair_rules": normalize_inventory_pair_rules(data.get("inventory_pair_rules") or old_profile.get("inventory_pair_rules") or []),
            "inventory_allocation_config": data.get("inventory_allocation_config") if isinstance(data.get("inventory_allocation_config"), dict) else old_profile.get("inventory_allocation_config", {}),
            "include_company_prefix": data.get("include_company_prefix") is not False,
        })
        save_config(cfg)
    except Exception as e:
        return jsonify({"error": str(e)}), 400
    return send_file(
        out,
        as_attachment=True,
        download_name=out.name,
        mimetype=XLS_MEDIA_TYPE,
    )


@app.route("/", defaults={"path": ""})
@app.route("/<path:path>")
def serve_frontend(path):
    if STATIC_DIR.exists():
        file_path = STATIC_DIR / path
        if path and file_path.exists() and file_path.is_file():
            return send_from_directory(str(STATIC_DIR), path)
        return send_from_directory(str(STATIC_DIR), "index.html")
    return jsonify({"status": "ok"})


def open_browser():
    for _ in range(60):
        try:
            with urlopen("http://127.0.0.1:5000/api/config", timeout=0.5):
                break
        except Exception:
            time.sleep(0.5)
    webbrowser.open("http://localhost:5000")


def run_tray_icon():
    try:
        import pystray
        from PIL import Image
    except Exception:
        return
    if not ICON_PATH.exists():
        return

    def open_app(icon, item):
        open_browser()

    def quit_app(icon, item):
        icon.stop()
        os._exit(0)

    image = Image.open(ICON_PATH)
    menu = pystray.Menu(
        pystray.MenuItem("Mở Product Code Formatter", open_app),
        pystray.MenuItem("Thoát", quit_app),
    )
    pystray.Icon("ProductCodeFormatter", image, "Product Code Formatter", menu).run()


def close_existing_instances():
    current_pid = str(os.getpid())
    try:
        output = subprocess.check_output(
            ["netstat", "-ano", "-p", "tcp"],
            text=True,
            stderr=subprocess.DEVNULL,
        )
    except Exception:
        return
    for line in output.splitlines():
        parts = line.split()
        if len(parts) < 5 or parts[0].lower() != "tcp":
            continue
        local_address = parts[1]
        state = parts[3].lower()
        pid = parts[4]
        if not local_address.endswith(":5000") or state != "listening":
            continue
        if pid and pid != current_pid:
            subprocess.run(["taskkill", "/F", "/PID", pid, "/T"], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)


def main():
    if getattr(sys, "frozen", False):
        close_existing_instances()
        threading.Thread(target=run_tray_icon, daemon=True).start()
        threading.Timer(1.0, open_browser).start()
    app.run(host="127.0.0.1", port=5000, debug=False, use_reloader=False)


if __name__ == "__main__":
    main()
