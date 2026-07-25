import unittest
import time
from io import BytesIO
from pathlib import Path
import sys

from openpyxl import Workbook, load_workbook

ROOT_DIR = Path(__file__).resolve().parents[1]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

from inventory_allocation_app import app as inventory_app
from inventory_allocation_app.app import DEFAULT_BAREM_MAP, allocate_stock, build_inventory_ledger, build_sales_report_rows_from_ledger, clean_mapping, create_output_workbook, find_sale_only_codes, generic_steel_sale_type, parse_barem_file, preview_workbook, read_lines, resolve_barem_weight, sales_summary_rows_for_ui, steel_coating, steel_kind, steel_profile_key


def line(kind, code, quantity, price=None, row_number=3, invoice_no="", invoice_date="", **extra):
    base = code.rsplit(".", 1)[0] if code.rsplit(".", 1)[-1].isdigit() else code
    suffix = int(code.rsplit(".", 1)[-1]) if code.rsplit(".", 1)[-1].isdigit() else None
    result = {
        "kind": kind,
        "row_number": row_number,
        "variant_code": code,
        "base_code": base,
        "suffix": suffix,
        "invoice_no": invoice_no,
        "invoice_date": invoice_date,
        "product_name": "Hang A",
        "quantity": quantity,
        "unit_price": price,
    }
    result.update(extra)
    return result


def invoice_file(code, quantity, price):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Invoices"
    sheet.cell(2, 12, "Ma VT")
    sheet.cell(2, 13, "Ten hang")
    sheet.cell(2, 15, "So luong")
    sheet.cell(2, 16, "Don gia")
    sheet.cell(2, 3, "So hoa don")
    sheet.cell(2, 4, "Ngay hoa don")
    sheet.cell(3, 12, code)
    sheet.cell(3, 13, "Hang A")
    sheet.cell(3, 15, quantity)
    sheet.cell(3, 16, price)
    sheet.cell(3, 3, "HD1")
    sheet.cell(3, 4, "02/01/2025")
    content = BytesIO()
    workbook.save(content)
    content.seek(0)
    return content


def invoice_file_with_name(code, product_name, quantity, price):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Invoices"
    sheet.cell(2, 12, "Ma VT")
    sheet.cell(2, 13, "Ten hang")
    sheet.cell(2, 15, "So luong")
    sheet.cell(2, 16, "Don gia")
    sheet.cell(2, 3, "So hoa don")
    sheet.cell(2, 4, "Ngay hoa don")
    sheet.cell(3, 12, code)
    sheet.cell(3, 13, product_name)
    sheet.cell(3, 15, quantity)
    sheet.cell(3, 16, price)
    sheet.cell(3, 3, "HD1")
    sheet.cell(3, 4, "02/01/2026")
    content = BytesIO()
    workbook.save(content)
    content.seek(0)
    return content


def up_sales_file(code, quantity, price):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "UP Bán ra"
    sheet.cell(1, 5, "Số chứng từ\n(so_ct)")
    sheet.cell(1, 6, "Ngày chứng từ\n(ngay_ct)")
    sheet.cell(1, 8, "Số lượng:Q\n(so_luong)")
    sheet.cell(1, 15, "Giá bán:P0\n(gia2)")
    sheet.cell(1, 23, "Thuế suất\n(thue_suat)")
    sheet.cell(1, 25, "Tiền thuế:N0\n(tien_thue)")
    sheet.cell(1, 33, "Mã vật tư\n(ma_vt)")
    sheet.cell(1, 34, "Diễn giải\n(dien_giai)")
    sheet.cell(2, 5, "HDSP1")
    sheet.cell(2, 6, "02/01/2026")
    sheet.cell(2, 8, quantity)
    sheet.cell(2, 15, price)
    sheet.cell(2, 23, 10)
    sheet.cell(2, 25, quantity * price * 0.1)
    sheet.cell(2, 33, code)
    sheet.cell(2, 34, "Xuất bán hàng")
    content = BytesIO()
    workbook.save(content)
    content.seek(0)
    return content


class AllocationTests(unittest.TestCase):
    def test_deferred_analysis_keeps_financial_fields_for_report_hydration(self):
        purchase = invoice_file("ITEM", 10, 40).getvalue()
        sales = invoice_file("ITEM", 3, 100).getvalue()

        result = inventory_app.analysis_payload(
            purchase,
            sales,
            None,
            {},
            {"company_profile": "son_phuong", "allow_negative_export": True},
            "sales.xlsx",
            defer_workbook=True,
            result_job_id="deferred-report-test",
        )

        self.assertNotIn("line_amount", result["allocations"][0])
        self.assertEqual(result["_allocations"][0]["line_amount"], 300)
        ledger = build_inventory_ledger(
            result["_opening_lines"],
            result["_purchase_lines"],
            result["_allocations"],
            sales_lines=result["_sales_lines"],
            company_profile="son_phuong",
        )
        report_rows = build_sales_report_rows_from_ledger(ledger)
        self.assertEqual(sum(row["sale_amount"] for row in report_rows), 300)

    def test_split_sales_ledger_preserves_source_sale_and_tax_amounts(self):
        purchase = [
            line("purchase", "PAPER", 5, 40, invoice_no="M1", invoice_date="01/01/2026", warehouse_code="KVT", warehouse_account="152")
        ]
        sales = [
            line(
                "sales",
                "PAPER",
                10,
                100,
                invoice_no="B1",
                invoice_date="02/01/2026",
                line_amount=1000.3,
                tax_amount=80.07,
            )
        ]
        allocations, _, _, _ = allocate_stock([], purchase, sales, {"allow_negative_export": True})

        ledger = build_inventory_ledger([], purchase, allocations, sales_lines=sales)
        sale_rows = [
            row
            for warehouse in ledger["warehouses"]
            for group in warehouse["groups"]
            for row in group["rows"]
            if row["type"] == "sale"
        ]

        self.assertEqual(len(sale_rows), 2)
        self.assertAlmostEqual(sum(row["sale_amount"] for row in sale_rows), 1000.3)
        self.assertAlmostEqual(sum(row["tax_amount"] for row in sale_rows), 80.07)

    def test_sale_warehouse_without_inbound_keeps_zero_cost_and_zero_margin(self):
        purchase = [line("purchase", "PAPER", 10, 50, invoice_no="M1", invoice_date="01/01/2026", warehouse_code="KVT", warehouse_account="152")]
        sales = [line("sales", "PAPER", 3, 120, invoice_no="B1", invoice_date="02/01/2026", party_name="Khach A", warehouse_code="KGCI", warehouse_account="1552")]
        allocations, _, _, _ = allocate_stock([], purchase, sales, {"allow_negative_export": True})

        ledger = build_inventory_ledger([], purchase, allocations, sales_lines=sales)
        report_rows = build_sales_report_rows_from_ledger(ledger)
        kgci_rows = [row for row in report_rows if row["warehouse_code"] == "KGCI"]

        self.assertEqual(len(kgci_rows), 1)
        self.assertEqual(kgci_rows[0]["cost_amount"], 0)
        self.assertEqual(kgci_rows[0]["profit_amount"], 0)
        self.assertTrue(kgci_rows[0]["cost_missing"])
        summary = sales_summary_rows_for_ui(report_rows)[0]
        self.assertEqual(summary["margin_percent"], 0)
        self.assertEqual(summary["profit_amount"], 0)

    def test_son_phuong_detects_pipe_and_box_products(self):
        self.assertEqual(steel_kind("Ống tôn mạ CN 30x60x1.4x6.0"), "box")
        self.assertEqual(steel_kind("Thép hộp mạ kẽm 20x20x1.1x6000"), "box")
        self.assertEqual(steel_kind("Ống tôn mạ vuông 40x1.4x6.0"), "box")
        self.assertEqual(steel_kind("Ống thép tròn Φ26.65x1.4"), "pipe")
        self.assertEqual(steel_kind("Ống thép F59.9x1.4"), "pipe")
        self.assertEqual(steel_kind("Ống thép D21.2x1.2"), "pipe")
        self.assertEqual(steel_coating("Thép hộp đen 20x20x1.1"), "black")
        self.assertEqual(steel_coating("Thép hộp 20x20x1.1"), "unknown")
        self.assertEqual(steel_coating("Thép hộp mạ kẽm 20x20x1.1"), "galvanized")
        self.assertEqual(steel_coating("Ong ton ma vuong 30x1.8x6.0"), "galvanized")
        self.assertEqual(steel_profile_key("Thép hộp 20x20x1.1"), "box|unknown|20x20x1.1")
        self.assertIsNone(inventory_app.explicit_steel_coating("Thép ống hộp các loại"))
        self.assertEqual(steel_kind("Thép cuộn cán nóng các loại", "ONGTRONGMA"), "unknown")
        self.assertEqual(generic_steel_sale_type("Thép ống hộp các loại"), "pipe_box")
        self.assertEqual(generic_steel_sale_type("Thép hộp các loại"), "box")
        self.assertEqual(generic_steel_sale_type("Thép ống các loại"), "pipe")

    def test_son_phuong_generic_sale_uses_allowed_steel_types_by_low_profit_then_stock(self):
        purchase = [
            line("purchase", "BOXA", 4, 90, product_name="Thép hộp 20x20x1.1x6000"),
            line("purchase", "BOXB", 12, 90, product_name="Thép hộp 30x60x1.4x6000"),
            line("purchase", "PIPEA", 20, 95, product_name="Ống thép tròn Φ26.65x1.4"),
        ]
        sales = [line("sales", "GENBOX", 10, 100, product_name="Thép hộp các loại")]

        allocations, _, summary, _ = allocate_stock(
            [],
            purchase,
            sales,
            {
                "company_profile": "son_phuong",
                "max_loss_percent": 10,
                "max_profit_percent": 25,
                "son_phuong_split_counts": {"box": 2, "pipe": 2, "pipe_box": 2},
            },
        )

        used_codes = [item["variant_code"] for item in allocations[0]["used"]]
        self.assertEqual(used_codes, ["BOXB", "BOXA", "BOXB"])
        self.assertEqual(summary["material_quantity"], 10)
        self.assertEqual(summary["finished_quantity"], 0)
        self.assertNotIn("PIPEA", used_codes)

    def test_son_phuong_priority_uses_low_profit_before_loss(self):
        purchase = [
            line("purchase", "PROFIT_LOW", 100, 98, product_name="Thep hop 20x20x1.1"),
            line("purchase", "PROFIT_HIGH", 100, 80, product_name="Thep hop 30x60x1.4"),
            line("purchase", "LOSS_LOW", 100, 102, product_name="Thep hop 25x50x1.4"),
            line("purchase", "LOSS_HIGH", 100, 130, product_name="Thep hop 40x80x1.4"),
        ]
        sales = [line("sales", "GENBOX", 100, 100, product_name="Thep hop cac loai")]

        allocations, _, _, _ = allocate_stock(
            [],
            purchase,
            sales,
            {"company_profile": "son_phuong", "barem_tolerance_percent": 100},
            barem_map={
                "by_code": {"PROFITLOW": 10, "PROFITHIGH": 10, "LOSSLOW": 10, "LOSSHIGH": 10},
                "by_profile": {},
            },
        )

        used_codes = [item["variant_code"] for item in allocations[0]["used"]]
        used_codes = list(dict.fromkeys(used_codes))
        self.assertEqual(used_codes, ["PROFIT_LOW", "PROFIT_HIGH"])

    def test_son_phuong_shape_steel_matches_by_product_family_across_codes(self):
        purchase = [line("purchase", "TCC.THV50", 100, 90, product_name="Thép hình V50")]
        sales = [line("sales", "OTHER.CODE", 25, 100, product_name="Thép V 50")]

        allocations, _, summary, _ = allocate_stock(
            [],
            purchase,
            sales,
            {"company_profile": "son_phuong"},
        )

        self.assertEqual(allocations[0]["used"][0]["variant_code"], "TCC.THV50")
        self.assertEqual(summary["material_quantity"], 25)
        self.assertEqual(summary["negative_export_quantity"], 0)

    def test_son_phuong_c100_uses_matching_purchase_then_ktp_for_shortage(self):
        purchase = [line(
            "purchase", "PURCHASE.C100", 5, 90,
            product_name="Thep C 100",
            invoice_date="01/01/2025",
            invoice_date_iso="2025-01-01",
        )]
        sales = [line(
            "sales", "SALE.C100", 8, 100,
            product_name="Thep C100",
            invoice_date="02/01/2025",
            invoice_date_iso="2025-01-02",
        )]

        allocations, _, summary, _ = allocate_stock([], purchase, sales, {"company_profile": "son_phuong"})

        self.assertEqual(allocations[0]["allocation_role"], "finished_goods")
        self.assertEqual(allocations[0]["used"][0]["variant_code"], "PURCHASE.C100")
        self.assertEqual(allocations[0]["material_quantity"], 5)
        self.assertEqual(allocations[0]["finished_quantity"], 3)
        self.assertEqual(allocations[0]["remainder_warehouse_code"], "KTP")
        self.assertEqual(summary["negative_export_quantity"], 0)

    def test_son_phuong_v40_uses_matching_purchase_across_name_styles(self):
        purchase = [line(
            "purchase", "PURCHASE.V40", 10, 90,
            product_name="Thep hinh V40",
            invoice_date="01/01/2025",
            invoice_date_iso="2025-01-01",
        )]
        sales = [line(
            "sales", "SALE.V40", 6, 100,
            product_name="Thep V 40",
            invoice_date="02/01/2025",
            invoice_date_iso="2025-01-02",
        )]

        allocations, _, summary, _ = allocate_stock([], purchase, sales, {"company_profile": "son_phuong"})

        self.assertEqual(allocations[0]["allocation_role"], "finished_goods")
        self.assertEqual(allocations[0]["used"][0]["variant_code"], "PURCHASE.V40")
        self.assertEqual(allocations[0]["material_quantity"], 6)
        self.assertEqual(allocations[0]["finished_quantity"], 0)
        self.assertEqual(summary["negative_export_quantity"], 0)

    def test_son_phuong_shape_steel_does_not_consume_another_size(self):
        purchase = [line(
            "purchase", "PURCHASE.V50", 10, 90,
            product_name="Thep hinh V50",
            invoice_date="01/01/2025",
            invoice_date_iso="2025-01-01",
        )]
        sales = [line(
            "sales", "SALE.V40", 6, 100,
            product_name="Thep V40",
            invoice_date="02/01/2025",
            invoice_date_iso="2025-01-02",
        )]

        allocations, _, summary, _ = allocate_stock([], purchase, sales, {"company_profile": "son_phuong"})

        self.assertEqual(allocations[0]["allocation_role"], "finished_goods")
        self.assertEqual(allocations[0]["used"], [])
        self.assertEqual(allocations[0]["material_quantity"], 0)
        self.assertEqual(allocations[0]["finished_quantity"], 6)
        self.assertEqual(summary["negative_export_quantity"], 0)

    def test_son_phuong_all_cold_sheet_products_are_always_khhvt(self):
        sales = [
            line(
                "sales", "TON-LANH-MA-MAU", 12, 100,
                product_name="Ton lanh ma mau",
                invoice_date="02/01/2025",
                invoice_date_iso="2025-01-02",
            ),
            line(
                "sales", "TONLANHMCL", 15, 100,
                product_name="Tôn lạnh màu các loại",
                invoice_date="02/01/2025",
                invoice_date_iso="2025-01-02",
            ),
            line(
                "sales", "TON-LANH", 18, 100,
                product_name="Tôn lạnh",
                invoice_date="02/01/2025",
                invoice_date_iso="2025-01-02",
            ),
        ]

        allocations, _, summary, warnings = allocate_stock([], [], sales, {"company_profile": "son_phuong"})
        ledger = build_inventory_ledger([], [], allocations, company_profile="son_phuong")

        self.assertEqual([row["allocation_role"] for row in allocations], ["materials", "materials", "materials"])
        self.assertTrue(all(not row["negative_warning"] for row in allocations))
        self.assertEqual([row["material_quantity"] for row in allocations], [0, 0, 0])
        self.assertEqual([row["unresolved_material_quantity"] for row in allocations], [12, 15, 18])
        self.assertEqual(summary["unresolved_material_quantity"], 45)
        self.assertTrue(all(row["finished_quantity"] == 0 for row in allocations))
    def test_son_phuong_generic_sale_increases_type_count_only_until_quantity_is_covered(self):
        purchase = [
            line("purchase", f"BOX{i}", 30, 95 + i, product_name=f"Thep hop {20 + i}x{20 + i}x1.1")
            for i in range(1, 9)
        ]
        sales = [line("sales", "GENBOX", 140, 100, product_name="Thep hop cac loai")]
        barem_map = {"by_code": {f"BOX{i}": 10 for i in range(1, 9)}, "by_profile": {}}

        allocations, _, summary, _ = allocate_stock(
            [],
            purchase,
            sales,
            {"company_profile": "son_phuong", "barem_tolerance_percent": 100},
            barem_map=barem_map,
        )

        used_codes = list(dict.fromkeys(item["variant_code"] for item in allocations[0]["used"]))
        self.assertEqual(len(used_codes), 5)
        self.assertEqual(sum(item["quantity"] for item in allocations[0]["used"]), 140)
        self.assertEqual(summary["negative_export_quantity"], 0)

    def test_son_phuong_generic_plan_groups_same_code_across_purchase_invoices(self):
        purchase = [
            line("purchase", "BOXA", 30, 99, row_number=3, invoice_no="M1", invoice_date="01/01/2026", invoice_date_iso="2026-01-01", product_name="Thep hop 20x20x1.1"),
            line("purchase", "BOXA", 30, 99, row_number=4, invoice_no="M2", invoice_date="01/01/2026", invoice_date_iso="2026-01-01", product_name="Thep hop 20x20x1.1"),
            line("purchase", "BOXB", 60, 98, row_number=5, invoice_no="M3", invoice_date="01/01/2026", invoice_date_iso="2026-01-01", product_name="Thep hop 30x60x1.4"),
        ]
        sales = [
            line("sales", "GENBOX", 100, 110, invoice_no="B1", invoice_date="02/01/2026", invoice_date_iso="2026-01-02", product_name="Thep hop cac loai"),
        ]

        allocations, _, summary, _ = allocate_stock(
            [],
            purchase,
            sales,
            {"company_profile": "son_phuong", "generic_min_type_count": 2},
            barem_map={"BOXA": 10, "BOXB": 10},
        )

        used = allocations[0]["used"]
        used_codes = list(dict.fromkeys(item["variant_code"] for item in used))
        box_a_invoices = {item["invoice_no"] for item in used if item["variant_code"] == "BOXA"}
        self.assertEqual(used_codes, ["BOXA", "BOXB"])
        self.assertEqual(box_a_invoices, {"M1", "M2"})
        self.assertEqual(sum(item["quantity"] for item in used), 100)
        self.assertEqual(summary["negative_export_quantity"], 0)

    def test_son_phuong_generic_plan_groups_supplier_prefixes_by_technical_profile(self):
        purchase = [
            line("purchase", "AA.HOP.MK.20X20X1.1", 30, 99, row_number=3, invoice_no="M1", invoice_date="01/01/2026", invoice_date_iso="2026-01-01", product_name="Thep hop 20x20x1.1"),
            line("purchase", "BB.HOP.MK.20X20X1.1", 30, 98, row_number=4, invoice_no="M2", invoice_date="01/01/2026", invoice_date_iso="2026-01-01", product_name="Thep hop 20x20x1.1"),
            line("purchase", "CC.HOP.MK.30X60X1.4", 60, 97, row_number=5, invoice_no="M3", invoice_date="01/01/2026", invoice_date_iso="2026-01-01", product_name="Thep hop 30x60x1.4"),
        ]
        sales = [
            line("sales", "GENBOX", 100, 110, invoice_no="B1", invoice_date="02/01/2026", invoice_date_iso="2026-01-02", product_name="Thep hop cac loai"),
        ]
        barem_map = {
            "by_code": {},
            "by_profile": {
                "box|unknown|20x20x1.1": 10,
                "box|galvanized|30x60x1.4": 10,
            },
        }

        allocations, _, summary, _ = allocate_stock(
            [],
            purchase,
            sales,
            {"company_profile": "son_phuong", "generic_min_type_count": 2, "scenario_count": 10},
            barem_map=barem_map,
        )

        used_codes = {item["variant_code"] for item in allocations[0]["used"]}
        self.assertEqual(used_codes, {
            "AA.HOP.MK.20X20X1.1",
            "BB.HOP.MK.20X20X1.1",
            "CC.HOP.MK.30X60X1.4",
        })
        self.assertIn("2 lo", allocations[0]["generic_plan_note"])
        self.assertEqual(summary["negative_export_quantity"], 0)

    def test_son_phuong_generic_plan_retries_with_more_types_without_mutating_stock(self):
        purchase = [
            line("purchase", code, 40, 90 + index, row_number=index + 3, invoice_no=f"M{index}", invoice_date="01/01/2026", invoice_date_iso="2026-01-01", product_name=f"Thep hop {20 + index}x{20 + index}x1.1")
            for index, code in enumerate(("BOXA", "BOXB", "BOXC"), start=1)
        ]
        sales = [
            line("sales", "GENBOX", 100, 110, invoice_no="B1", invoice_date="02/01/2026", invoice_date_iso="2026-01-02", product_name="Thep hop cac loai"),
        ]

        allocations, _, summary, _ = allocate_stock(
            [],
            purchase,
            sales,
            {"company_profile": "son_phuong", "generic_min_type_count": 2},
            barem_map={"BOXA": 10, "BOXB": 10, "BOXC": 10},
        )

        used_codes = list(dict.fromkeys(item["variant_code"] for item in allocations[0]["used"]))
        self.assertEqual(len(used_codes), 3)
        self.assertEqual(sum(item["quantity"] for item in allocations[0]["used"]), 100)
        self.assertIn("3 lo", allocations[0]["generic_plan_note"])
        self.assertEqual(summary["negative_export_quantity"], 0)

    def test_son_phuong_unknown_surface_stock_can_supply_explicit_black_or_galvanized_sales(self):
        purchase = [
            line("purchase", "P.GALV", 100, 90, row_number=3, invoice_no="M1", invoice_date="01/01/2026", invoice_date_iso="2026-01-01", product_name="Thep hop ma kem 20x20x1.1"),
            line("purchase", "P.UNKNOWN", 100, 91, row_number=4, invoice_no="M2", invoice_date="01/01/2026", invoice_date_iso="2026-01-01", product_name="Thep hop 20x20x1.1"),
        ]
        sales = [
            line("sales", "S.BLACK", 100, 110, row_number=3, invoice_no="B1", invoice_date="02/01/2026", invoice_date_iso="2026-01-02", product_name="Thep hop den 20x20x1.1"),
            line("sales", "S.GALV", 100, 110, row_number=4, invoice_no="B2", invoice_date="03/01/2026", invoice_date_iso="2026-01-03", product_name="Thep hop ma kem 20x20x1.1"),
        ]
        for row in purchase + sales:
            profile = inventory_app.steel_profile_summary(row["product_name"], row["variant_code"])
            row.update({
                "source_variant_code": row["variant_code"],
                "steel_profile_key": profile["profile_key"],
                "steel_profile_code": profile["profile_code"],
                "steel_kind": profile["kind"],
                "steel_coating": profile["coating"],
                "steel_dimension": profile["dimension"],
            })

        allocations, _, summary, _ = allocate_stock(
            [], purchase, sales, {"company_profile": "son_phuong"}
        )

        self.assertEqual(
            {item["variant_code"] for item in allocations[0]["used"]},
            {"P.UNKNOWN"},
        )
        self.assertEqual(
            {item["variant_code"] for item in allocations[1]["used"]},
            {"P.GALV"},
        )
        self.assertEqual(summary["negative_export_quantity"], 0)

    def test_son_phuong_specific_unknown_surface_can_use_explicit_surface_stock(self):
        purchase = [line(
            "purchase", "P.GALV", 100, 90,
            invoice_no="M1", invoice_date="01/01/2026", invoice_date_iso="2026-01-01",
            product_name="Thep hop ma kem 20x20x1.1",
        )]
        sales = [line(
            "sales", "S.UNKNOWN", 100, 110,
            invoice_no="B1", invoice_date="02/01/2026", invoice_date_iso="2026-01-02",
            product_name="Thep hop 20x20x1.1",
        )]
        for row in purchase + sales:
            profile = inventory_app.steel_profile_summary(row["product_name"], row["variant_code"])
            row.update({
                "source_variant_code": row["variant_code"],
                "steel_profile_key": profile["profile_key"],
                "steel_profile_code": profile["profile_code"],
                "steel_kind": profile["kind"],
                "steel_coating": profile["coating"],
                "steel_dimension": profile["dimension"],
            })

        allocations, _, summary, _ = allocate_stock(
            [], purchase, sales, {"company_profile": "son_phuong"}
        )

        self.assertFalse(allocations[0]["used"][0].get("negative_export", False))
        self.assertEqual(allocations[0]["used"][0]["variant_code"], "P.GALV")
        self.assertEqual(summary["negative_export_quantity"], 0)

    def test_son_phuong_unknown_surface_never_relaxes_kind_or_dimensions(self):
        exact_unknown = {"variant_code": "EXACT", "steel_coating": "unknown"}
        wrong_dimension = {"variant_code": "WRONG-DIM", "steel_coating": "unknown"}
        wrong_kind = {"variant_code": "WRONG-KIND", "steel_coating": "unknown"}
        by_profile = {
            "box|unknown|20x20x1.1": [exact_unknown],
            "box|unknown|30x30x1.1": [wrong_dimension],
            "pipe|unknown|20x1.1": [wrong_kind],
        }

        compatible = inventory_app.son_phuong_compatible_profile_lots(
            by_profile,
            "box|black|20x20x1.1",
        )

        self.assertEqual([lot["variant_code"] for lot in compatible], ["EXACT"])


    def test_son_phuong_generic_sale_preserves_stock_reserved_for_later_specific_sale(self):
        purchase = [
            line(
                "purchase", "AA.HOP.MK.20X20X1.1", 100, 99,
                row_number=3, invoice_no="M1", invoice_date="01/01/2026",
                invoice_date_iso="2026-01-01", product_name="Thep hop ma kem 20x20x1.1",
            ),
            line(
                "purchase", "BB.HOP.MK.30X60X1.4", 100, 90,
                row_number=4, invoice_no="M2", invoice_date="01/01/2026",
                invoice_date_iso="2026-01-01", product_name="Thep hop ma kem 30x60x1.4",
            ),
        ]
        sales = [
            line(
                "sales", "GENBOX", 100, 110,
                row_number=3, invoice_no="B1", invoice_date="02/01/2026",
                invoice_date_iso="2026-01-02", product_name="Thep hop cac loai",
            ),
            line(
                "sales", "HOP.MK.20X20X1.1", 100, 110,
                row_number=4, invoice_no="B2", invoice_date="03/01/2026",
                invoice_date_iso="2026-01-03", product_name="Thep hop ma kem 20x20x1.1",
            ),
        ]
        for row in purchase + sales:
            profile = inventory_app.steel_profile_summary(
                row.get("product_name", ""),
                row.get("variant_code", ""),
            )
            row.update({
                "source_variant_code": row.get("variant_code", ""),
                "steel_profile_key": profile.get("profile_key", ""),
                "steel_profile_code": profile.get("profile_code", ""),
                "steel_kind": profile.get("kind", ""),
                "steel_coating": profile.get("coating", ""),
                "steel_dimension": profile.get("dimension", ""),
            })

        allocations, _, summary, _ = allocate_stock(
            [],
            purchase,
            sales,
            {
                "company_profile": "son_phuong",
                "generic_min_type_count": 1,
                "scenario_count": 5,
            },
            barem_map={
                "by_code": {},
                "by_profile": {
                    "box|unknown|20x20x1.1": 10,
                    "box|galvanized|30x60x1.4": 10,
                },
            },
        )

        self.assertEqual([row["invoice_no"] for row in allocations], ["B1", "B2"])
        self.assertEqual(
            {item["variant_code"] for item in allocations[0]["used"]},
            {"BB.HOP.MK.30X60X1.4"},
        )
        self.assertEqual(
            {item["variant_code"] for item in allocations[1]["used"]},
            {"AA.HOP.MK.20X20X1.1"},
        )
        self.assertTrue(allocations[1]["used"][0].get("specific_stock_reserved"))
        self.assertEqual(summary["negative_export_quantity"], 0)

    def test_son_phuong_sales_are_ordered_by_date_invoice_and_update_stock_per_row(self):
        purchase = [
            line("purchase", "BOXA", 60, 90, invoice_no="M1", invoice_date="01/01/2026", invoice_date_iso="2026-01-01", product_name="Thep hop 20x20x1.1"),
            line("purchase", "BOXB", 60, 91, invoice_no="M2", invoice_date="01/01/2026", invoice_date_iso="2026-01-01", product_name="Thep hop 30x60x1.4"),
        ]
        sales = [
            line("sales", "GENBOX", 70, 110, row_number=3, invoice_no="HD10", invoice_date="02/01/2026", invoice_date_iso="2026-01-02", product_name="Thep hop cac loai"),
            line("sales", "GENBOX", 70, 110, row_number=4, invoice_no="HD2", invoice_date="02/01/2026", invoice_date_iso="2026-01-02", product_name="Thep hop cac loai"),
        ]

        allocations, _, summary, _ = allocate_stock(
            [],
            purchase,
            sales,
            {"company_profile": "son_phuong", "generic_min_type_count": 2},
            barem_map={"BOXA": 10, "BOXB": 10},
        )

        self.assertEqual([item["invoice_no"] for item in allocations], ["HD2", "HD10"])
        self.assertEqual(sum(item["quantity"] for item in allocations[0]["used"]), 70)
        self.assertEqual(sum(item["quantity"] for item in allocations[1]["inventory_before"]), 50)
        self.assertEqual(sum(item["quantity"] for item in allocations[1]["used"] if item.get("negative_export")), 0)
        self.assertEqual(allocations[1]["unresolved_material_quantity"], 20)
        self.assertEqual(summary["negative_export_quantity"], 0)

    def test_son_phuong_pipe_box_generic_preserves_pipe_for_pipe_only_generic(self):
        purchase = [
            line("purchase", "AA.ONG.MK.21.2X1.1", 100, 90, row_number=3, invoice_no="M1", invoice_date="01/01/2026", invoice_date_iso="2026-01-01", product_name="Thep ong ma kem 21.2x1.1"),
            line("purchase", "BB.HOP.MK.20X20X1.1", 100, 90, row_number=4, invoice_no="M2", invoice_date="01/01/2026", invoice_date_iso="2026-01-01", product_name="Thep hop ma kem 20x20x1.1"),
        ]
        sales = [
            line("sales", "GEN-PIPE-BOX", 100, 110, row_number=3, invoice_no="B1", invoice_date="02/01/2026", invoice_date_iso="2026-01-02", product_name="Thep ong hop cac loai"),
            line("sales", "GEN-PIPE", 100, 110, row_number=4, invoice_no="B2", invoice_date="03/01/2026", invoice_date_iso="2026-01-03", product_name="Thep ong cac loai"),
        ]
        for row in purchase + sales:
            profile = inventory_app.steel_profile_summary(row.get("product_name", ""), row.get("variant_code", ""))
            row.update({
                "source_variant_code": row.get("variant_code", ""),
                "steel_profile_key": profile.get("profile_key", ""),
                "steel_profile_code": profile.get("profile_code", ""),
                "steel_kind": profile.get("kind", ""),
                "steel_coating": profile.get("coating", ""),
                "steel_dimension": profile.get("dimension", ""),
            })

        allocations, _, summary, _ = allocate_stock(
            [], purchase, sales,
            {"company_profile": "son_phuong", "generic_min_type_count": 1, "scenario_count": 5},
            barem_map={"by_code": {}, "by_profile": {
                "pipe|galvanized|21.2x1.1": 10,
                "box|unknown|20x20x1.1": 10,
            }},
        )

        self.assertEqual({item["variant_code"] for item in allocations[0]["used"]}, {"BB.HOP.MK.20X20X1.1"})
        self.assertEqual({item["variant_code"] for item in allocations[1]["used"]}, {"AA.ONG.MK.21.2X1.1"})
        self.assertEqual(summary["negative_export_quantity"], 0)

    def test_son_phuong_future_pipe_stock_beats_smaller_barem_remainder(self):
        purchase = [
            line("purchase", "BOXA", 60, 90, row_number=3, invoice_no="M1", invoice_date="01/01/2026", invoice_date_iso="2026-01-01", product_name="Thep hop ma kem 20x20x1.1"),
            line("purchase", "BOXB", 60, 90, row_number=4, invoice_no="M2", invoice_date="01/01/2026", invoice_date_iso="2026-01-01", product_name="Thep hop ma kem 30x30x1.1"),
            line("purchase", "PIPEA", 50, 90, row_number=5, invoice_no="M3", invoice_date="01/01/2026", invoice_date_iso="2026-01-01", product_name="Thep ong ma kem 21.2x1.1"),
            line("purchase", "PIPEB", 50, 90, row_number=6, invoice_no="M4", invoice_date="01/01/2026", invoice_date_iso="2026-01-01", product_name="Thep ong ma kem 26.65x1.1"),
        ]
        sales = [
            line("sales", "GEN-PIPE-BOX", 100, 110, row_number=3, invoice_no="B1", invoice_date="02/01/2026", invoice_date_iso="2026-01-02", product_name="Thep ong hop cac loai"),
            line("sales", "GEN-PIPE", 100, 110, row_number=4, invoice_no="B2", invoice_date="03/01/2026", invoice_date_iso="2026-01-03", product_name="Thep ong cac loai"),
        ]
        for row in purchase + sales:
            profile = inventory_app.steel_profile_summary(row.get("product_name", ""), row.get("variant_code", ""))
            row.update({
                "source_variant_code": row.get("variant_code", ""),
                "steel_profile_key": profile.get("profile_key", ""),
                "steel_profile_code": profile.get("profile_code", ""),
                "steel_kind": profile.get("kind", ""),
                "steel_coating": profile.get("coating", ""),
                "steel_dimension": profile.get("dimension", ""),
            })

        allocations, _, summary, _ = allocate_stock(
            [], purchase, sales,
            {
                "company_profile": "son_phuong",
                "generic_min_type_count": 2,
                "scenario_count": 100,
                "barem_remainder_max_kg": 10,
            },
            barem_map={"by_code": {}, "by_profile": {
                "box|unknown|20x20x1.1": 30,
                "box|galvanized|30x30x1.1": 30,
                "pipe|galvanized|21.2x1.1": 25,
                "pipe|galvanized|26.65x1.1": 25,
            }},
        )

        self.assertEqual(
            {item["variant_code"] for item in allocations[0]["used"]},
            {"BOXA", "BOXB"},
        )
        self.assertEqual(
            {item["variant_code"] for item in allocations[1]["used"]},
            {"PIPEA", "PIPEB"},
        )
        self.assertEqual(summary["negative_export_quantity"], 0)

    def test_son_phuong_pipe_generic_does_not_use_box_but_pipe_box_uses_both(self):
        purchase = [
            line("purchase", "BOXA", 12, 90, product_name="Thép hộp 20x20x1.1x6000"),
            line("purchase", "PIPEA", 12, 90, product_name="Ống thép tròn Φ26.65x1.4"),
        ]
        policy = {
            "company_profile": "son_phuong",
            "max_loss_percent": 10,
            "max_profit_percent": 25,
            "son_phuong_split_counts": {"box": 2, "pipe": 2, "pipe_box": 2},
        }

        pipe_allocations, _, _, _ = allocate_stock(
            [], purchase, [line("sales", "GENPIPE", 5, 100, product_name="Thép ống các loại")], policy
        )
        self.assertEqual([item["variant_code"] for item in pipe_allocations[0]["used"]], ["PIPEA"])

        pipe_box_allocations, _, _, _ = allocate_stock(
            [], purchase, [line("sales", "GENPIPEBOX", 10, 100, product_name="Thép ống hộp các loại")], policy
        )
        self.assertEqual(
            {item["variant_code"] for item in pipe_box_allocations[0]["used"]},
            {"BOXA", "PIPEA"},
        )

    def test_son_phuong_barem_plan_uses_only_barem_multiples_and_reports_missing_barem(self):
        purchase = [
            line("purchase", "BOXA", 100, 90, product_name="Thép hộp 20x20x1.1x6000"),
            line("purchase", "BOXB", 100, 91, product_name="Thép hộp 30x60x1.4x6000"),
            line("purchase", "PIPEA", 100, 92, product_name="Ống thép tròn Φ26.65x1.4"),
            line("purchase", "PIPEMISS", 100, 93, product_name="Ống thép tròn Φ59.9x1.4"),
        ]
        sales = [line("sales", "GENPIPEBOX", 100, 100, product_name="Thép ống hộp các loại")]

        allocations, _, summary, _ = allocate_stock(
            [],
            purchase,
            sales,
            {
                "company_profile": "son_phuong",
                "barem_tolerance_percent": 5,
            },
            barem_map={"BOXA": 30, "BOXB": 30, "PIPEA": 30},
        )

        used = allocations[0]["used"]
        self.assertTrue(used)
        self.assertTrue(all(item["variant_code"] != "PIPEMISS" for item in used))
        remainders = [item["quantity"] % 30 for item in used]
        self.assertLessEqual(sum(1 for value in remainders if value > 0.000001), 1)
        self.assertEqual(summary["material_quantity"], 100)
        self.assertEqual(summary["finished_quantity"], 0)
        self.assertEqual(summary["missing_barem_count"], 1)
        self.assertEqual(summary["missing_barem_report"][0]["variant_code"], "PIPEMISS")

    def test_son_phuong_barem_tolerance_rounds_final_type_to_nearest_whole_bar(self):
        purchase = [
            line("purchase", "PIPEA", 200, 90, product_name="Thep ong D21.2x1.4"),
        ]
        sales = [line("sales", "GENPIPE", 74, 100, product_name="Thep ong cac loai")]

        allocations, _, summary, _ = allocate_stock(
            [],
            purchase,
            sales,
            {
                "company_profile": "son_phuong",
                "generic_min_type_count": 1,
                "barem_remainder_max_kg": 10,
                "scenario_count": 5,
            },
            barem_map={"PIPEA": 12.166},
        )

        used = allocations[0]["used"]
        self.assertEqual(len(used), 1)
        self.assertAlmostEqual(used[0]["quantity"], 74)
        self.assertEqual(used[0]["barem_multiple"], 6)
        self.assertAlmostEqual(used[0]["barem_remainder_kg"], 1.004, places=3)
        self.assertAlmostEqual(summary["material_quantity"], 74)
        self.assertEqual(summary["finished_quantity"], 0)
    def test_son_phuong_barem_plan_respects_min_max_per_code_per_sale_row(self):
        purchase = [
            line("purchase", "BOXA", 100, 90, product_name="Thep hop 20x20x1.1"),
            line("purchase", "BOXB", 100, 90, product_name="Thep hop 30x60x1.4"),
            line("purchase", "BOXC", 100, 90, product_name="Thep hop 25x50x1.4"),
        ]
        sales = [line("sales", "GENBOX", 90, 100, product_name="Thep hop cac loai")]

        allocations, _, summary, _ = allocate_stock(
            [],
            purchase,
            sales,
            {
                "company_profile": "son_phuong",
                "generic_min_take_quantity": 20,
                "generic_max_take_quantity": 30,
            },
            barem_map={"BOXA": 10, "BOXB": 10, "BOXC": 10},
        )

        used = allocations[0]["used"]
        self.assertEqual(summary["material_quantity"], 90)
        self.assertTrue(used)
        self.assertTrue(all(item["quantity"] <= 30 for item in used))
        self.assertTrue(all(item["quantity"] >= 20 for item in used if not item.get("barem_remainder")))

    def test_purchase_classification_preview_returns_unknown_rows_for_filter(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.cell(2, 12, "Ma VT")
        sheet.cell(2, 13, "Ten hang")
        sheet.cell(2, 15, "So luong")
        sheet.cell(2, 16, "Don gia")
        for index in range(3, 256):
            sheet.cell(index, 12, f"BOX{index}")
            sheet.cell(index, 13, "Thep hop 20x20x1.1")
            sheet.cell(index, 15, 1)
            sheet.cell(index, 16, 10)
        sheet.cell(256, 12, "UNKNOWN")
        sheet.cell(256, 13, "Thep can nong cac loai")
        sheet.cell(256, 15, 1)
        sheet.cell(256, 16, 10)
        content = BytesIO()
        workbook.save(content)
        content.seek(0)

        response = inventory_app.app.test_client().post(
            "/api/purchase-classification-preview",
            data={
                "file": (content, "purchase.xlsx"),
                "mapping": "{}",
                "policy": '{"company_profile": "son_phuong"}',
            },
            content_type="multipart/form-data",
        )
        data = response.get_json()

        self.assertEqual(response.status_code, 200)
        self.assertTrue(data["limited"])
        self.assertEqual(data["unknown_rows"][0]["variant_code"], "UNKNOWN")

    def test_barem_file_applies_by_steel_type_and_dimension_across_company_codes(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Mã VT", "Tên hàng", "Kg / barem"])
        sheet.append(["SRC-A", "Ống thép tròn Φ26.65x1.4", 12.5])
        sheet.append(["SRC-B", "Thép hộp 20x20x1.1x6000", 7.2])
        content = BytesIO()
        workbook.save(content)
        content.seek(0)

        barem_map = parse_barem_file(content.getvalue())

        self.assertEqual(steel_profile_key("Ống thép F26.65x1.4", "OTHER-A"), "pipe|unknown|26.65x1.4")
        self.assertEqual(steel_profile_key("Thép hộp 20x20x1.1", "OTHER-B"), "box|unknown|20x20x1.1")
        purchase = [
            line("purchase", "OTHER-A", 100, 90, product_name="Ống thép F26.65x1.4"),
            line("purchase", "OTHER-B", 100, 91, product_name="Thép hộp 20x20x1.1"),
        ]
        self.assertEqual(resolve_barem_weight(barem_map, purchase[0]), 12.5)
        self.assertEqual(resolve_barem_weight(barem_map, purchase[1]), 7.2)
        sales = [line("sales", "GENPIPEBOX", 40, 100, product_name="Thép ống hộp các loại")]

        allocations, _, summary, _ = allocate_stock(
            [],
            purchase,
            sales,
            {"company_profile": "son_phuong", "barem_tolerance_percent": 5},
            barem_map=barem_map,
        )

        remainder_rows = []
        for item in allocations[0]["used"]:
            barem = resolve_barem_weight(barem_map, {"variant_code": item["variant_code"], "product_name": next(row["product_name"] for row in purchase if row["variant_code"] == item["variant_code"])})
            remainder = item["quantity"] % barem
            if min(abs(remainder), abs(barem - remainder)) >= 0.000001:
                remainder_rows.append(item)
        self.assertLessEqual(len(remainder_rows), 1)
        self.assertTrue(all(item.get("barem_remainder") for item in remainder_rows))
        self.assertEqual(summary["material_quantity"], 40)
        self.assertEqual(summary["finished_quantity"], 0)
        self.assertEqual(summary["missing_barem_count"], 0)

    def test_builtin_barem_separates_coatings_and_unspecified_sale_accepts_both(self):
        galvanized = line("purchase", "PIPE-G", 100, 90, product_name="Ống thép D59.9x2")
        black = line("purchase", "PIPE-B", 100, 90, product_name="Ống thép đen D59.9x2")
        box = line("purchase", "BOX-G", 100, 90, product_name="Thép hộp 20x20x1.1")

        self.assertEqual(resolve_barem_weight(DEFAULT_BAREM_MAP, galvanized), 17.14)
        self.assertEqual(resolve_barem_weight(DEFAULT_BAREM_MAP, black), 17.13)
        self.assertEqual(resolve_barem_weight(DEFAULT_BAREM_MAP, box), 3.87)

        sales = [line("sales", "GENPIPE", 20, 100, product_name="Thép ống các loại")]
        allocations, _, _, _ = allocate_stock(
            [],
            [black, galvanized],
            sales,
            {"company_profile": "son_phuong", "barem_tolerance_percent": 5},
            barem_map=DEFAULT_BAREM_MAP,
        )
        self.assertEqual({item["variant_code"] for item in allocations[0]["used"]}, {"PIPE-G", "PIPE-B"})
        self.assertLessEqual(sum(1 for item in allocations[0]["used"] if item.get("barem_remainder")), 1)

    def test_son_phuong_material_shortage_stays_unresolved_without_negative_khhvt(self):
        purchase = [line(
            "purchase", "PIPE-G", 20, 90,
            product_name="Ống thép D59.9x2",
            invoice_date="01/01/2025",
            invoice_date_iso="2025-01-01",
        )]
        sales = [line(
            "sales", "THEP-CAC-LOAI", 50, 100,
            product_name="Thép ống các loại",

            invoice_date="01/01/2025",
            invoice_date_iso="2025-01-01",
        )]

        allocations, _, summary, _ = allocate_stock(
            [],
            purchase,
            sales,
            {"company_profile": "son_phuong"},
            barem_map=DEFAULT_BAREM_MAP,
        )
        ledger = build_inventory_ledger([], purchase, allocations, company_profile="son_phuong")

        self.assertEqual([warehouse["warehouse_code"] for warehouse in ledger["warehouses"]], ["KHHVT"])
        self.assertEqual(summary["finished_quantity"], 0)
        self.assertEqual(summary["material_quantity"], 20)
        self.assertEqual(summary["unresolved_material_quantity"], 30)
        self.assertEqual(summary["negative_export_quantity"], 0)
        self.assertEqual(allocations[0]["used"][0]["ledger_variant_code"], "PIPE-G")
        self.assertFalse(any(item.get("negative_export") for item in allocations[0]["used"]))
        self.assertEqual(allocations[0]["unresolved_material_quantity"], 30)

    def test_son_phuong_structural_shortage_moves_only_remainder_to_ktp(self):
        purchase = [line(
            "purchase", "THEP-V100", 5, 90,
            product_name="Thep V 50x50x5",
            invoice_date="01/01/2025",
            invoice_date_iso="2025-01-01",
        )]
        sales = [line(
            "sales", "THEP-V100", 8, 100,
            product_name="Thep V 50x50x5",
            invoice_date="02/01/2025",
            invoice_date_iso="2025-01-02",
        )]

        allocations, _, summary, _ = allocate_stock([], purchase, sales, {"company_profile": "son_phuong"})
        ledger = build_inventory_ledger([], purchase, allocations, company_profile="son_phuong")

        self.assertEqual(allocations[0]["allocation_role"], "finished_goods")
        self.assertEqual(allocations[0]["material_quantity"], 5)
        self.assertEqual(allocations[0]["finished_quantity"], 3)
        self.assertEqual(summary["negative_export_quantity"], 0)
        self.assertEqual([warehouse["warehouse_code"] for warehouse in ledger["warehouses"]], ["KHHVT", "KTP"])

    def test_son_phuong_solid_square_is_finished_goods_not_pipe_box(self):
        purchase = [line(
            "purchase", "THEP-VUONG-16", 5, 90,
            product_name="Thep vuong 16",
            invoice_date="01/01/2025",
            invoice_date_iso="2025-01-01",
        )]
        sales = [line(
            "sales", "THEP-VUONG-16", 8, 100,
            product_name="Thep vuong 16",
            invoice_date="02/01/2025",
            invoice_date_iso="2025-01-02",
        )]

        allocations, _, summary, _ = allocate_stock([], purchase, sales, {"company_profile": "son_phuong"})

        self.assertEqual(steel_kind("Thep vuong 16", "THEP-VUONG-16"), "unknown")
        self.assertEqual(allocations[0]["allocation_role"], "finished_goods")
        self.assertEqual(allocations[0]["material_quantity"], 5)
        self.assertEqual(allocations[0]["finished_quantity"], 3)
        self.assertEqual(summary["negative_export_quantity"], 0)
    def test_son_phuong_unclassified_product_uses_khock_159(self):
        sales = [line(
            "sales", "MAY-HAN", 2, 100,
            product_name="May han cong nghiep",
            invoice_date="02/01/2025",
            invoice_date_iso="2025-01-02",
        )]
        policy = {
            "company_profile": "son_phuong",
            "sales_inventory_pairs": [
                {"role": "materials", "ma_kho": "KHHVT", "tk_vat_tu": "156"},
                {"role": "finished_goods", "ma_kho": "KTP", "tk_vat_tu": "155"},
                {"role": "fallback", "ma_kho": "KHOCK", "tk_vat_tu": "159"},
            ],
        }

        allocations, _, summary, warnings = allocate_stock([], [], sales, policy)
        ledger = build_inventory_ledger([], [], allocations, company_profile="son_phuong")

        self.assertEqual(allocations[0]["allocation_role"], "fallback")
        self.assertEqual(allocations[0]["material_quantity"], 0)
        self.assertEqual(allocations[0]["finished_quantity"], 2)
        self.assertEqual(allocations[0]["remainder_warehouse_code"], "KHOCK")
        self.assertEqual(summary["negative_export_quantity"], 0)
        self.assertFalse(any("chua ghep du" in warning for warning in warnings))
        self.assertEqual([warehouse["warehouse_code"] for warehouse in ledger["warehouses"]], ["KHOCK"])
        self.assertEqual(ledger["warehouses"][0]["account"], "159")

    def test_son_phuong_export_removes_old_ktp_sheets_and_writes_only_khh_ledger(self):
        purchase = [line(
            "purchase", "PIPE-G", 10, 90, invoice_no="MUA1", invoice_date="01/01/2025",
            invoice_date_iso="2025-01-01", product_name="Ong thep D59.9x2", unit_name="kg", line_amount=900,
        )]
        sales = [line(
            "sales", "GENPIPE", 12, 100, invoice_no="BAN1", invoice_date="02/01/2025",
            invoice_date_iso="2025-01-02", product_name="Thep ong cac loai", unit_name="kg", line_amount=1200,
        )]
        allocations, stock_rows, summary, _ = allocate_stock(
            [],
            purchase,
            sales,
            {"company_profile": "son_phuong"},
            barem_map={"PIPE-G": 5},
        )
        ledger = build_inventory_ledger([], purchase, allocations, sales_lines=sales, company_profile="son_phuong")
        ledger["date_range"] = {"from": "2025-01-02", "to": "2025-01-02"}
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Invoices"
        sheet.cell(2, 12, "Ma VT")
        sheet.cell(2, 13, "Ten hang")
        sheet.cell(2, 15, "So luong")
        sheet.cell(2, 16, "Don gia")
        for old_name in ("SoChiTietKTP", "SoChiTietHH_TP", "TongHopNXT_KTP", "BaoCaoBH_KTP", "BangKeHDBH_KTP"):
            workbook.create_sheet(old_name)
        content = BytesIO()
        workbook.save(content)
        content.seek(0)

        result_content = create_output_workbook(
            content.getvalue(),
            {"sheet": "Invoices", "header_row": 2, "data_start_row": 3},
            allocations,
            stock_rows,
            summary,
            {"company_profile": "son_phuong"},
            purchase_lines=purchase,
            sales_lines=sales,
            ledger=ledger,
        )
        result = load_workbook(result_content)

        self.assertNotIn("SoChiTietKVT", result.sheetnames)
        self.assertIn("SoChiTietKHHVT", result.sheetnames)
        self.assertNotIn("SoChiTietKTP", result.sheetnames)
        self.assertNotIn("SoChiTietHH_TP", result.sheetnames)
        self.assertNotIn("TongHopNXT_KVT", result.sheetnames)
        self.assertIn("TongHopNXT_KHHVT", result.sheetnames)
        self.assertNotIn("TongHopNXT_KTP", result.sheetnames)
        self.assertNotIn("BaoCaoBH_KVT", result.sheetnames)
        self.assertIn("BaoCaoBH_KHHVT", result.sheetnames)
        self.assertNotIn("BaoCaoBH_KTP", result.sheetnames)
        self.assertNotIn("BangKeHDBH_KVT", result.sheetnames)
        self.assertIn("BangKeHDBH_KHHVT", result.sheetnames)
        self.assertNotIn("BangKeHDBH_KTP", result.sheetnames)

    def test_ambiguous_steel_report_does_not_treat_nong_as_ong(self):
        self.assertIsNone(inventory_app.steel_barem_ambiguity({
            "product_name": "Thép cuộn cán nóng các loại",
            "variant_code": "TH.TCCNCL",
        }))

    def test_default_barem_api_returns_four_reference_tables(self):
        response = inventory_app.app.test_client().get("/api/default-barem")
        data = response.get_json()

        self.assertEqual(response.status_code, 200)
        self.assertEqual(data["count"], len(data["rows"]))
        self.assertEqual({
            (row["kind"], row["coating"])
            for row in data["rows"]
        }, {
            ("box", "black"),
            ("pipe", "black"),
            ("pipe", "galvanized"),
            ("box", "galvanized"),
        })

    def test_uses_lowest_profit_split_code_first(self):
        purchase = [
            line("purchase", "A.001", 5, 90, invoice_no="MUA1"),
            line("purchase", "A.002", 10, 98, invoice_no="MUA2"),
        ]
        sales = [line("sales", "A", 12, 100)]

        allocations, stock, summary, _ = allocate_stock([], purchase, sales)

        used = allocations[0]["used"]
        self.assertEqual([item["variant_code"] for item in used], ["A.002", "A.001"])
        self.assertEqual([item["quantity"] for item in used], [10, 2])
        self.assertEqual([item["variant_code"] for item in allocations[0]["inventory_before"]], ["A.002", "A.001"])
        self.assertEqual(allocations[0]["inventory_after"][0]["variant_code"], "A.001")
        self.assertEqual(allocations[0]["inventory_after"][0]["quantity"], 3)
        self.assertEqual(summary["material_quantity"], 12)
        self.assertEqual(summary["finished_quantity"], 0)
        remaining = {item["variant_code"]: item["ending_quantity"] for item in stock}
        self.assertEqual(remaining, {"A.001": 3, "A.002": 0})
        self.assertTrue(all(item["opening_quantity"] == 0 for item in stock))

    def test_summarizes_same_purchase_code_with_weighted_average_cost(self):
        purchase = [
            line("purchase", "A.001", 2, 10, invoice_no="MUA1"),
            line("purchase", "A.001", 6, 20, invoice_no="MUA2"),
        ]
        sales = [line("sales", "A", 5, 30)]

        allocations, stock, _, _ = allocate_stock([], purchase, sales)

        used = allocations[0]["used"]
        self.assertEqual(len(used), 1)
        self.assertEqual(used[0]["variant_code"], "A.001")
        self.assertEqual(used[0]["quantity"], 5)
        self.assertEqual(used[0]["unit_cost"], 17.5)
        self.assertEqual(used[0]["summary_count"], 2)
        self.assertEqual(stock[0]["purchase_quantity"], 8)
        self.assertEqual(stock[0]["ending_quantity"], 3)

    def test_suffix_sales_rows_can_use_other_purchase_suffixes_under_same_base(self):
        purchase = [
            line("purchase", "A.001", 5, 90, invoice_no="MUA1"),
            line("purchase", "A.002", 10, 98, invoice_no="MUA2"),
        ]
        sales = [line("sales", "A.001", 6, 100)]

        allocations, stock, summary, _ = allocate_stock([], purchase, sales)

        self.assertEqual([item["variant_code"] for item in allocations[0]["used"]], ["A.002"])
        self.assertEqual(allocations[0]["material_quantity"], 6)
        self.assertEqual(allocations[0]["finished_quantity"], 0)
        remaining = {item["variant_code"]: item["ending_quantity"] for item in stock}
        self.assertEqual(remaining, {"A.001": 5, "A.002": 4})
        self.assertEqual(summary["material_quantity"], 6)

    def test_sends_shortage_to_finished_goods(self):
        opening = [line("opening", "B.001", 2, None)]
        purchase = [line("purchase", "B.002", 3, 60)]
        sales = [line("sales", "B", 8, 75)]

        allocations, _, summary, _ = allocate_stock(opening, purchase, sales)

        self.assertEqual(allocations[0]["material_quantity"], 5)
        self.assertEqual(allocations[0]["finished_quantity"], 3)
        self.assertEqual(summary["finished_quantity"], 3)

    def test_yen_thanh_shortage_exports_negative_in_ktp_not_khh(self):
        purchase = [line("purchase", "B.001", 3, 60)]
        sales = [line("sales", "B", 8, 75)]

        allocations, _, summary, _ = allocate_stock([], purchase, sales, {"allow_negative_export": True})
        ledger = build_inventory_ledger([], purchase, allocations)

        self.assertEqual(allocations[0]["material_quantity"], 3)
        self.assertEqual(allocations[0]["finished_quantity"], 5)
        self.assertEqual(summary["material_quantity"], 3)
        self.assertEqual(summary["finished_quantity"], 5)
        self.assertFalse(any(item.get("negative_export") for item in allocations[0]["used"]))
        ktp_group = ledger["warehouses"][1]["groups"][0]
        self.assertEqual(ktp_group["rows"][0]["type"], "sale")
        self.assertEqual(sum(row["qty_out"] for row in ktp_group["rows"]), 5)

    def test_future_purchase_does_not_cover_earlier_sale(self):
        purchase = [line(
            "purchase", "A.001", 500, 10, invoice_no="MUA42", invoice_date="30/06/2025",
            invoice_date_iso="2025-06-30",
        )]
        sales = [
            line(
                "sales", "A", 15, 20, invoice_no="BAN312", invoice_date="11/02/2025",
                invoice_date_iso="2025-02-11", row_number=3,
            ),
            line(
                "sales", "A", 20, 20, invoice_no="BAN2246", invoice_date="02/07/2025",
                invoice_date_iso="2025-07-02", row_number=4,
            ),
        ]

        allocations, _, summary, _ = allocate_stock([], purchase, sales)
        ledger = build_inventory_ledger([], purchase, allocations, sales_lines=sales)

        self.assertEqual(allocations[0]["material_quantity"], 0)
        self.assertEqual(allocations[0]["finished_quantity"], 15)
        self.assertEqual(allocations[1]["material_quantity"], 20)
        self.assertEqual(allocations[1]["finished_quantity"], 0)
        self.assertEqual(summary["material_quantity"], 20)
        self.assertEqual(summary["finished_quantity"], 15)
        khh_rows = ledger["warehouses"][0]["groups"][0]["rows"]
        ktp_rows = ledger["warehouses"][1]["groups"][0]["rows"]
        self.assertNotIn("11/02/2025", [row.get("date") for row in khh_rows if row["type"] == "sale"])
        self.assertEqual([row["date"] for row in ktp_rows], ["11/02/2025"])
        self.assertEqual(ktp_rows[0]["qty_in"], 0)
        self.assertEqual(ktp_rows[0]["qty_out"], 15)

    def test_yen_thanh_future_purchase_option_moves_purchase_before_sale(self):
        purchase = [line(
            "purchase", "A.001", 100, 10, invoice_no="MUA42", invoice_date="20/02/2025",
            invoice_date_iso="2025-02-20", row_number=42,
        )]
        sales = [line(
            "sales", "A", 30, 20, invoice_no="BAN312", invoice_date="11/02/2025",
            invoice_date_iso="2025-02-11", row_number=312,
        )]

        allocations, _, summary, _ = allocate_stock([], purchase, sales, {
            "company_profile": "yen_thanh",
            "allow_future_purchase_reorder": True,
            "future_purchase_window_days": 31,
        })
        ledger = build_inventory_ledger([], purchase, allocations, sales_lines=sales)
        khh_rows = ledger["warehouses"][0]["groups"][0]["rows"]

        self.assertEqual(allocations[0]["material_quantity"], 30)
        self.assertEqual(allocations[0]["finished_quantity"], 0)
        self.assertTrue(allocations[0]["used"][0]["future_purchase_reordered"])
        self.assertEqual(allocations[0]["used"][0]["future_reorder_days"], 9)
        self.assertEqual(summary["future_purchase_reorder_count"], 1)
        self.assertEqual(summary["future_purchase_reorder_quantity"], 30)
        self.assertEqual(khh_rows[0]["type"], "purchase_future_reorder")
        self.assertEqual(khh_rows[0]["date"], "11/02/2025")
        self.assertTrue(khh_rows[0]["future_purchase_reordered"])

    def test_yen_thanh_future_purchase_window_blocks_far_purchase(self):
        purchase = [line(
            "purchase", "A.001", 100, 10, invoice_no="MUA99", invoice_date="20/04/2025",
            invoice_date_iso="2025-04-20",
        )]
        sales = [line(
            "sales", "A", 30, 20, invoice_no="BAN312", invoice_date="11/02/2025",
            invoice_date_iso="2025-02-11",
        )]

        allocations, _, summary, _ = allocate_stock([], purchase, sales, {
            "company_profile": "yen_thanh",
            "allow_future_purchase_reorder": True,
            "future_purchase_window_days": 31,
        })

        self.assertEqual(allocations[0]["material_quantity"], 0)
        self.assertEqual(allocations[0]["finished_quantity"], 30)
        self.assertEqual(summary["future_purchase_reorder_count"], 0)

    def test_clears_fractional_rounding_remainder_when_stock_is_sufficient(self):
        purchase = [line("purchase", "R.001", 0.1, 90), line("purchase", "R.002", 0.2, 95)]
        sales = [line("sales", "R", 0.3, 100)]

        allocations, _, summary, _ = allocate_stock([], purchase, sales)

        self.assertEqual(allocations[0]["finished_quantity"], 0)
        self.assertEqual(summary["finished_quantity"], 0)

    def test_acceptance_range_excludes_excess_loss_and_excess_profit(self):
        purchase = [
            line("purchase", "A.001", 4, 120),
            line("purchase", "A.002", 4, 90),
            line("purchase", "A.003", 4, 60),
        ]
        sales = [line("sales", "A", 6, 100)]

        allocations, stock, summary, warnings = allocate_stock(
            [], purchase, sales, {"max_loss_percent": 10, "max_profit_percent": 25}
        )

        self.assertEqual([item["variant_code"] for item in allocations[0]["used"]], ["A.002"])
        self.assertEqual(allocations[0]["material_quantity"], 4)
        self.assertEqual(allocations[0]["finished_quantity"], 2)
        self.assertEqual({item["variant_code"] for item in allocations[0]["rejected"]}, {"A.001", "A.003"})
        self.assertEqual(summary["range_rejected_lines"], 1)
        self.assertTrue(warnings)
        remaining = {item["variant_code"]: item["ending_quantity"] for item in stock}
        self.assertEqual(remaining, {"A.001": 4, "A.002": 0, "A.003": 4})

    def test_loss_priority_uses_smallest_loss_before_larger_loss(self):
        purchase = [
            line("purchase", "A.001", 4, 105),
            line("purchase", "A.002", 4, 120),
        ]
        sales = [line("sales", "A", 5, 100)]

        allocations, _, _, _ = allocate_stock(
            [], purchase, sales, {"max_loss_percent": 30, "max_profit_percent": 25}
        )

        self.assertEqual([item["variant_code"] for item in allocations[0]["used"]], ["A.001", "A.002"])
        self.assertEqual([item["quantity"] for item in allocations[0]["used"]], [4, 1])

    def test_ktp_suffix_fallback_uses_best_purchase_suffix_even_when_rejected(self):
        purchase = [
            line("purchase", "A.001", 4, 105),
            line("purchase", "A.002", 4, 120),
        ]
        sales = [line("sales", "A", 3, 100)]

        allocations, _, _, _ = allocate_stock(
            [], purchase, sales, {"max_loss_percent": 1, "max_profit_percent": 25}
        )
        ledger = build_inventory_ledger([], purchase, allocations, sales_lines=sales)

        self.assertEqual(allocations[0]["material_quantity"], 0)
        self.assertEqual(allocations[0]["finished_quantity"], 3)
        self.assertEqual(allocations[0]["finished_variant_code"], "A.001")
        self.assertEqual(ledger["warehouses"][1]["groups"][0]["variant_code"], "A.001")

    def test_acceptance_range_rejects_missing_price_needed_for_validation(self):
        opening = [line("opening", "D.001", 3, None)]
        sales = [line("sales", "D", 2, 100)]

        allocations, _, _, _ = allocate_stock(
            opening, [], sales, {"max_loss_percent": 10, "max_profit_percent": 25}
        )

        self.assertEqual(allocations[0]["material_quantity"], 0)
        self.assertEqual(allocations[0]["finished_quantity"], 2)

    def test_lists_sales_codes_that_have_no_opening_or_purchase_stock(self):
        opening = [line("opening", "A.001", 1, 95)]
        purchase = [line("purchase", "B.002", 2, 90)]
        sales = [
            line("sales", "A", 1, 100, row_number=3),
            line("sales", "B", 2, 100, row_number=4),
            line("sales", "C", 3, 100, row_number=5),
            line("sales", "C", 4, 100, row_number=6),
        ]

        result = find_sale_only_codes(opening, purchase, sales)

        self.assertEqual(len(result), 1)
        self.assertEqual(result[0]["variant_code"], "C")
        self.assertEqual(result[0]["quantity"], 7)
        self.assertEqual(result[0]["rows"], [5, 6])
        self.assertEqual(result[0]["opening_quantity"], 0)
        self.assertEqual(result[0]["purchase_quantity"], 0)

        _, stock_rows, _, _ = allocate_stock(opening, purchase, sales)
        missing_stock = next(row for row in stock_rows if row["variant_code"] == "C")
        self.assertEqual(missing_stock["opening_quantity"], 0)
        self.assertEqual(missing_stock["purchase_quantity"], 0)
        self.assertEqual(missing_stock["ending_quantity"], 0)

    def test_output_preserves_sales_sheet_and_adds_two_columns(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Invoices"
        sheet.cell(2, 13, "Ma VT")
        sheet.cell(2, 16, "So luong")
        sheet.cell(3, 13, "A")
        sheet.cell(3, 16, 4)
        content = BytesIO()
        workbook.save(content)
        allocation = line("sales", "A", 4, 100)
        allocation.update({"material_quantity": 3, "finished_quantity": 1, "used": []})
        mapping = {"sheet": "", "header_row": 2, "data_start_row": 3}

        result = create_output_workbook(content.getvalue(), mapping, [allocation], [], {
            "opening_quantity": 0,
            "purchase_quantity": 3,
            "sales_quantity": 4,
            "material_quantity": 3,
            "finished_quantity": 1,
            "material_percent": 75,
        })
        output = load_workbook(result)
        original = output["Invoices"]

        self.assertEqual(original.cell(2, 17).value, "SL khớp từ mua vào")
        self.assertEqual(original.cell(2, 18).value, "SL xuất theo kho bán ra")
        self.assertEqual(original.cell(3, 17).value, 3)
        self.assertEqual(original.cell(3, 18).value, 1)
        self.assertEqual(original.cell(2, 19).value, "Tồn mua vào trước khi bán")
        self.assertEqual(original.cell(2, 21).value, "Tồn mua vào sau khi bán")
        self.assertNotIn("PhanBoKho", output.sheetnames)
        self.assertNotIn("TonKhoHangHoa", output.sheetnames)
        self.assertIn("MaChiBanRaKhongTon", output.sheetnames)
        self.assertIn("MauMuaVao", output.sheetnames)
        self.assertIn("MauBanRa", output.sheetnames)

    def test_output_adds_purchase_and_sales_template_sheets(self):
        purchase = [
            line(
                "purchase", "A.001", 2, 70, invoice_no="MUA1", invoice_date="01/01/2025",
                party_tax_code="NCC001", party_name="Nha cung cap 1", line_amount=140,
            ),
            line(
                "purchase", "A.002", 3, 90, invoice_no="MUA2", invoice_date="02/01/2025",
                party_tax_code="NCC002", party_name="Nha cung cap 2", line_amount=270,
            ),
        ]
        sales = [line(
            "sales", "A", 6, 120, invoice_no="BAN1", invoice_date="03/01/2025",
            party_tax_code="KH001", party_name="Khach hang 1", line_amount=720,
            tax_rate_percent=10, tax_amount=72,
        )]
        allocations, stock, summary, _ = allocate_stock([], purchase, sales)

        result = create_output_workbook(
            invoice_file("A", 6, 120).getvalue(),
            clean_mapping({}, "sales"),
            allocations,
            stock,
            summary,
            purchase_lines=purchase,
            sales_lines=sales,
        )
        output = load_workbook(result)
        purchase_sheet = output["MauMuaVao"]
        sales_sheet = output["MauBanRa"]

        self.assertEqual([purchase_sheet.cell(1, column).value for column in range(1, 10)], [
            "MST", "Tên công ty", "Số HĐ", "Ngày HĐ", "Mã kho", "Mã VT", "Số lượng", "Đơn giá mua", "Thành tiền mua",
        ])
        self.assertEqual(purchase_sheet.cell(2, 5).value, "KVT")
        self.assertEqual(purchase_sheet.cell(2, 6).value, "A.001")
        self.assertEqual(purchase_sheet.cell(3, 9).value, 270)

        self.assertEqual(sales_sheet.cell(2, 12).value, "KVT")
        self.assertEqual(sales_sheet.cell(2, 13).value, "A.002")
        self.assertEqual(sales_sheet.cell(2, 5).value, 3)
        self.assertEqual(sales_sheet.cell(2, 7).value, 360)
        self.assertEqual(sales_sheet.cell(2, 8).value, 90)
        self.assertEqual(sales_sheet.cell(2, 11).value, 36)
        self.assertEqual(sales_sheet.cell(3, 12).value, "KVT")
        self.assertEqual(sales_sheet.cell(3, 13).value, "A.001")
        self.assertEqual(sales_sheet.cell(4, 12).value, "KTP")
        self.assertEqual(sales_sheet.cell(4, 13).value, "A.002")
        self.assertEqual(sales_sheet.cell(4, 8).value, 0)
        self.assertEqual(sales_sheet.cell(4, 9).value, 0)
        self.assertEqual(sales_sheet.cell(4, 11).value, 12)

    def test_inventory_ledger_splits_khh_and_ktp_rows(self):
        purchase = [line(
            "purchase", "A.001", 10, 5, invoice_no="MUA1", invoice_date="01/01/2025",
            invoice_date_iso="2025-01-01", party_name="Nha cung cap", unit_name="Cai", line_amount=50,
        )]
        sales = [line(
            "sales", "A", 12, 20, invoice_no="BAN1", invoice_date="02/01/2025",
            invoice_date_iso="2025-01-02", party_name="Khach hang", unit_name="Cai", line_amount=80,
        )]
        allocations, _, _, _ = allocate_stock([], purchase, sales)

        ledger = build_inventory_ledger([], purchase, allocations)
        khh = ledger["warehouses"][0]
        ktp = ledger["warehouses"][1]
        khh_group = next(group for group in khh["groups"] if group["variant_code"] == "A.001")
        ktp_group = ktp["groups"][0]
        rows = khh_group["rows"]

        self.assertEqual(khh["warehouse_code"], "KVT")
        self.assertEqual(ktp["warehouse_code"], "KTP")
        self.assertEqual(khh_group["variant_code"], "A.001")
        self.assertEqual(khh_group["unit_name"], "Cai")
        self.assertEqual(rows[0]["type"], "purchase")
        self.assertEqual(rows[0]["date"], "01/01/2025")
        self.assertEqual(rows[0]["qty_in"], 10)
        self.assertEqual(rows[0]["amount_in"], 50)
        self.assertEqual(rows[1]["type"], "sale")
        self.assertEqual(rows[1]["date"], "02/01/2025")
        self.assertEqual(rows[1]["account"], "6321")
        self.assertEqual(rows[1]["variant_code"], "A.001")
        self.assertEqual(rows[1]["qty_out"], 10)
        self.assertEqual(rows[1]["amount_out"], 50)
        self.assertEqual(ktp_group["variant_code"], "A.001")
        self.assertEqual(ktp_group["rows"][0]["type"], "sale")
        self.assertEqual(ktp_group["rows"][0]["qty_in"], 0)
        self.assertEqual(ktp_group["rows"][0]["qty_out"], 2)
        self.assertEqual(ktp_group["rows"][0]["unit_price"], 0)
        self.assertEqual(ktp_group["rows"][0]["amount_in"], 0)
        self.assertEqual(ktp_group["rows"][0]["amount_out"], 0)
        self.assertEqual(ktp_group["rows"][0]["account"], "6321")

    def test_inventory_ledger_routes_matched_purchase_into_sales_warehouse_for_cost(self):
        purchase = [line(
            "purchase", "PAPER", 10, 50, invoice_no="MUA1", invoice_date="01/01/2025",
            invoice_date_iso="2025-01-01", party_name="Nha cung cap", unit_name="Kg",
            line_amount=500, warehouse_code="KTP", warehouse_account="1551",
        )]
        sales = [line(
            "sales", "PAPER", 3, 120, invoice_no="BAN1", invoice_date="02/01/2025",
            invoice_date_iso="2025-01-02", party_name="Khach hang", unit_name="Kg",
            line_amount=360, warehouse_code="KVT", warehouse_account="152",
        )]
        allocations, _, _, _ = allocate_stock([], purchase, sales)

        ledger = build_inventory_ledger([], purchase, allocations, sales_lines=sales)
        khh = next(warehouse for warehouse in ledger["warehouses"] if warehouse["warehouse_code"] == "KVT")
        khh_group = khh["groups"][0]

        self.assertEqual(khh["account"], "152")
        self.assertFalse(any(warehouse["warehouse_code"] == "KTP" for warehouse in ledger["warehouses"]))
        self.assertEqual(sum(row["qty_in"] for row in khh_group["rows"]), 10)
        self.assertEqual(sum(row["amount_in"] for row in khh_group["rows"]), 500)
        self.assertEqual(sum(row["qty_out"] for row in khh_group["rows"]), 3)
        purchase_row = next(row for row in khh_group["rows"] if row["type"] == "purchase")
        sale_row = next(row for row in khh_group["rows"] if row["type"] == "sale")
        self.assertEqual(purchase_row["qty_in"], 10)
        self.assertFalse(sale_row["cost_missing"])
        self.assertEqual(sale_row["unit_price"], 50)
        self.assertEqual(sale_row["amount_out"], 150)

    def test_inventory_ledger_moves_unmatched_khh_remainder_to_ktp(self):
        purchase = [line(
            "purchase", "PAPER", 10, 50, invoice_no="MUA1", invoice_date="05/01/2025",
            invoice_date_iso="2025-01-05", party_name="Nha cung cap", unit_name="Kg",
            line_amount=500, warehouse_code="KTP", warehouse_account="1551",
        )]
        sales = [line(
            "sales", "PAPER", 3, 120, invoice_no="BAN1", invoice_date="02/01/2025",
            invoice_date_iso="2025-01-02", party_name="Khach hang", unit_name="Kg",
            line_amount=360, warehouse_code="KVT", warehouse_account="152",
        )]
        allocations, _, _, _ = allocate_stock([], purchase, sales)

        ledger = build_inventory_ledger([], purchase, allocations, sales_lines=sales)
        warehouse_codes = [warehouse["warehouse_code"] for warehouse in ledger["warehouses"]]
        khh = next(warehouse for warehouse in ledger["warehouses"] if warehouse["warehouse_code"] == "KVT")
        ktp = next(warehouse for warehouse in ledger["warehouses"] if warehouse["warehouse_code"] == "KTP")
        sale_row = next(row for group in khh["groups"] for row in group["rows"] if row["type"] == "sale")
        purchase_row = next(row for group in ktp["groups"] for row in group["rows"] if row["type"] == "purchase")

        self.assertIn("KVT", warehouse_codes)
        self.assertEqual(sale_row["qty_out"], 3)
        self.assertEqual(sale_row["unit_price"], 0)
        self.assertTrue(sale_row["cost_missing"])
        self.assertEqual(purchase_row["qty_in"], 10)

    def test_inventory_ledger_puts_full_shortage_only_in_ktp(self):
        sales = [line(
            "sales", "NO_STOCK", 5, 20, invoice_no="BAN2", invoice_date="03/01/2025",
            invoice_date_iso="2025-01-03", party_name="Khach hang", unit_name="Cai", line_amount=100,
        )]
        allocations, _, _, _ = allocate_stock([], [], sales)

        ledger = build_inventory_ledger([], [], allocations)
        ktp = ledger["warehouses"][0]

        self.assertEqual(ktp["groups"][0]["variant_code"], "NO_STOCK")
        self.assertEqual([row["type"] for row in ktp["groups"][0]["rows"]], ["sale"])
        self.assertEqual(ktp["groups"][0]["rows"][0]["qty_in"], 0)
        self.assertEqual(ktp["groups"][0]["rows"][0]["qty_out"], 5)

    def test_sales_warehouse_from_processed_sales_drives_export_only_rows(self):
        sales = [line(
            "sales", "GC_ONLY", 7, 120, invoice_no="BAN-GC", invoice_date="04/01/2025",
            invoice_date_iso="2025-01-04", party_name="Khach gia cong", unit_name="Cai",
            line_amount=840, warehouse_code="KGCI", warehouse_account="1552",
        )]
        allocations, _, summary, _ = allocate_stock([], [], sales)
        ledger = build_inventory_ledger([], [], allocations, sales_lines=sales)
        report_rows = inventory_app.build_sales_report_rows(allocations, [], sales)

        self.assertEqual(summary["material_quantity"], 0)
        self.assertEqual(summary["finished_quantity"], 7)
        self.assertEqual([warehouse["warehouse_code"] for warehouse in ledger["warehouses"]], ["KGCI"])
        self.assertEqual(ledger["warehouses"][0]["account"], "1552")
        rows = ledger["warehouses"][0]["groups"][0]["rows"]
        self.assertEqual([row["type"] for row in rows], ["sale"])
        self.assertEqual(sum(row["qty_in"] for row in rows), 0)
        self.assertEqual(sum(row["qty_out"] for row in rows), 7)
        self.assertEqual(report_rows[0]["warehouse_code"], "KGCI")
        self.assertEqual(report_rows[0]["cost_amount"], 0)

    def test_ktp_uses_zero_cost_when_no_purchase_stock_exists(self):
        sales = [
            line(
                "sales", "KTP_ONLY", 2, 20, invoice_no="BAN1", invoice_date="03/01/2025",
                invoice_date_iso="2025-01-03", party_name="Khach hang", unit_name="Cai", line_amount=40,
                row_number=3,
            ),
            line(
                "sales", "KTP_ONLY", 6, 30, invoice_no="BAN2", invoice_date="04/01/2025",
                invoice_date_iso="2025-01-04", party_name="Khach hang", unit_name="Cai", line_amount=180,
                row_number=4,
            ),
        ]
        allocations, _, _, _ = allocate_stock([], [], sales)

        ledger = build_inventory_ledger([], [], allocations, sales_lines=sales)
        rows = ledger["warehouses"][0]["groups"][0]["rows"]

        self.assertEqual([row["type"] for row in rows], ["sale", "sale"])
        self.assertTrue(all(row["unit_price"] == 0 for row in rows))
        self.assertEqual(sum(row["qty_in"] for row in rows), 0)
        self.assertEqual(sum(row["qty_out"] for row in rows), 8)

    def test_output_reuses_existing_result_columns(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Invoices"
        sheet.cell(2, 13, "Ma VT")
        sheet.cell(2, 16, "So luong")
        sheet.cell(2, 17, "SL lấy từ kho hàng hóa")
        sheet.cell(2, 18, "SL lấy từ kho thành phẩm")
        content = BytesIO()
        workbook.save(content)
        allocation = line("sales", "A", 2, 100)
        allocation.update({"material_quantity": 2, "finished_quantity": 0, "used": []})
        result = create_output_workbook(
            content.getvalue(),
            {"sheet": "", "header_row": 2, "data_start_row": 3},
            [allocation],
            [],
            {
                "opening_quantity": 0, "purchase_quantity": 2, "sales_quantity": 2,
                "material_quantity": 2, "finished_quantity": 0, "material_percent": 100,
            },
        )
        output = load_workbook(result)

        self.assertEqual(output["Invoices"].max_column, 22)
        self.assertEqual(output["Invoices"].cell(3, 17).value, 2)

    def test_api_generates_downloadable_workbook(self):
        response = inventory_app.app.test_client().post(
            "/api/analyze",
            data={
                "purchase_file": (invoice_file("C.001", 3, 95), "mua.xlsx"),
                "sales_file": (invoice_file("C", 5, 100), "ban.xlsx"),
                "mapping": "{}",
                "policy": '{"max_loss_percent": 10, "max_profit_percent": 25}',
            },
            content_type="multipart/form-data",
        )
        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        self.assertEqual(payload["summary"]["material_quantity"], 3)
        self.assertEqual(payload["summary"]["finished_quantity"], 2)
        self.assertEqual(payload["sale_only_codes"], [])
        self.assertEqual(payload["allocations"][0]["used"][0]["variant_code"], "C.001")
        self.assertEqual(payload["allocations"][0]["used"][0]["sale_variant_code"], "C")
        self.assertEqual(payload["allocations"][0]["used"][0]["quantity"], 3)
        self.assertEqual(payload["allocations"][0]["invoice_no"], "HD1")
        self.assertEqual(payload["allocations"][0]["invoice_date"], "02/01/2025")
        self.assertEqual(payload["allocations"][0]["used"][0]["invoice_no"], "HD1")
        self.assertEqual(payload["allocations"][0]["used"][0]["invoice_date"], "02/01/2025")
        self.assertEqual(payload["allocations"][0]["sale_split_codes"], "C.001")
        self.assertNotIn("allocation_summaries", payload)
        self.assertTrue(payload["verification"])
        self.assertTrue(all(row["status"] == "OK" for row in payload["verification"]))

        download = inventory_app.app.test_client().get(f"/api/download/{payload['job_id']}")
        self.assertEqual(download.status_code, 200)
        result = load_workbook(BytesIO(download.data))
        self.assertIn("TongHopKho", result.sheetnames)
        self.assertIn("KiemTraDoiChieu", result.sheetnames)
        self.assertNotIn("TongHopNhapXuatTon", result.sheetnames)
        self.assertIn("TongHopNXT_KVT", result.sheetnames)
        self.assertIn("TongHopNXT_KTP", result.sheetnames)
        self.assertIn("BaoCaoBH_KVT", result.sheetnames)
        self.assertIn("BaoCaoBH_KTP", result.sheetnames)
        self.assertIn("BangKeHDBH_KVT", result.sheetnames)
        self.assertIn("BangKeHDBH_KTP", result.sheetnames)
        self.assertNotIn("PhanBoKho", result.sheetnames)
        self.assertNotIn("TonKhoHangHoa", result.sheetnames)
        self.assertIn("SoChiTietKVT", result.sheetnames)
        self.assertIn("SoChiTietKTP", result.sheetnames)
        self.assertIn("SoChiTietHH_TP", result.sheetnames)
        self.assertEqual(result["SoChiTietKVT"].cell(1, 1).value, "Ngày")
        self.assertEqual(result["SoChiTietHH_TP"].cell(1, 7).value, "Mã VT chi tiết")
        self.assertEqual(result["TongHopNXT_KVT"].cell(1, 1).value, "TỔNG HỢP NHẬP XUẤT TỒN")
        self.assertEqual(result["TongHopNXT_KTP"].cell(1, 1).value, "TỔNG HỢP NHẬP XUẤT TỒN")
        self.assertEqual(result["TongHopNXT_KVT"].cell(2, 1).value, "KHO: KVT - KHO VẬT TƯ, HÀNG HÓA")
        self.assertEqual(result["TongHopNXT_KTP"].cell(2, 1).value, "KHO: KTP - KHO THÀNH PHẨM")
        self.assertEqual(result["BaoCaoBH_KVT"].cell(1, 1).value, "BÁO CÁO TỔNG HỢP BÁN HÀNG")
        self.assertEqual(result["BaoCaoBH_KVT"].cell(5, 8).value, "TIỀN LÃI/LỖ")
        self.assertEqual(result["BaoCaoBH_KVT"].cell(5, 9).value, "% LÃI/LỖ")
        self.assertEqual(result["BaoCaoBH_KVT"].cell(7, 2).value, "Tổng cộng:")
        self.assertEqual(result["BangKeHDBH_KVT"].cell(1, 1).value, "BẢNG KÊ HOÁ ĐƠN BÁN HÀNG")
        verification_statuses = [
            result["KiemTraDoiChieu"].cell(row, 7).value
            for row in range(2, result["KiemTraDoiChieu"].max_row + 1)
        ]
        self.assertTrue(verification_statuses)
        self.assertTrue(all(status == "OK" for status in verification_statuses))
        download.close()

    def test_async_analyze_job_reports_progress_and_result(self):
        client = inventory_app.app.test_client()
        response = client.post(
            "/api/analyze-job",
            data={
                "purchase_file": (invoice_file("D.001", 3, 90), "mua.xlsx"),
                "sales_file": (invoice_file("D", 4, 120), "ban.xlsx"),
                "mapping": "{}",
                "policy": "{}",
            },
            content_type="multipart/form-data",
        )
        self.assertEqual(response.status_code, 200)
        analysis_job_id = response.get_json()["analysis_job_id"]

        status = None
        for _ in range(30):
            poll = client.get(f"/api/analyze-job/{analysis_job_id}")
            self.assertEqual(poll.status_code, 200)
            status = poll.get_json()
            self.assertIn("progress", status)
            if status["status"] == "complete":
                break
            time.sleep(0.1)

        self.assertEqual(status["status"], "complete")
        self.assertEqual(status["progress"], 100)
        self.assertEqual(status["result"]["summary"]["material_quantity"], 3)
        self.assertEqual(status["result"]["summary"]["finished_quantity"], 1)

    def test_new_application_uses_separate_port(self):
        self.assertEqual(inventory_app.APP_PORT, 5082)

    def test_preview_returns_excel_letters_with_selected_header_labels(self):
        preview = preview_workbook(invoice_file("Z", 2, 100).getvalue(), header_row=2)

        columns = {column["letter"]: column["header"] for column in preview["columns"]}
        self.assertEqual(preview["header_row"], 2)
        self.assertEqual(columns["L"], "Ma VT")
        self.assertEqual(columns["D"], "Ngay hoa don")
        self.assertEqual(columns["O"], "So luong")
        self.assertEqual(columns["P"], "Don gia")

    def test_read_lines_auto_detects_up_sales_layout(self):
        rows = read_lines(up_sales_file("HP.OTONMC13X26X1.2X6.0", 2142, 18590.91).getvalue(), clean_mapping({}, "sales"), "sales", company_profile="son_phuong")[1]

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["variant_code"], "HP.OTONMC13X26X1.2X6.0")
        self.assertEqual(rows[0]["quantity"], 2142)
        self.assertEqual(rows[0]["unit_price"], 18590.91)
        self.assertEqual(rows[0]["invoice_no"], "HDSP1")
        self.assertEqual(rows[0]["invoice_date_iso"], "2026-01-02")

    def test_read_lines_normalizes_dates_to_display_and_iso(self):
        rows = read_lines(invoice_file("Z", 2, 100).getvalue(), clean_mapping({}, "sales"), "sales")[1]

        self.assertEqual(rows[0]["invoice_date"], "02/01/2025")
        self.assertEqual(rows[0]["invoice_date_iso"], "2025-01-02")

    def test_son_phuong_read_lines_strips_company_prefix_from_product_code(self):
        purchase = read_lines(
            invoice_file("HA.OTMT21X2X1X6.0", 2, 100).getvalue(),
            clean_mapping({}, "purchase"),
            "purchase",
            company_profile="son_phuong",
        )[1]
        sales = read_lines(
            invoice_file("CD.THMK150X150X2.5X6000", 3, 120).getvalue(),
            clean_mapping({}, "sales"),
            "sales",
            company_profile="son_phuong",
        )[1]

        self.assertEqual(purchase[0]["original_variant_code"], "HA.OTMT21X2X1X6.0")
        self.assertEqual(purchase[0]["variant_code"], "HA.OTMT21X2X1X6.0")
        self.assertEqual(sales[0]["original_variant_code"], "CD.THMK150X150X2.5X6000")
        self.assertEqual(sales[0]["variant_code"], "CD.THMK150X150X2.5X6000")

    def test_son_phuong_read_lines_normalizes_steel_profile_across_company_codes(self):
        purchase_a = read_lines(
            invoice_file_with_name("CD.THMK150X150X2.5X6000", "Thep hop ma kem 150x150x2.5x6000", 5, 18000).getvalue(),
            clean_mapping({}, "purchase"),
            "purchase",
            company_profile="son_phuong",
        )[1][0]
        purchase_b = read_lines(
            invoice_file_with_name("TPHY.THMK150X150X2.5X6000", "Hop ma kem 150x150x2.5x6000", 7, 18100).getvalue(),
            clean_mapping({}, "purchase"),
            "purchase",
            company_profile="son_phuong",
        )[1][0]
        pipe = read_lines(
            invoice_file_with_name("HA.OTMT212X21X6.0", "Ong thep ma tron 21,2 x 2,1 x 6.0", 3, 22000).getvalue(),
            clean_mapping({}, "purchase"),
            "purchase",
            company_profile="son_phuong",
        )[1][0]

        self.assertEqual(purchase_a["variant_code"], "CD.HOP.MK.150X150X2.5")
        self.assertEqual(purchase_b["variant_code"], "TPHY.HOP.MK.150X150X2.5")
        self.assertEqual(purchase_a["source_variant_code"], "THMK150X150X2.5X6000")
        self.assertEqual(pipe["variant_code"], "HA.ONG.MK.21.2X2.1")
        self.assertEqual(pipe["steel_profile_key"], "pipe|galvanized|21.2x2.1")

    def test_son_phuong_exact_sale_matches_purchase_by_normalized_profile(self):
        purchase = read_lines(
            invoice_file_with_name("CD.THMK150X150X2.5X6000", "Thep hop ma kem 150x150x2.5x6000", 5, 10000).getvalue(),
            clean_mapping({}, "purchase"),
            "purchase",
            company_profile="son_phuong",
        )[1]
        sales = read_lines(
            invoice_file_with_name("TP.THMK150X150X2.5X6000", "Hop ma kem 150x150x2.5x6000", 3, 12000).getvalue(),
            clean_mapping({}, "sales"),
            "sales",
            company_profile="son_phuong",
        )[1]

        allocations, _, summary, _ = allocate_stock(
            [], purchase, sales, {"company_profile": "son_phuong"}, barem_map=DEFAULT_BAREM_MAP
        )

        self.assertEqual(allocations[0]["used"][0]["variant_code"], "CD.HOP.MK.150X150X2.5")
        self.assertEqual(summary["negative_export_quantity"], 0)
        self.assertEqual(summary["material_quantity"], 3)

    def test_son_phuong_non_steel_sale_matches_purchase_by_normalized_product_family(self):
        purchase = read_lines(
            invoice_file_with_name("TH.BANMA14", "Thep ban ma 14mm", 20, 9000).getvalue(),
            clean_mapping({}, "purchase"),
            "purchase",
            company_profile="son_phuong",
        )[1]
        sales = read_lines(
            invoice_file_with_name("BANM14MM", "Thep ban ma 14mm", 5, 12000).getvalue(),
            clean_mapping({}, "sales"),
            "sales",
            company_profile="son_phuong",
        )[1]

        allocations, _, summary, _ = allocate_stock(
            [], purchase, sales, {"company_profile": "son_phuong"}, barem_map=DEFAULT_BAREM_MAP
        )
        report_rows = inventory_app.build_sales_report_rows(allocations, purchase, sales)

        self.assertEqual(allocations[0]["used"][0]["variant_code"], "TH.BANMA14")
        self.assertEqual(summary["negative_export_quantity"], 0)
        self.assertEqual(report_rows[0]["cost_amount"], 45000)
        self.assertEqual(report_rows[0]["profit_amount"], 15000)

    def test_ignores_placeholder_zero_code_rows_from_old_formatter_output(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.cell(2, 12, "Ma VT")
        sheet.cell(2, 15, "So luong")
        sheet.cell(3, 12, 0)
        sheet.cell(3, 15, 5)
        sheet.cell(4, 12, "A.001")
        sheet.cell(4, 15, 3)
        content = BytesIO()
        workbook.save(content)

        _, rows = read_lines(content.getvalue(), clean_mapping({}, "purchase"), "purchase")

        self.assertEqual([(row["variant_code"], row["quantity"]) for row in rows], [("A.001", 3.0)])

    def test_normalizes_vietnamese_item_codes_like_formatter_before_matching(self):
        purchase = invoice_file("Ống.001", 3, 95)
        sales = invoice_file("ống", 2, 100)
        purchase_rows = read_lines(purchase.getvalue(), clean_mapping({}, "purchase"), "purchase")[1]
        sales_rows = read_lines(sales.getvalue(), clean_mapping({}, "sales"), "sales")[1]

        allocations, _, _, _ = allocate_stock([], purchase_rows, sales_rows)

        self.assertEqual(purchase_rows[0]["variant_code"], "ONG.001")
        self.assertEqual(sales_rows[0]["variant_code"], "ONG")
        self.assertEqual(allocations[0]["used"][0]["variant_code"], "ONG.001")


if __name__ == "__main__":
    unittest.main()
