import json
import math
import os
import random
import re
import subprocess
import sys
import threading
import time
import traceback
import unicodedata
import uuid
import webbrowser
from collections import defaultdict
from datetime import date, datetime
from io import BytesIO
from pathlib import Path

from flask import Flask, jsonify, request, send_file, send_from_directory
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import column_index_from_string, get_column_letter

try:
    import pystray
    from PIL import Image, ImageDraw
except Exception:
    pystray = None
    Image = None
    ImageDraw = None


if getattr(sys, "frozen", False):
    RESOURCE_DIR = Path(getattr(sys, "_MEIPASS", "."))
    BASE_DIR = Path(os.environ.get("LOCALAPPDATA") or Path.home()) / "InventoryAllocator"
else:
    RESOURCE_DIR = Path(__file__).resolve().parent
    BASE_DIR = RESOURCE_DIR

STATIC_DIR = RESOURCE_DIR / "static"
ASSET_DIR = RESOURCE_DIR / "assets"
ICON_PATH = ASSET_DIR / "black_coffee.ico"
OUTPUT_DIR = BASE_DIR / "outputs"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
DETECTION_RULE_CACHE_PATH = BASE_DIR / "company_detection_rules.json"

app = Flask(__name__, static_folder=None)
ANALYSIS_JOBS = {}
ANALYSIS_LOCK = threading.Lock()

APP_VERSION = "0.0"
VARIANT_SUFFIX = re.compile(r"^(.*?)(?:\.(\d{3}))$")
DIAMETER_CHARS = "ØøΦφ⌀"
QUANTITY_EPSILON = 0.00000001
APP_PORT = 5082

DEFAULT_MAPPING = {
    "purchase": {
        "sheet": "",
        "header_row": 2,
        "data_start_row": 3,
        "invoice_col": "C",
        "date_col": "D",
        "code_col": "L",
        "product_col": "M",
        "qty_col": "O",
        "price_col": "P",
    },
    "sales": {
        "sheet": "",
        "header_row": 2,
        "data_start_row": 3,
        "invoice_col": "C",
        "date_col": "D",
        "code_col": "L",
        "product_col": "M",
        "qty_col": "O",
        "price_col": "P",
    },
    "opening": {
        "sheet": "",
        "header_row": 1,
        "data_start_row": 2,
        "invoice_col": "",
        "date_col": "",
        "code_col": "A",
        "product_col": "B",
        "qty_col": "C",
        "price_col": "D",
    },
}

DEFAULT_POLICY = {
    "max_loss_percent": None,
    "max_profit_percent": None,
    "ignore_sale_suffix": False,
    "allow_negative_export": True,
    "company_profile": "yen_thanh",
    "son_phuong_split_counts": {
        "pipe_box": 2,
        "box": 2,
        "pipe": 2,
    },
    "generic_split_variance_percent": 0,
    "barem_tolerance_percent": 5,
    "barem_remainder_max_kg": 10,
    "generic_min_take_quantity": None,
    "generic_min_type_count": 2,
    "generic_max_take_quantity": None,
    "allow_future_purchase_reorder": False,
    "future_purchase_window_days": 31,
}


def text(value):
    return "" if value is None else str(value).strip()


def first_present(*values):
    for value in values:
        if value is not None and text(value) != "":
            return value
    return None


def number(value):
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    raw = str(value).strip().replace(" ", "")
    if not raw:
        return None
    if "," in raw and "." in raw:
        if raw.rfind(",") > raw.rfind("."):
            raw = raw.replace(".", "").replace(",", ".")
        else:
            raw = raw.replace(",", "")
    elif "," in raw:
        raw = raw.replace(",", ".")
    raw = re.sub(r"[^0-9.\-]", "", raw)
    try:
        return float(raw)
    except ValueError:
        return None


def column_index(column):
    raw = text(column).upper()
    if not re.fullmatch(r"[A-Z]+", raw):
        raise ValueError(f"Cột Excel không hợp lệ: {column}")
    return column_index_from_string(raw)


def normalize_code(value):
    value = text(value).replace("Đ", "D").replace("đ", "d").replace("Ä", "D").replace("Ä‘", "d")
    value = unicodedata.normalize("NFD", value)
    value = "".join(char for char in value if unicodedata.category(char) != "Mn").upper()
    return re.sub(r"[^A-Z0-9./]+", "", value)


def strip_son_phuong_company_prefix(value):
    code = normalize_code(value)
    if "." not in code:
        return code
    prefix, remainder = code.split(".", 1)
    if re.fullmatch(r"[A-Z]{1,4}", prefix) and re.search(r"[A-Z]", remainder) and re.search(r"\d", remainder):
        return remainder
    return code


def son_phuong_company_prefix(value):
    code = normalize_code(value)
    if "." not in code:
        return ""
    prefix, remainder = code.split(".", 1)
    if re.fullmatch(r"[A-Z]{1,4}", prefix) and re.search(r"[A-Z]", remainder) and re.search(r"\d", remainder):
        return prefix
    return ""


def apply_company_prefix(prefix, code):
    prefix = normalize_code(prefix)
    code = normalize_code(code)
    if prefix and code and not code.startswith(f"{prefix}."):
        return f"{prefix}.{code}"
    return code


def son_phuong_product_family(product_name):
    normalized = normalize_match_text(product_name)
    compact = re.sub(r"[^A-Z0-9]+", "", normalized)
    if not compact:
        return ""
    thickness_match = re.search(r"(\d+(?:[.,]\d+)?)\s*MM\b", normalized)
    thickness = ""
    if thickness_match:
        thickness = thickness_match.group(1).replace(",", ".").rstrip("0").rstrip(".")
    if "THEPBANMA" in compact or "BANMA" in compact:
        return f"THEP_BAN_MA:{thickness}" if thickness else "THEP_BAN_MA"
    if "CUONCANNONG" in compact or ("THEPCUON" in compact and "NONG" in compact):
        return "THEP_CUON_CAN_NONG"
    c_shape_match = re.search(r"\bTHEP\s*C\s*[-.]?\s*(\d+(?:[.,]\d+)?)\b", normalized)
    if not c_shape_match:
        c_shape_match = re.search(r"^THEPC(\d+(?:[.,]\d+)?)$", compact)
    if c_shape_match:
        shape_size = canonical_dimension_token(c_shape_match.group(1))
        return f"THEP_C:{shape_size}" if shape_size else "THEP_C"
    if re.search(r"\bTHEP\s*C\b", normalized) or compact == "THEPC":
        return "THEP_C"
    shape_match = re.search(r"\bTHEP\s+(?:HINH\s+|GOC\s+)?([VUIHL])\s*[-.]?\s*(\d+(?:[.,]\d+)?)\b", normalized)
    if not shape_match:
        shape_match = re.search(r"^THEP(?:HINH|GOC)?([VUIHL])(\d+(?:[.,]\d+)?)$", compact)
    if shape_match:
        shape_kind = shape_match.group(1)
        shape_size = canonical_dimension_token(shape_match.group(2))
        return f"THEP_HINH_{shape_kind}:{shape_size}" if shape_size else f"THEP_HINH_{shape_kind}"
    return generated_code_from_product(product_name)


def son_phuong_line_match_keys(line):
    keys = set()
    for field in ("variant_code", "base_code", "source_variant_code", "original_variant_code"):
        value = normalize_code(line.get(field, ""))
        if value:
            keys.add(value)
            stripped = strip_son_phuong_company_prefix(value)
            if stripped:
                keys.add(stripped)
    family = son_phuong_product_family(line.get("product_name", ""))
    if family:
        keys.add(f"NAME:{family}")
    return keys


def steel_profile_code(profile_key):
    if not profile_key:
        return ""
    parts = str(profile_key).split("|", 2)
    if len(parts) != 3:
        return ""
    kind, coating, dimensions = parts
    kind_label = {"box": "HOP", "pipe": "ONG"}.get(kind, kind.upper())
    coating_label = {"galvanized": "MK", "black": "DEN"}.get(coating, coating.upper())
    return f"{kind_label}.{coating_label}.{dimensions.upper()}"


def normalize_match_text(value):
    value = text(value).replace("Đ", "D").replace("đ", "d").replace("Ä", "D").replace("Ä‘", "d")
    value = unicodedata.normalize("NFD", value)
    value = "".join(char for char in value if unicodedata.category(char) != "Mn")
    return value.upper()


def dimension_part_count(value):
    raw = normalize_match_text(value).replace(",", ".")
    matches = re.findall(r"\d+(?:\.\d+)?(?:\s*[Xx\*]\s*\d+(?:\.\d+)?){1,4}", raw)
    counts = [len(re.split(r"\s*[Xx\*]\s*", match)) for match in matches]
    return max(counts) if counts else 0


def canonical_dimension_token(value):
    number_value = number(value)
    if number_value is None:
        return ""
    return f"{number_value:.4f}".rstrip("0").rstrip(".")


def is_decimal_token(value):
    return "." in canonical_dimension_token(value)


def extract_dimension_tokens(value):
    raw = normalize_match_text(value).replace(",", ".")
    matches = re.findall(r"\d+(?:\.\d+)?(?:\s*[Xx\*]\s*\d+(?:\.\d+)?){1,4}", raw)
    best = []
    for match in matches:
        tokens = [canonical_dimension_token(part) for part in re.split(r"\s*[Xx\*]\s*", match)]
        tokens = [token for token in tokens if token]
        if len(tokens) >= len(best):
            best = tokens
    if len(best) >= 3:
        last = number(best[-1])
        if last in {6, 6000}:
            best = best[:-1]
    pipe_matches = re.findall(r"(?:^|[^A-Z0-9])(?:D|F|PHI|Ø|Φ)\s*(\d+(?:\.\d+)?)(?:\s*[Xx\*]\s*(\d+(?:\.\d+)?))?", raw)
    for diameter, thickness in pipe_matches:
        tokens = [canonical_dimension_token(diameter)]
        if thickness:
            tokens.append(canonical_dimension_token(thickness))
        if len(tokens) >= len(best):
            best = [token for token in tokens if token]
    return best


def steel_coating(product_name="", variant_code=""):
    normalized = normalize_match_text(f"{product_name} {variant_code}")
    normalized_name = normalize_match_text(product_name)
    compact = re.sub(r"\s+", "", normalized)
    if "DEN" in normalized:
        return "black"
    if "MA KEM" in normalized or "MAKEM" in compact or re.search(r"\bMK\b", normalized) or " KEM" in normalized or re.search(r"\bMA\b", normalized_name):
        return "galvanized"
    return "unknown"


def explicit_steel_coating(product_name="", variant_code=""):
    normalized = normalize_match_text(f"{product_name} {variant_code}")
    normalized_name = normalize_match_text(product_name)
    compact = re.sub(r"\s+", "", normalized)
    if "DEN" in normalized:
        return "black"
    if "MA KEM" in normalized or "MAKEM" in compact or re.search(r"\bMK\b", normalized) or " KEM" in normalized or re.search(r"\bMA\b", normalized_name):
        return "galvanized"
    return None


def steel_profile_key(product_name="", variant_code=""):
    kind = steel_kind(product_name, variant_code)
    if kind not in {"pipe", "box"}:
        return None
    product_tokens = extract_dimension_tokens(product_name)
    tokens = product_tokens or extract_dimension_tokens(variant_code)
    if not product_tokens:
        normalized_name = normalize_match_text(product_name)
        has_steel_context = (
            "THEP" in normalized_name
            or bool(re.search(r"\bONG\b", normalized_name))
            or bool(re.search(r"\bHOP\b", normalized_name))
            or "VUONG" in normalized_name
            or bool(re.search(r"\bCN\b", normalized_name))
            or any(char in text(product_name) for char in DIAMETER_CHARS)
        )
        if not has_steel_context:
            return None
    if kind == "pipe" and len(tokens) < 2:
        return None
    if kind == "box" and len(tokens) == 2 and (number(tokens[0]) or 0) >= 8:
        tokens = [tokens[0], tokens[0], tokens[1]]
    if kind == "box" and len(tokens) < 3:
        return None
    return f"{kind}|{steel_coating(product_name, variant_code)}|{'x'.join(tokens)}"


def steel_profile_summary(product_name="", variant_code=""):
    key = steel_profile_key(product_name, variant_code)
    if not key:
        return {
            "profile_key": "",
            "profile_code": "",
            "kind": steel_kind(product_name, variant_code),
            "coating": steel_coating(product_name, variant_code),
            "dimension": "",
        }
    kind, coating, dimension = key.split("|", 2)
    return {
        "profile_key": key,
        "profile_code": steel_profile_code(key),
        "kind": kind,
        "coating": coating,
        "dimension": dimension,
    }


def steel_kind_detail(product_name="", variant_code=""):
    raw = text(product_name)
    normalized = normalize_match_text(raw)
    compact = re.sub(r"\s+", "", normalized)
    solid_square = (
        bool(re.search(r"\bTHEP\s+VUONG\b", normalized))
        and not re.search(r"\bONG\b", normalized)
        and not re.search(r"\bHOP\b", normalized)
    )
    if re.search(r"\bTHEP\s+(?:HINH\s+)?[UVCILH]\b", normalized) or solid_square:
        return "unknown", "Thép hình/vuông đặc không thuộc nhóm ống/hộp."
    has_ong_word = bool(re.search(r"\bONG\b", normalized))
    tokens = extract_dimension_tokens(raw) or extract_dimension_tokens(variant_code)
    explicit_pipe = (
        any(char in raw for char in DIAMETER_CHARS)
        or "PHI" in normalized
        or "TRON" in normalized
        or re.search(r"(^|[^A-Z])F\s*\d", normalized)
        or re.search(r"(^|[^A-Z])D\s*\d", normalized)
    )
    part_count = dimension_part_count(raw)
    if part_count >= 4:
        return "box", "Có kích thước 4 phần nên xác định là thép hộp."
    has_box_word = (
        "ONG HOP" in normalized
        or "ONGHOP" in compact
        or re.search(r"\bTHEP\s+HOP\b", normalized)
        or (
            re.search(r"\bHOP\b", normalized)
            and "HOP DONG" not in normalized
            and (tokens or "THEP" in normalized)
        )
    )
    if has_box_word or "VUONG" in normalized or re.search(r"\bCN\b", normalized):
        return "box", "Tên hàng có dấu hiệu hộp: ống hộp/hộp/vuông/CN."
    if explicit_pipe:
        return "pipe", "Tên hàng có dấu hiệu ống: Φ/Ø/F/phi/tròn."
    if has_ong_word:
        return "pipe", "Tên hàng có từ ống riêng biệt."
    if len(tokens) >= 2 and is_decimal_token(tokens[0]):
        return "pipe", "Không có chữ ống/hộp rõ, nhưng đường kính/cạnh đầu là số thập phân nên ưu tiên xác định là ống tròn."
    if len(tokens) >= 3:
        return "box", "Có đủ 2 cạnh và độ dày nên xác định là thép hộp."
    if part_count >= 2 and len(tokens) >= 2 and not is_decimal_token(tokens[0]) and (number(tokens[0]) or 0) >= 8:
        return "box", "Không có chữ ống/hộp rõ, nhưng cạnh đầu là số nguyên nên ưu tiên xác định là thép hộp."
    return "unknown", "Không có dấu hiệu đủ rõ để xếp vào ống/hộp."


def steel_kind(product_name="", variant_code=""):
    return steel_kind_detail(product_name, variant_code)[0]


def generic_steel_sale_type(product_name):
    normalized = normalize_match_text(product_name)
    compact = re.sub(r"\s+", "", normalized)
    if "THEP" not in normalized:
        return None
    is_explicit_generic = "CAC LOAI" in normalized
    has_dimensions = bool(extract_dimension_tokens(product_name))
    is_dimensionless_generic = not has_dimensions and (
        "ONG HOP" in normalized
        or "ONGHOP" in compact
        or bool(re.search(r"\bHOP\b", normalized))
        or bool(re.search(r"\bONG\b", normalized))
    )
    if not is_explicit_generic and not is_dimensionless_generic:
        return None
    if "ONG HOP" in normalized or "ONGHOP" in compact:
        return "pipe_box"
    if "HOP" in normalized:
        return "box"
    if re.search(r"\bONG\b", normalized):
        return "pipe"
    return None


def generic_allowed_kinds(generic_type):
    return {
        "pipe_box": {"pipe", "box"},
        "box": {"box"},
        "pipe": {"pipe"},
    }.get(generic_type, set())


def son_phuong_sales_pair_map(policy):
    defaults = {
        "materials": {"ma_kho": "KHHVT", "tk_vat_tu": "156"},
        "finished_goods": {"ma_kho": "KTP", "tk_vat_tu": "155"},
        "fallback": {"ma_kho": "KHOCK", "tk_vat_tu": "159"},
    }
    rows = policy.get("sales_inventory_pairs") if isinstance(policy, dict) else []
    if not isinstance(rows, list):
        rows = []
    result = {key: dict(value) for key, value in defaults.items()}
    for row in rows:
        if not isinstance(row, dict):
            continue
        role = text(row.get("role"))
        if role not in result:
            continue
        ma_kho = normalize_code(row.get("ma_kho"))
        tk_vat_tu = text(row.get("tk_vat_tu"))
        if not ma_kho or not tk_vat_tu:
            raise ValueError(f"Vai tro {role} phai co du Ma kho va TK vat tu.")
        result[role] = {"ma_kho": ma_kho, "tk_vat_tu": tk_vat_tu}
    return result


def son_phuong_sales_role(line):
    product_name = line.get("product_name", "")
    normalized = normalize_match_text(product_name)
    compact = re.sub(r"[^A-Z0-9]+", "", normalized)
    generic_type = generic_steel_sale_type(product_name)
    steel_kind_value = steel_kind(product_name, line.get("source_variant_code") or line.get("variant_code", ""))
    always_material_tokens = (
        "THEPCUON",
        "CUONCANNONG",
        "TONMAMAU",
        "TONMAU",
        "TONMAT",
        "TONLANH",
        "TONCACHNHIET",
    )
    if generic_type or any(token in compact for token in always_material_tokens):
        return "materials"
    structural_tokens = (
        "THEPHINH",
        "THEPTAM",
        "BANMA",
        "NEPDET",
        "THEPGOC",
        "THANH",
        "DAM",
        "RAY",
        "PHITRON",
        "VUONGDAC",
    )
    structural_pattern = re.search(r"\bTHEP\s+(?:HINH\s+)?[UVCILH](?=\s|[-.]|\d|$)", normalized)
    solid_square = (
        "THEPVUONG" in compact
        and not re.search(r"\bONG\b", normalized)
        and not re.search(r"\bHOP\b", normalized)
    )
    if structural_pattern or solid_square or any(token in compact for token in structural_tokens):
        return "finished_goods"
    if steel_kind_value in {"pipe", "box"}:
        return "materials"
    return "fallback"
def normalize_header(value):
    value = text(value).replace("Đ", "D").replace("đ", "d").replace("Ä", "D").replace("Ä‘", "d")
    value = unicodedata.normalize("NFD", value)
    value = "".join(char for char in value if unicodedata.category(char) != "Mn").lower()
    return re.sub(r"[^a-z0-9]+", " ", value).strip()


def valid_inventory_code(value):
    return normalize_code(value) not in {"", "0", "0.0"}


def rm_accents_for_code(value):
    value = text(value).replace("Đ", "D").replace("đ", "d").replace("Ä", "D").replace("Ä‘", "d")
    for char in DIAMETER_CHARS:
        value = value.replace(char, "F")
    value = unicodedata.normalize("NFD", value)
    return "".join(char for char in value if unicodedata.category(char) != "Mn")


def normalize_code_token(value, keep_slash=False):
    pattern = r"[^A-Z0-9./]+" if keep_slash else r"[^A-Z0-9.]+"
    return re.sub(pattern, "", rm_accents_for_code(value).upper())


def normalize_sep(value):
    return re.sub(r"(?<=\d)\s*([xX*])\s*(?=\d)", "x", text(value))


def code_words(value):
    return [word for word in re.split(r"\s+", normalize_sep(value).strip()) if word]


def remove_duplicate_phrases(words, phrases):
    seen = set()
    result = []
    normalized_phrases = {rm_accents_for_code(phrase).casefold().strip() for phrase in phrases}
    for word in words:
        key = rm_accents_for_code(word).casefold().strip()
        if key in normalized_phrases:
            if key in seen:
                continue
            seen.add(key)
        result.append(word)
    return result


def normalize_cao_thanh_inox_grade_words(words):
    result = []
    for word in words:
        current = normalize_code_token(word)
        previous = normalize_code_token(result[-1]) if result else ""
        if current == "304" and previous == "INOX":
            continue
        result.append(word)
    return result


def normalize_cao_thanh_con_reducer_words(words):
    if not words:
        return words
    first = normalize_code_token(words[0])
    last = normalize_code_token(words[-1], keep_slash=True)
    if first != "CON" or not re.search(r"\d+(?:[.,]\d+)?/\d+", last):
        return words
    if any(normalize_code_token(word) == "THU" for word in words):
        return words
    return [words[0], "thu", *words[1:]]


def generated_code_from_product(product_name):
    words = remove_duplicate_phrases(code_words(product_name), ["inox"])
    words = normalize_cao_thanh_inox_grade_words(words)
    words = normalize_cao_thanh_con_reducer_words(words)
    first = [normalize_code_token(word, keep_slash=True) for word in words[:2]]
    rest = []
    for word in words[2:]:
        compact = normalize_code_token(word, keep_slash=True)
        rest.append(compact if re.search(r"\d", word) else compact[:1])
    return re.sub(r"\.+", ".", "".join(first + rest)).strip(". ")


def code_parts(value):
    code = normalize_code(value)
    match = VARIANT_SUFFIX.match(code)
    if not match:
        return code, code, None
    return code, match.group(1), int(match.group(2))


def clean_mapping(raw, key):
    source = raw.get(key) if isinstance(raw, dict) else {}
    source = source if isinstance(source, dict) else {}
    result = dict(DEFAULT_MAPPING[key])
    result.update({field: value for field, value in source.items() if value is not None})
    result["header_row"] = max(1, int(result.get("header_row") or 1))
    result["data_start_row"] = max(result["header_row"] + 1, int(result.get("data_start_row") or result["header_row"] + 1))
    for field in ("code_col", "qty_col"):
        column_index(result[field])
    for field in ("invoice_col", "date_col", "product_col", "price_col"):
        if text(result.get(field)):
            column_index(result[field])
    return result


def optional_percent(value):
    if value is None or text(value) == "":
        return None
    result = number(value)
    if result is None or result < 0:
        raise ValueError("Khoảng lãi/lỗ phải là số không âm.")
    return result


def optional_quantity(value, field_name):
    if value is None or text(value) == "":
        return None
    result = number(value)
    if result is None or result < 0:
        raise ValueError(f"{field_name} phai la so khong am.")
    return result


def clean_policy(raw):
    raw = raw if isinstance(raw, dict) else {}
    ignore_sale_suffix = str(raw.get("ignore_sale_suffix", "")).lower() in {"1", "true", "yes", "on"}
    allow_negative_export = str(raw.get("allow_negative_export", DEFAULT_POLICY["allow_negative_export"])).lower() in {"1", "true", "yes", "on"}
    allow_future_purchase_reorder = str(raw.get("allow_future_purchase_reorder", "")).lower() in {"1", "true", "yes", "on"}
    company_profile = text(raw.get("company_profile") or DEFAULT_POLICY["company_profile"]).lower()
    if company_profile not in {"yen_thanh", "son_phuong"}:
        company_profile = "yen_thanh"
    max_loss_percent = optional_percent(raw.get("max_loss_percent"))
    max_profit_percent = optional_percent(raw.get("max_profit_percent"))
    split_counts_raw = raw.get("son_phuong_split_counts") if isinstance(raw.get("son_phuong_split_counts"), dict) else {}
    split_counts = {}
    remainder_max_kg = optional_quantity(raw.get("barem_remainder_max_kg"), "Kg le toi da")
    if remainder_max_kg is None:
        remainder_max_kg = DEFAULT_POLICY["barem_remainder_max_kg"]
    for key, default_value in DEFAULT_POLICY["son_phuong_split_counts"].items():
        try:
            split_counts[key] = max(1, int(number(split_counts_raw.get(key, default_value)) or default_value))
        except (TypeError, ValueError):
            split_counts[key] = default_value
    variance = optional_percent(raw.get("generic_split_variance_percent"))
    if variance is None:
        variance = DEFAULT_POLICY["generic_split_variance_percent"]
    barem_tolerance = optional_percent(raw.get("barem_tolerance_percent"))
    if barem_tolerance is None:
        barem_tolerance = DEFAULT_POLICY["barem_tolerance_percent"]
    min_take_quantity = optional_quantity(raw.get("generic_min_take_quantity"), "Khoi luong nho nhat")
    max_take_quantity = optional_quantity(raw.get("generic_max_take_quantity"), "Khoi luong lon nhat")
    if min_take_quantity is not None and max_take_quantity is not None and min_take_quantity > max_take_quantity:
        raise ValueError("Khoi luong nho nhat khong duoc lon hon khoi luong lon nhat.")
    try:
        min_type_count = max(1, int(number(raw.get("generic_min_type_count", DEFAULT_POLICY["generic_min_type_count"])) or DEFAULT_POLICY["generic_min_type_count"]))
    except (TypeError, ValueError):
        min_type_count = DEFAULT_POLICY["generic_min_type_count"]
    try:
        future_purchase_window_days = int(number(raw.get("future_purchase_window_days", DEFAULT_POLICY["future_purchase_window_days"])) or DEFAULT_POLICY["future_purchase_window_days"])
    except (TypeError, ValueError):
        future_purchase_window_days = DEFAULT_POLICY["future_purchase_window_days"]
    future_purchase_window_days = max(1, min(future_purchase_window_days, 366))
    try:
        scenario_count = max(1, min(1000, int(number(raw.get("scenario_count", 100)) or 100)))
    except (TypeError, ValueError):
        scenario_count = 100
    sales_inventory_pairs = son_phuong_sales_pair_map(raw) if company_profile == "son_phuong" else []

    if company_profile == "son_phuong":
        ignore_sale_suffix = False
        allow_negative_export = False
        allow_future_purchase_reorder = False
    if ignore_sale_suffix:
        max_loss_percent = None
        max_profit_percent = None
    return {
        "max_loss_percent": max_loss_percent,
        "max_profit_percent": max_profit_percent,
        "ignore_sale_suffix": ignore_sale_suffix,
        "allow_negative_export": allow_negative_export,
        "company_profile": company_profile,
        "son_phuong_split_counts": split_counts,
        "generic_split_variance_percent": variance,
        "barem_tolerance_percent": barem_tolerance,
        "generic_min_take_quantity": min_take_quantity,
        "barem_remainder_max_kg": remainder_max_kg,
        "generic_max_take_quantity": max_take_quantity,
        "generic_min_type_count": min_type_count,
        "allow_future_purchase_reorder": allow_future_purchase_reorder,
        "future_purchase_window_days": future_purchase_window_days,
        "sales_inventory_pairs": sales_inventory_pairs,
        "scenario_count": scenario_count,
    }


def policy_active(policy):
    return policy["max_loss_percent"] is not None or policy["max_profit_percent"] is not None


def normalize_sales_codes_to_base(sales_lines):
    normalized = []
    for line in sales_lines:
        item = dict(line)
        item["variant_code"] = item.get("base_code") or item.get("variant_code", "")
        item["suffix"] = None
        normalized.append(item)
    return normalized


def sheet_for_mapping(workbook, mapping):
    name = text(mapping.get("sheet"))
    if name and name in workbook.sheetnames:
        return workbook[name]
    return workbook[workbook.sheetnames[0]]


def header_column(sheet, header_row, labels, fallback=None):
    targets = {normalize_header(label) for label in labels}
    for column in range(1, (sheet.max_column or 128) + 1):
        if normalize_header(sheet.cell(header_row, column).value) in targets:
            return column
    return column_index(fallback) if fallback else None


def header_column_contains(sheet, header_row, labels):
    targets = [normalize_header(label) for label in labels]
    for column in range(1, (sheet.max_column or 128) + 1):
        header = normalize_header(sheet.cell(header_row, column).value)
        if header and any(target and target in header for target in targets):
            return column
    return None


def detect_sales_mapping(sheet, mapping):
    current_header = normalize_header(sheet.cell(mapping["header_row"], column_index(mapping["code_col"])).value)
    if current_header in {"ma vt", "ma vat tu"}:
        return mapping
    for header_row in range(1, min(sheet.max_row, 5) + 1):
        code_col = header_column_contains(sheet, header_row, ["Mã vật tư", "Mã VT", "ma_vt"])
        qty_col = header_column_contains(sheet, header_row, ["Số lượng", "so_luong"])
        if not code_col or not qty_col:
            continue
        result = dict(mapping)
        result["header_row"] = header_row
        result["data_start_row"] = header_row + 1
        result["code_col"] = get_column_letter(code_col)
        result["qty_col"] = get_column_letter(qty_col)
        invoice_col = header_column_contains(sheet, header_row, ["Số hóa đơn", "Số chứng từ", "so_ct"])
        date_col = header_column_contains(sheet, header_row, ["Ngày hóa đơn", "Ngày chứng từ", "ngay_ct"])
        product_col = header_column_contains(sheet, header_row, ["Tên hàng hóa", "Tên hàng", "Diễn giải", "dien_giai"])
        price_col = header_column_contains(sheet, header_row, ["Đơn giá", "Giá bán", "gia2"])
        if invoice_col:
            result["invoice_col"] = get_column_letter(invoice_col)
        if date_col:
            result["date_col"] = get_column_letter(date_col)
        if product_col:
            result["product_col"] = get_column_letter(product_col)
        if price_col:
            result["price_col"] = get_column_letter(price_col)
        return result
    return mapping


def detect_opening_mapping(sheet, mapping):
    for header_row in range(1, 11):
        code_col = header_column_contains(sheet, header_row, ["M\u00e3 VT", "M\u00e3 v\u1eadt t\u01b0", "M\u00e3 h\u00e0ng", "ma_vt"])
        qty_col = header_column_contains(sheet, header_row, [
            "S\u1ed1 l\u01b0\u1ee3ng \u0111\u1ea7u k\u1ef3", "T\u1ed3n \u0111\u1ea7u k\u1ef3", "SL t\u1ed3n \u0111\u1ea7u", "S\u1ed1 l\u01b0\u1ee3ng t\u1ed3n", "S\u1ed1 l\u01b0\u1ee3ng",
        ])
        if not code_col or not qty_col:
            continue
        result = dict(mapping)
        result["header_row"] = header_row
        result["data_start_row"] = header_row + 1
        result["code_col"] = get_column_letter(code_col)
        result["qty_col"] = get_column_letter(qty_col)
        product_col = header_column_contains(sheet, header_row, ["T\u00ean h\u00e0ng", "T\u00ean v\u1eadt t\u01b0", "T\u00ean s\u1ea3n ph\u1ea9m", "ten_hang"])
        price_col = header_column_contains(sheet, header_row, [
            "\u0110\u01a1n gi\u00e1 v\u1ed1n", "Gi\u00e1 v\u1ed1n", "\u0110\u01a1n gi\u00e1 t\u1ed3n", "\u0110\u01a1n gi\u00e1", "don_gia",
        ])
        if product_col:
            result["product_col"] = get_column_letter(product_col)
        if price_col:
            result["price_col"] = get_column_letter(price_col)
        return result
    raise ValueError(
        "File t?n ??u k? kh?ng t?m th?y c?t M? VT v? S? l??ng. "
        "H?y d?ng header theo t?n ho?c t?i file m?u t?n ??u k?."
    )


def cell_value(values, column):
    if not column:
        return None
    index = column - 1
    return values[index] if index < len(values) else None


def tax_rate_percent(value):
    result = number(value)
    if result is None:
        return None
    return result * 100 if abs(result) <= 1 else result


def parse_excel_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    raw = text(value)
    if not raw:
        return None
    raw = raw.split()[0]
    for date_format in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d.%m.%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(raw, date_format).date()
        except ValueError:
            pass
    return None


def date_iso(value):
    parsed = parse_excel_date(value)
    return parsed.isoformat() if parsed else ""


def date_display(value):
    parsed = parse_excel_date(value)
    return parsed.strftime("%d/%m/%Y") if parsed else text(value)


def read_lines(content, mapping, kind, company_profile="yen_thanh"):
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    sheet = sheet_for_mapping(workbook, mapping)
    if kind == "sales":
        mapping = detect_sales_mapping(sheet, mapping)
    if kind == "opening":
        mapping = detect_opening_mapping(sheet, mapping)
    code_col = column_index(mapping["code_col"])
    qty_col = column_index(mapping["qty_col"])
    invoice_col = column_index(mapping["invoice_col"]) if text(mapping.get("invoice_col")) else None
    date_col = column_index(mapping["date_col"]) if text(mapping.get("date_col")) else None
    product_col = column_index(mapping["product_col"]) if text(mapping.get("product_col")) else None
    price_col = column_index(mapping["price_col"]) if text(mapping.get("price_col")) else None
    header_row = mapping["header_row"]
    if kind == "purchase":
        party_name_col = header_column(sheet, header_row, ["Tên người bán", "Tên công ty", "Người bán"], "F")
        party_tax_col = header_column(sheet, header_row, ["MST người bán", "MST", "Mã số thuế người bán"], "G")
    elif kind == "sales":
        party_name_col = (
            header_column_contains(sheet, header_row, ["Tên người mua", "Người mua hàng", "ong_ba", "Tên khách hàng"])
            or header_column(sheet, header_row, ["Tên người mua", "Người mua hàng", "Người mua hàng (ong_ba)", "Tên khách hàng"], "B")
        )
        party_tax_col = (
            header_column_contains(sheet, header_row, ["MST người mua", "Mã ST", "Mã khách", "Mã số thuế người mua"])
            or header_column(sheet, header_row, ["MST người mua", "Mã ST", "Mã khách", "Mã số thuế người mua"], "A")
        )
    else:
        party_name_col = None
        party_tax_col = None
    amount_col = header_column(sheet, header_row, ["Tiền chưa thuế nguyên tệ", "Thành tiền mua", "Tiền bán", "Tiền bán:N0 (tien2)", "Thành tiền"], "S") if kind in {"purchase", "sales"} else None
    unit_col = header_column(sheet, header_row, ["Đơn vị tính", "ĐVT", "DVT"], "N") if kind in {"purchase", "sales"} else None
    warehouse_col = header_column(sheet, header_row, ["Mã kho", "Ma kho", "ma_kho"]) if kind in {"purchase", "sales", "opening"} else None
    warehouse_account_col = header_column(sheet, header_row, ["TK vật tư", "TK vat tu", "tk_vat_tu"]) if kind in {"purchase", "sales", "opening"} else None
    processing_group_col = header_column(sheet, header_row, ["Nh\u00f3m x\u1eed l\u00fd", "Nhom xu ly", "processing group", "processing_group"]) if company_profile == "son_phuong" and kind == "purchase" else None
    tax_rate_col = header_column(sheet, header_row, ["Thuế suất", "Thuế suất GTGT", "Thuế suất (thue_suat)"], "R") if kind == "sales" else None
    tax_amount_col = header_column(sheet, header_row, ["Tiền thuế nguyên tệ", "Tiền thuế", "Tiền thuế:N0 (tien_thue)"], "T") if kind == "sales" else None
    last_col = max(
        column for column in (
            code_col, qty_col, invoice_col, date_col, product_col, price_col,
            party_name_col, party_tax_col, amount_col, unit_col, warehouse_col, warehouse_account_col,
            tax_rate_col, tax_amount_col, processing_group_col,
        ) if column is not None
    )
    lines = []
    row_values = sheet.iter_rows(
        min_row=mapping["data_start_row"],
        max_col=last_col,
        values_only=True,
    )
    for row_number, values in enumerate(row_values, start=mapping["data_start_row"]):
        raw_code = cell_value(values, code_col)
        product_name = text(cell_value(values, product_col)) if product_col else ""
        if processing_group_col:
            processing_group = normalize_header(cell_value(values, processing_group_col)).replace(" ", "_")
            if processing_group and processing_group not in {"materials", "nhom_vat_tu"}:
                continue
        if (kind == "sales" or company_profile == "son_phuong") and not valid_inventory_code(raw_code) and product_name:
            raw_code = generated_code_from_product(product_name)
        quantity = number(cell_value(values, qty_col))
        if not valid_inventory_code(raw_code) or quantity is None or quantity <= 0:
            continue
        original_code = normalize_code(raw_code)
        source_code = original_code
        steel_profile = {"profile_key": "", "profile_code": "", "kind": "", "coating": "", "dimension": ""}
        if company_profile == "son_phuong":
            company_prefix = son_phuong_company_prefix(raw_code)
            raw_code = strip_son_phuong_company_prefix(raw_code)
            source_code = normalize_code(raw_code)
            steel_profile = steel_profile_summary(product_name, source_code)
            if steel_profile["profile_code"] and not (kind == "sales" and generic_steel_sale_type(product_name)):
                raw_code = apply_company_prefix(company_prefix, steel_profile["profile_code"])
            else:
                raw_code = original_code
        variant_code, base_code, suffix = code_parts(raw_code)
        unit_price = number(cell_value(values, price_col)) if price_col else None
        line_amount = number(cell_value(values, amount_col)) if amount_col else None
        if line_amount is None and unit_price is not None:
            line_amount = unit_price * quantity
        rate_percent = tax_rate_percent(cell_value(values, tax_rate_col)) if tax_rate_col else None
        tax_amount = number(cell_value(values, tax_amount_col)) if tax_amount_col else None
        if tax_amount is None and line_amount is not None and rate_percent is not None:
            tax_amount = line_amount * rate_percent / 100
        raw_invoice_date = cell_value(values, date_col) if date_col else None
        warehouse_code = normalize_warehouse_code(cell_value(values, warehouse_col)) if warehouse_col else ""
        warehouse_account = text(cell_value(values, warehouse_account_col)) if warehouse_account_col else ""
        if company_profile == "vietmax":
            if kind == "purchase":
                warehouse_code = MATERIAL_WAREHOUSE_CODE
                warehouse_account = default_warehouse_account(MATERIAL_WAREHOUSE_CODE)
            elif warehouse_code in LEGACY_MATERIAL_WAREHOUSE_CODES or warehouse_code == MATERIAL_WAREHOUSE_CODE:
                warehouse_code = MATERIAL_WAREHOUSE_CODE
                warehouse_account = default_warehouse_account(MATERIAL_WAREHOUSE_CODE)
        lines.append({
            "kind": kind,
            "row_number": row_number,
            "original_variant_code": original_code,
            "source_variant_code": source_code,
            "steel_profile_key": steel_profile.get("profile_key", ""),
            "steel_profile_code": steel_profile.get("profile_code", ""),
            "steel_kind": steel_profile.get("kind", ""),
            "steel_coating": steel_profile.get("coating", ""),
            "steel_dimension": steel_profile.get("dimension", ""),
            "variant_code": variant_code,
            "base_code": base_code,
            "suffix": suffix,
            "invoice_no": text(cell_value(values, invoice_col)) if invoice_col else "",
            "invoice_date": date_display(raw_invoice_date) if date_col else "",
            "invoice_date_iso": date_iso(raw_invoice_date) if date_col else "",
            "party_tax_code": text(cell_value(values, party_tax_col)) if party_tax_col else "",
            "party_name": text(cell_value(values, party_name_col)) if party_name_col else "",
            "product_name": product_name,
            "unit_name": text(cell_value(values, unit_col)) if unit_col else "",
            "warehouse_code": warehouse_code,
            "warehouse_account": warehouse_account,
            "quantity": quantity,
            "unit_price": unit_price,
            "line_amount": line_amount,
            "tax_rate_percent": rate_percent,
            "tax_amount": tax_amount,
        })
    workbook.close()
    return sheet.title, lines


def summarize_purchase_lines(purchase_lines):
    grouped = {}
    order = []
    for line in purchase_lines:
        key = (line["variant_code"], line.get("invoice_date_iso") or "")
        if key not in grouped:
            grouped[key] = {
                **line,
                "quantity": 0.0,
                "unit_price": None,
                "weighted_cost_total": 0.0,
                "weighted_cost_quantity": 0.0,
                "invoice_numbers": [],
                "invoice_dates": [],
                "row_numbers": [],
                "summary_count": 0,
            }
            order.append(key)
        item = grouped[key]
        item["quantity"] = clean_quantity(item["quantity"] + line["quantity"])
        item["summary_count"] += 1
        item["row_numbers"].append(line["row_number"])
        if line.get("invoice_no"):
            item["invoice_numbers"].append(line["invoice_no"])
        if line.get("invoice_date"):
            item["invoice_dates"].append(line["invoice_date"])
        if line.get("unit_price") is not None:
            item["weighted_cost_total"] += line["unit_price"] * line["quantity"]
            item["weighted_cost_quantity"] += line["quantity"]
        if not item.get("product_name") and line.get("product_name"):
            item["product_name"] = line["product_name"]
    result = []
    for key in order:
        item = grouped[key]
        if item["weighted_cost_quantity"] > QUANTITY_EPSILON:
            item["unit_price"] = item["weighted_cost_total"] / item["weighted_cost_quantity"]
        item["row_number"] = item["row_numbers"][0]
        invoices = list(dict.fromkeys(item["invoice_numbers"]))
        invoice_dates = list(dict.fromkeys(item["invoice_dates"]))
        item["invoice_no"] = ", ".join(str(value) for value in invoices[:3])
        if len(invoices) > 3:
            item["invoice_no"] += f" +{len(invoices) - 3}"
        item["invoice_date"] = ", ".join(str(value) for value in invoice_dates[:3])
        if len(invoice_dates) > 3:
            item["invoice_date"] += f" +{len(invoice_dates) - 3}"
        item.pop("weighted_cost_total")
        item.pop("weighted_cost_quantity")
        result.append(item)
    return result


def make_lots(opening_lines, purchase_lines):
    lots = []
    for sequence, line in enumerate(opening_lines + purchase_lines):
        lot = dict(line)
        lot.update({
            "source": "Tồn đầu kỳ" if line["kind"] == "opening" else "Mua vào",
            "sequence": sequence,
            "initial_quantity": line["quantity"],
            "remaining_quantity": line["quantity"],
            "allocated_quantity": 0.0,
        })
        lots.append(lot)
    return lots


def sale_sort_key(sale, include_invoice=False):
    sale_date = sale.get("invoice_date_iso") or "9999-12-31"
    row_number = sale.get("row_number", 0)
    if not include_invoice:
        return (sale_date, row_number)
    invoice_parts = tuple(
        (0, int(part)) if part.isdigit() else (1, part.casefold())
        for part in re.split(r"(\d+)", text(sale.get("invoice_no")))
        if part
    )
    return (sale_date, invoice_parts, row_number)


def lot_available_for_sale(lot, sale):
    if lot.get("kind") == "opening":
        return True
    lot_date = lot.get("invoice_date_iso")
    sale_date = sale.get("invoice_date_iso")
    if not lot_date or not sale_date:
        return True
    return lot_date <= sale_date


def parse_iso_date(value):
    if not value:
        return None
    try:
        return datetime.strptime(str(value), "%Y-%m-%d").date()
    except (TypeError, ValueError):
        return None


def future_purchase_days(lot, sale):
    lot_date = parse_iso_date(lot.get("invoice_date_iso"))
    sale_date = parse_iso_date(sale.get("invoice_date_iso"))
    if not lot_date or not sale_date or lot_date <= sale_date:
        return None
    return (lot_date - sale_date).days


def future_lot_allowed_for_sale(lot, sale, policy):
    if not policy.get("allow_future_purchase_reorder") or lot.get("kind") == "opening":
        return False
    days = future_purchase_days(lot, sale)
    if days is None:
        return False
    return days <= int(policy.get("future_purchase_window_days") or DEFAULT_POLICY["future_purchase_window_days"])


def lot_priority(lot, sale_price):
    suffix = lot["suffix"] if lot["suffix"] is not None else 999999
    if lot["unit_price"] is None or sale_price is None:
        return (2, suffix, lot["sequence"])
    profit_per_unit = sale_price - lot["unit_price"]
    if profit_per_unit >= 0:
        return (0, profit_per_unit, suffix, lot["sequence"])
    return (1, abs(profit_per_unit), suffix, lot["sequence"])


def generic_lot_priority(lot, sale_price):
    return generic_lot_priority_for_quantity(lot, sale_price, lot_unreserved_quantity(lot))


def generic_lot_priority_for_quantity(lot, sale_price, available_quantity):
    suffix = lot["suffix"] if lot["suffix"] is not None else 999999
    stock_priority = -available_quantity
    if lot["unit_price"] is None or sale_price is None:
        return (2, stock_priority, suffix, lot["sequence"])
    profit_per_unit = sale_price - lot["unit_price"]
    if profit_per_unit >= 0:
        return (0, profit_per_unit, stock_priority, suffix, lot["sequence"])
    return (1, abs(profit_per_unit), stock_priority, suffix, lot["sequence"])


def generic_lot_quality_rank(lot, sale_price):
    if lot.get("unit_price") is None or sale_price is None:
        return 2
    return 0 if sale_price - lot["unit_price"] >= 0 else 1


def lot_margin_percent(lot, sale_price):
    if lot["unit_price"] is None or sale_price is None or sale_price <= 0:
        return None
    return (sale_price - lot["unit_price"]) / sale_price * 100


def lot_acceptance(lot, sale_price, policy):
    if not policy_active(policy):
        return True, ""
    margin_percent = lot_margin_percent(lot, sale_price)
    if margin_percent is None:
        return False, "không có đủ đơn giá để kiểm tra khoảng lãi/lỗ"
    if policy["max_loss_percent"] is not None and margin_percent < -policy["max_loss_percent"]:
        return False, f"lỗ {format_number(abs(margin_percent))}% vượt mức {format_number(policy['max_loss_percent'])}%"
    if policy["max_profit_percent"] is not None and margin_percent > policy["max_profit_percent"]:
        return False, f"lãi {format_number(margin_percent)}% vượt mức {format_number(policy['max_profit_percent'])}%"
    return True, ""


def lot_unreserved_quantity(lot):
    return clean_quantity(max(
        0,
        clean_quantity(lot.get("remaining_quantity", 0))
        - clean_quantity(lot.get("specific_reserved_quantity", 0)),
    ))


def son_phuong_sale_reservation_key(sale):
    return (
        sale.get("invoice_date_iso") or "",
        text(sale.get("invoice_no")),
        sale.get("row_number", 0),
        normalize_code(sale.get("variant_code", "")),
    )
def son_phuong_compatible_profile_lots(by_steel_profile, profile_key):
    """Return exact-surface lots first, then unknown-surface fallback lots."""
    parts = text(profile_key).split("|", 2)
    if len(parts) != 3:
        return []
    kind, coating, dimension = parts
    compatible_keys = [profile_key]
    if coating in {"black", "galvanized"}:
        compatible_keys.append(f"{kind}|unknown|{dimension}")
    elif coating == "unknown":
        compatible_keys.extend((f"{kind}|black|{dimension}", f"{kind}|galvanized|{dimension}"))
    result = []
    seen = set()
    for key in compatible_keys:
        for lot in by_steel_profile.get(key, []):
            lot_id = id(lot)
            if lot_id in seen:
                continue
            seen.add(lot_id)
            result.append(lot)
    return result


def son_phuong_surface_priority(lot, requested_profile_key):
    parts = text(requested_profile_key).split("|", 2)
    requested_coating = parts[1] if len(parts) == 3 else "unknown"
    lot_coating = lot.get("steel_coating") or "unknown"
    return 0 if lot_coating == requested_coating else 1




def reserve_son_phuong_specific_sales(ordered_sales, by_steel_profile, policy):
    """Reserve dated stock for exact pipe/box rows before generic rows plan."""
    reservations = defaultdict(list)
    for sale in ordered_sales:
        if generic_steel_sale_type(sale.get("product_name", "")):
            continue
        if son_phuong_sales_role(sale) != "materials":
            continue
        profile_key = sale.get("steel_profile_key") or steel_profile_key(
            sale.get("product_name", ""),
            sale.get("source_variant_code") or sale.get("variant_code", ""),
        )
        if not profile_key:
            continue
        quantity_left = clean_quantity(sale.get("quantity", 0))
        if quantity_left <= QUANTITY_EPSILON:
            continue
        candidates = []
        for lot in son_phuong_compatible_profile_lots(by_steel_profile, profile_key):
            available = clean_quantity(
                lot.get("initial_quantity", 0)
                - lot.get("specific_reserved_quantity", 0)
            )
            if available <= QUANTITY_EPSILON or not lot_available_for_sale(lot, sale):
                continue
            accepted, _reason = lot_acceptance(lot, sale.get("unit_price"), policy)
            if accepted:
                candidates.append(lot)
        candidates.sort(key=lambda lot: (
            son_phuong_surface_priority(lot, profile_key),
            lot_priority(lot, sale.get("unit_price")),
        ))
        reservation_key = son_phuong_sale_reservation_key(sale)
        for lot in candidates:
            if quantity_left <= QUANTITY_EPSILON:
                break
            available = clean_quantity(
                lot.get("initial_quantity", 0)
                - lot.get("specific_reserved_quantity", 0)
            )
            amount = clean_quantity(min(quantity_left, available))
            if amount <= QUANTITY_EPSILON:
                continue
            lot["specific_reserved_quantity"] = clean_quantity(
                lot.get("specific_reserved_quantity", 0) + amount
            )
            reservations[reservation_key].append((lot, amount))
            quantity_left = clean_quantity(quantity_left - amount)
    return reservations

def lot_within_generic_variance(lot, selected_lots, variance_percent):
    if not variance_percent or not selected_lots:
        return True
    unit_price = lot.get("unit_price")
    if unit_price is None or unit_price <= 0:
        return True
    prices = [item.get("unit_price") for item in selected_lots if item.get("unit_price")]
    if not prices:
        return True
    low = min(prices + [unit_price])
    high = max(prices + [unit_price])
    if low <= 0:
        return True
    return (high - low) / low * 100 <= variance_percent


def son_phuong_lot_pool(lots, sale):
    generic_type = generic_steel_sale_type(sale.get("product_name", ""))
    if not generic_type:
        return None, []
    allowed = generic_allowed_kinds(generic_type)
    sale_coating = explicit_steel_coating(sale.get("product_name", ""), sale.get("source_variant_code") or sale.get("variant_code", ""))
    return generic_type, [
        lot for lot in lots
        if (lot.get("steel_kind") or steel_kind(lot.get("product_name", ""), lot.get("source_variant_code") or lot.get("variant_code", ""))) in allowed
        and (not sale_coating or (lot.get("steel_coating") or steel_coating(lot.get("product_name", ""), lot.get("source_variant_code") or lot.get("variant_code", ""))) in {sale_coating, "unknown"})
    ]


BUILTIN_BAREM_TABLE = """
box black 12x12 0.8:1.66 0.9:1.85
box black 14x14 0.9:2.19 1:2.41 1.1:2.63 1.2:2.84 1.4:3.25
box black 16x16 0.9:2.53 1.2:3.29 1.4:3.78 1.5:4.01
box black 20x20 0.9:3.21 1.2:4.20 1.4:4.83 1.8:6.05 2:6.63
box black 25x25 0.8:3.62 0.9:4.06 1:4.48 1.1:4.91 1.2:5.33 1.4:6.15 1.8:7.75 2:8.52
box black 30x30 1:5.43 1.1:5.94 1.2:6.46 1.4:7.47 1.8:9.44 2:10.40 2.5:12.72 3:14.92
box black 40x40 1.1:8.02 1.4:10.11 1.5:10.80 1.8:12.83 2:14.17 2.5:17.43 2.8:19.33 3:20.57
box black 50x50 1.4:12.74 1.5:13.62 1.8:16.22 2:17.94 2.3:20.47 2.5:22.14 2.8:24.60 3:26.23 4:34.06
box black 60x60 2:21.70 2.5:26.85 2.8:29.88 3:31.88
box black 75x75 1.8:24.70 2:27.36 2.5:33.91 3:40.36 4:52.90
box black 90x90 2:33.01 2.5:40.98 3:48.83 4:64.21
box black 100x100 2.5:45.69 3:54.49 4:71.74 4.5:80.20 5:88.55
box black 200x200 4:147.10 4.5:164.98 5:182.75 6:217.94 8:286.97
box black 13x26 0.8:2.79 0.9:3.12 1.1:3.77 1.2:4.08 1.4:4.70
box black 20x40 1.2:6.46 1.4:7.47 1.8:9.44 2:10.40 2.5:12.72
box black 25x50 1.1:7.50 1.2:8.15 1.4:9.45 1.5:10.09 1.8:11.98 2:13.23 2.5:16.25
box black 30x60 1.4:11.43 1.5:12.21 1.8:14.53 2:16.05 2.3:18.30 2.5:19.78 3:23.40
box black 40x80 1.2:13.24 1.4:15.38 1.8:19.61 2:21.70 2.3:24.80 2.5:26.85 2.8:29.88 3:31.88 4:41.56
box black 50x100 1.4:19.34 1.8:24.70 2:27.36 2.3:31.30 2.5:33.91 2.8:37.79 3:40.36 4:52.90
box black 60x120 1.8:29.79 2:33.01 2.3:37.80 2.5:40.98 2.8:45.70 3:48.83 4:64.21
box black 150x150 2.5:69.24 2.8:77.36 3:82.75 3.5:96.14 4:109.42 4.5:122.59
box black 100x150 2.5:57.46 2.8:64.17 3:68.62 3.5:79.66 4:90.58 4.5:101.40
box black 100x200 2.5:69.24 2.8:77.36 3:82.75 3.5:96.14 4:109.42 4.5:122.59
box black 250x250 4:184.78 5:229.85 6:274.46 8:362.33 10:448.39
box black 200x300 5:229.85 6:274.46
box galvanized 12x12 0.8:1.66
box galvanized 14x14 1:2.41 1.1:2.63 1.2:2.84 1.4:3.25
box galvanized 16x16 0.9:2.53 1.2:3.29 1.4:3.78
box galvanized 20x20 1:3.54 1.1:3.87 1.2:4.20 1.4:4.83 1.8:6.05 2:6.63
box galvanized 25x25 1:4.48 1.1:4.91 1.2:5.33 1.4:6.15 1.8:7.75 2:8.52
box galvanized 30x30 1:5.43 1.1:5.94 1.2:6.46 1.4:7.47 1.8:9.44 2:10.40
box galvanized 40x40 1.1:8.02 1.2:8.72 1.4:10.11 1.5:10.80 1.8:12.83 2:14.17
box galvanized 50x50 1.2:10.98 1.4:12.74 1.5:13.62 1.8:16.22 2:17.94
box galvanized 60x60 1.4:15.38 1.8:19.61 2:21.70
box galvanized 75x75 1.4:19.34 1.8:24.70 2:27.36
box galvanized 90x90 1.4:23.30 1.8:29.79 2:33.01
box galvanized 100x100 1.8:33.18 2:36.78 2.3:42.14
box galvanized 13x26 0.8:2.79 0.9:3.12 1:3.45 1.1:3.77 1.2:4.08 1.4:4.70
box galvanized 20x40 0.9:4.90 1:5.43 1.1:5.94 1.2:6.46 1.4:7.47 1.5:7.97 1.8:9.44 2:10.40
box galvanized 25x50 1:6.84 1.2:8.15 1.4:9.45 1.5:10.09 1.8:11.98 2:13.23
box galvanized 30x60 0.9:7.45 1:8.25 1.1:9.05 1.2:9.85 1.4:11.43 1.5:12.21 1.8:14.53 2:16.05 2.3:18.30
box galvanized 30x90 1.2:13.24 1.4:15.38 1.8:19.61
box galvanized 40x80 1.1:12.16 1.2:13.24 1.4:15.38 1.5:16.45 1.8:19.61 2:21.70 2.3:24.80
box galvanized 50x100 1.2:16.63 1.4:19.34 1.5:20.69 1.8:24.70 2:27.36 2.3:31.30
box galvanized 60x120 1.4:23.30 1.8:29.79 2:33.01 2.3:37.80 2.5:40.98
pipe black 21.2 1:2.99 1.2:3.55 1.4:4.10 1.5:4.37 2:5.68 2.2:6.19 2.5:6.92 2.6:7.16
pipe black 26.65 1.2:4.52 1.5:5.58 1.8:6.62 2:7.29 2.5:8.93 2.6:9.25
pipe black 33.5 1.5:7.10 1.8:8.44 2:9.32 2.5:11.47 2.6:11.89 2.9:13.13 3:13.54 3.2:14.35 3.5:15.54
pipe black 42.2 1.4:8.45 1.5:9.03 1.8:10.76 2:11.90 2.3:13.58 2.5:14.69 2.6:15.24 2.8:16.32 2.9:16.86 3:17.40 3.2:18.47 3.5:20.04
pipe black 48.1 1.5:10.34 2:13.64 2.3:15.59 2.5:16.87 2.6:17.50 2.9:19.40 3:20.02 3.2:21.26 3.5:23.10 3.8:24.91 4:26.10 4.5:29.03 5:31.89
pipe black 59.9 1.4:12.12 1.5:12.96 1.8:15.47 2:17.13 2.3:19.60 2.5:21.23 2.6:22.04 2.9:24.46 3:25.26 3.2:26.85 3.5:29.21 3.6:29.99 3.8:31.54 4:33.09 4.5:36.89 5:40.62
pipe black 75.6 1.8:19.66 2:21.78 2.3:24.95 2.5:27.04 3:32.23 3.2:34.28 3.5:37.34 3.6:38.35 3.8:40.37 4:42.38 4.5:47.34 5:52.23
pipe black 88.3 2:25.54 2.3:29.27 2.5:31.74 3:37.87 3.2:40.30 3.5:43.92 3.6:45.12 4:49.90 4.5:55.80 5:61.63 5.5:67.39 6:73.07
pipe black 113.5 1.8:29.75 2:33.00 2.3:37.84 2.5:41.06 3:49.05 3.2:52.23 3.5:56.97 3.6:58.54 4:64.81 4.5:72.58 5:80.27 5.5:87.89 6:95.44
pipe black 141.3 2.5:51.35 3:61.39 3.96:80.46 4.78:96.54 5:100.84 5.16:103.95 5.56:111.66 6.35:126.79 6.55:130.62
pipe black 168.3 2.5:61.33 3:73.38 3.96:96.30 4.78:115.62 5:120.82 5.16:124.56 5.56:133.86 6.35:152.16
pipe black 219.1 3.96:126.06 4.78:151.56 5.16:163.32 5.56:175.68 6.35:199.86 8.18:255.30
pipe black 273 4.78:189.72 5.16:204.48 5.56:220.02 6.35:250.50 9.27:361.74 10.31:478.20 11.13:514.92
pipe black 323.8 5.16:243.30 5.56:261.78 6.35:298.26 7.92:370.14
pipe galvanized 21.2 1:2.99 1.1:3.27 1.2:3.55 1.4:4.10 1.5:4.37 1.9:5.43 2.1:5.94 2.5:6.92
pipe galvanized 26.65 1:3.80 1.1:4.16 1.2:4.52 1.4:5.23 1.5:5.58 1.9:6.96 2.1:7.63
pipe galvanized 33.5 0.9:13.56 1:4.81 1.1:5.27 1.2:5.74 1.4:6.65 1.5:7.10 1.9:8.88 2.1:9.76 2.3:10.62
pipe galvanized 42.2 1.2:7.28 1.4:8.45 1.5:9.03 1.8:10.76 2:11.90 2.3:13.58 2.5:14.69
pipe galvanized 48.1 1.4:9.67 1.5:10.34 1.8:12.33 2:13.64 2.3:15.59
pipe galvanized 59.9 1.4:12.12 1.5:12.96 1.8:15.47 2:17.14 2.3:19.60
pipe galvanized 75.6 1.4:15.37 1.5:16.45 1.8:19.66 2:21.78 2.3:24.95
pipe galvanized 88.3 1.4:18.00 1.5:19.27 1.8:23.04 2:25.54 2.3:29.27
pipe galvanized 113.5 1.8:29.75 2:33.00 2.3:37.84
"""


def profile_key_from_parts(kind, coating, dimension, thickness):
    dimension_tokens = [canonical_dimension_token(part) for part in str(dimension).split("x")]
    dimension_tokens.append(canonical_dimension_token(thickness))
    return f"{kind}|{coating}|{'x'.join(token for token in dimension_tokens if token)}"


def builtin_barem_map():
    by_profile = {}
    for raw_line in BUILTIN_BAREM_TABLE.strip().splitlines():
        parts = raw_line.split()
        if len(parts) < 4:
            continue
        kind, coating, dimension, *pairs = parts
        for pair in pairs:
            if ":" not in pair:
                continue
            thickness, weight = pair.split(":", 1)
            key = profile_key_from_parts(kind, coating, dimension, thickness)
            by_profile[key] = float(weight)
    return {"by_code": {}, "by_profile": by_profile}


def builtin_barem_rows():
    rows = []
    for raw_line in BUILTIN_BAREM_TABLE.strip().splitlines():
        parts = raw_line.split()
        if len(parts) < 4:
            continue
        kind, coating, dimension, *pairs = parts
        for pair in pairs:
            if ":" not in pair:
                continue
            thickness, weight = pair.split(":", 1)
            rows.append({
                "kind": kind,
                "coating": coating,
                "dimension": dimension,
                "thickness": canonical_dimension_token(thickness),
                "weight": float(weight),
            })
    return rows


def merge_barem_maps(base_map, override_map=None):
    result = {
        "by_code": dict((base_map or {}).get("by_code") or {}),
        "by_profile": dict((base_map or {}).get("by_profile") or {}),
    }
    if override_map:
        result["by_code"].update(override_map.get("by_code") or {})
        result["by_profile"].update(override_map.get("by_profile") or {})
    return result


DEFAULT_BAREM_MAP = builtin_barem_map()


def parse_barem_file(content):
    if not content:
        return {"by_code": {}, "by_profile": {}}

    def barem_number_cell(value):
        if isinstance(value, (int, float)):
            return float(value)
        raw = text(value)
        if not raw or not re.fullmatch(r"[-+]?\d[\d.,\s]*", raw):
            return None
        return number(raw)

    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    result = {"by_code": {}, "by_profile": {}}
    try:
        for row in sheet.iter_rows(values_only=True):
            if not row:
                continue
            values = [cell for cell in row if cell is not None and text(cell) != ""]
            if not values:
                continue
            text_values = [text(cell) for cell in values if text(cell)]
            numeric_values = [(index, barem_number_cell(cell)) for index, cell in enumerate(row) if barem_number_cell(cell) is not None and barem_number_cell(cell) > 0]
            weight = None
            if len(row) > 1 and barem_number_cell(row[1]) is not None and barem_number_cell(row[1]) > 0:
                weight = barem_number_cell(row[1])
            elif numeric_values:
                weight = numeric_values[-1][1]
            if weight is None or weight <= 0:
                continue
            code = normalize_code(text_values[0] if text_values else "")
            product_name = max(text_values, key=len, default="")
            if code and code not in {"0", "0.0"}:
                result["by_code"][code] = weight
            profile_key = steel_profile_key(product_name, code)
            if profile_key:
                result["by_profile"][profile_key] = weight
    finally:
        workbook.close()
    return result


def theoretical_barem_weight(profile_key):
    parts = text(profile_key).split("|", 2)
    if len(parts) != 3:
        return None
    kind, _coating, dimension = parts
    values = [number(token) for token in dimension.split("x")]
    if any(value is None or value <= 0 for value in values):
        return None
    if kind == "pipe" and len(values) == 2:
        diameter, thickness = values
        if thickness * 2 >= diameter:
            return None
        area_mm2 = math.pi * (diameter * thickness - thickness * thickness)
    elif kind == "box" and len(values) == 3:
        width, height, thickness = values
        if thickness * 2 >= min(width, height):
            return None
        area_mm2 = 2 * thickness * (width + height - 2 * thickness)
    else:
        return None
    # 7.85 g/cm3, one standard bar is 6 metres.
    return area_mm2 * 0.0471


def barem_profile_parts(profile_key):
    parts = text(profile_key).split("|", 2)
    if len(parts) != 3:
        return None
    dimensions = [number(token) for token in parts[2].split("x")]
    if any(value is None for value in dimensions):
        return None
    return parts[0], parts[1], dimensions


def compatible_barem_profile_keys(profile_key):
    parts = text(profile_key).split("|", 2)
    if len(parts) != 3:
        return []
    kind, coating, dimension = parts
    keys = [profile_key]
    if coating == "unknown":
        keys.extend((f"{kind}|galvanized|{dimension}", f"{kind}|black|{dimension}"))
    elif coating in {"black", "galvanized"}:
        keys.append(f"{kind}|unknown|{dimension}")
    return keys


def infer_barem_weight(by_profile, profile_key):
    target = barem_profile_parts(profile_key)
    if not target:
        return None
    kind, coating, dimensions = target
    thickness = dimensions[-1]
    same_base = []
    calibration = []
    for candidate_key, raw_weight in (by_profile or {}).items():
        weight = number(raw_weight)
        candidate = barem_profile_parts(candidate_key)
        if weight is None or weight <= 0 or not candidate:
            continue
        candidate_kind, candidate_coating, candidate_dimensions = candidate
        if candidate_kind != kind or candidate_coating != coating or len(candidate_dimensions) != len(dimensions):
            continue
        if all(abs(left - right) <= QUANTITY_EPSILON for left, right in zip(candidate_dimensions[:-1], dimensions[:-1])):
            same_base.append((candidate_dimensions[-1], weight, candidate_key))
        theoretical = theoretical_barem_weight(candidate_key)
        if theoretical and theoretical > QUANTITY_EPSILON:
            distance = sum(abs(left - right) / max(abs(right), 1) for left, right in zip(candidate_dimensions, dimensions))
            calibration.append((distance, weight / theoretical, candidate_key))

    same_base.sort(key=lambda item: item[0])
    if len(same_base) >= 2:
        lower = max((item for item in same_base if item[0] <= thickness), default=same_base[0], key=lambda item: item[0])
        upper = min((item for item in same_base if item[0] >= thickness), default=same_base[-1], key=lambda item: item[0])
        if abs(upper[0] - lower[0]) <= QUANTITY_EPSILON:
            lower, upper = sorted(sorted(same_base, key=lambda item: abs(item[0] - thickness))[:2], key=lambda item: item[0])
        if abs(upper[0] - lower[0]) > QUANTITY_EPSILON:
            ratio = (thickness - lower[0]) / (upper[0] - lower[0])
            inferred = lower[1] + ratio * (upper[1] - lower[1])
            if inferred > QUANTITY_EPSILON:
                return {"weight": round(inferred, 6), "source": "inferred", "method": "N?i suy/ngo?i suy theo ?? d?y c?ng k?ch th??c", "references": [lower[2], upper[2]], "confidence": "high" if lower[0] <= thickness <= upper[0] else "medium"}
    if len(same_base) == 1 and same_base[0][0] > QUANTITY_EPSILON:
        inferred = same_base[0][1] * thickness / same_base[0][0]
        if inferred > QUANTITY_EPSILON:
            return {"weight": round(inferred, 6), "source": "inferred", "method": "T? l? theo ?? d?y c?ng k?ch th??c", "references": [same_base[0][2]], "confidence": "medium"}

    theoretical_target = theoretical_barem_weight(profile_key)
    if theoretical_target and calibration:
        nearest = sorted(calibration, key=lambda item: item[0])[:12]
        ratios = sorted(item[1] for item in nearest)
        middle = len(ratios) // 2
        correction = ratios[middle] if len(ratios) % 2 else (ratios[middle - 1] + ratios[middle]) / 2
        inferred = theoretical_target * correction
        if inferred > QUANTITY_EPSILON:
            return {"weight": round(inferred, 6), "source": "inferred", "method": "Suy ra t? ti?t di?n v? hi?u ch?nh theo c?c barem g?n nh?t", "references": [item[2] for item in nearest[:3]], "confidence": "medium"}
    return None


def resolve_barem_details(barem_map, lot):
    if not barem_map:
        return None
    code = normalize_code(lot.get("variant_code", ""))
    if not isinstance(barem_map, dict) or not ("by_code" in barem_map or "by_profile" in barem_map):
        direct = barem_map.get(code) if hasattr(barem_map, "get") else None
        return {"weight": direct, "source": "code", "method": "Barem theo M? VT", "references": [code], "confidence": "exact"} if direct is not None else None
    by_code = barem_map.get("by_code") or {}
    by_profile = barem_map.get("by_profile") or {}
    direct = by_code.get(code)
    if direct is not None:
        return {"weight": direct, "source": "code", "method": "Barem theo M? VT", "references": [code], "confidence": "exact"}
    profile_key = lot.get("steel_profile_key") or steel_profile_key(lot.get("product_name", ""), lot.get("source_variant_code") or lot.get("variant_code", ""))
    for compatible_key in compatible_barem_profile_keys(profile_key):
        profile_weight = by_profile.get(compatible_key)
        if profile_weight is not None:
            return {
                "weight": profile_weight,
                "source": "profile",
                "method": "Barem theo khoa ky thuat",
                "references": [compatible_key],
                "confidence": "exact" if compatible_key == profile_key else "high",
            }
    for compatible_key in compatible_barem_profile_keys(profile_key):
        inferred = infer_barem_weight(by_profile, compatible_key)
        if inferred:
            return inferred
    return None


def resolve_barem_weight(barem_map, lot):
    if not barem_map:
        return None
    code = normalize_code(lot.get("variant_code", ""))
    if isinstance(barem_map, dict) and ("by_code" in barem_map or "by_profile" in barem_map):
        by_code = barem_map.get("by_code") or {}
        by_profile = barem_map.get("by_profile") or {}
        direct = by_code.get(code)
        if direct is not None:
            return direct
        profile_key = lot.get("steel_profile_key") or steel_profile_key(
            lot.get("product_name", ""),
            lot.get("source_variant_code") or lot.get("variant_code", ""),
        )
        details = resolve_barem_details(barem_map, lot)
        return details.get("weight") if details else None
    return barem_map.get(code) if hasattr(barem_map, "get") else None


def steel_barem_ambiguity(line):
    product_name = line.get("product_name", "")
    variant_code = line.get("source_variant_code") or line.get("variant_code", "")
    kind, kind_reason = steel_kind_detail(product_name, variant_code)
    tokens = extract_dimension_tokens(product_name) or extract_dimension_tokens(variant_code)
    normalized = normalize_match_text(product_name)
    has_relevant_hint = (
        bool(re.search(r"\bONG\b", normalized))
        or bool(re.search(r"\bHOP\b", normalized))
        or "VUONG" in normalized
        or bool(re.search(r"\bCN\b", normalized))
    )
    has_dimension_pattern = bool(re.search(r"\d+(?:[.,]\d+)?\s*[Xx\*]\s*\d+", normalized))
    if kind == "unknown" and (has_relevant_hint or has_dimension_pattern):
        return {
            "kind": "unknown",
            "dimension": "x".join(tokens),
            "reason": "Không xác định rõ là ống tròn hay hộp/CN/vuông.",
            "kind_reason": kind_reason,
        }
    return None


def build_ambiguous_steel_rows(purchase_lines):
    rows = []
    for line in purchase_lines:
        ambiguity = steel_barem_ambiguity(line)
        if not ambiguity:
            continue
        rows.append({
            "company": line.get("party_name") or "(Không có tên công ty)",
            "tax_code": line.get("party_tax_code", ""),
            "variant_code": line.get("source_variant_code") or line.get("variant_code", ""),
            "profile_code": line.get("steel_profile_code", ""),
            "product_name": line.get("product_name", ""),
            "quantity": line.get("quantity", 0),
            "unit_price": line.get("unit_price"),
            "invoice_no": line.get("invoice_no", ""),
            "invoice_date": line.get("invoice_date", ""),
            "row_number": line.get("row_number", ""),
            "detected_kind": ambiguity["kind"],
            "dimension": ambiguity["dimension"],
            "reason": ambiguity["reason"],
            "kind_reason": ambiguity["kind_reason"],
        })
    rows.sort(key=lambda row: (normalize_match_text(row["company"]), row.get("row_number") or 0, row.get("variant_code", "")))
    return rows


def build_company_detection_rules(purchase_lines):
    companies = {}
    for line in purchase_lines:
        company = line.get("party_name") or "(Không có tên công ty)"
        item = companies.setdefault(company, {
            "company": company,
            "tax_code": line.get("party_tax_code", ""),
            "total": 0,
            "pipe": 0,
            "box": 0,
            "unknown": 0,
            "profiles": {},
            "examples": [],
        })
        kind = line.get("steel_kind") or steel_kind(line.get("product_name", ""), line.get("source_variant_code") or line.get("variant_code", ""))
        item["total"] += 1
        item[kind if kind in {"pipe", "box"} else "unknown"] += 1
        profile_code = line.get("steel_profile_code") or ""
        if profile_code:
            profile = item["profiles"].setdefault(profile_code, {
                "profile_code": profile_code,
                "profile_key": line.get("steel_profile_key", ""),
                "kind": kind,
                "coating": line.get("steel_coating", ""),
                "dimension": line.get("steel_dimension", ""),
                "count": 0,
                "quantity": 0,
                "examples": [],
            })
            profile["count"] += 1
            profile["quantity"] = clean_quantity(profile["quantity"] + clean_quantity(line.get("quantity", 0)))
            if len(profile["examples"]) < 3:
                profile["examples"].append({
                    "row_number": line.get("row_number", ""),
                    "source_variant_code": line.get("source_variant_code") or line.get("variant_code", ""),
                    "product_name": line.get("product_name", ""),
                })
        if len(item["examples"]) < 8:
            item["examples"].append({
                "row_number": line.get("row_number", ""),
                "source_variant_code": line.get("source_variant_code") or line.get("variant_code", ""),
                "profile_code": profile_code,
                "product_name": line.get("product_name", ""),
                "kind": kind,
                "reason": steel_kind_detail(line.get("product_name", ""), line.get("source_variant_code") or line.get("variant_code", ""))[1],
            })
    result = []
    for item in companies.values():
        profiles = sorted(
            item.pop("profiles").values(),
            key=lambda row: (-row["quantity"], row["profile_code"]),
        )
        item["profile_count"] = len(profiles)
        item["top_profiles"] = profiles[:12]
        result.append(item)
    result.sort(key=lambda row: (-row["total"], normalize_match_text(row["company"])))
    return result


def save_company_detection_rules(profile, rules):
    if profile != "son_phuong":
        return
    try:
        cache = {}
        if DETECTION_RULE_CACHE_PATH.exists():
            cache = json.loads(DETECTION_RULE_CACHE_PATH.read_text(encoding="utf-8"))
        cache[profile] = {
            "updated_at": datetime.now().isoformat(timespec="seconds"),
            "rules": rules,
        }
        DETECTION_RULE_CACHE_PATH.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")
    except Exception:
        pass


def missing_barem_item(lot, reason="Thiếu khối lượng 1 barem"):
    profile = steel_profile_summary(lot.get("product_name", ""), lot.get("source_variant_code") or lot.get("variant_code", ""))
    return {
        "variant_code": lot.get("variant_code", ""),
        "product_name": lot.get("product_name", ""),
        "quantity": clean_quantity(lot.get("remaining_quantity", 0)),
        "invoice_no": lot.get("invoice_no", ""),
        "invoice_date": lot.get("invoice_date", ""),
        "row_number": lot.get("row_number", ""),
        "reason": reason,
        "profile_key": profile.get("profile_key", ""),
        "steel_kind": profile.get("kind", ""),
        "steel_coating": profile.get("coating", ""),
        "steel_dimension": profile.get("dimension", ""),
        "status": "unresolved",
    }


def deterministic_random(seed_parts):
    seed = "|".join(text(part) for part in seed_parts)
    return random.Random(seed)


def random_percent_options(option_count, split_count, seed_parts):
    rng = deterministic_random(seed_parts)
    options = []
    for _ in range(option_count):
        weights = [rng.random() + 0.05 for _ in range(split_count)]
        total = sum(weights) or 1
        options.append([weight / total for weight in weights])
    return options


def barem_multiple_quantity(target, available, barem, min_quantity=None, max_quantity=None):
    if barem is None or barem <= QUANTITY_EPSILON:
        return 0
    limit = min(target, available)
    if max_quantity is not None:
        limit = min(limit, max_quantity)
    amount = clean_quantity(int(limit / barem) * barem)
    if min_quantity is not None and amount + QUANTITY_EPSILON < min_quantity:
        return 0
    return amount


def choose_barem_generic_plan_legacy(sale, candidates, policy, barem_map):
    if not candidates:
        return [], [], [], "Không có mã ứng viên."
    missing_barem = []
    usable = []
    seen_missing = set()
    for lot in candidates:
        code = normalize_code(lot.get("variant_code", ""))
        barem = resolve_barem_weight(barem_map, lot)
        if barem is None or barem <= 0:
            if code not in seen_missing:
                profile_key = lot.get("steel_profile_key") or steel_profile_key(
                    lot.get("product_name", ""),
                    lot.get("source_variant_code") or lot.get("variant_code", ""),
                )
                reason = "Thiếu khối lượng 1 barem"
                if profile_key:
                    reason += f" cho {profile_key.replace('|', ' ')}"
                else:
                    reason += "; không nhận diện được dimension để dò theo barem chung"
                missing_barem.append(missing_barem_item(lot, reason))
                seen_missing.add(code)
            continue
        lot["barem_weight"] = barem
        usable.append(lot)
    if not usable:
        return [], missing_barem, [], "Không có mã nào có barem hợp lệ."

    required = clean_quantity(sale.get("quantity", 0))
    if required <= QUANTITY_EPSILON:
        return [], missing_barem, [], "Số lượng bán ra bằng 0."
    min_take_quantity = policy.get("generic_min_take_quantity")
    max_take_quantity = policy.get("generic_max_take_quantity")

    def lot_capacity(lot):
        capacity = clean_quantity(lot.get("remaining_quantity", 0))
        if max_take_quantity is not None:
            capacity = min(capacity, max_take_quantity)
        return clean_quantity(capacity)

    usable = [
        lot for lot in usable
        if lot_capacity(lot) > QUANTITY_EPSILON
        and (min_take_quantity is None or lot_capacity(lot) + QUANTITY_EPSILON >= min_take_quantity)
    ]
    if not usable:
        return [], missing_barem, [], "Khong co ma nao dat gioi han khoi luong min/max."

    has_profitable_lots = any(generic_lot_quality_rank(lot, sale.get("unit_price")) == 0 for lot in usable)
    source_lots = usable
    best_plan = []
    planned_amounts = defaultdict(float)
    total = 0.0
    base_target = required / max(1, len(source_lots))
    for lot in source_lots:
        amount = barem_multiple_quantity(
            base_target,
            clean_quantity(lot_capacity(lot) - planned_amounts.get(id(lot), 0)),
            lot.get("barem_weight"),
            min_take_quantity,
            max_take_quantity,
        )
        if amount <= QUANTITY_EPSILON:
            continue
        best_plan.append((lot, amount, 0, False))
        planned_amounts[id(lot)] = clean_quantity(planned_amounts[id(lot)] + amount)
        total = clean_quantity(total + amount)
    for lot in source_lots:
        shortage = clean_quantity(required - total)
        if shortage <= QUANTITY_EPSILON:
            break
        amount = barem_multiple_quantity(
            shortage,
            clean_quantity(lot_capacity(lot) - planned_amounts.get(id(lot), 0)),
            lot.get("barem_weight"),
            min_take_quantity,
            max_take_quantity,
        )
        if amount <= QUANTITY_EPSILON:
            continue
        best_plan.append((lot, amount, 0, False))
        planned_amounts[id(lot)] = clean_quantity(planned_amounts[id(lot)] + amount)
        total = clean_quantity(total + amount)
    rounded_shortage = clean_quantity(max(0, required - total))
    if rounded_shortage > QUANTITY_EPSILON:
        def spare_capacity(lot):
            return clean_quantity(lot_capacity(lot) - planned_amounts.get(id(lot), 0))

        remainder_lot = None
        for lot in source_lots:
            if spare_capacity(lot) + QUANTITY_EPSILON >= rounded_shortage:
                remainder_lot = lot
                break
        if remainder_lot is not None:
            best_plan.append((remainder_lot, rounded_shortage, 0, True))
            rounded_shortage = 0
    selected_codes = [lot.get("variant_code", "") for lot, _amount, _ratio, _is_remainder in best_plan]
    split_count = len(dict.fromkeys(selected_codes))
    remainder_quantity = clean_quantity(sum(
        amount for _lot, amount, _ratio, is_remainder in best_plan if is_remainder
    ))
    if remainder_quantity > QUANTITY_EPSILON:
        note = f"Chon {split_count} ma khong gioi han so loai; {format_number(remainder_quantity)} kg du cuoi duoc gop vao ma phu hop nhat."
    elif rounded_shortage == 0:
        note = f"Chon {split_count} ma khong gioi han so loai; toan bo khoi luong dung boi so barem."
    else:
        note = f"Chon {split_count} ma khong gioi han so loai, thieu {format_number(rounded_shortage)} kg sau khi lam tron theo barem."
    if not has_profitable_lots:
        note = "Khong co ma lai hop le; buoc phai chon ma lo thap nhat. " + note
    top_lot = source_lots[0]
    split_count = 4 if lot_capacity(top_lot) >= required * 0.25 else 6
    selected_lots = source_lots[:split_count]
    if not selected_lots:
        return [], missing_barem, [], "Không đủ mã ứng viên có barem."
    if len(selected_lots) < split_count:
        split_count = len(selected_lots)
    tolerance = required * (policy.get("barem_tolerance_percent") or 0) / 100
    best_plan = None
    best_score = None
    for option in random_percent_options(policy.get("scenario_count") or 100, split_count, [
        sale.get("invoice_no", ""),
        sale.get("row_number", ""),
        sale.get("variant_code", ""),
        split_count,
    ]):
        plan = []
        total = 0.0
        for lot, ratio in zip(selected_lots, option):
            amount = barem_multiple_quantity(
                required * ratio,
                lot.get("remaining_quantity", 0),
                lot.get("barem_weight"),
                min_take_quantity,
                max_take_quantity,
            )
            if amount <= QUANTITY_EPSILON:
                continue
            plan.append((lot, amount, ratio, False))
            total = clean_quantity(total + amount)
        shortage = clean_quantity(max(0, required - total))
        accepted_penalty = 0 if shortage <= tolerance + QUANTITY_EPSILON else 1
        score = (accepted_penalty, -len(plan), shortage)
        if best_score is None or score < best_score:
            best_score = score
            best_plan = plan
    if best_plan and best_score and best_score[2] > QUANTITY_EPSILON:
        shortage = clean_quantity(best_score[2])
        planned_amounts = defaultdict(float)
        for lot, amount, _ratio, _is_remainder in best_plan:
            planned_amounts[id(lot)] += amount
        plan_lots = set(planned_amounts)

        def spare_capacity(lot):
            return clean_quantity(lot_capacity(lot) - planned_amounts.get(id(lot), 0))

        remainder_lot = None
        for lot in selected_lots:
            if id(lot) not in plan_lots and spare_capacity(lot) + QUANTITY_EPSILON >= shortage:
                remainder_lot = lot
                break
        if remainder_lot is None:
            for lot, amount, _ratio, _is_remainder in best_plan:
                if spare_capacity(lot) + QUANTITY_EPSILON >= shortage:
                    remainder_lot = lot
                    break
        if remainder_lot is not None:
            best_plan.append((remainder_lot, shortage, 0, True))
            best_score = (0, best_score[1], 0)
    selected_codes = [lot.get("variant_code", "") for lot, _amount, _ratio, _is_remainder in (best_plan or [])]
    rounded_shortage = best_score[2] if best_score else required
    remainder_quantity = clean_quantity(sum(
        amount for _lot, amount, _ratio, is_remainder in (best_plan or []) if is_remainder
    ))
    if remainder_quantity > QUANTITY_EPSILON:
        note = f"Chon {split_count} ma; {format_number(remainder_quantity)} kg du cuoi duoc gop vao ma phu hop nhat."
    elif best_score and best_score[2] == 0:
        note = f"Chọn {split_count} mã; phần dư nhỏ cuối cùng được gộp vào mã phù hợp nhất."
    else:
        note = f"Chọn {split_count} mã, thiếu {format_number(rounded_shortage)} kg sau khi làm tròn theo barem."
    note = f"Da danh gia {policy.get('scenario_count') or 100} phuong an tai lap. {note}"
    if remainder_quantity <= QUANTITY_EPSILON and best_score and best_score[2] == 0:
        note = f"Chon {split_count} ma; toan bo khoi luong dung boi so barem."
    if not has_profitable_lots:
        note = "Khong co ma lai hop le; buoc phai chon ma lo thap nhat. " + note
    return best_plan or [], missing_barem, selected_codes, note


def choose_barem_generic_plan(sale, candidates, policy, barem_map):
    """Plan by material type, then expand the accepted plan back to purchase lots."""
    if not candidates:
        return [], [], [], "Kh\u00f4ng c\u00f3 m\u00e3 \u1ee9ng vi\u00ean t\u1ea1i th\u1eddi \u0111i\u1ec3m b\u00e1n."

    required = clean_quantity(sale.get("quantity", 0))
    if required <= QUANTITY_EPSILON:
        return [], [], [], "S\u1ed1 l\u01b0\u1ee3ng b\u00e1n ra b\u1eb1ng 0."

    min_take = policy.get("generic_min_take_quantity")
    max_take = policy.get("generic_max_take_quantity")
    min_type_count = max(1, int(policy.get("generic_min_type_count") or 2))
    scenario_count = max(1, int(policy.get("scenario_count") or 100))
    missing_barem = []
    missing_codes = set()
    grouped = {}
    available_by_lot = {}
    priority_by_lot = dict(policy.get("_generic_priority_by_lot") or {})

    def cached_priority(lot):
        lot_id = id(lot)
        if lot_id not in priority_by_lot:
            priority = generic_lot_priority(lot, sale.get("unit_price"))
            priority_by_lot[lot_id] = priority
        return priority_by_lot[lot_id]

    for lot in candidates:
        code = lot.get("_normalized_variant_code") or normalize_code(lot.get("variant_code", ""))
        lot["_normalized_variant_code"] = code
        remaining = lot_unreserved_quantity(lot)
        if remaining <= QUANTITY_EPSILON:
            continue
        available_by_lot[id(lot)] = remaining
        if "_son_phuong_barem_details" in lot:
            barem_details = lot.get("_son_phuong_barem_details") or None
        else:
            barem_details = resolve_barem_details(barem_map, lot)
            lot["_son_phuong_barem_details"] = barem_details or {}
        barem = barem_details.get("weight") if barem_details else None
        if barem is None or barem <= QUANTITY_EPSILON:
            missing_key = code or f"row:{lot.get('row_number', '')}"
            if missing_key not in missing_codes:
                missing_barem.append(missing_barem_item(
                    lot,
                    "Thi\u1ebfu kh\u1ed1i l\u01b0\u1ee3ng barem cho m\u00e3/kh\u00f3a k\u1ef9 thu\u1eadt.",
                ))
                missing_codes.add(missing_key)
            continue
        lot["barem_weight"] = barem
        lot["barem_source"] = barem_details.get("source", "")
        lot["barem_method"] = barem_details.get("method", "")
        lot["barem_references"] = list(barem_details.get("references") or [])
        profile_key = lot.get("steel_profile_key") or steel_profile_key(
            lot.get("product_name", ""),
            lot.get("source_variant_code") or lot.get("variant_code", ""),
        )
        material_key = (
            f"profile:{profile_key}|barem:{barem:.12g}"
            if profile_key else f"code:{strip_son_phuong_company_prefix(code)}|barem:{barem:.12g}"
        )
        group = grouped.setdefault(material_key, {
            "code": code,
            "material_key": material_key,
            "steel_kind": lot.get("steel_kind") or steel_kind(
                lot.get("product_name", ""),
                lot.get("source_variant_code") or lot.get("variant_code", ""),
            ),
            "lots": [],
            "barem_weight": barem,
            "barem_source": lot.get("barem_source", ""),
            "barem_method": lot.get("barem_method", ""),
            "barem_references": lot.get("barem_references", []),
            "capacity": 0.0,
        })
        group["lots"].append(lot)
        group["capacity"] = clean_quantity(group["capacity"] + remaining)

    groups = []
    for group in grouped.values():
        if max_take is not None:
            group["capacity"] = clean_quantity(min(group["capacity"], max_take))
        if group["capacity"] <= QUANTITY_EPSILON:
            continue
        if min_take is not None and group["capacity"] + QUANTITY_EPSILON < min_take:
            continue
        group["lots"].sort(key=cached_priority)
        group["priority"] = cached_priority(group["lots"][0])
        groups.append(group)

    generic_type = generic_steel_sale_type(sale.get("product_name", ""))
    kind_capacities = defaultdict(float)
    for group in groups:
        kind_capacities[group.get("steel_kind", "unknown")] += group["capacity"]
    future_kind_demand = policy.get("_future_generic_kind_demand") or {}

    groups.sort(key=lambda group: (
        generic_lot_quality_rank(group["lots"][0], sale.get("unit_price")),
        -max(
            0,
            kind_capacities.get(group.get("steel_kind", "unknown"), 0)
            - clean_quantity(future_kind_demand.get(group.get("steel_kind", "unknown"), 0)),
        ) if generic_type == "pipe_box" else 0,
        group["priority"],
        -group["capacity"],
        group["code"],
    ))
    if not groups:
        return [], missing_barem, [], "Kh\u00f4ng c\u00f3 m\u00e3 n\u00e0o c\u00f3 barem v\u00e0 t\u1ed3n h\u1ee3p l\u1ec7."

    def minimum_amount(group):
        barem = group["barem_weight"]
        minimum = max(barem, min_take or 0)
        multiples = max(1, int(math.ceil((minimum - QUANTITY_EPSILON) / barem)))
        return clean_quantity(multiples * barem)

    def targets_for_weights(selected, weights):
        if len(selected) < min_type_count:
            return None
        if sum(group["capacity"] for group in selected) + QUANTITY_EPSILON < required:
            return None
        weight_total = sum(weights) or 1
        targets = []
        for group, weight in zip(selected, weights):
            target = required * weight / weight_total
            amount = barem_multiple_quantity(
                target,
                group["capacity"],
                group["barem_weight"],
            )
            minimum = minimum_amount(group)
            if amount + QUANTITY_EPSILON < minimum:
                amount = minimum if minimum <= group["capacity"] + QUANTITY_EPSILON else 0
            if amount <= QUANTITY_EPSILON:
                return None
            targets.append({
                "group": group,
                "amount": clean_quantity(amount),
                "remainder": 0.0,
            })
        total = clean_quantity(sum(item["amount"] for item in targets))
        if total > required + QUANTITY_EPSILON:
            return None

        shortage = clean_quantity(required - total)
        for item in sorted(
            targets,
            key=lambda row: (
                -(row["group"]["capacity"] - row["amount"]),
                row["group"]["priority"],
            ),
        ):
            if shortage <= QUANTITY_EPSILON:
                break
            spare = clean_quantity(item["group"]["capacity"] - item["amount"])
            addition = barem_multiple_quantity(
                shortage,
                spare,
                item["group"]["barem_weight"],
            )
            if addition <= QUANTITY_EPSILON:
                continue
            item["amount"] = clean_quantity(item["amount"] + addition)
            shortage = clean_quantity(shortage - addition)

        if shortage > QUANTITY_EPSILON:
            remainder_limit = policy.get("barem_remainder_max_kg")
            remainder_options = []
            for item in targets:
                if item["group"]["capacity"] - item["amount"] + QUANTITY_EPSILON < shortage:
                    continue
                barem = item["group"]["barem_weight"]
                desired_amount = clean_quantity(item["amount"] + shortage)
                barem_multiple = max(1, int(math.floor((desired_amount / barem) + 0.5)))
                theoretical_amount = clean_quantity(barem_multiple * barem)
                tolerance_kg = clean_quantity(desired_amount - theoretical_amount)
                if remainder_limit is not None and abs(tolerance_kg) > remainder_limit + QUANTITY_EPSILON:
                    continue
                remainder_options.append((item, tolerance_kg))
            if not remainder_options:
                return None
            remainder_target, tolerance_kg = sorted(
                remainder_options,
                key=lambda option: (
                    -option[0]["group"]["barem_weight"],
                    abs(option[1]),
                    option[0]["group"]["priority"],
                ),
            )[0]
            remainder_target["amount"] = clean_quantity(remainder_target["amount"] + shortage)
            remainder_target["remainder"] = tolerance_kg
            shortage = 0.0

        return targets if shortage <= QUANTITY_EPSILON else None

    def expand_targets(targets):
        expanded = []
        for target in sorted(targets, key=lambda row: row["group"]["priority"]):
            group = target["group"]
            quantity_left = target["amount"]
            ratio = target["amount"] / required if required else 0
            remainder_left = target["remainder"]
            for lot in group["lots"]:
                if quantity_left <= QUANTITY_EPSILON:
                    break
                amount = clean_quantity(min(quantity_left, available_by_lot.get(id(lot), 0)))
                if amount <= QUANTITY_EPSILON:
                    continue
                quantity_left = clean_quantity(quantity_left - amount)
                is_remainder = abs(remainder_left) > QUANTITY_EPSILON and quantity_left <= QUANTITY_EPSILON
                expanded.append((lot, amount, ratio, remainder_left if is_remainder else 0.0))
                if is_remainder:
                    remainder_left = 0.0
            if quantity_left > QUANTITY_EPSILON:
                return None
        return expanded

    def plan_score(targets, expanded):
        shortage = clean_quantity(max(0, required - sum(item["amount"] for item in targets)))
        remainder = clean_quantity(sum(abs(item["remainder"]) for item in targets))
        scarcity_cost = 0.0
        used_by_kind = defaultdict(float)
        positive_profit = 0.0
        loss = 0.0
        fifo = 0.0
        sale_price = sale.get("unit_price")
        for lot, amount, _ratio, _is_remainder in expanded:
            fifo += (lot.get("sequence", 0) or 0) * amount
            if generic_type == "pipe_box":
                lot_kind = lot.get("steel_kind") or "unknown"
                used_by_kind[lot_kind] += amount
                kind_capacity = kind_capacities.get(lot_kind, 0)
                if kind_capacity > QUANTITY_EPSILON:
                    scarcity_cost += amount / kind_capacity
            if sale_price is None or lot.get("unit_price") is None:
                continue
            difference = (sale_price - lot["unit_price"]) * amount
            if difference >= 0:
                positive_profit += difference
            else:
                loss += abs(difference)
        preservation_shortage = clean_quantity(sum(
            max(
                0,
                clean_quantity(future_kind_demand.get(kind, 0))
                - clean_quantity(kind_capacities.get(kind, 0) - used_by_kind.get(kind, 0)),
            )
            for kind in ("pipe", "box")
        )) if generic_type == "pipe_box" else 0.0
        # Avoid a tiny barem improvement consuming stock required by constrained future sales.
        # For KHHVT, actual loss is more important than barem, scarcity, or
        # FIFO preferences.  Profit is minimized only after loss is minimized.
        return (
            shortage,
            clean_quantity(loss),
            1 if loss > QUANTITY_EPSILON else 0,
            clean_quantity(positive_profit),
            preservation_shortage,
            remainder,
            clean_quantity(scarcity_cost),
            clean_quantity(fifo),
        )

    best_plan = None
    best_targets = None
    best_score = None
    chosen_type_count = 0
    attempts = 0
    max_types = len(groups)
    first_type_count = min(min_type_count, max_types)
    total_capacity = clean_quantity(sum(group["capacity"] for group in groups))
    if total_capacity + QUANTITY_EPSILON >= required:
        capacity_total = 0.0
        capacity_type_count = 0
        for group in sorted(groups, key=lambda item: item["capacity"], reverse=True):
            capacity_total = clean_quantity(capacity_total + group["capacity"])
            capacity_type_count += 1
            if capacity_total + QUANTITY_EPSILON >= required:
                break
        first_type_count = max(first_type_count, capacity_type_count)
    rng = deterministic_random([
        sale.get("invoice_date_iso", ""),
        sale.get("invoice_no", ""),
        sale.get("row_number", ""),
        sale.get("variant_code", ""),
    ])

    type_counts = range(first_type_count, max_types + 1) if total_capacity + QUANTITY_EPSILON >= required else ()
    for type_count in type_counts:
        exact_plan_found = False
        for option_index in range(scenario_count):
            attempts += 1
            if option_index == 0:
                selected = groups[:type_count]
                weights = [1.0] * type_count
            elif option_index == 1:
                selected = sorted(groups, key=lambda group: group["capacity"], reverse=True)[:type_count]
                weights = [group["capacity"] for group in selected]
            elif option_index == 2:
                selected = groups[:type_count]
                weights = [group["capacity"] for group in selected]
            else:
                selected = rng.sample(groups, type_count)
                weights = [rng.random() + 0.05 for _ in selected]
            targets = targets_for_weights(selected, weights)
            if not targets:
                continue
            expanded = expand_targets(targets)
            if not expanded:
                continue
            score = plan_score(targets, expanded)
            if best_score is None or score < best_score:
                best_score = score
                best_plan = expanded
                best_targets = targets
                chosen_type_count = type_count
            if score[0] <= QUANTITY_EPSILON:
                exact_plan_found = True
        if best_plan:
            break

    generated_new_plan = False
    if not best_plan:
        generated_new_plan = True
        selected = groups
        capacity_weights = [group["capacity"] for group in selected]
        targets = targets_for_weights(selected, capacity_weights)
        if targets:
            best_targets = targets
            best_plan = expand_targets(targets)
            chosen_type_count = len(targets)

    if not best_plan:
        generated_new_plan = True
        best_plan = []
        quantity_left = required
        selected_codes = []
        allocated_by_group = defaultdict(float)
        for group in groups:
            if quantity_left <= QUANTITY_EPSILON:
                break
            group_target = clean_quantity(min(group["capacity"], quantity_left))
            group_target = barem_multiple_quantity(group_target, group["capacity"], group["barem_weight"])
            if group_target <= QUANTITY_EPSILON:
                continue
            selected_codes.append(group["code"])
            for lot in group["lots"]:
                if group_target <= QUANTITY_EPSILON:
                    break
                amount = clean_quantity(min(group_target, available_by_lot.get(id(lot), 0)))
                if amount <= QUANTITY_EPSILON:
                    continue
                best_plan.append((lot, amount, amount / required if required else 0, False))
                allocated_by_group[group["material_key"]] = clean_quantity(
                    allocated_by_group[group["material_key"]] + amount
                )
                group_target = clean_quantity(group_target - amount)
                quantity_left = clean_quantity(quantity_left - amount)

        # Keep the final small balance in real KHHVT stock. It may be assigned
        # as the one permitted non-barem remainder, including to a second type
        # whose remaining quantity is smaller than one full barem.
        remainder_limit = policy.get("barem_remainder_max_kg")
        remaining_stock_capacity = clean_quantity(sum(
            max(0, group["capacity"] - allocated_by_group.get(group["material_key"], 0))
            for group in groups
        ))
        remainder_quantity = clean_quantity(min(quantity_left, remaining_stock_capacity))
        if (
            remainder_quantity > QUANTITY_EPSILON
            and (remainder_limit is None or remainder_quantity <= remainder_limit + QUANTITY_EPSILON)
        ):
            selected_group_keys = {
                key for key, amount in allocated_by_group.items()
                if amount > QUANTITY_EPSILON
            }
            remainder_groups = sorted(
                groups,
                key=lambda group: (
                    0 if group["material_key"] not in selected_group_keys else 1,
                    group["priority"],
                ),
            )
            for group in remainder_groups:
                group_used = allocated_by_group.get(group["material_key"], 0)
                if clean_quantity(group["capacity"] - group_used) + QUANTITY_EPSILON < remainder_quantity:
                    continue
                amount_left = remainder_quantity
                for lot in group["lots"]:
                    lot_used = clean_quantity(sum(
                        amount
                        for planned_lot, amount, _ratio, _remainder in best_plan
                        if id(planned_lot) == id(lot)
                    ))
                    lot_spare = clean_quantity(available_by_lot.get(id(lot), 0) - lot_used)
                    amount = clean_quantity(min(amount_left, lot_spare))
                    if amount <= QUANTITY_EPSILON:
                        continue
                    amount_left = clean_quantity(amount_left - amount)
                    is_last = amount_left <= QUANTITY_EPSILON
                    best_plan.append((
                        lot,
                        amount,
                        amount / required if required else 0,
                        remainder_quantity if is_last else 0.0,
                    ))
                    if is_last:
                        break
                if amount_left <= QUANTITY_EPSILON:
                    if group["code"] not in selected_codes:
                        selected_codes.append(group["code"])
                    quantity_left = clean_quantity(quantity_left - remainder_quantity)
                    break
        note = (
            f"Kh\u00f4ng c\u00f3 m\u1eabu \u0111\u1ee7 {format_number(required)} kg. "
            f"\u0110\u00e3 t\u1ea1o m\u1eabu m\u1edbi t\u1eeb {len(selected_codes)} lo\u1ea1i c\u00f2n t\u1ed3n; "
            f"c\u00f2n {format_number(max(0, quantity_left))} kg ch\u01b0a ph\u00e2n b\u1ed5 v\u00e0 kh\u00f4ng xu\u1ea5t \u00e2m KHHVT."
        )
        return best_plan, missing_barem, selected_codes, note

    selected_codes = list(dict.fromkeys(
        lot.get("_normalized_variant_code") or normalize_code(lot.get("variant_code", ""))
        for lot, _amount, _ratio, _is_remainder in best_plan
    ))
    allocated_quantity = clean_quantity(sum(item["amount"] for item in (best_targets or [])))
    unallocated_quantity = clean_quantity(max(0, required - allocated_quantity))
    source = "m\u1eabu m\u1edbi" if generated_new_plan else "m\u1eabu th\u1eed"
    note = (
        f"Ch\u1ecdn {source} {chosen_type_count} lo\u1ea1i sau {attempts} l\u1ea7n ki\u1ec3m tra; "
        f"\u0111\u00e3 g\u1ed9p t\u1ed3n theo m\u00e3 t\u1eeb nhi\u1ec1u h\u00f3a \u0111\u01a1n mua."
    )
    if unallocated_quantity > QUANTITY_EPSILON:
        note += f" Ch\u01b0a ph\u00e2n b\u1ed5 {format_number(unallocated_quantity)} kg v\u00ec kh\u00f4ng th\u1ec3 bi\u1ec3u di\u1ec5n b\u1eb1ng b\u1ed9i s\u1ed1 barem; ph\u1ea7n n\u00e0y ph\u1ea3i xu\u1ea5t \u00e2m v\u00e0 c\u1ea3nh b\u00e1o."
    return best_plan, missing_barem, selected_codes, note

def inventory_snapshot(lots, sale_price):
    return [{
        "variant_code": lot["variant_code"],
        "source": lot["source"],
        "invoice_no": lot.get("invoice_no", ""),
        "invoice_date": lot.get("invoice_date", ""),
        "row_number": lot["row_number"],
        "quantity": clean_quantity(lot["remaining_quantity"]),
        "unit_cost": lot["unit_price"],
        "profit_percent": lot_margin_percent(lot, sale_price),
        "summary_count": lot.get("summary_count", 1),
    } for lot in lots if lot["remaining_quantity"] > QUANTITY_EPSILON]


def allocate_stock(opening_lines, purchase_lines, sales_lines, policy=None, progress_callback=None, barem_map=None):
    policy = clean_policy(policy or DEFAULT_POLICY)
    is_son_phuong = policy.get("company_profile") == "son_phuong"
    sales_pair_map = policy.get("sales_inventory_pairs") if is_son_phuong else {}
    material_pair = sales_pair_map.get("materials", {"ma_kho": "KHHVT", "tk_vat_tu": "156"}) if is_son_phuong else {}
    # Sơn Phương material rows must not be completed by a negative KHHVT
    # export. A shortfall remains unresolved and blocks FDI generation.
    allow_negative_export = False if is_son_phuong else bool(policy.get("allow_negative_export"))
    if is_son_phuong:
        purchase_lines = sorted((dict(line) for line in purchase_lines), key=lambda line: sale_sort_key(line, True))
        opening_lines = sorted((dict(line) for line in opening_lines), key=lambda line: sale_sort_key(line, True))
    else:
        purchase_lines = summarize_purchase_lines(purchase_lines)
    lots = make_lots(opening_lines, purchase_lines)
    by_code = defaultdict(list)
    by_variant = defaultdict(list)
    by_steel_profile = defaultdict(list)
    by_steel_kind = defaultdict(list)
    by_steel_kind_coating = defaultdict(list)
    by_son_phuong_match_key = defaultdict(list)
    for lot in lots:
        lot["_normalized_variant_code"] = normalize_code(lot.get("variant_code", ""))
        by_code[lot["base_code"]].append(lot)
        by_variant[lot["variant_code"]].append(lot)
        if is_son_phuong and lot.get("steel_profile_key"):
            by_steel_profile[lot["steel_profile_key"]].append(lot)
        if is_son_phuong:
            lot_kind = lot.get("steel_kind") or steel_kind(
                lot.get("product_name", ""),
                lot.get("source_variant_code") or lot.get("variant_code", ""),
            )
            lot_coating = lot.get("steel_coating") or steel_coating(
                lot.get("product_name", ""),
                lot.get("source_variant_code") or lot.get("variant_code", ""),
            )
            lot["steel_kind"] = lot_kind
            lot["steel_coating"] = lot_coating
            by_steel_kind[lot_kind].append(lot)
            by_steel_kind_coating[(lot_kind, lot_coating)].append(lot)
            for key in son_phuong_line_match_keys(lot):
                by_son_phuong_match_key[key].append(lot)

    allocations = []
    warnings = []
    missing_barem_report = []
    total_sales = len(sales_lines)
    ordered_sales = sorted(sales_lines, key=lambda sale: sale_sort_key(sale, is_son_phuong))
    remaining_generic_kind_demand = defaultdict(float)
    if is_son_phuong:
        for pending_sale in ordered_sales:
            pending_type = generic_steel_sale_type(pending_sale.get("product_name", ""))
            if pending_type in {"pipe", "box"}:
                remaining_generic_kind_demand[pending_type] = clean_quantity(
                    remaining_generic_kind_demand[pending_type] + pending_sale.get("quantity", 0)
                )
    specific_reservations = (
        reserve_son_phuong_specific_sales(ordered_sales, by_steel_profile, policy)
        if is_son_phuong else {}
    )
    for sale_index, sale in enumerate(ordered_sales, start=1):
        if progress_callback and (sale_index == 1 or sale_index == total_sales or sale_index % 250 == 0):
            progress_callback(sale_index, total_sales)
        required = sale["quantity"]
        remaining = required
        sale_role = son_phuong_sales_role(sale) if is_son_phuong else ""
        used = []
        # A sales row can be split back to any accepted purchase suffix under the same
        # base code (.001/.002/.003). If none passes date and margin checks, the
        # remainder is pushed to KTP.
        generic_type = generic_steel_sale_type(sale.get("product_name", "")) if is_son_phuong else None
        if generic_type and (sale_index == 1 or sale_index % 50 == 0):
            for lot_kind, kind_lots in list(by_steel_kind.items()):
                by_steel_kind[lot_kind] = [lot for lot in kind_lots if lot_unreserved_quantity(lot) > QUANTITY_EPSILON]
            for kind_coating, coating_lots in list(by_steel_kind_coating.items()):
                by_steel_kind_coating[kind_coating] = [lot for lot in coating_lots if lot_unreserved_quantity(lot) > QUANTITY_EPSILON]
        if generic_type in {"pipe", "box"}:
            remaining_generic_kind_demand[generic_type] = clean_quantity(max(
                0,
                remaining_generic_kind_demand.get(generic_type, 0) - required,
            ))
        generic_pool = []
        if generic_type:
            allowed_kinds = sorted(generic_allowed_kinds(generic_type))
            sale_coating = explicit_steel_coating(
                sale.get("product_name", ""),
                sale.get("source_variant_code") or sale.get("variant_code", ""),
            )
            for allowed_kind in allowed_kinds:
                if sale_coating:
                    generic_pool.extend(by_steel_kind_coating.get((allowed_kind, sale_coating), []))
                    generic_pool.extend(by_steel_kind_coating.get((allowed_kind, "unknown"), []))
                else:
                    generic_pool.extend(by_steel_kind.get(allowed_kind, []))
        if is_son_phuong and sale_role == "fallback":
            generic_type, generic_pool = None, []
        is_generic_sale = bool(generic_type)
        specific_reservation_plan = specific_reservations.get(
            son_phuong_sale_reservation_key(sale),
            [],
        ) if is_son_phuong else []
        sale_profile_key = sale.get("steel_profile_key") or steel_profile_key(
            sale.get("product_name", ""),
            sale.get("source_variant_code") or sale.get("variant_code", ""),
        ) if is_son_phuong and not is_generic_sale else ""
        if is_son_phuong and sale_role == "fallback":
            lot_pool = []
        elif is_generic_sale:
            lot_pool = generic_pool
        elif is_son_phuong and sale_profile_key:
            lot_pool = son_phuong_compatible_profile_lots(by_steel_profile, sale_profile_key)
        elif is_son_phuong:
            lot_pool = []
            seen_lots = set()
            for key in son_phuong_line_match_keys(sale):
                for lot in by_son_phuong_match_key.get(key, []):
                    lot_id = id(lot)
                    if lot_id in seen_lots:
                        continue
                    seen_lots.add(lot_id)
                    lot_pool.append(lot)
        else:
            lot_pool = by_code.get(sale["base_code"], [])
        available_lots = []
        generic_priority_by_lot = {}
        for lot in lot_pool:
            available_quantity = (
                clean_quantity(lot.get("remaining_quantity", 0))
                if specific_reservation_plan or not is_son_phuong
                else lot_unreserved_quantity(lot)
            )
            if available_quantity <= QUANTITY_EPSILON or not lot_available_for_sale(lot, sale):
                continue
            available_lots.append(lot)
            if is_generic_sale:
                generic_priority_by_lot[id(lot)] = generic_lot_priority_for_quantity(
                    lot,
                    sale.get("unit_price"),
                    available_quantity,
                )
        future_reorder_lots = [
            lot for lot in lot_pool
            if lot["remaining_quantity"] > 0 and future_lot_allowed_for_sale(lot, sale, policy)
        ] if not is_son_phuong else []
        priority_fn = generic_lot_priority if is_generic_sale else lot_priority
        def current_priority(lot):
            if is_generic_sale and id(lot) in generic_priority_by_lot:
                return generic_priority_by_lot[id(lot)]
            return priority_fn(lot, sale["unit_price"])
        if is_son_phuong and sale_profile_key:
            available_lots.sort(key=lambda lot: (
                son_phuong_surface_priority(lot, sale_profile_key),
                current_priority(lot),
            ))
        else:
            available_lots.sort(key=current_priority)
        future_reorder_lots.sort(key=current_priority)
        future_reorder_ids = {id(lot) for lot in future_reorder_lots}
        candidate_pool = available_lots + future_reorder_lots
        inventory_before = inventory_snapshot(candidate_pool, sale["unit_price"])
        candidates = []
        rejected = []
        for lot in candidate_pool:
            accepted, reason = lot_acceptance(lot, sale["unit_price"], policy)
            if accepted:
                candidates.append(lot)
            else:
                rejected.append({
                    "variant_code": lot["variant_code"],
                    "sale_variant_code": sale["variant_code"],
                "invoice_no": lot.get("invoice_no", ""),
                "invoice_date": lot.get("invoice_date", ""),
                "invoice_date_iso": lot.get("invoice_date_iso", ""),
                "quantity": lot["remaining_quantity"],
                "reason": reason,
                "margin_percent": lot_margin_percent(lot, sale["unit_price"]),
                "summary_count": lot.get("summary_count", 1),
                "future_purchase_reordered": id(lot) in future_reorder_ids,
                "future_reorder_days": future_purchase_days(lot, sale),
            })
        selected_generic_codes = None
        generic_first_pass_targets = {}
        generic_barem_plan = None
        generic_plan_note = ""
        if is_generic_sale:
            if is_son_phuong and barem_map is not None:
                generic_policy = policy
                if generic_type == "pipe_box":
                    generic_policy = dict(policy)
                    generic_policy["_future_generic_kind_demand"] = dict(remaining_generic_kind_demand)
                else:
                    generic_policy = dict(generic_policy)
                generic_policy["_generic_priority_by_lot"] = generic_priority_by_lot
                generic_barem_plan, missing_barem, selected_codes, generic_plan_note = choose_barem_generic_plan(
                    sale,
                    candidates,
                    generic_policy,
                    barem_map,
                )
                missing_barem_report.extend(missing_barem)
                selected_generic_codes = set(selected_codes)
                candidates = [lot for lot in candidates if lot.get("_normalized_variant_code") in selected_generic_codes]
            else:
                selected_lots = []
                selected_codes = []
                max_types = policy["son_phuong_split_counts"].get(generic_type, 1)
                for lot in candidates:
                    code = lot.get("variant_code", "")
                    if code in selected_codes:
                        continue
                    if not lot_within_generic_variance(lot, selected_lots, policy.get("generic_split_variance_percent")):
                        rejected.append({
                            "variant_code": lot["variant_code"],
                            "sale_variant_code": sale["variant_code"],
                            "invoice_no": lot.get("invoice_no", ""),
                            "invoice_date": lot.get("invoice_date", ""),
                            "quantity": lot["remaining_quantity"],
                            "reason": f"đơn giá chênh hơn {format_number(policy.get('generic_split_variance_percent'))}% so với nhóm đã chọn",
                            "margin_percent": lot_margin_percent(lot, sale["unit_price"]),
                            "summary_count": lot.get("summary_count", 1),
                        })
                        continue
                    selected_lots.append(lot)
                    selected_codes.append(code)
                    if len(selected_codes) >= max_types:
                        break
                selected_generic_codes = set(selected_codes)
                if selected_generic_codes:
                    quota = required / len(selected_generic_codes)
                    generic_first_pass_targets = {code: quota for code in selected_generic_codes}
                    candidates = [lot for lot in candidates if lot.get("variant_code") in selected_generic_codes]

        def consume_lot(lot, amount, barem_remainder=False):
            nonlocal remaining
            if amount <= QUANTITY_EPSILON:
                return
            lot["remaining_quantity"] = clean_quantity(lot["remaining_quantity"] - amount)
            lot["allocated_quantity"] = clean_quantity(lot["allocated_quantity"] + amount)
            remaining = clean_quantity(remaining - amount)
            margin = None
            if sale["unit_price"] is not None and lot["unit_price"] is not None:
                margin = sale["unit_price"] - lot["unit_price"]
            barem_weight = lot.get("barem_weight")
            tolerance_kg = (
                clean_quantity(barem_remainder)
                if isinstance(barem_remainder, (int, float)) and not isinstance(barem_remainder, bool)
                else 0.0
            )
            theoretical_amount = clean_quantity(amount - tolerance_kg)
            barem_multiple = (
                max(0, int(math.floor((theoretical_amount / barem_weight) + 0.5)))
                if barem_weight else None
            )
            used.append({
                "variant_code": lot["variant_code"],
                "purchase_variant_code": lot["variant_code"],
                "purchase_source_variant_code": lot.get("source_variant_code", lot["variant_code"]),
                "sale_variant_code": sale["variant_code"],
                "sale_source_variant_code": sale.get("source_variant_code", sale["variant_code"]),
                # Matched stock must keep the purchase variant that supplied
                # both its inventory quantity and cost.
                "ledger_variant_code": lot["variant_code"],
                "steel_profile_key": lot.get("steel_profile_key", ""),
                "source": lot["source"],
                "invoice_no": lot.get("invoice_no", ""),
                "invoice_date": lot.get("invoice_date", ""),
                "invoice_date_iso": lot.get("invoice_date_iso", ""),
                "row_number": lot["row_number"],
                "purchase_product_name": lot.get("product_name", ""),
                "purchase_party_name": lot.get("party_name", ""),
                "purchase_party_tax_code": lot.get("party_tax_code", ""),
                "purchase_unit_name": lot.get("unit_name", ""),
                "quantity": amount,
                "unit_cost": lot["unit_price"],
                "profit_per_unit": margin,
                "profit_percent": lot_margin_percent(lot, sale["unit_price"]),
                "summary_count": lot.get("summary_count", 1),
                "barem_weight": barem_weight,
                "barem_multiple": barem_multiple,
                "barem_remainder_kg": tolerance_kg,
                "barem_source": lot.get("barem_source", ""),
                "barem_method": lot.get("barem_method", ""),
                "barem_references": lot.get("barem_references", []),
                "barem_remainder": barem_remainder,
                "future_purchase_reordered": id(lot) in future_reorder_ids,
                "future_reorder_days": future_purchase_days(lot, sale) if id(lot) in future_reorder_ids else None,
                "original_purchase_date": lot.get("invoice_date", "") if id(lot) in future_reorder_ids else "",
                "original_purchase_date_iso": lot.get("invoice_date_iso", "") if id(lot) in future_reorder_ids else "",
                "effective_purchase_date": sale.get("invoice_date", "") if id(lot) in future_reorder_ids else "",
                "effective_purchase_date_iso": sale.get("invoice_date_iso", "") if id(lot) in future_reorder_ids else "",
                "future_reorder_sale_invoice_no": sale.get("invoice_no", "") if id(lot) in future_reorder_ids else "",
                "future_reorder_sale_row": sale.get("row_number", "") if id(lot) in future_reorder_ids else "",
                "logic_note": (
                    f"Dung sai barem cuoi {format_number(tolerance_kg)} kg; "
                    f"hoa don {format_number(amount)} kg, ly thuyet {format_number(theoretical_amount)} kg."
                    if barem_remainder else ""
                ),
            })
            if id(lot) in future_reorder_ids:
                note = (
                    f"HD mua vao ngay {lot.get('invoice_date', '')} duoc dua len truoc "
                    f"HD ban ra ngay {sale.get('invoice_date', '')} trong {future_purchase_days(lot, sale)} ngay."
                )
                used[-1]["logic_note"] = f"{used[-1].get('logic_note', '')} {note}".strip()
        if specific_reservation_plan:
            accepted_lot_ids = {id(lot) for lot in candidates}
            for lot, reserved_amount in specific_reservation_plan:
                if remaining <= QUANTITY_EPSILON:
                    break
                if id(lot) not in accepted_lot_ids or not lot_available_for_sale(lot, sale):
                    continue
                amount = clean_quantity(min(
                    remaining,
                    reserved_amount,
                    lot.get("remaining_quantity", 0),
                ))
                if amount <= QUANTITY_EPSILON:
                    continue
                consume_lot(lot, amount)
                lot["specific_reserved_quantity"] = clean_quantity(max(
                    0,
                    lot.get("specific_reserved_quantity", 0) - amount,
                ))
                used[-1]["specific_stock_reserved"] = True

        def consumable_quantity(lot):
            if is_son_phuong:
                return lot_unreserved_quantity(lot)
            return clean_quantity(lot.get("remaining_quantity", 0))
        if is_generic_sale and generic_barem_plan is not None:
            for lot, planned_amount, _ratio, is_remainder in generic_barem_plan:
                if remaining <= QUANTITY_EPSILON:
                    break
                consume_lot(lot, min(remaining, planned_amount, consumable_quantity(lot)), barem_remainder=is_remainder)
        elif is_generic_sale and generic_first_pass_targets:
            for lot in candidates:
                if remaining <= QUANTITY_EPSILON:
                    break
                code = lot.get("variant_code", "")
                target = generic_first_pass_targets.get(code, 0)
                already_used = sum(item["quantity"] for item in used if item.get("variant_code") == code)
                amount = min(remaining, consumable_quantity(lot), max(0, target - already_used))
                consume_lot(lot, amount)
        if generic_barem_plan is None:
            for lot in candidates:
                if remaining <= QUANTITY_EPSILON:
                    break
                if selected_generic_codes is not None and lot.get("variant_code") not in selected_generic_codes:
                    continue
                amount = min(remaining, consumable_quantity(lot))
                consume_lot(lot, amount)
        if is_son_phuong and sale_role == "materials" and remaining > QUANTITY_EPSILON:
            if not lot_pool:
                shortage_reason = "Khong co ma mua vao dung nhom ong/hop va den/ma kem."
            elif not available_lots:
                shortage_reason = "Co ma mua vao dung nhom, nhung khong co ton hop le tai ngay ban."
            elif is_generic_sale and generic_barem_plan is not None:
                shortage_reason = generic_plan_note or "Khong ghep du theo barem/min-max."
            else:
                shortage_reason = "Khong ghep du so luong tu cac ma ung vien."
            warnings.append(f"{sale['variant_code']}: chua ghep du {format_number(remaining)} kg; {shortage_reason}")
        else:
            shortage_reason = ""
        if is_son_phuong and sale_role == "materials" and allow_negative_export and remaining > QUANTITY_EPSILON:
            fallback_lots = sorted(lot_pool, key=lambda lot: priority_fn(lot, sale["unit_price"]))
            fallback_lot = fallback_lots[0] if fallback_lots else {}
            unit_cost = fallback_lot.get("unit_price")
            used.append({
                "variant_code": fallback_lot.get("variant_code", sale["variant_code"]),
                "purchase_variant_code": fallback_lot.get("variant_code", sale["variant_code"]),
                "purchase_source_variant_code": fallback_lot.get("source_variant_code", fallback_lot.get("variant_code", sale["variant_code"])),
                "sale_variant_code": sale["variant_code"],
                "sale_source_variant_code": sale.get("source_variant_code", sale["variant_code"]),
                "ledger_variant_code": fallback_lot.get("variant_code", sale["variant_code"]) if is_son_phuong else sale["variant_code"],
                "steel_profile_key": fallback_lot.get("steel_profile_key", ""),
                "source": "Chua ghep du KHH" if is_son_phuong else "Xuất âm",
                "invoice_no": fallback_lot.get("invoice_no", ""),
                "invoice_date": fallback_lot.get("invoice_date", ""),
                "row_number": fallback_lot.get("row_number", sale.get("row_number", 0)),
                "purchase_product_name": fallback_lot.get("product_name", ""),
                "purchase_party_name": fallback_lot.get("party_name", ""),
                "purchase_party_tax_code": fallback_lot.get("party_tax_code", ""),
                "purchase_unit_name": fallback_lot.get("unit_name", ""),
                "quantity": remaining,
                "unit_cost": unit_cost,
                "profit_per_unit": (sale["unit_price"] - unit_cost) if sale["unit_price"] is not None and unit_cost is not None else None,
                "profit_percent": lot_margin_percent(fallback_lot, sale["unit_price"]) if fallback_lot else None,
                "summary_count": fallback_lot.get("summary_count", 1),
                "negative_export": True,
                "barem_unallocated": bool(is_generic_sale),
                "barem_weight": fallback_lot.get("barem_weight"),
                "logic_note": shortage_reason if is_son_phuong else "",
            })
            remaining = 0
        material_quantity = clean_quantity(required - remaining)
        if candidates and any(lot["unit_price"] is None for lot in candidates):
            warnings.append(f"{sale['variant_code']}: có tồn kho thiếu đơn giá vốn; ưu tiên lãi chỉ áp dụng được cho phần có đơn giá.")
        if rejected:
            warnings.append(f"{sale['variant_code']}: có {len(rejected)} lô kho không đạt khoảng lãi/lỗ chấp nhận.")
        if generic_plan_note:
            warnings.append(f"{sale['variant_code']}: {generic_plan_note}")
        unresolved_material_quantity = clean_quantity(remaining) if is_son_phuong and sale_role == "materials" else 0.0
        if unresolved_material_quantity > QUANTITY_EPSILON:
            warnings.append(
                f"{sale['variant_code']}: còn thiếu {format_number(unresolved_material_quantity)} kg KHHVT thực; "
                "không tạo xuất âm và không thể tạo FDI bán ra cho đến khi xử lý xong."
            )
        finished_variant_code = sale["variant_code"]
        if remaining > QUANTITY_EPSILON:
            fallback_lots = sorted(lot_pool, key=lambda lot: priority_fn(lot, sale["unit_price"]))
            if fallback_lots:
                finished_variant_code = fallback_lots[0].get("variant_code") or finished_variant_code
        remainder_role = "finished_goods" if sale_role == "finished_goods" else "fallback"
        remainder_pair = sales_pair_map.get(remainder_role, {"ma_kho": "KTP", "tk_vat_tu": "155"}) if is_son_phuong else {}

        allocations.append({
            **sale,
            "allocation_role": sale_role,
            "warehouse_code": material_pair.get("ma_kho", "") if is_son_phuong else sale.get("warehouse_code", ""),
            "warehouse_account": material_pair.get("tk_vat_tu", "") if is_son_phuong else sale.get("warehouse_account", ""),
            "remainder_warehouse_code": remainder_pair.get("ma_kho", "") if is_son_phuong else "",
            "remainder_warehouse_account": remainder_pair.get("tk_vat_tu", "") if is_son_phuong else "",
            "negative_warning": bool(is_son_phuong and sale_role == "materials" and any(item.get("negative_export") for item in used)),
            "unresolved_material_quantity": unresolved_material_quantity,
            "material_quantity": material_quantity,
            "finished_quantity": 0.0 if unresolved_material_quantity > QUANTITY_EPSILON else clean_quantity(remaining),
            "finished_variant_code": finished_variant_code,
            "sale_split_codes": ", ".join(dict.fromkeys(
                item.get("ledger_variant_code") or item.get("purchase_variant_code") or item.get("variant_code", "")
                for item in used
                if not item.get("barem_unallocated")
            )),
            "used": used,
            "rejected": rejected,
            "generic_plan_note": generic_plan_note,
            "inventory_before": inventory_before,
                "inventory_after": inventory_snapshot(
                sorted(
                    [lot for lot in lot_pool if lot_available_for_sale(lot, sale)],
                    # Snapshot after sale must include any future invoice that was treated as
                    # effective stock for this sale; otherwise the UI shows an artificial KTP/negative gap.
                    key=lambda lot: priority_fn(lot, sale["unit_price"]),
                ) + sorted(
                    [lot for lot in lot_pool if future_lot_allowed_for_sale(lot, sale, policy)],
                    key=lambda lot: priority_fn(lot, sale["unit_price"]),
                ),
                sale["unit_price"],
            ),
        })

    stock_rows = []
    grouped = defaultdict(lambda: {
        "opening_quantity": 0.0,
        "purchase_quantity": 0.0,
        "allocated_quantity": 0.0,
        "ending_quantity": 0.0,
        "costs": set(),
    })
    for lot in lots:
        item = grouped[lot["variant_code"]]
        item["base_code"] = lot["base_code"]
        item["variant_code"] = lot["variant_code"]
        item["product_name"] = item.get("product_name") or lot["product_name"]
        target = "opening_quantity" if lot["kind"] == "opening" else "purchase_quantity"
        item[target] = clean_quantity(item[target] + lot["initial_quantity"])
        item["allocated_quantity"] = clean_quantity(item["allocated_quantity"] + lot["allocated_quantity"])
        item["ending_quantity"] = clean_quantity(item["ending_quantity"] + lot["remaining_quantity"])
        if lot["unit_price"] is not None:
            item["costs"].add(lot["unit_price"])
    stocked_base_codes = set(by_code.keys())
    stocked_variant_codes = set(by_variant.keys())
    for sale in sales_lines:
        if sale["suffix"] is not None:
            has_stock = sale["variant_code"] in stocked_variant_codes
        else:
            has_stock = sale["base_code"] in stocked_base_codes
        if has_stock:
            continue
        item = grouped[sale["variant_code"]]
        item["base_code"] = sale["base_code"]
        item["variant_code"] = sale["variant_code"]
        item["product_name"] = item.get("product_name") or sale["product_name"]
    for item in grouped.values():
        item["unit_costs"] = ", ".join(format_number(value) for value in sorted(item.pop("costs")))
        stock_rows.append(item)
    stock_rows.sort(key=lambda item: (item["base_code"], item["variant_code"]))

    summary = {
        "opening_quantity": clean_quantity(sum(line["quantity"] for line in opening_lines)),
        "purchase_quantity": clean_quantity(sum(line["quantity"] for line in purchase_lines)),
        "sales_quantity": clean_quantity(sum(line["quantity"] for line in sales_lines)),
        "material_quantity": clean_quantity(sum(line["material_quantity"] for line in allocations)),
        "finished_quantity": clean_quantity(sum(line["finished_quantity"] for line in allocations)),
        "negative_export_quantity": clean_quantity(sum(
            used.get("quantity", 0)
            for line in allocations
            for used in line.get("used", [])
            if used.get("negative_export")
        )),
        "unresolved_material_quantity": clean_quantity(sum(
            line.get("unresolved_material_quantity", 0) or 0
            for line in allocations
        )),
    }
    summary["material_percent"] = (
        summary["material_quantity"] / summary["sales_quantity"] * 100
        if summary["sales_quantity"] else 0
    )
    summary["range_rejected_lines"] = sum(1 for line in allocations if line["rejected"])
    unique_missing_barem = list({item["variant_code"]: item for item in missing_barem_report}.values())
    summary["missing_barem_count"] = len(unique_missing_barem)
    summary["missing_barem_report"] = unique_missing_barem
    future_purchase_reorder_report = build_future_purchase_reorder_report(allocations)
    summary["future_purchase_reorder_count"] = len(future_purchase_reorder_report)
    summary["future_purchase_reorder_quantity"] = clean_quantity(sum(
        row.get("quantity", 0) or 0 for row in future_purchase_reorder_report
    ))
    summary["future_purchase_reorder_report"] = future_purchase_reorder_report
    return allocations, stock_rows, summary, list(dict.fromkeys(warnings))


def find_sale_only_codes(opening_lines, purchase_lines, sales_lines):
    stocked_base_codes = {
        line["base_code"]
        for line in opening_lines + purchase_lines
        if line["base_code"]
    }
    stocked_variant_codes = {
        line["variant_code"]
        for line in opening_lines + purchase_lines
        if line["variant_code"]
    }
    grouped = {}
    for line in sales_lines:
        if line["suffix"] is not None:
            has_stock = line["variant_code"] in stocked_variant_codes
        else:
            has_stock = line["base_code"] in stocked_base_codes
        if has_stock:
            continue
        key = line["variant_code"]
        item = grouped.setdefault(key, {
            "variant_code": line["variant_code"],
            "base_code": line["base_code"],
            "product_name": line["product_name"],
            "row_count": 0,
            "quantity": 0.0,
            "opening_quantity": 0.0,
            "purchase_quantity": 0.0,
            "rows": [],
        })
        item["row_count"] += 1
        item["quantity"] = clean_quantity(item["quantity"] + line["quantity"])
        item["rows"].append(line["row_number"])
        if not item["product_name"]:
            item["product_name"] = line["product_name"]
    result = list(grouped.values())
    result.sort(key=lambda item: (item["base_code"], item["variant_code"]))
    return result


def format_number(value):
    if value is None:
        return ""
    if abs(value - round(value)) < 0.0000001:
        return f"{int(round(value)):,}"
    return f"{value:,.4f}".rstrip("0").rstrip(".")


def clean_quantity(value):
    if abs(value) < QUANTITY_EPSILON:
        return 0.0
    return round(value, 10)


def detail_text(items):
    result = []
    for item in items:
        detail = f"{item['variant_code']}: {format_number(item['quantity'])}"
        source = item.get("source", "")
        invoice_no = item.get("invoice_no", "")
        if source:
            detail += f" từ {source}"
            if item.get("summary_count", 1) > 1:
                detail += f" ({item['summary_count']} dòng mua, giá vốn TB)"
            elif invoice_no:
                detail += f" HD {invoice_no}"
            else:
                detail += f" dòng {item.get('row_number', '')}"
        if item.get("profit_percent") is not None:
            detail += f" ({format_number(item['profit_percent'])}% lãi/lỗ)"
        result.append(detail)
    return "; ".join(result)


def inventory_text(items):
    if not items:
        return "Không có tồn kho liên quan"
    result = []
    for item in items:
        detail = f"{item['variant_code']}: {format_number(item['quantity'])}"
        detail += f" [{item['source']}"
        if item.get("summary_count", 1) > 1:
            detail += f" tổng hợp {item['summary_count']} dòng mua"
        elif item.get("invoice_no"):
            detail += f" HD {item['invoice_no']}"
        else:
            detail += f" dòng {item['row_number']}"
        detail += "]"
        if item.get("unit_cost") is not None:
            detail += f", vốn {format_number(item['unit_cost'])}"
        if item.get("profit_percent") is not None:
            detail += f", lãi/lỗ {format_number(item['profit_percent'])}%"
        result.append(detail)
    return "; ".join(result)


def rejected_text(items):
    return "; ".join(f"{item['variant_code']}: {item['reason']}" for item in items)


def build_future_purchase_reorder_report(allocations):
    grouped = {}
    for allocation in allocations or []:
        for used in allocation.get("used", []):
            if not used.get("future_purchase_reordered"):
                continue
            key = (
                used.get("purchase_variant_code", ""),
                used.get("invoice_no", ""),
                used.get("original_purchase_date_iso") or used.get("invoice_date_iso", ""),
                used.get("row_number", ""),
                allocation.get("invoice_no", ""),
                allocation.get("row_number", ""),
            )
            item = grouped.setdefault(key, {
                "purchase_variant_code": used.get("purchase_variant_code", ""),
                "sale_variant_code": allocation.get("variant_code", ""),
                "product_name": allocation.get("product_name", ""),
                "purchase_invoice_no": used.get("invoice_no", ""),
                "purchase_original_date": used.get("original_purchase_date") or used.get("invoice_date", ""),
                "purchase_original_date_iso": used.get("original_purchase_date_iso") or used.get("invoice_date_iso", ""),
                "effective_date": used.get("effective_purchase_date") or allocation.get("invoice_date", ""),
                "effective_date_iso": used.get("effective_purchase_date_iso") or allocation.get("invoice_date_iso", ""),
                "sale_invoice_no": allocation.get("invoice_no", ""),
                "sale_date": allocation.get("invoice_date", ""),
                "sale_date_iso": allocation.get("invoice_date_iso", ""),
                "sale_row_number": allocation.get("row_number", ""),
                "purchase_row_number": used.get("row_number", ""),
                "quantity": 0.0,
                "unit_cost": used.get("unit_cost"),
                "future_reorder_days": used.get("future_reorder_days"),
                "logic_note": used.get("logic_note", ""),
            })
            item["quantity"] = clean_quantity(item["quantity"] + clean_quantity(used.get("quantity", 0)))
    rows = list(grouped.values())
    rows.sort(key=lambda row: (
        row.get("effective_date_iso") or "",
        row.get("sale_invoice_no") or "",
        row.get("purchase_original_date_iso") or "",
        row.get("purchase_variant_code") or "",
    ))
    return rows


def replace_sheet(workbook, name):
    if name in workbook.sheetnames:
        del workbook[name]
    return workbook.create_sheet(name)


def style_report_sheet(sheet):
    header_fill = PatternFill("solid", fgColor="E7F2FF")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="12304B")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for idx, column in enumerate(sheet.columns, start=1):
        values = [len(str(cell.value or "")) for cell in list(column)[:80]]
        sheet.column_dimensions[get_column_letter(idx)].width = min(max(max(values, default=8) + 2, 12), 38)


def write_table(sheet, headers, rows, progress_callback=None, status_callback=None):
    sheet.append(headers)
    total = len(rows) if hasattr(rows, "__len__") else 0
    for index, row in enumerate(rows, start=1):
        sheet.append(row)
        if progress_callback and (index == 1 or index == total or index % 250 == 0):
            progress_callback(index, total)
    if status_callback:
        status_callback()
    style_report_sheet(sheet)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.####'
                cell.alignment = Alignment(horizontal="right")


PURCHASE_EXPORT_HEADERS = [
    "MST",
    "Tên công ty",
    "Số HĐ",
    "Ngày HĐ",
    "Mã kho",
    "Mã VT",
    "Số lượng",
    "Đơn giá mua",
    "Thành tiền mua",
]

SALES_EXPORT_HEADERS = [
    "Mã ST",
    "Người mua hàng\n(ong_ba)",
    "Số HD\n(so_ct)",
    "Ngày HD",
    "Số lượng:Q\n(so_luong)",
    "Giá bán:P0\n(gia2)",
    "Tiền bán:N0\n(tien2)",
    "Giá vốn:P0\n(gia)",
    "Tiền vốn:N0\n(tien)",
    "Thuế suất\n(thue_suat)",
    "Tiền thuế:N0\n(tien_thue)",
    "Mã kho\n(ma_kho)",
    "Mã vật tư\n(ma_vt)",
]


def line_total_amount(line):
    amount = line.get("line_amount")
    if amount is not None:
        return amount
    unit_price = line.get("unit_price")
    quantity = line.get("quantity")
    if unit_price is not None and quantity is not None:
        return unit_price * quantity
    return None


def line_unit_price(line):
    if line.get("unit_price") is not None:
        return line["unit_price"]
    amount = line_total_amount(line)
    quantity = line.get("quantity")
    if amount is not None and quantity:
        return amount / quantity
    return None


def amount_for_quantity(line, quantity):
    unit_price = line_unit_price(line)
    if unit_price is not None:
        return unit_price * quantity
    total_amount = line_total_amount(line)
    total_quantity = line.get("quantity")
    if total_amount is not None and total_quantity:
        return total_amount * quantity / total_quantity
    return None


def split_amounts_for_quantities(line, quantities):
    cleaned_quantities = [clean_quantity(quantity or 0) for quantity in quantities]
    total_amount = line_total_amount(line)
    if total_amount is None:
        return [amount_for_quantity(line, quantity) or 0 for quantity in cleaned_quantities]
    if not cleaned_quantities:
        return []
    assigned = 0
    amounts = []
    split_total_quantity = sum(cleaned_quantities)
    line_quantity = clean_quantity(line.get("quantity", 0) or 0)
    for index, quantity in enumerate(cleaned_quantities):
        if index == len(cleaned_quantities) - 1 and abs(split_total_quantity - line_quantity) <= QUANTITY_EPSILON:
            amount = total_amount - assigned
        else:
            amount = amount_for_quantity(line, quantity)
            if amount is None:
                base_quantity = split_total_quantity or line_quantity
                amount = total_amount * quantity / base_quantity if base_quantity else 0
        assigned += amount
        amounts.append(amount)
    return amounts


def split_tax_amounts_for_sale_amounts(line, sale_amounts):
    sale_amounts = [amount or 0 for amount in sale_amounts]
    total_tax = line.get("tax_amount")
    if total_tax is None:
        rate = sale_tax_rate(line)
        return [amount * rate / 100 for amount in sale_amounts]
    if not sale_amounts:
        return []
    assigned = 0
    tax_amounts = []
    total_sale_amount = sum(sale_amounts)
    for index, sale_amount in enumerate(sale_amounts):
        if index == len(sale_amounts) - 1:
            tax_amount = total_tax - assigned
        else:
            tax_amount = total_tax * sale_amount / total_sale_amount if total_sale_amount else 0
        assigned += tax_amount
        tax_amounts.append(tax_amount)
    return tax_amounts


def iso_in_line_range(*line_sets):
    lines = []
    for line_set in line_sets:
        lines.extend(line_set or [])
    values = sorted({line.get("invoice_date_iso", "") for line in lines if line.get("invoice_date_iso")})
    if not values:
        return {"from": "", "to": "", "from_display": "", "to_display": ""}
    return {
        "from": values[0],
        "to": values[-1],
        "from_display": date_display(values[0]),
        "to_display": date_display(values[-1]),
    }


def iso_in_sales_range(lines):
    return iso_in_line_range(lines)


def ledger_sort_key(row):
    row_date = row.get("date_iso") or "0000-00-00"
    type_order = {"opening": 0, "purchase": 1, "purchase_future_reorder": 1, "sale": 2}
    return (row_date, type_order.get(row.get("type"), 9), row.get("sequence", 0))


def combined_ledger_sort_key(row):
    row_date = row.get("date_iso") or "0000-00-00"
    type_order = {"opening": 0, "purchase": 1, "purchase_future_reorder": 1, "sale": 2}
    return (
        row_date,
        row.get("doc_no", ""),
        type_order.get(row.get("type"), 9),
        warehouse_sort_key(row.get("warehouse_code")),
        row.get("sequence", 0),
    )


MATERIAL_WAREHOUSE_CODE = "KVT"
LEGACY_MATERIAL_WAREHOUSE_CODES = {"KHH"}


def normalize_warehouse_code(value):
    code = normalize_code(value)
    return MATERIAL_WAREHOUSE_CODE if code in LEGACY_MATERIAL_WAREHOUSE_CODES else code


def warehouse_sort_key(warehouse_code):
    warehouse_order = {MATERIAL_WAREHOUSE_CODE: 0, "KHHVT": 0, "KTP": 1, "KHOCK": 2, "KHH": 0}
    code = normalize_warehouse_code(warehouse_code) or ""
    return (warehouse_order.get(code, 20), code)


def line_warehouse_code(line, fallback=MATERIAL_WAREHOUSE_CODE):
    fallback_code = normalize_warehouse_code(fallback) or MATERIAL_WAREHOUSE_CODE
    return normalize_warehouse_code(line.get("warehouse_code") or fallback_code) or fallback_code


def line_warehouse_account(line, fallback=""):
    return text(line.get("warehouse_account")) or fallback


def default_warehouse_name(code):
    code = normalize_warehouse_code(code)
    known = {
        MATERIAL_WAREHOUSE_CODE: "KHO VẬT TƯ, HÀNG HÓA",
        "KHH": "KHO VẬT TƯ, HÀNG HÓA",
        "KTP": "KHO THÀNH PHẨM",
        "KHHVT": "KHO HANG HOA VAT TU",
        "KHOCK": "KHO SAN PHAM CON LAI",
    }
    return known.get(code, f"KHO {code}")


def default_warehouse_account(code):
    code = normalize_warehouse_code(code)
    known = {
        MATERIAL_WAREHOUSE_CODE: "152",
        "KHH": "152",
        "KTP": "155",
        "KHHVT": "156",
        "KHOCK": "159",
    }
    return known.get(code, "")


def allocation_sale_warehouse_code(allocation, fallback):
    return line_warehouse_code(allocation, fallback)


def allocation_sale_warehouse_account(allocation, fallback=""):
    return line_warehouse_account(allocation, fallback or default_warehouse_account(allocation_sale_warehouse_code(allocation, "")))


def allocation_remainder_warehouse_code(allocation, fallback="KTP"):
    explicit = normalize_code(allocation.get("remainder_warehouse_code"))
    if explicit:
        return explicit
    sale_warehouse_code = allocation_sale_warehouse_code(allocation, "")
    if normalize_warehouse_code(sale_warehouse_code) == MATERIAL_WAREHOUSE_CODE:
        return normalize_code(fallback) or "KTP"
    return sale_warehouse_code or fallback


def allocation_remainder_warehouse_account(allocation, fallback=""):
    explicit = text(allocation.get("remainder_warehouse_account"))
    if explicit:
        return explicit
    warehouse_code = allocation_remainder_warehouse_code(allocation, "KTP")
    return line_warehouse_account(allocation, fallback or default_warehouse_account(warehouse_code))


def build_inventory_ledger(opening_lines, purchase_lines, allocations, sales_lines=None, company_profile="yen_thanh"):
    warehouses = {}
    material_warehouse_code = "KHHVT" if company_profile == "son_phuong" else MATERIAL_WAREHOUSE_CODE
    material_warehouse_account = "156" if company_profile == "son_phuong" else default_warehouse_account(material_warehouse_code)

    def ensure_warehouse(warehouse_code, account=""):
        code = normalize_warehouse_code(warehouse_code or MATERIAL_WAREHOUSE_CODE) or MATERIAL_WAREHOUSE_CODE
        item = warehouses.setdefault(code, {
            "warehouse_code": code,
            "warehouse_name": default_warehouse_name(code),
            "account": account or default_warehouse_account(code),
            "groups": {},
        })
        if account and not item.get("account"):
            item["account"] = account
        return item

    def group_for(warehouse_code, code, product_name="", unit_name=""):
        warehouse = ensure_warehouse(warehouse_code)
        groups = warehouse["groups"]
        item = groups.setdefault(code, {
            "variant_code": code,
            "product_name": product_name,
            "unit_name": unit_name,
            "account": warehouse["account"],
            "rows": [],
        })
        if not item.get("product_name") and product_name:
            item["product_name"] = product_name
        if not item.get("unit_name") and unit_name:
            item["unit_name"] = unit_name
        return item

    def remainder_logic_note(allocation, remainder_warehouse_code="KTP"):
        quantity = clean_quantity(allocation.get("quantity", 0))
        material_quantity = clean_quantity(allocation.get("material_quantity", 0))
        finished_quantity = clean_quantity(allocation.get("finished_quantity", 0))
        parts = [
            f"Khớp mua vào lấy {format_number(material_quantity)}/{format_number(quantity)}; kho {remainder_warehouse_code} lấy {format_number(finished_quantity)}."
        ]
        rejected = allocation.get("rejected", [])
        if rejected:
            details = []
            for item in rejected[:5]:
                margin = item.get("margin_percent")
                margin_text = f", lai/lo {format_number(margin)}%" if margin is not None else ""
                details.append(f"{item.get('variant_code', '')}: {item.get('reason', '')}{margin_text}")
            if len(rejected) > 5:
                details.append(f"+{len(rejected) - 5} lo khac")
            parts.append("Lô mua vào bị loại: " + "; ".join(details))
        elif material_quantity <= QUANTITY_EPSILON:
            parts.append("Không có tồn mua vào khả dụng theo ngày hóa đơn.")
        else:
            parts.append("Tồn mua vào không đủ số lượng theo ngày hóa đơn.")
        return " ".join(parts)

    def ktp_margin_percent(allocation):
        rejected = [
            item.get("margin_percent")
            for item in allocation.get("rejected", [])
            if item.get("margin_percent") is not None
        ]
        if not rejected:
            return ""
        return max(rejected, key=lambda value: abs(value))

    future_reorder_by_purchase = {}
    purchase_route_destinations = {}
    for allocation in allocations or []:
        sale_warehouse_code = allocation_sale_warehouse_code(allocation, MATERIAL_WAREHOUSE_CODE)
        sale_warehouse_account = allocation_sale_warehouse_account(
            allocation,
            default_warehouse_account(sale_warehouse_code),
        )
        for used in allocation.get("used", []):
            if company_profile != "son_phuong" and normalize_warehouse_code(sale_warehouse_code) == MATERIAL_WAREHOUSE_CODE:
                quantity = clean_quantity(used.get("quantity", 0))
                purchase_code = used.get("purchase_variant_code") or used.get("variant_code") or ""
                purchase_date = used.get("invoice_date_iso") or ""
                if quantity > QUANTITY_EPSILON and purchase_code:
                    purchase_route_destinations[(purchase_code, purchase_date)] = {
                        "warehouse_code": sale_warehouse_code,
                        "warehouse_account": sale_warehouse_account,
                    }
            if not used.get("future_purchase_reordered"):
                continue
            key = (used.get("purchase_variant_code", ""), used.get("original_purchase_date_iso") or used.get("invoice_date_iso", ""))
            sale_sequence = allocation.get("row_number", 0) * 1000 - 1
            candidate = {
                "effective_date": used.get("effective_purchase_date") or allocation.get("invoice_date", ""),
                "effective_date_iso": used.get("effective_purchase_date_iso") or allocation.get("invoice_date_iso", ""),
                "original_date": used.get("original_purchase_date") or used.get("invoice_date", ""),
                "original_date_iso": used.get("original_purchase_date_iso") or used.get("invoice_date_iso", ""),
                "sale_invoice_no": allocation.get("invoice_no", ""),
                "sale_date": allocation.get("invoice_date", ""),
                "sale_row_number": allocation.get("row_number", ""),
                "days": used.get("future_reorder_days"),
                "sequence": sale_sequence,
            }
            current = future_reorder_by_purchase.get(key)
            if not current or (candidate["effective_date_iso"], candidate["sequence"]) < (current["effective_date_iso"], current["sequence"]):
                future_reorder_by_purchase[key] = candidate

    def append_purchase_ledger_row(line, quantity, warehouse_code, account, row_type, row_date_iso, row_date, sequence, future_reorder=None, route_note=""):
        quantity = clean_quantity(quantity)
        if quantity <= QUANTITY_EPSILON:
            return
        unit_price = line_unit_price(line)
        line_amount = line_total_amount(line)
        total_quantity = clean_quantity(line.get("quantity", 0))
        amount = None
        if line_amount is not None and total_quantity > QUANTITY_EPSILON:
            amount = line_amount * quantity / total_quantity
        elif unit_price is not None:
            amount = unit_price * quantity
        logic_note = ""
        if future_reorder:
            logic_note = (
                f"HD mua vao goc ngay {future_reorder.get('original_date', '')} duoc dua len truoc "
                f"HD ban ra {future_reorder.get('sale_invoice_no', '')} ngay {future_reorder.get('sale_date', '')} "
                f"trong {future_reorder.get('days', '')} ngay."
            )
        if route_note:
            logic_note = f"{logic_note} {route_note}".strip()
        ensure_warehouse(warehouse_code, account)
        group_for(warehouse_code, line["variant_code"], line.get("product_name", ""), line.get("unit_name", ""))["rows"].append({
            "type": row_type,
            "date_iso": row_date_iso,
            "date": row_date,
            "doc_no": line.get("invoice_no", ""),
            "variant_code": line.get("variant_code", ""),
            "customer": line.get("party_name", ""),
            "description": f"Mua hàng nhập kho HD{line.get('invoice_no', '')}".strip(),
            "account": "331",
            "unit_price": unit_price,
            "sale_unit_price": "",
            "sale_amount": "",
            "qty_in": quantity,
            "amount_in": amount or 0,
            "qty_out": 0,
            "amount_out": 0,
            "future_purchase_reordered": bool(future_reorder),
            "original_date": future_reorder.get("original_date", "") if future_reorder else "",
            "original_date_iso": future_reorder.get("original_date_iso", "") if future_reorder else "",
            "logic_note": logic_note,
            "sequence": sequence,
        })

    for line in opening_lines or []:
        unit_price = line_unit_price(line)
        quantity = clean_quantity(line.get("quantity", 0))
        if quantity <= QUANTITY_EPSILON:
            continue
        amount = unit_price * quantity if unit_price is not None else 0
        warehouse_code = line_warehouse_code(line, material_warehouse_code)
        account = line_warehouse_account(line, material_warehouse_account or default_warehouse_account(warehouse_code))
        ensure_warehouse(warehouse_code, account)
        group_for(warehouse_code, line["variant_code"], line.get("product_name", ""), line.get("unit_name", ""))["rows"].append({
            "type": "opening",
            "date_iso": "",
            "date": "",
            "doc_no": line.get("invoice_no", ""),
            "variant_code": line.get("variant_code", ""),
            "customer": line.get("party_name", ""),
            "description": "Tồn đầu kỳ",
            "account": "",
            "unit_price": unit_price,
            "sale_unit_price": "",
            "sale_amount": "",
            "qty_in": quantity,
            "amount_in": amount,
            "qty_out": 0,
            "amount_out": 0,
            "sequence": line.get("row_number", 0),
        })

    for line in purchase_lines or []:
        quantity = clean_quantity(line.get("quantity", 0))
        if quantity <= QUANTITY_EPSILON:
            continue
        future_reorder = future_reorder_by_purchase.get((line.get("variant_code", ""), line.get("invoice_date_iso", "")))
        row_type = "purchase_future_reorder" if future_reorder else "purchase"
        row_date_iso = future_reorder.get("effective_date_iso") if future_reorder else line.get("invoice_date_iso", "")
        row_date = future_reorder.get("effective_date") if future_reorder else line.get("invoice_date", "")
        sequence = future_reorder.get("sequence") if future_reorder else line.get("row_number", 0)
        warehouse_code = line_warehouse_code(line, material_warehouse_code)
        account = line_warehouse_account(line, material_warehouse_account or default_warehouse_account(warehouse_code))
        route_key = (line.get("variant_code", ""), line.get("invoice_date_iso", ""))
        route = purchase_route_destinations.get(route_key)
        if route:
            route_warehouse_code = route.get("warehouse_code") or warehouse_code
            route_account = route.get("warehouse_account") or account or default_warehouse_account(route_warehouse_code)
            route_note = ""
            if normalize_code(route_warehouse_code) != normalize_code(warehouse_code):
                route_note = f"Dong mua vao duoc dua vao so kho {route_warehouse_code} vi da duoc dung lam gia von cho ban ra kho {route_warehouse_code}."
            append_purchase_ledger_row(
                line,
                quantity,
                route_warehouse_code,
                route_account,
                row_type,
                row_date_iso,
                row_date,
                sequence,
                future_reorder=future_reorder,
                route_note=route_note,
            )
        else:
            append_purchase_ledger_row(
                line,
                quantity,
                warehouse_code,
                account,
                row_type,
                row_date_iso,
                row_date,
                sequence,
                future_reorder=future_reorder,
            )

    def append_source_sale_ledger_row(allocation):
        quantity = clean_quantity(allocation.get("quantity", 0))
        if quantity <= QUANTITY_EPSILON:
            return
        sale_warehouse_code = allocation_sale_warehouse_code(allocation, MATERIAL_WAREHOUSE_CODE)
        sale_warehouse_account = allocation_sale_warehouse_account(allocation, default_warehouse_account(sale_warehouse_code))
        code = allocation.get("variant_code", "")
        ensure_warehouse(sale_warehouse_code, sale_warehouse_account)
        group = group_for(sale_warehouse_code, code, allocation.get("product_name", ""), allocation.get("unit_name", ""))
        sale_amount = line_total_amount(allocation)
        if sale_amount is None:
            sale_amount = amount_for_quantity(allocation, quantity) or 0
        rate = sale_tax_rate(allocation)
        tax_amount = line_tax_amount(allocation)
        if tax_amount is None:
            tax_amount = sale_amount * rate / 100 if sale_amount is not None else 0
        total_cost_quantity = 0
        total_cost_amount = 0
        logic_notes = []
        for used in allocation.get("used", []):
            used_quantity = clean_quantity(used.get("quantity", 0))
            if used_quantity <= QUANTITY_EPSILON:
                continue
            unit_cost = used.get("unit_cost")
            if unit_cost is None:
                continue
            total_cost_quantity = clean_quantity(total_cost_quantity + used_quantity)
            total_cost_amount += unit_cost * used_quantity
            if used.get("logic_note"):
                logic_notes.append(used.get("logic_note", ""))
        has_inbound = any((row.get("qty_in", 0) or 0) > QUANTITY_EPSILON for row in group.get("rows", []))
        cost_missing = not has_inbound or total_cost_quantity <= QUANTITY_EPSILON
        if cost_missing:
            unit_cost = 0
            cost_amount = 0
            missing_note = f"Kho {sale_warehouse_code} khong co nhap/ton dau ky cho ma nay; gia von de 0 va khong tinh % lai."
            logic_notes.append(missing_note)
        else:
            cost_amount = total_cost_amount
            unit_cost = cost_amount / quantity if quantity > QUANTITY_EPSILON else 0
            if total_cost_quantity + QUANTITY_EPSILON < quantity:
                logic_notes.append(
                    f"Chi khop duoc {format_number(total_cost_quantity)}/{format_number(quantity)} theo file mua vao; tien von tinh theo phan da khop."
                )
        group["rows"].append({
            "type": "sale",
            "date_iso": allocation.get("invoice_date_iso", ""),
            "date": allocation.get("invoice_date", ""),
            "doc_no": allocation.get("invoice_no", ""),
            "variant_code": code,
            "tax_code": allocation.get("party_tax_code", ""),
            "customer": allocation.get("party_name", ""),
            "description": f"Xuat ban cho Khach (kho {sale_warehouse_code})",
            "account": "6321",
            "unit_price": unit_cost,
            "sale_unit_price": line_unit_price(allocation),
            "sale_amount": sale_amount,
            "tax_rate": rate,
            "tax_amount": tax_amount,
            "total_amount": sale_amount + tax_amount,
            "qty_in": 0,
            "amount_in": 0,
            "qty_out": quantity,
            "amount_out": cost_amount,
            "logic_note": " ".join(dict.fromkeys(logic_notes)),
            "cost_missing": cost_missing,
            "sequence": allocation.get("row_number", 0) * 1000,
        })

    for allocation in allocations or []:
        if company_profile != "son_phuong" and normalize_code(allocation.get("warehouse_code", "")):
            append_source_sale_ledger_row(allocation)
            continue
        sale_split_entries = [
            ("used", used)
            for used in allocation.get("used", [])
            if clean_quantity(used.get("quantity", 0)) > QUANTITY_EPSILON
        ]
        finished_quantity = clean_quantity(allocation.get("finished_quantity", 0))
        if finished_quantity > QUANTITY_EPSILON:
            sale_split_entries.append(("finished", None))
        sale_split_quantities = [
            clean_quantity(item.get("quantity", 0)) if kind == "used" else finished_quantity
            for kind, item in sale_split_entries
        ]
        sale_split_amounts = split_amounts_for_quantities(allocation, sale_split_quantities)
        sale_split_tax_amounts = split_tax_amounts_for_sale_amounts(allocation, sale_split_amounts)

        for index, (kind, used) in enumerate(sale_split_entries):
            if kind != "used":
                continue
            quantity = clean_quantity(used.get("quantity", 0))
            if quantity <= QUANTITY_EPSILON:
                continue
            code = used.get("ledger_variant_code") or used.get("sale_variant_code") or allocation.get("variant_code", "")
            detail_code = used.get("purchase_variant_code") or used.get("variant_code") or code
            unit_price = used.get("unit_cost")
            sale_amount = sale_split_amounts[index] if index < len(sale_split_amounts) else (amount_for_quantity(allocation, quantity) or 0)
            sale_warehouse_code = allocation_sale_warehouse_code(allocation, MATERIAL_WAREHOUSE_CODE)
            sale_warehouse_account = allocation_sale_warehouse_account(allocation, default_warehouse_account(sale_warehouse_code))
            ensure_warehouse(sale_warehouse_code, sale_warehouse_account)
            group = group_for(sale_warehouse_code, code, allocation.get("product_name", ""), allocation.get("unit_name", ""))
            has_inbound = any((row.get("qty_in", 0) or 0) > QUANTITY_EPSILON for row in group.get("rows", []))
            cost_missing = not has_inbound
            if cost_missing:
                unit_price = 0
                amount = 0
            else:
                amount = unit_price * quantity if unit_price is not None else 0
            rate = sale_tax_rate(allocation)
            tax_amount = sale_split_tax_amounts[index] if index < len(sale_split_tax_amounts) else (sale_amount * rate / 100 if sale_amount is not None else 0)
            logic_note = used.get("logic_note", "")
            if cost_missing:
                missing_note = f"Kho {sale_warehouse_code} không có nhập/tồn đầu kỳ cho mã này; giá vốn để 0 và không tính % lãi."
                logic_note = f"{logic_note} {missing_note}".strip()
            group["rows"].append({
                "type": "sale",
                "date_iso": allocation.get("invoice_date_iso", ""),
                "date": allocation.get("invoice_date", ""),
                "doc_no": allocation.get("invoice_no", ""),
                "variant_code": detail_code,
                "tax_code": allocation.get("party_tax_code", ""),
                "customer": allocation.get("party_name", ""),
                "description": f"Xuất bán cho Khách (kho {sale_warehouse_code})",
                "account": "6321",
                "unit_price": unit_price,
                "sale_unit_price": line_unit_price(allocation),
                "sale_amount": sale_amount,
                "tax_rate": rate,
                "tax_amount": tax_amount,
                "total_amount": sale_amount + tax_amount,
                "qty_in": 0,
                "amount_in": 0,
                "qty_out": quantity,
                "amount_out": amount,
                "logic_note": logic_note,
                "future_purchase_reordered": bool(used.get("future_purchase_reordered")),
                "cost_missing": cost_missing,
                "sequence": allocation.get("row_number", 0) * 1000 + index,
            })
        if finished_quantity > QUANTITY_EPSILON:
            finished_index = len(sale_split_entries) - 1
            code = allocation.get("finished_variant_code") or allocation.get("variant_code", "")
            unit_price = 0
            sale_warehouse_code = allocation_remainder_warehouse_code(allocation, "KTP")
            sale_warehouse_account = allocation_remainder_warehouse_account(allocation, default_warehouse_account(sale_warehouse_code))
            logic_note = remainder_logic_note(allocation, sale_warehouse_code)
            ensure_warehouse(sale_warehouse_code, sale_warehouse_account)
            group = group_for(sale_warehouse_code, code, allocation.get("product_name", ""), allocation.get("unit_name", ""))
            amount = 0
            sale_amount = sale_split_amounts[finished_index] if finished_index < len(sale_split_amounts) else (amount_for_quantity(allocation, finished_quantity) or 0)
            rate = sale_tax_rate(allocation)
            tax_amount = sale_split_tax_amounts[finished_index] if finished_index < len(sale_split_tax_amounts) else (sale_amount * rate / 100 if sale_amount is not None else 0)
            group["rows"].append({
                "type": "sale",
                "date_iso": allocation.get("invoice_date_iso", ""),
                "date": allocation.get("invoice_date", ""),
                "doc_no": allocation.get("invoice_no", ""),
                "variant_code": code,
                "tax_code": allocation.get("party_tax_code", ""),
                "customer": allocation.get("party_name", ""),
                "description": f"Xuất bán cho Khách (kho {sale_warehouse_code})",
                "account": "6321",
                "unit_price": unit_price,
                "sale_unit_price": line_unit_price(allocation),
                "sale_amount": sale_amount,
                "tax_rate": rate,
                "tax_amount": tax_amount,
                "total_amount": sale_amount + tax_amount,
                "qty_in": 0,
                "amount_in": 0,
                "qty_out": finished_quantity,
                "amount_out": amount,
                "logic_note": logic_note,
                "margin_percent": 0,
                "cost_missing": True,
                "sequence": allocation.get("row_number", 0) * 1000 + 999,
            })

    result = []
    for warehouse in warehouses.values():
        groups = []
        for group in warehouse["groups"].values():
            group["rows"].sort(key=ledger_sort_key)
            if not group["unit_name"]:
                group["unit_name"] = "-"
            groups.append(group)
        groups.sort(key=lambda item: (item["variant_code"], item.get("product_name", "")))
        result.append({
            "warehouse_code": warehouse["warehouse_code"],
            "warehouse_name": warehouse["warehouse_name"],
            "account": warehouse["account"],
            "groups": groups,
        })
    result.sort(key=lambda warehouse: warehouse_sort_key(warehouse.get("warehouse_code")))
    return {
        "warehouses": result,
        "warehouse_code": result[0]["warehouse_code"] if result else MATERIAL_WAREHOUSE_CODE,
        "warehouse_name": result[0]["warehouse_name"] if result else default_warehouse_name(MATERIAL_WAREHOUSE_CODE),
        "account": result[0]["account"] if result else default_warehouse_account(MATERIAL_WAREHOUSE_CODE),
        "groups": result[0]["groups"] if result else [],
    }


def average_costs_by_variant(purchase_lines, sales_lines):
    totals = defaultdict(lambda: {"amount": 0.0, "quantity": 0.0})
    for line in (purchase_lines or []) + (sales_lines or []):
        quantity = line.get("quantity")
        amount = line_total_amount(line)
        if not line.get("variant_code") or quantity is None or quantity <= QUANTITY_EPSILON or amount is None:
            continue
        item = totals[line["variant_code"]]
        item["amount"] += amount
        item["quantity"] += quantity
    return {
        code: item["amount"] / item["quantity"]
        for code, item in totals.items()
        if item["quantity"] > QUANTITY_EPSILON
    }


def average_cost_by_variant_until(lines, variant_code, date_iso):
    total_amount = 0.0
    total_quantity = 0.0
    for line in lines or []:
        if line.get("variant_code") != variant_code:
            continue
        line_date = line.get("invoice_date_iso")
        if date_iso and line_date and line_date > date_iso:
            continue
        quantity = line.get("quantity")
        amount = line_total_amount(line)
        if quantity is None or quantity <= QUANTITY_EPSILON or amount is None:
            continue
        total_amount += amount
        total_quantity += quantity
    if total_quantity > QUANTITY_EPSILON:
        return total_amount / total_quantity
    return None


def build_purchase_export_rows(purchase_lines):
    rows = []
    for line in purchase_lines or []:
        quantity = line.get("quantity")
        if quantity is None or quantity <= QUANTITY_EPSILON:
            continue
        unit_price = line_unit_price(line)
        amount = line_total_amount(line)
        rows.append([
            line.get("party_tax_code", ""),
            line.get("party_name", ""),
            line.get("invoice_no", ""),
            line.get("invoice_date", ""),
            line_warehouse_code(line, MATERIAL_WAREHOUSE_CODE),
            line.get("variant_code", ""),
            clean_quantity(quantity),
            unit_price,
            amount,
        ])
    return rows


def sale_tax_rate(line):
    rate = line.get("tax_rate_percent")
    if rate is not None:
        return rate
    amount = line_total_amount(line)
    tax_amount = line.get("tax_amount")
    if amount:
        return tax_amount / amount * 100 if tax_amount is not None else 0
    return 0


def build_sales_export_row(line, quantity, warehouse_code, variant_code, average_costs, fallback_cost=None):
    quantity = clean_quantity(quantity)
    sale_price = line_unit_price(line)
    sale_amount = amount_for_quantity(line, quantity)
    cost = first_present(fallback_cost, average_costs.get(variant_code), average_costs.get(line.get("variant_code")))
    cost_amount = cost * quantity if cost is not None else None
    rate = sale_tax_rate(line)
    tax_amount = sale_amount * rate / 100 if sale_amount is not None else None
    return [
        line.get("party_tax_code", ""),
        line.get("party_name", ""),
        line.get("invoice_no", ""),
        line.get("invoice_date", ""),
        quantity,
        sale_price,
        sale_amount,
        cost,
        cost_amount,
        rate,
        tax_amount,
        warehouse_code,
        variant_code,
    ]


def build_sales_export_rows(allocations, purchase_lines=None, sales_lines=None):
    average_costs = average_costs_by_variant(purchase_lines, [])
    rows = []
    for line in allocations:
        for used in line.get("used", []):
            quantity = used.get("quantity", 0)
            if quantity <= QUANTITY_EPSILON:
                continue
            warehouse_code = allocation_sale_warehouse_code(line, MATERIAL_WAREHOUSE_CODE)
            rows.append(build_sales_export_row(
                line,
                quantity,
                warehouse_code,
                used.get("ledger_variant_code") or used.get("purchase_variant_code") or used.get("variant_code") or line.get("variant_code", ""),
                average_costs,
                used.get("unit_cost"),
            ))
        finished_quantity = line.get("finished_quantity", 0)
        if finished_quantity > QUANTITY_EPSILON:
            warehouse_code = allocation_sale_warehouse_code(line, "KTP")
            rows.append(build_sales_export_row(
                line,
                finished_quantity,
                warehouse_code,
                line.get("finished_variant_code") or line.get("variant_code", ""),
                average_costs,
                0,
            ))
    return rows


def build_sales_report_row(line, quantity, warehouse_code, variant_code, average_costs, fallback_cost=None):
    quantity = clean_quantity(quantity)
    sale_price = line_unit_price(line)
    sale_amount = amount_for_quantity(line, quantity) or 0
    cost = first_present(fallback_cost, average_costs.get(variant_code), average_costs.get(line.get("variant_code")))
    cost_amount = cost * quantity if cost is not None else 0
    profit_amount = sale_amount - cost_amount
    rate = sale_tax_rate(line)
    tax_amount = sale_amount * rate / 100
    return {
        "tax_code": line.get("party_tax_code", ""),
        "customer": line.get("party_name", ""),
        "invoice_no": line.get("invoice_no", ""),
        "invoice_date": line.get("invoice_date", ""),
        "invoice_date_iso": line.get("invoice_date_iso", ""),
        "warehouse_code": warehouse_code,
        "variant_code": variant_code,
        "product_name": line.get("product_name", ""),
        "unit_name": line.get("unit_name", ""),
        "quantity": quantity,
        "sale_price": sale_price or 0,
        "sale_amount": sale_amount,
        "cost_price": cost or 0,
        "cost_amount": cost_amount,
        "profit_amount": profit_amount,
        "tax_rate": rate,
        "tax_amount": tax_amount,
        "total_amount": sale_amount + tax_amount,
        "row_number": line.get("row_number", 0),
    }


def build_sales_report_rows_from_ledger(ledger):
    rows = []
    for warehouse in (ledger or {}).get("warehouses", []):
        warehouse_code = warehouse.get("warehouse_code", "")
        for group in warehouse.get("groups", []):
            for row in group.get("rows", []):
                if row.get("type") != "sale":
                    continue
                quantity = clean_quantity(row.get("qty_out", 0))
                if quantity <= QUANTITY_EPSILON:
                    continue
                sale_amount = row.get("sale_amount")
                if sale_amount in (None, ""):
                    sale_unit_price = row.get("sale_unit_price") or 0
                    sale_amount = sale_unit_price * quantity
                cost_amount = row.get("amount_out", 0) or 0
                cost_missing = bool(row.get("cost_missing"))
                cost_price = 0 if cost_missing else (cost_amount / quantity if quantity else 0)
                profit_amount = 0 if cost_missing else (sale_amount or 0) - cost_amount
                tax_amount = row.get("tax_amount") if row.get("tax_amount") not in (None, "") else 0
                rows.append({
                    "tax_code": row.get("tax_code", ""),
                    "customer": row.get("customer", ""),
                    "invoice_no": row.get("doc_no", ""),
                    "invoice_date": row.get("date", "") or date_display(row.get("date_iso", "")),
                    "invoice_date_iso": row.get("date_iso", ""),
                    "warehouse_code": warehouse_code,
                    "variant_code": group.get("variant_code", ""),
                    "product_name": group.get("product_name", ""),
                    "unit_name": group.get("unit_name", ""),
                    "quantity": quantity,
                    "sale_price": row.get("sale_unit_price") or ((sale_amount or 0) / quantity if quantity else 0),
                    "sale_amount": sale_amount or 0,
                    "cost_price": cost_price,
                    "cost_amount": cost_amount,
                    "profit_amount": profit_amount,
                    "tax_rate": row.get("tax_rate", 0) or 0,
                    "tax_amount": tax_amount or 0,
                    "total_amount": row.get("total_amount", (sale_amount or 0) + (tax_amount or 0)) or 0,
                    "row_number": row.get("sequence", 0),
                    "cost_missing": cost_missing,
                })
    rows.sort(key=lambda row: (row.get("invoice_date_iso") or "", text(row.get("invoice_no")), row.get("row_number", 0), row.get("warehouse_code", ""), row.get("variant_code", "")))
    return rows


def build_sales_report_rows(allocations, purchase_lines=None, sales_lines=None):
    average_costs = average_costs_by_variant(purchase_lines, [])
    rows = []
    for line in allocations:
        for used in line.get("used", []):
            quantity = used.get("quantity", 0)
            if quantity <= QUANTITY_EPSILON:
                continue
            warehouse_code = allocation_sale_warehouse_code(line, MATERIAL_WAREHOUSE_CODE)
            rows.append(build_sales_report_row(
                line,
                quantity,
                warehouse_code,
                used.get("ledger_variant_code") or used.get("purchase_variant_code") or used.get("variant_code") or line.get("variant_code", ""),
                average_costs,
                used.get("unit_cost"),
            ))
        finished_quantity = line.get("finished_quantity", 0)
        if finished_quantity > QUANTITY_EPSILON:
            warehouse_code = allocation_sale_warehouse_code(line, "KTP")
            rows.append(build_sales_report_row(
                line,
                finished_quantity,
                warehouse_code,
                line.get("finished_variant_code") or line.get("variant_code", ""),
                average_costs,
                0,
            ))
    rows.sort(key=lambda row: (row.get("invoice_date_iso") or "", text(row.get("invoice_no")), row.get("row_number", 0), row.get("variant_code", "")))
    return rows


def sales_report_summary_key(row):
    return "|||".join([
        text(row.get("warehouse_code")),
        text(row.get("variant_code")),
        text(row.get("product_name")),
        text(row.get("unit_name")),
    ])


def inventory_report_summary_key(warehouse_code, variant_code):
    return "|||".join([text(warehouse_code), text(variant_code)])


def sales_summary_rows_for_ui(sales_report_rows):
    grouped = {}
    for row in sales_report_rows or []:
        key = sales_report_summary_key(row)
        item = grouped.setdefault(key, {
            "key": key,
            "warehouse_code": row.get("warehouse_code", ""),
            "variant_code": row.get("variant_code", ""),
            "product_name": row.get("product_name", ""),
            "unit_name": row.get("unit_name", ""),
            "quantity": 0.0,
            "cost_amount": 0.0,
            "sale_amount": 0.0,
            "profit_amount": 0.0,
            "tax_amount": 0.0,
            "total_amount": 0.0,
            "row_count": 0,
        })
        item["quantity"] += row.get("quantity", 0) or 0
        item["cost_amount"] += row.get("cost_amount", 0) or 0
        item["sale_amount"] += row.get("sale_amount", 0) or 0
        cost_missing = bool(row.get("cost_missing"))
        item.setdefault("margin_sale_amount", 0.0)
        item.setdefault("cost_missing_count", 0)
        item["profit_amount"] += 0 if cost_missing else (row.get("profit_amount", (row.get("sale_amount", 0) or 0) - (row.get("cost_amount", 0) or 0)) or 0)
        if cost_missing:
            item["cost_missing_count"] += 1
        else:
            item["margin_sale_amount"] += row.get("sale_amount", 0) or 0
        item["tax_amount"] += row.get("tax_amount", 0) or 0
        item["total_amount"] += row.get("total_amount", 0) or 0
        item["row_count"] += 1
    rows = sorted(grouped.values(), key=lambda item: (warehouse_sort_key(item.get("warehouse_code")), item.get("variant_code", ""), item.get("product_name", "")))
    for index, item in enumerate(rows, start=1):
        item["index"] = index
        item["margin_percent"] = margin_percent_from_profit(item.get("profit_amount"), item.get("margin_sale_amount"))
    return rows


def sales_detail_rows_for_ui(sales_report_rows):
    rows = []
    for index, row in enumerate(sales_report_rows or [], start=1):
        rows.append({
            "key": f"{sales_report_summary_key(row)}|||{index}",
            "summary_key": sales_report_summary_key(row),
            "warehouse_code": row.get("warehouse_code", ""),
            "invoice_date": row.get("invoice_date", ""),
            "invoice_date_iso": row.get("invoice_date_iso", ""),
            "invoice_no": row.get("invoice_no", ""),
            "customer": row.get("customer", ""),
            "tax_code": row.get("tax_code", ""),
            "variant_code": row.get("variant_code", ""),
            "product_name": row.get("product_name", ""),
            "unit_name": row.get("unit_name", ""),
            "quantity": row.get("quantity", 0) or 0,
            "sale_price": row.get("sale_price", 0) or 0,
            "sale_amount": row.get("sale_amount", 0) or 0,
            "cost_price": row.get("cost_price", 0) or 0,
            "cost_amount": row.get("cost_amount", 0) or 0,
            "profit_amount": row.get("profit_amount", 0) or 0,
            "tax_rate": row.get("tax_rate", 0) or 0,
            "tax_amount": row.get("tax_amount", 0) or 0,
            "total_amount": row.get("total_amount", 0) or 0,
            "cost_missing": bool(row.get("cost_missing")),
        })
    rows.sort(key=lambda row: (warehouse_sort_key(row.get("warehouse_code")), row.get("invoice_date_iso") or "", text(row.get("invoice_no")), row.get("variant_code", "")))
    return rows


def inventory_summary_rows_for_ui(ledger):
    result = []
    date_range = (ledger or {}).get("date_range", {})
    from_date = date_range.get("from", "")
    to_date = date_range.get("to", "")
    for warehouse in sorted((ledger or {}).get("warehouses", []), key=lambda item: warehouse_sort_key(item.get("warehouse_code"))):
        warehouse_code = warehouse.get("warehouse_code", "")
        for group in warehouse.get("groups", []):
            view = ledger_view_for_range(group, from_date, to_date)
            if not has_ledger_activity(view):
                continue
            variant_code = group.get("variant_code", "")
            result.append({
                "key": inventory_report_summary_key(warehouse_code, variant_code),
                "warehouse_code": warehouse_code,
                "warehouse_name": warehouse.get("warehouse_name", ""),
                "account": group.get("account") or warehouse.get("account") or "",
                "variant_code": variant_code,
                "product_name": group.get("product_name", ""),
                "unit_name": group.get("unit_name", ""),
                "opening_qty": view["opening_qty"],
                "opening_amount": view["opening_amount"],
                "in_qty": view["in_qty"],
                "in_amount": view["in_amount"],
                "out_qty": view["out_qty"],
                "out_amount": view["out_amount"],
                "ending_qty": view["ending_qty"],
                "ending_amount": view["ending_amount"],
                "row_count": len(view["period_rows"]),
            })
    result.sort(key=lambda row: (warehouse_sort_key(row.get("warehouse_code")), row.get("variant_code", ""), row.get("product_name", "")))
    for index, row in enumerate(result, start=1):
        row["index"] = index
    return result


def ledger_detail_rows_for_ui(ledger):
    result = []
    date_range = (ledger or {}).get("date_range", {})
    from_date = date_range.get("from", "")
    to_date = date_range.get("to", "")
    for warehouse in sorted((ledger or {}).get("warehouses", []), key=lambda item: warehouse_sort_key(item.get("warehouse_code"))):
        warehouse_code = warehouse.get("warehouse_code", "")
        for group in warehouse.get("groups", []):
            variant_code = group.get("variant_code", "")
            summary_key = inventory_report_summary_key(warehouse_code, variant_code)
            view = ledger_view_for_range(group, from_date, to_date)
            if not has_ledger_activity(view):
                continue
            running_qty = view["opening_qty"]
            running_amount = view["opening_amount"]
            if abs(running_qty) > QUANTITY_EPSILON or abs(running_amount) > QUANTITY_EPSILON:
                result.append({
                    "key": f"{summary_key}|||opening",
                    "summary_key": summary_key,
                    "row_type": "opening",
                    "warehouse_code": warehouse_code,
                    "variant_code": variant_code,
                    "product_name": group.get("product_name", ""),
                    "unit_name": group.get("unit_name", ""),
                    "date": "",
                    "date_iso": "",
                    "doc_no": "",
                    "customer": "",
                    "description": "Tồn đầu kỳ",
                    "account": "",
                    "unit_price": "",
                    "sale_unit_price": "",
                    "sale_amount": "",
                    "tax_rate": "",
                    "tax_amount": "",
                    "total_amount": "",
                    "cost_missing": False,
                    "qty_in": running_qty,
                    "amount_in": running_amount,
                    "qty_out": 0,
                    "amount_out": 0,
                    "running_qty": running_qty,
                    "running_amount": running_amount,
                    "logic_note": "",
                })
            for index, row in enumerate(view["period_rows"], start=1):
                running_qty += (row.get("qty_in", 0) or 0) - (row.get("qty_out", 0) or 0)
                running_amount += (row.get("amount_in", 0) or 0) - (row.get("amount_out", 0) or 0)
                result.append({
                    "key": f"{summary_key}|||{index}",
                    "summary_key": summary_key,
                    "row_type": row.get("type", ""),
                    "warehouse_code": warehouse_code,
                    "variant_code": row.get("variant_code", variant_code),
                    "product_name": group.get("product_name", ""),
                    "unit_name": group.get("unit_name", ""),
                    "date": row.get("date", "") or date_display(row.get("date_iso", "")),
                    "date_iso": row.get("date_iso", ""),
                    "doc_no": row.get("doc_no", ""),
                    "customer": row.get("customer", ""),
                    "description": row.get("description", ""),
                    "account": row.get("account", ""),
                    "unit_price": row.get("unit_price", ""),
                    "sale_unit_price": row.get("sale_unit_price", ""),
                    "sale_amount": row.get("sale_amount", ""),
                    "tax_rate": row.get("tax_rate", ""),
                    "tax_amount": row.get("tax_amount", ""),
                    "total_amount": row.get("total_amount", ""),
                    "cost_missing": bool(row.get("cost_missing")),
                    "qty_in": row.get("qty_in", 0) or 0,
                    "amount_in": row.get("amount_in", 0) or 0,
                    "qty_out": row.get("qty_out", 0) or 0,
                    "amount_out": row.get("amount_out", 0) or 0,
                    "running_qty": running_qty,
                    "running_amount": running_amount,
                    "logic_note": row.get("logic_note", ""),
                })
    return result


def report_view_for_ui(ledger, sales_report_rows):
    warehouse_map = {}
    for warehouse in (ledger or {}).get("warehouses", []):
        code = warehouse.get("warehouse_code", "")
        if code:
            warehouse_map[code] = {
                "warehouse_code": code,
                "warehouse_name": warehouse.get("warehouse_name", ""),
                "account": warehouse.get("account", ""),
            }
    for row in sales_report_rows or []:
        code = row.get("warehouse_code", "")
        if code and code not in warehouse_map:
            warehouse_map[code] = {
                "warehouse_code": code,
                "warehouse_name": default_warehouse_name(code),
                "account": default_warehouse_account(code),
            }
    warehouses = sorted(warehouse_map.values(), key=lambda item: warehouse_sort_key(item.get("warehouse_code")))
    return {
        "date_range": (ledger or {}).get("date_range", {}),
        "warehouses": warehouses,
        "sales_summary_rows": sales_summary_rows_for_ui(sales_report_rows),
        "sales_detail_rows": sales_detail_rows_for_ui(sales_report_rows),
        "inventory_summary_rows": inventory_summary_rows_for_ui(ledger),
        "ledger_detail_rows": ledger_detail_rows_for_ui(ledger),
    }


def format_export_sheet(sheet, quantity_columns, money_columns, percent_columns=()):
    for row in sheet.iter_rows(min_row=2):
        for column in quantity_columns:
            row[column - 1].number_format = '#,##0.####'
        for column in money_columns:
            row[column - 1].number_format = '#,##0'
        for column in percent_columns:
            row[column - 1].number_format = '+0.####;-0.####;0'


def fill_ledger_activity_rows(sheet, in_qty_column, out_qty_column, warehouse_column=None, default_warehouse=""):
    fills = {
        (MATERIAL_WAREHOUSE_CODE, "in"): PatternFill("solid", fgColor="E6F1FF"),
        (MATERIAL_WAREHOUSE_CODE, "out"): PatternFill("solid", fgColor="FFF0D7"),
        ("KTP", "in"): PatternFill("solid", fgColor="E9FBFF"),
        ("KTP", "out"): PatternFill("solid", fgColor="FFE7E1"),
        ("", "in"): PatternFill("solid", fgColor="EAF3FF"),
        ("", "out"): PatternFill("solid", fgColor="FFF1DF"),
        ("", "mixed"): PatternFill("solid", fgColor="F3ECFF"),
    }
    for row in sheet.iter_rows(min_row=2):
        in_qty = row[in_qty_column - 1].value or 0
        out_qty = row[out_qty_column - 1].value or 0
        if not isinstance(in_qty, (int, float)) and not isinstance(out_qty, (int, float)):
            continue
        warehouse_code = default_warehouse
        if warehouse_column:
            warehouse_code = text(row[warehouse_column - 1].value).upper() or warehouse_code
        warehouse_code = normalize_warehouse_code(warehouse_code)
        warehouse_code = warehouse_code if warehouse_code in {MATERIAL_WAREHOUSE_CODE, "KTP"} else ""
        fill = None
        if isinstance(in_qty, (int, float)) and abs(in_qty) > QUANTITY_EPSILON:
            fill = fills.get((warehouse_code, "in"), fills[("", "in")])
        if isinstance(out_qty, (int, float)) and abs(out_qty) > QUANTITY_EPSILON:
            fill = fills[("", "mixed")] if fill else fills.get((warehouse_code, "out"), fills[("", "out")])
        if fill:
            for cell in row:
                cell.fill = fill


def fill_future_reorder_rows(sheet):
    future_fill = PatternFill("solid", fgColor="F3E8FF")
    for row in sheet.iter_rows(min_row=2):
        row_text = " ".join(text(cell.value).lower() for cell in row)
        if "dua len truoc" in row_text or "đưa lên trước" in row_text:
            for cell in row:
                cell.fill = future_fill


LEDGER_EXPORT_HEADERS = [
    "Ngày",
    "Số",
    "Khách hàng",
    "Diễn giải",
    "TK đối ứng",
    "Đơn giá",
    "Nhập số lượng",
    "Nhập thành tiền",
    "Xuất số lượng",
    "Xuất thành tiền",
    "Ton so luong",
    "Ton thanh tien",
]


def ledger_export_rows(warehouse):
    rows = []
    for group in warehouse.get("groups", []):
        rows.append([
            f"KHO: {warehouse.get('warehouse_code', '')} - {warehouse.get('warehouse_name', '')}",
            "", "", "", "", "", "", "", "", "", "", "",
        ])
        rows.append([
            f"Vật tư: {group.get('variant_code', '')} - {group.get('product_name', '')}, Đvt: {group.get('unit_name', '-')}, TK: {group.get('account') or warehouse.get('account') or ''}",
            "", "", "", "", "", "", "", "", "", "", "",
        ])
        opening_rows = [row for row in group.get("rows", []) if row.get("type") == "opening"]
        period_rows = [row for row in group.get("rows", []) if row.get("type") != "opening"]
        opening_qty = sum(row.get("qty_in", 0) or 0 for row in opening_rows) - sum(row.get("qty_out", 0) or 0 for row in opening_rows)
        opening_amount = sum(row.get("amount_in", 0) or 0 for row in opening_rows) - sum(row.get("amount_out", 0) or 0 for row in opening_rows)
        in_qty = sum(row.get("qty_in", 0) or 0 for row in period_rows)
        in_amount = sum(row.get("amount_in", 0) or 0 for row in period_rows)
        out_qty = sum(row.get("qty_out", 0) or 0 for row in period_rows)
        out_amount = sum(row.get("amount_out", 0) or 0 for row in period_rows)
        running_qty = opening_qty
        running_amount = opening_amount
        summary_rows = [
            ["", "", "", "Tồn đầu kỳ", "", "", opening_qty, opening_amount, "", "", opening_qty, opening_amount],
            ["", "", "", "Nhập trong kỳ", "", "", in_qty, in_amount, "", "", "", ""],
            ["", "", "", "Xuất trong kỳ", "", "", "", "", out_qty, out_amount, "", ""],
            ["", "", "", "Tồn cuối kỳ", "", "", opening_qty + in_qty - out_qty, opening_amount + in_amount - out_amount, "", "", opening_qty + in_qty - out_qty, opening_amount + in_amount - out_amount],
        ]
        rows.extend(summary_rows)
        for row in period_rows:
            running_qty += (row.get("qty_in", 0) or 0) - (row.get("qty_out", 0) or 0)
            running_amount += (row.get("amount_in", 0) or 0) - (row.get("amount_out", 0) or 0)
            description = row.get("description", "")
            if row.get("future_purchase_reordered") and row.get("logic_note"):
                description = f"{description} - {row.get('logic_note')}"
            rows.append([
                row.get("date", "") or date_display(row.get("date_iso", "")),
                row.get("doc_no", ""),
                row.get("customer", ""),
                description,
                row.get("account", ""),
                row.get("unit_price", ""),
                row.get("qty_in", 0),
                row.get("amount_in", 0),
                row.get("qty_out", 0),
                row.get("amount_out", 0),
                running_qty,
                running_amount,
            ])
    return rows


def write_ledger_export_sheet(workbook, warehouse, progress_callback=None, status_callback=None):
    sheet_name = f"SoChiTiet{warehouse.get('warehouse_code', '')}"[:31]
    sheet = replace_sheet(workbook, sheet_name)
    rows = ledger_export_rows(warehouse)
    write_table(sheet, LEDGER_EXPORT_HEADERS, rows, progress_callback=progress_callback, status_callback=status_callback)
    for row in sheet.iter_rows(min_row=2):
        first_value = text(row[0].value)
        if first_value.startswith("KHO:") or first_value.startswith("Vật tư:"):
            for cell in row:
                cell.font = Font(bold=True)
            row[0].alignment = Alignment(wrap_text=True)
        if text(row[3].value) in {"Tồn đầu kỳ", "Nhập trong kỳ", "Xuất trong kỳ", "Tồn cuối kỳ"}:
            for cell in row:
                cell.font = Font(bold=True)
    format_export_sheet(sheet, quantity_columns=(7, 9, 11), money_columns=(8, 10, 12))
    fill_ledger_activity_rows(sheet, 7, 9, default_warehouse=warehouse.get("warehouse_code", ""))
    fill_future_reorder_rows(sheet)
    sheet.column_dimensions["A"].width = 16
    sheet.column_dimensions["B"].width = 16
    sheet.column_dimensions["C"].width = 34
    sheet.column_dimensions["D"].width = 44
    sheet.column_dimensions["E"].width = 12
    sheet.column_dimensions["F"].width = 14
    sheet.column_dimensions["K"].width = 14
    sheet.column_dimensions["L"].width = 16
    return len(rows)


COMBINED_LEDGER_HEADERS = [
    "Ngày",
    "Số",
    "Khách hàng",
    "Diễn giải",
    "TK đối ứng",
    "Mã kho",
    "Mã VT chi tiết",
    "Đơn giá",
    "Nhập số lượng",
    "Nhập thành tiền",
    "Xuất số lượng",
    "Xuất thành tiền",
    "Đơn Giá bán",
    "TT bán",
    "Tỉ lệ lãi/lỗ",
    "Ton so luong",
    "Ton thanh tien",
    "Giai thich logic",
]


def combined_ledger_sections(ledger):
    sections = []
    grouped = {}
    order = []
    for warehouse in (ledger or {}).get("warehouses", []):
        for group in warehouse.get("groups", []):
            _, base_code, _ = code_parts(group.get("variant_code", ""))
            key = base_code
            if key not in grouped:
                grouped[key] = {
                    "warehouse": {"warehouse_code": "TONG", "warehouse_name": "TỔNG HỢP CÁC KHO", "account": ""},
                    "base_code": base_code,
                    "product_name": group.get("product_name", ""),
                    "unit_name": group.get("unit_name", ""),
                    "account": "",
                    "detail_codes": [],
                    "rows": [],
                }
                order.append(key)
            section = grouped[key]
            if not section.get("product_name") and group.get("product_name"):
                section["product_name"] = group.get("product_name", "")
            if not section.get("unit_name") and group.get("unit_name"):
                section["unit_name"] = group.get("unit_name", "")
            if group.get("variant_code") not in section["detail_codes"]:
                section["detail_codes"].append(group.get("variant_code", ""))
            for row in group.get("rows", []):
                section["rows"].append({**row, "warehouse_code": warehouse.get("warehouse_code", "")})
                detail_code = row.get("variant_code")
                if detail_code and detail_code not in section["detail_codes"]:
                    section["detail_codes"].append(detail_code)
    for key in order:
        section = grouped[key]
        section["rows"].sort(key=combined_ledger_sort_key)
        section["detail_codes"].sort()
        sections.append(section)
    return sections


def combined_ledger_export_rows(ledger):
    rows = []
    for section in combined_ledger_sections(ledger):
        warehouse = section["warehouse"]
        opening_rows = [row for row in section["rows"] if row.get("type") == "opening"]
        period_rows = [row for row in section["rows"] if row.get("type") != "opening"]
        opening_qty = sum(row.get("qty_in", 0) or 0 for row in opening_rows) - sum(row.get("qty_out", 0) or 0 for row in opening_rows)
        opening_amount = sum(row.get("amount_in", 0) or 0 for row in opening_rows) - sum(row.get("amount_out", 0) or 0 for row in opening_rows)
        in_qty = sum(row.get("qty_in", 0) or 0 for row in period_rows)
        in_amount = sum(row.get("amount_in", 0) or 0 for row in period_rows)
        out_qty = sum(row.get("qty_out", 0) or 0 for row in period_rows)
        out_amount = sum(row.get("amount_out", 0) or 0 for row in period_rows)
        running_balances = defaultdict(lambda: {"qty": 0.0, "amount": 0.0})
        for row in opening_rows:
            balance_key = row.get("warehouse_code") or ""
            running_balances[balance_key]["qty"] += (row.get("qty_in", 0) or 0) - (row.get("qty_out", 0) or 0)
            running_balances[balance_key]["amount"] += (row.get("amount_in", 0) or 0) - (row.get("amount_out", 0) or 0)
        rows.append([
            f"KHO: {warehouse.get('warehouse_code', '')} - {warehouse.get('warehouse_name', '')}",
            "", "", "", "", "", "", "", "", "", "", "", "", "", "",
        ])
        rows.append([
            f"Vật tư: {section['base_code']} - {section.get('product_name', '')}, Đvt: {section.get('unit_name') or '-'}, TK: {section.get('account') or warehouse.get('account') or ''}",
            "Ghi chú lấy mã VT chung không tách .001/.002",
            "", "", "", "", ", ".join(section["detail_codes"]), "", "", "", "", "", "", "", "",
        ])
        rows.extend([
            ["", "", "", "Tồn đầu kỳ", "", "", "", "", opening_qty, opening_amount, "", "", "", "", "", opening_qty, opening_amount, ""],
            ["", "", "", "Nhập trong kỳ", "", "", "", "", in_qty, in_amount, "", "", "", "", "", "", "", ""],
            ["", "", "", "Xuất trong kỳ", "", "", "", "", "", "", out_qty, out_amount, "", "", "", "", "", ""],
            ["", "", "", "Tồn cuối kỳ", "", "", "", "", opening_qty + in_qty - out_qty, opening_amount + in_amount - out_amount, "", "", "", "", "", opening_qty + in_qty - out_qty, opening_amount + in_amount - out_amount, ""],
        ])
        for row in period_rows:
            sale_amount = row.get("sale_amount", "")
            cost_amount = row.get("amount_out", 0) or 0
            margin_percent = ""
            if row.get("margin_percent") not in ("", None):
                margin_percent = row.get("margin_percent")
            elif row.get("type") == "sale" and sale_amount not in ("", None) and sale_amount:
                margin_percent = report_margin_percent(row.get("warehouse_code"), sale_amount, cost_amount) or ""
            balance_key = row.get("warehouse_code") or ""
            running_balances[balance_key]["qty"] += (row.get("qty_in", 0) or 0) - (row.get("qty_out", 0) or 0)
            running_balances[balance_key]["amount"] += (row.get("amount_in", 0) or 0) - (row.get("amount_out", 0) or 0)
            rows.append([
                row.get("date", "") or date_display(row.get("date_iso", "")),
                row.get("doc_no", ""),
                row.get("customer", ""),
                row.get("description", ""),
                row.get("account", ""),
                row.get("warehouse_code") or warehouse.get("warehouse_code", ""),
                row.get("variant_code", ""),
                row.get("unit_price", ""),
                row.get("qty_in", 0),
                row.get("amount_in", 0),
                row.get("qty_out", 0),
                row.get("amount_out", 0),
                row.get("sale_unit_price", ""),
                sale_amount,
                margin_percent,
                running_balances[balance_key]["qty"],
                running_balances[balance_key]["amount"],
                row.get("logic_note", ""),
            ])
    return rows


def write_combined_ledger_sheet(workbook, ledger, progress_callback=None, status_callback=None):
    if len((ledger or {}).get("warehouses", [])) <= 1:
        return
    sheet = replace_sheet(workbook, "SoChiTietHH_TP")
    rows = combined_ledger_export_rows(ledger)
    write_table(sheet, COMBINED_LEDGER_HEADERS, rows, progress_callback=progress_callback, status_callback=status_callback)
    sheet.auto_filter.ref = f"A1:R{max(1, sheet.max_row)}"
    sheet.freeze_panes = "A2"
    for row in sheet.iter_rows(min_row=2):
        first_value = text(row[0].value)
        if first_value.startswith("KHO:") or first_value.startswith("Vật tư:"):
            for cell in row:
                cell.font = Font(bold=True)
            row[0].alignment = Alignment(wrap_text=True)
            row[6].alignment = Alignment(wrap_text=True)
        if text(row[3].value) in {"Tồn đầu kỳ", "Nhập trong kỳ", "Xuất trong kỳ", "Tồn cuối kỳ"}:
            for cell in row:
                cell.font = Font(bold=True)
    for column in (8, 13):
        for row in sheet.iter_rows(min_row=2):
            row[column - 1].number_format = '#,##0.####'
    for row in sheet.iter_rows(min_row=2):
        row[14].number_format = '+0.####;-0.####;0'
    for column in (9, 11, 16):
        for row in sheet.iter_rows(min_row=2):
            row[column - 1].number_format = '#,##0.####'
    for column in (10, 12, 14, 17):
        for row in sheet.iter_rows(min_row=2):
            row[column - 1].number_format = '#,##0'
    fill_ledger_activity_rows(sheet, 9, 11, warehouse_column=6)
    fill_future_reorder_rows(sheet)
    widths = {
        "A": 14, "B": 14, "C": 32, "D": 42, "E": 11,
        "F": 10, "G": 20, "H": 13, "I": 14, "J": 15,
        "K": 14, "L": 15, "M": 13, "N": 14, "O": 12, "P": 14,
        "Q": 15, "R": 46,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    return len(rows)


def ledger_view_for_range(group, from_date="", to_date=""):
    rows = group.get("rows", [])
    opening_rows = [
        row for row in rows
        if row.get("type") == "opening" or (from_date and row.get("date_iso") and row.get("date_iso") < from_date)
    ]
    period_rows = [
        row for row in rows
        if row.get("type") != "opening"
        and not (from_date and row.get("date_iso") and row.get("date_iso") < from_date)
        and (not from_date or not row.get("date_iso") or row.get("date_iso") >= from_date)
        and (not to_date or not row.get("date_iso") or row.get("date_iso") <= to_date)
    ]
    opening_qty = sum(row.get("qty_in", 0) or 0 for row in opening_rows) - sum(row.get("qty_out", 0) or 0 for row in opening_rows)
    opening_amount = sum(row.get("amount_in", 0) or 0 for row in opening_rows) - sum(row.get("amount_out", 0) or 0 for row in opening_rows)
    in_qty = sum(row.get("qty_in", 0) or 0 for row in period_rows)
    in_amount = sum(row.get("amount_in", 0) or 0 for row in period_rows)
    out_qty = sum(row.get("qty_out", 0) or 0 for row in period_rows)
    out_amount = sum(row.get("amount_out", 0) or 0 for row in period_rows)
    return {
        "opening_qty": opening_qty,
        "opening_amount": opening_amount,
        "in_qty": in_qty,
        "in_amount": in_amount,
        "out_qty": out_qty,
        "out_amount": out_amount,
        "ending_qty": opening_qty + in_qty - out_qty,
        "ending_amount": opening_amount + in_amount - out_amount,
        "period_rows": period_rows,
    }


def has_ledger_activity(view):
    return (
        view["period_rows"]
        or abs(view["opening_qty"]) > QUANTITY_EPSILON
        or abs(view["opening_amount"]) > QUANTITY_EPSILON
        or abs(view["ending_qty"]) > QUANTITY_EPSILON
        or abs(view["ending_amount"]) > QUANTITY_EPSILON
    )


def inventory_summary_rows_for_warehouse(warehouse, from_date="", to_date=""):
    rows = []
    for group in warehouse.get("groups", []):
        view = ledger_view_for_range(group, from_date, to_date)
        if not has_ledger_activity(view):
            continue
        rows.append([
            group.get("variant_code", ""),
            group.get("product_name", ""),
            group.get("unit_name", ""),
            view["opening_qty"],
            view["opening_amount"],
            view["in_qty"],
            view["in_amount"],
            view["out_qty"],
            view["out_amount"],
            view["ending_qty"],
            view["ending_amount"],
        ])
    return rows


def inventory_summary_row_count(ledger):
    if not ledger:
        return 0
    date_range = ledger.get("date_range", {})
    from_date = date_range.get("from", "")
    to_date = date_range.get("to", "")
    return sum(
        len(inventory_summary_rows_for_warehouse(warehouse, from_date, to_date))
        for warehouse in ledger.get("warehouses", [])
    )


def write_inventory_summary_sheet(workbook, warehouse, from_date="", to_date="", progress_callback=None, status_callback=None):
    progress_callback = progress_callback or (lambda _done, _total: None)
    warehouse_code = warehouse.get("warehouse_code", "")
    sheet = replace_sheet(workbook, f"TongHopNXT_{warehouse_code}"[:31])
    rows = inventory_summary_rows_for_warehouse(warehouse, from_date, to_date)
    total_rows = max(1, len(rows))
    done = 0
    current_row = 1
    header_fill = PatternFill("solid", fgColor="DFF0DF")
    subheader_fill = PatternFill("solid", fgColor="E9F6E9")

    def style_header(row_number):
        for cell in sheet[row_number]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=12)
    sheet.cell(current_row, 1, "TỔNG HỢP NHẬP XUẤT TỒN")
    sheet.cell(current_row, 1).font = Font(bold=True, size=14)
    sheet.cell(current_row, 1).alignment = Alignment(horizontal="center")
    current_row += 1
    sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=12)
    sheet.cell(current_row, 1, f"KHO: {warehouse.get('warehouse_code', '')} - {warehouse.get('warehouse_name', '')}")
    sheet.cell(current_row, 1).alignment = Alignment(horizontal="center")
    current_row += 1
    sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=12)
    sheet.cell(current_row, 1, f"TỪ NGÀY: {date_display(from_date)} ĐẾN NGÀY: {date_display(to_date)}")
    sheet.cell(current_row, 1).alignment = Alignment(horizontal="center")
    current_row += 2

    header_row = current_row
    subheader_row = current_row + 1
    single_headers = [
        ("STT", 1), ("MÃ VẬT TƯ", 2), ("TÊN VẬT TƯ", 3), ("ĐVT", 4),
    ]
    for label, column in single_headers:
        sheet.merge_cells(start_row=header_row, start_column=column, end_row=subheader_row, end_column=column)
        sheet.cell(header_row, column, label)
    grouped_headers = [
        ("TỒN ĐẦU KỲ", 5, 6),
        ("NHẬP TRONG KỲ", 7, 8),
        ("XUẤT TRONG KỲ", 9, 10),
        ("TỒN CUỐI KỲ", 11, 12),
    ]
    for label, start_column, end_column in grouped_headers:
        sheet.merge_cells(start_row=header_row, start_column=start_column, end_row=header_row, end_column=end_column)
        sheet.cell(header_row, start_column, label)
    for column, label in enumerate(["SỐ LƯỢNG", "TIỀN", "SỐ LƯỢNG", "TIỀN", "SỐ LƯỢNG", "TIỀN", "SỐ LƯỢNG", "TIỀN"], start=5):
        sheet.cell(subheader_row, column, label)
    style_header(header_row)
    for cell in sheet[subheader_row]:
        cell.font = Font(bold=True)
        cell.fill = subheader_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    current_row += 2

    for index, row in enumerate(rows, start=1):
        values = [index, *row]
        for column, value in enumerate(values, start=1):
            sheet.cell(current_row, column, value)
        for column in (5, 7, 9, 11):
            sheet.cell(current_row, column).number_format = '#,##0.00'
        for column in (6, 8, 10, 12):
            sheet.cell(current_row, column).number_format = '#,##0'
        done += 1
        if done == 1 or done == total_rows or done % 250 == 0:
            progress_callback(done, total_rows)
        current_row += 1
    if rows:
        totals = [sum((row[column_index] or 0) for row in rows) for column_index in range(3, 11)]
        values = ["", "Tổng cộng:", "", "", *totals]
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(current_row, column, value)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="DDE8F6")
        for column in (5, 7, 9, 11):
            sheet.cell(current_row, column).number_format = '#,##0.00'
        for column in (6, 8, 10, 12):
            sheet.cell(current_row, column).number_format = '#,##0'
        current_row += 1
    if not rows:
        sheet.cell(current_row, 1, "Không có phát sinh trong kho này theo khoảng ngày.")
        progress_callback(1, total_rows)
    if status_callback:
        status_callback()

    widths = {
        "A": 7, "B": 18, "C": 34, "D": 9,
        "E": 13, "F": 15, "G": 13, "H": 15,
        "I": 13, "J": 15, "K": 13, "L": 15,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(
                horizontal="right" if cell.column >= 5 and cell.row > 4 else cell.alignment.horizontal,
                vertical="center",
                wrap_text=True,
            )
    return total_rows


def style_sales_report_header(sheet, row_number, max_column):
    for column in range(1, max_column + 1):
        cell = sheet.cell(row_number, column)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DFF0DF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def margin_percent_from_amounts(sale_amount, cost_amount):
    if not sale_amount:
        return 0
    return ((sale_amount or 0) - (cost_amount or 0)) / sale_amount * 100


def margin_percent_from_profit(profit_amount, base_amount):
    if not base_amount:
        return 0
    return (profit_amount or 0) / base_amount * 100


def report_margin_percent(warehouse_code, sale_amount, cost_amount):
    if abs(cost_amount or 0) <= QUANTITY_EPSILON:
        return 0
    return margin_percent_from_amounts(sale_amount, cost_amount)


def write_sales_summary_report_sheet(workbook, warehouse_code, rows, from_date="", to_date="", progress_callback=None, status_callback=None):
    progress_callback = progress_callback or (lambda _done, _total: None)
    sheet = replace_sheet(workbook, f"BaoCaoBH_{warehouse_code}"[:31])
    rows = [row for row in rows if row.get("warehouse_code") == warehouse_code]
    grouped = {}
    for row in rows:
        key = (row.get("variant_code", ""), row.get("product_name", ""), row.get("unit_name", ""))
        item = grouped.setdefault(key, {
            "variant_code": row.get("variant_code", ""),
            "product_name": row.get("product_name", ""),
            "unit_name": row.get("unit_name", ""),
            "quantity": 0.0,
            "cost_amount": 0.0,
            "sale_amount": 0.0,
            "profit_amount": 0.0,
            "tax_amount": 0.0,
            "total_amount": 0.0,
        })
        item["quantity"] += row.get("quantity", 0) or 0
        item["cost_amount"] += row.get("cost_amount", 0) or 0
        item["sale_amount"] += row.get("sale_amount", 0) or 0
        cost_missing = bool(row.get("cost_missing"))
        item.setdefault("margin_sale_amount", 0.0)
        item.setdefault("cost_missing_count", 0)
        item["profit_amount"] += 0 if cost_missing else (row.get("profit_amount", (row.get("sale_amount", 0) or 0) - (row.get("cost_amount", 0) or 0)) or 0)
        if cost_missing:
            item["cost_missing_count"] += 1
        else:
            item["margin_sale_amount"] += row.get("sale_amount", 0) or 0
        item["tax_amount"] += row.get("tax_amount", 0) or 0
        item["total_amount"] += row.get("total_amount", 0) or 0
    summary_rows = sorted(grouped.values(), key=lambda item: (item["variant_code"], item["product_name"]))

    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11)
    sheet.cell(1, 1, "BÁO CÁO TỔNG HỢP BÁN HÀNG")
    sheet.cell(1, 1).font = Font(bold=True, size=14)
    sheet.cell(1, 1).alignment = Alignment(horizontal="center")
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=11)
    sheet.cell(2, 1, f"KHO: {warehouse_code}")
    sheet.cell(2, 1).alignment = Alignment(horizontal="center")
    sheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=11)
    sheet.cell(3, 1, f"TỪ NGÀY: {date_display(from_date)} ĐẾN NGÀY: {date_display(to_date)}")
    sheet.cell(3, 1).alignment = Alignment(horizontal="center")
    headers = ["STT", "MÃ VẬT TƯ", "TÊN VẬT TƯ", "ĐVT", "SỐ LƯỢNG", "TIỀN VỐN", "TIỀN HÀNG", "TIỀN LÃI/LỖ", "% LÃI/LỖ", "TIỀN THUẾ", "TỔNG TIỀN TT"]
    for column, header in enumerate(headers, start=1):
        sheet.cell(5, column, header)
    style_sales_report_header(sheet, 5, len(headers))
    total_rows = max(1, len(summary_rows))
    for index, row in enumerate(summary_rows, start=1):
        excel_row = index + 5
        values = [
            index,
            row["variant_code"],
            row["product_name"],
            row["unit_name"],
            row["quantity"],
            row["cost_amount"],
            row["sale_amount"],
            row["profit_amount"],
            margin_percent_from_profit(row.get("profit_amount"), row.get("margin_sale_amount")),
            row["tax_amount"],
            row["total_amount"],
        ]
        for column, value in enumerate(values, start=1):
            sheet.cell(excel_row, column, value)
        sheet.cell(excel_row, 5).number_format = '#,##0.00'
        for column in (6, 7, 8, 10, 11):
            sheet.cell(excel_row, column).number_format = '#,##0'
        sheet.cell(excel_row, 9).number_format = '+0.####;-0.####;0'
        if index == 1 or index == total_rows or index % 250 == 0:
            progress_callback(index, total_rows)
    if not summary_rows:
        sheet.cell(6, 1, "Không có dữ liệu bán hàng trong kho này.")
        progress_callback(1, total_rows)
    if summary_rows:
        footer_row = len(summary_rows) + 6
        totals = {
            "quantity": sum(row["quantity"] for row in summary_rows),
            "cost_amount": sum(row["cost_amount"] for row in summary_rows),
            "sale_amount": sum(row["sale_amount"] for row in summary_rows),
            "profit_amount": sum(row["profit_amount"] for row in summary_rows),
            "tax_amount": sum(row["tax_amount"] for row in summary_rows),
            "total_amount": sum(row["total_amount"] for row in summary_rows),
        }
        footer_values = [
            "", "Tổng cộng:", "", "",
            totals["quantity"],
            totals["cost_amount"],
            totals["sale_amount"],
            totals["profit_amount"],
            margin_percent_from_profit(totals["profit_amount"], sum(row.get("margin_sale_amount", 0) for row in summary_rows)),
            totals["tax_amount"],
            totals["total_amount"],
        ]
        for column, value in enumerate(footer_values, start=1):
            cell = sheet.cell(footer_row, column, value)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="DDEBF7")
        sheet.cell(footer_row, 5).number_format = '#,##0.00'
        for column in (6, 7, 8, 10, 11):
            sheet.cell(footer_row, column).number_format = '#,##0'
        sheet.cell(footer_row, 9).number_format = '+0.####;-0.####;0'
    if status_callback:
        status_callback()
    for column, width in {"A": 7, "B": 18, "C": 38, "D": 9, "E": 13, "F": 15, "G": 15, "H": 16, "I": 12, "J": 15, "K": 16}.items():
        sheet.column_dimensions[column].width = width
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(
                horizontal="right" if cell.column >= 5 and cell.row >= 6 else cell.alignment.horizontal,
                vertical="center",
                wrap_text=True,
            )
    return total_rows


def sales_invoice_groups(rows, warehouse_code):
    groups = []
    by_key = {}
    for row in rows:
        if row.get("warehouse_code") != warehouse_code:
            continue
        key = (row.get("invoice_date_iso", ""), row.get("invoice_date", ""), row.get("invoice_no", ""), row.get("customer", ""), row.get("tax_rate", 0))
        if key not in by_key:
            group = {
                "invoice_date_iso": row.get("invoice_date_iso", ""),
                "invoice_date": row.get("invoice_date", ""),
                "invoice_no": row.get("invoice_no", ""),
                "customer": row.get("customer", ""),
                "tax_rate": row.get("tax_rate", 0),
                "rows": [],
            }
            by_key[key] = group
            groups.append(group)
        by_key[key]["rows"].append(row)
    groups.sort(key=lambda group: (group["invoice_date_iso"], text(group["invoice_no"])))
    return groups


def write_sales_invoice_report_sheet(workbook, warehouse_code, rows, from_date="", to_date="", progress_callback=None, status_callback=None):
    progress_callback = progress_callback or (lambda _done, _total: None)
    sheet = replace_sheet(workbook, f"BangKeHDBH_{warehouse_code}"[:31])
    groups = sales_invoice_groups(rows, warehouse_code)
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
    sheet.cell(1, 1, "BẢNG KÊ HOÁ ĐƠN BÁN HÀNG")
    sheet.cell(1, 1).font = Font(bold=True, size=14)
    sheet.cell(1, 1).alignment = Alignment(horizontal="center")
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=9)
    sheet.cell(2, 1, f"KHO: {warehouse_code}")
    sheet.cell(2, 1).alignment = Alignment(horizontal="center")
    sheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=9)
    sheet.cell(3, 1, f"TỪ NGÀY: {date_display(from_date)} ĐẾN NGÀY: {date_display(to_date)}")
    sheet.cell(3, 1).alignment = Alignment(horizontal="center")
    headers = ["NGÀY", "SỐ", "DIỄN GIẢI", "MÃ BP\nĐVT", "MÃ KHO\nMÃ NX", "SỐ LƯỢNG", "GIÁ VỐN", "TIỀN VỐN", "GIÁ BÁN", "TIỀN HÀNG"]
    for column, header in enumerate(headers, start=1):
        sheet.cell(5, column, header)
    style_sales_report_header(sheet, 5, len(headers))
    current_row = 6
    done = 0
    total_rows = max(1, sum(len(group["rows"]) for group in groups))
    for group in groups:
        group_rows = group["rows"]
        sale_total = sum(row.get("sale_amount", 0) or 0 for row in group_rows)
        tax_total = sum(row.get("tax_amount", 0) or 0 for row in group_rows)
        payment_total = sale_total + tax_total
        sheet.cell(current_row, 1, date_display(group.get("invoice_date_iso")) if group.get("invoice_date_iso") else group.get("invoice_date", ""))
        sheet.cell(current_row, 2, group.get("invoice_no", ""))
        sheet.cell(current_row, 3, group.get("customer", ""))
        sheet.cell(current_row + 1, 3, "Xuất bán cho khách")
        sheet.cell(current_row, 5, warehouse_code)
        sheet.cell(current_row + 1, 5, "131")
        current_row += 2
        for row in group_rows:
            sheet.cell(current_row, 3, f"{row.get('variant_code', '')} - {row.get('product_name', '')}".strip(" -"))
            sheet.cell(current_row, 4, row.get("unit_name", ""))
            sheet.cell(current_row, 6, row.get("quantity", 0)).number_format = '#,##0.00'
            sheet.cell(current_row, 7, row.get("cost_price", 0)).number_format = '#,##0.00'
            sheet.cell(current_row, 8, row.get("cost_amount", 0)).number_format = '#,##0'
            sheet.cell(current_row, 9, row.get("sale_price", 0)).number_format = '#,##0.00'
            sheet.cell(current_row, 10, row.get("sale_amount", 0)).number_format = '#,##0'
            done += 1
            if done == 1 or done == total_rows or done % 250 == 0:
                progress_callback(done, total_rows)
            current_row += 1
        for label, value in (("Tiền hàng:", sale_total), ("Tiền thuế:", tax_total), ("Tổng tiền thanh toán:", payment_total)):
            sheet.cell(current_row, 9, label)
            sheet.cell(current_row, 9).font = Font(bold=True)
            sheet.cell(current_row, 10, value)
            sheet.cell(current_row, 10).font = Font(bold=True)
            sheet.cell(current_row, 10).number_format = '#,##0'
            current_row += 1
        current_row += 1
    if not groups:
        sheet.cell(6, 1, "Không có dữ liệu bán hàng trong kho này.")
        progress_callback(1, total_rows)
    if status_callback:
        status_callback()
    for column, width in {"A": 10, "B": 10, "C": 42, "D": 12, "E": 12, "F": 13, "G": 14, "H": 15, "I": 14, "J": 15}.items():
        sheet.column_dimensions[column].width = width
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(
                horizontal="right" if cell.column >= 6 and cell.row >= 6 else cell.alignment.horizontal,
                vertical="center",
                wrap_text=True,
            )
    return total_rows


def line_tax_amount(line):
    if line.get("tax_amount") is not None:
        return line.get("tax_amount") or 0
    amount = line_total_amount(line)
    rate = sale_tax_rate(line)
    return (amount or 0) * rate / 100


def sum_values(rows, field):
    return sum(row.get(field, 0) or 0 for row in rows or [])


def ledger_totals_for_type(ledger, warehouse_code, row_type):
    totals = {"qty_in": 0.0, "amount_in": 0.0, "qty_out": 0.0, "amount_out": 0.0}
    for warehouse in (ledger or {}).get("warehouses", []):
        if warehouse.get("warehouse_code") != warehouse_code:
            continue
        for group in warehouse.get("groups", []):
            for row in group.get("rows", []):
                if row.get("type") != row_type:
                    continue
                for field in totals:
                    totals[field] += row.get(field, 0) or 0
    return totals


def verification_status(original_value, processed_value, tolerance):
    diff = (processed_value or 0) - (original_value or 0)
    return diff, "OK" if abs(diff) <= tolerance else "Lệch"


def build_verification_rows(purchase_lines, sales_lines, allocations, ledger, sales_report_rows, policy=None):
    original_sales_qty = sum(line.get("quantity", 0) or 0 for line in sales_lines or [])
    original_sales_amount = sum(line_total_amount(line) or 0 for line in sales_lines or [])
    original_sales_tax = sum(line_tax_amount(line) for line in sales_lines or [])
    original_purchase_qty = sum(line.get("quantity", 0) or 0 for line in purchase_lines or [])
    original_purchase_amount = sum(line_total_amount(line) or 0 for line in purchase_lines or [])

    processed_sales_qty = sum_values(sales_report_rows, "quantity")
    processed_sales_amount = sum_values(sales_report_rows, "sale_amount")
    processed_sales_tax = sum_values(sales_report_rows, "tax_amount")
    allocated_qty = sum((row.get("material_quantity", 0) or 0) + (row.get("finished_quantity", 0) or 0) for row in allocations or [])

    warehouse_codes = sorted({
        row.get("warehouse_code", "")
        for row in sales_report_rows or []
        if row.get("warehouse_code")
    } | {
        warehouse.get("warehouse_code", "")
        for warehouse in (ledger or {}).get("warehouses", [])
        if warehouse.get("warehouse_code")
    })
    purchase_ledger_totals = {"qty_in": 0.0, "amount_in": 0.0}
    for warehouse_code in warehouse_codes:
        for ledger_type in ("purchase", "purchase_future_reorder"):
            totals = ledger_totals_for_type(ledger, warehouse_code, ledger_type)
            purchase_ledger_totals["qty_in"] += totals["qty_in"]
            purchase_ledger_totals["amount_in"] += totals["amount_in"]
    checks = [
        ("Bán ra", "SL gốc = SL đã phân bổ theo kho", original_sales_qty, processed_sales_qty, 0.0001, "Đảm bảo không mất hoặc nhân đôi số lượng bán ra khi tách kho."),
        ("Bán ra", "Tiền hàng gốc = tiền hàng đã tách", original_sales_amount, processed_sales_amount, 1, "Đảm bảo doanh thu chưa thuế sau xử lý khớp file bán ra gốc."),
        ("Bán ra", "Tiền thuế gốc = tiền thuế đã tách", original_sales_tax, processed_sales_tax, 1, "Đảm bảo thuế được phân bổ theo tiền hàng sau tách kho."),
        ("Phân bổ", "SL gốc = SL phân bổ theo các kho bán ra", original_sales_qty, allocated_qty, 0.0001, "Đảm bảo mỗi dòng bán ra đều được phân bổ đủ sang kho ghi trên file bán ra."),
        ("Mua vào", "SL mua gốc = SL nhập các kho", original_purchase_qty, purchase_ledger_totals["qty_in"], 0.0001, "Đảm bảo hóa đơn mua vào được đưa vào sổ chi tiết kho."),
        ("Mua vào", "Tiền mua gốc = tiền nhập các kho", original_purchase_amount, purchase_ledger_totals["amount_in"], 1, "Đảm bảo giá trị mua vào khớp nhập kho."),
    ]
    for warehouse_code in warehouse_codes:
        report_rows = [row for row in sales_report_rows if row.get("warehouse_code") == warehouse_code]
        sale_ledger = ledger_totals_for_type(ledger, warehouse_code, "sale")
        checks.extend([
            (warehouse_code, f"SL bán hàng {warehouse_code} = SL xuất sổ chi tiết {warehouse_code}", sum_values(report_rows, "quantity"), sale_ledger["qty_out"], 0.0001, f"Đối chiếu báo cáo bán hàng {warehouse_code} với sổ chi tiết {warehouse_code}."),
            (warehouse_code, f"Tiền vốn {warehouse_code} = xuất thành tiền sổ chi tiết {warehouse_code}", sum_values(report_rows, "cost_amount"), sale_ledger["amount_out"], 1, f"Đối chiếu giá vốn bán hàng {warehouse_code} với xuất kho {warehouse_code}."),
        ])
    policy = clean_policy(policy or DEFAULT_POLICY)
    if policy_active(policy):
        outside_range = 0
        for row in sales_report_rows or []:
            sale_amount = row.get("sale_amount", 0) or 0
            if sale_amount <= QUANTITY_EPSILON:
                continue
            margin_percent = report_margin_percent(row.get("warehouse_code"), sale_amount, row.get("cost_amount", 0) or 0)
            if margin_percent is None:
                continue
            if policy["max_loss_percent"] is not None and margin_percent < -policy["max_loss_percent"]:
                outside_range += 1
            if policy["max_profit_percent"] is not None and margin_percent > policy["max_profit_percent"]:
                outside_range += 1
        checks.append((
            "Báo cáo bán hàng",
            "Dòng ngoài khoảng lãi/lỗ",
            0,
            outside_range,
            0,
            "Kiểm tra các dòng báo cáo bán hàng có tiền hàng và giá vốn.",
        ))
    rows = []
    for group, check_name, original_value, processed_value, tolerance, explanation in checks:
        diff, status = verification_status(original_value, processed_value, tolerance)
        rows.append([
            group,
            check_name,
            original_value,
            processed_value,
            diff,
            tolerance,
            status,
            explanation,
        ])
    return rows


def write_verification_sheet(workbook, verification_rows):
    sheet = replace_sheet(workbook, "KiemTraDoiChieu")
    headers = ["Nhóm", "Kiểm tra", "Giá trị file gốc", "Giá trị sau xử lý", "Chênh lệch", "Ngưỡng", "Kết quả", "Ý nghĩa"]
    write_table(sheet, headers, verification_rows)
    for row in sheet.iter_rows(min_row=2):
        status = text(row[6].value)
        row[6].font = Font(bold=True, color="0B774B" if status == "OK" else "AD2C2C")
        for column in (3, 4, 5, 6):
            row[column - 1].number_format = '#,##0.####'
    sheet.column_dimensions["A"].width = 14
    sheet.column_dimensions["B"].width = 34
    sheet.column_dimensions["C"].width = 18
    sheet.column_dimensions["D"].width = 18
    sheet.column_dimensions["E"].width = 16
    sheet.column_dimensions["F"].width = 12
    sheet.column_dimensions["G"].width = 10
    sheet.column_dimensions["H"].width = 58
    return len(verification_rows)


RESULT_COLUMN_ALIASES = {
    "SL khớp từ mua vào": ["SL lấy từ kho hàng hóa"],
    "SL xuất theo kho bán ra": ["SL lấy từ kho thành phẩm"],
    "Tồn mua vào trước khi bán": ["Tồn kho trước khi bán"],
    "Chi tiết khớp mua vào": ["Chi tiết lấy từ kho hàng hóa"],
    "Tồn mua vào sau khi bán": ["Tồn kho sau khi bán"],
    "Lô mua vào không đạt khoảng lãi/lỗ": ["Lô không đạt khoảng lãi/lỗ"],
}


def result_column(sheet, header_row, header, fallback):
    accepted_headers = {header, *RESULT_COLUMN_ALIASES.get(header, [])}
    for column in range(1, sheet.max_column + 1):
        if text(sheet.cell(header_row, column).value) in accepted_headers:
            return column
    return fallback


def create_output_workbook(
    sales_content,
    sales_mapping,
    allocations,
    stock_rows,
    summary,
    policy=None,
    sale_only_codes=None,
    purchase_lines=None,
    sales_lines=None,
    progress_callback=None,
    ledger=None,
    missing_barem_report=None,
    ambiguous_steel_rows=None,
):
    policy = clean_policy(policy or DEFAULT_POLICY)
    sale_only_codes = sale_only_codes or []
    missing_barem_report = missing_barem_report or []
    ambiguous_steel_rows = ambiguous_steel_rows or []
    progress_callback = progress_callback or (lambda _done, _total, _label: None)
    purchase_export_rows = build_purchase_export_rows(purchase_lines)
    sales_export_rows = build_sales_export_rows(allocations, purchase_lines, sales_lines)
    sales_report_rows = build_sales_report_rows_from_ledger(ledger) if ledger else build_sales_report_rows(allocations, purchase_lines, sales_lines)
    future_reorder_report = (summary or {}).get("future_purchase_reorder_report", [])
    ledger_warehouses = (ledger or {}).get("warehouses", [])
    verification_rows = build_verification_rows(purchase_lines, sales_lines, allocations, ledger, sales_report_rows, policy)
    ledger_row_count = sum(len(ledger_export_rows(warehouse)) for warehouse in ledger_warehouses)
    should_write_combined_ledger = bool(ledger and len(ledger_warehouses) > 1)
    combined_ledger_row_count = len(combined_ledger_export_rows(ledger)) if should_write_combined_ledger else 0
    nxt_row_count = inventory_summary_row_count(ledger)
    output_total = max(
        1,
        len(allocations)
        + len(sale_only_codes)
        + len(missing_barem_report)
        + len(ambiguous_steel_rows)
        + len(future_reorder_report)
        + len(purchase_export_rows)
        + len(sales_export_rows)
        + len(sales_report_rows) * 2
        + len(verification_rows)
        + ledger_row_count
        + combined_ledger_row_count
        + nxt_row_count
        + 5,
    )
    output_done = 0

    def report(delta, label):
        nonlocal output_done
        output_done = min(output_total, output_done + max(0, delta))
        progress_callback(output_done, output_total, label)

    def row_reporter(label):
        last = 0

        def callback(done, total):
            nonlocal last
            report(done - last, f"{label}: {done}/{total} dòng...")
            last = done

        return callback

    def status_reporter(label):
        return lambda: report(0, label)

    workbook = load_workbook(BytesIO(sales_content))
    generated_sheet_names = {
        "PhanBoKho", "TonKhoHangHoa", "TongHopNhapXuatTon", "TongHopKho", "KiemTraDoiChieu",
        "MaChiBanRaKhongTon", "MaThieuBarem", "MaThepKhongRo", "MauMuaVao", "MauBanRa",
        "HDMuaVaoDuaLenTruoc", "ChiTietPhanKho", "PhuongAnOngHop", "XuatAmKHHVT",
    }
    generated_sheet_prefixes = ("SoChiTiet", "TongHopNXT_", "BaoCaoBH_", "BangKeHDBH_")
    for removed_sheet in list(workbook.sheetnames):
        if removed_sheet in generated_sheet_names or removed_sheet.startswith(generated_sheet_prefixes):
            del workbook[removed_sheet]
    sales_sheet = sheet_for_mapping(workbook, sales_mapping)
    header_row = sales_mapping["header_row"]
    column_headers = [
        "SL khớp từ mua vào",
        "SL xuất theo kho bán ra",
        "Tồn mua vào trước khi bán",
        "Chi tiết khớp mua vào",
        "Tồn mua vào sau khi bán",
        "Lô mua vào không đạt khoảng lãi/lỗ",
    ]
    output_columns = {}
    for header in column_headers:
        output_columns[header] = result_column(sales_sheet, header_row, header, sales_sheet.max_column + 1)
        sales_sheet.cell(header_row, output_columns[header], header)
    for cell in (sales_sheet.cell(header_row, column) for column in output_columns.values()):
        cell.font = Font(bold=True, color="12304B")
        cell.fill = PatternFill("solid", fgColor="E7F2FF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sales_sheet.column_dimensions[get_column_letter(output_columns["SL khớp từ mua vào"])].width = 22
    sales_sheet.column_dimensions[get_column_letter(output_columns["SL xuất theo kho bán ra"])].width = 23
    for header in column_headers[2:]:
        sales_sheet.column_dimensions[get_column_letter(output_columns[header])].width = 44
    total_allocations = len(allocations)
    last_reported_allocation = 0
    for index, line in enumerate(allocations, start=1):
        sales_sheet.cell(line["row_number"], output_columns["SL khớp từ mua vào"], line["material_quantity"]).number_format = '#,##0.####'
        sales_sheet.cell(line["row_number"], output_columns["SL xuất theo kho bán ra"], line["finished_quantity"]).number_format = '#,##0.####'
        sales_sheet.cell(line["row_number"], output_columns["Tồn mua vào trước khi bán"], inventory_text(line.get("inventory_before", [])))
        sales_sheet.cell(line["row_number"], output_columns["Chi tiết khớp mua vào"], detail_text(line["used"]))
        sales_sheet.cell(line["row_number"], output_columns["Tồn mua vào sau khi bán"], inventory_text(line.get("inventory_after", [])))
        sales_sheet.cell(line["row_number"], output_columns["Lô mua vào không đạt khoảng lãi/lỗ"], rejected_text(line.get("rejected", [])))
        if index == 1 or index == total_allocations or index % 250 == 0:
            report(index - last_reported_allocation, f"Đang ghi kết quả vào hóa đơn bán ra: {index}/{total_allocations} dòng...")
            last_reported_allocation = index
    if not allocations:
        report(1, "Đang ghi kết quả vào hóa đơn bán ra...")

    summary_sheet = replace_sheet(workbook, "TongHopKho")
    write_table(summary_sheet, ["Chỉ tiêu", "Giá trị"], [
        ["Tồn đầu kỳ", summary["opening_quantity"]],
        ["Nhập mua vào", summary["purchase_quantity"]],
        ["Số lượng bán ra", summary["sales_quantity"]],
        ["SL khớp từ mua vào", summary["material_quantity"]],
        ["SL xuất theo kho bán ra", summary["finished_quantity"]],
        ["Tỷ lệ khớp từ mua vào (%)", summary["material_percent"]],
        ["Dòng có lô bị loại theo khoảng lãi/lỗ", summary.get("range_rejected_lines", 0)],
        ["HD mua vào đưa lên trước", summary.get("future_purchase_reorder_count", 0)],
        ["SL từ HD mua vào đưa lên trước", summary.get("future_purchase_reorder_quantity", 0)],
        ["Mã VT chỉ bán ra mà không có tồn mua vào", summary.get("sale_only_code_count", len(sale_only_codes))],
        ["Lỗ tối đa chấp nhận (%)", policy["max_loss_percent"] if policy["max_loss_percent"] is not None else "Không giới hạn"],
        ["Lãi tối đa chấp nhận (%)", policy["max_profit_percent"] if policy["max_profit_percent"] is not None else "Không giới hạn"],
    ], status_callback=status_reporter("Đang định dạng tổng hợp kho..."))
    report(1, "Đã ghi tổng hợp kho.")

    if policy.get("company_profile") == "son_phuong":
        allocation_detail_rows = []
        negative_rows = []
        scenario_rows = []
        for allocation in allocations:
            common = [
                allocation.get("invoice_date", ""),
                allocation.get("invoice_no", ""),
                allocation.get("row_number", ""),
                allocation.get("product_name", ""),
                allocation.get("variant_code", ""),
                allocation.get("quantity", 0),
                allocation.get("allocation_role", ""),
            ]
            for used in allocation.get("used", []):
                warehouse_code = allocation.get("warehouse_code", "KHHVT")
                warehouse_account = allocation.get("warehouse_account", "156")
                detail_row = common + [
                    used.get("ledger_variant_code") or used.get("variant_code", ""),
                    used.get("quantity", 0),
                    warehouse_code,
                    warehouse_account,
                    used.get("invoice_date", ""),
                    used.get("invoice_no", ""),
                    used.get("row_number", ""),
                    "X" if used.get("negative_export") else "",
                    used.get("logic_note", ""),
                ]
                allocation_detail_rows.append(detail_row)
                if used.get("negative_export"):
                    negative_rows.append(detail_row)
            if allocation.get("finished_quantity", 0) > QUANTITY_EPSILON:
                allocation_detail_rows.append(common + [
                    allocation.get("finished_variant_code") or allocation.get("variant_code", ""),
                    allocation.get("finished_quantity", 0),
                    allocation.get("remainder_warehouse_code", ""),
                    allocation.get("remainder_warehouse_account", ""),
                    "", "", "", "",
                    allocation.get("generic_plan_note", ""),
                ])
            if allocation.get("generic_plan_note"):
                before_by_code = {}
                after_by_code = {}
                for snapshot_item in allocation.get("inventory_before", []):
                    snapshot_code = normalize_code(snapshot_item.get("variant_code", ""))
                    before_by_code[snapshot_code] = clean_quantity(
                        before_by_code.get(snapshot_code, 0) + (snapshot_item.get("quantity", 0) or 0)
                    )
                for snapshot_item in allocation.get("inventory_after", []):
                    snapshot_code = normalize_code(snapshot_item.get("variant_code", ""))
                    after_by_code[snapshot_code] = clean_quantity(
                        after_by_code.get(snapshot_code, 0) + (snapshot_item.get("quantity", 0) or 0)
                    )
                for used in allocation.get("used", []):
                    profile_key = used.get("steel_profile_key", "")
                    profile_parts = profile_key.split("|", 2) if profile_key else []
                    split_code = used.get("ledger_variant_code") or used.get("variant_code", "")
                    split_quantity = used.get("quantity", 0) or 0
                    sale_amount = amount_for_quantity(allocation, split_quantity)
                    unit_cost = used.get("unit_cost")
                    cost_amount = unit_cost * split_quantity if unit_cost is not None else None
                    normalized_split_code = normalize_code(split_code)
                    before_qty = before_by_code.get(normalized_split_code, 0)
                    after_qty = after_by_code.get(normalized_split_code, 0)
                    scenario_rows.append([
                        allocation.get("invoice_date", ""), allocation.get("invoice_no", ""), allocation.get("row_number", ""),
                        allocation.get("party_tax_code", ""), allocation.get("party_name", ""),
                        allocation.get("product_name", ""), allocation.get("variant_code", ""), allocation.get("quantity", 0),
                        allocation.get("unit_name", ""), allocation.get("unit_price", ""), allocation.get("line_amount", ""),
                        profile_parts[0] if len(profile_parts) == 3 else "", profile_parts[1] if len(profile_parts) == 3 else "", profile_parts[2] if len(profile_parts) == 3 else "",
                        split_code, used.get("purchase_product_name", ""), split_quantity,
                        used.get("barem_weight", ""), used.get("barem_multiple", ""), used.get("barem_remainder_kg", 0), used.get("barem_source", ""), used.get("barem_method", ""), ", ".join(used.get("barem_references") or []),
                        allocation.get("warehouse_code", "KHHVT"), allocation.get("warehouse_account", "156"),
                        used.get("invoice_date", ""), used.get("invoice_no", ""), used.get("row_number", ""), used.get("purchase_party_tax_code", ""), used.get("purchase_party_name", ""),
                        unit_cost, cost_amount, sale_amount, (sale_amount - cost_amount) if sale_amount is not None and cost_amount is not None else None,
                        before_qty, after_qty, "X" if used.get("negative_export") else "", "X" if used.get("barem_unallocated") else "",
                        used.get("logic_note", "") or allocation.get("generic_plan_note", ""),
                    ])
            if False and allocation.get("generic_plan_note"):
                scenario_rows.append([
                    allocation.get("invoice_date", ""),
                    allocation.get("invoice_no", ""),
                    allocation.get("row_number", ""),
                    allocation.get("product_name", ""),
                    allocation.get("quantity", 0),
                    allocation.get("sale_split_codes", ""),
                    allocation.get("generic_plan_note", ""),
                ])

        scenario_headers = [
            "Ng\u00e0y H\u0110 b\u00e1n", "S\u1ed1 H\u0110 b\u00e1n", "D\u00f2ng b\u00e1n", "MST kh\u00e1ch", "Kh\u00e1ch h\u00e0ng",
            "T\u00ean h\u00e0ng b\u00e1n g\u1ed1c", "M\u00e3 VT b\u00e1n g\u1ed1c", "SL b\u00e1n g\u1ed1c", "\u0110VT b\u00e1n", "\u0110\u01a1n gi\u00e1 b\u00e1n", "Ti\u1ec1n b\u00e1n g\u1ed1c",
            "Lo\u1ea1i", "B\u1ec1 m\u1eb7t", "K\u00edch th\u01b0\u1edbc / kh\u00f3a k\u1ef9 thu\u1eadt", "M\u00e3 VT \u0111\u00e3 t\u00e1ch", "T\u00ean v\u1eadt t\u01b0 mua",
            "Kh\u1ed1i l\u01b0\u1ee3ng t\u00e1ch", "Kg / barem", "S\u1ed1 thanh nguy\u00ean", "Kg l\u1ebb cu\u1ed1i", "Ngu\u1ed3n barem", "Ph\u01b0\u01a1ng ph\u00e1p barem", "Barem tham chi\u1ebfu",
            "M\u00e3 kho", "TK v\u1eadt t\u01b0", "Ng\u00e0y H\u0110 mua", "S\u1ed1 H\u0110 mua", "D\u00f2ng mua", "MST nh\u00e0 cung c\u1ea5p", "Nh\u00e0 cung c\u1ea5p",
            "\u0110\u01a1n gi\u00e1 v\u1ed1n", "Ti\u1ec1n v\u1ed1n", "Ti\u1ec1n b\u00e1n ph\u00e2n b\u1ed5", "L\u00e3i/l\u1ed7", "T\u1ed3n tr\u01b0\u1edbc", "T\u1ed3n sau",
            "Xu\u1ea5t \u00e2m KHHVT", "Ch\u01b0a ph\u00e2n b\u1ed5 theo barem", "L\u00fd do / ghi ch\u00fa",
        ]

        allocation_headers = [
            "Ng\u00e0y H\u0110 b\u00e1n", "S\u1ed1 H\u0110 b\u00e1n", "D\u00f2ng b\u00e1n", "T\u00ean h\u00e0ng b\u00e1n",
            "M\u00e3 VT b\u00e1n", "SL b\u00e1n", "Vai tr\u00f2", "M\u00e3 VT ph\u00e2n kho", "SL ph\u00e2n kho",
            "M\u00e3 kho", "TK v\u1eadt t\u01b0", "Ng\u00e0y H\u0110 mua", "S\u1ed1 H\u0110 mua", "D\u00f2ng mua",
            "Xu\u1ea5t \u00e2m KHHVT", "Ghi ch\u00fa",
        ]
        write_table(
            replace_sheet(workbook, "ChiTietPhanKho"),
            allocation_headers,
            allocation_detail_rows,
            progress_callback=row_reporter("\u0110ang ghi chi ti\u1ebft Ph\u00e2n kho"),
            status_callback=status_reporter("\u0110ang \u0111\u1ecbnh d\u1ea1ng chi ti\u1ebft Ph\u00e2n kho..."),
        )
        write_table(
            replace_sheet(workbook, "PhuongAnOngHop"),
            scenario_headers,
            scenario_rows,
            progress_callback=row_reporter("\u0110ang ghi ph\u01b0\u01a1ng \u00e1n \u1ed1ng/h\u1ed9p"),
            status_callback=status_reporter("\u0110ang \u0111\u1ecbnh d\u1ea1ng ph\u01b0\u01a1ng \u00e1n \u1ed1ng/h\u1ed9p..."),
        )
        write_table(
            replace_sheet(workbook, "XuatAmKHHVT"),
            allocation_headers,
            negative_rows,
            progress_callback=row_reporter("\u0110ang ghi xu\u1ea5t \u00e2m KHHVT"),
            status_callback=status_reporter("\u0110ang \u0111\u1ecbnh d\u1ea1ng xu\u1ea5t \u00e2m KHHVT..."),
        )

    write_verification_sheet(workbook, verification_rows)
    report(len(verification_rows), "Đã ghi kiểm tra đối chiếu dữ liệu.")

    sale_only_sheet = replace_sheet(workbook, "MaChiBanRaKhongTon")
    write_table(sale_only_sheet, [
        "Mã VT bán ra", "Mã gốc", "Tên hàng", "SL tồn đầu kỳ", "SL mua vào",
        "Số dòng bán ra", "Tổng SL bán", "Các dòng trong hóa đơn bán ra",
    ], [[
        row["variant_code"], row["base_code"], row["product_name"], row["opening_quantity"],
        row["purchase_quantity"], row["row_count"], row["quantity"],
        ", ".join(str(row_number) for row_number in row["rows"]),
    ] for row in sale_only_codes], progress_callback=row_reporter("Đang ghi mã chỉ bán ra"), status_callback=status_reporter("Đang định dạng mã chỉ bán ra..."))

    missing_barem_sheet = replace_sheet(workbook, "MaThieuBarem")
    write_table(missing_barem_sheet, [
        "Mã VT", "Tên hàng", "SL tồn còn lại", "Số HĐ", "Ngày HĐ", "Dòng Excel", "Lý do",
    ], [[
        row.get("variant_code", ""),
        row.get("product_name", ""),
        row.get("quantity", 0),
        row.get("invoice_no", ""),
        row.get("invoice_date", ""),
        row.get("row_number", ""),
        row.get("reason", ""),
    ] for row in missing_barem_report], progress_callback=row_reporter("Đang ghi mã thiếu barem"), status_callback=status_reporter("Đang định dạng mã thiếu barem..."))

    ambiguous_sheet = replace_sheet(workbook, "MaThepKhongRo")
    write_table(ambiguous_sheet, [
        "Công ty", "MST", "Mã VT", "Tên hàng", "SL", "Đơn giá", "Số HĐ", "Ngày HĐ",
        "Dòng Excel", "Loại nhận diện", "Dimension đọc được", "Lý do", "Ghi chú nhận diện",
    ], [[
        row.get("company", ""),
        row.get("tax_code", ""),
        row.get("variant_code", ""),
        row.get("product_name", ""),
        row.get("quantity", 0),
        row.get("unit_price", ""),
        row.get("invoice_no", ""),
        row.get("invoice_date", ""),
        row.get("row_number", ""),
        row.get("detected_kind", ""),
        row.get("dimension", ""),
        row.get("reason", ""),
        row.get("kind_reason", ""),
    ] for row in ambiguous_steel_rows], progress_callback=row_reporter("Đang ghi mã thép không rõ"), status_callback=status_reporter("Đang định dạng mã thép không rõ..."))

    future_reorder_sheet = replace_sheet(workbook, "HDMuaVaoDuaLenTruoc")
    write_table(future_reorder_sheet, [
        "Ma VT mua vao", "Ma VT ban ra", "Ten hang", "So HD mua vao",
        "Ngay mua goc", "Ngay hieu luc", "So HD ban ra", "Ngay ban ra",
        "Dong ban ra", "Dong mua vao", "SL lay", "Don gia von", "So ngay keo len", "Ghi chu",
    ], [[
        row.get("purchase_variant_code", ""),
        row.get("sale_variant_code", ""),
        row.get("product_name", ""),
        row.get("purchase_invoice_no", ""),
        row.get("purchase_original_date", ""),
        row.get("effective_date", ""),
        row.get("sale_invoice_no", ""),
        row.get("sale_date", ""),
        row.get("sale_row_number", ""),
        row.get("purchase_row_number", ""),
        row.get("quantity", 0),
        row.get("unit_cost", ""),
        row.get("future_reorder_days", ""),
        row.get("logic_note", ""),
    ] for row in future_reorder_report], progress_callback=row_reporter("Đang ghi HD mua vào đưa lên trước"), status_callback=status_reporter("Đang định dạng HD mua vào đưa lên trước..."))

    purchase_export_sheet = replace_sheet(workbook, "MauMuaVao")
    write_table(purchase_export_sheet, PURCHASE_EXPORT_HEADERS, purchase_export_rows, progress_callback=row_reporter("Đang ghi mẫu mua vào"), status_callback=status_reporter("Đang định dạng mẫu mua vào..."))
    format_export_sheet(purchase_export_sheet, quantity_columns=(7,), money_columns=(8, 9))
    report(1, "Đã định dạng mẫu mua vào.")

    sales_export_sheet = replace_sheet(workbook, "MauBanRa")
    write_table(sales_export_sheet, SALES_EXPORT_HEADERS, sales_export_rows, progress_callback=row_reporter("Đang ghi mẫu bán ra"), status_callback=status_reporter("Đang định dạng mẫu bán ra..."))
    format_export_sheet(
        sales_export_sheet,
        quantity_columns=(5,),
        money_columns=(6, 7, 8, 9, 11),
        percent_columns=(10,),
    )
    report(1, "Đã định dạng mẫu bán ra.")

    for warehouse in ledger_warehouses:
        write_ledger_export_sheet(
            workbook,
            warehouse,
            progress_callback=row_reporter(f"Đang ghi sổ chi tiết {warehouse.get('warehouse_code', '')}"),
            status_callback=status_reporter(f"Đang định dạng sổ chi tiết {warehouse.get('warehouse_code', '')}..."),
        )
    if ledger and ledger_warehouses:
        date_range = ledger.get("date_range", {})
        from_date = date_range.get("from", "")
        to_date = date_range.get("to", "")
        write_combined_ledger_sheet(
            workbook,
            ledger,
            progress_callback=row_reporter("Đang ghi sổ chi tiết tổng hợp các kho"),
            status_callback=status_reporter("Đang định dạng sổ chi tiết tổng hợp các kho..."),
        )
        for warehouse in ledger_warehouses:
            write_inventory_summary_sheet(
                workbook,
                warehouse,
                from_date,
                to_date,
                progress_callback=row_reporter(f"Đang ghi tổng hợp nhập xuất tồn {warehouse.get('warehouse_code', '')}"),
                status_callback=status_reporter(f"Đang định dạng tổng hợp nhập xuất tồn {warehouse.get('warehouse_code', '')}..."),
            )
            warehouse_code = warehouse.get("warehouse_code", "")
            write_sales_summary_report_sheet(
                workbook,
                warehouse_code,
                sales_report_rows,
                from_date,
                to_date,
                progress_callback=row_reporter(f"Đang ghi báo cáo tổng hợp bán hàng {warehouse_code}"),
                status_callback=status_reporter(f"Đang định dạng báo cáo tổng hợp bán hàng {warehouse_code}..."),
            )
            write_sales_invoice_report_sheet(
                workbook,
                warehouse_code,
                sales_report_rows,
                from_date,
                to_date,
                progress_callback=row_reporter(f"Đang ghi bảng kê hóa đơn bán hàng {warehouse_code}"),
                status_callback=status_reporter(f"Đang định dạng bảng kê hóa đơn bán hàng {warehouse_code}..."),
            )

    stream = BytesIO()
    workbook.save(stream)
    report(1, "Đã lưu workbook vào bộ nhớ.")
    stream.seek(0)
    return stream


def output_filename(original_name):
    stem = Path(original_name or "hoa_don_ban_ra").stem
    stem = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", stem).strip(" .") or "hoa_don_ban_ra"
    return f"{stem}_phan_bo_kho.xlsx"


def save_output(stream, original_name):
    job_id = uuid.uuid4().hex
    filename = output_filename(original_name)
    path = OUTPUT_DIR / f"{job_id}_{filename}"
    path.write_bytes(stream.getvalue())
    return job_id, filename, path


def preview_workbook(content, header_row=1):
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    header_row = max(1, min(int(header_row or 1), sheet.max_row))
    displayed_columns = min(sheet.max_column, 18)
    rows = []
    for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 8), values_only=True):
        rows.append([text(value) for value in row[:displayed_columns]])
    columns = [{
        "letter": get_column_letter(column),
        "header": text(sheet.cell(header_row, column).value),
    } for column in range(1, sheet.max_column + 1)]
    result = {
        "sheets": workbook.sheetnames,
        "active_sheet": sheet.title,
        "max_column": get_column_letter(sheet.max_column),
        "header_row": header_row,
        "preview_letters": [get_column_letter(column) for column in range(1, displayed_columns + 1)],
        "columns": columns,
        "rows": rows,
    }
    workbook.close()
    return result


@app.post("/api/preview")
def preview():
    uploaded = request.files.get("file")
    if not uploaded:
        return jsonify({"error": "Chưa chọn file Excel."}), 400
    try:
        header_row = request.form.get("header_row") or 1
        return jsonify(preview_workbook(uploaded.read(), header_row))
    except Exception as exc:
        return jsonify({"error": str(exc)}), 400


@app.post("/api/purchase-classification-preview")
def purchase_classification_preview():
    uploaded = request.files.get("file")
    if not uploaded:
        return jsonify({"error": "Chưa chọn file mua vào."}), 400
    try:
        raw_mapping = json.loads(request.form.get("mapping") or "{}")
        policy = clean_policy(json.loads(request.form.get("policy") or "{}"))
        mapping = clean_mapping(raw_mapping, "purchase")
        _, lines = read_lines(
            uploaded.read(),
            mapping,
            "purchase",
            company_profile=policy.get("company_profile", "yen_thanh"),
        )
        counts = {"pipe": 0, "box": 0, "unknown": 0}
        rows = []
        unknown_rows = []
        company_rules = build_company_detection_rules(lines) if policy.get("company_profile") == "son_phuong" else []
        save_company_detection_rules(policy.get("company_profile", "yen_thanh"), company_rules)
        for line in lines:
            kind, reason = steel_kind_detail(line.get("product_name", ""), line.get("source_variant_code") or line.get("variant_code", ""))
            counts[kind] = counts.get(kind, 0) + 1
            if kind == "unknown":
                unknown_rows.append({
                    "row_number": line.get("row_number", ""),
                    "source_variant_code": line.get("source_variant_code", ""),
                    "profile_code": line.get("steel_profile_code", ""),
                    "variant_code": line.get("variant_code", ""),
                    "product_name": line.get("product_name", ""),
                    "quantity": line.get("quantity", 0),
                    "unit_price": line.get("unit_price"),
                    "kind": kind,
                    "kind_label": "Khong phan loai",
                    "reason": reason,
                })
            if len(rows) < 250:
                rows.append({
                    "row_number": line.get("row_number", ""),
                    "source_variant_code": line.get("source_variant_code", ""),
                    "profile_code": line.get("steel_profile_code", ""),
                    "profile_key": line.get("steel_profile_key", ""),
                    "variant_code": line.get("variant_code", ""),
                    "product_name": line.get("product_name", ""),
                    "quantity": line.get("quantity", 0),
                    "unit_price": line.get("unit_price"),
                    "kind": kind,
                    "kind_label": {"pipe": "Thép ống", "box": "Thép hộp", "unknown": "Không phân loại"}.get(kind, kind),
                    "reason": reason,
                })
        return jsonify({
            "profile": policy.get("company_profile", "yen_thanh"),
            "total": len(lines),
            "counts": counts,
            "rows": rows,
            "unknown_rows": unknown_rows,
            "company_rules": company_rules,
            "limited": len(lines) > len(rows),
        })
    except Exception as exc:
        return jsonify({"error": str(exc)}), 400


def analysis_payload(
    purchase_content,
    sales_content,
    opening_content,
    raw_mapping,
    policy,
    sales_filename,
    progress=None,
    barem_content=None,
    defer_workbook=False,
    result_job_id=None,
):
    progress = progress or (lambda _percent, _label, *_row_progress: None)
    progress(5, "Đang đọc cấu hình cột...")
    purchase_mapping = clean_mapping(raw_mapping, "purchase")
    sales_mapping = clean_mapping(raw_mapping, "sales")
    opening_mapping = clean_mapping(raw_mapping, "opening")

    policy = clean_policy(policy or DEFAULT_POLICY)
    company_profile = policy.get("company_profile", "yen_thanh")
    progress(12, "Đang đọc hóa đơn mua vào...")
    _, purchase_lines = read_lines(purchase_content, purchase_mapping, "purchase", company_profile=company_profile)
    progress(18, "Đang đọc file barem...")
    barem_map = merge_barem_maps(DEFAULT_BAREM_MAP, parse_barem_file(barem_content) if barem_content else None)
    progress(24, "Đang đọc hóa đơn bán ra...")
    _, sales_lines = read_lines(sales_content, sales_mapping, "sales", company_profile=company_profile)
    if policy.get("ignore_sale_suffix"):
        sales_lines = normalize_sales_codes_to_base(sales_lines)
    progress(32, "Đang đọc tồn đầu kỳ...")
    opening_lines = read_lines(opening_content, opening_mapping, "opening", company_profile=company_profile)[1] if opening_content else []
    if not purchase_lines and not opening_lines:
        raise ValueError("Không đọc được tồn kho từ hóa đơn mua vào hoặc file tồn đầu kỳ. Kiểm tra mapping cột.")
    if not sales_lines:
        raise ValueError("Không đọc được dòng bán ra. Kiểm tra mapping cột.")

    progress(38, "Đang rà soát mã chỉ bán ra...")
    sale_only_codes = find_sale_only_codes(opening_lines, purchase_lines, sales_lines)
    ambiguous_steel_rows = build_ambiguous_steel_rows(purchase_lines) if company_profile == "son_phuong" else []
    company_detection_rules = build_company_detection_rules(purchase_lines) if company_profile == "son_phuong" else []
    save_company_detection_rules(company_profile, company_detection_rules)

    def allocation_progress(done, total):
        ratio = done / total if total else 1
        progress(40 + ratio * 30, f"Đang phân bổ kho: {done}/{total} dòng bán ra...", done, total)

    allocations, stock_rows, summary, warnings = allocate_stock(
        opening_lines,
        purchase_lines,
        sales_lines,
        policy,
        progress_callback=allocation_progress,
        barem_map=barem_map,
    )
    summary["sale_only_code_count"] = len(sale_only_codes)
    missing_barem_report = summary.get("missing_barem_report", [])

    if defer_workbook:
        ledger = {}
        sales_report_rows = []
        verification_rows = []
        progress(74, "Đã hoàn tất tính Phân kho. Sổ chi tiết sẽ được dựng khi mở Báo cáo.")
    else:
        progress(74, "Đang dựng sổ chi tiết theo các kho trong file bán ra...")
        ledger = build_inventory_ledger(
            opening_lines,
            purchase_lines,
            allocations,
            sales_lines=sales_lines,
            company_profile=company_profile,
        )
        ledger["date_range"] = iso_in_line_range(purchase_lines, sales_lines)
        sales_report_rows = build_sales_report_rows_from_ledger(ledger) if ledger else build_sales_report_rows(allocations, purchase_lines, sales_lines)
        verification_rows = build_verification_rows(
            purchase_lines,
            sales_lines,
            allocations,
            ledger,
            sales_report_rows,
            policy,
        )

    filename = output_filename(sales_filename)
    if defer_workbook:
        job_id = result_job_id or uuid.uuid4().hex
        progress(94, "Đã hoàn tất Phân kho. File báo cáo sẽ được tạo khi bấm Xuất.")
    else:
        def output_progress(done, total, label):
            ratio = done / total if total else 1
            progress(78 + ratio * 16, label, done, total)

        progress(78, "Đang tạo file Excel xuất...")
        output = create_output_workbook(
            sales_content,
            sales_mapping,
            allocations,
            stock_rows,
            summary,
            policy,
            sale_only_codes,
            purchase_lines=purchase_lines,
            sales_lines=sales_lines,
            progress_callback=output_progress,
            ledger=ledger,
            missing_barem_report=missing_barem_report,
            ambiguous_steel_rows=ambiguous_steel_rows,
        )
        progress(94, "Đang lưu file kết quả...")
        job_id, filename, _ = save_output(output, sales_filename)
    if defer_workbook:
        report_view = None
        progress(99, "Đã sẵn sàng Review Phân kho.")
    else:
        progress(99, "Đang chuẩn bị giao diện báo cáo...")
        report_view = report_view_for_ui(ledger, sales_report_rows)
    result = {
        "job_id": job_id,
        "filename": filename,
        "summary": summary,
        "policy": policy,
        "warnings": warnings[:30],
        "missing_barem_report": missing_barem_report,
        "future_purchase_reorder_report": summary.get("future_purchase_reorder_report", []),
        "ambiguous_steel_rows": ambiguous_steel_rows,
        "company_detection_rules": company_detection_rules,
        "verification": [{
            "group": row[0],
            "check": row[1],
            "original_value": row[2],
            "processed_value": row[3],
            "difference": row[4],
            "tolerance": row[5],
            "status": row[6],
            "explanation": row[7],
        } for row in verification_rows],
        "sales_report_rows": sales_report_rows,
        "report_view": report_view,
        "allocation_count": len(allocations),
        "stock_count": len(stock_rows),
        "sale_only_codes": sale_only_codes,
        "ledger": ledger,
        "allocations": [{
            "row_number": item["row_number"],
            "variant_code": item["variant_code"],
            "base_code": item["base_code"],
            "product_name": item["product_name"],
            "quantity": item["quantity"],
            "invoice_no": item.get("invoice_no", ""),
            "invoice_date": item.get("invoice_date", ""),
            "invoice_date_iso": item.get("invoice_date_iso", ""),
            "sale_split_codes": item.get("sale_split_codes", ""),
            "material_quantity": item["material_quantity"],
            "unresolved_material_quantity": item.get("unresolved_material_quantity", 0),
            "finished_quantity": item["finished_quantity"],
            "finished_variant_code": item.get("finished_variant_code", item.get("variant_code", "")),
            "allocation_role": item.get("allocation_role", ""),
            "warehouse_code": item.get("warehouse_code", ""),
            "warehouse_account": item.get("warehouse_account", ""),
            "remainder_warehouse_code": item.get("remainder_warehouse_code", ""),
            "remainder_warehouse_account": item.get("remainder_warehouse_account", ""),
            "negative_warning": bool(item.get("negative_warning")),
            "generic_plan_note": item.get("generic_plan_note", ""),
            "inventory_before": item["inventory_before"],
            "used": item["used"],
            "inventory_after": item["inventory_after"],
            "rejected": item["rejected"],
            "inventory_before_detail": inventory_text(item["inventory_before"]),
            "detail": detail_text(item["used"]),
            "inventory_after_detail": inventory_text(item["inventory_after"]),
            "rejected_detail": rejected_text(item["rejected"]),
        } for item in allocations],
        "stock_rows": stock_rows,
    }
    if defer_workbook:
        result.update({
            # Keep the complete allocation rows server-side. The compact public
            # rows intentionally omit financial fields, so they cannot rebuild
            # revenue/tax reports later without this private copy.
            "_allocations": allocations,
            "_sales_content": sales_content,
            "_sales_mapping": sales_mapping,
            "_purchase_lines": purchase_lines,
            "_sales_lines": sales_lines,
            "_opening_lines": opening_lines,
            "_ambiguous_steel_rows": ambiguous_steel_rows,
        })
    return result


def update_analysis_job(job_id, **fields):
    with ANALYSIS_LOCK:
        job = ANALYSIS_JOBS.setdefault(job_id, {})
        job.update(fields)


def get_analysis_job(job_id):
    with ANALYSIS_LOCK:
        job = ANALYSIS_JOBS.get(job_id)
        return dict(job) if job else None


def run_analysis_job(
    job_id,
    purchase_content,
    sales_content,
    opening_content,
    raw_mapping,
    policy,
    sales_filename,
    barem_content=None,
    defer_workbook=False,
):
    def progress(percent, label, done=None, total=None):
        fields = {"status": "running", "progress": round(percent), "label": label}
        if done is not None and total is not None:
            safe_total = max(0, int(total or 0))
            safe_done = max(0, min(int(done or 0), safe_total)) if safe_total else 0
            fields["done"] = safe_done
            fields["total"] = safe_total
            fields["progress"] = round((safe_done / safe_total) * 100) if safe_total else 0
        else:
            fields["done"] = 0
            fields["total"] = 0
        update_analysis_job(job_id, **fields)

    try:
        result = analysis_payload(
            purchase_content,
            sales_content,
            opening_content,
            raw_mapping,
            policy,
            sales_filename,
            progress=progress,
            barem_content=barem_content,
            defer_workbook=defer_workbook,
            result_job_id=job_id,
        )
        update_analysis_job(job_id, status="complete", progress=100, done=0, total=0, label="Hoàn tất phân bổ tồn kho.", result=result)
    except Exception as exc:
        update_analysis_job(job_id, status="error", progress=0, done=0, total=0, label=str(exc), error=str(exc))


@app.post("/api/analyze")
def analyze():
    purchase_file = request.files.get("purchase_file")
    sales_file = request.files.get("sales_file")
    opening_file = request.files.get("opening_file")
    barem_file = request.files.get("barem_file")
    if not purchase_file or not sales_file:
        return jsonify({"error": "Cần chọn hóa đơn mua vào và hóa đơn bán ra."}), 400
    try:
        raw_mapping = json.loads(request.form.get("mapping") or "{}")
        policy = clean_policy(json.loads(request.form.get("policy") or "{}"))
        purchase_content = purchase_file.read()
        sales_content = sales_file.read()
        opening_content = opening_file.read() if opening_file and opening_file.filename else None
        barem_content = barem_file.read() if barem_file and barem_file.filename else None
        return jsonify(analysis_payload(
            purchase_content,
            sales_content,
            opening_content,
            raw_mapping,
            policy,
            sales_file.filename,
            barem_content=barem_content,
        ))
    except Exception as exc:
        return jsonify({"error": str(exc)}), 400


@app.post("/api/analyze-job")
def analyze_job():
    purchase_file = request.files.get("purchase_file")
    sales_file = request.files.get("sales_file")
    opening_file = request.files.get("opening_file")
    barem_file = request.files.get("barem_file")
    if not purchase_file or not sales_file:
        return jsonify({"error": "Cần chọn hóa đơn mua vào và hóa đơn bán ra."}), 400
    try:
        raw_mapping = json.loads(request.form.get("mapping") or "{}")
        policy = clean_policy(json.loads(request.form.get("policy") or "{}"))
        purchase_content = purchase_file.read()
        sales_content = sales_file.read()
        opening_content = opening_file.read() if opening_file and opening_file.filename else None
        barem_content = barem_file.read() if barem_file and barem_file.filename else None
        job_id = uuid.uuid4().hex
        update_analysis_job(job_id, status="queued", progress=0, done=0, total=0, label="Đã tải file lên server. Đang xếp hàng xử lý...")
        worker = threading.Thread(
            target=run_analysis_job,
            args=(job_id, purchase_content, sales_content, opening_content, raw_mapping, policy, sales_file.filename, barem_content),
            daemon=True,
        )
        worker.start()
        return jsonify({"analysis_job_id": job_id})
    except Exception as exc:
        return jsonify({"error": str(exc)}), 400


@app.get("/api/analyze-job/<job_id>")
def analyze_job_status(job_id):
    if not re.fullmatch(r"[a-f0-9]{32}", job_id):
        return jsonify({"error": "Mã xử lý không hợp lệ."}), 404
    job = get_analysis_job(job_id)
    if not job:
        return jsonify({"error": "Không tìm thấy tiến trình xử lý."}), 404
    return jsonify(job)


@app.get("/api/download/<job_id>")
def download(job_id):
    if not re.fullmatch(r"[a-f0-9]{32}", job_id):
        return jsonify({"error": "Mã kết quả không hợp lệ."}), 404
    matching = list(OUTPUT_DIR.glob(f"{job_id}_*.xlsx"))
    if not matching:
        return jsonify({"error": "Không tìm thấy file kết quả."}), 404
    path = matching[0]
    filename = path.name[len(job_id) + 1:]
    return send_file(
        path,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def ps_quote(value):
    return "'" + str(value).replace("'", "''") + "'"


def installed_exe_path():
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve()
    return (RESOURCE_DIR / "deploy" / "InventoryAllocator.exe").resolve()


def write_update_script(update_exe, target_exe):
    script_path = BASE_DIR / "update_inventory_allocator.ps1"
    script_path.parent.mkdir(parents=True, exist_ok=True)
    desktop_shortcut = "Inventory Allocator.lnk"
    current_pid = os.getpid()
    script = f"""
$ErrorActionPreference = 'SilentlyContinue'
$UpdateExe = {ps_quote(update_exe)}
$TargetExe = {ps_quote(target_exe)}
$CurrentPid = {current_pid}
$ShortcutName = {ps_quote(desktop_shortcut)}
Start-Sleep -Seconds 2

Get-Process -Name 'InventoryAllocator' -ErrorAction SilentlyContinue |
  Where-Object {{ $_.Id -ne $PID }} |
  Stop-Process -Force -ErrorAction SilentlyContinue
for ($i = 0; $i -lt 40; $i++) {{
  $stillRunning = Get-Process -Name 'InventoryAllocator' -ErrorAction SilentlyContinue
  if (-not $stillRunning) {{ break }}
  Start-Sleep -Milliseconds 250
}}

$desktopPaths = @(
  [Environment]::GetFolderPath('Desktop'),
  [Environment]::GetFolderPath('CommonDesktopDirectory')
) | Where-Object {{ $_ -and (Test-Path $_) }} | Select-Object -Unique

$shell = New-Object -ComObject WScript.Shell
foreach ($desktop in $desktopPaths) {{
  Get-ChildItem -LiteralPath $desktop -Filter '*.lnk' -ErrorAction SilentlyContinue | ForEach-Object {{
    $remove = $false
    try {{
      $shortcut = $shell.CreateShortcut($_.FullName)
      $target = [string]$shortcut.TargetPath
      $icon = [string]$shortcut.IconLocation
      if ($_.BaseName -like '*Inventory Allocator*' -or $_.BaseName -like '*InventoryAllocator*') {{ $remove = $true }}
      if ($target -like '*InventoryAllocator.exe*' -or $icon -like '*InventoryAllocator.exe*') {{ $remove = $true }}
    }} catch {{}}
    if ($remove) {{ Remove-Item -LiteralPath $_.FullName -Force -ErrorAction SilentlyContinue }}
  }}
}}

New-Item -ItemType Directory -Force -Path (Split-Path -Parent $TargetExe) | Out-Null
for ($i = 0; $i -lt 40; $i++) {{
  if (Test-Path $TargetExe) {{
    Remove-Item -LiteralPath $TargetExe -Force -ErrorAction SilentlyContinue
  }}
  if (-not (Test-Path $TargetExe)) {{ break }}
  Start-Sleep -Milliseconds 250
}}
if (Test-Path $TargetExe) {{
  throw "Không xóa được exe cũ: $TargetExe"
}}
Copy-Item -LiteralPath $UpdateExe -Destination $TargetExe -Force

$userDesktop = [Environment]::GetFolderPath('Desktop')
if ($userDesktop -and (Test-Path $TargetExe)) {{
  $shortcutPath = Join-Path $userDesktop $ShortcutName
  $link = $shell.CreateShortcut($shortcutPath)
  $link.TargetPath = $TargetExe
  $link.WorkingDirectory = Split-Path -Parent $TargetExe
  $link.IconLocation = $TargetExe + ',0'
  $link.Description = 'Inventory Allocator version {APP_VERSION}'
  $link.Save()
}}

Start-Process -FilePath $TargetExe -WorkingDirectory (Split-Path -Parent $TargetExe)
Start-Sleep -Seconds 2
Remove-Item -LiteralPath $UpdateExe -Force -ErrorAction SilentlyContinue
Remove-Item -LiteralPath $MyInvocation.MyCommand.Path -Force -ErrorAction SilentlyContinue
"""
    script_path.write_text(script, encoding="utf-8")
    return script_path


@app.get("/api/version")
def version():
    return jsonify({
        "version": APP_VERSION,
        "pid": os.getpid(),
        "exe": str(installed_exe_path()),
    })


@app.get("/api/default-barem")
def default_barem():
    return jsonify({
        "rows": builtin_barem_rows(),
        "count": len(DEFAULT_BAREM_MAP["by_profile"]),
    })


@app.post("/api/update-exe")
def update_exe():
    uploaded = request.files.get("exe_file")
    if not uploaded or not uploaded.filename:
        return jsonify({"error": "Cần chọn file .exe để cập nhật."}), 400
    if not uploaded.filename.lower().endswith(".exe"):
        return jsonify({"error": "File cập nhật phải là .exe."}), 400
    update_dir = BASE_DIR / "updates"
    update_dir.mkdir(parents=True, exist_ok=True)
    update_path = update_dir / f"InventoryAllocator_update_{uuid.uuid4().hex}.exe"
    uploaded.save(update_path)
    if update_path.stat().st_size <= 0:
        update_path.unlink(missing_ok=True)
        return jsonify({"error": "File .exe cập nhật bị rỗng."}), 400
    target_exe = installed_exe_path()
    script_path = write_update_script(update_path, target_exe)
    subprocess.Popen(
        [
            "powershell.exe",
            "-NoProfile",
            "-ExecutionPolicy",
            "Bypass",
            "-File",
            str(script_path),
        ],
        creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0),
        cwd=str(target_exe.parent),
    )
    return jsonify({
        "message": "Đã nhận file cập nhật. Ứng dụng sẽ tự dừng, thay exe mới, tạo lại shortcut và mở lại.",
        "version": APP_VERSION,
    })


@app.get("/api/opening-template")
def opening_template():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "TonDauKy"
    write_table(sheet, ["Mã VT", "Tên hàng", "Số lượng đầu kỳ", "Đơn giá vốn (tùy chọn)"], [
        ["MAHANG.001", "Ví dụ hàng hóa", 0, ""],
    ])
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return send_file(
        stream,
        as_attachment=True,
        download_name="mau_ton_dau_ky.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/", defaults={"path": ""})
@app.route("/<path:path>")
def frontend(path):
    candidate = STATIC_DIR / path
    if path and candidate.exists() and candidate.is_file():
        return send_from_directory(STATIC_DIR, path)
    return send_from_directory(STATIC_DIR, "index.html")


def open_browser():
    webbrowser.open(f"http://127.0.0.1:{APP_PORT}")


def tray_icon_image():
    if Image is None:
        return None
    if ICON_PATH.exists():
        with Image.open(ICON_PATH) as image:
            return image.convert("RGBA").copy()
    image = Image.new("RGBA", (64, 64), (0, 0, 0, 0))
    draw = ImageDraw.Draw(image)
    draw.rounded_rectangle((8, 8, 56, 56), radius=13, fill=(7, 10, 14, 255))
    draw.rounded_rectangle((17, 29, 42, 44), radius=5, fill=(248, 250, 252, 255))
    draw.rounded_rectangle((41, 32, 52, 41), radius=5, fill=(248, 250, 252, 255))
    draw.rounded_rectangle((44, 34, 49, 39), radius=3, fill=(7, 10, 14, 255))
    draw.ellipse((14, 43, 47, 51), fill=(248, 250, 252, 255))
    for x in (24, 31, 38):
        draw.arc((x - 4, 14, x + 4, 29), 100, 260, fill=(248, 250, 252, 230), width=2)
    return image


def open_from_tray(icon=None, item=None):
    open_browser()


def quit_from_tray(icon=None, item=None):
    if icon is not None:
        icon.stop()
    os._exit(0)


def run_tray_icon():
    if pystray is None:
        return False
    image = tray_icon_image()
    if image is None:
        return False
    menu = pystray.Menu(
        pystray.MenuItem("Mở Inventory Allocator", open_from_tray, default=True),
        pystray.MenuItem("Dừng ứng dụng", quit_from_tray),
    )
    icon = pystray.Icon("InventoryAllocator", image, "Inventory Allocator", menu)
    icon.run()
    return True


def write_startup_log(message):
    BASE_DIR.mkdir(parents=True, exist_ok=True)
    with (BASE_DIR / "startup.log").open("a", encoding="utf-8") as handle:
        handle.write(f"[{datetime.now().isoformat(timespec='seconds')}] {message}\n")


def run_server():
    try:
        write_startup_log(f"Starting Flask server on 127.0.0.1:{APP_PORT}")
        app.run(host="127.0.0.1", port=APP_PORT, debug=False, use_reloader=False)
    except BaseException:
        write_startup_log(traceback.format_exc())


if __name__ == "__main__":
    threading.Thread(target=run_server, daemon=True).start()
    threading.Timer(0.8, open_browser).start()
    if not run_tray_icon():
        try:
            while True:
                time.sleep(3600)
        except KeyboardInterrupt:
            pass
