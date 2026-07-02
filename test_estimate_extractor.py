import unittest
from io import BytesIO
import os
from pathlib import Path
from tempfile import TemporaryDirectory

import xlrd
from openpyxl import Workbook, load_workbook

from workflows.estimate_extractor import logic
from workflows.estimate_extractor.logic import analyze_estimate_workbook, create_estimate_output_workbook, sample_path_from_env


class EstimateExtractorLogicTest(unittest.TestCase):
    def test_ho_guom_sample_matches_visible_du_thau_totals(self):
        sample_path = sample_path_from_env()
        if not sample_path:
            self.skipTest("Ho Guom estimate sample workbook is not available on this machine")

        result = analyze_estimate_workbook(sample_path)

        self.assertEqual(result.bid_rows, 172)
        self.assertEqual(result.detail_blocks, 172)
        self.assertEqual(result.identity_mismatches, [])
        self.assertEqual(result.calculation_mismatches, [])
        self.assertEqual(result.unclassified_rows, [])

    def test_known_helper_conflicts_are_detected_but_section_text_wins(self):
        sample_path = sample_path_from_env()
        if not sample_path:
            self.skipTest("Ho Guom estimate sample workbook is not available on this machine")

        result = analyze_estimate_workbook(sample_path)

        rows = {item["row"]: item for item in result.helper_mismatches}
        self.assertEqual(rows[713]["label"], "NC")
        self.assertEqual(rows[713]["helper"], "MTC")
        self.assertEqual(rows[765]["label"], "VL")
        self.assertEqual(rows[765]["helper"], "NC")
        self.assertEqual(rows[778]["label"], "VL")
        self.assertEqual(rows[778]["helper"], "NC")

    def test_ho_guom_sample_thvt_formula_matches_existing_material_summary(self):
        sample_path = sample_path_from_env()
        if not sample_path:
            self.skipTest("Ho Guom estimate sample workbook is not available on this machine")

        result = analyze_estimate_workbook(sample_path)

        self.assertEqual(result.thvt_rows, 326)
        self.assertEqual(result.generated_thvt_rows, 328)
        self.assertEqual(result.thvt_mismatches, [])
        self.assertEqual(result.thvt_extra_rows, [])
        self.assertEqual(len(result.thvt_key_mismatches), 4)
        self.assertEqual(
            [(item["source_no"], item["source_detail_row"]) for item in result.thvt_missing_rows],
            [("47", 765), ("48", 778)],
        )

    def test_output_workbook_is_openable_xlsx(self):
        sample_path = sample_path_from_env()
        if not sample_path:
            self.skipTest("Ho Guom estimate sample workbook is not available on this machine")

        content, summary = create_estimate_output_workbook(sample_path)
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)

        self.assertEqual(workbook.sheetnames, ["Dự thầu", "Chiết tính", "THVT", "Tổng hợp THVT"])
        self.assertEqual(summary["bid_rows"], 172)
        self.assertGreater(workbook["THVT"].max_row, 100)
        self.assertGreater(workbook["Tổng hợp THVT"].max_row, 50)

    def test_output_thvt_summary_groups_same_material_and_price(self):
        sample_path = sample_path_from_env()
        if not sample_path:
            self.skipTest("Ho Guom estimate sample workbook is not available on this machine")

        content, _summary = create_estimate_output_workbook(sample_path)
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        detail = workbook["THVT"]
        summary = workbook["Tổng hợp THVT"]

        detail_totals = {}
        for row in detail.iter_rows(min_row=2, values_only=True):
            key = (
                str(row[3] or "").upper(),
                logic._fold(row[5]),
                logic._fold(row[6]),
                logic._fold(row[7]),
                round(float(row[11] or 0), 8),
            )
            bucket = detail_totals.setdefault(key, {"qty": 0.0, "amount": 0.0, "count": 0})
            bucket["qty"] += float(row[10] or 0)
            bucket["amount"] += float(row[12] or 0)
            bucket["count"] += 1

        summary_totals = {}
        for row in summary.iter_rows(min_row=2, values_only=True):
            key = (
                str(row[1] or "").upper(),
                logic._fold(row[2]),
                logic._fold(row[3]),
                logic._fold(row[4]),
                round(float(row[6] or 0), 8),
            )
            summary_totals[key] = {
                "qty": float(row[5] or 0),
                "amount": float(row[7] or 0),
                "count": int(row[8] or 0),
            }

        self.assertEqual(set(summary_totals), set(detail_totals))
        self.assertTrue(any(item["count"] > 1 for item in summary_totals.values()))
        for key, expected in detail_totals.items():
            actual = summary_totals[key]
            self.assertEqual(actual["count"], expected["count"], key)
            self.assertAlmostEqual(actual["qty"], expected["qty"], delta=0.000001, msg=str(key))
            self.assertAlmostEqual(actual["amount"], expected["amount"], delta=0.01, msg=str(key))

    def test_output_du_thau_generated_columns_match_sample_for_all_labels(self):
        sample_path = sample_path_from_env()
        if not sample_path:
            self.skipTest("Ho Guom estimate sample workbook is not available on this machine")

        source = xlrd.open_workbook(sample_path, formatting_info=True)
        sample_bid = source.sheet_by_index(0)
        sample_detail = source.sheet_by_index(1)
        bid_header = logic._find_header_row(sample_bid, ["ma so", "ten cong tac", "khoi luong"])
        detail_header = logic._find_header_row(sample_detail, ["ma so", "thanh phan hao phi", "thanh tien"])
        bid_cols = logic._bid_columns(sample_bid, bid_header)
        detail_cols = logic._detail_columns(sample_detail, detail_header)
        bid_rows = logic._bid_item_rows(sample_bid, bid_header + 1, bid_cols)
        detail_blocks = logic._detail_blocks(sample_detail, detail_header + 1, detail_cols)
        price_start = _find_generated_group_start(sample_bid, bid_header, bid_cols.amount_total + 1)
        amount_start = _find_generated_group_start(sample_bid, bid_header, price_start + len(logic.COST_LABEL_ORDER))

        content, _summary = create_estimate_output_workbook(sample_path)
        output_bid = load_workbook(BytesIO(content), read_only=True, data_only=True).worksheets[0]

        self.assertEqual(output_bid.cell(row=bid_header + 1, column=price_start + 1).value, "VL")
        self.assertEqual(output_bid.cell(row=bid_header + 1, column=amount_start + 1).value, "VL")

        for block_index, row in enumerate(bid_rows):
            excel_row = row + 1
            start, end = detail_blocks[block_index]
            qty = logic._number(logic._cell(sample_bid, row, bid_cols.qty))
            sums = logic._detail_block_sums(sample_detail, start + 1, end, detail_cols)
            total = sums["TC"] or sum(sums[label] for label in logic.COST_LABEL_ORDER if label != "TC")
            sums["TC"] = total
            for offset, label in enumerate(logic.COST_LABEL_ORDER):
                unit_expected = sums[label]
                amount_expected = unit_expected * qty
                unit_actual = output_bid.cell(row=excel_row, column=price_start + offset + 1).value or 0
                amount_actual = output_bid.cell(row=excel_row, column=amount_start + offset + 1).value or 0
                self.assertAlmostEqual(
                    float(unit_actual),
                    unit_expected,
                    delta=1.0,
                    msg=f"{label} unit-price mismatch at Excel row {excel_row}",
                )
                self.assertAlmostEqual(
                    float(amount_actual),
                    amount_expected,
                    delta=2.0,
                    msg=f"{label} amount mismatch at Excel row {excel_row}",
                )

    def test_tba_sample_uses_plain_bid_sheet_and_parent_multipliers(self):
        sample_path = tba_sample_path_from_env()
        if not sample_path:
            self.skipTest("Ho Guom TBA estimate sample workbook is not available on this machine")

        result = analyze_estimate_workbook(sample_path)

        self.assertEqual(result.bid_rows, 784)
        self.assertEqual(result.detail_blocks, 784)
        self.assertEqual(result.identity_mismatches, [])
        self.assertEqual(result.calculation_mismatches, [])
        self.assertEqual(result.unclassified_rows, [])

    def test_tba_output_multiplies_child_quantities_and_keeps_bid_units(self):
        sample_path = tba_sample_path_from_env()
        if not sample_path:
            self.skipTest("Ho Guom TBA estimate sample workbook is not available on this machine")

        content, _summary = create_estimate_output_workbook(sample_path)
        output_bid = load_workbook(BytesIO(content), read_only=True, data_only=True).worksheets[0]
        amount_start = _find_openpyxl_generated_group_start(output_bid, 4, 23)
        tc_amount_col = amount_start + len(logic.COST_LABEL_ORDER) - 1

        self.assertAlmostEqual(float(output_bid.cell(row=38, column=5).value), 1.95, delta=0.000001)
        self.assertAlmostEqual(float(output_bid.cell(row=38, column=10).value), 68463.802005768, delta=0.01)
        self.assertAlmostEqual(float(output_bid.cell(row=38, column=tc_amount_col).value), 68463.802005768, delta=0.01)
        self.assertAlmostEqual(float(output_bid.cell(row=37, column=10).value), 830282.2873806191, delta=0.01)
        self.assertAlmostEqual(
            sum(float(output_bid.cell(row=row, column=10).value or 0) for row in range(38, 42)),
            float(output_bid.cell(row=37, column=10).value),
            delta=0.01,
        )

        self.assertAlmostEqual(float(output_bid.cell(row=44, column=5).value), 12.4, delta=0.000001)
        self.assertAlmostEqual(float(output_bid.cell(row=44, column=10).value), 6179648.768402505, delta=0.01)
        self.assertAlmostEqual(float(output_bid.cell(row=44, column=tc_amount_col).value), 6179648.768402505, delta=0.01)

        self.assertAlmostEqual(float(output_bid.cell(row=69, column=10).value), 946271.6916545785, delta=0.01)
        self.assertAlmostEqual(
            sum(float(output_bid.cell(row=row, column=10).value or 0) for row in range(70, 74)),
            float(output_bid.cell(row=69, column=10).value),
            delta=0.01,
        )
        self.assertEqual(output_bid.cell(row=31, column=1).value, "2.2.8")
        self.assertEqual(output_bid.cell(row=37, column=1).value, "2.3.1")
        self.assertIsNone(output_bid.cell(row=38, column=1).value)
        self.assertEqual(output_bid.cell(row=43, column=1).value, "2.3.2")
        self.assertEqual(output_bid.cell(row=50, column=1).value, "2.3.3")
        self.assertIsNone(output_bid.cell(row=51, column=1).value)
        self.assertEqual(output_bid.cell(row=83, column=1).value, "2.5.1")
        self.assertEqual(output_bid.cell(row=84, column=4).value, "10 mốc")
        self.assertAlmostEqual(float(output_bid.cell(row=762, column=tc_amount_col).value), 1084445.29050624, delta=0.01)

    def test_unit_mismatch_converts_detail_unit_to_bid_unit(self):
        with TemporaryDirectory() as tmp:
            path = Path(tmp) / "unit_conversion.xlsx"
            workbook = Workbook()
            bid = workbook.active
            bid.title = "Du thau"
            detail = workbook.create_sheet("Chiet tinh")

            bid.append([])
            bid.append([])
            bid.append([])
            bid.append(["STT", "Ma so", "Ten cong tac", "Don vi", "Khoi luong", "Vat lieu", "Nhan cong", "May", "Don gia", "Thanh tien"])
            bid.append([1, "AB.001", "Dao dat", "m3", 2, 50, 30, 20, 100, 200])

            detail.append([])
            detail.append([])
            detail.append([])
            detail.append(["STT", "Ma so", "Thanh phan hao phi", "Don vi", "Dinh muc", "Don gia", "He so", "Thanh tien"])
            detail.append([1, "AB.001", "Dao dat", "100m3", 1, None, None, None])
            detail.append([None, None, "a) Vat lieu", None, None, None, None, 5000])
            detail.append([None, "A1", "Cat", "m3", 2, 2500, 1, 5000])
            detail.append([None, None, "b) Nhan cong", None, None, None, None, 3000])
            detail.append([None, "N1", "Nhan cong", "cong", 3, 1000, 1, 3000])
            detail.append([None, None, "c) May thi cong", None, None, None, None, 2000])
            detail.append([None, "M1", "May dao", "ca", 1, 2000, 1, 2000])
            detail.append([None, None, "Cong chi phi truc tiep (VL+NC+M)", "T", None, None, None, 10000])
            workbook.save(path)

            content, summary = create_estimate_output_workbook(str(path))
            self.assertEqual(summary["calculation_mismatches"], 0)
            output = load_workbook(BytesIO(content), read_only=True, data_only=True)
            output_bid = output.worksheets[0]
            output_thvt = output["THVT"]
            price_start = _find_openpyxl_generated_group_start(output_bid, 4, 11)
            amount_start = price_start + len(logic.COST_LABEL_ORDER)
            tc_amount_col = amount_start + len(logic.COST_LABEL_ORDER) - 1

            self.assertEqual(output_bid.cell(row=5, column=4).value, "m3")
            self.assertAlmostEqual(float(output_bid.cell(row=5, column=price_start).value), 50, delta=0.000001)
            self.assertAlmostEqual(float(output_bid.cell(row=5, column=price_start + 1).value), 30, delta=0.000001)
            self.assertAlmostEqual(float(output_bid.cell(row=5, column=price_start + 2).value), 20, delta=0.000001)
            self.assertAlmostEqual(float(output_bid.cell(row=5, column=amount_start).value), 100, delta=0.000001)
            self.assertAlmostEqual(float(output_bid.cell(row=5, column=amount_start + 1).value), 60, delta=0.000001)
            self.assertAlmostEqual(float(output_bid.cell(row=5, column=amount_start + 2).value), 40, delta=0.000001)
            self.assertAlmostEqual(float(output_bid.cell(row=5, column=tc_amount_col).value), 200, delta=0.000001)

            self.assertAlmostEqual(float(output_thvt.cell(row=2, column=10).value), 2, delta=0.000001)
            self.assertAlmostEqual(float(output_thvt.cell(row=2, column=11).value), 0.04, delta=0.000001)
            self.assertAlmostEqual(float(output_thvt.cell(row=2, column=13).value), 100, delta=0.000001)

    def test_other_material_uses_amount_column_as_unit_price(self):
        with TemporaryDirectory() as tmp:
            path = Path(tmp) / "other_material.xlsx"
            workbook = Workbook()
            bid = workbook.active
            bid.title = "Du thau"
            detail = workbook.create_sheet("Chiet tinh")

            bid.append([])
            bid.append([])
            bid.append([])
            bid.append(["STT", "Ma so", "Ten cong tac", "Don vi", "Khoi luong", "Vat lieu", "Nhan cong", "May", "Don gia", "Thanh tien"])
            bid.append([1, "AB.002", "Cong tac co vat lieu khac", "bo", 2, 300, 0, 0, 300, 600])

            detail.append([])
            detail.append([])
            detail.append([])
            detail.append(["STT", "Ma so", "Thanh phan hao phi", "Don vi", "Dinh muc", "Don gia", "He so", "Thanh tien"])
            detail.append([1, "AB.002", "Cong tac co vat lieu khac", "bo", 1, None, None, None])
            detail.append([None, None, "a) Vat lieu", None, None, None, None, 300])
            detail.append([None, "Vat lieu khac", "Vat lieu khac", "bo", 3, 10, 2, 300])
            detail.append([None, None, "Cong chi phi truc tiep (VL+NC+M)", "T", None, None, None, 300])
            workbook.save(path)

            content, summary = create_estimate_output_workbook(str(path))
            self.assertEqual(summary["calculation_mismatches"], 0)
            output = load_workbook(BytesIO(content), read_only=True, data_only=True)
            output_thvt = output["THVT"]

            self.assertEqual(output_thvt.cell(row=2, column=6).value, "Vat lieu khac")
            self.assertEqual(output_thvt.cell(row=2, column=7).value, "Vat lieu khac")
            self.assertAlmostEqual(float(output_thvt.cell(row=2, column=9).value), 3, delta=0.000001)
            self.assertAlmostEqual(float(output_thvt.cell(row=2, column=10).value), 2, delta=0.000001)
            self.assertAlmostEqual(float(output_thvt.cell(row=2, column=11).value), 2, delta=0.000001)
            self.assertAlmostEqual(float(output_thvt.cell(row=2, column=12).value), 300, delta=0.000001)
            self.assertAlmostEqual(float(output_thvt.cell(row=2, column=13).value), 600, delta=0.000001)

    def test_generated_bid_vl_amount_total_matches_thvt_vl_amount_total(self):
        with TemporaryDirectory() as tmp:
            path = Path(tmp) / "vl_total_match.xlsx"
            workbook = Workbook()
            bid = workbook.active
            bid.title = "Du thau"
            detail = workbook.create_sheet("Chiet tinh")

            bid.append([])
            bid.append([])
            bid.append([])
            bid.append(["STT", "Ma so", "Ten cong tac", "Don vi", "Khoi luong", "Vat lieu", "Nhan cong", "May", "Don gia", "Thanh tien"])
            bid.append([1, "AB.003", "Cong tac co vat lieu thuong va vat lieu khac", "bo", 2, 350, 0, 0, 350, 700])

            detail.append([])
            detail.append([])
            detail.append([])
            detail.append(["STT", "Ma so", "Thanh phan hao phi", "Don vi", "Dinh muc", "Don gia", "He so", "Thanh tien"])
            detail.append([1, "AB.003", "Cong tac co vat lieu thuong va vat lieu khac", "bo", 1, None, None, None])
            detail.append([None, None, "a) Vat lieu", None, None, None, None, 350])
            detail.append([None, "CAT", "Cat", "m3", 1, 50, 1, 50])
            detail.append([None, "Vat lieu khac", "Vat lieu khac", "bo", 3, 10, 2, 300])
            detail.append([None, None, "Cong chi phi truc tiep (VL+NC+M)", "T", None, None, None, 350])
            workbook.save(path)

            content, summary = create_estimate_output_workbook(str(path))
            self.assertEqual(summary["calculation_mismatches"], 0)
            output = load_workbook(BytesIO(content), read_only=True, data_only=True)
            output_bid = output.worksheets[0]
            output_thvt = output["THVT"]
            price_start = _find_openpyxl_generated_group_start(output_bid, 4, 11)
            amount_start = price_start + len(logic.COST_LABEL_ORDER)

            bid_vl_amount_total = sum(float(output_bid.cell(row=row, column=amount_start).value or 0) for row in range(5, output_bid.max_row + 1))
            thvt_vl_amount_total = sum(
                float(row[12] or 0)
                for row in output_thvt.iter_rows(min_row=2, values_only=True)
                if str(row[3] or "").upper() == "VL"
            )

            self.assertAlmostEqual(bid_vl_amount_total, 700, delta=0.000001)
            self.assertAlmostEqual(thvt_vl_amount_total, 700, delta=0.000001)
            self.assertAlmostEqual(thvt_vl_amount_total, bid_vl_amount_total, delta=0.000001)


def _find_generated_group_start(sheet, header_row: int, start_col: int) -> int:
    for col in range(start_col, sheet.ncols):
        if logic._text(logic._cell(sheet, header_row, col)).upper() == "VL":
            return col
    raise AssertionError("Cannot find generated VL group in sample workbook")


def _find_openpyxl_generated_group_start(worksheet, header_row: int, start_col: int) -> int:
    for col in range(start_col, worksheet.max_column + 1):
        if str(worksheet.cell(row=header_row, column=col).value or "").upper() == "VL":
            return col
    raise AssertionError("Cannot find generated VL group in output workbook")


def tba_sample_path_from_env() -> str | None:
    explicit = os.environ.get("ESTIMATE_TBA_SAMPLE_XLS")
    if explicit and os.path.exists(explicit):
        return explicit
    root = Path("E:/Excel Mom")
    if not root.exists():
        return None
    for doc_dir in root.iterdir():
        if not doc_dir.is_dir() or not doc_dir.name.startswith("D"):
            continue
        folder = doc_dir / "Ho Guom"
        if not folder.exists():
            continue
        for candidate in folder.iterdir():
            if candidate.suffix.lower() == ".xls" and "TBA 22KV" in candidate.name:
                return str(candidate)
    return None


if __name__ == "__main__":
    unittest.main()
