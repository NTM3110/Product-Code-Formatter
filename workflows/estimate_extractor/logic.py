"""Prototype logic for extracting estimate buckets from Du thau/Chiet tinh sheets.

This module is intentionally UI/API-free. It reads a workbook, pairs each Du thau
line with the matching Chiet tinh block, classifies Chiet tinh rows into cost
buckets, and validates the totals against the visible Du thau columns.
"""

from dataclasses import dataclass, field
from io import BytesIO
import os
import re
import unicodedata

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


DIRECT_LABELS = {"VL", "NC", "MTC"}
INDIRECT_LABELS = {"CPC", "LT", "CPK", "TNCT", "VAT", "TC"}
ALL_LABELS = DIRECT_LABELS | INDIRECT_LABELS
COST_LABEL_ORDER = ["VL", "NC", "MTC", "CPC", "LT", "CPK", "TNCT", "VAT", "TC"]


@dataclass
class SheetColumns:
    stt: int
    code: int | None
    name: int
    unit: int
    qty: int | None = None
    vl: int | None = None
    nc: int | None = None
    mtc: int | None = None
    unit_total: int | None = None
    amount_total: int | None = None
    norm: int | None = None
    price: int | None = None
    coef: int | None = None
    amount: int | None = None
    helper: int | None = None


@dataclass
class EstimateAnalysisResult:
    path: str
    sheet_names: list[str]
    bid_rows: int
    detail_blocks: int
    identity_mismatches: list[dict] = field(default_factory=list)
    calculation_mismatches: list[dict] = field(default_factory=list)
    helper_mismatches: list[dict] = field(default_factory=list)
    unclassified_rows: list[dict] = field(default_factory=list)
    thvt_rows: int = 0
    generated_thvt_rows: int = 0
    thvt_mismatches: list[dict] = field(default_factory=list)
    thvt_key_mismatches: list[dict] = field(default_factory=list)
    thvt_missing_rows: list[dict] = field(default_factory=list)
    thvt_extra_rows: list[dict] = field(default_factory=list)
    label_totals: dict[str, float] = field(default_factory=dict)

    @property
    def ok(self) -> bool:
        return (
            self.bid_rows == self.detail_blocks
            and not self.identity_mismatches
            and not self.calculation_mismatches
            and not self.unclassified_rows
        )

    def summary(self) -> dict:
        return {
            "path": self.path,
            "bid_rows": self.bid_rows,
            "detail_blocks": self.detail_blocks,
            "identity_mismatches": len(self.identity_mismatches),
            "calculation_mismatches": len(self.calculation_mismatches),
            "helper_mismatches": len(self.helper_mismatches),
            "unclassified_rows": len(self.unclassified_rows),
            "thvt_rows": self.thvt_rows,
            "generated_thvt_rows": self.generated_thvt_rows,
            "thvt_mismatches": len(self.thvt_mismatches),
            "thvt_key_mismatches": len(self.thvt_key_mismatches),
            "thvt_missing_rows": len(self.thvt_missing_rows),
            "thvt_extra_rows": len(self.thvt_extra_rows),
            "ok": self.ok,
        }


@dataclass(frozen=True)
class ThvtRow:
    source_no: str
    group_key: str
    item_key: str
    label: str
    sequence: str
    code: str
    name: str
    unit: str
    norm: float
    coef: float
    bid_qty: float
    total_qty: float
    unit_price: float
    amount: float
    source_bid_row: int | None = None
    source_detail_row: int | None = None

    def data_key(self) -> tuple:
        return (
            self.source_no,
            self.label.upper(),
            _fold(self.code),
            _fold(self.name),
            _fold(self.unit),
        )

    def compare_key(self) -> tuple:
        return (
            self.source_no,
            self.label.upper(),
            self.sequence,
            _fold(self.code),
            _fold(self.name),
            _fold(self.unit),
        )

    def as_dict(self) -> dict:
        return {
            "source_no": self.source_no,
            "group_key": self.group_key,
            "item_key": self.item_key,
            "label": self.label,
            "sequence": self.sequence,
            "code": self.code,
            "name": self.name,
            "unit": self.unit,
            "norm": self.norm,
            "coef": self.coef,
            "bid_qty": self.bid_qty,
            "total_qty": self.total_qty,
            "unit_price": self.unit_price,
            "amount": self.amount,
            "source_bid_row": self.source_bid_row,
            "source_detail_row": self.source_detail_row,
        }


@dataclass(frozen=True)
class ThvtSummaryRow:
    label: str
    code: str
    name: str
    unit: str
    unit_price: float
    total_qty: float
    amount: float
    source_count: int
    source_refs: str


def analyze_estimate_workbook(
    path: str,
    bid_sheet_index: int | None = None,
    detail_sheet_index: int | None = None,
    thvt_sheet_index: int | None = None,
    bid_header_row: int | None = None,
    detail_header_row: int | None = None,
    bid_columns: dict | None = None,
    detail_columns: dict | None = None,
) -> EstimateAnalysisResult:
    """Analyze a Ho Guom-style estimate workbook.

    The current validated sample is an Excel 97-2003 .xls file. Future workflow
    code can wrap the same classification logic for .xlsx input once the UI
    contract is finalized.
    """

    workbook = _open_workbook(path, formatting_info=True)
    bid_sheet_index, detail_sheet_index = _resolve_estimate_sheet_indexes(workbook, bid_sheet_index, detail_sheet_index)
    bid_sheet = workbook.sheet_by_index(bid_sheet_index)
    detail_sheet = workbook.sheet_by_index(detail_sheet_index)

    bid_header = _resolve_header_row(bid_header_row, bid_sheet, ["ten cong tac", "khoi luong"])
    detail_header = _resolve_header_row(detail_header_row, detail_sheet, ["ma so", "thanh phan hao phi", "thanh tien"])
    bid_cols = _bid_columns(bid_sheet, bid_header, bid_columns)
    detail_cols = _detail_columns(detail_sheet, detail_header, detail_columns)

    bid_rows = _bid_item_rows(bid_sheet, bid_header + 1, bid_cols)
    detail_blocks = _detail_blocks(detail_sheet, detail_header + 1, detail_cols)
    bid_multipliers = _bid_parent_multipliers(bid_sheet, bid_header + 1, bid_cols)

    result = EstimateAnalysisResult(
        path=path,
        sheet_names=workbook.sheet_names(),
        bid_rows=len(bid_rows),
        detail_blocks=len(detail_blocks),
        label_totals={label: 0.0 for label in COST_LABEL_ORDER},
    )

    for block_index, (start, end) in enumerate(detail_blocks):
        if block_index >= len(bid_rows):
            result.identity_mismatches.append(
                {"block": block_index + 1, "detail_row": start + 1, "reason": "missing Du thau row"}
            )
            continue

        bid_row = bid_rows[block_index]
        _validate_identity(result, block_index, bid_sheet, bid_row, bid_cols, detail_sheet, start, detail_cols)

        sums = _classify_detail_block(result, detail_sheet, start + 1, end, detail_cols)
        for key, value in sums.items():
            result.label_totals[key] += value
        _validate_bid_totals(result, block_index, bid_sheet, bid_row, bid_cols, detail_sheet, start, detail_cols, sums, bid_multipliers)

    if thvt_sheet_index is None:
        thvt_sheet_index = _find_optional_exact_sheet_index(workbook, "thvt")
    if thvt_sheet_index is not None and thvt_sheet_index < workbook.nsheets:
        thvt_sheet = workbook.sheet_by_index(thvt_sheet_index)
        _validate_thvt(result, bid_sheet, bid_rows, bid_cols, detail_sheet, detail_blocks, detail_cols, thvt_sheet, bid_multipliers)

    return result


def list_estimate_workbook_sheets(path: str) -> dict:
    """Return workbook sheets and best-effort defaults for the sheet-choice stage."""

    workbook = _open_workbook(path, formatting_info=False)
    sheets = []
    for index, name in enumerate(workbook.sheet_names()):
        sheet = workbook.sheet_by_index(index)
        sheets.append({
            "index": index,
            "name": name,
            "rows": int(getattr(sheet, "nrows", 0) or 0),
            "cols": int(getattr(sheet, "ncols", 0) or 0),
        })
    suggested: dict[str, object | None] = {"bid_sheet_index": None, "detail_sheet_index": None}
    try:
        bid_index = _find_sheet_index(workbook, ["du thau", "du thau (2)"], ["ten cong tac", "khoi luong"])
        bid_sheet = workbook.sheet_by_index(bid_index)
        bid_header = _find_header_row(bid_sheet, ["ten cong tac", "khoi luong"])
        suggested["bid_sheet_index"] = bid_index
        suggested["bid_header_row"] = bid_header + 1
        suggested["bid_columns"] = _columns_to_letters(_bid_columns(bid_sheet, bid_header))
    except ValueError:
        pass
    try:
        detail_index = _find_sheet_index(workbook, ["chiet tinh"], ["ma so", "thanh phan hao phi", "thanh tien"])
        detail_sheet = workbook.sheet_by_index(detail_index)
        detail_header = _find_header_row(detail_sheet, ["ma so", "thanh phan hao phi", "thanh tien"])
        suggested["detail_sheet_index"] = detail_index
        suggested["detail_header_row"] = detail_header + 1
        suggested["detail_columns"] = _columns_to_letters(_detail_columns(detail_sheet, detail_header))
    except ValueError:
        pass
    return {
        "sheet_names": workbook.sheet_names(),
        "sheets": sheets,
        "suggested_sheets": suggested,
    }


def create_estimate_output_workbook(
    path: str,
    bid_sheet_index: int | None = None,
    detail_sheet_index: int | None = None,
    bid_header_row: int | None = None,
    detail_header_row: int | None = None,
    bid_columns: dict | None = None,
    detail_columns: dict | None = None,
) -> tuple[bytes, dict]:
    """Create the Ho Guom extracted workbook as modern .xlsx bytes."""

    source = _open_workbook(path, formatting_info=True)
    bid_sheet_index, detail_sheet_index = _resolve_estimate_sheet_indexes(source, bid_sheet_index, detail_sheet_index)
    result = analyze_estimate_workbook(
        path,
        bid_sheet_index=bid_sheet_index,
        detail_sheet_index=detail_sheet_index,
        thvt_sheet_index=None,
        bid_header_row=bid_header_row,
        detail_header_row=detail_header_row,
        bid_columns=bid_columns,
        detail_columns=detail_columns,
    )
    bid_sheet = source.sheet_by_index(bid_sheet_index)
    detail_sheet = source.sheet_by_index(detail_sheet_index)

    bid_header = _resolve_header_row(bid_header_row, bid_sheet, ["ten cong tac", "khoi luong"])
    detail_header = _resolve_header_row(detail_header_row, detail_sheet, ["ma so", "thanh phan hao phi", "thanh tien"])
    bid_cols = _bid_columns(bid_sheet, bid_header, bid_columns)
    detail_cols = _detail_columns(detail_sheet, detail_header, detail_columns)
    display_bid_sheet = _preferred_display_bid_sheet(source, bid_sheet_index)
    display_bid_header = _find_header_row(display_bid_sheet, ["ten cong tac", "khoi luong"]) if display_bid_sheet is not None else None
    display_bid_cols = _bid_columns(display_bid_sheet, display_bid_header) if display_bid_sheet is not None and display_bid_header is not None else None
    bid_rows = _bid_item_rows(bid_sheet, bid_header + 1, bid_cols)
    detail_blocks = _detail_blocks(detail_sheet, detail_header + 1, detail_cols)
    bid_multipliers = _bid_parent_multipliers(bid_sheet, bid_header + 1, bid_cols)

    workbook = Workbook()
    ws_bid = workbook.active
    ws_bid.title = "Dự thầu"
    ws_detail = workbook.create_sheet("Chiết tính")
    ws_thvt = workbook.create_sheet("THVT")
    ws_thvt_summary = workbook.create_sheet("Tổng hợp THVT")

    _write_bid_output_sheet(
        ws_bid,
        bid_sheet,
        bid_rows,
        bid_cols,
        detail_sheet,
        detail_blocks,
        detail_cols,
        bid_header,
        bid_multipliers,
        display_bid_sheet=display_bid_sheet,
        display_bid_header=display_bid_header,
        display_bid_cols=display_bid_cols,
    )
    _copy_sheet_values(ws_detail, detail_sheet, max_cols=min(detail_sheet.ncols, 13))
    thvt_rows = _generate_thvt_rows(bid_sheet, bid_rows, bid_cols, detail_sheet, detail_blocks, detail_cols, bid_multipliers)
    _write_thvt_sheet(ws_thvt, thvt_rows)
    _write_thvt_summary_sheet(ws_thvt_summary, _summarize_thvt_rows(thvt_rows))

    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue(), result.summary()


class _OpenpyxlSheetAdapter:
    def __init__(self, worksheet):
        self._worksheet = worksheet
        self.nrows = worksheet.max_row or 0
        self.ncols = worksheet.max_column or 0

    def cell_value(self, row: int, col: int):
        if row < 0 or col < 0:
            return ""
        value = self._worksheet.cell(row=row + 1, column=col + 1).value
        return "" if value is None else value


class _OpenpyxlWorkbookAdapter:
    def __init__(self, workbook):
        self._workbook = workbook
        self.nsheets = len(workbook.worksheets)

    def sheet_names(self):
        return list(self._workbook.sheetnames)

    def sheet_by_index(self, index: int):
        return _OpenpyxlSheetAdapter(self._workbook.worksheets[index])


def _open_workbook(path: str, formatting_info: bool = False):
    suffix = os.path.splitext(str(path))[1].lower()
    if suffix == ".xls":
        try:
            import xlrd
        except ModuleNotFoundError as exc:
            raise RuntimeError("Reading legacy .xls estimate workbooks requires xlrd. Save the file as .xlsx and upload again.") from exc
        return xlrd.open_workbook(path, formatting_info=formatting_info)
    return _OpenpyxlWorkbookAdapter(load_workbook(path, data_only=True))


def _cell(sheet, row: int, col: int | None):
    if col is None or row < 0 or col < 0 or row >= sheet.nrows or col >= sheet.ncols:
        return ""
    value = sheet.cell_value(row, col)
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def _text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _fold(value) -> str:
    text = unicodedata.normalize("NFD", _text(value))
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    text = text.replace("đ", "d").replace("Đ", "D")
    text = text.replace("đ", "d").replace("Đ", "D")
    return re.sub(r"\s+", " ", text).strip().lower()


def _number(value) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if not text:
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def _is_numeric_stt(value) -> bool:
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value).is_integer() and value > 0
    return bool(re.fullmatch(r"\d+", _text(value)))


def _is_bid_stt(value) -> bool:
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return value > 0
    return bool(re.fullmatch(r"\d+(?:\.\d+)*\.?", _text(value)))


def _excel_col_to_index(value) -> int | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value - 1 if value > 0 else None
    if isinstance(value, float) and value.is_integer():
        return int(value) - 1 if value > 0 else None
    text = _text(value).strip().upper()
    if not text:
        return None
    if text.isdigit():
        number = int(text)
        return number - 1 if number > 0 else None
    if not re.fullmatch(r"[A-Z]+", text):
        raise ValueError(f"Invalid Excel column: {value}")
    result = 0
    for char in text:
        result = result * 26 + (ord(char) - ord("A") + 1)
    return result - 1


def _index_to_excel_col(index: int | None) -> str:
    if index is None or index < 0:
        return ""
    return get_column_letter(index + 1)


def _columns_to_letters(columns: SheetColumns) -> dict[str, str]:
    return {field: _index_to_excel_col(getattr(columns, field)) for field in SheetColumns.__dataclass_fields__}


def _normalize_column_overrides(overrides: dict | None) -> dict[str, int | None]:
    if not isinstance(overrides, dict):
        return {}
    valid = set(SheetColumns.__dataclass_fields__)
    normalized: dict[str, int | None] = {}
    for key, value in overrides.items():
        if key in valid:
            normalized[key] = _excel_col_to_index(value)
    return normalized


def _columns_from_overrides(overrides: dict | None, required_fields: list[str]) -> SheetColumns | None:
    normalized = _normalize_column_overrides(overrides)
    if not normalized:
        return None
    if any(normalized.get(field) is None for field in required_fields):
        return None
    values = {field: None for field in SheetColumns.__dataclass_fields__}
    values.update(normalized)
    return SheetColumns(**values)


def _apply_column_overrides(columns: SheetColumns, overrides: dict | None, required_fields: list[str]) -> SheetColumns:
    normalized = _normalize_column_overrides(overrides)
    if not normalized:
        return columns
    values = {field: getattr(columns, field) for field in SheetColumns.__dataclass_fields__}
    values.update(normalized)
    missing = [field for field in required_fields if values.get(field) is None]
    if missing:
        raise ValueError(f"Missing required column config: {', '.join(missing)}")
    return SheetColumns(**values)


def _resolve_header_row(value, sheet, required_labels: list[str]) -> int:
    if value is None or value == "":
        return _find_header_row(sheet, required_labels)
    try:
        row = int(value) - 1
    except (TypeError, ValueError) as exc:
        raise ValueError(f"Invalid header row: {value}") from exc
    if row < 0 or row >= sheet.nrows:
        raise ValueError(f"Header row is outside sheet range: {value}")
    return row


def _find_header_row(sheet, required_labels: list[str]) -> int:
    folded_required = set(required_labels)
    best_row = None
    best_score = -1
    for row in range(min(sheet.nrows, 80)):
        labels = {_fold(_cell(sheet, row, col)) for col in range(sheet.ncols)}
        score = len(labels & folded_required)
        if score > best_score:
            best_score = score
            best_row = row
        if folded_required <= labels:
            return row
    if best_row is None or best_score <= 0:
        raise ValueError(f"Cannot find header row for labels: {required_labels}")
    return best_row


def _find_sheet_index(workbook, preferred_names, required_labels: list[str]) -> int:
    sheet_names = workbook.sheet_names()
    folded_names = [_fold(name) for name in sheet_names]
    for preferred_name in list(preferred_names):
        preferred_folded = _fold(preferred_name)
        for index, folded_name in enumerate(folded_names):
            if folded_name == preferred_folded:
                return index
    for index in range(workbook.nsheets):
        try:
            _find_header_row(workbook.sheet_by_index(index), required_labels)
            return index
        except ValueError:
            continue
    raise ValueError(f"Cannot find worksheet with headers: {required_labels}")


def _resolve_estimate_sheet_indexes(workbook, bid_sheet_index: int | None, detail_sheet_index: int | None) -> tuple[int, int]:
    if bid_sheet_index is None:
        bid_sheet_index = _find_sheet_index(workbook, ["du thau", "du thau (2)"], ["ten cong tac", "khoi luong"])
    else:
        bid_sheet_index = _validate_sheet_index(workbook, bid_sheet_index, "Dự thầu")
    if detail_sheet_index is None:
        detail_sheet_index = _find_sheet_index(workbook, ["chiet tinh"], ["ma so", "thanh phan hao phi", "thanh tien"])
    else:
        detail_sheet_index = _validate_sheet_index(workbook, detail_sheet_index, "Chiết tính")
    return bid_sheet_index, detail_sheet_index


def _find_optional_exact_sheet_index(workbook, name: str) -> int | None:
    target = _fold(name)
    for index, sheet_name in enumerate(workbook.sheet_names()):
        if _fold(sheet_name) == target:
            return index
    return None


def _preferred_display_bid_sheet(workbook, bid_sheet_index: int):
    display_index = _find_optional_exact_sheet_index(workbook, "du thau (2)")
    if display_index is None or display_index == bid_sheet_index:
        return None
    try:
        return workbook.sheet_by_index(display_index)
    except Exception:
        return None


def _validate_sheet_index(workbook, index: int, label: str) -> int:
    if not isinstance(index, int) or index < 0 or index >= workbook.nsheets:
        raise ValueError(f"Sheet {label} không hợp lệ.")
    return index


def _first_header_col(sheet, header_row: int, aliases: list[str], start: int = 0) -> int:
    folded_aliases = set(aliases)
    for col in range(start, sheet.ncols):
        if _fold(_cell(sheet, header_row, col)) in folded_aliases:
            return col
    raise ValueError(f"Cannot find header column: {aliases}")


def _optional_header_col(sheet, header_row: int, aliases: list[str], start: int = 0) -> int | None:
    try:
        return _first_header_col(sheet, header_row, aliases, start)
    except ValueError:
        return None


def _bid_columns(sheet, header_row: int, overrides: dict | None = None) -> SheetColumns:
    manual = _columns_from_overrides(overrides, ["stt", "name", "unit", "qty", "unit_total", "amount_total"])
    if manual is not None:
        return manual
    stt = _first_header_col(sheet, header_row, ["stt"])
    code = _optional_header_col(sheet, header_row, ["ma so"])
    name = _first_header_col(sheet, header_row, ["ten cong tac"])
    unit = _first_header_col(sheet, header_row, ["don vi"])
    qty = _first_header_col(sheet, header_row, ["khoi luong"])
    vl = _optional_header_col(sheet, header_row, ["vat lieu (d)", "vat lieu"], qty + 1)
    nc = _optional_header_col(sheet, header_row, ["nhan cong (d)", "nhan cong"], (vl + 1) if vl is not None else qty + 1)
    mtc = _optional_header_col(sheet, header_row, ["may (d)", "may"], (nc + 1) if nc is not None else qty + 1)
    unit_total = _first_header_col(sheet, header_row, ["don gia (d)", "don gia"], (mtc + 1) if mtc is not None else qty + 1)
    amount_total = _first_header_col(sheet, header_row, ["thanh tien (d)", "thanh tien"], unit_total + 1)
    columns = SheetColumns(stt, code, name, unit, qty, vl, nc, mtc, unit_total, amount_total)
    return _apply_column_overrides(columns, overrides, ["stt", "name", "unit", "qty", "unit_total", "amount_total"])


def _detail_columns(sheet, header_row: int, overrides: dict | None = None) -> SheetColumns:
    manual = _columns_from_overrides(overrides, ["stt", "code", "name", "unit", "amount"])
    if manual is not None:
        return manual
    code = _first_header_col(sheet, header_row, ["ma so"])
    stt = _first_header_col(sheet, header_row, ["stt"])
    name = _first_header_col(sheet, header_row, ["thanh phan hao phi"])
    unit = _first_header_col(sheet, header_row, ["don vi"])
    norm = _optional_header_col(sheet, header_row, ["dinh muc"])
    price = _optional_header_col(sheet, header_row, ["don gia (d)", "don gia"])
    coef = _optional_header_col(sheet, header_row, ["he so"])
    amount = _first_header_col(sheet, header_row, ["thanh tien (d)", "thanh tien"], unit + 1)
    helper = _detect_helper_column(sheet, header_row + 1, code)
    columns = SheetColumns(stt, code, name, unit, norm=norm, price=price, coef=coef, amount=amount, helper=helper)
    return _apply_column_overrides(columns, overrides, ["stt", "code", "name", "unit", "amount"])


def _detect_helper_column(sheet, data_start: int, before_col: int | None) -> int | None:
    if before_col is None:
        return None
    best_col = None
    best_count = 0
    for col in range(before_col):
        count = 0
        for row in range(data_start, sheet.nrows):
            if _text(_cell(sheet, row, col)).upper() in ALL_LABELS:
                count += 1
        if count > best_count:
            best_col = col
            best_count = count
    return best_col if best_count else None


def _bid_item_rows(sheet, data_start: int, columns: SheetColumns) -> list[int]:
    rows = []
    for row in range(data_start, sheet.nrows):
        if not _is_bid_stt(_cell(sheet, row, columns.stt)):
            continue
        code = _text(_cell(sheet, row, columns.code)) if columns.code is not None else ""
        if code == "*":
            continue
        if not _text(_cell(sheet, row, columns.name)) or not _text(_cell(sheet, row, columns.unit)):
            continue
        if _bid_row_has_cost(sheet, row, columns):
            rows.append(row)
    return rows


def _detail_blocks(sheet, data_start: int, columns: SheetColumns) -> list[tuple[int, int]]:
    starts = []
    for row in range(data_start, sheet.nrows):
        if _is_numeric_stt(_cell(sheet, row, columns.stt)):
            starts.append(row)
    blocks = [(start, starts[index + 1] if index + 1 < len(starts) else sheet.nrows) for index, start in enumerate(starts)]
    return [(start, end) for start, end in blocks if _detail_block_has_cost(sheet, start, end, columns)]


def _bid_row_has_cost(sheet, row: int, columns: SheetColumns) -> bool:
    for col in [columns.vl, columns.nc, columns.mtc, columns.unit_total, columns.amount_total]:
        if col is not None and _number(_cell(sheet, row, col)):
            return True
    return False


def _detail_block_has_cost(sheet, start: int, end: int, columns: SheetColumns) -> bool:
    sums = _detail_block_sums(sheet, start + 1, end, columns)
    return any(abs(value) > 0 for value in sums.values())


def _is_bid_total_row(sheet, row: int, columns: SheetColumns) -> bool:
    return _fold(_cell(sheet, row, columns.name)).startswith("tong cong")


def _is_bid_parent_row(sheet, row: int, columns: SheetColumns) -> bool:
    code = _text(_cell(sheet, row, columns.code)) if columns.code is not None else ""
    if code != "*":
        return False
    if _is_bid_total_row(sheet, row, columns):
        return False
    if _bid_row_has_cost(sheet, row, columns):
        return False
    return bool(_number(_cell(sheet, row, columns.qty)))


def _bid_parent_multipliers(sheet, data_start: int, columns: SheetColumns) -> dict[int, float]:
    multipliers: dict[int, float] = {}
    current_multiplier = 1.0
    for row in range(data_start, sheet.nrows):
        code = _text(_cell(sheet, row, columns.code)) if columns.code is not None else ""
        if _is_bid_total_row(sheet, row, columns):
            current_multiplier = 1.0
            continue
        if _is_bid_parent_row(sheet, row, columns):
            current_multiplier = _number(_cell(sheet, row, columns.qty)) or 1.0
            continue
        if code == "*" and not _number(_cell(sheet, row, columns.qty)):
            current_multiplier = 1.0
            continue
        if _bid_row_has_cost(sheet, row, columns):
            multipliers[row] = current_multiplier or 1.0
    return multipliers


def _bid_parent_groups(sheet, data_start: int, columns: SheetColumns) -> dict[int, list[int]]:
    groups: dict[int, list[int]] = {}
    current_parent: int | None = None
    for row in range(data_start, sheet.nrows):
        code = _text(_cell(sheet, row, columns.code)) if columns.code is not None else ""
        if _is_bid_total_row(sheet, row, columns):
            current_parent = None
            continue
        if _is_bid_parent_row(sheet, row, columns):
            current_parent = row
            groups.setdefault(row, [])
            continue
        if code == "*" and not _number(_cell(sheet, row, columns.qty)):
            current_parent = None
            continue
        if current_parent is not None and _bid_row_has_cost(sheet, row, columns):
            groups.setdefault(current_parent, []).append(row)
    return {parent: children for parent, children in groups.items() if children}


def _bid_multiplier(multipliers: dict[int, float] | None, row: int) -> float:
    if not multipliers:
        return 1.0
    return multipliers.get(row, 1.0) or 1.0


def _effective_bid_qty(sheet, row: int, columns: SheetColumns, multipliers: dict[int, float] | None = None) -> float:
    return _number(_cell(sheet, row, columns.qty)) * _bid_multiplier(multipliers, row)


def _detail_to_bid_unit_factor(
    bid_sheet,
    bid_row: int,
    bid_cols: SheetColumns,
    sums: dict[str, float],
    bid_multipliers: dict[int, float] | None = None,
) -> float:
    detail_unit_total = sums.get("TC") or sum(sums.get(label, 0.0) for label in COST_LABEL_ORDER if label != "TC")
    if not detail_unit_total:
        return 1.0
    bid_unit_total = _number(_cell(bid_sheet, bid_row, bid_cols.unit_total))
    if bid_unit_total:
        return bid_unit_total / detail_unit_total
    bid_amount_total = _number(_cell(bid_sheet, bid_row, bid_cols.amount_total))
    effective_qty = _effective_bid_qty(bid_sheet, bid_row, bid_cols, bid_multipliers)
    if bid_amount_total and effective_qty:
        return bid_amount_total / (detail_unit_total * effective_qty)
    return 1.0


def _convert_detail_sums_to_bid_unit(
    bid_sheet,
    bid_row: int,
    bid_cols: SheetColumns,
    sums: dict[str, float],
    bid_multipliers: dict[int, float] | None = None,
) -> dict[str, float]:
    factor = _detail_to_bid_unit_factor(bid_sheet, bid_row, bid_cols, sums, bid_multipliers)
    converted = {label: sums.get(label, 0.0) * factor for label in COST_LABEL_ORDER}
    converted["TC"] = converted.get("TC") or sum(converted[label] for label in COST_LABEL_ORDER if label != "TC")
    return converted


def _classify_row(name, unit, current_direct: str | None):
    name_key = _fold(name)
    unit_key = _fold(unit).upper()
    if not name_key:
        return None, current_direct, "blank"

    if re.match(r"^a\s*[\.\)]*\s*vat\s*lieu\b", name_key):
        return None, "VL", "direct_header"
    if re.match(r"^b\s*[\.\)]*\s*nhan\s*cong\b", name_key):
        return None, "NC", "direct_header"
    if re.match(r"^c\s*[\.\)]*\s*may\s*thi\s*cong\b", name_key):
        return None, "MTC", "direct_header"

    if name_key == "cong" or "cong chi phi truc tiep" in name_key:
        return None, current_direct, "summary"
    if name_key.startswith("chi phi gian tiep") or "cong chi phi gian tiep" in name_key:
        return None, current_direct, "summary"
    if "chi phi xay dung truoc thue" in name_key:
        return None, current_direct, "summary"

    if "chi phi chung" in name_key or unit_key == "C":
        return "CPC", current_direct, "indirect"
    if "chi phi nha tam" in name_key or unit_key == "LT":
        return "LT", current_direct, "indirect"
    if "chi phi mot so" in name_key or "chi phi khac" in name_key or unit_key == "TT":
        return "CPK", current_direct, "indirect"
    if "thu nhap chiu thue" in name_key or unit_key == "TL":
        return "TNCT", current_direct, "indirect"
    if "thue gia tri gia tang" in name_key or unit_key == "GTGT":
        return "VAT", current_direct, "indirect"
    if "chi phi xay dung sau thue" in name_key or unit_key == "GXD":
        return "TC", current_direct, "indirect"

    if current_direct in DIRECT_LABELS:
        return current_direct, current_direct, "direct_item"
    return None, current_direct, "unclassified"


def _classify_detail_block(result: EstimateAnalysisResult, sheet, start: int, end: int, columns: SheetColumns) -> dict[str, float]:
    return _detail_block_sums(sheet, start, end, columns, result)


def _detail_block_sums(sheet, start: int, end: int, columns: SheetColumns, result: EstimateAnalysisResult | None = None) -> dict[str, float]:
    sums = {label: 0.0 for label in COST_LABEL_ORDER}
    current_direct = None
    for row in range(start, end):
        label, current_direct, reason = _classify_row(
            _cell(sheet, row, columns.name),
            _cell(sheet, row, columns.unit),
            current_direct,
        )
        helper = _text(_cell(sheet, row, columns.helper)).upper() if columns.helper is not None else ""
        amount = _number(_cell(sheet, row, columns.amount))

        if label:
            sums[label] += amount
            if result is not None and helper in ALL_LABELS and helper != label:
                result.helper_mismatches.append(_row_note(sheet, row, columns, label=label, helper=helper, amount=amount, reason=reason))
        elif result is not None and helper in ALL_LABELS and reason not in {"direct_header", "summary", "blank"}:
            result.helper_mismatches.append(_row_note(sheet, row, columns, label=None, helper=helper, amount=amount, reason=reason))

        if result is not None and reason == "unclassified" and (_text(_cell(sheet, row, columns.name)) or amount):
            result.unclassified_rows.append(_row_note(sheet, row, columns, label=None, helper=helper, amount=amount, reason=reason))
    return sums


def _row_note(sheet, row: int, columns: SheetColumns, label, helper, amount: float, reason: str) -> dict:
    return {
        "row": row + 1,
        "name": _text(_cell(sheet, row, columns.name)),
        "unit": _text(_cell(sheet, row, columns.unit)),
        "amount": amount,
        "label": label,
        "helper": helper,
        "reason": reason,
    }


def _validate_identity(result, block_index, bid_sheet, bid_row, bid_cols, detail_sheet, detail_row, detail_cols) -> None:
    checks = [
        ("name", _cell(detail_sheet, detail_row, detail_cols.name), _cell(bid_sheet, bid_row, bid_cols.name)),
    ]
    if bid_cols.code is not None:
        checks.insert(0, ("code", _cell(detail_sheet, detail_row, detail_cols.code), _cell(bid_sheet, bid_row, bid_cols.code)))
    if any(_fold(left) != _fold(right) for _, left, right in checks):
        result.identity_mismatches.append(
            {
                "block": block_index + 1,
                "detail_row": detail_row + 1,
                "bid_row": bid_row + 1,
                "checks": {key: {"detail": _text(left), "bid": _text(right)} for key, left, right in checks},
            }
        )


def _validate_bid_totals(result, block_index, bid_sheet, bid_row, bid_cols, detail_sheet, detail_row, detail_cols, sums, bid_multipliers=None) -> None:
    qty = _effective_bid_qty(bid_sheet, bid_row, bid_cols, bid_multipliers)
    converted_sums = _convert_detail_sums_to_bid_unit(bid_sheet, bid_row, bid_cols, sums, bid_multipliers)
    got = {
        "unit_total": converted_sums["TC"] or sum(converted_sums[label] for label in ["VL", "NC", "MTC", "CPC", "LT", "CPK", "TNCT", "VAT"]),
    }
    for key, col in [("VL", bid_cols.vl), ("NC", bid_cols.nc), ("MTC", bid_cols.mtc)]:
        if col is not None:
            got[key] = converted_sums[key]
    got["amount_total"] = got["unit_total"] * qty
    expected = {
        "unit_total": _number(_cell(bid_sheet, bid_row, bid_cols.unit_total)),
        "amount_total": _number(_cell(bid_sheet, bid_row, bid_cols.amount_total)),
    }
    for key, col in [("VL", bid_cols.vl), ("NC", bid_cols.nc), ("MTC", bid_cols.mtc)]:
        if col is not None:
            expected[key] = _number(_cell(bid_sheet, bid_row, col))
    tolerances = {"amount_total": 2.0}
    for key, got_value in got.items():
        tolerance = tolerances.get(key, 1.0)
        if abs(got_value - expected[key]) > tolerance:
            result.calculation_mismatches.append(
                {
                    "block": block_index + 1,
                    "detail_row": detail_row + 1,
                    "bid_row": bid_row + 1,
                    "code": _text(_cell(detail_sheet, detail_row, detail_cols.code)),
                    "name": _text(_cell(detail_sheet, detail_row, detail_cols.name)),
                    "field": key,
                    "calculated": got_value,
                    "expected": expected[key],
                    "diff": got_value - expected[key],
                }
            )
            return


def _validate_thvt(result, bid_sheet, bid_rows, bid_cols, detail_sheet, detail_blocks, detail_cols, thvt_sheet, bid_multipliers=None) -> None:
    generated = _generate_thvt_rows(bid_sheet, bid_rows, bid_cols, detail_sheet, detail_blocks, detail_cols, bid_multipliers)
    actual = _read_thvt_rows(thvt_sheet)
    result.generated_thvt_rows = len(generated)
    result.thvt_rows = len(actual)

    generated_by_key = _rows_by_thvt_data_key(generated)
    actual_by_key = _rows_by_thvt_data_key(actual)

    for key, generated_rows in generated_by_key.items():
        actual_rows = actual_by_key.get(key, [])
        for index, generated_row in enumerate(generated_rows):
            if index >= len(actual_rows):
                result.thvt_missing_rows.append(generated_row.as_dict())
                continue
            actual_row = actual_rows[index]
            mismatch = _compare_thvt_row(generated_row, actual_row)
            if mismatch:
                result.thvt_mismatches.append(mismatch)
            key_mismatch = _compare_thvt_helper_key(generated_row, actual_row)
            if key_mismatch:
                result.thvt_key_mismatches.append(key_mismatch)

    for key, actual_rows in actual_by_key.items():
        generated_rows = generated_by_key.get(key, [])
        if len(actual_rows) > len(generated_rows):
            for actual_row in actual_rows[len(generated_rows) :]:
                result.thvt_extra_rows.append(actual_row.as_dict())


def _rows_by_thvt_data_key(rows: list[ThvtRow]) -> dict[tuple, list[ThvtRow]]:
    grouped: dict[tuple, list[ThvtRow]] = {}
    for row in rows:
        grouped.setdefault(row.data_key(), []).append(row)
    for values in grouped.values():
        values.sort(key=lambda row: (_number(row.sequence), _fold(row.group_key), _fold(row.item_key), row.amount))
    return grouped


def _compare_thvt_helper_key(generated: ThvtRow, actual: ThvtRow) -> dict | None:
    diffs = {}
    for field in ["group_key", "item_key", "sequence"]:
        left = getattr(generated, field)
        right = getattr(actual, field)
        if _text(left).upper() != _text(right).upper():
            diffs[field] = {"generated": left, "actual": right}
    if not diffs:
        return None
    return {
        "key": generated.data_key(),
        "generated": generated.as_dict(),
        "actual": actual.as_dict(),
        "diffs": diffs,
    }


def _is_other_material_row(sheet, row: int, columns: SheetColumns) -> bool:
    code = _fold(_cell(sheet, row, columns.code))
    name = _fold(_cell(sheet, row, columns.name))
    return "vat lieu khac" in code or "vat lieu khac" in name


def _detail_material_unit_price(sheet, row: int, columns: SheetColumns) -> float:
    if _is_other_material_row(sheet, row, columns) and columns.amount is not None:
        return _number(_cell(sheet, row, columns.amount))
    return _number(_cell(sheet, row, columns.price))


def _detail_coef(sheet, row: int, columns: SheetColumns) -> float:
    if columns.coef is None:
        return 1.0
    raw = _cell(sheet, row, columns.coef)
    if _text(raw) == "":
        return 1.0
    return _number(raw)


def _detail_material_total_qty(sheet, row: int, columns: SheetColumns, bid_qty: float, detail_qty: float, norm: float, coef: float) -> float:
    if _is_other_material_row(sheet, row, columns):
        return bid_qty
    return norm * detail_qty * coef


def _detail_material_amount(sheet, row: int, columns: SheetColumns, detail_qty: float, total_qty: float, unit_price: float) -> float:
    if _is_other_material_row(sheet, row, columns):
        return total_qty * unit_price
    if columns.amount is not None:
        return _number(_cell(sheet, row, columns.amount)) * detail_qty
    return total_qty * unit_price


def _generate_thvt_rows(bid_sheet, bid_rows, bid_cols, detail_sheet, detail_blocks, detail_cols, bid_multipliers=None) -> list[ThvtRow]:
    rows: list[ThvtRow] = []
    for block_index, (start, end) in enumerate(detail_blocks):
        if block_index >= len(bid_rows):
            continue

        bid_row = bid_rows[block_index]
        source_no = _source_no_for_bid_row(bid_sheet, bid_row, bid_cols, block_index)
        bid_qty = _effective_bid_qty(bid_sheet, bid_row, bid_cols, bid_multipliers)
        block_sums = _detail_block_sums(detail_sheet, start + 1, end, detail_cols)
        detail_qty = bid_qty * _detail_to_bid_unit_factor(bid_sheet, bid_row, bid_cols, block_sums, bid_multipliers)
        direct_rows = _direct_detail_rows(detail_sheet, start + 1, end, detail_cols)
        label_counts = {label: sum(1 for item in direct_rows if item["label"] == label) for label in DIRECT_LABELS}
        label_seen = {label: 0 for label in DIRECT_LABELS}

        for item in direct_rows:
            label = item["label"]
            if label != "VL":
                continue
            label_seen[label] += 1
            sequence = str(label_seen[label]) if label_counts[label] > 1 else ""
            norm = _number(_cell(detail_sheet, item["row"], detail_cols.norm))
            coef = _detail_coef(detail_sheet, item["row"], detail_cols)
            unit_price = _detail_material_unit_price(detail_sheet, item["row"], detail_cols)
            total_qty = _detail_material_total_qty(detail_sheet, item["row"], detail_cols, bid_qty, detail_qty, norm, coef)
            amount = _detail_material_amount(detail_sheet, item["row"], detail_cols, detail_qty, total_qty, unit_price)
            rows.append(
                ThvtRow(
                    source_no=source_no,
                    group_key=f"{source_no}{label}",
                    item_key=f"{source_no}{label}{sequence}",
                    label=label,
                    sequence=sequence,
                    code=_text(_cell(detail_sheet, item["row"], detail_cols.code)),
                    name=_text(_cell(detail_sheet, item["row"], detail_cols.name)),
                    unit=_text(_cell(detail_sheet, item["row"], detail_cols.unit)),
                    norm=norm,
                    coef=coef,
                    bid_qty=bid_qty,
                    total_qty=total_qty,
                    unit_price=unit_price,
                    amount=amount,
                    source_bid_row=bid_row + 1,
                    source_detail_row=item["row"] + 1,
                )
            )

    return sorted(rows, key=lambda row: (_fold(row.name), _fold(row.code), _fold(row.unit), _number(row.source_no), row.label, _number(row.sequence)))


def _source_no_for_bid_row(bid_sheet, bid_row, bid_cols, block_index: int) -> str:
    # In the Ho Guom sample, Du thau column B is the global work-item number
    # used by THVT, while column A resets inside section headings.
    if bid_cols.code is not None:
        likely_global_col = bid_cols.code - 1
        if likely_global_col >= 0:
            value = _text(_cell(bid_sheet, bid_row, likely_global_col))
            if _is_numeric_stt(value):
                return str(int(_number(value)))
    value = _text(_cell(bid_sheet, bid_row, bid_cols.stt)).rstrip(".")
    if _is_bid_stt(value):
        return value
    return str(block_index + 1)


def _direct_detail_rows(sheet, start: int, end: int, columns: SheetColumns) -> list[dict]:
    rows = []
    current_direct = None
    for row in range(start, end):
        label, current_direct, reason = _classify_row(
            _cell(sheet, row, columns.name),
            _cell(sheet, row, columns.unit),
            current_direct,
        )
        has_detail_value = any(
            [
                _text(_cell(sheet, row, columns.code)),
                _number(_cell(sheet, row, columns.norm)),
                _number(_cell(sheet, row, columns.price)),
                _number(_cell(sheet, row, columns.amount)),
            ]
        )
        if label in DIRECT_LABELS and reason == "direct_item" and has_detail_value:
            rows.append({"row": row, "label": label})
    return rows


def _read_thvt_rows(sheet) -> list[ThvtRow]:
    header_row = _find_header_row(sheet, ["ma so", "thanh phan hao phi", "kl tong"])
    headers = {_fold(_cell(sheet, header_row, col)): col for col in range(sheet.ncols)}
    has_coef = "he so" in headers
    coef_col = headers.get("he so")
    bid_qty_col = headers.get("khoi luong", 9)
    total_qty_col = headers.get("kl tong", 10)
    unit_price_col = headers.get("don gia", 11)
    amount_col = headers.get("thanh tien", 12)
    rows: list[ThvtRow] = []
    for row in range(header_row + 1, sheet.nrows):
        label = _text(_cell(sheet, row, 3)).upper()
        if label not in DIRECT_LABELS:
            continue
        rows.append(
            ThvtRow(
                source_no=_text(_cell(sheet, row, 0)),
                group_key=_text(_cell(sheet, row, 1)),
                item_key=_text(_cell(sheet, row, 2)),
                label=label,
                sequence=_text(_cell(sheet, row, 4)),
                code=_text(_cell(sheet, row, 5)),
                name=_text(_cell(sheet, row, 6)),
                unit=_text(_cell(sheet, row, 7)),
                norm=_number(_cell(sheet, row, 8)),
                coef=_number(_cell(sheet, row, coef_col)) if has_coef else 1.0,
                bid_qty=_number(_cell(sheet, row, bid_qty_col)),
                total_qty=_number(_cell(sheet, row, total_qty_col)),
                unit_price=_number(_cell(sheet, row, unit_price_col)),
                amount=_number(_cell(sheet, row, amount_col)),
            )
        )
    return rows


def _compare_thvt_row(generated: ThvtRow, actual: ThvtRow) -> dict | None:
    fields = [
        ("norm", generated.norm, actual.norm, 0.000001),
        ("bid_qty", generated.bid_qty, actual.bid_qty, 0.000001),
        ("total_qty", generated.total_qty, actual.total_qty, 0.000001),
        ("unit_price", generated.unit_price, actual.unit_price, 0.000001),
        ("amount", generated.amount, actual.amount, 0.01),
    ]
    diffs = {}
    for field, left, right, tolerance in fields:
        if tolerance is None:
            if _text(left).upper() != _text(right).upper():
                diffs[field] = {"generated": left, "actual": right}
        elif abs(float(left) - float(right)) > tolerance:
            diffs[field] = {"generated": left, "actual": right, "diff": float(left) - float(right)}
    if not diffs:
        return None
    return {
        "key": generated.compare_key(),
        "generated": generated.as_dict(),
        "actual": actual.as_dict(),
        "diffs": diffs,
    }


def _copy_sheet_values(worksheet, source_sheet, max_cols: int | None = None) -> None:
    limit_cols = max_cols or source_sheet.ncols
    for row in range(source_sheet.nrows):
        for col in range(limit_cols):
            worksheet.cell(row=row + 1, column=col + 1, value=_cell(source_sheet, row, col))
    _style_basic_sheet(worksheet, min(limit_cols, 20))


def _write_bid_output_sheet(
    worksheet,
    bid_sheet,
    bid_rows,
    bid_cols,
    detail_sheet,
    detail_blocks,
    detail_cols,
    header_row: int,
    bid_multipliers=None,
    display_bid_sheet=None,
    display_bid_header: int | None = None,
    display_bid_cols: SheetColumns | None = None,
) -> None:
    base_cols = min(max(11, (bid_cols.amount_total or 10) + 1), bid_sheet.ncols)
    _copy_sheet_values(worksheet, bid_sheet, max_cols=base_cols)
    if display_bid_sheet is not None and display_bid_header is not None and display_bid_cols is not None:
        _apply_display_stt(worksheet, bid_sheet, header_row + 1, bid_cols, display_bid_sheet, display_bid_header + 1, display_bid_cols)

    price_start = base_cols + 3
    amount_start = price_start + len(COST_LABEL_ORDER)
    title_row = max(1, header_row)
    label_row = header_row + 1

    worksheet.merge_cells(
        start_row=title_row,
        start_column=price_start,
        end_row=title_row,
        end_column=price_start + len(COST_LABEL_ORDER) - 1,
    )
    worksheet.merge_cells(
        start_row=title_row,
        start_column=amount_start,
        end_row=title_row,
        end_column=amount_start + len(COST_LABEL_ORDER) - 1,
    )
    worksheet.cell(row=title_row, column=price_start, value="Don gia")
    worksheet.cell(row=title_row, column=amount_start, value="Thanh tien")
    for offset, label in enumerate(COST_LABEL_ORDER):
        worksheet.cell(row=label_row, column=price_start + offset, value=label)
        worksheet.cell(row=label_row, column=amount_start + offset, value=label)

    row_amounts_by_label: dict[int, dict[str, float]] = {}
    for block_index, (start, end) in enumerate(detail_blocks):
        if block_index >= len(bid_rows):
            continue
        bid_row = bid_rows[block_index]
        excel_row = bid_row + 1
        qty = _effective_bid_qty(bid_sheet, bid_row, bid_cols, bid_multipliers)
        multiplier = _bid_multiplier(bid_multipliers, bid_row)
        if multiplier != 1.0 and bid_cols.qty is not None:
            worksheet.cell(row=excel_row, column=bid_cols.qty + 1, value=qty)
        sums = _detail_block_sums(detail_sheet, start + 1, end, detail_cols)
        converted_sums = _convert_detail_sums_to_bid_unit(bid_sheet, bid_row, bid_cols, sums, bid_multipliers)
        row_amounts_by_label[bid_row] = {label: converted_sums[label] * qty for label in COST_LABEL_ORDER}
        if bid_cols.amount_total is not None:
            worksheet.cell(row=excel_row, column=bid_cols.amount_total + 1, value=row_amounts_by_label[bid_row]["TC"])
        for offset, label in enumerate(COST_LABEL_ORDER):
            unit_price = converted_sums[label]
            worksheet.cell(row=excel_row, column=price_start + offset, value=unit_price)
            worksheet.cell(row=excel_row, column=amount_start + offset, value=row_amounts_by_label[bid_row][label])

    parent_groups = _bid_parent_groups(bid_sheet, header_row + 1, bid_cols)
    for parent_row, child_rows in parent_groups.items():
        parent_excel_row = parent_row + 1
        parent_qty = _number(_cell(bid_sheet, parent_row, bid_cols.qty)) or 1.0
        aggregate_amounts = {label: 0.0 for label in COST_LABEL_ORDER}
        for child_row in child_rows:
            for label, amount in row_amounts_by_label.get(child_row, {}).items():
                aggregate_amounts[label] += amount
        total_amount = aggregate_amounts["TC"] or sum(aggregate_amounts[label] for label in COST_LABEL_ORDER if label != "TC")
        aggregate_amounts["TC"] = total_amount
        if bid_cols.unit_total is not None:
            worksheet.cell(row=parent_excel_row, column=bid_cols.unit_total + 1, value=total_amount / parent_qty)
        if bid_cols.amount_total is not None:
            worksheet.cell(row=parent_excel_row, column=bid_cols.amount_total + 1, value=total_amount)
        for key, label in [("vl", "VL"), ("nc", "NC"), ("mtc", "MTC")]:
            col = getattr(bid_cols, key)
            if col is not None:
                worksheet.cell(row=parent_excel_row, column=col + 1, value=aggregate_amounts[label] / parent_qty)
        for offset, label in enumerate(COST_LABEL_ORDER):
            worksheet.cell(row=parent_excel_row, column=price_start + offset, value=aggregate_amounts[label] / parent_qty)
            worksheet.cell(row=parent_excel_row, column=amount_start + offset, value=aggregate_amounts[label])

    _style_basic_sheet(worksheet, amount_start + len(COST_LABEL_ORDER))
    _style_bid_generated_columns(worksheet, title_row, label_row, price_start, amount_start)
    for col in range(price_start, price_start + len(COST_LABEL_ORDER)):
        worksheet.column_dimensions[get_column_letter(col)].width = 13
    for col in range(amount_start, amount_start + len(COST_LABEL_ORDER)):
        worksheet.column_dimensions[get_column_letter(col)].width = 15
    worksheet.freeze_panes = worksheet.cell(row=header_row + 2, column=1).coordinate


def _style_bid_generated_columns(worksheet, title_row: int, label_row: int, price_start: int, amount_start: int) -> None:
    price_fill = PatternFill("solid", fgColor="CDBE7D")
    amount_fill = PatternFill("solid", fgColor="C6E0B4")
    label_fill = PatternFill("solid", fgColor="D9EAD3")
    border = Border(
        left=Side(style="thin", color="808080"),
        right=Side(style="thin", color="808080"),
        top=Side(style="thin", color="808080"),
        bottom=Side(style="thin", color="808080"),
    )
    for start_col, fill in [(price_start, price_fill), (amount_start, amount_fill)]:
        title = worksheet.cell(row=title_row, column=start_col)
        title.fill = fill
        title.font = Font(bold=True)
        title.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        title.border = border
        for col in range(start_col, start_col + len(COST_LABEL_ORDER)):
            header = worksheet.cell(row=label_row, column=col)
            header.fill = label_fill
            header.font = Font(bold=True)
            header.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            header.border = border


def _write_thvt_sheet(worksheet, rows: list[ThvtRow]) -> None:
    headers = [
        "STT",
        "Ma nhom",
        "Ma khoan",
        "Loai",
        "Thu tu",
        "Ma so",
        "Thanh phan hao phi",
        "Don vi",
        "Dinh muc",
        "He so",
        "Khoi luong",
        "KL tong",
        "Don gia",
        "Thanh tien",
        "Dong du thau",
        "Dong chiet tinh",
    ]
    worksheet.append(headers)
    for index, row in enumerate(rows, start=1):
        worksheet.append(
            [
                index,
                row.group_key,
                row.item_key,
                row.label,
                row.sequence,
                row.code,
                row.name,
                row.unit,
                row.norm,
                row.coef,
                row.bid_qty,
                row.total_qty,
                row.unit_price,
                row.amount,
                row.source_bid_row,
                row.source_detail_row,
            ]
        )
    _style_basic_sheet(worksheet, len(headers))
    worksheet.freeze_panes = "A2"
    for col, width in enumerate([8, 12, 14, 8, 8, 14, 42, 10, 12, 10, 12, 14, 14, 16, 14, 14], start=1):
        worksheet.column_dimensions[get_column_letter(col)].width = width


def _thvt_summary_key(row: ThvtRow) -> tuple:
    return (
        row.label.upper(),
        _fold(row.code),
        _fold(row.name),
        _fold(row.unit),
        round(float(row.unit_price or 0), 8),
    )


def _summarize_thvt_rows(rows: list[ThvtRow]) -> list[ThvtSummaryRow]:
    grouped: dict[tuple, dict] = {}
    for row in rows:
        key = _thvt_summary_key(row)
        if key not in grouped:
            grouped[key] = {
                "first": row,
                "total_qty": 0.0,
                "amount": 0.0,
                "source_count": 0,
                "source_refs": [],
            }
        bucket = grouped[key]
        bucket["total_qty"] += row.total_qty
        bucket["amount"] += row.amount
        bucket["source_count"] += 1
        if row.source_detail_row:
            bucket["source_refs"].append(str(row.source_detail_row))

    summary_rows: list[ThvtSummaryRow] = []
    for bucket in grouped.values():
        first = bucket["first"]
        source_refs = ", ".join(bucket["source_refs"])
        summary_rows.append(
            ThvtSummaryRow(
                label=first.label,
                code=first.code,
                name=first.name,
                unit=first.unit,
                unit_price=first.unit_price,
                total_qty=bucket["total_qty"],
                amount=bucket["amount"],
                source_count=bucket["source_count"],
                source_refs=source_refs,
            )
        )
    return summary_rows


def _write_thvt_summary_sheet(worksheet, rows: list[ThvtSummaryRow]) -> None:
    headers = [
        "STT",
        "Loai",
        "Ma so",
        "Thanh phan hao phi",
        "Don vi",
        "KL tong",
        "Don gia",
        "Thanh tien",
        "So dong THVT",
        "Dong chiet tinh",
    ]
    worksheet.append(headers)
    for index, row in enumerate(rows, start=1):
        worksheet.append(
            [
                index,
                row.label,
                row.code,
                row.name,
                row.unit,
                row.total_qty,
                row.unit_price,
                row.amount,
                row.source_count,
                row.source_refs,
            ]
        )

    _style_basic_sheet(worksheet, len(headers))
    worksheet.freeze_panes = "A2"
    for col, width in enumerate([8, 8, 14, 42, 10, 14, 14, 16, 14, 32], start=1):
        worksheet.column_dimensions[get_column_letter(col)].width = width


def _display_stt_match_key(sheet, row: int, columns: SheetColumns) -> tuple | None:
    name = _fold(_cell(sheet, row, columns.name))
    if not name:
        return None
    code = _fold(_cell(sheet, row, columns.code)) if columns.code is not None else ""
    if _is_bid_total_row(sheet, row, columns):
        return ("total", name)
    if not code or code == "*":
        return ("group", name)
    return ("item", code, name)


def _apply_display_stt(worksheet, target_sheet, target_start: int, target_cols: SheetColumns, display_sheet, display_start: int, display_cols: SheetColumns) -> None:
    display_entries: list[tuple[tuple, str]] = []
    for row in range(display_start, display_sheet.nrows):
        key = _display_stt_match_key(display_sheet, row, display_cols)
        if key is None:
            continue
        display_entries.append((key, _text(_cell(display_sheet, row, display_cols.stt))))

    cursor = 0
    stt_column = target_cols.stt + 1
    max_lookahead = 80
    for row in range(target_start, target_sheet.nrows):
        key = _display_stt_match_key(target_sheet, row, target_cols)
        if key is None:
            continue
        matched_index = None
        for index in range(cursor, min(len(display_entries), cursor + max_lookahead)):
            if display_entries[index][0] == key:
                matched_index = index
                break
        if matched_index is None:
            continue
        worksheet.cell(row=row + 1, column=stt_column, value=display_entries[matched_index][1])
        cursor = matched_index + 1
    _clear_nested_child_stt(worksheet, target_sheet, target_start, target_cols)
    _fill_missing_parent_stt_from_neighbors(worksheet, target_sheet, target_start, target_cols)


def _clear_nested_child_stt(worksheet, sheet, data_start: int, columns: SheetColumns) -> None:
    in_parent_group = False
    stt_column = columns.stt + 1
    for row in range(data_start, sheet.nrows):
        code = _text(_cell(sheet, row, columns.code)) if columns.code is not None else ""
        if _is_bid_total_row(sheet, row, columns):
            in_parent_group = False
            continue
        if _is_bid_parent_row(sheet, row, columns):
            in_parent_group = True
            continue
        if code == "*" and not _number(_cell(sheet, row, columns.qty)):
            in_parent_group = False
            continue
        if in_parent_group and _bid_row_has_cost(sheet, row, columns):
            worksheet.cell(row=row + 1, column=stt_column, value="")


def _parse_dotted_stt(value) -> list[int] | None:
    text = _text(value).rstrip(".")
    if not re.fullmatch(r"\d+(?:\.\d+)*", text):
        return None
    return [int(part) for part in text.split(".")]


def _format_dotted_stt(parts: list[int]) -> str:
    return ".".join(str(part) for part in parts)


def _fill_missing_parent_stt_from_neighbors(worksheet, sheet, data_start: int, columns: SheetColumns) -> None:
    stt_column = columns.stt + 1
    parent_rows: list[tuple[int, list[int] | None]] = []
    for row in range(data_start, sheet.nrows):
        if _is_bid_parent_row(sheet, row, columns):
            parent_rows.append((row, _parse_dotted_stt(worksheet.cell(row=row + 1, column=stt_column).value)))

    index = 0
    while index < len(parent_rows):
        if parent_rows[index][1] is not None:
            index += 1
            continue
        start = index
        while index < len(parent_rows) and parent_rows[index][1] is None:
            index += 1
        if start == 0 or index >= len(parent_rows):
            continue
        previous_parts = parent_rows[start - 1][1]
        next_parts = parent_rows[index][1]
        if previous_parts is None or next_parts is None:
            continue
        if len(previous_parts) != len(next_parts) or previous_parts[:-1] != next_parts[:-1]:
            continue
        missing_count = index - start
        available_gap = next_parts[-1] - previous_parts[-1] - 1
        if available_gap < missing_count:
            continue
        for offset in range(missing_count):
            row = parent_rows[start + offset][0]
            filled_parts = previous_parts[:-1] + [previous_parts[-1] + offset + 1]
            worksheet.cell(row=row + 1, column=stt_column, value=_format_dotted_stt(filled_parts))


def _style_basic_sheet(worksheet, max_cols: int) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAD3")
    title_fill = PatternFill("solid", fgColor="EAF4FF")
    border = Border(
        left=Side(style="thin", color="C9DDF0"),
        right=Side(style="thin", color="C9DDF0"),
        top=Side(style="thin", color="C9DDF0"),
        bottom=Side(style="thin", color="C9DDF0"),
    )
    for col in range(1, max_cols + 1):
        cell = worksheet.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        worksheet.column_dimensions[get_column_letter(col)].width = max(worksheet.column_dimensions[get_column_letter(col)].width or 0, 12)
    for row in worksheet.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0.###"
    if worksheet.max_row > 1:
        for col in range(1, max_cols + 1):
            worksheet.cell(row=2, column=col).fill = title_fill


def sample_path_from_env() -> str | None:
    path = os.environ.get("ESTIMATE_SAMPLE_XLS") or "E:\\Excel Mom\\D\u00f3c\\Ho Guom\\boctachdutoan.xls"
    return path if os.path.exists(path) else None
